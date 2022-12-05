import shutil
from time import sleep
import ctypes
from openpyxl.descriptors.serialisable import Serialisable
# import PIL
import ssl
import random2 as random
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
from openpyxl.cell import cell
from openpyxl.chart import LineChart, Reference
import xml.etree.ElementTree as ET
from flask import flash
from flask import Flask, render_template, request, send_from_directory, after_this_request, redirect, url_for
from openpyxl.descriptors import (
	String,
	Sequence,
	Integer,
	)
from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.styles import Alignment, alignment
from string import ascii_uppercase
import openpyxl
# import datetime
# from datetime import datetime
import os
app=Flask(__name__)
app.secret_key = "GT ROMANIA Delivery Center"
var_list=[]
var_rute=[]
# app.app_context().push()
	
# @app.route("/")
# def home():
#     # text=request.form.get('client')
#     # print(text)
#     return render_template("index.html")
@app.route('/Ageing/Instructions', methods=['GET'])
def downloadAgeing():
		filepath = "/home/auditappnexia/output/Ageing"
 
		return send_from_directory(filepath,"Instructions - Ageing.docx", as_attachment=True)
@app.route('/Ageing/GTv51bfdGKkuaKo9ggrm7plxbjn')
def Ageing():
	return render_template('Ageing.html')
@app.route('/Ageing/GTv51bfdGKkuaKo9ggrm7plxbjn', methods=['POST', 'GET'])
def Ageing_process():
	# global column_description, coloana_opTB_tb, coloana_cr_tb, coloana_db_tb, coloana_clTB_tb, column_tb, row_tb, clientName
	clientname1 = request.form['client']
	yearEnd1 = datetime.datetime.strptime(
		request.form['yearEnd'],
		'%Y-%m-%d')
	preparedBy1 = request.form['preparedBy']
	datePrepared1 = datetime.datetime.strptime(
		request.form['preparedDate'],
		'%Y-%m-%d')
	refference1 = request.form['reff']

	if request.method == 'POST':
		file_Details = request.files["details"]
		file_TB = request.files["TB"]
		val0 = request.form['notDue']
		val1 = request.form['bucket1']
		val2 = request.form['bucket2']
		val3 = request.form['bucket3']
		val4 = request.form['bucket4']
		val5 = request.form['bucket5']

		# for i in file_Details:
		# 	i.save(secure_filename(i.filename))
		# print(isChecked)



#colors
		ft1 = Font(name = 'Tahoma', size = 8, bold = True)
		ft1_1 = Font(name = 'Tahoma', size = 8)		
		ft1_1b = Font(name = 'Tahoma', size = 8, bold=True)
		f_testname = Font(name = 'Tahoma', size = 8, bold = True, underline = 'single', italic = True)
		cap_tabel = Font(name = 'Tahoma', size = 8, color = "FFFFFF", bold = True)
		cap_tabel_color_PBC = PatternFill(start_color = '808080', end_color ='808080', fill_type = 'solid') #grey
		cap_tabel_color_GT = PatternFill(start_color = '00AEAC', end_color ='00AEAC', fill_type = 'solid') #indigo
		fprocentaj = Font(name = 'Tahoma', size = 8, color = "FF0000", bold = True)
		font_worksheet = Font(name = 'Tahoma', size = 8)
		check_font = Font(name = 'Tahoma', size = 8, color = "FF0000", bold = True)
		thin = Side(border_style='double', color='000000')
		border_bottom = Border(bottom=thin)
		cap_tabel_galben = PatternFill(start_color = 'FFFF00', end_color ='FFFF00', fill_type = 'solid')
		doubleborder = Border(bottom=Side(style='double'))

		# variable = StringVar(app)
		# variable.set("(None)")
		#
		# app.mainloop()
		# try:
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "TR Aging Invoice level"
		Sheet1 = wb["TR Aging Invoice level"]
		Sheet1.sheet_view.showGridLines = False
		# Sheet1.font = ft1
		
		# try:
		details = openpyxl.load_workbook(file_Details, data_only = True)
		# except:
		# 	None
		ws = details.active

		#header
		Sheet1.cell(row = 1, column = 1).value = "Client name:"
		Sheet1.cell(row = 1, column = 2).value = clientname1
		Sheet1.cell(row = 2, column = 1).value = "Period end:"
		Sheet1.cell(row = 2, column = 2).value = yearEnd1
		Sheet1.cell(row = 2, column = 2).number_format = 'mm/dd/yyyy'

		#test name
		Sheet1.cell(row = 4, column = 1).value = "Receivables ageing recomputation"
		Sheet1.cell(row = 12, column = 1).value = "YE Date"
		Sheet1.cell(row = 13, column = 1).value = yearEnd1
		Sheet1.cell(row = 13, column = 1).number_format = 'mm/dd/yyyy'

		#testing table
		Sheet1.cell(row = 16, column = 1).value = "As per Detail"
		Sheet1.cell(row = 17, column = 1).value = "Client code"
		Sheet1.cell(row = 17, column = 2).value = "Client"
		Sheet1.cell(row = 17, column = 3).value = "Account"
		Sheet1.cell(row = 17, column = 4).value = "Due date"
		Sheet1.cell(row = 17, column = 5).value = "Invoice number"
		Sheet1.cell(row = 17, column = 6).value = "Amount in LC"
		Sheet1.cell(row = 16, column = 8).value = "As per Nexia"
		Sheet1.cell(row = 17, column = 7).value = "Invoice value"
		Sheet1.cell(row = 17, column = 8).value = "No of days"
		Sheet1.cell(row = 17, column = 9).value = "Not due"
		Sheet1.cell(row = 17, column = 10).value = "<="+str(val1)+" days"
		Sheet1.cell(row = 17, column = 11).value = str(val1)+"-"+str(val2)+" days"
		Sheet1.cell(row = 17, column = 12).value = str(val2)+"-"+str(val3)+" days"
		Sheet1.cell(row = 17, column = 13).value = str(val3)+"-"+str(val4)+" days"
		Sheet1.cell(row = 17, column = 14).value = str(val4)+"-"+str(val5)+" days"
		Sheet1.cell(row = 17, column = 15).value = "Over "+str(val5)+" days"
		Sheet1.cell(row = 17, column = 16).value = "Check"
		# Sheet1.cell(row = 17, column = 17).value = "Random for IPE"




		Sheet1.cell(row = 5, column = 5).value = "Account"
		Sheet1.cell(row = 5, column = 6).value = "As per Detail"
		Sheet1.cell(row = 5, column = 7).value = "As per TB"
		Sheet1.cell(row = 5, column = 8).value = "Difference"

		#bucket zone
		Sheet1.cell(row = 12, column = 9).value = "Not due"
		Sheet1.cell(row = 12, column = 10).value = "<="+str(val1)+" days"
		Sheet1.cell(row = 12, column = 11).value = str(val1)+"-"+str(val2)+" days"
		Sheet1.cell(row = 12, column = 12).value = str(val2)+"-"+str(val3)+" days"
		Sheet1.cell(row = 12, column = 13).value = str(val3)+"-"+str(val4)+" days"
		Sheet1.cell(row = 12, column = 14).value = str(val4)+"-"+str(val5)+" days"
		Sheet1.cell(row = 12, column = 15).value = "Over "+str(val5)+" days"
		Sheet1.cell(row = 12, column = 16).value = "Total"

		#suma procentelor
		Sheet1.cell(row = 13, column = 16).value = "=SUM(I13:O13)".format(18)
		Sheet1.cell(row=13, column=16).number_format = '#,##0_);(#,##0)'

		#procent din total
		Sheet1.cell(row = 14, column = 9).value = "=I13/P13".format(12)
		Sheet1.cell(row = 14, column = 10).value = "=J13/P13".format(12)
		Sheet1.cell(row = 14, column = 11).value = "=K13/P13".format(12)
		Sheet1.cell(row = 14, column = 12).value = "=L13/P13".format(12)
		Sheet1.cell(row = 14, column = 13).value = "=M13/P13".format(12)
		Sheet1.cell(row = 14, column = 14).value = "=N13/P13".format(12)
		Sheet1.cell(row = 14, column = 15).value = "=O13/P13".format(12)

		#prepared and date
		Sheet1.cell(row = 1, column = 15).value = "Processed by"
		Sheet1.cell(row = 1, column = 16).value = preparedBy1
		Sheet1.cell(row = 2, column = 15).value = "Date"
		Sheet1.cell(row = 2, column = 16).value = datePrepared1
		Sheet1.cell(row = 2, column = 16).number_format = 'mm/dd/yyyy'
		Sheet1.cell(row = 3, column = 15).value = "Refference"
		Sheet1.cell(row = 3, column = 16).value = refference1

		for row in ws.iter_rows():
			for cell in row:
				if cell.value == "Client code":
					row_customer = cell.row
					column_client = cell.column
					lun = len(ws[cell.column])
		try:
			clientCode = [b.value for b in ws[column_client][row_customer:lun]]
		except:
			flash("Please insert correct header for Client code in Aging file")
			return render_template("index.html")
		for row in ws.iter_rows():
			for cell in row:
				if cell.value == "Client":
					row_customer = cell.row
					column_customer = cell.column
					lun = len(ws[cell.column])
		try:
			clientName = [b.value for b in ws[column_customer][row_customer:lun]]
		except:
			flash("Please insert correct header for Client in Aging file")
			return render_template("index.html")
		for row in ws.iter_rows():
			for cell in row:
				if cell.value == "Account":
					row_customer = cell.row
					column_account = cell.column
					lun = len(ws[cell.column])
		try:
			clientAccount = [b.value for b in ws[column_account][row_customer:lun]]
		except:
			flash("Please insert correct header for Account in Aging file")
			return render_template("index.html")
			# messagebox.showerror('Error', 'The "Account" header in details file was not created properly')

		for row in ws.iter_rows():
			for cell in row:
				if cell.value == "Due date":
					row_customer = cell.row
					column_due = cell.column
					lun = len(ws[cell.column])
		try:
			dueDate = [b.value for b in ws[column_due][row_customer:lun]]
		except:
			flash("Please insert correct header for Due date in Aging file")
			return render_template("index.html")
			# messagebox.showerror('Error', 'The "Due date" header in details file was not created properly')

		for row in ws.iter_rows():
			for cell in row:
				if cell.value == "Invoice no":
					row_customer = cell.row
					column_invno = cell.column
					lun = len(ws[cell.column])
		try:
			invoiceNo = [b.value for b in ws[column_invno][row_customer:lun]]
		except:
			flash("Please insert correct header for Invoice no in Aging file")
			return render_template("index.html")
			# messagebox.showerror('Error', 'The "Invoice no" header in details file was not created properly')

		for row in ws.iter_rows():
			for cell in row:
				if cell.value == "Amount in LC":
					row_customer = cell.row
					column_amountinlc = cell.column
					lun = len(ws[cell.column])
		try:
			amtLC = [b.value for b in ws[column_amountinlc][row_customer:lun]]
		except:
			flash("Please insert correct header for Amount in LC in Aging file")
			return render_template("index.html")
			# messagebox.showerror('Error', 'The "Amount in LC" header in details file was not created properly')

		for row in ws.iter_rows():
			for cell in row:
				if cell.value == "Invoice value":
					row_customer = cell.row
					column_invoicev = cell.column
					lun = len(ws[cell.column])
		try:
			invoiceVal = [b.value for b in ws[column_invoicev][row_customer:lun]]
		except:
			flash("Please insert correct header for Invoice value in Aging file")
			return render_template("index.html")
			# messagebox.showerror('Error', 'The "Invoice value" header in details file was not created properly')


		# starting_row=18
		# for i in invoiceNo:
		# 	Sheet1.cell(row= starting_row, column= 17).value="=CHAR(RANDBETWEEN(87,88))"
		# 	starting_row=starting_row+1 


		# if(len(invoiceNo)<250):
		# 	starting_row=18
		# 	for i in range(0,10):
		# 		Sheet1.cell(row= i+starting_row, column= 17).value=choice(invoiceNo)
				

			

		for i in range(0, len(clientName)):
			Sheet1.cell(row = 18 + i, column = 2).value = clientName[i]

		for i in range(0, len(clientCode)):
			Sheet1.cell(row = 18 + i, column = 1).value = clientCode[i]

		for i in range(0, len(clientAccount)):
			Sheet1.cell(row = 18 + i, column = 3).value = clientAccount[i]

		for i in range(0, len(dueDate)):
			Sheet1.cell(row = 18 + i, column = 4).value = dueDate[i]

		for i in range(0, len(invoiceNo)):
			Sheet1.cell(row = 18 + i, column = 5).value = invoiceNo[i]

		for i in range(0, len(amtLC)):
			Sheet1.cell(row = 18 + i, column = 6).value = amtLC[i]

		for i in range(0, len(invoiceVal)):
			Sheet1.cell(row = 18 + i, column = 7).value = invoiceVal[i]

		# print("=IF(AND(H{0}<=" + str(val1) + ",H{0}>" + str(val0) + "),F{0},0)")

		for i in range(0, len(dueDate)):
			Sheet1.cell(row = 18 + i, column = 8).value = "=$A$13-D"+str(18 + i)
			Sheet1.cell(row = 18 + i, column = 9).value = "=IF(H"+str(18 + i)+"<="+str(val0)+",F"+str(18 + i)+",0)".format(18 + i)
			Sheet1.cell(row = 18 + i, column = 10).value = "=IF(AND(H"+str(18 + i)+"<="+str(val1)+",H"+str(18 + i)+">"+str(val0)+"),F"+str(18 + i)+",0)".format(18 + i)
			Sheet1.cell(row = 18 + i, column = 11).value = "=IF(AND(H"+str(18 + i)+"<="+str(val2)+",H"+str(18 + i)+">"+str(val1)+"),F"+str(18 + i)+",0)".format(18 + i)
			Sheet1.cell(row = 18 + i, column = 12).value = "=IF(AND(H"+str(18 + i)+"<="+str(val3)+",H"+str(18 + i)+">"+str(val2)+"),F"+str(18 + i)+",0)".format(18 + i)
			Sheet1.cell(row = 18 + i, column = 13).value = "=IF(AND(H"+str(18 + i)+"<="+str(val4)+",H"+str(18 + i)+">"+str(val3)+"),F"+str(18 + i)+",0)".format(18 + i)
			Sheet1.cell(row = 18 + i, column = 14).value = "=IF(AND(H"+str(18 + i)+"<="+str(val5)+",H"+str(18 + i)+">"+str(val4)+"),F"+str(18 + i)+",0)".format(18 + i)
			Sheet1.cell(row = 18 + i, column = 15).value = "=IF(H"+str(18 + i)+">"+str(val5)+",F"+str(18 + i)+",0)".format(18 + i)
			Sheet1.cell(row = 18 + i, column = 16).value = "=F"+str(18 + i)+"-SUM(I"+str(18+i)+":O"+str(18+i)+")"

			Sheet1.cell(row=18 + i, column=9).number_format = '#,##0_);(#,##0)'
			Sheet1.cell(row=18 + i, column=10).number_format = '#,##0_);(#,##0)'
			Sheet1.cell(row=18 + i, column=11).number_format = '#,##0_);(#,##0)'
			Sheet1.cell(row=18 + i, column=12).number_format = '#,##0_);(#,##0)'
			Sheet1.cell(row=18 + i, column=13).number_format = '#,##0_);(#,##0)'
			Sheet1.cell(row=18 + i, column=14).number_format = '#,##0_);(#,##0)'
			Sheet1.cell(row=18 + i, column=15).number_format = '#,##0_);(#,##0)'
			Sheet1.cell(row=18 + i, column=16).number_format = '#,##0_);(#,##0)'

			#bucket zone formulas
			Sheet1.cell(row = 13, column = 9).value = "=SUM(I18:I{0})".format(18 + i)
			Sheet1.cell(row = 13, column = 10).value = "=SUM(J18:J{0})".format(18 + i)
			Sheet1.cell(row = 13, column = 11).value = "=SUM(K18:K{0})".format(18 + i)
			Sheet1.cell(row = 13, column = 12).value = "=SUM(L18:L{0})".format(18 + i)
			Sheet1.cell(row = 13, column = 13).value = "=SUM(M18:M{0})".format(18 + i)
			Sheet1.cell(row = 13, column = 14).value = "=SUM(N18:N{0})".format(18 + i)
			Sheet1.cell(row = 13, column = 15).value = "=SUM(O18:O{0})".format(18 + i)

			#FORMAT
			for row in Sheet1["I13:O13"]:
				for cell in row:
					cell.number_format = '#,##0_);(#,##0)'
					cell.font = font_worksheet

			#reconciliation zone
			Sheet1.cell(row = 6, column = 5).value = "411"
			Sheet1.cell(row = 6, column = 6).value = "=P13".format(18 + i)
			Sheet1.cell(row = 6, column = 7).value ="=SUMIF('TB Robot'!A:A,""411"",'TB Robot'!H:H)".format(18 + i)
			Sheet1.cell(row = 6, column = 8).value ="=F6-G6".format(18 + i)

			Sheet1.cell(row=6, column=6).number_format = '#,##0_);(#,##0)'
			Sheet1.cell(row=6, column=7).number_format = '#,##0_);(#,##0)'
			Sheet1.cell(row=6, column=8).number_format = '#,##0_);(#,##0)'

			Sheet1.cell(row=6, column=5).font = font_worksheet
			Sheet1.cell(row=6, column=6).font = font_worksheet
			Sheet1.cell(row=6, column=7).font = font_worksheet
			Sheet1.cell(row=6, column=8).font = check_font

		for i in range(0, len(clientName)):
			Sheet1.cell(row = 18 + i, column = 2).font = font_worksheet

		for i in range(0, len(clientCode)):
			Sheet1.cell(row = 18 + i, column = 1).font = font_worksheet

		for i in range(0, len(clientAccount)):
			Sheet1.cell(row = 18 + i, column = 3).font = font_worksheet

		for i in range(0, len(dueDate)):
			Sheet1.cell(row = 18 + i, column = 4).font = font_worksheet

		for i in range(0, len(invoiceNo)):
			Sheet1.cell(row = 18 + i, column = 5).font = font_worksheet

		for i in range(0, len(amtLC)):
			Sheet1.cell(row = 18 + i, column = 6).number_format = '#,##0_);(#,##0)'

		for i in range(0, len(amtLC)):
			Sheet1.cell(row = 18 + i, column = 6).font = font_worksheet

		for i in range(0, len(invoiceVal)):
			Sheet1.cell(row = 18 + i, column = 7).number_format = '#,##0_);(#,##0)'

		for i in range(0, len(invoiceVal)):
			Sheet1.cell(row = 18 + i, column = 7).font = font_worksheet

		for i in range(0, len(dueDate)):
			Sheet1.cell(row = 18 + i, column = 8).font = font_worksheet
			Sheet1.cell(row = 18 + i, column = 9).font = font_worksheet
			Sheet1.cell(row = 18 + i, column = 10).font = font_worksheet
			Sheet1.cell(row = 18 + i, column = 11).font = font_worksheet
			Sheet1.cell(row = 18 + i, column = 12).font = font_worksheet
			Sheet1.cell(row = 18 + i, column = 13).font = font_worksheet
			Sheet1.cell(row = 18 + i, column = 14).font = font_worksheet
			Sheet1.cell(row = 18 + i, column = 15).font = font_worksheet
			Sheet1.cell(row = 18 + i, column = 16).font = font_worksheet

		Sheet1.freeze_panes = 'C18'

		Sheet1.column_dimensions['A'].width = 10
		Sheet1.column_dimensions['B'].width = 20
		Sheet1.column_dimensions['C'].width = 11
		Sheet1.column_dimensions['D'].width = 12
		Sheet1.column_dimensions['E'].width = 14
		Sheet1.column_dimensions['F'].width = 12
		Sheet1.column_dimensions['G'].width = 10
		Sheet1.column_dimensions['H'].width = 9
		Sheet1.column_dimensions['I'].width = 9
		Sheet1.column_dimensions['J'].width = 9
		Sheet1.column_dimensions['K'].width = 9
		Sheet1.column_dimensions['L'].width = 9
		Sheet1.column_dimensions['M'].width = 9
		Sheet1.column_dimensions['N'].width = 9
		Sheet1.column_dimensions['O'].width = 12
		Sheet1.column_dimensions['P'].width = 15

		for cell in Sheet1["D"]:
			cell.number_format = "mm/dd/yyyy"

		Sheet1.auto_filter.ref = "A17:P17"

		#FONT / FORMAT
		#header
		Sheet1.cell(row=1, column=1).font = ft1
		Sheet1.cell(row=1, column=2).font = ft1
		Sheet1.cell(row=2, column=1).font = ft1
		Sheet1.cell(row=2, column=2).font = ft1

		Sheet1.cell(row=1, column=15).font = ft1
		Sheet1.cell(row=1, column=16).font = ft1
		Sheet1.cell(row=2, column=15).font = ft1
		Sheet1.cell(row=2, column=16).font = ft1
		Sheet1.cell(row=3, column=15).font = ft1
		Sheet1.cell(row=3, column=16).font = check_font

		Sheet1.cell(row=5, column=5).font = cap_tabel
		Sheet1.cell(row=5, column=6).font = cap_tabel
		Sheet1.cell(row=5, column=7).font = cap_tabel
		Sheet1.cell(row=5, column=8).font = check_font

		# test name
		Sheet1.cell(row=4, column=1).font = f_testname
		Sheet1.cell(row=12, column=1).font = f_testname
		Sheet1.cell(row=13, column=1).font = ft1

		Sheet1.cell(row=13, column=16).font = ft1

		Sheet1.cell(row=16, column=8).font = font_worksheet
		Sheet1.cell(row=17, column=16).font = check_font

		for row in Sheet1["A17:O17"]:
			for cell in row:
				cell.font = cap_tabel

		for row in Sheet1["A17:G17"]:
			for cell in row:
				cell.fill = cap_tabel_color_PBC

		for row in Sheet1["H17:P17"]:
			for cell in row:
				cell.fill = cap_tabel_color_GT

		Sheet1.cell(row=5, column=5).fill = cap_tabel_color_GT
		Sheet1.cell(row=5, column=6).fill = cap_tabel_color_GT
		Sheet1.cell(row=5, column=7).fill = cap_tabel_color_GT
		Sheet1.cell(row=5, column=8).fill = cap_tabel_color_GT

		# bucket zone
		for row in Sheet1["I12:P12"]:
			for cell in row:
				cell.font = cap_tabel

		for row in Sheet1["I12:P12"]:
			for cell in row:
				cell.fill = cap_tabel_color_GT

		# suma procentelor
		Sheet1.cell(row=13, column=16).font = ft1
		Sheet1.cell(row=13, column=16).number_format = '#,##0_);(#,##0)'

		# procent din total
		for row in Sheet1['I14:O14']:
			for cell in row:
				cell.style='Percent'

		for row in Sheet1['I14:O14']:
			for cell in row:
				cell.font = fprocentaj

		# prepared and date
		Sheet1.cell(row=1, column=15).font = ft1
		Sheet1.cell(row=1, column=16).font = ft1
		Sheet1.cell(row=2, column=15).font = ft1
		Sheet1.cell(row=2, column=16).font = ft1
		Sheet1.cell(row=3, column=15).font = ft1
		Sheet1.cell(row=3, column=16).font = check_font

		Sheet11 = wb.create_sheet("IPE")
		Sheet11.sheet_view.showGridLines = False
		
		# try:
		details = openpyxl.load_workbook(file_Details, data_only = True)
		# except:
		# 	None
		ws = details.active

		#header
		Sheet11.cell(row = 1, column = 1).value = "Client name:"
		Sheet11.cell(row = 1, column = 2).value = clientname1
		Sheet11.cell(row = 2, column = 1).value = "Period end:"
		Sheet11.cell(row = 2, column = 2).value = yearEnd1
		Sheet11.cell(row = 2, column = 2).number_format = 'mm/dd/yyyy'

		#test name
		# Sheet11.cell(row = 4, column = 1).value = "Receivables ageing recomputation"
		# Sheet11.cell(row = 12, column = 1).value = "YE Date"
		# Sheet11.cell(row = 13, column = 1).value = yearEnd1
		# Sheet11.cell(row = 13, column = 1).number_format = 'mm/dd/yyyy'

		#testing table
		# Sheet11.cell(row = 16, column = 1).value = "As per Detail"
		# Sheet11.cell(row = 17, column = 1).value = "Client code"
		# Sheet11.cell(row = 17, column = 2).value = "Client"
		# Sheet11.cell(row = 17, column = 3).value = "Account"
		# Sheet11.cell(row = 17, column = 4).value = "Due date"
		# Sheet11.cell(row = 17, column = 5).value = "Invoice number"
		# Sheet11.cell(row = 17, column = 6).value = "Amount in LC"
		# Sheet11.cell(row = 16, column = 8).value = "As per Nexia"
		# Sheet11.cell(row = 17, column = 7).value = "Invoice value"
		# Sheet11.cell(row = 17, column = 8).value = "No of days"
		# Sheet11.cell(row = 17, column = 9).value = "Not due"
		# Sheet11.cell(row = 17, column = 10).value = "<="+str(val1)+" days"
		# Sheet11.cell(row = 17, column = 11).value = str(val1)+"-"+str(val2)+" days"
		# Sheet11.cell(row = 17, column = 12).value = str(val2)+"-"+str(val3)+" days"
		# Sheet11.cell(row = 17, column = 13).value = str(val3)+"-"+str(val4)+" days"
		# Sheet11.cell(row = 17, column = 14).value = str(val4)+"-"+str(val5)+" days"
		# Sheet11.cell(row = 17, column = 15).value = "Over "+str(val5)+" days"
		# Sheet11.cell(row = 17, column = 16).value = "Check"
		Sheet11.cell(row = 6, column = 1).value = "Random invoice number for IPE"
		Sheet11.cell(row = 6, column = 2).value = "Client code"
		Sheet11.cell(row = 6, column = 3).value = "Client"
		Sheet11.cell(row = 6, column = 4).value = "Account"
		Sheet11.cell(row = 6, column = 5).value = "Due date"
		Sheet11.cell(row = 6, column = 6).value = "Amount in LC"
		Sheet11.cell(row = 6, column = 7).value = "Invoice value"
		# Sheet11.cell(row = 6, column = 8).value = "Amount in FC"

		Sheet11.cell(row = 6, column = 1).fill=cap_tabel_color_GT
		Sheet11.cell(row = 6, column = 2).fill=cap_tabel_color_GT
		Sheet11.cell(row = 6, column = 3).fill=cap_tabel_color_GT
		Sheet11.cell(row = 6, column = 4).fill=cap_tabel_color_GT
		Sheet11.cell(row = 6, column = 5).fill=cap_tabel_color_GT
		Sheet11.cell(row = 6, column = 6).fill=cap_tabel_color_GT
		Sheet11.cell(row = 6, column = 7).fill=cap_tabel_color_GT
		# Sheet11.cell(row = 6, column = 8).fill=cap_tabel_color_GT



		# Sheet11.cell(row = 5, column = 5).value = "Account"
		# Sheet11.cell(row = 5, column = 6).value = "As per Detail"
		# Sheet11.cell(row = 5, column = 7).value = "As per TB"
		# Sheet11.cell(row = 5, column = 8).value = "Difference"

		#bucket zone
		# Sheet11.cell(row = 12, column = 9).value = "Not due"
		# Sheet11.cell(row = 12, column = 10).value = "<="+str(val1)+" days"
		# Sheet11.cell(row = 12, column = 11).value = str(val1)+"-"+str(val2)+" days"
		# Sheet11.cell(row = 12, column = 12).value = str(val2)+"-"+str(val3)+" days"
		# Sheet11.cell(row = 12, column = 13).value = str(val3)+"-"+str(val4)+" days"
		# Sheet11.cell(row = 12, column = 14).value = str(val4)+"-"+str(val5)+" days"
		# Sheet11.cell(row = 12, column = 15).value = "Over "+str(val5)+" days"
		# Sheet11.cell(row = 12, column = 16).value = "Total"

		#suma procentelor
		# Sheet11.cell(row = 13, column = 16).value = "=SUM(I13:O13)".format(18)
		# Sheet11.cell(row=13, column=16).number_format = '#,##0_);(#,##0)'

		#procent din total
		# Sheet11.cell(row = 14, column = 9).value = "=I13/P13".format(12)
		# Sheet11.cell(row = 14, column = 10).value = "=J13/P13".format(12)
		# Sheet11.cell(row = 14, column = 11).value = "=K13/P13".format(12)
		# Sheet11.cell(row = 14, column = 12).value = "=L13/P13".format(12)
		# Sheet11.cell(row = 14, column = 13).value = "=M13/P13".format(12)
		# Sheet11.cell(row = 14, column = 14).value = "=N13/P13".format(12)
		# Sheet11.cell(row = 14, column = 15).value = "=O13/P13".format(12)

		#prepared and date
		Sheet11.cell(row = 1, column = 15).value = "Processed by"
		Sheet11.cell(row = 1, column = 16).value = preparedBy1
		Sheet11.cell(row = 2, column = 15).value = "Date"
		Sheet11.cell(row = 2, column = 16).value = datePrepared1
		Sheet11.cell(row = 2, column = 16).number_format = 'mm/dd/yyyy'
		Sheet11.cell(row = 3, column = 15).value = "Refference"
		Sheet11.cell(row = 3, column = 16).value = refference1

		for row in ws.iter_rows():
			for cell in row:
				if cell.value == "Client code":
					row_customer = cell.row
					column_client = cell.column
					lun = len(ws[cell.column])
		try:
			clientCode = [b.value for b in ws[column_client][row_customer:lun]]
		except:
			flash("Please insert correct header for Client code in Aging file")
			return render_template("index.html")
		for row in ws.iter_rows():
			for cell in row:
				if cell.value == "Client":
					row_customer = cell.row
					column_customer = cell.column
					lun = len(ws[cell.column])
		try:
			clientName = [b.value for b in ws[column_customer][row_customer:lun]]
		except:
			flash("Please insert correct header for Client in Aging file")
			return render_template("index.html")
		for row in ws.iter_rows():
			for cell in row:
				if cell.value == "Account":
					row_customer = cell.row
					column_account = cell.column
					lun = len(ws[cell.column])
		try:
			clientAccount = [b.value for b in ws[column_account][row_customer:lun]]
		except:
			flash("Please insert correct header for Account in Aging file")
			return render_template("index.html")
			# messagebox.showerror('Error', 'The "Account" header in details file was not created properly')

		for row in ws.iter_rows():
			for cell in row:
				if cell.value == "Due date":
					row_customer = cell.row
					column_due = cell.column
					lun = len(ws[cell.column])
		try:
			dueDate = [b.value for b in ws[column_due][row_customer:lun]]
		except:
			flash("Please insert correct header for Due date in Aging file")
			return render_template("index.html")
			# messagebox.showerror('Error', 'The "Due date" header in details file was not created properly')

		for row in ws.iter_rows():
			for cell in row:
				if cell.value == "Invoice no":
					row_customer = cell.row
					column_invno = cell.column
					lun = len(ws[cell.column])
		try:
			invoiceNo = [b.value for b in ws[column_invno][row_customer:lun]]
		except:
			flash("Please insert correct header for Invoice no in Aging file")
			return render_template("index.html")
			# messagebox.showerror('Error', 'The "Invoice no" header in details file was not created properly')

		for row in ws.iter_rows():
			for cell in row:
				if cell.value == "Amount in LC":
					row_customer = cell.row
					column_amountinlc = cell.column
					lun = len(ws[cell.column])
		try:
			amtLC = [b.value for b in ws[column_amountinlc][row_customer:lun]]
		except:
			flash("Please insert correct header for Amount in LC in Aging file")
			return render_template("index.html")
			# messagebox.showerror('Error', 'The "Amount in LC" header in details file was not created properly')

		for row in ws.iter_rows():
			for cell in row:
				if cell.value == "Invoice value":
					row_customer = cell.row
					column_invoicev = cell.column
					lun = len(ws[cell.column])
		try:
			invoiceVal = [b.value for b in ws[column_invoicev][row_customer:lun]]
		except:
			flash("Please insert correct header for Invoice value in Aging file")
			return render_template("index.html")
			# messagebox.showerror('Error', 'The "Invoice value" header in details file was not created properly')


		# for row in ws.iter_rows():
		# 	for cell in row:
		# 		if cell.value == "Amount in FC":
		# 			row_customer = cell.row
		# 			column_amountinfc = cell.column
		# 			lun = len(ws[cell.column])
		# try:
		# 	amtFC = [b.value for b in ws[column_amountinfc][row_customer:lun]]
		# except:
		# 	flash("Please insert correct header for Amount in FC in Aging file")
		# 	return render_template("index.html")
			# messagebox.showerror('Error', 'The "Amount in LC" header in details file was not created properly')

		# starting_row=18
		# for i in invoiceNo:
		# 	Sheet1.cell(row= starting_row, column= 17).value="=CHAR(RANDBETWEEN(87,88))"
		# 	starting_ro	w=starting_row+1 


		if(len(invoiceNo)<250):
			starting_row=7
			for i in range(0,10):
				Sheet11.cell(row= i+starting_row, column= 1).value=choice(invoiceNo)
				Sheet11.cell(row= 7, column= 2).value="=_xlfn.XLOOKUP(A7,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 8, column= 2).value="=_xlfn.XLOOKUP(A8,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 9, column= 2).value="=_xlfn.XLOOKUP(A9,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 10, column= 2).value="=_xlfn.XLOOKUP(A10,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 11, column= 2).value="=_xlfn.XLOOKUP(A11,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 12, column= 2).value="=_xlfn.XLOOKUP(A12,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 13, column= 2).value="=_xlfn.XLOOKUP(A13,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 14, column= 2).value="=_xlfn.XLOOKUP(A14,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 15, column= 2).value="=_xlfn.XLOOKUP(A15,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 16, column= 2).value="=_xlfn.XLOOKUP(A16,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"

				Sheet11.cell(row= 7, column= 3).value="=_xlfn.XLOOKUP(A7,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 8, column= 3).value="=_xlfn.XLOOKUP(A8,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 9, column= 3).value="=_xlfn.XLOOKUP(A9,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 10, column= 3).value="=_xlfn.XLOOKUP(A10,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 11, column= 3).value="=_xlfn.XLOOKUP(A11,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 12, column= 3).value="=_xlfn.XLOOKUP(A12,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 13, column= 3).value="=_xlfn.XLOOKUP(A13,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 14, column= 3).value="=_xlfn.XLOOKUP(A14,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 15, column= 3).value="=_xlfn.XLOOKUP(A15,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 16, column= 3).value="=_xlfn.XLOOKUP(A16,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"

				Sheet11.cell(row= 7, column= 4).value="=_xlfn.XLOOKUP(A7,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 8, column= 4).value="=_xlfn.XLOOKUP(A8,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 9, column= 4).value="=_xlfn.XLOOKUP(A9,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 10, column= 4).value="=_xlfn.XLOOKUP(A10,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 11, column= 4).value="=_xlfn.XLOOKUP(A11,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 12, column= 4).value="=_xlfn.XLOOKUP(A12,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 13, column= 4).value="=_xlfn.XLOOKUP(A13,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 14, column= 4).value="=_xlfn.XLOOKUP(A14,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 15, column= 4).value="=_xlfn.XLOOKUP(A15,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 16, column= 4).value="=_xlfn.XLOOKUP(A16,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"

				Sheet11.cell(row= 7, column= 5).value="=_xlfn.XLOOKUP(A7,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!D:D,0,0,1)"
				Sheet11.cell(row= 8, column= 5).value="=_xlfn.XLOOKUP(A8,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 9, column= 5).value="=_xlfn.XLOOKUP(A9,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 10, column= 5).value="=_xlfn.XLOOKUP(A10,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 11, column= 5).value="=_xlfn.XLOOKUP(A11,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 12, column= 5).value="=_xlfn.XLOOKUP(A12,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 13, column= 5).value="=_xlfn.XLOOKUP(A13,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 14, column= 5).value="=_xlfn.XLOOKUP(A14,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 15, column= 5).value="=_xlfn.XLOOKUP(A15,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 16, column= 5).value="=_xlfn.XLOOKUP(A16,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"

				Sheet11.cell(row= 7, column= 6).value="=_xlfn.XLOOKUP(A7,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 8, column= 6).value="=_xlfn.XLOOKUP(A8,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 9, column= 6).value="=_xlfn.XLOOKUP(A9,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 10, column= 6).value="=_xlfn.XLOOKUP(A10,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 11, column= 6).value="=_xlfn.XLOOKUP(A11,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 12, column= 6).value="=_xlfn.XLOOKUP(A12,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 13, column= 6).value="=_xlfn.XLOOKUP(A13,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 14, column= 6).value="=_xlfn.XLOOKUP(A14,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 15, column= 6).value="=_xlfn.XLOOKUP(A15,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 16, column= 6).value="=_xlfn.XLOOKUP(A16,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"

				Sheet11.cell(row= 7, column= 7).value="=_xlfn.XLOOKUP(A7,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 8, column= 7).value="=_xlfn.XLOOKUP(A8,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 9, column= 7).value="=_xlfn.XLOOKUP(A9,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 10, column= 7).value="=_xlfn.XLOOKUP(A10,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 11, column= 7).value="=_xlfn.XLOOKUP(A11,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 12, column= 7).value="=_xlfn.XLOOKUP(A12,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 13, column= 7).value="=_xlfn.XLOOKUP(A13,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 14, column= 7).value="=_xlfn.XLOOKUP(A14,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 15, column= 7).value="=_xlfn.XLOOKUP(A15,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 16, column= 7).value="=_xlfn.XLOOKUP(A16,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"

				# Sheet11.cell(row= 7, column= 8).value="=_xlfn.XLOOKUP(A7,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 8, column= 8).value="=_xlfn.XLOOKUP(A8,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 9, column= 8).value="=_xlfn.XLOOKUP(A9,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 10, column= 8).value="=_xlfn.XLOOKUP(A10,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 11, column= 8).value="=_xlfn.XLOOKUP(A11,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 12, column= 8).value="=_xlfn.XLOOKUP(A12,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 13, column= 8).value="=_xlfn.XLOOKUP(A13,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 14, column= 8).value="=_xlfn.XLOOKUP(A14,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 15, column= 8).value="=_xlfn.XLOOKUP(A15,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 16, column= 8).value="=_xlfn.XLOOKUP(A16,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"


				Sheet11.cell(row= 7, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 8, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 9, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 10, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 11, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 12, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 13, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 14, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 15, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 16, column= 6).number_format = '#,##0_);(#,##0)'

				Sheet11.cell(row= 7, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 8, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 9, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 10, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 11, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 12, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 13, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 14, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 15, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 16, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row=13, column=16).number_format = '#,##0_);(#,##0)'

				Sheet11.cell(row= 7, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 8, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 9, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 10, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 11, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 12, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 13, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 14, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 15, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 16, column= 8).number_format = '#,##0_);(#,##0)'
		else:
			starting_row=7		
			for i in range(0,25):
				Sheet11.cell(row= i+starting_row, column= 1).value=choice(invoiceNo)
				Sheet11.cell(row= 7, column= 2).value="=_xlfn.XLOOKUP(A7,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 8, column= 2).value="=_xlfn.XLOOKUP(A8,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 9, column= 2).value="=_xlfn.XLOOKUP(A9,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 10, column= 2).value="=_xlfn.XLOOKUP(A10,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 11, column= 2).value="=_xlfn.XLOOKUP(A11,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 12, column= 2).value="=_xlfn.XLOOKUP(A12,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 13, column= 2).value="=_xlfn.XLOOKUP(A13,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 14, column= 2).value="=_xlfn.XLOOKUP(A14,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 15, column= 2).value="=_xlfn.XLOOKUP(A15,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 16, column= 2).value="=_xlfn.XLOOKUP(A16,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 17, column= 2).value="=_xlfn.XLOOKUP(A17,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 18, column= 2).value="=_xlfn.XLOOKUP(A18,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 19, column= 2).value="=_xlfn.XLOOKUP(A19,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 20, column= 2).value="=_xlfn.XLOOKUP(A20,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 21, column= 2).value="=_xlfn.XLOOKUP(A21,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 22, column= 2).value="=_xlfn.XLOOKUP(A22,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 23, column= 2).value="=_xlfn.XLOOKUP(A23,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 24, column= 2).value="=_xlfn.XLOOKUP(A24,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 25, column= 2).value="=_xlfn.XLOOKUP(A25,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 26, column= 2).value="=_xlfn.XLOOKUP(A26,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 27, column= 2).value="=_xlfn.XLOOKUP(A27,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 28, column= 2).value="=_xlfn.XLOOKUP(A28,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 29, column= 2).value="=_xlfn.XLOOKUP(A29,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 30, column= 2).value="=_xlfn.XLOOKUP(A30,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"
				Sheet11.cell(row= 31, column= 2).value="=_xlfn.XLOOKUP(A31,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!A:A,0,0,1)"

				Sheet11.cell(row= 7, column= 3).value="=_xlfn.XLOOKUP(A7,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 8, column= 3).value="=_xlfn.XLOOKUP(A8,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 9, column= 3).value="=_xlfn.XLOOKUP(A9,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 10, column= 3).value="=_xlfn.XLOOKUP(A10,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 11, column= 3).value="=_xlfn.XLOOKUP(A11,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 12, column= 3).value="=_xlfn.XLOOKUP(A12,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 13, column= 3).value="=_xlfn.XLOOKUP(A13,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 14, column= 3).value="=_xlfn.XLOOKUP(A14,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 15, column= 3).value="=_xlfn.XLOOKUP(A15,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 16, column= 3).value="=_xlfn.XLOOKUP(A16,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 17, column= 3).value="=_xlfn.XLOOKUP(A17,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 18, column= 3).value="=_xlfn.XLOOKUP(A18,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 19, column= 3).value="=_xlfn.XLOOKUP(A19,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 20, column= 3).value="=_xlfn.XLOOKUP(A20,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 21, column= 3).value="=_xlfn.XLOOKUP(A21,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 22, column= 3).value="=_xlfn.XLOOKUP(A22,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 23, column= 3).value="=_xlfn.XLOOKUP(A23,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 24, column= 3).value="=_xlfn.XLOOKUP(A24,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 25, column= 3).value="=_xlfn.XLOOKUP(A25,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 26, column= 3).value="=_xlfn.XLOOKUP(A26,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 27, column= 3).value="=_xlfn.XLOOKUP(A27,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 28, column= 3).value="=_xlfn.XLOOKUP(A28,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 29, column= 3).value="=_xlfn.XLOOKUP(A29,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 30, column= 3).value="=_xlfn.XLOOKUP(A30,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"
				Sheet11.cell(row= 31, column= 3).value="=_xlfn.XLOOKUP(A31,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!b:b,0,0,1)"

				Sheet11.cell(row= 7, column= 4).value="=_xlfn.XLOOKUP(A7,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 8, column= 4).value="=_xlfn.XLOOKUP(A8,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 9, column= 4).value="=_xlfn.XLOOKUP(A9,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 10, column= 4).value="=_xlfn.XLOOKUP(A10,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 11, column= 4).value="=_xlfn.XLOOKUP(A11,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 12, column= 4).value="=_xlfn.XLOOKUP(A12,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 13, column= 4).value="=_xlfn.XLOOKUP(A13,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 14, column= 4).value="=_xlfn.XLOOKUP(A14,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 15, column= 4).value="=_xlfn.XLOOKUP(A15,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 16, column= 4).value="=_xlfn.XLOOKUP(A16,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 17, column= 4).value="=_xlfn.XLOOKUP(A17,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 18, column= 4).value="=_xlfn.XLOOKUP(A18,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 19, column= 4).value="=_xlfn.XLOOKUP(A19,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 20, column= 4).value="=_xlfn.XLOOKUP(A20,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 21, column= 4).value="=_xlfn.XLOOKUP(A21,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 22, column= 4).value="=_xlfn.XLOOKUP(A22,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 23, column= 4).value="=_xlfn.XLOOKUP(A23,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 24, column= 4).value="=_xlfn.XLOOKUP(A24,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 25, column= 4).value="=_xlfn.XLOOKUP(A25,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 26, column= 4).value="=_xlfn.XLOOKUP(A26,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 27, column= 4).value="=_xlfn.XLOOKUP(A27,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 28, column= 4).value="=_xlfn.XLOOKUP(A28,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 29, column= 4).value="=_xlfn.XLOOKUP(A29,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 30, column= 4).value="=_xlfn.XLOOKUP(A30,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"
				Sheet11.cell(row= 31, column= 4).value="=_xlfn.XLOOKUP(A31,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!c:c,0,0,1)"


				Sheet11.cell(row= 7, column= 5).value="=_xlfn.XLOOKUP(A7,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!D:D,0,0,1)"
				Sheet11.cell(row= 8, column= 5).value="=_xlfn.XLOOKUP(A8,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 9, column= 5).value="=_xlfn.XLOOKUP(A9,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 10, column= 5).value="=_xlfn.XLOOKUP(A10,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 11, column= 5).value="=_xlfn.XLOOKUP(A11,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 12, column= 5).value="=_xlfn.XLOOKUP(A12,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 13, column= 5).value="=_xlfn.XLOOKUP(A13,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 14, column= 5).value="=_xlfn.XLOOKUP(A14,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 15, column= 5).value="=_xlfn.XLOOKUP(A15,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 16, column= 5).value="=_xlfn.XLOOKUP(A16,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 17, column= 5).value="=_xlfn.XLOOKUP(A17,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!D:D,0,0,1)"
				Sheet11.cell(row= 18, column= 5).value="=_xlfn.XLOOKUP(A18,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 19, column= 5).value="=_xlfn.XLOOKUP(A19,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 20, column= 5).value="=_xlfn.XLOOKUP(A20,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 21, column= 5).value="=_xlfn.XLOOKUP(A21,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 22, column= 5).value="=_xlfn.XLOOKUP(A22,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 23, column= 5).value="=_xlfn.XLOOKUP(A23,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 24, column= 5).value="=_xlfn.XLOOKUP(A24,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 25, column= 5).value="=_xlfn.XLOOKUP(A25,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 26, column= 5).value="=_xlfn.XLOOKUP(A26,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 27, column= 5).value="=_xlfn.XLOOKUP(A27,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 28, column= 5).value="=_xlfn.XLOOKUP(A28,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 29, column= 5).value="=_xlfn.XLOOKUP(A29,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 30, column= 5).value="=_xlfn.XLOOKUP(A30,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"
				Sheet11.cell(row= 31, column= 5).value="=_xlfn.XLOOKUP(A31,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!d:d,0,0,1)"


				Sheet11.cell(row= 7, column= 6).value="=_xlfn.XLOOKUP(A7,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 8, column= 6).value="=_xlfn.XLOOKUP(A8,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 9, column= 6).value="=_xlfn.XLOOKUP(A9,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 10, column= 6).value="=_xlfn.XLOOKUP(A10,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 11, column= 6).value="=_xlfn.XLOOKUP(A11,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 12, column= 6).value="=_xlfn.XLOOKUP(A12,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 13, column= 6).value="=_xlfn.XLOOKUP(A13,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 14, column= 6).value="=_xlfn.XLOOKUP(A14,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 15, column= 6).value="=_xlfn.XLOOKUP(A15,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 16, column= 6).value="=_xlfn.XLOOKUP(A16,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 17, column= 6).value="=_xlfn.XLOOKUP(A17,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 18, column= 6).value="=_xlfn.XLOOKUP(A18,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 19, column= 6).value="=_xlfn.XLOOKUP(A19,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 20, column= 6).value="=_xlfn.XLOOKUP(A20,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 21, column= 6).value="=_xlfn.XLOOKUP(A21,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 22, column= 6).value="=_xlfn.XLOOKUP(A22,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 23, column= 6).value="=_xlfn.XLOOKUP(A23,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 24, column= 6).value="=_xlfn.XLOOKUP(A24,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 25, column= 6).value="=_xlfn.XLOOKUP(A25,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 26, column= 6).value="=_xlfn.XLOOKUP(A26,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 27, column= 6).value="=_xlfn.XLOOKUP(A27,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 28, column= 6).value="=_xlfn.XLOOKUP(A28,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 29, column= 6).value="=_xlfn.XLOOKUP(A29,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 30, column= 6).value="=_xlfn.XLOOKUP(A30,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"
				Sheet11.cell(row= 31, column= 6).value="=_xlfn.XLOOKUP(A31,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!f:f,0,0,1)"


				Sheet11.cell(row= 7, column= 7).value="=_xlfn.XLOOKUP(A7,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 8, column= 7).value="=_xlfn.XLOOKUP(A8,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 9, column= 7).value="=_xlfn.XLOOKUP(A9,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 10, column= 7).value="=_xlfn.XLOOKUP(A10,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 11, column= 7).value="=_xlfn.XLOOKUP(A11,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 12, column= 7).value="=_xlfn.XLOOKUP(A12,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 13, column= 7).value="=_xlfn.XLOOKUP(A13,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 14, column= 7).value="=_xlfn.XLOOKUP(A14,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 15, column= 7).value="=_xlfn.XLOOKUP(A15,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 16, column= 7).value="=_xlfn.XLOOKUP(A16,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 17, column= 7).value="=_xlfn.XLOOKUP(A17,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 18, column= 7).value="=_xlfn.XLOOKUP(A18,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 19, column= 7).value="=_xlfn.XLOOKUP(A19,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 20, column= 7).value="=_xlfn.XLOOKUP(A20,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 21, column= 7).value="=_xlfn.XLOOKUP(A21,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 22, column= 7).value="=_xlfn.XLOOKUP(A22,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 23, column= 7).value="=_xlfn.XLOOKUP(A23,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 24, column= 7).value="=_xlfn.XLOOKUP(A24,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 25, column= 7).value="=_xlfn.XLOOKUP(A25,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 26, column= 7).value="=_xlfn.XLOOKUP(A26,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 27, column= 7).value="=_xlfn.XLOOKUP(A27,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 28, column= 7).value="=_xlfn.XLOOKUP(A28,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 29, column= 7).value="=_xlfn.XLOOKUP(A29,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 30, column= 7).value="=_xlfn.XLOOKUP(A30,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"
				Sheet11.cell(row= 31, column= 7).value="=_xlfn.XLOOKUP(A31,'TR Aging Invoice level'!E:E,'TR Aging Invoice level'!g:g,0,0,1)"

				# Sheet11.cell(row= 7, column= 8).value="=_xlfn.XLOOKUP(A7,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 8, column= 8).value="=_xlfn.XLOOKUP(A8,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 9, column= 8).value="=_xlfn.XLOOKUP(A9,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 10, column= 8).value="=_xlfn.XLOOKUP(A10,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 11, column= 8).value="=_xlfn.XLOOKUP(A11,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 12, column= 8).value="=_xlfn.XLOOKUP(A12,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 13, column= 8).value="=_xlfn.XLOOKUP(A13,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 14, column= 8).value="=_xlfn.XLOOKUP(A14,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 15, column= 8).value="=_xlfn.XLOOKUP(A15,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 16, column= 8).value="=_xlfn.XLOOKUP(A16,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 17, column= 8).value="=_xlfn.XLOOKUP(A17,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 18, column= 8).value="=_xlfn.XLOOKUP(A18,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 19, column= 8).value="=_xlfn.XLOOKUP(A19,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 20, column= 8).value="=_xlfn.XLOOKUP(A20,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 21, column= 8).value="=_xlfn.XLOOKUP(A21,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 22, column= 8).value="=_xlfn.XLOOKUP(A22,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 23, column= 8).value="=_xlfn.XLOOKUP(A23,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 24, column= 8).value="=_xlfn.XLOOKUP(A24,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 25, column= 8).value="=_xlfn.XLOOKUP(A25,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 26, column= 8).value="=_xlfn.XLOOKUP(A26,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 27, column= 8).value="=_xlfn.XLOOKUP(A27,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 28, column= 8).value="=_xlfn.XLOOKUP(A28,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 29, column= 8).value="=_xlfn.XLOOKUP(A29,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 30, column= 8).value="=_xlfn.XLOOKUP(A30,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"
				# Sheet11.cell(row= 31, column= 8).value="=_xlfn.XLOOKUP(A31,'Details PBC'!E:E,'Details PBC'!H:H,0,0,1)"


				Sheet11.cell(row= 7, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 8, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 9, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 10, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 11, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 12, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 13, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 14, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 15, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 16, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 17, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 18, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 19, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 20, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 21, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 22, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 23, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 24, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 25, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 26, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 27, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 28, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 29, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 30, column= 6).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 31, column= 6).number_format = '#,##0_);(#,##0)'


				Sheet11.cell(row= 7, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 8, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 9, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 10, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 11, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 12, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 13, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 14, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 15, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 16, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row=13, column=16).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 17, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 18, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 19, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 20, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 21, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 22, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 23, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 24, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 25, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 26, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 27, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 28, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 29, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 30, column= 7).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 31, column= 7).number_format = '#,##0_);(#,##0)'
			

			
				Sheet11.cell(row= 7, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 8, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 9, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 10, column=87).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 11, column=8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 12, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 13, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 14, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 15, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 16, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row=13, column=16).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 17, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 18, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 19, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 20, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 21, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 22, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 23, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 24, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 25, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 26, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 27, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 28, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 29, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 30, column= 8).number_format = '#,##0_);(#,##0)'
				Sheet11.cell(row= 31, column= 8).number_format = '#,##0_);(#,##0)'
			

		# for i in range(0, len(clientName)):
		# 	Sheet11.cell(row = 18 + i, column = 2).value = clientName[i]

		# for i in range(0, len(clientCode)):
		# 	Sheet11.cell(row = 18 + i, column = 1).value = clientCode[i]

		# for i in range(0, len(clientAccount)):
		# 	Sheet11.cell(row = 18 + i, column = 3).value = clientAccount[i]

		# for i in range(0, len(dueDate)):
		# 	Sheet11.cell(row = 18 + i, column = 4).value = dueDate[i]

		# for i in range(0, len(invoiceNo)):
		# 	Sheet11.cell(row = 18 + i, column = 5).value = invoiceNo[i]

		# for i in range(0, len(amtFC)):
		# 	Sheet11.cell(row = 7 + i, column = 8).value = amtFC[i]

		# for i in range(0, len(invoiceVal)):
		# 	Sheet11.cell(row = 18 + i, column = 7).value = invoiceVal[i]

		# print("=IF(AND(H{0}<=" + str(val1) + ",H{0}>" + str(val0) + "),F{0},0)")

		# for i in range(0, len(dueDate)):
		# 	Sheet11.cell(row = 18 + i, column = 8).value = "=$A$13-D"+str(18 + i)
		# 	Sheet11.cell(row = 18 + i, column = 9).value = "=IF(H"+str(18 + i)+"<="+str(val0)+",F"+str(18 + i)+",0)".format(18 + i)
		# 	Sheet11.cell(row = 18 + i, column = 10).value = "=IF(AND(H"+str(18 + i)+"<="+str(val1)+",H"+str(18 + i)+">"+str(val0)+"),F"+str(18 + i)+",0)".format(18 + i)
		# 	Sheet11.cell(row = 18 + i, column = 11).value = "=IF(AND(H"+str(18 + i)+"<="+str(val2)+",H"+str(18 + i)+">"+str(val1)+"),F"+str(18 + i)+",0)".format(18 + i)
		# 	Sheet11.cell(row = 18 + i, column = 12).value = "=IF(AND(H"+str(18 + i)+"<="+str(val3)+",H"+str(18 + i)+">"+str(val2)+"),F"+str(18 + i)+",0)".format(18 + i)
		# 	Sheet11.cell(row = 18 + i, column = 13).value = "=IF(AND(H"+str(18 + i)+"<="+str(val4)+",H"+str(18 + i)+">"+str(val3)+"),F"+str(18 + i)+",0)".format(18 + i)
		# 	Sheet11.cell(row = 18 + i, column = 14).value = "=IF(AND(H"+str(18 + i)+"<="+str(val5)+",H"+str(18 + i)+">"+str(val4)+"),F"+str(18 + i)+",0)".format(18 + i)
		# 	Sheet11.cell(row = 18 + i, column = 15).value = "=IF(H"+str(18 + i)+">"+str(val5)+",F"+str(18 + i)+",0)".format(18 + i)
		# 	Sheet11.cell(row = 18 + i, column = 16).value = "=F"+str(18 + i)+"-SUM(I"+str(18+i)+":O"+str(18+i)+")"

		# 	Sheet11.cell(row=18 + i, column=9).number_format = '#,##0_);(#,##0)'
		# 	Sheet11.cell(row=18 + i, column=10).number_format = '#,##0_);(#,##0)'
		# 	Sheet11.cell(row=18 + i, column=11).number_format = '#,##0_);(#,##0)'
		# 	Sheet11.cell(row=18 + i, column=12).number_format = '#,##0_);(#,##0)'
		# 	Sheet11.cell(row=18 + i, column=13).number_format = '#,##0_);(#,##0)'
		# 	Sheet11.cell(row=18 + i, column=14).number_format = '#,##0_);(#,##0)'
		# 	Sheet11.cell(row=18 + i, column=15).number_format = '#,##0_);(#,##0)'
		# 	Sheet11.cell(row=18 + i, column=16).number_format = '#,##0_);(#,##0)'

			#bucket zone formulas
			# Sheet11.cell(row = 13, column = 9).value = "=SUM(I18:I{0})".format(18 + i)
			# Sheet11.cell(row = 13, column = 10).value = "=SUM(J18:J{0})".format(18 + i)
			# Sheet11.cell(row = 13, column = 11).value = "=SUM(K18:K{0})".format(18 + i)
			# Sheet11.cell(row = 13, column = 12).value = "=SUM(L18:L{0})".format(18 + i)
			# Sheet11.cell(row = 13, column = 13).value = "=SUM(M18:M{0})".format(18 + i)
			# Sheet11.cell(row = 13, column = 14).value = "=SUM(N18:N{0})".format(18 + i)
			# Sheet11.cell(row = 13, column = 15).value = "=SUM(O18:O{0})".format(18 + i)

			#FORMAT
			# for row in Sheet11["I13:O13"]:
			# 	for cell in row:
			# 		cell.number_format = '#,##0_);(#,##0)'
			# 		cell.font = font_worksheet

			#reconciliation zone
		# 	Sheet11.cell(row = 6, column = 5).value = "411"
		# 	Sheet11.cell(row = 6, column = 6).value = "=P13".format(18 + i)
		# 	Sheet11.cell(row = 6, column = 7).value ="=SUMIF('TB Robot'!A:A,""411"",'TB Robot'!H:H)".format(18 + i)
		# 	Sheet11.cell(row = 6, column = 8).value ="=F6-G6".format(18 + i)

		# 	Sheet11.cell(row=6, column=6).number_format = '#,##0_);(#,##0)'
		# 	Sheet11.cell(row=6, column=7).number_format = '#,##0_);(#,##0)'
		# 	Sheet11.cell(row=6, column=8).number_format = '#,##0_);(#,##0)'

		# 	Sheet11.cell(row=6, column=5).font = font_worksheet
		# 	Sheet11.cell(row=6, column=6).font = font_worksheet
		# 	Sheet11.cell(row=6, column=7).font = font_worksheet
		# 	Sheet11.cell(row=6, column=8).font = check_font

		# for i in range(0, len(clientName)):
		# 	Sheet11.cell(row = 18 + i, column = 2).font = font_worksheet

		# for i in range(0, len(clientCode)):
		# 	Sheet11.cell(row = 18 + i, column = 1).font = font_worksheet

		# for i in range(0, len(clientAccount)):
		# 	Sheet11.cell(row = 18 + i, column = 3).font = font_worksheet

		# for i in range(0, len(dueDate)):
		# 	Sheet11.cell(row = 18 + i, column = 4).font = font_worksheet

		# for i in range(0, len(invoiceNo)):
		# 	Sheet11.cell(row = 18 + i, column = 5).font = font_worksheet

		# for i in range(0, len(amtLC)):
		# 	Sheet11.cell(row = 18 + i, column = 6).number_format = '#,##0_);(#,##0)'

		# for i in range(0, len(amtLC)):
		# 	Sheet11.cell(row = 18 + i, column = 6).font = font_worksheet

		# for i in range(0, len(invoiceVal)):
		# 	Sheet11.cell(row = 18 + i, column = 7).number_format = '#,##0_);(#,##0)'

		# for i in range(0, len(invoiceVal)):
		# 	Sheet11.cell(row = 18 + i, column = 7).font = font_worksheet

		# for i in range(0, len(dueDate)):
		# 	Sheet11.cell(row = 18 + i, column = 8).font = font_worksheet
		# 	Sheet11.cell(row = 18 + i, column = 9).font = font_worksheet
		# 	Sheet11.cell(row = 18 + i, column = 10).font = font_worksheet
		# 	Sheet11.cell(row = 18 + i, column = 11).font = font_worksheet
		# 	Sheet11.cell(row = 18 + i, column = 12).font = font_worksheet
		# 	Sheet11.cell(row = 18 + i, column = 13).font = font_worksheet
		# 	Sheet11.cell(row = 18 + i, column = 14).font = font_worksheet
		# 	Sheet11.cell(row = 18 + i, column = 15).font = font_worksheet
		# 	Sheet11.cell(row = 18 + i, column = 16).font = font_worksheet

		# Sheet11.freeze_panes = 'C18'

		Sheet11.column_dimensions['A'].width = 10
		Sheet11.column_dimensions['B'].width = 20
		Sheet11.column_dimensions['C'].width = 11
		Sheet11.column_dimensions['D'].width = 12
		Sheet11.column_dimensions['E'].width = 14
		Sheet11.column_dimensions['F'].width = 12
		Sheet11.column_dimensions['G'].width = 12
		Sheet11.column_dimensions['H'].width = 12
		Sheet11.column_dimensions['I'].width = 9
		Sheet11.column_dimensions['J'].width = 9
		Sheet11.column_dimensions['K'].width = 9
		Sheet11.column_dimensions['L'].width = 9
		Sheet11.column_dimensions['M'].width = 9
		Sheet11.column_dimensions['N'].width = 9
		Sheet11.column_dimensions['O'].width = 12
		Sheet11.column_dimensions['P'].width = 15

		for cell in Sheet11["E"]:
			cell.number_format = "mm/dd/yyyy"

		# Sheet11.auto_filter.ref = "A17:P17"

		#FONT / FORMAT
		#header
		Sheet11.cell(row=1, column=1).font = ft1
		Sheet11.cell(row=1, column=2).font = ft1
		Sheet11.cell(row=2, column=1).font = ft1
		Sheet11.cell(row=2, column=2).font = ft1

		# Sheet11.cell(row=1, column=15).font = ft1
		# Sheet11.cell(row=1, column=16).font = ft1
		# Sheet11.cell(row=2, column=15).font = ft1
		# Sheet11.cell(row=2, column=16).font = ft1
		# Sheet11.cell(row=3, column=15).font = ft1
		# Sheet11.cell(row=3, column=16).font = check_font

		# Sheet11.cell(row=5, column=5).font = cap_tabel
		# Sheet11.cell(row=5, column=6).font = cap_tabel
		# Sheet11.cell(row=5, column=7).font = cap_tabel
		# Sheet11.cell(row=5, column=8).font = check_font

		# # test name
		# Sheet11.cell(row=4, column=1).font = f_testname
		# Sheet11.cell(row=12, column=1).font = f_testname
		# Sheet11.cell(row=13, column=1).font = ft1

		# Sheet11.cell(row=13, column=16).font = ft1

		# Sheet11.cell(row=16, column=8).font = font_worksheet
		# Sheet11.cell(row=17, column=16).font = check_font

		# for row in Sheet11["A17:O17"]:
		# 	for cell in row:
		# 		cell.font = cap_tabel

		# for row in Sheet11["A17:G17"]:
		# 	for cell in row:
		# 		cell.fill = cap_tabel_color_PBC

		# for row in Sheet11["H17:P17"]:
		# 	for cell in row:
		# 		cell.fill = cap_tabel_color_GT

		# Sheet11.cell(row=5, column=5).fill = cap_tabel_color_GT
		# Sheet11.cell(row=5, column=6).fill = cap_tabel_color_GT
		# Sheet11.cell(row=5, column=7).fill = cap_tabel_color_GT
		# Sheet11.cell(row=5, column=8).fill = cap_tabel_color_GT

		# bucket zone
		# for row in Sheet11["I12:P12"]:
		# 	for cell in row:
		# 		cell.font = cap_tabel

		# for row in Sheet11["I12:P12"]:
		# 	for cell in row:
		# 		cell.fill = cap_tabel_color_GT

		# suma procentelor
		Sheet11.cell(row=13, column=16).font = ft1
		Sheet11.cell(row=13, column=16).number_format = '#,##0_);(#,##0)'

		# procent din total
		# for row in Sheet11['I14:O14']:
		# 	for cell in row:
		# 		cell.style='Percent'

		# for row in Sheet11['I14:O14']:
		# 	for cell in row:
		# 		cell.font = fprocentaj

		# prepared and date
		Sheet11.cell(row=1, column=15).font = ft1
		Sheet11.cell(row=1, column=16).font = ft1
		Sheet11.cell(row=2, column=15).font = ft1
		Sheet11.cell(row=2, column=16).font = ft1
		Sheet11.cell(row=3, column=15).font = ft1
		Sheet11.cell(row=3, column=16).font = check_font

		ws1 = wb.active
		Sheet2 = wb.create_sheet("TR Aging Client level")
		Sheet2.sheet_view.showGridLines = False
		Sheet2.freeze_panes = 'C12'
		Sheet2.auto_filter.ref = "A11:L11"
		# Sheet2.auto_filter.ref = "Amount in LC"

		#header
		Sheet2.cell(row = 1, column = 1).value = "Client name:"
		Sheet2.cell(row = 1, column = 2).value = clientname1
		Sheet2.cell(row = 2, column = 1).value = "Period end:"
		Sheet2.cell(row = 2, column = 2).value = yearEnd1
		Sheet2.cell(row = 2, column = 2).number_format = 'mm/dd/yyyy'            

		#test name
		Sheet2.cell(row = 4, column = 1).value = "Receivables ageing recomputation"

		#table
		Sheet2.cell(row = 7, column = 1).value = "Unusual items"
		Sheet2.cell(row = 7, column = 2).value ='=SUMIF(L:L,"yes",C:C)'.format(0)
		Sheet2.cell(row = 11, column = 1).value = "Client"
		Sheet2.cell(row = 11, column = 2).value = "Client code"
		Sheet2.cell(row = 11, column = 3).value = "Amount in LC"
		Sheet2.cell(row = 11, column = 4).value = "Not due"
		Sheet2.cell(row = 11, column = 5).value = "<="+str(val1)+" days"
		Sheet2.cell(row = 11, column = 6).value = str(val1)+"-"+str(val2)+" days"
		Sheet2.cell(row = 11, column = 7).value = str(val2)+"-"+str(val3)+" days"
		Sheet2.cell(row = 11, column = 8).value = str(val3)+"-"+str(val4)+" days"
		Sheet2.cell(row = 11, column = 9).value = str(val4)+"-"+str(val5)+" days"
		Sheet2.cell(row = 11, column = 10).value = "Over "+str(val5)+" days"
		Sheet2.cell(row = 11, column = 11).value = "Check"
		Sheet2.cell(row = 11, column = 12).value = "Unusual items"

		#prepared and date
		Sheet2.cell(row = 1, column = 10).value = "Processed by"
		Sheet2.cell(row = 1, column = 11).value = preparedBy1
		Sheet2.cell(row = 2, column = 10).value = "Date"
		Sheet2.cell(row = 2, column = 11).value = datePrepared1
		Sheet2.cell(row = 2, column = 11).number_format = 'mm/dd/yyyy'
		Sheet2.cell(row = 3, column = 10).value = "Refference"
		Sheet2.cell(row = 3, column = 11).value = refference1

		#bucket zone
		Sheet2.cell(row = 6, column = 4).value = "Not due"
		Sheet2.cell(row = 6, column = 5).value = "<="+str(val1)+" days"
		Sheet2.cell(row = 6, column = 6).value = str(val1)+"-"+str(val2)+" days"
		Sheet2.cell(row = 6, column = 7).value = str(val2)+"-"+str(val3)+" days"
		Sheet2.cell(row = 6, column = 8).value = str(val3)+"-"+str(val4)+" days"
		Sheet2.cell(row = 6, column = 9).value = str(val4)+"-"+str(val5)+" days"
		Sheet2.cell(row = 6, column = 10).value = "Over "+str(val5)+" days"
		Sheet2.cell(row = 6, column = 11).value = "Total"

		for cell in Sheet2['C']:
			cell.number_format = '#,##0_);(#,##0)'

		#FORMAT & COLORS
		# header
		
		Sheet2.cell(row=1, column=2).font = ft1
		Sheet2.cell(row=1, column=1).font = ft1
		Sheet2.cell(row=2, column=2).font = ft1
		Sheet2.cell(row=2, column=1).font = ft1		
		Sheet2.cell(row = 11, column = 12).font =check_font
		Sheet2.cell(row = 11, column = 12).fill = cap_tabel_galben

		# test name
		Sheet2.cell(row=4, column=1).font = f_testname

		# table
		Sheet2.cell(row=11, column=11).font = check_font

		for row in Sheet2["A11:J11"]:
			for cell in row:
				cell.font = cap_tabel

		for row in Sheet2["A11:C11"]:
			for cell in row:
				cell.fill = cap_tabel_color_PBC

		for row in Sheet2['D11:K11']:
			for cell in row:
				cell.fill = cap_tabel_color_GT

		# prepared and date
		Sheet2.cell(row=1, column=11).font = ft1
		Sheet2.cell(row=1, column=10).font = ft1
		Sheet2.cell(row=2, column=11).font = ft1
		Sheet2.cell(row=2, column=10).font = ft1
		Sheet2.cell(row=3, column=11).font = check_font
		Sheet2.cell(row=3, column=10).font = ft1

		# bucket zone
		for row in Sheet2["D6:K6"]:
			for cell in row:
				cell.font=cap_tabel
				cell.fill = cap_tabel_color_GT

		Sheet2.cell(row=7, column=1).font =check_font
		Sheet2.cell(row=7, column=1).fill =cap_tabel_galben

		Sheet2.cell(row=7, column=2).font =check_font
		Sheet2.cell(row=7, column=2).fill =cap_tabel_galben
		Sheet2.cell(row=7, column=2).number_format ='#,##0_);(#,##0)'
		Sheet2.column_dimensions['B'].width = 10

		cod_clienti = list(set(clientCode))
		# cod_clienti.remove("None")
		cod_clienti1 = []
		for val in cod_clienti:
			if val != None:
				cod_clienti1.append(val)
		nume_clienti = list(set(clientName))

		# print(nume_clienti)
		print(cod_clienti1)

		if len(cod_clienti1) == 0:
			print("e goala")
			for i in range(0, len(nume_clienti)):
				Sheet2.cell(row=12 + i, column=2).value = nume_clienti[i]
				Sheet2.cell(row=12 + i, column=3).value = "=SUMIF('TR Aging Invoice level'!B:B,'TR Aging Client level'!B{0},'TR Aging Invoice level'!F:F)".format(12 + i)
				Sheet2.cell(row=12 + i,	column=4).value = "=SUMIF('TR Aging Invoice level'!B:B,'TR Aging Client level'!B{0},'TR Aging Invoice level'!I:I)".format(12 + i)
				Sheet2.cell(row=12 + i, column=5).value = "=SUMIF('TR Aging Invoice level'!B:B,'TR Aging Client level'!B{0},'TR Aging Invoice level'!J:J)".format(12 + i)
				Sheet2.cell(row=12 + i,column=6).value = "=SUMIF('TR Aging Invoice level'!B:B,'TR Aging Client level'!B{0},'TR Aging Invoice level'!K:K)".format(12 + i)
				Sheet2.cell(row=12 + i,column=7).value = "=SUMIF('TR Aging Invoice level'!B:B,'TR Aging Client level'!B{0},'TR Aging Invoice level'!L:L)".format(12 + i)
				Sheet2.cell(row=12 + i,column=8).value = "=SUMIF('TR Aging Invoice level'!B:B,'TR Aging Client level'!B{0},'TR Aging Invoice level'!M:M)".format(12 + i)
				Sheet2.cell(row=12 + i,column=9).value = "=SUMIF('TR Aging Invoice level'!B:B,'TR Aging Client level'!B{0},'TR Aging Invoice level'!N:N)".format(12 + i)
				Sheet2.cell(row=12 + i,column=10).value = "=SUMIF('TR Aging Invoice level'!B:B,'TR Aging Client level'!B{0},'TR Aging Invoice level'!O:O)".format(12 + i)
				Sheet2.cell(row=12 + i, column=11).value = "=C{0}-SUM(D{0}:J{0})".format(12 + i)
				Sheet2.cell(row=12 + i, column=12).value = """=IF('TR Aging Invoice level'!$E$6="411",IF('TR Aging Client level'!C{0}<0,"Yes","No"),IF('TR Aging Client level'!C{0}>0,"Yes","No"))""".format(12+i)

				Sheet2.cell(row=12 + i, column=2).font = font_worksheet
				Sheet2.cell(row=12 + i, column=3).font = font_worksheet
				Sheet2.cell(row=12 + i, column=4).font = font_worksheet
				Sheet2.cell(row=12 + i, column=5).font = font_worksheet
				Sheet2.cell(row=12 + i, column=6).font = font_worksheet
				Sheet2.cell(row=12 + i, column=7).font = font_worksheet
				Sheet2.cell(row=12 + i, column=8).font = font_worksheet
				Sheet2.cell(row=12 + i, column=9).font = font_worksheet
				Sheet2.cell(row=12 + i, column=10).font = font_worksheet
				Sheet2.cell(row=12 + i, column=11).font = font_worksheet

				Sheet2.cell(row=12 + i, column=3).number_format = '#,##0_);(#,##0)'
				Sheet2.cell(row=12 + i, column=4).number_format = '#,##0_);(#,##0)'
				Sheet2.cell(row=12 + i, column=5).number_format = '#,##0_);(#,##0)'
				Sheet2.cell(row=12 + i, column=6).number_format = '#,##0_);(#,##0)'
				Sheet2.cell(row=12 + i, column=7).number_format = '#,##0_);(#,##0)'
				Sheet2.cell(row=12 + i, column=8).number_format = '#,##0_);(#,##0)'
				Sheet2.cell(row=12 + i, column=9).number_format = '#,##0_);(#,##0)'
				Sheet2.cell(row=12 + i, column=10).number_format = '#,##0_);(#,##0)'
				Sheet2.cell(row=12 + i, column=11).number_format = '#,##0_);(#,##0)'

		else:
			print("nu e goala")
			for i in range(0, len(cod_clienti1)):
				Sheet2.cell(row = 12 + i, column = 1).value = cod_clienti1[i]
				Sheet2.cell(row = 12 + i, column = 2).value = "=VLOOKUP(A{0},'TR Aging Invoice level'!A:B,2,0)".format(12 + i)
				Sheet2.cell(row = 12 + i, column = 3).value = "=SUMIF('TR Aging Invoice level'!A:A,'TR Aging Client level'!A{0},'TR Aging Invoice level'!F:F)".format(12 + i)
				Sheet2.cell(row = 12 + i, column = 4).value = "=SUMIF('TR Aging Invoice level'!A:A,'TR Aging Client level'!A{0},'TR Aging Invoice level'!I:I)".format(12 + i)
				Sheet2.cell(row = 12 + i, column = 5).value = "=SUMIF('TR Aging Invoice level'!A:A,'TR Aging Client level'!A{0},'TR Aging Invoice level'!J:J)".format(12 + i)
				Sheet2.cell(row = 12 + i, column = 6).value = "=SUMIF('TR Aging Invoice level'!A:A,'TR Aging Client level'!A{0},'TR Aging Invoice level'!K:K)".format(12 + i)
				Sheet2.cell(row = 12 + i, column = 7).value = "=SUMIF('TR Aging Invoice level'!A:A,'TR Aging Client level'!A{0},'TR Aging Invoice level'!L:L)".format(12 + i)
				Sheet2.cell(row = 12 + i, column = 8).value = "=SUMIF('TR Aging Invoice level'!A:A,'TR Aging Client level'!A{0},'TR Aging Invoice level'!M:M)".format(12 + i)
				Sheet2.cell(row = 12 + i, column = 9).value = "=SUMIF('TR Aging Invoice level'!A:A,'TR Aging Client level'!A{0},'TR Aging Invoice level'!N:N)".format(12 + i)
				Sheet2.cell(row = 12 + i, column = 10).value = "=SUMIF('TR Aging Invoice level'!A:A,'TR Aging Client level'!A{0},'TR Aging Invoice level'!O:O)".format(12 + i)
				Sheet2.cell(row = 12 + i, column = 11).value = "=C{0}-SUM(D{0}:J{0})".format(12 + i)
				Sheet2.cell(row = 12 + i, column = 12).value = """=IF('TR Aging Invoice level'!$E$6="411",IF('TR Aging Client level'!C{0}<0,"Yes","No"),IF('TR Aging Client level'!C{0}>0,"Yes","No"))""".format(12+i)

				Sheet2.cell(row=12 + i, column=1).font = font_worksheet
				Sheet2.cell(row=12 + i, column=2).font = font_worksheet
				Sheet2.cell(row=12 + i, column=3).font = font_worksheet
				Sheet2.cell(row=12 + i, column=4).font = font_worksheet
				Sheet2.cell(row=12 + i, column=5).font = font_worksheet
				Sheet2.cell(row=12 + i, column=6).font = font_worksheet
				Sheet2.cell(row=12 + i, column=7).font = font_worksheet
				Sheet2.cell(row=12 + i, column=8).font = font_worksheet
				Sheet2.cell(row=12 + i, column=9).font = font_worksheet
				Sheet2.cell(row=12 + i, column=10).font = font_worksheet
				Sheet2.cell(row=12 + i, column=11).font = font_worksheet

				Sheet2.cell(row=12 + i, column=3).number_format = '#,##0_);(#,##0)'
				Sheet2.cell(row=12 + i, column=4).number_format = '#,##0_);(#,##0)'
				Sheet2.cell(row=12 + i, column=5).number_format = '#,##0_);(#,##0)'
				Sheet2.cell(row=12 + i, column=6).number_format = '#,##0_);(#,##0)'
				Sheet2.cell(row=12 + i, column=7).number_format = '#,##0_);(#,##0)'
				Sheet2.cell(row=12 + i, column=8).number_format = '#,##0_);(#,##0)'
				Sheet2.cell(row=12 + i, column=9).number_format = '#,##0_);(#,##0)'
				Sheet2.cell(row=12 + i, column=10).number_format = '#,##0_);(#,##0)'
				Sheet2.cell(row=12 + i, column=11).number_format = '#,##0_);(#,##0)'


		Sheet2.cell(row = 7, column = 4).value = "=SUM(D11:D103)".format(18 + i)
		Sheet2.cell(row = 7, column = 5).value = "=SUM(E11:E103)".format(18 + i)
		Sheet2.cell(row = 7, column = 6).value = "=SUM(F11:F103)".format(18 + i)
		Sheet2.cell(row = 7, column = 7).value = "=SUM(G11:G103)".format(18 + i)
		Sheet2.cell(row = 7, column = 8).value = "=SUM(H11:H103)".format(18 + i)
		Sheet2.cell(row = 7, column = 9).value = "=SUM(I11:I103)".format(18 + i)
		Sheet2.cell(row = 7, column = 10).value = "=SUM(J11:J103)".format(18 + i)
		Sheet2.cell(row = 7, column = 11).value = "=SUM(D7:J7)".format(18 + i)

		Sheet2.cell(row=7, column=11).font = ft1
		Sheet2.cell(row=7, column=11).number_format = '#,##0_);(#,##0)'

		for row in Sheet2["D7:J7"]:
			for cell in row:
				cell.font = font_worksheet
				cell.number_format = '#,##0_);(#,##0)'


		#procente buckets
		Sheet2.cell(row=8, column=4).value = "=D7/$K$7".format(18 + i)
		Sheet2.cell(row=8, column=5).value = "=E7/$K$7".format(18 + i)
		Sheet2.cell(row=8, column=6).value = "=F7/$K$7".format(18 + i)
		Sheet2.cell(row=8, column=7).value = "=G7/$K$7".format(18 + i)
		Sheet2.cell(row=8, column=8).value = "=H7/$K$7".format(18 + i)
		Sheet2.cell(row=8, column=9).value = "=I7/$K$7".format(18 + i)
		Sheet2.cell(row=8, column=10).value = "=J7/$K$7".format(18 + i)

		for row in Sheet2["D8:K8"]:
			for cell in row:
				cell.style = 'Percent'
				cell.font = fprocentaj

		Sheet2.column_dimensions['A'].width = 10
		Sheet2.column_dimensions['B'].width = 10
		Sheet2.column_dimensions['C'].width = 15
		Sheet2.column_dimensions['D'].width = 10
		Sheet2.column_dimensions['E'].width = 10
		Sheet2.column_dimensions['F'].width = 10
		Sheet2.column_dimensions['G'].width = 10
		Sheet2.column_dimensions['H'].width = 10
		Sheet2.column_dimensions['I'].width = 10
		Sheet2.column_dimensions['J'].width = 10
		Sheet2.column_dimensions['K'].width = 15
		print("a ajuns aici")

		topClients=wb.create_sheet("Top 10 clients")
		# sheetAscuns=wb.create_sheet("top 10 ")
		# sheetAscuns.sheet_state = 'hidden'
		topClients.sheet_view.showGridLines = False
		# sheetAscuns.cell(row = 1, column = 2).value = "suma"
		# sheetAscuns.cell(row = 1, column = 1).value = "partener"

		topClients.cell(row = 1, column = 1).value = "Client name:"
		topClients.cell(row = 1, column = 2).value = clientname1
		topClients.cell(row = 1, column = 2).font =ft1_1
		topClients.cell(row = 2, column = 1).value = "Period end:"
		topClients.cell(row = 2, column = 2).value = yearEnd1
		topClients.cell(row = 2, column = 2).font = ft1_1
		topClients.cell(row = 2, column = 2).number_format = 'mm/dd/yyyy'
		topClients.cell(row = 3, column = 1).value = "Top Clients"
		topClients.cell(row = 3, column = 1).font =ft1_1

		topClients.cell(row = 1, column = 10).value = "Processed by"
		topClients.cell(row = 1, column = 11).value = preparedBy1
		topClients.cell(row = 1, column = 11).font = ft1_1
		topClients.cell(row = 2, column = 10).value = "Date"
		topClients.cell(row = 2, column = 11).value = datePrepared1
		topClients.cell(row = 2, column = 11).font =ft1_1
		topClients.cell(row = 2, column = 11).number_format = 'mm/dd/yyyy'
		topClients.cell(row = 3, column = 10).value = "Refference"
		topClients.cell(row = 3, column = 10).font = check_font
		topClients.cell(row = 3, column = 11).value = refference1
		topClients.cell(row = 3, column = 11).font =check_font

		topClients.cell(row = 7, column = 1).value = "No crt"
		topClients.cell(row = 7, column = 2).value = "Client"
		topClients.cell(row = 7, column = 3).value = "Client code"
		topClients.cell(row = 7, column = 4).value = "Amount in LC CY"
		topClients.cell(row = 7, column = 5).value = "Amount in PY (manual)"
		topClients.cell(row = 7, column = 6).value = "Variation"
		topClients.cell(row = 7, column = 7).value = "Variation %"
		topClients.cell(row = 7, column = 8).value = "Weight"
		# a=0
		# for i in range(0, 10):
		# 	a=a+1
		# 	topClients.cell(row=8+i, column=1).value=a
		# 	# print(a,'numar')
		
		# topClients.cell(row = 19, column = 3).value = "Others"

		clientNameUni=list(set(clientName))
		clientCodeUni=list(set(clientCode))
		
		amountTip=[]
		amountTips=[]
		for i in range(0, len(clientNameUni)):
			s=0
			for z in range(0, len(clientName)):
				if clientNameUni[i] == clientName[z]:
					s=s+amtLC[z]
			amountTip.append(s)
			amountTips.append(s)

		# # print(len(amountTip))
		amountTips.sort(reverse=True)
		# # print(amountTip, 'amountTip')			# print(len(amountTipSort))
		x=min(len(amountTips),10)
		
		top10am=[]
		top10c=[]
		for jj in range(0, x):
			top10am.append(amountTips[jj])
		for m in range(0,len(top10am)):
			topClients.cell(row = 8+m, column = 4).value = top10am[m]
		
		
		for k in range(0,len(top10am)):
			for p in range(0,len(amountTip)):
				if(top10am[k]==amountTip[p]):
					top10c.append(clientNameUni[p])
					# topClients.cell(row = 7+i, column = 4).value = top10c[k]
		for n in range(0, len(top10c)):
			topClients.cell(row=8+n, column=2).value=top10c[n]

		for j in range(1, len(top10am)+1):
			topClients.cell(row=7+j, column=1).value=j



		for i in range(0, len(top10am)):
			topClients.cell(row = 8+i, column = 3).value = "=_xlfn.XLOOKUP(B"+str(8+i)+",'TR Aging Invoice level'!B:B,'TR Aging Invoice level'!A:A,0,0,1)".format(8+i)
			topClients.cell(row = 8+i, column = 6).value ="=D{0}-E{0}".format(8+i)
			topClients.cell(row = 8+i, column = 7).value ="=F{0}/E{0}".format(8+i)
			topClients.cell(row = 8+i, column = 8).value ="=D{0}/'TR Aging Client level'!$K$7".format(8+i)

		topClients.cell(row = 8+len(top10am), column = 4).value ="=SUM(D8:D"+str(len(top10am)+7)+")"
		topClients.cell(row = 8+len(top10am), column = 5).value ="=SUM(E8:E"+str(len(top10am)+7)+")"
		topClients.cell(row = 8+len(top10am), column = 6).value ="=SUM(F8:F"+str(len(top10am)+7)+")"
		topClients.cell(row = 8+len(top10am), column = 7).value ="=F"+str(len(top10am)+8)+"/E"+str(len(top10am)+8)
		# topClients.cell(row = 8+len(top10am), column = 8).value ="=D18/'TR Aging Client level'!$K$7"

		topClients.cell(row = 7+len(top10am), column = 1).border = doubleborder
		topClients.cell(row = 7+len(top10am), column = 2).border = doubleborder
		topClients.cell(row = 7+len(top10am), column = 3).border = doubleborder
		topClients.cell(row = 7+len(top10am), column = 4).border = doubleborder
		topClients.cell(row = 7+len(top10am), column = 5).border = doubleborder
		topClients.cell(row = 7+len(top10am), column = 6).border = doubleborder
		topClients.cell(row = 7+len(top10am), column = 7).border = doubleborder
		topClients.cell(row = 7+len(top10am), column = 8).border = doubleborder

		topClients.cell(row = 7+len(top10am), column = 1).font = ft1_1
		topClients.cell(row = 7+len(top10am), column = 2).font = ft1_1
		topClients.cell(row = 7+len(top10am), column = 3).font = ft1_1
		topClients.cell(row = 7+len(top10am), column = 4).font = ft1_1
		topClients.cell(row = 7+len(top10am), column = 5).font = ft1_1
		topClients.cell(row = 7+len(top10am), column = 6).font = ft1_1
		topClients.cell(row = 7+len(top10am), column = 7).font = ft1_1
		topClients.cell(row = 7+len(top10am), column = 8).font = ft1_1
		# topClients.cell(row = 19, column = 4).value ="='TR Aging Client level'!$K$7-'Top 10 Clients'!D18"
		# topClients.cell(row = 19, column = 5).value ="Manual"
		# topClients.cell(row = 19, column = 6).value ="=SUM(F9:F18)"
		# topClients.cell(row = 19, column = 7).value ="=F19/E19"
		# topClients.cell(row = 19, column = 8).value ="=D19/'TR Aging Client level'!$K$7"
		topClients.cell(row =  8+len(top10am), column = 3).value = "Total"
		for row in topClients["A7:H7"]:
			for cell in row:
				cell.font = cap_tabel
				cell.fill = cap_tabel_color_GT
		for row in topClients['D8:D19']:
			for cell in row:
				cell.number_format = '#,##0_);(#,##0)'

		for row in topClients['E8:E19']:
			for cell in row:
				cell.number_format = '#,##0_);(#,##0)'
		for row in topClients['F8:F19']:
			for cell in row:
				cell.number_format = '#,##0_);(#,##0)'
		for row in topClients['H8:H19']:
			for cell in row:
				cell.number_format = '0.00%'
		for row in topClients['G8:G19']:
			for cell in row:
				cell.number_format = '0.00%'

		for row in topClients['A8:H19']:
			for cell in row:
				cell.font = ft1_1
		
		for row in topClients['C18:H19']:
			for cell in row:
				cell.font = ft1_1b

		# for row in topClients['A17:H17']:
		# 	for cell in row:
		# 		cell.border = doubleborder

		
		
		topClients.cell(row=1, column=1).font=ft1
		topClients.cell(row=2, column=1).font=ft1

		topClients.cell(row=1, column=10).font=ft1
		topClients.cell(row=2, column=10).font=ft1
		
		topClients.cell(row=1, column=1).font=ft1
		topClients.cell(row=2, column=1).font=ft1

		
		
		topClients.column_dimensions['B'].width = 16
		topClients.column_dimensions['C'].width = 10
		topClients.column_dimensions['D'].width = 15
		topClients.column_dimensions['E'].width = 15
		topClients.column_dimensions['F'].width = 15




	# except:
	# messagebox.showerror("Error!", "Details file not found.")
		# app.destroy()

	# try:
		#tb
		tb = openpyxl.load_workbook(file_TB, data_only = True) #deschidem TB-ul
		tb1 = tb.active
		Sheet3 = wb.create_sheet("TB Robot")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "Account":
					row_tb = cell.row
					column_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			account = [b.value for b in tb1[column_tb][row_tb:lun]]
		except:
			flash("Please insert the correct header for Account in Trial Balance file")
			return render_template("index.html")                     
		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "Description":
					row_tb = cell.row
					column_description = cell.column
					lun = len(tb1[cell.column])
		try:
			descr = [b.value for b in tb1[column_description][row_tb:lun]]
		except:
			flash("Please insert the correct header for Description in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error!", "The 'Description' value is  not correctly written")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "OB":
					row_tb = cell.row
					coloana_opTB_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			opTB = [b.value for b in tb1[coloana_opTB_tb][row_tb:lun]]
		except:
			flash("Please insert the correct header for OB in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error!", "OB' value is  not correctly written")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "CM":
					row_tb = cell.row
					coloana_cr_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			cr_mv = [b.value for b in tb1[coloana_cr_tb][row_tb:lun]]
		except:
			flash("Please insert the correct header for CM in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error!", "CM' value is  not correctly written")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "DM":
					row_tb = cell.row
					coloana_db_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			dr_mv = [b.value for b in tb1[coloana_db_tb][row_tb:lun]]
		except:
			flash("Please insert the correct header for DM in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error!", "DM' value is  not correctly written")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "CB":
					row_tb = cell.row
					coloana_clTB_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			clTB = [b.value for b in tb1[coloana_clTB_tb][row_tb:lun]]
		
		except:
			flash("Please insert the correct header for CB in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error!", "CB' value is  not correctly written")

		Sheet3.cell(row = 1, column = 1).value = "Synt(3)"
		Sheet3.cell(row = 1, column = 2).value = "Synt(4)"
		Sheet3.cell(row = 1, column = 3).value = "Account"
		Sheet3.cell(row = 1, column = 4).value = "Description"
		Sheet3.cell(row = 1, column = 5).value = "Opening Balance"
		Sheet3.cell(row = 1, column = 6).value = "Debit Movement"
		Sheet3.cell(row = 1, column = 7).value = "Credit Movement"
		Sheet3.cell(row = 1, column = 8).value = "Closing Balance"
		Sheet3.cell(row = 1, column = 9).value = "Synt(2)"

		#FORMAT
		for row in Sheet3['A1:I1']:
			for cell in row:
				cell.font = cap_tabel
				cell.fill = cap_tabel_color_GT

		for i in range(0,len(account)):
			Sheet3.cell(row = 2 + i, column = 3).value = account[i]
			Sheet3.cell(row = 2 + i, column = 2).value = str(account[i])[:4]   #in Excel =left("celula", 4)

		for i in range(0, len(account)):
			Sheet3.cell(row = 2 + i, column = 1).value = str(account[i])[:3] #in Excel =left("celula", 3)

		for i in range(0, len(descr)):
			Sheet3.cell(row = 2 + i, column = 4).value = descr[i]

		for i in range(0, len(opTB)):
			Sheet3.cell(row = 2 + i, column = 5).value = opTB[i]

		for i in range(0, len(cr_mv)):
			Sheet3.cell(row = 2 + i, column = 6).value = cr_mv[i]

		for i in range(0, len(dr_mv)):
			Sheet3.cell(row = 2 + i, column = 7).value = dr_mv[i]

		for i in range(0, len(clTB)):
			Sheet3.cell(row = 2 + i, column = 8).value = clTB[i]

		for i in range(0, len(account)):
			Sheet3.cell(row = 2 + i, column = 9).value =  str(account[i])[:2]

		Sheet4 = wb.create_sheet("Supporting docs --->")
		Sheet5 = wb.create_sheet("TB PBC")

		mr = tb1.max_row
		mc = tb1.max_column
		# copying the cell values from source
		# excel file to destination excel file
		for i in range (1, mr + 1):
			for j in range (1, mc + 1):
				# reading cell value from source excel file
				c = tb1.cell(row = i, column = j)
				# writing the read value to destination excel file
				Sheet5.cell(row = i, column = j).value = c.value

		Sheet6 = wb.create_sheet("Details PBC")

		am = ws.max_row
		an = ws.max_column

		for i in range (1, am +1):
			for j in range (1, an + 1):
				a = ws.cell(row = i, column = j)
				Sheet6.cell(row = i, column = j).value = a.value


		folderpath = "home/auditappnexia/output/Ageing" 
		file_pathFS = os.path.join(folderpath, "Ageing Test"+" "+clientname1+".xlsx")
		wb.save(file_pathFS)
		# print("ceva")
	return send_from_directory(folderpath, "Ageing Test" + " " + clientname1 + ".xlsx", as_attachment=True)

@app.route('/VAT/Instructions', methods=['GET'])
def downloadVAT():
		filepath = "/home/auditappnexia/output/vat"
 
		return send_from_directory(filepath,"Instructions - VAT.docx", as_attachment=True)
@app.route('/VAT/GTbJY47MKf1oajfEqntaRFSt8fw')
def my_formVAT():
	return render_template('VAT.html')
@app.route('/VAT/GTbJY47MKf1oajfEqntaRFSt8fw', methods=['POST', 'GET'])
def my_form_post():
	yearEnd1 = datetime.datetime.strptime(
		request.form['yearEnd'],
		'%Y-%m-%d')
	preparedBy1 = request.form['preparedBy']
	clientname1 = request.form['client']   
	datePrepared1 = datetime.datetime.strptime(
		request.form['preparedDate'],
		'%Y-%m-%d')
	refference1 = request.form['reff']
#
# 	# yearEnd = str(request.form['yearEnd'])
# 	# processed_text = client.upper()
# 	# fisier=request.files.get('monthlyTB')
	if request.method == 'POST':
		def getAttachments(reader):
			catalog = reader.trailer["/Root"]
			fileNames = catalog['/Names']['/EmbeddedFiles']['/Names']
			attachments = {}
			for f in fileNames:
				if isinstance(f, str):
					name = f
					dataIndex = fileNames.index(f) + 1
					fDict = fileNames[dataIndex].getObject()
					fData = fDict['/EF']['/F'].getData()  
					attachments[name] = fData

			return attachments
		# file_TemplateXML = request.files('vatXML')
		file_TemplateXML = request.files.getlist('vatXML')      
		file_TB = request.files["TB"]

		# for i in file_TemplateXML:
		# 	i.save(secure_filename(i.filename))

		# fonts and colors
		ft1 = Font(name='Arial', size=10, bold=True)
		f_testname = Font(name='Arial', size=15, color='614C77', bold=True)
		f_info = Font(name='Arial', size=10, color='614C77', bold=True)
		cap_tabel = Font(name='Arial', size=10, color="FFFFFF", bold=True)
		cap_tabel_color_PBC = PatternFill(start_color='808080', end_color='808080', fill_type='solid')  # grey
		cap_tabel_color_GT = PatternFill(start_color='00AEAC', end_color='00AEAC', fill_type='solid')  # indigo #B1A0C7
		fprocentaj = Font(name='Arial', size=10, color="FF0000", bold=True)
		font_worksheet = Font(name='Arial', size=10)
		check_font = Font(name='Arial', size=10, color="FF0000", bold=True)
		check_font_1 = Font(name='Arial', size=10, color="FF0000", bold=False)
		cap_tabel_color_GT_movdeschis = PatternFill(start_color='00AEAC', end_color='00AEAC', fill_type='solid')
		cap_tabel_color_GT_movinchis = PatternFill(start_color='3BBCCA', end_color='3BBCCA', fill_type='solid')
		blue_bold_font = Font(name='Arial', size=10, color="0070C0", bold=True)
		blue_thin_font = Font(name='Arial', size=10, color="0070C0", bold=False)
		# TB_font = Font(name='Arial', size=10, color='0070C0', bold=True)

		thin = Side(border_style='thin', color='000000')
		border = Border(left=thin, right=thin, top=thin, bottom=thin)

		thin = Side(border_style='thin', color='000000')
		border_left = Border(left=thin, right=None, top=thin, bottom=thin)

		thin = Side(border_style='thin', color='000000')
		border_right = Border(left=None, right=thin, top=thin, bottom=thin)

		thin = Side(border_style='thin', color='000000')
		border_centered = Border(left=None, right=None, top=thin, bottom=thin)

		thin = Side(border_style='thin', color='000000')
		border_upperleft = Border(left=thin, top=thin)

		thin = Side(border_style='thin', color='000000')
		border_lowerleft = Border(left=thin, right=None, top=None, bottom=thin)

		thin = Side(border_style='thin', color='000000')
		border_upperright = Border(right=thin, top=thin)

		thin = Side(border_style='thin', color='000000')
		border_lowerright = Border(right=thin, bottom=thin)

		thin = Side(border_style='thin', color='000000')
		border_left1 = Border(left=thin)

		thin = Side(border_style='thin', color='000000')
		border_right1 = Border(right=thin)

		thin = Side(border_style='thin', color='000000')
		border_top = Border(top=thin)

		thin = Side(border_style='thin', color='000000')
		border_bottom = Border(bottom=thin)

		# app.mainloop()

		wb = openpyxl.Workbook()
		ws = wb.active

		Sheet1 = wb.create_sheet("Lead Schedule")
		Sheet1.sheet_view.showGridLines = False

		Sheet1.cell(row = 1, column = 2).value ="Client Name:"
		Sheet1.cell(row = 2, column = 2).value ="Period End:"
		Sheet1.cell(row = 1, column = 15).value ="Prepared By:"
		Sheet1.cell(row = 2, column = 15).value ="Date:"
		Sheet1.cell(row = 3, column = 15).value ="Reviewed by:"
		Sheet1.cell(row = 5, column = 2).value ="Account Reconciliation & SA "

		Sheet1.cell(row=7, column=2).value = "Class"
		Sheet1.cell(row=7, column=3).value = "Synt 3"
		Sheet1.cell(row=7, column=4).value = "Synt 4"
		Sheet1.cell(row=7, column=5).value = "Account"
		Sheet1.cell(row=7, column=6).value = "Descriere"
		Sheet1.cell(row=7, column=7).value = "OB"
		Sheet1.cell(row=7, column=8).value = "DR"
		Sheet1.cell(row=7, column=9).value = "CR"
		Sheet1.cell(row=7, column=10).value = "CB"
		Sheet1.cell(row=7, column=11).value = "Variation"
		Sheet1.cell(row=7, column=12).value = "Variation %"

		Sheet1.cell(row=7, column=2).value = "Class"
		Sheet1.cell(row=7, column=3).value = "Synt 3"
		Sheet1.cell(row=7, column=4).value = "Synt 4"
		Sheet1.cell(row=7, column=5).value = "Account"
		Sheet1.cell(row=7, column=6).value = "Descriere"

		Sheet1.cell(row=8, column=2).value ="4"
		Sheet1.cell(row=9, column=2).value ="4"
		Sheet1.cell(row=10, column=2).value ="4"
		Sheet1.cell(row=11, column=2).value ="4"

		Sheet1.cell(row=8, column=3).value ="442"
		Sheet1.cell(row=9, column=3).value ="442"
		Sheet1.cell(row=10, column=3).value ="442"
		Sheet1.cell(row=11, column=3).value ="442"

		Sheet1.cell(row=8, column=4).value ="4423"
		Sheet1.cell(row=9, column=4).value ="4426"
		Sheet1.cell(row=10, column=4).value ="4427"
		Sheet1.cell(row=11, column=4).value ="4428"

		Sheet1.cell(row=8, column=5).value ="Taxa pe valoarea adaugata de plata"
		Sheet1.cell(row=9, column=5).value ="Taxa pe valoarea adaugata deductibil"
		Sheet1.cell(row=10, column=5).value ="Taxa pe valoarea adaugata colectat"
		Sheet1.cell(row=11, column=5).value ="Taxa pe valoarea adaugata neexigibila"

		Sheet1.cell(row=12, column=6).value = "Total"

		#content
		Sheet1.cell(row=8, column=7).value = "=SUMIF('TB Robot'!B:B,4423,'TB Robot'!E:E)".format(16)
		Sheet1.cell(row=9, column=7).value = "=SUMIF('TB Robot'!B:B,4426,'TB Robot'!E:E)".format(16)
		Sheet1.cell(row=10, column=7).value = "=SUMIF('TB Robot'!B:B,4427,'TB Robot'!E:E)".format(16)
		Sheet1.cell(row=11, column=7).value = "=SUMIF('TB Robot'!B:B,4428,'TB Robot'!E:E)".format(16)
		Sheet1.cell(row=12, column=7).value = "=SUM(G8:G11)".format(16)

		Sheet1.cell(row=8, column=8).value = "=SUMIF('TB Robot'!B:B,4423,'TB Robot'!F:F)".format(16)
		Sheet1.cell(row=9, column=8).value = "=SUMIF('TB Robot'!B:B,4426,'TB Robot'!F:F)".format(16)
		Sheet1.cell(row=10, column=8).value = "=SUMIF('TB Robot'!B:B,4427,'TB Robot'!F:F)".format(16)
		Sheet1.cell(row=11, column=8).value = "=SUMIF('TB Robot'!B:B,4428,'TB Robot'!F:F)".format(16)
		Sheet1.cell(row=12, column=8).value = "=SUM(H8:H11)".format(16)

		Sheet1.cell(row=8, column=9).value = "=SUMIF('TB Robot'!B:B,4423,'TB Robot'!G:G)".format(16)
		Sheet1.cell(row=9, column=9).value = "=SUMIF('TB Robot'!B:B,4426,'TB Robot'!G:G)".format(16)
		Sheet1.cell(row=10, column=9).value = "=SUMIF('TB Robot'!B:B,4427,'TB Robot'!G:G)".format(16)
		Sheet1.cell(row=11, column=9).value = "=SUMIF('TB Robot'!B:B,4428,'TB Robot'!G:G)".format(16)
		Sheet1.cell(row=12, column=9).value = "=SUM(I8:I11)".format(16)

		Sheet1.cell(row=8, column=10).value = "=SUMIF('TB Robot'!B:B,4423,'TB Robot'!H:H)".format(16)
		Sheet1.cell(row=9, column=10).value = "=SUMIF('TB Robot'!B:B,4426,'TB Robot'!H:H)".format(16)
		Sheet1.cell(row=10, column=10).value = "=SUMIF('TB Robot'!B:B,4427,'TB Robot'!H:H)".format(16)
		Sheet1.cell(row=11, column=10).value = "=SUMIF('TB Robot'!B:B,4428,'TB Robot'!H:H)".format(16)
		Sheet1.cell(row=12, column=10).value = "=SUM(J8:J11)".format(16)

		Sheet1.cell(row=8, column=11).value = "=J8-G8".format(16)
		Sheet1.cell(row=9, column=11).value = "=J9-G9".format(16)
		Sheet1.cell(row=10, column=11).value = "=J10-G10".format(16)
		Sheet1.cell(row=11, column=11).value = "=J11-G11".format(16)
		Sheet1.cell(row=12, column=11).value = "=SUM(K8:K11)".format(16)

		Sheet1.cell(row=8, column=12).value = "=IFERROR(K8/G8,1)".format(16)
		Sheet1.cell(row=9, column=12).value = "=IFERROR(K9/G9,1)".format(16)
		Sheet1.cell(row=10, column=12).value = "=IFERROR(K10/G10,1)".format(16)
		Sheet1.cell(row=11, column=12).value = "=IFERROR(K11/G11,1)".format(16)
		Sheet1.cell(row=12, column=12).value = "=IFERROR(K12/G12,1)".format(16)

		Sheet1.cell(row = 1, column = 3).value =clientname1
		Sheet1.cell(row = 1, column = 3).value =clientname1
		Sheet1.cell(row = 2, column = 3).value =yearEnd1
		Sheet1.cell(row = 2, column = 3).number_format = 'mm/dd/yyyy'
		Sheet1.cell(row = 1, column = 16).value =preparedBy1
		Sheet1.cell(row = 2, column = 16).value =datePrepared1
		Sheet1.cell(row = 2, column = 16).number_format = 'mm/dd/yyyy'

		#design
		Sheet1.cell(row = 1, column = 2).font =ft1
		Sheet1.cell(row = 2, column = 2).font =ft1
		Sheet1.cell(row = 1, column = 15).font =ft1
		Sheet1.cell(row = 2, column = 15).font =ft1
		Sheet1.cell(row = 3, column = 15).font =ft1
		Sheet1.cell(row = 5, column = 2).font = f_testname

		for row in Sheet1['B7:L7']:
			for cell in row:
				cell.fill = cap_tabel_color_GT

		for row in Sheet1['B7:L7']:
			for cell in row:
				cell.font = cap_tabel

		for row in Sheet1['B8:L11']:
			for cell in row:
				cell.font = font_worksheet

		for row in Sheet1['G8:L12']:
			for cell in row:
				cell.number_format = '#,##0_);(#,##0)'

		for row in Sheet1['B7:L11']:
			for cell in row:
				cell.border = border

		Sheet1.column_dimensions['F'].hidden = True
		Sheet1.column_dimensions['F'].width = 40
		Sheet1.column_dimensions['G'].width = 10
		Sheet1.column_dimensions['H'].width = 10
		Sheet1.column_dimensions['I'].width = 10
		Sheet1.column_dimensions['J'].width = 10
		Sheet1.column_dimensions['K'].width = 14
		Sheet1.column_dimensions['L'].width = 14

		Sheet2 = wb.create_sheet("T10 VAT Test")
		Sheet2.sheet_view.showGridLines = False

		Sheet2.cell(row = 1, column = 2).value ="Client Name:"
		Sheet2.cell(row = 2, column = 2).value ="Period End:"
		Sheet2.cell(row = 1, column = 10).value ="Prepared By:"
		Sheet2.cell(row = 2, column = 10).value ="Date:"
		Sheet2.cell(row = 3, column = 10).value ="Reviewed by:"

		Sheet2.cell(row = 5, column = 2).value ="VAT Reconciliation Summary"
		Sheet2.cell(row = 7, column = 2).value ="Procedures"
		Sheet2.cell(row = 8, column = 2).value ="We have obtained the monthly VAT statement prepared for Local Budget (no. 300)."
		Sheet2.cell(row = 9, column = 2).value ="We have reconciled the VAT from the VAT statement and VAT from the sales and acquisition journals with the value from trial balance."

		Sheet2.cell(row = 11, column = 2).value ="Work Done"
		Sheet2.cell(row = 12, column = 2).value ="Please see below:"

		Sheet2.cell(row = 16, column = 2).value = "Month"
		Sheet2.cell(row = 16, column = 3).value = "Refference"
		Sheet2.cell(row = 16, column = 5).value = "Sales 19% as per VAT Statement"
		Sheet2.cell(row = 16, column = 6).value = "VAT collected 19% as per Nexia"
		Sheet2.cell(row = 16, column = 7).value = "Sales 9% as per VAT Statement"
		Sheet2.cell(row = 16, column = 8).value = "VAT collected 9% as per Nexia"
		Sheet2.cell(row = 16, column = 9).value = "Sales 5% as per VAT Statement"
		Sheet2.cell(row = 16, column = 10).value = "VAT collected 5% as per Nexia"
		Sheet2.cell(row = 16, column = 11).value = "Sales w/o VAT as per VAT Statement"
		Sheet2.cell(row = 16, column = 12).value = "Reverse Taxation as per VAT Statement"
		Sheet2.cell(row = 16, column = 13).value = "VAT collected for Reverse Taxation as per Nexia (19%)"
		Sheet2.cell(row = 16, column = 14).value = "Regularisation as per VAT Statement"
		Sheet2.cell(row = 16, column = 15).value = "VAT collected for Regularization as per Nexia (19%/9%/5%)"
		Sheet2.cell(row = 16, column = 17).value = "Total sales as per VAT Statement"
		Sheet2.cell(row = 16, column = 18).value = "Total sales as per Nexia Calculation"
		Sheet2.cell(row = 16, column = 19).value = "Difference on Total Sales"
		Sheet2.cell(row = 16, column = 21).value ="Total VAT Collected as per VAT Statement"
		Sheet2.cell(row = 16, column = 22).value ="Total VAT collected as per Nexia"
		Sheet2.cell(row = 16, column = 23).value ="Difference on VAT Collected Company vs Nexia"
		Sheet2.cell(row = 16, column = 25).value ="Purchases  19% as per VAT Statement"
		Sheet2.cell(row = 16, column = 26).value ="VAT deductible  as per Nexia"
		Sheet2.cell(row = 16, column = 27).value ="Purchases 9% as per VAT Statement"
		Sheet2.cell(row = 16, column = 28).value ="VAT 9% as per Nexia"
		Sheet2.cell(row = 16, column = 29).value ="Purchases 5% as per VAT Statement"
		Sheet2.cell(row = 16, column = 30).value ="VAT 5% as per Nexia"
		Sheet2.cell(row = 16, column = 31).value ="Reverse Taxation as per VAT Statement"
		Sheet2.cell(row = 16, column = 32).value ="VAT deducted for Reverse Taxation as per Nexia (19%)"
		Sheet2.cell(row = 16, column = 33).value ="Total Purchases w/o VAT as VAT per Statement"
		Sheet2.cell(row = 16, column = 34).value ="Regularisation as per VAT Statement"
		Sheet2.cell(row = 16, column = 35).value ="VAT deducted for Regularization as per Nexia (19%/9%/5%)"
		Sheet2.cell(row = 16, column = 37).value ="Total Purchases with VAT as per VAT Statement"
		Sheet2.cell(row = 16, column = 38).value ="Total Purchases with VAT as per Nexia"
		Sheet2.cell(row = 16, column = 39).value ="Difference on Total Purchases with VAT Company vs Nexia"
		Sheet2.cell(row = 16, column = 41).value ="Total VAT deductible as per VAT Statement"
		Sheet2.cell(row = 16, column = 42).value ="Total VAT deductible as per Nexia"
		Sheet2.cell(row = 16, column = 43).value ="Difference on Total VAT deductible Company vs Nexia"
		Sheet2.cell(row = 16, column = 45).value ="VAT (due)/to be recovered per client"
		Sheet2.cell(row = 16, column = 46).value ="VAT (due)/to be recovered per Nexia"
		Sheet2.cell(row = 16, column = 47).value ="Difference Company vs Nexia"
		Sheet2.cell(row = 16, column = 48).value ="VAT deducted as per statement"
		Sheet2.cell(row = 16, column = 49).value ="VAT deductible vs VAT deducted"
		Sheet2.cell(row = 16, column = 50).value ="Remaining difference statement vs Nexia"
		Sheet2.cell(row = 16, column = 52).value ="Payment order/ Request for reimbursement or compensation no/date during the month"
		Sheet2.cell(row = 16, column = 53).value ="Amount "
		Sheet2.cell(row = 16, column = 54).value ="VAT balance"

		Sheet2.cell(row=34, column=19).value ="Credit Movement of account #4427 as per T/B"
		Sheet2.cell(row=35, column=19).value ="Difference"

		Sheet2.cell(row=34, column=39).value = "Debit Movement of account #4426 as per T/B"
		Sheet2.cell(row=35, column=39).value = "Difference Detail vs TB"

		Sheet2.cell(row=37, column=39).value = "Debit Movement of account #4423 as per T/B"
		Sheet2.cell(row=38, column=39).value = "Difference Detail vs TB"

		Sheet2.cell(row=43, column=6).value = "Total Sales excluding Reverse Taxation as per VAT Statement"
		Sheet2.cell(row=44, column=6).value = "Total Sales as per TB"
		Sheet2.cell(row=45, column=6).value = "Mvm 70x"
		Sheet2.cell(row=46, column=6).value = "Variation 419 (Cr-Db)"
		Sheet2.cell(row=47, column=6).value = "Variation 418 -(Db-Cr)"
		Sheet2.cell(row=48, column=6).value = "Variation 472 (Cr-Db)"
		Sheet2.cell(row=49, column=6).value = "Mvm 758X"
		Sheet2.cell(row=50, column=6).value = "Total"
		Sheet2.cell(row=52, column=5).value ="Difference"
		Sheet2.cell(row=53, column=5).value ="Maximum impact on VAT"

		Sheet2.cell(row=58, column=6).value = "Total Purchases as per VAT Statement"
		Sheet2.cell(row=59, column=6).value = "Total Purchases as per TB"
		Sheet2.cell(row=60, column=6).value = "Mvm 60x "
		Sheet2.cell(row=61, column=6).value = "Variation #3xx other than #39x"
		Sheet2.cell(row=62, column=6).value = "FA acqusitions"
		Sheet2.cell(row=63, column=6).value = "Mvm 61X"
		Sheet2.cell(row=64, column=6).value = "Mvm 62x"
		Sheet2.cell(row=65, column=6).value = "Mvm 658X"
		Sheet2.cell(row=66, column=6).value = "Total"

		Sheet2.cell(row=68, column=5).value = "Difference"
		Sheet2.cell(row=69, column=5).value = "Maximum impact on VAT"

		#content
		Sheet2.cell(row = 1, column = 3).value =clientname1
		Sheet2.cell(row = 2, column = 3).value =yearEnd1
		Sheet2.cell(row = 2, column = 3).number_format='mm/dd/yyyy'
		Sheet2.cell(row = 1, column = 11).value =preparedBy1
		Sheet2.cell(row = 2, column = 11).value =datePrepared1
		Sheet2.cell(row = 2, column = 11).number_format='mm/dd/yyyy'

		#retinem variabilele din XML
		files=list(file_TemplateXML)
		def changeWord(word):
			for letter in word:
				if letter == "b'":
					word = word.replace(letter,'')
			return word
		nr=0
		abc=[]
		now=datetime.datetime.now()
		for i in files:
			nr=nr+1
			reader = PyPDF2.PdfFileReader(i)
			dictionary = getAttachments(reader)
			a=str(*dictionary.values())
			b=a.replace('\\n','').replace("b'","").replace("></declaratie300>'","></declaratie300>").replace('encoding="utf-8"', "").replace('''"/>''', "</declaratie300>").replace("</declaratie300>'", '"></declaratie300>').replace('\\r', '')
			# print(b)
			f=open("home/auditappnexia/output/vat/"+str(nr)+str(now.year)+str(now.month)+str(now.day)+str(now.hour)+str(now.minute)+str(now.second)+".xml","w").write(b)
			abc.append("home/auditappnexia/output/vat/"+str(nr)+str(now.year)+str(now.month)+str(now.day)+str(now.hour)+str(now.minute)+str(now.second)+".xml")
			
		fisiere=list(abc)
		for v in range(0,len(fisiere)):
			tree = ET.parse(fisiere[v])
			root = tree.getroot()
			month = int(root.attrib["luna"])
			print(month)
			for j in range(1, month+1):
				Sheet2.cell(row=16 + j, column=2).value = j
				# Sheet2.cell(row=36 + j, column=2).value = j
				# Sheet2.cell(row=54 + j, column=2).value = j
			# a=0
			# try:
			# 	a = a + int(root.attrib['R9_1'])
			# except:
			# 	print("n a mers")
			#
			# try:
			# 	a = a + int(root.attrib['R17_1'])
			# except:
			# 	print("n a mers 2")

			try:
				Sheet2.cell(row=16 + month, column=5).value = int(root.attrib['R9_1'])
			except:
				Sheet2.cell(row=16 + month, column=5).value = 0


			# Sheet2.cell(row=15+i, column=2).value = int(root.attrib["luna"])

			Sheet2.cell(row = 16 + month, column = 6).value = "=(E{0}*0.19)*1".format(16 + month)


			try:
				Sheet2.cell(row = 16 + month, column = 7).value = int(root.attrib['R10_1'])
			except:
				Sheet2.cell(row = 16 + month, column = 7).value = 0

			Sheet2.cell(row = 16 + month, column = 8).value = "=(G{0}*0.09)*1".format(16 + month)

			try:
				Sheet2.cell(row = 16 + month, column = 9).value = int(root.attrib['R11_1'])
			except:
				Sheet2.cell(row = 16 + month, column = 9).value = 0

			Sheet2.cell(row = 16 + month, column = 10).value = "=(I{0}*0.05)*1".format(16 + month)

			x=0
			try:
				x = x + int(root.attrib['R1_1'])
			except:
				print('b')

			try:
				x = x + int(root.attrib['R2_1'])
			except:
				print('b')

			try:
				x = x + int(root.attrib['R3_1'])
			except:
				print('b')

			try:
				x = x + int(root.attrib['R4_1'])
			except:
				print('b')

			try:
				x = x + int(root.attrib['R13_1'])
			except:
				print('b')

			try:
				x = x + int(root.attrib['R14_1'])
			except:
				print('b')

			try:
				x = x + int(root.attrib['R15_1'])
			except:
				print('b')

			Sheet2.cell(row = 16 + month, column = 11).value = x

			x = 0
			# try:
			# 	x = x + int(root.attrib['R5_1'])
			# except:
			# 	print('5_1')

			try:
				x = x + int(root.attrib['R6_1'])
			except:
				print('6_1')

			try:
				x = x + int(root.attrib['R7_1'])
			except:
				print('7_1')

			try:
				x = x + int(root.attrib['R8_1'])
			except:
				print('8_1')

			try:
				x = x + int(root.attrib['R12_1'])
			except:
				print('12_1')
			Sheet2.cell(row = 16 + month, column = 12).value = x
			Sheet2.cell(row = 16 + month, column = 13).value = "=(L{0}*0.19)*1".format(16 + month)

			b= 0
			try:
				b = b + int(root.attrib['R16_1'])
			except:
				print("nu merge 16_1")

			try:
				b = b + int(root.attrib['R18_1'])
			except:
				print("nu merge 18_1")

			try:
				Sheet2.cell(row = 16 + month, column = 14).value = b
			except:
				Sheet2.cell(row = 16 + month, column = 14).value = 0

			Sheet2.cell(row = 16 + month, column = 15).value = "=(N{0}*0.19)*1".format(16 + month)

			try:
				Sheet2.cell(row = 16 + month, column = 17).value = int(root.attrib['R17_1']) #C1
			except:
				Sheet2.cell(row = 16 + month, column = 17).value = int(root.attrib['R19_1'])

			Sheet2.cell(row = 16 + month, column = 18).value = "=(E{0}+G{0}+K{0}+L{0}+N{0}+I{0})*1".format(16 + month)
			Sheet2.cell(row = 16 + month, column = 19).value = "=Q{0}-R{0}".format(16 + month)

			try:
				Sheet2.cell(row = 16 + month, column = 21).value = int(root.attrib['R17_2']) #C2
			except:
				Sheet2.cell(row = 16 + month, column = 21).value = 0

			Sheet2.cell(row = 16 + month, column = 22).value = "=(F{0}+J{0}+H{0}+M{0}+O{0})*1".format(16 + month)
			Sheet2.cell(row = 16 + month, column = 23).value = "=(U{0}-V{0})*1".format(16 + month)

			try:
				Sheet2.cell(row = 16 + month, column = 25).value = int(root.attrib['R22_1'])
			except:
				Sheet2.cell(row = 16 + month, column = 25).value = 0
			Sheet2.cell(row = 16 + month, column = 26).value = "=(Y{0}*0.19)*1".format(16 + month)

			try:
				Sheet2.cell(row = 16 + month, column = 27).value = int(root.attrib['R23_1'])
			except:
				Sheet2.cell(row = 16 + month, column = 27).value = 0
			Sheet2.cell(row = 16 + month, column = 28).value = "=(AA{0}*0.09)*1".format(16 + month)

			try:
				Sheet2.cell(row = 16 + month, column = 29).value = int(root.attrib['R24_1'])
			except:
				Sheet2.cell(row = 16 + month, column = 29).value = 0
			Sheet2.cell(row = 16 + month, column = 30).value = "=(AC{0}*0.05)*1".format(16 + month)

			x = 0
			try:
				x = x + int(root.attrib['R18_1'])
			except:
				print('a')

			try:
				x = x + int(root.attrib['R19_1'])
			except:
				print('a')

			try:
				x = x + int(root.attrib['R20_1'])
			except:
				print('a')
			try:
				x = x + int(root.attrib['R21_1'])
			except:
				print('a')

			try:
				x = x + int(root.attrib['R25_1'])
			except:
				print('a')

			Sheet2.cell(row = 16 + month, column = 31).value = x
			Sheet2.cell(row = 16 + month, column = 32).value = "=(AE{0}*0.19)*1".format(16 + month)

			try:
				Sheet2.cell(row = 16 + month, column = 33).value = int(root.attrib['R26_1'])
			except:
				Sheet2.cell(row = 16 + month, column = 33).value = 0

			try:
				Sheet2.cell(row = 16 + month, column = 34).value = int(root.attrib['R30_2'])
			except:
				Sheet2.cell(row = 16 + month, column = 34).value = 0
			try:
				Sheet2.cell(row=16 + month, column=35).value = int(root.attrib['R30_2'])
			except:
				Sheet2.cell(row=16 + month, column=35).value = 0

			try:
				Sheet2.cell(row = 16 + month, column = 37).value = int(root.attrib['R27_1']) #C1
			except:
				Sheet2.cell(row = 16 + month, column = 37).value = 0

			Sheet2.cell(row = 16 + month, column = 38).value = "=(Y{0}+AA{0}+AE{0}+AC{0})*1".format(16 + month)
			Sheet2.cell(row = 16 + month, column = 39).value = "=(ABS(AK{0}-AL{0}))*1".format(16 + month)

			try:
				Sheet2.cell(row = 16 + month, column = 41).value = int(root.attrib['R27_2']) #C2
			except:
				Sheet2.cell(row = 16 + month, column = 41).value = 0

			Sheet2.cell(row = 16 + month, column = 42).value = "=(Z{0}+AB{0}+AF{0}+AD{0})*1".format(16 + month)
			Sheet2.cell(row = 16 + month, column = 43).value = "=(AO{0}-AP{0})*1".format(16 + month)

			# try:
			# 	Sheet2.cell(row = 16 + month, column = 45).value = int(root.attrib['R33_2'])
			# except:
			# 	if int(root.attrib['R33_2']) == 0:
			# 		Sheet2.cell(row=16 + month, column=45).value = int(root.attrib['R34_2'])
			# 	else:
			# 		Sheet2.cell(row=16 + month, column=45).value = 0

			# Sheet2.cell(row = 16 + month, column = 45).value = int(root.attrib['R33_2'])
			# if int(root.attrib['R33_2']) == 0:
			# 	Sheet2.cell(row=16 + month, column=45).value = int(-int(root.attrib['R34_2']))
			# else:
			# 	Sheet2.cell(row=16 + month, column=45).value = int(root.attrib['R33_2'])

			try:
				Sheet2.cell(row=16 + month, column=45).value = int(-int(root.attrib['R34_2']))
			except:
				Sheet2.cell(row=16 + month, column=45).value = int(root.attrib['R33_2'])

			Sheet2.cell(row = 16 + month, column = 46).value = "=(AP{0}-V{0}+AI{0})*1".format(16 + month)
			Sheet2.cell(row = 16 + month, column = 47).value = "=(AS{0}-AT{0})*1".format(16 + month)

			try:
				Sheet2.cell(row=16 + month, column=48).value = int(root.attrib['R28_2'])
			except:
				Sheet2.cell(row=16 + month, column=48).value = 0

			Sheet2.cell(row=16 + month, column=49).value ="=AP{0}-AV{0}".format(16 + month)
			Sheet2.cell(row=16 + month, column=50).value = "=AU{0}+AW{0}".format(16 + month)

		#calculam totalurile
			Sheet2.cell(row = 30, column = 5).value = "=sum(E17:E28)".format(16 + month)
			Sheet2.cell(row = 30, column = 6).value = "=sum(F17:F28)".format(16 + month)
			Sheet2.cell(row = 30, column = 7).value = "=sum(G17:G28)".format(16 + month)
			Sheet2.cell(row = 30, column = 8).value = "=sum(H17:H28)".format(16 + month)
			Sheet2.cell(row = 30, column = 9).value = "=sum(I17:I28)".format(16 + month)
			Sheet2.cell(row = 30, column = 10).value = "=sum(J17:J28)".format(16 + month)
			Sheet2.cell(row = 30, column = 11).value = "=sum(K17:K28)".format(16 + month)
			Sheet2.cell(row = 30, column = 12).value = "=sum(L17:L28)".format(16 + month)
			Sheet2.cell(row = 30, column = 13).value = "=sum(M17:M28)".format(16 + month)
			Sheet2.cell(row = 30, column = 14).value = "=sum(N17:N28)".format(16 + month)
			Sheet2.cell(row = 30, column = 15).value = "=sum(O17:O28)".format(16 + month)
			Sheet2.cell(row = 30, column = 17).value = "=sum(Q17:Q28)".format(16 + month)
			Sheet2.cell(row = 30, column = 18).value = "=sum(R17:R28)".format(16 + month)
			Sheet2.cell(row = 30, column = 19).value = "=sum(S17:S28)".format(16 + month)
			Sheet2.cell(row = 30, column = 21).value = "=sum(U17:U28)".format(16 + month)
			Sheet2.cell(row = 30, column = 22).value = "=sum(V17:V28)".format(16 + month)
			Sheet2.cell(row = 30, column = 23).value = "=sum(W17:W28)".format(16 + month)
			Sheet2.cell(row = 30, column = 25).value = "=sum(Y17:Y28)".format(16 + month)
			Sheet2.cell(row = 30, column = 26).value = "=sum(Z17:Z28)".format(16 + month)
			Sheet2.cell(row = 30, column = 27).value = "=sum(AA17:AA28)".format(16 + month)
			Sheet2.cell(row = 30, column = 28).value = "=sum(AB17:AB28)".format(16 + month)
			Sheet2.cell(row = 30, column = 29).value = "=sum(AC17:AC28)".format(16 + month)
			Sheet2.cell(row = 30, column = 30).value = "=sum(AD17:AD28)".format(16 + month)
			Sheet2.cell(row = 30, column = 31).value = "=sum(AE17:AE28)".format(16 + month)
			Sheet2.cell(row = 30, column = 32).value = "=sum(AF17:AF28)".format(16 + month)
			Sheet2.cell(row = 30, column = 33).value = "=sum(AG17:AG28)".format(16 + month)
			Sheet2.cell(row = 30, column = 34).value = "=sum(AH17:AH28)".format(16 + month)
			Sheet2.cell(row = 30, column = 35).value = "=sum(AI17:AI28)".format(16 + month)
			Sheet2.cell(row = 30, column = 37).value = "=sum(AK17:AK28)".format(16 + month)
			Sheet2.cell(row = 30, column = 38).value = "=sum(AL17:AL28)".format(16 + month)
			Sheet2.cell(row = 30, column = 39).value = "=sum(AM17:AM28)".format(16 + month)
			Sheet2.cell(row = 30, column = 41).value = "=sum(AO17:AO28)".format(16 + month)
			Sheet2.cell(row = 30, column = 42).value = "=sum(AP17:AP28)".format(16 + month)
			Sheet2.cell(row = 30, column = 43).value = "=sum(AQ17:AQ28)".format(16 + month)
			Sheet2.cell(row = 30, column = 45).value = "=sum(AS17:AS28)".format(16 + month)
			Sheet2.cell(row = 30, column = 46).value = "=sum(AT17:AT28)".format(16 + month)
			Sheet2.cell(row = 30, column = 47).value = "=sum(AU17:AU28)".format(16 + month)
			Sheet2.cell(row = 30, column = 48).value = "=sum(AV17:AV28)".format(16 + month)
			Sheet2.cell(row = 30, column = 49).value = "=sum(AW17:AW28)".format(16 + month)
			Sheet2.cell(row = 30, column = 50).value = "=sum(AX17:AX28)".format(16 + month)

			Sheet2.cell(row=34, column=21).value = "=SUMIF('Lead Schedule'!D:D,4427,'Lead Schedule'!I:I)".format(16 + month)
			Sheet2.cell(row=35, column=21).value = "=V30-U34".format(16 + month)

			Sheet2.cell(row=34, column=41).value = "=SUMIF('Lead Schedule'!D:D,4426,'Lead Schedule'!I:I)".format(16 + month)
			Sheet2.cell(row=35, column=41).value = "=AP30-AO34".format(16 + month)

			Sheet2.cell(row=37, column=41).value = "=SUMIF('Lead Schedule'!D:D,4423,'Lead Schedule'!I:I)".format(16 + month)
			Sheet2.cell(row=38, column=41).value = "=AS30+AO37".format(16 + month)

			Sheet2.cell(row=43, column=7).value = "=Q30-L30".format(16 + month)
			Sheet2.cell(row=44, column=7).value = "=SUM(G45:G49)".format(16 + month)
			Sheet2.cell(row=45, column=7).value = "=-SUMIF('TB Robot'!I:I,70,'TB Robot'!H:H)".format(16 + month)
			Sheet2.cell(row=46, column=7).value = "=SUMIF('TB Robot'!A:A,419,'TB Robot'!G:G)-SUMIF('TB Robot'!A:A,419,'TB Robot'!F:F)".format(16 + month)
			Sheet2.cell(row=47, column=7).value = "=SUMIF('TB Robot'!A:A,418,'TB Robot'!G:G)-SUMIF('TB Robot'!A:A,418,'TB Robot'!F:F)".format(16 + month)
			Sheet2.cell(row=48, column=7).value = "=SUMIF('TB Robot'!A:A,472,'TB Robot'!G:G)-SUMIF('TB Robot'!A:A,472,'TB Robot'!F:F)".format(16 + month)
			Sheet2.cell(row=49, column=7).value = "=SUMIF('TB Robot'!A:A,758,'TB Robot'!H:H)".format(16 + month)
			Sheet2.cell(row=50, column=7).value = "=SUM(G45:G49)".format(16 + month)
			Sheet2.cell(row=52, column=7).value = "=G43-G44".format(16 + month)
			Sheet2.cell(row=53, column=7).value = "=G52*19%".format(16 + month)

			Sheet2.cell(row=58, column=7).value = "=AL30".format(16 + month)
			Sheet2.cell(row=59, column=7).value = "=SUM(G60:G65)".format(16 + month)
			Sheet2.cell(row=60, column=7).value = "=-SUMIF('TB Robot'!I:I,60,'TB Robot'!H:H)".format(16 + month)
			Sheet2.cell(row=61, column=7).value = "=SUMIF('TB Robot'!J:J,3,'TB Robot'!H:H)-SUMIF('TB Robot'!I:I,39,'TB Robot'!H:H)".format(16 + month)
			Sheet2.cell(row=62, column=7).value = "0".format(16 + month)
			Sheet2.cell(row=63, column=7).value = "=SUMIF('TB Robot'!I:I,61,'TB Robot'!H:H)".format(16 + month)
			Sheet2.cell(row=64, column=7).value = "=SUMIF('TB Robot'!I:I,62,'TB Robot'!H:H)".format(16 + month)
			Sheet2.cell(row=65, column=7).value = "=SUMIF('TB Robot'!A:A,658,'TB Robot'!H:H)".format(16 + month)
			# Sheet2.cell(row=66, column=7).value = "Total"

			Sheet2.cell(row=68, column=7).value = "=G58-G59".format(16 + month)
			Sheet2.cell(row=69, column=7).value = "=G68*19%".format(16 + month)

		#design
		Sheet2.cell(row = 1, column = 2).font =ft1
		Sheet2.cell(row = 2, column = 2).font =ft1
		Sheet2.cell(row = 1, column = 10).font =ft1
		Sheet2.cell(row = 2, column = 10).font =ft1
		Sheet2.cell(row = 3, column = 10).font =ft1

		Sheet2.cell(row = 5, column = 2).font =f_testname
		Sheet2.cell(row = 7, column = 2).font =ft1
		Sheet2.cell(row = 11, column = 2).font =ft1
		Sheet2.cell(row = 12, column = 2).font =ft1

		for row in Sheet2['B17:BB30']:
			for cell in row:
				cell.font = font_worksheet

		for row in Sheet2['B16:C16']:
			for cell in row:
				cell.fill = cap_tabel_color_GT_movinchis

		Sheet2.cell(row=16, column=5).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=7).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=9).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=11).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=12).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=14).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=17).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=21).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=25).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=27).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=29).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=31).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=33).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=34).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=37).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=41).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=45).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=48).fill=cap_tabel_color_GT_movinchis

		Sheet2.cell(row=16, column=6).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=8).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=10).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=13).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=15).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=18).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=19).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=22).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=23).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=26).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=28).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=30).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=32).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=35).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=38).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=39).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=42).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=43).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=46).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=47).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=49).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=50).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=52).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=53).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=54).fill=cap_tabel_color_GT_movdeschis

		#tables
		for row in Sheet2['A30:AW30']:
			for cell in row:
				cell.font = ft1

		for row in Sheet2['B16:BB16']:
			for cell in row:
				cell.font = cap_tabel

		for row in Sheet2['B17:AY30']:
			for cell in row:
				cell.number_format = '#,##0_);(#,##0)'

		for row in Sheet2['B16:C28']:
			for cell in row:
				cell.border = border

		for row in Sheet2['E16:O28']:
			for cell in row:
				cell.border = border

		for row in Sheet2['Q16:S28']:
			for cell in row:
				cell.border = border

		for row in Sheet2['U16:W28']:
			for cell in row:
				cell.border = border

		for row in Sheet2['Y16:AI28']:
			for cell in row:
				cell.border = border

		for row in Sheet2['AK16:AM28']:
			for cell in row:
				cell.border = border

		for row in Sheet2['AO16:AQ28']:
			for cell in row:
				cell.border = border

		for row in Sheet2['AS16:AU28']:
			for cell in row:
				cell.border = border

		for row in Sheet2['AV16:AX28']:
			for cell in row:
				cell.border = border

		for row in Sheet2['AZ16:BB28']:
			for cell in row:
				cell.border = border

		#COLORS
		for row in Sheet2['S17:S28']:
			for cell in row:
				cell.font = check_font
		Sheet2.cell(row=30,column=19).font = check_font

		for row in Sheet2['W17:W28']:
			for cell in row:
				cell.font = check_font
		Sheet2.cell(row=30,column=23).font = check_font

		for row in Sheet2['AM17:AM28']:
			for cell in row:
				cell.font = check_font
		Sheet2.cell(row=30,column=39).font = check_font

		for row in Sheet2['AQ17:AQ28']:
			for cell in row:
				cell.font = check_font
		Sheet2.cell(row=30,column=43).font = check_font

		for row in Sheet2['AX17:AX30']:
			for cell in row:
				cell.font = check_font
		Sheet2.cell(row=30,column=50).font = check_font

		for row in Sheet2['E43:G69']:
			for cell in row:
				cell.font = font_worksheet

		for row in Sheet2['G43:G69']:
			for cell in row:
				cell.number_format = '#,##0_);(#,##0)'

		for row in Sheet2['E43:G44']:
			for cell in row:
				cell.font = ft1

		for row in Sheet2['F50:G50']:
			for cell in row:
				cell.font = ft1

		for row in Sheet2['E52:G52']:
			for cell in row:
				cell.font = check_font

		for row in Sheet2['F58:G59']:
			for cell in row:
				cell.font = ft1

		for row in Sheet2['F66:G66']:
			for cell in row:
				cell.font = ft1

		for row in Sheet2['F68:G68']:
			for cell in row:
				cell.font = check_font

		Sheet2.cell(row=34, column=21).number_format = '#,##0_);(#,##0)'
		Sheet2.cell(row=35, column=21).number_format = '#,##0_);(#,##0)'

		Sheet2.cell(row=34, column=41).number_format = '#,##0_);(#,##0)'
		Sheet2.cell(row=35, column=41).number_format = '#,##0_);(#,##0)'

		Sheet2.cell(row=37, column=41).number_format = '#,##0_);(#,##0)'
		Sheet2.cell(row=38, column=41).number_format = '#,##0_);(#,##0)'

		#column dimensions
		Sheet2.column_dimensions['C'].width = 20
		Sheet2.column_dimensions['E'].width = 20
		Sheet2.column_dimensions['F'].width = 20
		Sheet2.column_dimensions['G'].width = 20
		Sheet2.column_dimensions['H'].width = 20
		Sheet2.column_dimensions['I'].width = 20
		Sheet2.column_dimensions['J'].width = 20
		Sheet2.column_dimensions['K'].width = 20
		Sheet2.column_dimensions['L'].width = 20
		Sheet2.column_dimensions['M'].width = 20
		Sheet2.column_dimensions['N'].width = 20
		Sheet2.column_dimensions['O'].width = 20
		Sheet2.column_dimensions['Q'].width = 20
		Sheet2.column_dimensions['R'].width = 20
		Sheet2.column_dimensions['S'].width = 20
		Sheet2.column_dimensions['U'].width = 20
		Sheet2.column_dimensions['V'].width = 20
		Sheet2.column_dimensions['W'].width = 20
		Sheet2.column_dimensions['Y'].width = 20
		Sheet2.column_dimensions['Z'].width = 20
		Sheet2.column_dimensions['AA'].width = 20
		Sheet2.column_dimensions['AB'].width = 20
		Sheet2.column_dimensions['AC'].width = 20
		Sheet2.column_dimensions['AD'].width = 20
		Sheet2.column_dimensions['AE'].width = 20
		Sheet2.column_dimensions['AF'].width = 20
		Sheet2.column_dimensions['AG'].width = 20
		Sheet2.column_dimensions['AH'].width = 20
		Sheet2.column_dimensions['AI'].width = 20
		Sheet2.column_dimensions['AK'].width = 20
		Sheet2.column_dimensions['AL'].width = 20
		Sheet2.column_dimensions['AM'].width = 20
		Sheet2.column_dimensions['AO'].width = 20
		Sheet2.column_dimensions['AP'].width = 20
		Sheet2.column_dimensions['AQ'].width = 20
		Sheet2.column_dimensions['AS'].width = 20
		Sheet2.column_dimensions['AT'].width = 20
		Sheet2.column_dimensions['AU'].width = 20
		Sheet2.column_dimensions['AV'].width = 20
		Sheet2.column_dimensions['AW'].width = 30
		Sheet2.column_dimensions['AX'].width = 20
		Sheet2.column_dimensions['AY'].width = 20
		Sheet2.column_dimensions['AZ'].width = 20
		Sheet2.column_dimensions['BA'].width = 20
		Sheet2.column_dimensions['BB'].width = 20

		Sheet2['B16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['C16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['E16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['F16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['G16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['H16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['I16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['J16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['K16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['L16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['M16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['N16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['O16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['Q16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['R16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['S16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['U16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['V16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['W16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['Y16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['Z16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AA16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AB16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AC16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AD16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AE16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AF16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AG16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AH16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AI16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AK16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AL16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AM16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AO16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AP16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AQ16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AS16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AT16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AU16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AV16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AW16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AX16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AY16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AZ16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['BA16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['BB16'].alignment = Alignment(wrapText=True, vertical='center')

		Sheet3 = wb.create_sheet("T10.1 VAT Test")
		Sheet3.sheet_view.showGridLines = False

		Sheet3.cell(row = 1, column = 1).value ="Client Name:"
		Sheet3.cell(row = 2, column = 1).value ="Period End:"
		Sheet3.cell(row = 1, column = 14).value ="Prepared By:"
		Sheet3.cell(row = 2, column = 14).value ="Date:"
		Sheet3.cell(row = 3, column = 14).value ="Reviewed by:"

		Sheet3.cell(row = 4, column = 1).value ="VAT Reconciliation Summary"
		Sheet3.cell(row = 6, column = 1).value ="Objective:"
		Sheet3.cell(row = 8, column = 1).value ="Work Done:"
		Sheet3.cell(row = 13, column = 1).value ="Findings:"
		Sheet3.cell(row = 15, column = 1).value ="Conclusion:"

		Sheet3.cell(row = 18, column = 1).value ="As per VAT Statement (300)"
		Sheet3.cell(row = 27, column = 1).value ="As per Monthly TB"
		Sheet3.cell(row = 32, column = 1).value ="Check Statement 300 vs Monthly TB"

		Sheet3.cell(row = 37, column = 1).value ="As per Sales and Acquisitions Journals"
		Sheet3.cell(row = 43, column = 1).value ="Check Journals vs Trial Balance"
		Sheet3.cell(row = 49, column = 1).value ="Check Statement 300 vs Journals"

		Sheet3.cell(row = 6, column = 2).value ="To test the completeness of the VAT balances."
		Sheet3.cell(row = 8, column = 2).value ="We have obtained the Monthly VAT returns prepared for Local Budget (no. 300)."
		Sheet3.cell(row = 9, column = 2).value ="We have obtained the monthly sales and acquisition journals in order to extract the VAT and reconciled with the value presented in VAT statement. "
		Sheet3.cell(row = 10, column = 2).value ="We have reconciled the VAT from the VAT statement and VAT from the sales and acquisition journals with the value from trial balance."

		#bilding tables
		#table 1
		Sheet3.cell(row = 18, column = 2).value ="Type of VAT"
		Sheet3.cell(row = 18, column = 3).value ="January"
		Sheet3.cell(row = 18, column = 4).value ="February"
		Sheet3.cell(row = 18, column = 5).value ="March"
		Sheet3.cell(row = 18, column = 6).value ="April"
		Sheet3.cell(row = 18, column = 7).value ="May"
		Sheet3.cell(row = 18, column = 8).value ="June"
		Sheet3.cell(row = 18, column = 9).value ="July"
		Sheet3.cell(row = 18, column = 10).value ="August"
		Sheet3.cell(row = 18, column = 11).value ="September"
		Sheet3.cell(row = 18, column = 12).value ="October"
		Sheet3.cell(row = 18, column = 13).value ="November"
		Sheet3.cell(row = 18, column = 14).value ="December"
		Sheet3.cell(row = 18, column = 15).value ="Total"

		Sheet3.cell(row = 19, column = 2).value ="VAT Collectable"
		Sheet3.cell(row = 20, column = 2).value ="VAT Deductible"
		# Sheet3.cell(row = 21, column = 2).value ="Reverse charge"
		# Sheet3.cell(row = 22, column = 2).value ="Exempted"
		# Sheet3.cell(row = 23, column = 2).value ="Regularization collectable"
		Sheet3.cell(row = 24, column = 2).value ="Regularization deduction"
		Sheet3.cell(row = 25, column = 2).value ="VAT Payable/ recoverable"

		#table2
		Sheet3.cell(row = 27, column = 2).value ="Type of VAT"
		Sheet3.cell(row = 27, column = 3).value ="January"
		Sheet3.cell(row = 27, column = 4).value ="February"
		Sheet3.cell(row = 27, column = 5).value ="March"
		Sheet3.cell(row = 27, column = 6).value ="April"
		Sheet3.cell(row = 27, column = 7).value ="May"
		Sheet3.cell(row = 27, column = 8).value ="June"
		Sheet3.cell(row = 27, column = 9).value ="July"
		Sheet3.cell(row = 27, column = 10).value ="August"
		Sheet3.cell(row = 27, column = 11).value ="September"
		Sheet3.cell(row = 27, column = 12).value ="October"
		Sheet3.cell(row = 27, column = 13).value ="November"
		Sheet3.cell(row = 27, column = 14).value ="December"
		Sheet3.cell(row = 27, column = 15).value ="Total"

		Sheet3.cell(row = 28, column = 2).value ="VAT Collectable 4427"
		Sheet3.cell(row = 29, column = 2).value ="VAT Deductible 4426"
		Sheet3.cell(row = 30, column = 2).value ="VAT Payable/ recoverable"

		#table3
		Sheet3.cell(row = 32, column = 2).value ="Type of VAT"
		Sheet3.cell(row = 32, column = 3).value ="January"
		Sheet3.cell(row = 32, column = 4).value ="February"
		Sheet3.cell(row = 32, column = 5).value ="March"
		Sheet3.cell(row = 32, column = 6).value ="April"
		Sheet3.cell(row = 32, column = 7).value ="May"
		Sheet3.cell(row = 32, column = 8).value ="June"
		Sheet3.cell(row = 32, column = 9).value ="July"
		Sheet3.cell(row = 32, column = 10).value ="August"
		Sheet3.cell(row = 32, column = 11).value ="September"
		Sheet3.cell(row = 32, column = 12).value ="October"
		Sheet3.cell(row = 32, column = 13).value ="November"
		Sheet3.cell(row = 32, column = 14).value ="December"
		Sheet3.cell(row = 32, column = 15).value ="Total"

		Sheet3.cell(row = 33, column = 2).value ="VAT Collectable 4427"
		Sheet3.cell(row = 34, column = 2).value ="VAT Deductible 4426"
		Sheet3.cell(row = 35, column = 2).value ="VAT Payable/ recoverable"

		#table4
		Sheet3.cell(row = 37, column = 2).value ="Type of VAT"
		Sheet3.cell(row = 37, column = 3).value ="January"
		Sheet3.cell(row = 37, column = 4).value ="February"
		Sheet3.cell(row = 37, column = 5).value ="March"
		Sheet3.cell(row = 37, column = 6).value ="April"
		Sheet3.cell(row = 37, column = 7).value ="May"
		Sheet3.cell(row = 37, column = 8).value ="June"
		Sheet3.cell(row = 37, column = 9).value ="July"
		Sheet3.cell(row = 37, column = 10).value ="August"
		Sheet3.cell(row = 37, column = 11).value ="September"
		Sheet3.cell(row = 37, column = 12).value ="October"
		Sheet3.cell(row = 37, column = 13).value ="November"
		Sheet3.cell(row = 37, column = 14).value ="December"
		Sheet3.cell(row = 37, column = 15).value ="Total"

		Sheet3.cell(row = 38, column = 2).value ="VAT Collectable 4427"
		Sheet3.cell(row = 39, column = 2).value ="VAT Deductible 4426"
		Sheet3.cell(row = 40, column = 2).value ="VAT Payable/ recoverable"

		#table5
		Sheet3.cell(row = 43, column = 2).value ="Type of VAT"
		Sheet3.cell(row = 43, column = 3).value ="January"
		Sheet3.cell(row = 43, column = 4).value ="February"
		Sheet3.cell(row = 43, column = 5).value ="March"
		Sheet3.cell(row = 43, column = 6).value ="April"
		Sheet3.cell(row = 43, column = 7).value ="May"
		Sheet3.cell(row = 43, column = 8).value ="June"
		Sheet3.cell(row = 43, column = 9).value ="July"
		Sheet3.cell(row = 43, column = 10).value ="August"
		Sheet3.cell(row = 43, column = 11).value ="September"
		Sheet3.cell(row = 43, column = 12).value ="October"
		Sheet3.cell(row = 43, column = 13).value ="November"
		Sheet3.cell(row = 43, column = 14).value ="December"
		Sheet3.cell(row = 43, column = 15).value ="Total"

		Sheet3.cell(row = 44, column = 2).value ="VAT Collectable 4427"
		Sheet3.cell(row = 45, column = 2).value ="VAT Deductible 4426"
		Sheet3.cell(row = 46, column = 2).value ="VAT Payable/ recoverable"

		#table6
		Sheet3.cell(row = 49, column = 2).value ="Type of VAT"
		Sheet3.cell(row = 49, column = 3).value ="January"
		Sheet3.cell(row = 49, column = 4).value ="February"
		Sheet3.cell(row = 49, column = 5).value ="March"
		Sheet3.cell(row = 49, column = 6).value ="April"
		Sheet3.cell(row = 49, column = 7).value ="May"
		Sheet3.cell(row = 49, column = 8).value ="June"
		Sheet3.cell(row = 49, column = 9).value ="July"
		Sheet3.cell(row = 49, column = 10).value ="August"
		Sheet3.cell(row = 49, column = 11).value ="September"
		Sheet3.cell(row = 49, column = 12).value ="October"
		Sheet3.cell(row = 49, column = 13).value ="November"
		Sheet3.cell(row = 49, column = 14).value ="December"
		Sheet3.cell(row = 49, column = 15).value ="Total"

		Sheet3.cell(row = 50, column = 2).value ="VAT Collectable 4427"
		Sheet3.cell(row = 51, column = 2).value ="VAT Deductible 4426"
		Sheet3.cell(row = 52, column = 2).value ="VAT Payable/ recoverable"

		#content
		Sheet3.cell(row = 1, column = 2).value =clientname1
		Sheet3.cell(row = 2, column = 2).value =yearEnd1
		Sheet3.cell(row = 2, column = 2).number_format='mm/dd/yyyy'
		Sheet3.cell(row = 1, column = 15).value =preparedBy1
		Sheet3.cell(row = 2, column = 15).value =datePrepared1
		Sheet3.cell(row = 2, column = 15).number_format='mm/dd/yyyy'
		
		Sheet3.cell(row = 19, column = 3).value ="='T10 VAT Test'!V17"
		Sheet3.cell(row = 19, column = 4).value ="='T10 VAT Test'!V18"
		Sheet3.cell(row = 19, column = 5).value ="='T10 VAT Test'!V19"
		Sheet3.cell(row = 19, column = 6).value ="='T10 VAT Test'!V20"
		Sheet3.cell(row = 19, column = 7).value ="='T10 VAT Test'!V21"
		Sheet3.cell(row = 19, column = 8).value ="='T10 VAT Test'!V22"
		Sheet3.cell(row = 19, column = 9).value ="='T10 VAT Test'!V23"
		Sheet3.cell(row = 19, column = 10).value ="='T10 VAT Test'!V24"
		Sheet3.cell(row = 19, column = 11).value ="='T10 VAT Test'!V25"
		Sheet3.cell(row = 19, column = 12).value ="='T10 VAT Test'!V26"
		Sheet3.cell(row = 19, column = 13).value ="='T10 VAT Test'!V27"
		Sheet3.cell(row = 19, column = 14).value ="='T10 VAT Test'!V28"
		Sheet3.cell(row = 19, column = 15).value ="=SUM(C19:N19)"

		Sheet3.cell(row = 20, column = 3).value ="='T10 VAT Test'!AP17"
		Sheet3.cell(row = 20, column = 4).value ="='T10 VAT Test'!AP18"
		Sheet3.cell(row = 20, column = 5).value ="='T10 VAT Test'!AP19"
		Sheet3.cell(row = 20, column = 6).value ="='T10 VAT Test'!AP20"
		Sheet3.cell(row = 20, column = 7).value ="='T10 VAT Test'!AP21"
		Sheet3.cell(row = 20, column = 8).value ="='T10 VAT Test'!AP22"
		Sheet3.cell(row = 20, column = 9).value ="='T10 VAT Test'!AP23"
		Sheet3.cell(row = 20, column = 10).value ="='T10 VAT Test'!AP24"
		Sheet3.cell(row = 20, column = 11).value ="='T10 VAT Test'!AP25"
		Sheet3.cell(row = 20, column = 12).value ="='T10 VAT Test'!AP26"
		Sheet3.cell(row = 20, column = 13).value ="='T10 VAT Test'!AP27"
		Sheet3.cell(row = 20, column = 14).value ="='T10 VAT Test'!AP28"
		Sheet3.cell(row = 20, column = 15).value ="=SUM(C20:N20)"

		# Sheet3.cell(row = 22, column = 3).value ="='T10 VAT Test'!K17"
		# Sheet3.cell(row = 22, column = 4).value ="='T10 VAT Test'!K18"
		# Sheet3.cell(row = 22, column = 5).value ="='T10 VAT Test'!K19"
		# Sheet3.cell(row = 22, column = 6).value ="='T10 VAT Test'!K20"
		# Sheet3.cell(row = 22, column = 7).value ="='T10 VAT Test'!K21"
		# Sheet3.cell(row = 22, column = 8).value ="='T10 VAT Test'!K22"
		# Sheet3.cell(row = 22, column = 9).value ="='T10 VAT Test'!K23"
		# Sheet3.cell(row = 22, column = 10).value ="='T10 VAT Test'!K24"
		# Sheet3.cell(row = 22, column = 11).value ="='T10 VAT Test'!K25"
		# Sheet3.cell(row = 22, column = 12).value ="='T10 VAT Test'!K26"
		# Sheet3.cell(row = 22, column = 13).value ="='T10 VAT Test'!K27"
		# Sheet3.cell(row = 22, column = 14).value ="='T10 VAT Test'!K28"
		# Sheet3.cell(row = 22, column = 15).value ="=SUM(C22:N22)"

		Sheet3.cell(row = 25, column = 3).value ="=C24+C20-C19"
		Sheet3.cell(row = 25, column = 4).value ="=D24+D20-D19"
		Sheet3.cell(row = 25, column = 5).value ="=E24+E20-E19"
		Sheet3.cell(row = 25, column = 6).value ="=F24+F20-F19"
		Sheet3.cell(row = 25, column = 7).value ="=G24+G20-G19"
		Sheet3.cell(row = 25, column = 8).value ="=H24+H20-H19"
		Sheet3.cell(row = 25, column = 9).value ="=I24+I20-I19"
		Sheet3.cell(row = 25, column = 10).value ="=J24+J20-J19"
		Sheet3.cell(row = 25, column = 11).value ="=K24+K20-K19"
		Sheet3.cell(row = 25, column = 12).value ="=L24+L20-L19"
		Sheet3.cell(row = 25, column = 13).value ="=M24+M20-M19"
		Sheet3.cell(row = 25, column = 14).value ="=N24+N20-N19"
		Sheet3.cell(row = 25, column = 15).value ="=SUM(C25:N25)"

		#table3
		Sheet3.cell(row = 33, column = 3).value ="=C19-C28"
		Sheet3.cell(row = 33, column = 4).value ="=D19-D28"
		Sheet3.cell(row = 33, column = 5).value ="=E19-E28"
		Sheet3.cell(row = 33, column = 6).value ="=F19-F28"
		Sheet3.cell(row = 33, column = 7).value ="=G19-G28"
		Sheet3.cell(row = 33, column = 8).value ="=H19-H28"
		Sheet3.cell(row = 33, column = 9).value ="=I19-I28"
		Sheet3.cell(row = 33, column = 10).value ="=J19-J28"
		Sheet3.cell(row = 33, column = 11).value ="=K19-K28"
		Sheet3.cell(row = 33, column = 12).value ="=L19-L28"
		Sheet3.cell(row = 33, column = 13).value ="=M19-M28"
		Sheet3.cell(row = 33, column = 14).value ="=N19-N28"
		Sheet3.cell(row = 33, column = 15).value ="=SUM(C33:N33)"

		Sheet3.cell(row = 34, column = 3).value ="=C20+C24-C29"
		Sheet3.cell(row = 34, column = 4).value ="=D20+D24-D29"
		Sheet3.cell(row = 34, column = 5).value ="=E20+E24-E29"
		Sheet3.cell(row = 34, column = 6).value ="=F20+F24-F29"
		Sheet3.cell(row = 34, column = 7).value ="=G20+G24-G29"
		Sheet3.cell(row = 34, column = 8).value ="=H20+H24-H29"
		Sheet3.cell(row = 34, column = 9).value ="=I20+I24-I29"
		Sheet3.cell(row = 34, column = 10).value ="=J20+J24-J29"
		Sheet3.cell(row = 34, column = 11).value ="=K20+K24-K29"
		Sheet3.cell(row = 34, column = 12).value ="=L20+L24-L29"
		Sheet3.cell(row = 34, column = 13).value ="=M20+M24-M29"
		Sheet3.cell(row = 34, column = 14).value ="=N20+N24-N29"
		Sheet3.cell(row = 34, column = 15).value ="=SUM(C34:N34)"

		Sheet3.cell(row = 35, column = 3).value ="=SUM(C33:C34)"
		Sheet3.cell(row = 35, column = 4).value ="=SUM(D33:D34)"
		Sheet3.cell(row = 35, column = 5).value ="=SUM(E33:E34)"
		Sheet3.cell(row = 35, column = 6).value ="=SUM(F33:F34)"
		Sheet3.cell(row = 35, column = 7).value ="=SUM(G33:G34)"
		Sheet3.cell(row = 35, column = 8).value ="=SUM(H33:H34)"
		Sheet3.cell(row = 35, column = 9).value ="=SUM(I33:I34)"
		Sheet3.cell(row = 35, column = 10).value ="=SUM(J33:J34)"
		Sheet3.cell(row = 35, column = 11).value ="=SUM(K33:K34)"
		Sheet3.cell(row = 35, column = 12).value ="=SUM(L33:L34)"
		Sheet3.cell(row = 35, column = 13).value ="=SUM(M33:M34)"
		Sheet3.cell(row = 35, column = 14).value ="=SUM(N33:N34)"
		Sheet3.cell(row = 35, column = 15).value ="=SUM(C35:N35)"

		#table 5
		Sheet3.cell(row = 44, column = 3).value ="=C38-C28"
		Sheet3.cell(row = 44, column = 4).value ="=D38-D28"
		Sheet3.cell(row = 44, column = 5).value ="=E38-E28"
		Sheet3.cell(row = 44, column = 6).value ="=F38-F28"
		Sheet3.cell(row = 44, column = 7).value ="=G38-G28"
		Sheet3.cell(row = 44, column = 8).value ="=H38-H28"
		Sheet3.cell(row = 44, column = 9).value ="=I38-I28"
		Sheet3.cell(row = 44, column = 10).value ="=J38-J28"
		Sheet3.cell(row = 44, column = 11).value ="=K38-K28"
		Sheet3.cell(row = 44, column = 12).value ="=L38-L28"
		Sheet3.cell(row = 44, column = 13).value ="=M38-M28"
		Sheet3.cell(row = 44, column = 14).value ="=N38-N28"
		Sheet3.cell(row = 44, column = 15).value ="=SUM(C44:N44)"

		Sheet3.cell(row = 45, column = 3).value ="=C39-C29"
		Sheet3.cell(row = 45, column = 4).value ="=D39-D29"
		Sheet3.cell(row = 45, column = 5).value ="=E39-E29"
		Sheet3.cell(row = 45, column = 6).value ="=F39-F29"
		Sheet3.cell(row = 45, column = 7).value ="=G39-G29"
		Sheet3.cell(row = 45, column = 8).value ="=H39-H29"
		Sheet3.cell(row = 45, column = 9).value ="=I39-I29"
		Sheet3.cell(row = 45, column = 10).value ="=J39-J29"
		Sheet3.cell(row = 45, column = 11).value ="=K39-K29"
		Sheet3.cell(row = 45, column = 12).value ="=L39-L29"
		Sheet3.cell(row = 45, column = 13).value ="=M39-M29"
		Sheet3.cell(row = 45, column = 14).value ="=N39-N29"
		Sheet3.cell(row = 45, column = 15).value ="=SUM(C45:N45)"

		Sheet3.cell(row = 46, column = 3).value ="=C40-C30"
		Sheet3.cell(row = 46, column = 4).value ="=D40-D30"
		Sheet3.cell(row = 46, column = 5).value ="=E40-E30"
		Sheet3.cell(row = 46, column = 6).value ="=F40-F30"
		Sheet3.cell(row = 46, column = 7).value ="=G40-G30"
		Sheet3.cell(row = 46, column = 8).value ="=H40-H30"
		Sheet3.cell(row = 46, column = 9).value ="=I40-I30"
		Sheet3.cell(row = 46, column = 10).value ="=J40-J30"
		Sheet3.cell(row = 46, column = 11).value ="=K40-K30"
		Sheet3.cell(row = 46, column = 12).value ="=L40-L30"
		Sheet3.cell(row = 46, column = 13).value ="=M40-M30"
		Sheet3.cell(row = 46, column = 14).value ="=N40-N30"
		Sheet3.cell(row = 46, column = 15).value ="=SUM(C46:N46)"

		#table6
		Sheet3.cell(row = 50, column = 3).value ="=C19-C38"
		Sheet3.cell(row = 50, column = 4).value ="=D19-D38"
		Sheet3.cell(row = 50, column = 5).value ="=E19-E38"
		Sheet3.cell(row = 50, column = 6).value ="=F19-F38"
		Sheet3.cell(row = 50, column = 7).value ="=G19-G38"
		Sheet3.cell(row = 50, column = 8).value ="=H19-H38"
		Sheet3.cell(row = 50, column = 9).value ="=I19-I38"
		Sheet3.cell(row = 50, column = 10).value ="=J19-J38"
		Sheet3.cell(row = 50, column = 11).value ="=K19-K38"
		Sheet3.cell(row = 50, column = 12).value ="=L19-L38"
		Sheet3.cell(row = 50, column = 13).value ="=M19-M38"
		Sheet3.cell(row = 50, column = 14).value ="=N19-N38"
		Sheet3.cell(row = 50, column = 15).value ="=SUM(C50:N50)"

		Sheet3.cell(row = 51, column = 3).value ="=C20+C24-C39"
		Sheet3.cell(row = 51, column = 4).value ="=D20+D24-D39"
		Sheet3.cell(row = 51, column = 5).value ="=E20+E24-E39"
		Sheet3.cell(row = 51, column = 6).value ="=F20+F24-F39"
		Sheet3.cell(row = 51, column = 7).value ="=G20+G24-G39"
		Sheet3.cell(row = 51, column = 8).value ="=H20+H24-H39"
		Sheet3.cell(row = 51, column = 9).value ="=I20+I24-I39"
		Sheet3.cell(row = 51, column = 10).value ="=J20+J24-J39"
		Sheet3.cell(row = 51, column = 11).value ="=K20+K24-K39"
		Sheet3.cell(row = 51, column = 12).value ="=L20+L24-L39"
		Sheet3.cell(row = 51, column = 13).value ="=M20+M24-M39"
		Sheet3.cell(row = 51, column = 14).value ="=N20+N24-N39"
		Sheet3.cell(row = 51, column = 15).value ="=SUM(C51:N51)"

		Sheet3.cell(row = 52, column = 3).value ="=C25-C40"
		Sheet3.cell(row = 52, column = 4).value ="=D25-D40"
		Sheet3.cell(row = 52, column = 5).value ="=E25-E40"
		Sheet3.cell(row = 52, column = 6).value ="=F25-F40"
		Sheet3.cell(row = 52, column = 7).value ="=G25-G40"
		Sheet3.cell(row = 52, column = 8).value ="=H25-H40"
		Sheet3.cell(row = 52, column = 9).value ="=I25-I40"
		Sheet3.cell(row = 52, column = 10).value ="=J25-J40"
		Sheet3.cell(row = 52, column = 11).value ="=K25-K40"
		Sheet3.cell(row = 52, column = 12).value ="=L25-L40"
		Sheet3.cell(row = 52, column = 13).value ="=M25-M40"
		Sheet3.cell(row = 52, column = 14).value ="=N25-N40"
		Sheet3.cell(row = 52, column = 15).value ="=SUM(C52:N52)"

		#format
		for row in Sheet3['C19:O52']:
			for cell in row:
				cell.number_format = '#,##0_);(#,##0)'

		for row in Sheet3['B18:O25']:
			for cell in row:
				cell.border = border

		for row in Sheet3['B27:O30']:
			for cell in row:
				cell.border = border

		for row in Sheet3['B32:O35']:
			for cell in row:
				cell.border = border

		for row in Sheet3['B37:O40']:
			for cell in row:
				cell.border = border

		for row in Sheet3['B43:O46']:
			for cell in row:
				cell.border = border

		for row in Sheet3['B49:O52']:
			for cell in row:
				cell.border = border

		#COLORS
		Sheet3.cell(row = 1, column = 1).font =ft1
		Sheet3.cell(row = 2, column = 1).font =ft1
		Sheet3.cell(row = 1, column = 14).font =ft1
		Sheet3.cell(row = 2, column = 14).font =ft1
		Sheet3.cell(row = 3, column = 14).font =ft1

		Sheet3.cell(row = 4, column = 1).font = f_testname
		Sheet3.cell(row = 6, column = 1).font =ft1
		Sheet3.cell(row = 8, column = 1).font =ft1
		Sheet3.cell(row = 13, column = 1).font =ft1
		Sheet3.cell(row = 15, column = 1).font =ft1

		Sheet3.cell(row = 18, column = 1).font =ft1
		Sheet3.cell(row = 27, column = 1).font =blue_bold_font
		Sheet3.cell(row = 32, column = 1).font =check_font

		Sheet3.cell(row = 37, column = 1).font =ft1
		Sheet3.cell(row = 43, column = 1).font =check_font
		Sheet3.cell(row = 49, column = 1).font =check_font

		#table2 content
		for row in Sheet3['B28:O29']:
			for cell in row:
				cell.font = blue_thin_font

		#table3 content
		for row in Sheet3['C33:O35']:
			for cell in row:
				cell.font = check_font_1

		#table5 content
		for row in Sheet3['C44:O46']:
			for cell in row:
				cell.font = check_font_1

		#table6 content
		for row in Sheet3['C50:O52']:
			for cell in row:
				cell.font = check_font_1

		#header months
		for row in Sheet3['B18:O18']:
			for cell in row:
				cell.font = ft1

		for row in Sheet3['B25:O25']:
			for cell in row:
				cell.font = ft1

		for row in Sheet3['B27:O27']:
			for cell in row:
				cell.font = blue_bold_font

		for row in Sheet3['B32:O32']:
			for cell in row:
				cell.font = ft1

		for row in Sheet3['B37:O37']:
			for cell in row:
				cell.font = ft1

		for row in Sheet3['B43:O43']:
			for cell in row:
				cell.font = ft1

		for row in Sheet3['B49:O49']:
			for cell in row:
				cell.font = ft1

		#totals
		for row in Sheet3['B25:O25']:
			for cell in row:
				cell.font = ft1

		for row in Sheet3['B30:O30']:
			for cell in row:
				cell.font = blue_bold_font
		Sheet3.cell(row=35,column=2).font=ft1

		for row in Sheet3['B40:O40']:
			for cell in row:
				cell.font = ft1
		Sheet3.cell(row=46,column=2).font=ft1

		Sheet3.column_dimensions['A'].width = 38
		Sheet3.column_dimensions['B'].width = 26
		Sheet3.column_dimensions['C'].width = 13
		Sheet3.column_dimensions['D'].width = 13
		Sheet3.column_dimensions['E'].width = 13
		Sheet3.column_dimensions['F'].width = 13
		Sheet3.column_dimensions['G'].width = 13
		Sheet3.column_dimensions['H'].width = 13
		Sheet3.column_dimensions['I'].width = 13
		Sheet3.column_dimensions['J'].width = 13
		Sheet3.column_dimensions['K'].width = 13
		Sheet3.column_dimensions['L'].width = 13
		Sheet3.column_dimensions['M'].width = 13
		Sheet3.column_dimensions['N'].width = 13
		Sheet3.column_dimensions['O'].width = 13
		Sheet3.column_dimensions['Q'].width = 13

		Sheet4 = wb.create_sheet("TB Robot")
		Sheet4.sheet_view.showGridLines = False

		tb = openpyxl.load_workbook(file_TB, data_only = 'True') #deschidem TB-ul
		tb1 = tb.active

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "Account":
					rand_tb = cell.row
					coloana_acc_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			account = [b.value for b in tb1[coloana_acc_tb][rand_tb:lun]]           
		except:
			flash("Please insert the correct header for Account in Trial Balance file")
			return render_template("index.html")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "Description":
					rand_tb = cell.row
					coloana_descr_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			descr = [b.value for b in tb1[coloana_descr_tb][rand_tb:lun]]          
		except:
			flash("Please insert the correct header for Description in Trial Balance file")
			return render_template("index.html")


		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "OB":
					rand_tb = cell.row
					coloana_opTB_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			opTB = [b.value for b in tb1[coloana_opTB_tb][rand_tb:lun]]          
		except:
			flash("Please insert the correct header for OB in Trial Balance file")
			return render_template("index.html")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "CM":
					rand_tb = cell.row
					coloana_cr_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			cr_mv = [b.value for b in tb1[coloana_cr_tb][rand_tb:lun]]         
		except:
			flash("Please insert the correct header for CM in Trial Balance file")
			return render_template("index.html")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "DM":
					rand_tb = cell.row
					coloana_db_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			dr_mv = [b.value for b in tb1[coloana_db_tb][rand_tb:lun]]         
		except:
			flash("Please insert the correct header for DM in Trial Balance file")
			return render_template("index.html")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "CB":
					rand_tb = cell.row
					coloana_clTB_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			clTB = [b.value for b in tb1[coloana_clTB_tb][rand_tb:lun]]       
		except:
			flash("Please insert the correct header for CB in Trial Balance file")
			return render_template("index.html")

		Sheet4.cell(row = 1, column = 1).value = "Synt(3)"
		Sheet4.cell(row = 1, column = 2).value = "Synt(4)"
		Sheet4.cell(row = 1, column = 3).value = "Account"
		Sheet4.cell(row = 1, column = 4).value = "Description"
		Sheet4.cell(row = 1, column = 5).value = "Opening Balance"
		Sheet4.cell(row = 1, column = 6).value = "Debit Movement"
		Sheet4.cell(row = 1, column = 7).value = "Credit Movement"
		Sheet4.cell(row = 1, column = 8).value = "Closing Balance"
		Sheet4.cell(row = 1, column = 9).value = "Synt(2)"
		Sheet4.cell(row=1, column=10).value = "Class"

		# # ....adaugi tu restul adica: synt 4, account,descript,ob,dm,cm,cbp
		for i in range(0,len(account)):
			Sheet4.cell(row = 2 + i, column = 3).value = account[i]
			Sheet4.cell(row = 2 + i, column = 2).value = str(account[i])[:4]   #in Excel =left("celula", 4)

		for i in range(0, len(account)):
			Sheet4.cell(row = 2 + i, column = 1).value =  str(account[i])[:3] #in Excel =left("celula", 3)

		for i in range(0, len(descr)):
			Sheet4.cell(row = 2 + i, column = 4).value = descr[i]

		for i in range(0, len(opTB)):
			Sheet4.cell(row = 2 + i, column = 5).value = opTB[i]

		for i in range(0, len(cr_mv)):
			Sheet4.cell(row = 2 + i, column = 6).value = cr_mv[i]

		for i in range(0, len(dr_mv)):
			Sheet4.cell(row = 2 + i, column = 7).value = dr_mv[i]

		for i in range(0, len(clTB)):
			Sheet4.cell(row = 2 + i, column = 8).value = clTB[i]

		for i in range(0, len(account)):
			Sheet4.cell(row = 2 + i, column = 9).value =  str(account[i])[:2]

		for i in range(0, len(account)):
			Sheet4.cell(row=2 + i, column=10).value = str(account[i])[:1]

		Sheet5 = wb.create_sheet("TB PBC")

		mr = tb1.max_row
		mc = tb1.max_column
		# copying the cell values from source
		# excel file to destination excel file
		for i in range (1, mr + 1):
			for j in range (1, mc + 1):
		# reading cell value from source excel file
				c = tb1.cell(row = i, column = j)
		# writing the read value to destination excel file
				Sheet5.cell(row = i, column = j).value = c.value

		std = wb["Sheet"]
		wb.remove(std)
		folderpath = "home/auditappnexia/output/vat"
		file_pathFS = os.path.join(folderpath, "VAT test"+" "+clientname1+".xlsx")
		wb.save(file_pathFS)
		# out.save(folderpath + "/" + "T10 - VAT Test" + ".xlsx")
		return send_from_directory(folderpath, "VAT Test" + " " + clientname1 + ".xlsx", as_attachment=True)

	return render_template("VAT.html")
		
@app.route('/')
def home():
	# path_output="C:\\Users\\Cristian.Iordache\\Documents\\Automation Projects\\Web apps-v2\\Output"
	# shutil.rmtree(path_output)
	# os.mkdir(path_output)
	return render_template('dashboard.html')
@app.route('/MonthlyPL/GTbSJ64rDiFqoNeQyowqNh9pzcE')
def Monthly():
	return render_template('MonthlyPL.html')
@app.route('/MonthlyPL/Instructions', methods=['GET'])
def downloadMonthly():
		filepath = "/home/auditappnexia/output/MonthlyPL"
 
		return send_from_directory(filepath,"Instructions - Monthly P&L.docx", as_attachment=True)
@app.route('/MonthlyPL/GTbSJ64rDiFqoNeQyowqNh9pzcE', methods=['POST', 'GET'])
def Monthly_process():
   
	clientname1 = request.form['client']
	yearEnd1= datetime.datetime.strptime(
					 request.form['yearEnd'],
					 '%Y-%m-%d')
	preparedBy1=request.form['preparedBy']
	# datePrepared1= datetime.datetime.strptime(
	#                  request.form['preparedDate'],
	#                  '%Y-%m-%d')
	datePrepared1= datetime.datetime.now().date()
	refference1 = request.form['reff']
	isChecked1=request.form.get("tbSubtotals")
	print(isChecked1)
	if isChecked1=="": #daca e bifat
		isChecked=0
	else:
		isChecked=1
	


	
	# yearEnd = str(request.form['yearEnd'])
	# processed_text = client.upper()
	# fisier=request.files.get('monthlyTB')
	if request.method == 'POST':
		file_Details = request.files.getlist("monthlyTB")
		file_TB=request.files["TB"]
		
		
		for i in file_Details:
			i.save(secure_filename(i.filename))
		print(isChecked)
		ft1 = Font(name = 'Tahoma', size = 8, bold = True)
		f_testname = Font(name = 'Tahoma', size = 8, bold = True, underline = 'single', italic = True)
		cap_tabel = Font(name = 'Tahoma', size = 8, color = "FFFFFF", bold = True)
		cap_tabel_color_PBC = PatternFill(start_color = '808080', end_color ='808080', fill_type = 'solid') #grey
		cap_tabel_color_GT = PatternFill(start_color = '00AEAC', end_color ='00AEAC', fill_type = 'solid') #indigo
		fprocentaj = Font(name = 'Tahoma', size = 8, color = "FF0000", bold = True)
		font_worksheet = Font(name = 'Tahoma', size = 8)
		check_font = Font(name = 'Tahoma', size = 8, color = "FF0000", bold = True)
		wb = openpyxl.Workbook()
		details = openpyxl.load_workbook(file_Details[0], data_only = 'True')
		details1 = details.active
		for row in details1.iter_rows():
			for cell in row:
				if cell.value == "Cont":
					row_det = cell.row
					column_cont = cell.column
					lungime = len(details1[cell.column])
		try:
			account = [b.value for b in details1[column_cont][row_det:lungime]]
		except:
			flash("Please insert the correct header for Cont in first file monthly trial balance")
			return render_template("index.html")
		for row in details1.iter_rows():
			for cell in row:
				if cell.value == "Denumire cont":
					row_det = cell.row
					column_den = cell.column
					lun = len(details1[cell.column])
		try:
			denumire = [b.value for b in details1[column_den][row_det:lun]]
		except:
			flash("Please insert the correct header for Denumire cont in first file monthly trial balance")
			return render_template("index.html")

		for row in details1.iter_rows():
			for cell in row:
				if cell.value == "Debitor inceput an":
					row_det = cell.row
					column_debi = cell.column
					lun = len(details1[cell.column])
		try:
			DB = [b.value for b in details1[column_debi][row_det:lun]]
		except:
			flash("Please insert the correct header for Debitor inceput an cont in first file monthly trial balance")
			return render_template("index.html")
		for row in details1.iter_rows():
			for cell in row:
				if cell.value == "Creditor inceput an":
					row_det = cell.row
					column_crei = cell.column
					lun = len(details1[cell.column])
		try:
			CR = [b.value for b in details1[column_crei][row_det:lun]]
		except:
			flash("Please insert the correct header for Creditor inceput an cont in first file monthly trial balance")
			return render_template("index.html")

		for row in details1.iter_rows():
			for cell in row:
				if cell.value == "Debitor luna curenta":
					row_det = cell.row
					column_deblc = cell.column
					lun = len(details1[cell.column])
		try:
			DB_curr_month = [b.value for b in details1[column_deblc][row_det:lun]]
		except:
			flash("Please insert the correct header for Debitor luna curenta cont in first file monthly trial balance")
			return render_template("index.html")

		for row in details1.iter_rows():
			for cell in row:
				if cell.value == "Creditor luna curenta":
					row_det = cell.row
					column_crlc = cell.column
					lun = len(details1[cell.column])
					
		try:
			CR_curr_month = [b.value for b in details1[column_crlc][row_det:lun]]
		except:
			flash("Please insert the correct header for Creditor luna curenta cont in first file monthly trial balance")
			return render_template("index.html")
		for row in details1.iter_rows():
			for cell in row:
				if cell.value == "Debitor cumulat":
					row_det = cell.row
					column_debcu = cell.column
					lun = len(details1[cell.column])
		try:
			DR_cumulat = [b.value for b in details1[column_crlc][row_det:lun]]
		except:
			flash("Please insert the correct header for Debitor cumulat in first file monthly trial balance")
			return render_template("index.html")
		for row in details1.iter_rows():
			for cell in row:
				if cell.value == "Creditor cumulat":
					row_det = cell.row
					column_crecu = cell.column
					lun = len(details1[cell.column])
		try:
			CR_cumulat = [b.value for b in details1[column_crecu][row_det:lun]]
		except:
			flash("Please insert the correct header for Creditor cumulat in first file monthly trial balance")
			return render_template("index.html")

		for row in details1.iter_rows():
			for cell in row:
				if cell.value == "Total debitor":
					row_det = cell.row
					column_totd = cell.column
					lun = len(details1[cell.column])
		try:
			total_DB = [b.value for b in details1[column_totd][row_det:lun]]
		except:
			flash("Please insert the correct header for Total debitor in first file monthly trial balance")
			return render_template("index.html")
		for row in details1.iter_rows():
			for cell in row:
				if cell.value == "Total creditor":
					row_det = cell.row
					column_totc = cell.column
					lun = len(details1[cell.column])
		total_CR = [b.value for b in details1[column_totc][row_det:lun]]
		try:
			total_CR = [b.value for b in details1[column_totd][row_det:lun]]
		except:
			flash("Please insert the correct header for Total creditor in first file monthly trial balance")
			return render_template("index.html")
		for row in details1.iter_rows():
			for cell in row:
				if cell.value == "Sold final debitor":
					row_det = cell.row
					column_scd = cell.column
					lun = len(details1[cell.column])
		try:
			DB_curr_sold = [b.value for b in details1[column_scd][row_det:lun]]
		except:
			flash("Please insert the correct header for Sold final debitor in first file monthly trial balance")
			return render_template("index.html")
		for row in details1.iter_rows():
			for cell in row:
				if cell.value == "Sold final creditor":
					row_det = cell.row
					column_scc = cell.column
					lun = len(details1[cell.column])
		try:
			DR_curr_sold = [b.value for b in details1[column_scc][row_det:lun]]
		except:
			flash("Please insert the correct header for Sold final creditor in first file monthly trial balance")
			return render_template("index.html")

		denNo = 0
		#aducem tb monthly
		for a in file_Details:
			details = openpyxl.load_workbook(a, data_only = 'True')
			details1 = details.active

			list1=[]
			denNo = denNo+1
			Sheet1 = wb.create_sheet(str(denNo))

			lun=details1.max_row
			account = [b.value for b in details1[column_cont][row_det:lun]]
			for i in range(0, len(account)):
				list1.append(account[i])

			denumire = [b.value for b in details1[column_den][row_det:lun]]

			
			DB = [b.value for b in details1[column_debi][row_det:lun]]

			
			CR = [b.value for b in details1[column_crei][row_det:lun]]

			DB_curr_month = [b.value for b in details1[column_deblc][row_det:lun]]

			CR_curr_month = [b.value for b in details1[column_crlc][row_det:lun]]

			DR_cumulat = [b.value for b in details1[column_debcu][row_det:lun]]

			CR_cumulat = [b.value for b in details1[column_crecu][row_det:lun]]

			total_DB = [b.value for b in details1[column_totd][row_det:lun]]

			total_CR = [b.value for b in details1[column_totc][row_det:lun]]

			DB_curr_sold = [b.value for b in details1[column_scd][row_det:lun]]

			CR_curr_sold = [b.value for b in details1[column_scc][row_det:lun]]

			Sheet1.cell(row=1, column=1).value = 'Cont'
			Sheet1.cell(row=1, column=2).value = 'Nume cont'
			Sheet1.cell(row=1, column=3).value = "Sold debitor inceput an"
			Sheet1.cell(row=1, column=4).value = "Sold creditor inceput an"
			Sheet1.cell(row=1, column=17).value = "For Monthly P&L"
			Sheet1.cell(row=1, column=5).value ="Rulaj debitor curent"
			Sheet1.cell(row=1, column=6).value ="Rulaj creditor curent"
			Sheet1.cell(row=1, column=7).value ="Sold debitor luna curenta"
			Sheet1.cell(row=1, column=8).value ="Sold creditor luna curenta"
			Sheet1.cell(row=1, column=9).value ="Sold total debitor"
			Sheet1.cell(row=1, column=10).value ="Sold total creditor"
			Sheet1.cell(row=1, column=11).value ="Sold final debitor"
			Sheet1.cell(row=1, column=12).value ="Sold final creditor"

			for i in range(0, len(account)):
				Sheet1.cell(row=2 + i, column=1).value = account[i]

				Sheet1.cell(row=2+i, column=14).value = '=LEFT(A'+str(2+i)+',3)'		
				Sheet1.cell(row=2+i, column=17).value = '=IF(LEFT(A'+str(2+i)+',1)="6",E'+str(2+i)+',IF(LEFT(A'+str(2+i)+',1)="7",-F'+str(2+i)+',IF(LEFT(A'+str(2+i)+',3)="121",E'+str(2+i)+'-F'+str(2+i)+',0)))'
				# Sheet1.cell(row=4 + i, column=17).value = "=IF(LEFT(A"+str(4+i)+","+str(1)+")="+str(6)+",E"+str(4+i)+",IF(LEFT(A"+str(4+i)+","+str(1)+")="+str(7)+",-F"+str(4+i)+",IF(LEFT(A"+str(4+i)+","+str(3)+")="+str(121)+",E"+str(4+i)+"-F"+str(4+i)+","+str(0)+"))"

			for i in range(0, len(denumire)):
				Sheet1.cell(row=2 + i, column=2).value = denumire[i]

			for i in range(0, len(DB)):
				Sheet1.cell(row=2 + i, column=3).value = DB[i]
				Sheet1.cell(row=2 + i, column=3).number_format = '#,##0_);(#,##0)'

			for i in range(0, len(CR)):
				Sheet1.cell(row=2 + i, column=4).value = CR[i]
				Sheet1.cell(row=2 + i, column=4).number_format = '#,##0_);(#,##0)'

			for i in range(0, len(DB_curr_month)):
				Sheet1.cell(row=2 + i, column=5).value = DB_curr_month[i]
				Sheet1.cell(row=2 + i, column=5).number_format = '#,##0_);(#,##0)'

			for i in range(0, len(CR_curr_month)):
				Sheet1.cell(row=2 + i, column=6).value = CR_curr_month[i]
				Sheet1.cell(row=2 + i, column=6).number_format = '#,##0_);(#,##0)'

			for i in range(0, len(DR_cumulat)):
				Sheet1.cell(row=2 + i, column=7).value = DR_cumulat[i]
				Sheet1.cell(row=2 + i, column=7).number_format = '#,##0_);(#,##0)'

			for i in range(0, len(CR_cumulat)):
				Sheet1.cell(row=2 + i, column=8).value = CR_cumulat[i]
				Sheet1.cell(row=2 + i, column=8).number_format = '#,##0_);(#,##0)'

			for i in range(0, len(total_DB)):
				Sheet1.cell(row=2 + i, column=9).value = total_DB[i]
				Sheet1.cell(row=2 + i, column=9).number_format = '#,##0_);(#,##0)'

			for i in range(0, len(total_CR)):
				Sheet1.cell(row=2 + i, column=10).value = total_CR[i]
				Sheet1.cell(row=2 + i, column=10).number_format = '#,##0_);(#,##0)'

			for i in range(0, len(DB_curr_sold)):
				Sheet1.cell(row=2 + i, column=11).value = DB_curr_sold[i]
				Sheet1.cell(row=2 + i, column=11).number_format = '#,##0_);(#,##0)'

			for i in range(0, len(CR_curr_sold)):
				Sheet1.cell(row=2 + i, column=12).value = CR_curr_sold[i]
				Sheet1.cell(row=2 + i, column=12).number_format = '#,##0_);(#,##0)'

			Sheet1.cell(row=14, column=16).value="=K14-L14"

			# formatare

			for row in Sheet1['A1:L1']:
				for cell in row:
					cell.fill = cap_tabel_color_GT

			for row in Sheet1['A1:L1']:
				for cell in row:
					cell.font = cap_tabel

			Sheet1.cell(row=1, column=17).fill = cap_tabel_color_GT
			Sheet1.cell(row=1, column=17).font = cap_tabel

			Sheet1.column_dimensions['A'].width = 10
			Sheet1.column_dimensions['B'].width = 20
			Sheet1.column_dimensions['C'].width = 20
			Sheet1.column_dimensions['D'].width = 20
			Sheet1.column_dimensions['E'].width = 20
			Sheet1.column_dimensions['F'].width = 20
			Sheet1.column_dimensions['G'].width = 20
			Sheet1.column_dimensions['H'].width = 20
			Sheet1.column_dimensions['I'].width = 20
			Sheet1.column_dimensions['J'].width = 20
			Sheet1.column_dimensions['K'].width = 20
			Sheet1.column_dimensions['L'].width = 20


		# import TB general
		tb = openpyxl.load_workbook(file_TB, data_only='True')  # deschidem TB-ul
		tb1 = tb.active
		Sheet2 = wb.create_sheet("TB Robot")

		list2=[]
		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "Account":
					row_acc = cell.row 
					column_acc = cell.column
					lun = len(tb1[cell.column])
		account_tb = [b.value for b in tb1[column_acc][row_acc:lun]]
		print(account_tb)
		for i in range(0, len(account_tb)):
			# try:
				if(int(str(account_tb[i])[:1])>5):
					print(account_tb[i])
					list2.append(account_tb[i])
			# except:
			#     pass

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "Description":
					row_descr = cell.row
					column_descr = cell.column
					lun = len(tb1[cell.column])
		descr = [b.value for b in tb1[column_descr][row_descr:lun]]
		# list1.append(descr)
		
		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "OB":
					row_opTB_tb = cell.row
					coloana_opTB_tb = cell.column
					lun = len(tb1[cell.column])
		opTB = [b.value for b in tb1[coloana_opTB_tb][row_opTB_tb:lun]]
		# list1.append(opTB)

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "CM":
					row_tb = cell.row
					coloana_cr_tb = cell.column
					lun = len(tb1[cell.column])
		cr_mv = [b.value for b in tb1[coloana_cr_tb][row_tb:lun]]
		# list1.append(cr_mv)

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "DM":
					row_tb = cell.row
					coloana_db_tb = cell.column
					lun = len(tb1[cell.column])
		dr_mv = [b.value for b in tb1[coloana_db_tb][row_tb:lun]]
		# list1.append(dr_mv)

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "CB":
					row_tb = cell.row
					coloana_clTB_tb = cell.column
					lun = len(tb1[cell.column])
		clTB = [b.value for b in tb1[coloana_clTB_tb][row_tb:lun]]

		for i in range(0, len(account_tb)):
			Sheet2.cell(row=2 + i, column=3).value = account_tb[i]
			Sheet2.cell(row=2 + i, column=2).value = str(account_tb[i])[:4]
			Sheet2.cell(row=2 + i, column=2).number_format = '#,##0_);(#,##0)'

		for i in range(0, len(account_tb)):
			Sheet2.cell(row=2 + i, column=1).value = str(account_tb[i])[:3]  # in Excel =left("celula", 3)
			Sheet2.cell(row=2 + i, column=1).number_format = '#,##0_);(#,##0)'

		for i in range(0, len(descr)):
			Sheet2.cell(row=2 + i, column=4).value = descr[i]
			Sheet2.cell(row=2 + i, column=4).number_format = '#,##0_);(#,##0)'

		for i in range(0, len(opTB)):
			Sheet2.cell(row=2 + i, column=5).value = opTB[i]
			Sheet2.cell(row=2 + i, column=5).number_format = '#,##0_);(#,##0)'

		for i in range(0, len(cr_mv)):
			Sheet2.cell(row=2 + i, column=6).value = cr_mv[i]
			Sheet2.cell(row=2 + i, column=6).number_format = '#,##0_);(#,##0)'

		for i in range(0, len(dr_mv)):
			Sheet2.cell(row=2 + i, column=7).value = dr_mv[i]
			Sheet2.cell(row=2 + i, column=7).number_format = '#,##0_);(#,##0)'

		for i in range(0, len(clTB)):
			Sheet2.cell(row=2 + i, column=8).value = clTB[i]
			Sheet2.cell(row=2 + i, column=8).number_format = '#,##0_);(#,##0)'

		for i in range(0, len(account_tb)):
			Sheet2.cell(row=2 + i, column=9).value = str(account_tb[i])[:2]
			Sheet2.cell(row=2 + i, column=9).number_format = '#,##0_);(#,##0)'

		#creare sheet
		Sheet2.cell(row=1, column=1).value = "Synt(3)"
		Sheet2.cell(row=1, column=2).value = "Synt(4)"
		Sheet2.cell(row=1, column=3).value = "Account"
		Sheet2.cell(row=1, column=4).value = "Description"
		Sheet2.cell(row=1, column=5).value = "Opening Balance"
		Sheet2.cell(row=1, column=6).value = "Debit Movement"
		Sheet2.cell(row=1, column=7).value = "Credit Movement"
		Sheet2.cell(row=1, column=8).value = "Closing Balance"
		Sheet2.cell(row=1, column=9).value = "Synt(2)"

		#format sheet
		for row in Sheet2['A1:I1']:
			for cell in row:
				cell.font = cap_tabel

		for row in Sheet2['A1:I1']:
			for cell in row:
				cell.fill = cap_tabel_color_GT

		Sheet2.column_dimensions['D'].width = 20
		Sheet2.column_dimensions['E'].width = 20
		Sheet2.column_dimensions['F'].width = 20
		Sheet2.column_dimensions['G'].width = 20
		Sheet2.column_dimensions['H'].width = 20

		Sheet3 = wb.create_sheet('Monthly P&L')
		Sheet3.sheet_view.showGridLines = False

		#building the worksheet
		Sheet3.cell(row = 1, column = 1).value = "Client name"
		Sheet3.cell(row = 2, column = 1).value = "Period end"

		Sheet3.cell(row=5, column=5).value = 'January'
		Sheet3.cell(row=5, column=6).value = "February"
		Sheet3.cell(row=5, column=7).value = "March"
		Sheet3.cell(row=5, column=8).value = "April"
		Sheet3.cell(row=5, column=9).value = "May"
		Sheet3.cell(row=5, column=10).value = "June"
		Sheet3.cell(row=5, column=11).value ="July"
		Sheet3.cell(row=5, column=12).value ="August"
		Sheet3.cell(row=5, column=13).value ="September"
		Sheet3.cell(row=5, column=14).value ="October"
		Sheet3.cell(row=5, column=15).value ="November"
		Sheet3.cell(row=5, column=16).value ="December"
		Sheet3.cell(row=5, column=17).value ="Total"
		Sheet3.cell(row=5, column=18).value ="As per TB"
		Sheet3.cell(row=5, column=19).value ="Difference"

		Sheet3.cell(row=1, column=18).value = "Processed by"
		Sheet3.cell(row=2, column=18).value = "Date"
		Sheet3.cell(row=3, column=18).value = "Refference"

		#variables
		Sheet3.cell(row = 1, column = 2).value = clientname1
		Sheet3.cell(row = 2, column = 2).value = yearEnd1
		Sheet3.cell(row = 2, column = 2).number_format='mm/dd/yyyy'
		Sheet3.cell(row=1, column=19).value = preparedBy1
		Sheet3.cell(row=2, column=19).value = datePrepared1
		Sheet3.cell(row=2, column=19).number_format='mm/dd/yyyy'
		Sheet3.cell(row=3, column=19).value = refference1

		#formulas
		Sheet3.cell(row=6, column=4).value = "Expenses"
		Sheet3.cell(row=7, column=4).value = "Revenues"
		Sheet3.cell(row=8, column=4).value = "Result"
		Sheet3.cell(row=9, column=4).value ="Acc 121"
		Sheet3.cell(row=10, column=4).value ="Check"

		#formatare + cap tabel general
		Sheet3.cell(row=13, column=1).value = "Class"
		Sheet3.cell(row=13, column=2).value = "Synthentic"
		Sheet3.cell(row=13, column=3).value = "Account"
		Sheet3.cell(row=13, column=4).value = "Description"

		Sheet3.cell(row=12, column=5).value = '1'
		Sheet3.cell(row=12, column=6).value = "2"
		Sheet3.cell(row=12, column=7).value = "3"
		Sheet3.cell(row=12, column=8).value = "4"
		Sheet3.cell(row=12, column=9).value = "5"
		Sheet3.cell(row=12, column=10).value = "6"
		Sheet3.cell(row=12, column=11).value = "7"
		Sheet3.cell(row=12, column=12).value = "8"
		Sheet3.cell(row=12, column=13).value = "9"
		Sheet3.cell(row=12, column=14).value = "10"
		Sheet3.cell(row=12, column=15).value = "11"
		Sheet3.cell(row=12, column=16).value = "12"
		Sheet3.cell(row=12, column=18).value = "TB"

		Sheet3.cell(row=13, column=5).value = 'January'
		Sheet3.cell(row=13, column=6).value = "February"
		Sheet3.cell(row=13, column=7).value = "March"
		Sheet3.cell(row=13, column=8).value = "April"
		Sheet3.cell(row=13, column=9).value = "May"
		Sheet3.cell(row=13, column=10).value = "June"
		Sheet3.cell(row=13, column=11).value = "July"
		Sheet3.cell(row=13, column=12).value = "August"
		Sheet3.cell(row=13, column=13).value = "September"
		Sheet3.cell(row=13, column=14).value = "October"
		Sheet3.cell(row=13, column=15).value = "November"
		Sheet3.cell(row=13, column=16).value = "December"
		Sheet3.cell(row=13, column=17).value = "Total"
		Sheet3.cell(row=13, column=18).value = "As per TB"
		Sheet3.cell(row=13, column=19).value = "Difference"

		#continut
		# for i in range(0, len(descr)):
		# 	Sheet3.cell(row=14 + i, column=4).value = descr[i]

		Sheet3.cell(row=9, column=5).value = "=SUMIF('1'!N:N,121,'1'!Q:Q)"
		Sheet3.cell(row=9, column=6).value = "=SUMIF('2'!N:N,121,'2'!Q:Q)"
		Sheet3.cell(row=9, column=7).value = "=SUMIF('3'!N:N,121,'3'!Q:Q)"
		Sheet3.cell(row=9, column=8).value = "=SUMIF('4'!N:N,121,'4'!Q:Q)"
		Sheet3.cell(row=9, column=9).value = "=SUMIF('5'!N:N,121,'5'!Q:Q)"
		Sheet3.cell(row=9, column=10).value = "=SUMIF('6'!N:N,121,'6'!Q:Q)"
		Sheet3.cell(row=9, column=11).value = "=SUMIF('7'!N:N,121,'7'!Q:Q)"
		Sheet3.cell(row=9, column=12).value = "=SUMIF('8'!N:N,121,'8'!Q:Q)"
		Sheet3.cell(row=9, column=13).value = "=SUMIF('9'!N:N,121,'9'!Q:Q)"
		Sheet3.cell(row=9, column=14).value = "=SUMIF('10'!N:N,121,'10'!Q:Q)"
		Sheet3.cell(row=9, column=15).value = "=SUMIF('11'!N:N,121,'11'!Q:Q)"
		Sheet3.cell(row=9, column=16).value = "=SUMIF('12'!N:N,121,'12'!Q:Q)"
		Sheet3.cell(row=9, column=17).value ="=SUM(E9:P9)"
		Sheet3.cell(row=9, column=18).value ="=SUMIF('TB Robot'!A:A,121,'TB Robot'!H:H)"
		Sheet3.cell(row=9, column=19).value ="=Q9-R9"

		# [0 if v is None else v for v in list2]
		# for i in range(0, len(list2)):
		# 	if i>6:
		# 		Sheet3.cell(row=14 + i, column=1).value = str(account[i])[:1]
		# 		Sheet3.cell(row=14 + i, column=2).value = str(account[i])[:3]
		# 		Sheet3.cell(row=14 + i, column=3).value = account[i]
		# 	else:
		# 		None

		Sheet3.cell(row=6, column=5).value ="=SUMIF($A$14:$A$1000,6,E14:E1000)"
		Sheet3.cell(row=6, column=6).value ="=SUMIF($A$14:$A$1000,6,F14:F1000)"
		Sheet3.cell(row=6, column=7).value = "=SUMIF($A$14:$A$1000,6,G14:G1000)"
		Sheet3.cell(row=6, column=8).value ="=SUMIF($A$14:$A$1000,6,H14:H1000)"
		Sheet3.cell(row=6, column=9).value ="=SUMIF($A$14:$A$1000,6,I14:I1000)"
		Sheet3.cell(row=6, column=10).value ="=SUMIF($A$14:$A$1000,6,J14:J1000)"
		Sheet3.cell(row=6, column=11).value ="=SUMIF($A$14:$A$1000,6,K14:K1000)"
		Sheet3.cell(row=6, column=12).value ="=SUMIF($A$14:$A$1000,6,L14:L1000)"
		Sheet3.cell(row=6, column=13).value ="=SUMIF($A$14:$A$1000,6,M14:M1000)"
		Sheet3.cell(row=6, column=14).value ="=SUMIF($A$14:$A$1000,6,N14:N1000)"
		Sheet3.cell(row=6, column=15).value ="=SUMIF($A$14:$A$1000,6,O14:O1000)"
		Sheet3.cell(row=6, column=16).value ="=SUMIF($A$14:$A$1000,6,P14:P1000)"
		Sheet3.cell(row=6, column=17).value ="=SUM(E6:P6)"
		Sheet3.cell(row=6, column=18).value ="=SUMIF($A$14:$A$1000,6,R14:R1000)"
		Sheet3.cell(row=6, column=19).value ="=Q6-R6"

		Sheet3.cell(row=7, column=5).value = "=SUMIF($A$14:$A$1000,7,E14:E1000)"
		Sheet3.cell(row=7, column=6).value = "=SUMIF($A$14:$A$1000,7,F14:F1000)"
		Sheet3.cell(row=7, column=7).value = "=SUMIF($A$14:$A$1000,7,G14:G1000)"
		Sheet3.cell(row=7, column=8).value = "=SUMIF($A$14:$A$1000,7,H14:H1000)"
		Sheet3.cell(row=7, column=9).value = "=SUMIF($A$14:$A$1000,7,I14:I1000)"
		Sheet3.cell(row=7, column=10).value = "=SUMIF($A$14:$A$1000,7,J14:J1000)"
		Sheet3.cell(row=7, column=11).value = "=SUMIF($A$14:$A$1000,7,K14:K1000)"
		Sheet3.cell(row=7, column=12).value = "=SUMIF($A$14:$A$1000,7,L14:L1000)"
		Sheet3.cell(row=7, column=13).value = "=SUMIF($A$14:$A$1000,7,M14:M1000)"
		Sheet3.cell(row=7, column=14).value = "=SUMIF($A$14:$A$1000,7,N14:N1000)"
		Sheet3.cell(row=7, column=15).value = "=SUMIF($A$14:$A$1000,7,O14:O1000)"
		Sheet3.cell(row=7, column=16).value = "=SUMIF($A$14:$A$1000,7,P14:P1000)"
		Sheet3.cell(row=7, column=17).value = "=SUM(E7:P7)"
		Sheet3.cell(row=7, column=18).value = "=SUMIF($A$14:$A$1000,7,R14:R1000)"
		Sheet3.cell(row=7, column=19).value = "=Q7-R7"

		Sheet3.cell(row=8, column=5).value = "=SUM(E6:E7)"
		Sheet3.cell(row=8, column=6).value ="=SUM(F6:F7)"
		Sheet3.cell(row=8, column=7).value ="=SUM(G6:G7)"
		Sheet3.cell(row=8, column=8).value ="=SUM(H6:H7)"
		Sheet3.cell(row=8, column=9).value ="=SUM(I6:I7)"
		Sheet3.cell(row=8, column=10).value ="=SUM(J6:J7)"
		Sheet3.cell(row=8, column=11).value ="=SUM(K6:K7)"
		Sheet3.cell(row=8, column=12).value ="=SUM(L6:L7)"
		Sheet3.cell(row=8, column=13).value ="=SUM(M6:M7)"
		Sheet3.cell(row=8, column=14).value ="=SUM(N6:N7)"
		Sheet3.cell(row=8, column=15).value ="=SUM(O6:O7)"
		Sheet3.cell(row=8, column=16).value ="=SUM(P6:P7)"
		Sheet3.cell(row=8, column=17).value ="=SUM(Q6:Q7)"
		Sheet3.cell(row=8, column=18).value ="=SUM(R6:R7)"
		Sheet3.cell(row=8, column=19).value ="=Q8-R8"

		Sheet3.cell(row=10, column=5).value ="=E8-E9"
		Sheet3.cell(row=10, column=6).value = "=F8-F9"
		Sheet3.cell(row=10, column=7).value = "=G8-G9"
		Sheet3.cell(row=10, column=8).value = "=H8-H9"
		Sheet3.cell(row=10, column=9).value = "=I8-I9"
		Sheet3.cell(row=10, column=10).value = "=J8-J9"
		Sheet3.cell(row=10, column=11).value = "=K8-K9"
		Sheet3.cell(row=10, column=12).value = "=L8-L9"
		Sheet3.cell(row=10, column=13).value = "=M8-M9"
		Sheet3.cell(row=10, column=14).value = "=N8-N9"
		Sheet3.cell(row=10, column=15).value = "=O8-O9"
		Sheet3.cell(row=10, column=16).value = "=P8-P9"
		Sheet3.cell(row=10, column=17).value = "=Q8-Q9"
		Sheet3.cell(row=10, column=18).value = "=SUM(E10:P10)"
		Sheet3.cell(row=10, column=19).value = "=R10-Q10"

		Sheet3.cell(row = 8, column = 1).value = "Monthly PL @ 31-12-20XX"

		# for i in range(0, len(account)):
		# 	Sheet3.cell(row=14+i, column=5).value = '=SUMIF(INDIRECT(E12&"!$A:$A"),C'+str(14+i)+',INDIRECT(E12&"!$Q:$Q"))'
		# 	Sheet3.cell(row=14+i, column=6).value = '=SUMIF(INDIRECT(F$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(F$12&"!$Q:$Q"))'
		# 	Sheet3.cell(row=14 + i, column=7).value = '=SUMIF(INDIRECT(G$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(G$12&"!$Q:$Q"))'
		# 	Sheet3.cell(row=14 + i, column=8).value = '=SUMIF(INDIRECT(H$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(H$12&"!$Q:$Q"))'
		# 	Sheet3.cell(row=14 + i, column=9).value = '=SUMIF(INDIRECT(I$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(I$12&"!$Q:$Q"))'
		# 	Sheet3.cell(row=14 + i, column=10).value = '=SUMIF(INDIRECT(J$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(J$12&"!$Q:$Q"))'
		# 	Sheet3.cell(row=14 + i, column=11).value = '=SUMIF(INDIRECT(K$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(K$12&"!$Q:$Q"))'
		# 	Sheet3.cell(row=14 + i, column=12).value = '=SUMIF(INDIRECT(L$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(L$12&"!$Q:$Q"))'
		# 	Sheet3.cell(row=14 + i, column=13).value = '=SUMIF(INDIRECT(M$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(M$12&"!$Q:$Q"))'
		# 	Sheet3.cell(row=14 + i, column=14).value = '=SUMIF(INDIRECT(N$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(N$12&"!$Q:$Q"))'
		# 	Sheet3.cell(row=14 + i, column=15).value = '=SUMIF(INDIRECT(O$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(O$12&"!$Q:$Q"))'
		# 	Sheet3.cell(row=14 + i, column=16).value = '=SUMIF(INDIRECT(P$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(P$12&"!$Q:$Q"))'
		# 	Sheet3.cell(row=14 + i, column=17).value = '=SUM(E'+str(14+i)+':P'+str(14+i)+')'
		# 	Sheet3.cell(row=14 + i, column=18).value ="=SUMIF('TB Python'!A:A,'Monthly P&L'!B"+str(14+i)+",'TB Python'!G:G)"
		# 	Sheet3.cell(row=14 + i, column=19).value ='=Q'+str(14+i)+'-R'+str(14+i)+''

		#formats
		Sheet3.cell(row = 1, column = 1).font = ft1
		Sheet3.cell(row=1, column=2).font = ft1
		Sheet3.cell(row = 2, column = 1).font = ft1
		Sheet3.cell(row=2, column=2).font = ft1
		Sheet3.cell(row=2, column=2).number_format = 'YYYY-MM-DD'
		Sheet3.cell(row=5, column=5).font = cap_tabel

		Sheet3.cell(row=6, column=4).font = font_worksheet
		Sheet3.cell(row=7, column=4).font = font_worksheet
		Sheet3.cell(row=8, column=4).font = font_worksheet
		Sheet3.cell(row=9, column=4).font = font_worksheet
		Sheet3.cell(row=10, column=4).font = check_font

		Sheet3.cell(row=13, column=1).font = cap_tabel
		Sheet3.cell(row=13, column=2).font = cap_tabel
		Sheet3.cell(row=13, column=3).font = cap_tabel
		Sheet3.cell(row=13, column=4).font = cap_tabel

		Sheet3.cell(row=13, column=1).fill = cap_tabel_color_GT
		Sheet3.cell(row=13, column=2).fill = cap_tabel_color_GT
		Sheet3.cell(row=13, column=3).fill = cap_tabel_color_GT
		Sheet3.cell(row=13, column=4).fill = cap_tabel_color_GT
		Sheet3.cell(row = 8, column = 1).font = f_testname

		for row in Sheet3['E10:S10']:
			for cell in row:
				cell.font = check_font

		for row in Sheet3['A13:S13']:
			for cell in row:
				cell.fill = cap_tabel_color_GT

		for row in Sheet3['A13:R13']:
			for cell in row:
				cell.font = cap_tabel
		Sheet3.cell(row=13, column=19).font = check_font


		for row in Sheet3['E5:S5']:
			for cell in row:
				cell.fill = cap_tabel_color_GT

		for row in Sheet3['E5:S5']:
			for cell in row:
				cell.font = cap_tabel

		Sheet3.cell(row=5, column=19).font = check_font

		Sheet3.cell(row=1, column=18).font = ft1
		Sheet3.cell(row=1, column=19).font = ft1
		Sheet3.cell(row=2, column=18).font = ft1
		Sheet3.cell(row=2, column=19).font = ft1
		Sheet3.cell(row=3, column=18).font = ft1
		Sheet3.cell(row=3, column=19).font = check_font

		for row in Sheet3['E6:S10']:
			for cell in row:
				cell.number_format = '#,##0_);(#,##0)'

		for i in range(0, len(account)):
			Sheet3.cell(row=14+i, column=5).number_format = '#,##0_);(#,##0)'
			Sheet3.cell(row=14+i, column=6).number_format = '#,##0_);(#,##0)'
			Sheet3.cell(row=14 + i, column=7).number_format = '#,##0_);(#,##0)'
			Sheet3.cell(row=14 + i, column=8).number_format = '#,##0_);(#,##0)'
			Sheet3.cell(row=14 + i, column=9).number_format = '#,##0_);(#,##0)'
			Sheet3.cell(row=14 + i, column=10).number_format = '#,##0_);(#,##0)'
			Sheet3.cell(row=14 + i, column=11).number_format = '#,##0_);(#,##0)'
			Sheet3.cell(row=14 + i, column=12).number_format = '#,##0_);(#,##0)'
			Sheet3.cell(row=14 + i, column=13).number_format = '#,##0_);(#,##0)'
			Sheet3.cell(row=14 + i, column=14).number_format = '#,##0_);(#,##0)'
			Sheet3.cell(row=14 + i, column=15).number_format = '#,##0_);(#,##0)'
			Sheet3.cell(row=14 + i, column=16).number_format = '#,##0_);(#,##0)'
			Sheet3.cell(row=14 + i, column=17).number_format = '#,##0_);(#,##0)'
			Sheet3.cell(row=14 + i, column=18).number_format = '#,##0_);(#,##0)'
			Sheet3.cell(row=14 + i, column=19).number_format = '#,##0_);(#,##0)'
			Sheet3.cell(row=14 + i, column=19).font=check_font

		Sheet3.cell(row=6, column=19).font=check_font
		Sheet3.cell(row=7, column=19).font=check_font
		Sheet3.cell(row=8, column=19).font=check_font
		Sheet3.cell(row=9, column=19).font=check_font

		Sheet3.freeze_panes = 'D14'
		Sheet3.auto_filter.ref = "A13:S13"

		Sheet3.column_dimensions['A'].width = 10
		Sheet3.column_dimensions['B'].width = 15
		Sheet3.column_dimensions['C'].width = 11
		Sheet3.column_dimensions['D'].width = 12
		Sheet3.column_dimensions['E'].width = 14
		Sheet3.column_dimensions['F'].width = 12
		Sheet3.column_dimensions['G'].width = 10
		Sheet3.column_dimensions['H'].width = 12
		Sheet3.column_dimensions['I'].width = 12
		Sheet3.column_dimensions['J'].width = 12
		Sheet3.column_dimensions['K'].width = 12
		Sheet3.column_dimensions['L'].width = 12
		Sheet3.column_dimensions['M'].width = 12
		Sheet3.column_dimensions['N'].width = 12
		Sheet3.column_dimensions['O'].width = 12
		Sheet3.column_dimensions['P'].width = 15
		Sheet3.column_dimensions['Q'].width = 15
		Sheet3.column_dimensions['R'].width = 15
		Sheet3.column_dimensions['S'].width = 15


		list2=list(set(list2))
		list1=list(set(list1))
		if isChecked:
			list2nou=[]
			print(list2)
			for i in range(0,len(list2)):
				list2nou.append(str(list2[i]))
			list2nou.sort()
			for i in range(0, len(list2nou)):
				Sheet3.cell(row=14 + i, column=3).value = list2nou[i]
			for i in range(0, len(list2nou)):
				Sheet3.cell(row=14+i, column=5).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14+i, column=6).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=7).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=8).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=9).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=10).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=11).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=12).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=13).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=14).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=15).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=16).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=17).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=18).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=19).number_format = '#,##0_);(#,##0)'
			for i in range(0, len(list2nou)):
				Sheet3.cell(row=14+i,column=1).value='=left(C'+str(14+i)+',1)'
				Sheet3.cell(row=14+i,column=2).value='=left(C'+str(14+i)+',3)'
				Sheet3.cell(row=14+i,column=4).value="=vlookup(C"+str(14+i)+",'TB Robot'!C:D,2,0)"
				Sheet3.cell(row=14+i, column=5).value = '=SUMIF(INDIRECT(E$12&"!$A:$A"),C'+str(14+i)+',INDIRECT(E12&"!$Q:$Q"))'
				Sheet3.cell(row=14+i, column=6).value = '=SUMIF(INDIRECT(F$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(F$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=7).value = '=SUMIF(INDIRECT(G$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(G$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=8).value = '=SUMIF(INDIRECT(H$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(H$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=9).value = '=SUMIF(INDIRECT(I$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(I$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=10).value = '=SUMIF(INDIRECT(J$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(J$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=11).value = '=SUMIF(INDIRECT(K$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(K$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=12).value = '=SUMIF(INDIRECT(L$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(L$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=13).value = '=SUMIF(INDIRECT(M$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(M$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=14).value = '=SUMIF(INDIRECT(N$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(N$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=15).value = '=SUMIF(INDIRECT(O$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(O$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=16).value = '=SUMIF(INDIRECT(P$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(P$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=17).value = '=SUM(E'+str(14+i)+':P'+str(14+i)+')'
				if(str(list2[i])[:1]=="6"):
					Sheet3.cell(row=14 + i, column=18).value ="=SUMIF('TB Robot'!C:C,'Monthly P&L'!C"+str(14+i)+",'TB Robot'!H:H)"
				else:
					Sheet3.cell(row=14 + i, column=18).value ="=SUMIF('TB Robot'!C:C,'Monthly P&L'!C"+str(14+i)+",'TB Robot'!H:H)"
				Sheet3.cell(row=14 + i, column=19).value ='=Q'+str(14+i)+'-R'+str(14+i)+''

		else:
			print(list1)
			listanoua=[]
			for jj in range(0,len(list1)):
				if(list[jj] is not None):
					listanoua.append(str(list1))
			listanoua.sort()
			print(listanoua)
			for i in range(0, len(listanoua)):
				Sheet3.cell(row=14 + i, column=3).value = listanoua[i]
			for i in range(0, len(listanoua)):
				Sheet3.cell(row=14+i, column=5).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14+i, column=6).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=7).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=8).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=9).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=10).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=11).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=12).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=13).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=14).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=15).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=16).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=17).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=18).number_format = '#,##0_);(#,##0)'
				Sheet3.cell(row=14 + i, column=19).number_format = '#,##0_);(#,##0)'
			for i in range(0, len(listanoua)):
				Sheet3.cell(row=14+i,column=1).value='=left(C'+str(14+i)+',1)'
				Sheet3.cell(row=14+i,column=2).value='=left(C'+str(14+i)+',3)'
				Sheet3.cell(row=14+i,column=4).value="=vlookup(C"+str(14+i)+",'TB Robot'!C:D,2,0)"
				Sheet3.cell(row=14+i, column=5).value = '=SUMIF(INDIRECT(E12&"!$A:$A"),C'+str(14+i)+',INDIRECT(E12&"!$Q:$Q"))'
				Sheet3.cell(row=14+i, column=6).value = '=SUMIF(INDIRECT(F$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(F$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=7).value = '=SUMIF(INDIRECT(G$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(G$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=8).value = '=SUMIF(INDIRECT(H$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(H$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=9).value = '=SUMIF(INDIRECT(I$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(I$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=10).value = '=SUMIF(INDIRECT(J$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(J$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=11).value = '=SUMIF(INDIRECT(K$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(K$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=12).value = '=SUMIF(INDIRECT(L$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(L$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=13).value = '=SUMIF(INDIRECT(M$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(M$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=14).value = '=SUMIF(INDIRECT(N$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(N$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=15).value = '=SUMIF(INDIRECT(O$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(O$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=16).value = '=SUMIF(INDIRECT(P$12&"!$A:$A"),$C'+str(14+i)+',INDIRECT(P$12&"!$Q:$Q"))'
				Sheet3.cell(row=14 + i, column=17).value = '=SUM(E'+str(14+i)+':P'+str(14+i)+')'
				if(int(listanoua[i][:1])==6):
					Sheet3.cell(row=14 + i, column=18).value ="=SUMIF('TB Robot'!C:C,'Monthly P&L'!C"+str(14+i)+",'TB Robot'!H:H)"
				else:
					Sheet3.cell(row=14 + i, column=18).value ="=SUMIF('TB Robot'!C:C,'Monthly P&L'!C"+str(14+i)+",'TB Robot'!H:H)"
				Sheet3.cell(row=14 + i, column=19).value ='=Q'+str(14+i)+'-R'+str(14+i)+''
				# Sheet1.cell(row=2 + i, column=2).value = str(account_tb[i])[:4]
				# Sheet1.cell(row=2 + i, column=2).value = denumire[i]
				# Sheet1.cell(row=4 + i, column=3).value = DB[i]
				# Sheet1.cell(row=4 + i, column=4).value = CR[i]
				# Sheet1.cell(row=4 + i, column=5).value = DB_curr_month[i]
				# Sheet1.cell(row=4 + i, column=6).value = CR_curr_month[i]
				# Sheet1.cell(row=4 + i, column=7).value = DR_cumulat[i]
				# Sheet1.cell(row=4 + i, column=8).value = CR_cumulat[i]
				# Sheet1.cell(row=4 + i, column=9).value = total_DB[i]
				# Sheet1.cell(row=4 + i, column=10).value = total_CR[i]
				# Sheet1.cell(row=4 + i, column=11).value = DB_curr_sold[i]
				# Sheet1.cell(row=4 + i, column=12).value = CR_curr_sold[i]

		std = wb["Sheet"]
		wb.remove(std)
		folderpath="home/auditappnexia/output/MonthlyPL"
		file_pathFS = os.path.join(folderpath, "Monthly P&L" + " " + clientname1 + ".xlsx")
		wb.save(file_pathFS)
		
		for i in file_Details:
			os.remove(secure_filename(i.filename))
		
				
		
		return send_from_directory(folderpath, "Monthly P&L" + " " + clientname1 + ".xlsx",as_attachment=True)

	   


		# celula1=details1.cell(row=1, column=1).value
		# print(celula1)

	  
	# print(client)
	# print(yearEnd)
	# print(f)
	
	return render_template('MonthlyPL.html')
@app.route('/Reco/Instructions', methods=['GET'])
def downloadReco():
        filepath = "/home/auditappnexia/output/Reco"
 
        return send_from_directory(filepath,"Instructions - Reconciliation details vs TB.docx", as_attachment=True)
@app.route('/Reco/GT2ls9r79v3pr7L844am9wHNPCp')
def Reco():
    return render_template('Reco.html')
@app.route('/Reco/GT2ls9r79v3pr7L844am9wHNPCp', methods=['POST', 'GET'])
def Reco_process():
    
    def make_archive(source, destination):
        base = os.path.basename(destination)
        name = base.split('.')[0]
        format = base.split('.')[1]
        archive_from = os.path.dirname(source)
        archive_to = os.path.basename(source.strip(os.sep))
        shutil.make_archive(name, format, archive_from, archive_to)
        shutil.move('%s.%s'%(name,format), destination)

    clientname1 = request.form['client']
    filename="home/auditappnexia/output/Reco/"
    os.mkdir(filename+str(clientname1))
    yearEnd1= datetime.datetime.strptime(
                     request.form['yearEnd'],
                     '%Y-%m-%d')
    preparedBy1=request.form['preparedBy']
    if request.method == 'POST':
        ft2=Font(name='Tahoma',size=8,bold=True)
        ft1=Font(name='KPMG Logo',size=8)
        solidborder = Border(bottom=Side(style='thick'))
        blueFill = PatternFill(start_color='09b4ae',
                        end_color='ccd9ff',
                        fill_type='solid')
        headersblue= PatternFill(start_color='09b4ae',
                        end_color='09b4ae',
                        fill_type='solid')
        thin_border2 = Border(top=Side(style='double'))
        fontRed = Font(name='Tahoma', size=8, bold=True, color= 'FF0000')
        tah=Font(name='Tahoma',size=8)
        fontRedDiff=Font(color='FF0000', name='Tahoma', size=8)
        file_Details = request.files.getlist('detAcc')
        file_TB=request.files["TB"]
        
        
        for i in file_Details:
    
            i.save(secure_filename(i.filename))
        fisiere=list(file_Details)

        y=0
        start="/"

        y=str(fisiere[0]).count(start)

        fisiere2=[]
        for i in range(0,len(fisiere)):
            fisiere2.append(str(fisiere[i]).replace("/","",y-1))

        fisiere3=[]
        for i in range(0,len(fisiere)):
            fisiere3.append(fisiere2[i][fisiere2[i].find("/")+len("/"):fisiere2[i].rfind(".")]) 

        for z in range(0, len(fisiere)):
            
            wb=openpyxl.load_workbook(fisiere[z], data_only="True")
            ws=wb.active
            ws.title="PBC"
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value=="Account":
                        raa=cell.row
                        caa=cell.column
            lun=len(ws[caa])




            for row in ws.iter_rows():
                for cell in row:
                    if cell.value=="Date":
                        rdate=cell.row
                        cdate=cell.column

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value=="Document Number":
                        rdoc=cell.row
                        cdoc=cell.column

            ramount=None
            camount=None
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value=="Amount in FX":
                        ramount=cell.row
                        camount=cell.column

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value=="Amount":
                        rsold=cell.row
                        csold=cell.column

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value=="Description":
                        rdesc=cell.row
                        cdesc=cell.column
            ccur=None
            rcur=None         
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value=="Currency":
                        rcur=cell.row
                        ccur=cell.column
            cpart=None
            rpart=None         
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value=="Partner Name":
                        rpart=cell.row
                        cpart=cell.column
                    



            altaccount=[b.value for b in ws[caa][raa:lun]]
            listaunica=[]
            listaunica=list(set(altaccount))
            #listaunica.sort()
            curr=[]
            part=[]
            amountfx=[]
            documentno=[b.value for b in ws[cdoc][rdoc:lun]]
            date=[b.value for b in ws[cdate][rdate:lun]]
            if camount is not None:
                amountfx=[b.value for b in ws[camount][ramount:lun]]
            amount=[b.value for b in ws[csold][rsold:lun]]
            desc=[b.value for b in ws[cdesc][rdesc:lun]]
            if ccur is not None:
                curr=[b.value for b in ws[ccur][rcur:lun]]
            if cpart is not None:
                part=[b.value for b in ws[cpart][rpart:lun]]


            ws7=wb.create_sheet("Reconciliation")
            ws7.sheet_view.showGridLines = False

            ws7.cell(row=1, column=1).value="Client:"
            ws7.cell(row=1, column=2).value=clientname1
            ws7.cell(row=1, column=1).font=ft2
            ws7.cell(row=1, column=2).font=ft2
            
            ws7.cell(row=2, column=1).value="Period end:"
            ws7.cell(row=2, column=2).value=yearEnd1
            ws7.cell(row=2, column=2).number_format='mm/dd/yyyy'
            ws7.cell(row=2, column=1).font=ft2
            ws7.cell(row=2, column=2).font=ft2
            ws7.cell(row=1, column=6).value="Prepared by:"
            ws7.cell(row=1, column=7).value=preparedBy1
            ws7.cell(row=1, column=6).font=ft2
            ws7.cell(row=1, column=7).font=ft2
            ws7.cell(row=2, column=6).value="Date:"
            ws7.cell(row=2, column=7).value =datetime.date.today()
            ws7.cell(row=2, column=6).font=ft2
            ws7.cell(row=2, column=7).font=ft2
            ws7.cell(row=2, column=7).alignment = Alignment (horizontal='left')
            ws7.cell(row=2, column=7).number_format='mm/dd/yyyy'
            ws7.cell(row=3, column=6).value="Ref: "
            ws7.cell(row=3, column=6).font=fontRed
            ws7.cell(row=3, column=7).font=fontRed
            ws7.cell(row=5, column=3).value="Reconciliation of accounts with TB"

            ws7.cell(row=5, column=3).font=ft2
            ws7.cell(row=12, column=1).value="As per detail "
            ws7.cell(row=12, column=1).font=ft2
            ws7.cell(row=12, column=5).value="As per TB"
            ws7.cell(row=12, column=5).font=ft2
            ws7.cell(row=14, column=1).value="Account no."
            ws7.cell(row=14, column=1).font=ft2
            ws7.cell(row=14, column=1).border=solidborder
            ws7.cell(row=14, column=1).fill=headersblue
            ws7.cell(row=14, column=2).value="Description "
            ws7.cell(row=14, column=2).border=solidborder
            ws7.cell(row=14, column=2).font=ft2
            ws7.cell(row=14, column=2).fill=headersblue
            ws7.cell(row=14, column=3).value="Amount (As per Detail)"
            ws7.cell(row=14, column=3).fill=headersblue
            ws7.cell(row=14, column=3).border=solidborder
            ws7.cell(row=14, column=3).font=ft2
            ws7.cell(row=14, column=4).value="Amount (As per TB)"
            ws7.cell(row=14, column=4).fill=headersblue
            ws7.cell(row=14, column=4).border=solidborder
            ws7.cell(row=14, column=4).font=ft2
            ws7.cell(row=14, column=5).value="Check/Difference"
            ws7.cell(row=14, column=5).fill=headersblue
            ws7.cell(row=14, column=5).border=solidborder
            ws7.cell(row=14, column=5).font=fontRed
            "Adjust Column Width---------------------------------------------------------------------------------------------------------------------------------."
            

            for col in ws7.columns:
                max_length=20
                for cell in col:
                    if cell.coordinate in ws7.merged_cells:
                        continue
                    try:
                        if len(str(cell.value))>max_length:
                            max_length=len(cell.value)
                    except:
                        pass
                adjusted_width=(max_length-1)

            listaAdjusted=['H', 'C', 'F', 'I', 'D']
            for column in listaAdjusted:
                for i in listaAdjusted:
                    if (column==i):
                        ws7.column_dimensions[column].width=adjusted_width		


            for i in range(0, len(listaunica)):
                for j in range(1,9):
                    ws7.cell(row=i+15, column=1).value=listaunica[i]
                    ws7.cell(row=i+15,column=j).font=tah
            for row in ws7.iter_rows():
                for cell in row:
                    if cell.value=="Account no.":
                        raccountno=cell.row
                        caccountno=cell.column

            ws7.cell(row=15+len(listaunica), column = 2).value="Total"
            ws7.cell(row=15+len(listaunica), column = 2).font=ft2
            ws7.cell(row=15+len(listaunica), column=2).border=thin_border2
            ws7.cell(row=15+len(listaunica), column=1).border=thin_border2
            ws7.cell(row=15+len(listaunica), column=3).border=thin_border2
            ws7.cell(row=15+len(listaunica), column=4).border=thin_border2
            ws7.cell(row=15+len(listaunica), column=5).border=thin_border2

            for i in range(15,15+len(listaunica)):
                for j in range (3,30):
                        ws7.cell(row=i, column=j).number_format='#,##0_);(#,##0)'

            x=str(len(listaunica)+14)
            ws7.cell(row=len(listaunica)+15, column=3).value="=sum(C15:C"+x+")"
            ws7.cell(row=len(listaunica)+15, column=3).font=ft2
            ws7.cell(row=len(listaunica)+15, column=3).number_format="#,##0"
            ws7.cell(row=len(listaunica)+15, column=4).value="=sum(D15:D"+x+")"
            ws7.cell(row=len(listaunica)+15, column=4).font=ft2
            ws7.cell(row=len(listaunica)+15, column=4).number_format="#,##0"
            ws7.cell(row=len(listaunica)+15, column=5).value="=sum(E15:E"+x+")"
            ws7.cell(row=len(listaunica)+15, column=5).font=fontRed
            ws7.cell(row=len(listaunica)+15, column=5).number_format="#,##0"

            ws8=wb.create_sheet("Detail of account")
			# ws8.sheet_view.showGridLines = False
            ws8.cell(row=9, column=1).value="Account"
			# ws8.sheet_view.showGridLines = False

            ws8.cell(row=9, column=4).value="Date"
            ws8.cell(row=9, column=3).value="Description"
            ws8.cell(row=9, column=5).value="Amount"
            ws8.cell(row=9, column=2).value="Document Number"
            ws8.cell(row=9, column=9).value="Month"
            

            for j in range(1,10):
                ws8.cell(row=9, column=j).font=ft2
                ws8.cell(row=9,column=j).fill=headersblue
                ws8.cell(row=9, column=j).border=solidborder

            



            for i in range(0, len(listaunica)):
                ws7.cell(row=i+15, column=2).value="=VLOOKUP(A{0},'Trial Balance'!A:B,2,TRUE)".format(i+15)
                ws7.cell(row=i+15, column=3).value="=SUMIF('Detail of account'!A:A,A{0},'Detail of Account'!E:E)".format(i+15)
                ws7.cell(row=i+15, column=4).value="=SUMIF('Trial Balance'!A:A,Reconciliation!A{0},'Trial Balance'!C:C)".format(i+15)
                ws7.cell(row=i+15, column=5).value="=C{0}-D{0}".format(i+15)
                ws7.cell(row=i+15, column=5).font=fontRedDiff

            ws7.cell(row=44+len(listaunica), column=1).value="Findings:"
            ws7.cell(row=44+len(listaunica), column=1).font=ft2

            ws7.cell(row=len(listaunica)+17, column=1).value="Account"
            ws7.cell(row=len(listaunica)+17, column=2).value="January"
            ws7.cell(row=len(listaunica)+17, column=3).value="February"
            ws7.cell(row=len(listaunica)+17, column=4).value="March"
            ws7.cell(row=len(listaunica)+17, column=5).value="April"
            ws7.cell(row=len(listaunica)+17, column=6).value="May"
            ws7.cell(row=len(listaunica)+17, column=7).value="June"
            ws7.cell(row=len(listaunica)+17, column=8).value="July"
            ws7.cell(row=len(listaunica)+17, column=9).value="August"
            ws7.cell(row=len(listaunica)+17, column=10).value="September"
            ws7.cell(row=len(listaunica)+17, column=11).value="October"
            ws7.cell(row=len(listaunica)+17, column=12).value="November"
            ws7.cell(row=len(listaunica)+17, column=13).value="December"
            ws7.cell(row=len(listaunica)+17, column=14).value="Total"


            ws7.cell(row=len(listaunica)+18, column=2).value="=SUMIF('Detail of account'!F:F,1,'Detail of account'!E:E)"
            ws7.cell(row=len(listaunica)+18, column=3).value="=SUMIF('Detail of account'!F:F,2,'Detail of account'!E:E)"
            ws7.cell(row=len(listaunica)+18, column=4).value="=SUMIF('Detail of account'!F:F,3,'Detail of account'!E:E)"
            ws7.cell(row=len(listaunica)+18, column=5).value="=SUMIF('Detail of account'!F:F,4,'Detail of account'!E:E)"
            ws7.cell(row=len(listaunica)+18, column=6).value="=SUMIF('Detail of account'!F:F,5,'Detail of account'!E:E)"
            ws7.cell(row=len(listaunica)+18, column=7).value="=SUMIF('Detail of account'!F:F,6,'Detail of account'!E:E)"
            ws7.cell(row=len(listaunica)+18, column=8).value="=SUMIF('Detail of account'!F:F,7,'Detail of account'!E:E)"
            ws7.cell(row=len(listaunica)+18, column=9).value="=SUMIF('Detail of account'!F:F,8,'Detail of account'!E:E)"
            ws7.cell(row=len(listaunica)+18, column=10).value="=SUMIF('Detail of account'!F:F,9,'Detail of account'!E:E)"
            ws7.cell(row=len(listaunica)+18, column=11).value="=SUMIF('Detail of account'!F:F,10,'Detail of account'!E:E)"
            ws7.cell(row=len(listaunica)+18, column=12).value="=SUMIF('Detail of account'!F:F,11,'Detail of account'!E:E)"
            ws7.cell(row=len(listaunica)+18, column=13).value="=SUMIF('Detail of account'!F:F,12,'Detail of account'!E:E)"
            ws7.cell(row=len(listaunica)+18, column=14).value="=SUM(B"+str(len(listaunica)+18)+":M"+str(len(listaunica)+18)+")"

            for i in range(1,15):
                ws7.cell(row=len(listaunica)+17, column=i).font=ft2
                ws7.cell(row=len(listaunica)+17,column=i).fill=headersblue
                ws7.cell(row=len(listaunica)+17, column=i).border=solidborder

            for ii in range(2, 15):
                ws7.cell(row=len(listaunica)+17, column=ii).number_format="#,##0"


                # y = str(len(listaunica)+19)
                # ws7.cell(row=20+len(listaunica), column=2).value="=LEFT(A"+y+",A"+")"

            # incep=ws7.max_row+5
            # for i in range(0,len(listaunica)):
            # 		ws7.cell(row=incep+i, column=1).value=listaunica[i]
            
            listaMihai=[]
            for i in range(0,len(listaunica)):
                    listaMihai.append(str(listaunica[i]))
            
            listaMihai2=[]
            for i in range(0, len(listaunica)):listaMihai2.append((listaMihai[i][0:3]))
            #print(listaMihai2)

            listaMihai3=[]
            #for i in range(0, len(listaMihai2)):listaMihai3.append(set(listaMihai2))
            listaMihai3=list(set(listaMihai2))
            print(listaMihai3)
            listaMihai4=[] 
            for i in range(0, len(listaMihai3)):listaMihai4.append(int(listaMihai3[i]))
            #for i in range(0, len(listaMihai3)):ws7.cell(row=4, column=8+i).value=listaMihai4[i]
            string = '; '.join(map(str, listaMihai4))

            ws7.cell(row=8, column=1).value="Work done:"
            ws7.cell(row=8, column=1).font=ft2
            # ws7.cell(row=8, column=2).value="We reconciliated the detail of account " +string+" with TB @ "+str(yearEnd1)+"."
            ws7.cell(row=8, column=2).value="We reconciliated the detail of account " +string+" with TB@ 31.12.20XX."
            ws7.cell(row=9, column=2).value="Please see the work below:"
			# ws7.cell(row=len(listaunica)+18, column=1).value=string
            ws7.cell(row=len(listaunica)+18, column=1).value=string            

                    #ws7.cell(row=incep+i, column=1).value=listaMihai[i][0:3]

            # my = [1234, 3456, 6789]
            # mylist = my


                
            # listaMihai=[]
            # for i in range(0, len(mylist)):
            # 	listaMihai.append(str(mylist[i]))

            # print(listaMihai)
            # for i in range(0, len(listaMihai)):
            # 	print(listaMihai[i][0:3])


            # for i in range(incep, incep+len(listaunica)):
            # 	listaMihai.append(ws7.cell(row=i,column=2).value)

            # print(listaMihai)

            ws8.cell(row=5, column=3).value="Total as per detail:"
            ws8.cell(row=5,column=3).font=ft2
            ws8.cell(row=6, column=3).value="Total as per TB:"
            ws8.cell(row=6,column=3).font=ft2
            ws8.cell(row=7, column=3).value="Total Diff:"
            ws8.cell(row=7,column=3).font=fontRed
            ws8.cell(row=6,column=3).border=solidborder
            ws8.cell(row=6,column=4).border=solidborder

            ws8.cell(row=5, column=4).value="=SUM(E10:E"+str(10+len(altaccount))+")"
            ws8.cell(row=5, column=4).font=tah
            ws8.cell(row=6, column=4).value="=Reconciliation!D"+str(len(listaunica)+15)
            ws8.cell(row=6, column=4).font=tah
            ws8.cell(row=7, column=4).value="=D6-D5"
            ws8.cell(row=7, column=4).font=fontRed

            ws8.cell(row=1, column=1).font=ft2
            ws8.cell(row=1, column=1).value="Client:"
            ws8.cell(row=1, column=2).font=ft2
            ws8.cell(row=1, column=2).value=clientname1

            ws8.cell(row=2, column=2).font=ft2
            ws8.cell(row=2, column=1).value="PE:"
            ws8.cell(row=2, column=1).font=ft2
            ws8.cell(row=2, column=2).value=yearEnd1
            ws8.cell(row=2, column=2).number_format='mm/dd/yyyy'
            
            ws8.cell(row=3, column=2).font=ft2
            ws8.cell(row=1, column=6).value="Prepared by:"
            ws8.cell(row=1, column=6).font=ft2
            ws8.cell(row=1, column=7).value=preparedBy1
            ws8.cell(row=1, column=7).font=tah
            ws8.cell(row=1, column=7).font=ft2
            ws8.cell(row=2, column=6).value="Date:"
            ws8.cell(row=2, column=6).font=ft2
            ws8.cell(row=2, column=7).value=datetime.date.today()
            ws8.cell(row=2, column=7).font=ft2
            ws8.cell(row=2, column=7).number_format='mm/dd/yyyy'
            ws8.cell(row=2, column=7).alignment = Alignment (horizontal='left')
            ws8.cell(row=3, column=6).value="Ref:"
            ws8.cell(row=3, column=6).font=fontRed
            ws8.cell(row=3, column=7).font=fontRed

            for i in range(10,len(altaccount)+11):
                for j in range(1,9):
                    ws8.cell(row=i,column=j).font=tah

            for i in range(5,8):
                ws8.cell(row=i,column=5).number_format="#,##0_);(#,##0)"      
            for i in range(10,len(altaccount)+11):
                ws8.cell(row=i,column=5).number_format="#,##0_);(#,##0)"
                ws8.cell(row=i,column=6).number_format="#,##0_);(#,##0)"
            

            
            for i in range(0, len(altaccount)):
                ws8.cell(row=i+10, column=1).value=altaccount[i]
                ws8.cell(row=i+10, column=4).value=date[i]
                ws8.cell(row=i+10, column=4).number_format='mm/dd/yyyy'
                ws8.cell(row=i+10, column=6).value='=MONTH(D{0})'.format(i+10)
                ws8.cell(row=i+10, column=3).value=desc[i]
                ws8.cell(row=i+10, column=5).value=amount[i]
                ws8.cell(row=i+10, column=2).value=documentno[i]
                if len(amountfx)>0:
                    ws8.cell(row=9,column=7).value="Amount in Doc Currency"
                    ws8.cell(row=i+10, column=7).value=amountfx[i]
                    ws8.cell(row=9,column=7).font=ft2
                    ws8.cell(row=9,column=7).fill=headersblue
                    ws8.cell(row=9,column=7).border=solidborder
                        

                    if len(curr)>0:
                        ws8.cell(row=9,column=8).value="Currency"
                        ws8.cell(row=9,column=8).font=ft2
                        ws8.cell(row=9,column=8).fill=headersblue
                        ws8.cell(row=9,column=8).border=solidborder
                        
                        ws8.cell(row=i+10, column=8).value=curr[i]
                        if len(part)>0:
                            ws8.cell(row=9,column=9).value="Partner name"
                            ws8.cell(row=9,column=9).font=ft2
                            ws8.cell(row=9,column=9).fill=headersblue
                            ws8.cell(row=9,column=9).border=solidborder
                            ws8.cell(row=i+10,column=9).value=part[i]
                else:
                    if len(part)>0:
                        ws8.cell(row=9,column=10).value="Partner name"
                        ws8.cell(row=9,column=10).font=ft2
                        ws8.cell(row=9,column=10).fill=headersblue
                        ws8.cell(row=9,column=10).border=solidborder
                        ws8.cell(row=i+10, column=10).value=part[i]
                
            column_count=ws7.max_row	


            tb=openpyxl.load_workbook(file_TB, data_only=True)
            wstb=tb.active
            wstt=wb.create_sheet("Trial Balance")
            wstt.cell(row=1, column=1).value="Account"
            wstt.cell(row=1, column=2).value="Description"
            wstt.cell(row=1, column=3).value="CB"
            wstt.cell(row=1, column=4).value="Synth(3)"

            for i in range(1, 5):
                wstt.cell(row=1,column=i).font=ft2
                wstt.cell(row=1,column=i).fill=headersblue
                wstt.cell(row=1,column=i).border=solidborder


            for row in wstb.iter_rows():
                for cell in row:
                    if cell.value=="Account":
                        raaa=cell.row
                        caaa=cell.column

            for row in wstb.iter_rows():
                for cell in row:
                    if cell.value=="Description":
                        rdesc=cell.row
                        cdesc=cell.column

            for  row in wstb.iter_rows():
                for cell in row:
                    if cell.value=="CB":
                        rclos=cell.row
                        cclos=cell.column

            luntb=len(wstb[caaa])
            account=[c.value for c in wstb[caaa][raaa:luntb]]
            desc=[c.value for c in wstb[cdesc][rdesc:luntb]]
            closing=[c.value for c in wstb[cclos][rclos:luntb]]

            for i in range(0,len(account)):
                wstt.cell(row=i+2, column=1).value=account[i]
                wstt.cell(row=i+2, column=2).value=desc[i]
                wstt.cell(row=i+2, column=3).value=closing[i]
                wstt.cell(row=i+2, column=4).value=str(account[i])[:3]

            rand_max=wstt.max_row
            for i in range(2,rand_max):
                for j in range (3,4):
                    wstt.cell(row=i, column=j).number_format='#,##0_);(#,##0)'

            Sheet = "Sheet"
            ws7.cell(row=ws7.max_row,column=3).number_format="#,##0_);(#,##0)"
            ws7.cell(row=ws7.max_row,column=5).number_format="#,##0_);(#,##0)"
            ws7.cell(row=ws7.max_row,column=7).number_format="#,##0_);(#,##0)"  

            myorder=[1,2,3,0]
            wb._sheets=[wb._sheets[i] for i in myorder]
            c=ws7['A15']
            ws7.freeze_panes = c

            c=ws8['A10']
            ws8.freeze_panes = c


            
            ws8.cell(row=5,column=4).number_format="#,##0"
            ws8.cell(row=5,column=4).font=ft2
            ws8.cell(row=6,column=4).number_format="#,##0"
            ws8.cell(row=6, column=4).font=ft2
            ws8.cell(row=7, column=4).number_format="#,##0"
            ws8.column_dimensions['B'].width=17
            ws8.column_dimensions['C'].width=20
            ws8.column_dimensions['D'].width=15
            ws7.column_dimensions['A'].width=16
            ws7.column_dimensions['B'].width=16
            ws7.column_dimensions['G'].width=18
            ws7.column_dimensions['E'].width=18

            ws9 = wb.create_sheet("Transpose")
            ws9.sheet_state = 'hidden'

            ws9.cell(row=1, column=1).value="Month"
            ws9.cell(row=2, column=1).value="January"
            ws9.cell(row=3, column=1).value="February"
            ws9.cell(row=4, column=1).value="March"
            ws9.cell(row=5, column=1).value="April"
            ws9.cell(row=6, column=1).value="May"
            ws9.cell(row=7, column=1).value="June"
            ws9.cell(row=8, column=1).value="July"
            ws9.cell(row=9, column=1).value="August"
            ws9.cell(row=10, column=1).value="September"
            ws9.cell(row=11, column=1).value="October"
            ws9.cell(row=12, column=1).value="November"
            ws9.cell(row=13, column=1).value="December"

            # ws9.cell(row=1, column=2).value="Monthly Trend"
            ws9.cell(row=2, column=2).value="=Reconciliation!B"+str(len(listaunica)+18)
            ws9.cell(row=3, column=2).value="=Reconciliation!C"+str(len(listaunica)+18)
            ws9.cell(row=4, column=2).value="=Reconciliation!D"+str(len(listaunica)+18)
            ws9.cell(row=5, column=2).value="=Reconciliation!E"+str(len(listaunica)+18)
            ws9.cell(row=6, column=2).value="=Reconciliation!F"+str(len(listaunica)+18)
            ws9.cell(row=7, column=2).value="=Reconciliation!G"+str(len(listaunica)+18)
            ws9.cell(row=8, column=2).value="=Reconciliation!H"+str(len(listaunica)+18)
            ws9.cell(row=9, column=2).value="=Reconciliation!I"+str(len(listaunica)+18)
            ws9.cell(row=10, column=2).value="=Reconciliation!J"+str(len(listaunica)+18)
            ws9.cell(row=11, column=2).value="=Reconciliation!K"+str(len(listaunica)+18)
            ws9.cell(row=12, column=2).value="=Reconciliation!L"+str(len(listaunica)+18)
            ws9.cell(row=13, column=2).value="=Reconciliation!M"+str(len(listaunica)+18)

            values = Reference(ws9,
                    min_col=2,  # I
                    max_col=2,  # T
                    min_row=1,
                    max_row=13)
            labels=Reference(ws9,
                    min_col=1,  # I
                    max_col=1,  # T
                    min_row=2,
                    max_row=13)

            chart = LineChart()
            chart.add_data(values, titles_from_data=True)
            chart.set_categories(labels)

            chart.title = "Monthly Variation of Account " +str(altaccount[0])[:3]
            chart.x_axis.title = ""
            chart.y_axis.title = "" 
            ws7.add_chart(chart, "A"+str(len(listaunica)+20))

            file_path=os.path.join(filename+str(clientname1), "Detail of account "+str(altaccount[0])[:3]+".xlsx")
            wb.save(file_path)
        make_archive(filename+str(clientname1),filename+str(clientname1)+".zip")
        # return send_from_directory("C:\\Users\\Cristin.Iordache\\Documents\\Automation Projects\\Web apps-v2-login\\Output\\Reco\\",clientname1+".zip",as_attachment=True)
        return send_from_directory("home/auditappnexia/output/Reco/",clientname1+".zip",as_attachment=True)
    return render_template('Reco.html')

        # 	filename="C:\\Users\\Cristian.Iordache\\Documents\\Automation Projects\\Web apps-v2-login\\Output\\Reco"
            
        # 	file_path=os.path.join(filename, "Reconciliation with TB "+fisiere3[z]+".xlsx")
        # 	wb.save(file_path)
        # return send_from_directory(file_path, "Reconciliation with TB "+fisiere3[z]+".xlsx",as_attachment=True)

@app.route('/TrialBalances/Instructions', methods=['GET'])
def downloadTB():
		filepath = "/home/auditappnexia/output/tb"
 
		return send_from_directory(filepath,"Instructions - Trial Balance.docx", as_attachment=True)
@app.route('/TrialBalances/GT3SjGyxpbcxV35PeSUpKJQIOgY')
def TB():
	return render_template('TB.html')
@app.route('/TrialBalances/GT3SjGyxpbcxV35PeSUpKJQIOgY', methods=['POST', 'GET'])
def TB_process():

	namec = request.form['client']
	ant= datetime.datetime.strptime(
					 request.form['yearEnd'],
					 '%Y-%m-%d')
	threshol = request.form['threshold']
	preparedBy1 = request.form['preparedBy']
	isChecked1=request.form.get("Stdmapp")
	isChecked2=request.form.get("forml")
	isChecked3=request.form.get("forms")
	isChecked4=request.form.get("pyEx")


		# if isChecked1=="": #daca e bifat
	#     isChecked=1
	# else:
	#     isChecked=0
	
	folderpath="home/auditappnexia/output/tb"

	def make_archive(source, destination):
		base = os.path.basename(destination)
		name = base.split('.')[0]
		format = base.split('.')[1]
		archive_from = os.path.dirname(source)
		archive_to = os.path.basename(source.strip(os.sep))
		shutil.make_archive(name, format, archive_from, archive_to)
		shutil.move('%s.%s'%(name,format), destination)
	# yearEnd = str(request.form['yearEnd'])
	# processed_text = client.upper()
	# fisier=request.files.get('monthlyTB')
	if request.method == 'POST':
		workingsblue2= Font(bold=True, italic=True, name='Tahoma', size=8,color='FFFFFF')
		lbluefill = PatternFill(start_color='7030A0',
							end_color='7030A0',
							fill_type='solid')
		grifill=PatternFill(start_color='c4d79b',end_color='c4d79b',fill_type='solid')
		yellow=PatternFill(start_color='ffff00',end_color='ffff00',fill_type='solid')
		blueFill = PatternFill(start_color='00AEAC',
							end_color='00AEAC',
							fill_type='solid')
		doubleborder = Border(bottom=Side(style='double'))
		solidborder = Border(bottom=Side(style='thick'))
		solidborderstanga = Border(left=Side(style='thin'))
		rightborder = Border(right=Side(style='thin'))
		rightdouble = Border (right=Side(style='thin'), bottom=Side(style='double'))
		rightmedium = Border (right=Side(style='thin'), bottom=Side(style='medium'))
		solidborderdreapta = Border(right=Side(style='thin'))
		solidbordersus = Border(top=Side(style='thin'))
		fontitalic = Font(name='Tahoma', size=8, bold=True, italic=True)
		font = Font(name='Tahoma', size=8, bold=True)
		font1 = Font(name='Tahoma', size=8)
		font2 = Font(name='Tahoma', size=10, bold=True)
		fontRed = Font(name='Tahoma', size=10, bold=True, color= 'FF0000')
		fontRedDiff=Font(name="Tahoma", color='FF0000', size=11, )
		fontGT = Font (name='GT Logo', size=8)
		workingsblue = Font(color='2F75B5', bold=True, name='Tahoma', size=8 )
		headers= Font(bold=True, italic=True, name='Tahoma', size=8,color='FFFFFF') 
		headersblue = PatternFill(start_color='7030A0',
						end_color='7030A0',
						fill_type='solid')
		headerspurple= PatternFill(start_color='65CDCC',
							end_color='65CDCC',
							fill_type='solid')
		total=PatternFill(start_color='DDD9C4',
						end_color='DDD9C4',
						fill_type='solid')
		greenbolditalic= Font(bold=True, italic=True,  color='C0504D', name='Tahoma', size=8)
		greenbolditalic= Font(bold=True, italic=True,  color='00af50')
		fontalb = Font(italic=True, color="bfbfbf", size=8, name='Tahoma')
		trialc=request.files["trialBalCYPBC"]
		trialp=request.files["trialBalPYPBC"]
		TBCY = openpyxl.load_workbook(trialc,data_only=True)
		TBCY1 = TBCY.active

		# if isChecked4=="":
		# 	try:
		
		"Open files"




		"Iterate from imported PBC's:"


		'Iterate from CY TB'

		for row in TBCY1.iter_rows():
				for cell in row:
					if cell.value=="Account":
						tbCyAcount=cell.column
						tbrow=cell.row

		for row in TBCY1.iter_rows():
			for cell in row:
				if cell.value=="Description":
					tbCyDescription=cell.column

		for row in TBCY1.iter_rows():
			for cell in row:
				if cell.value=="SID":
					tbCySID=cell.column

		for row in TBCY1.iter_rows():

			for cell in row:
				if cell.value=="SIC":
					tbCySIC=cell.column
				
		for row in TBCY1.iter_rows():
			for cell in row:
				if cell.value=="RCD":
					tbCyRCD=cell.column

		for row in TBCY1.iter_rows():
			for cell in row:
				if cell.value=="RCC":
					tbCyRCC=cell.column

		for row in TBCY1.iter_rows():
			for cell in row:
				if cell.value=="SFD":
					tbCySFD=cell.column

		for row in TBCY1.iter_rows():
			for cell in row:
				if cell.value=="SFC":
					tbCySFC=cell.column


		try:
			luntb=len(TBCY1[tbCyAcount])
		except:
			flash("Please insert the correct header for Account in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Account'")
			# sys.exit()
		try:
			Account=[b.value for b in TBCY1[tbCyAcount][tbrow:luntb+1]]
		except:
			flash("Please insert the correct header for Account in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Account'")
			# sys.exit()

		try:
			Description=[b.value for b in TBCY1[tbCyDescription][tbrow:luntb+1]]
		except:
			flash("Please insert the correct header for Description in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Description'")
			# sys.exit()
		try:
			SID=[b.value for b in TBCY1[tbCySID][tbrow:luntb+1]]
		except:
			flash("Please insert the correct header for Sold Initial Debit in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Sold Initial Debit'")
			# sys.exit()
		try:
			SIC=[b.value for b in TBCY1[tbCySIC][tbrow:luntb+1]]
		except:
			flash("Please insert the correct header for Sold Initial Credit in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Sold Initial Credit'")
			# sys.exit()
		try:
			RCD=[b.value for b in TBCY1[tbCyRCD][tbrow:luntb+1]]
		except:
			flash("Please insert the correct header for Rulaj Curent Debit in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Rulaj Curent Debit'")
			# sys.exit()
		try:
			RCC=[b.value for b in TBCY1[tbCyRCC][tbrow:luntb+1]]
		except:
			flash("Please insert the correct header for Rulaj Curent Credit in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Rulaj Curent credit'")
			# sys.exit()
		try:
			SFD=[b.value for b in TBCY1[tbCySFD][tbrow:luntb+1]]
		except:
			flash("Please insert the correct header for Sold Final Debit in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Sold Final Debit'")
			# sys.exit()
		try: 
			SFC=[b.value for b in TBCY1[tbCySFC][tbrow:luntb+1]]
		except:
			flash("Please insert the correct header for Sold Final Credit in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Sold Final Credit'")
			# sys.exit()
		"Create CY PBC"

		if isChecked4=="":
			TBPY = openpyxl.load_workbook(trialp,data_only=True)
			TBPY1 = TBPY.active
			try:
				for row in TBPY1.iter_rows():
						for cell in row:
							if cell.value=="Account":
								tbPyAcount=cell.column
								tbPYrow=cell.row

				for row in TBPY1.iter_rows():
					for cell in row:
						if cell.value=="Description":
							tbPyDescription=cell.column


				for row in TBPY1.iter_rows():
					for cell in row:
						if cell.value=="CB":
							tbPySFD=cell.column


				try:
					luntbp=len(TBPY1[tbPyAcount])
				except:
					flash("Please insert the correct header for Account in Trial Balance Prior Year file")
					return render_template("index.html")
					# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Account'")
					# sys.exit()
				try:
					Accountp=[b.value for b in TBPY1[tbPyAcount][tbPYrow:luntb+1]]
				except:
					flash("Please insert the correct header for Account in Trial Balance Prior Year file")
					return render_template("index.html")
					# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Account'")
					# sys.exit()

				try:
					Descriptionp=[b.value for b in TBPY1[tbPyDescription][tbPYrow:luntb+1]]
				except:
					flash("Please insert the correct header for Description in Trial Balance Prior Year file")
					return render_template("index.html")
					# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Description'")
					# sys.exit()

				try:
					CB=[b.value for b in TBPY1[tbPySFD][tbPYrow:luntb+1]]
				except:
					flash("Please insert the correct header for CB in Trial Balance Prior Year file")
					return render_template("index.html")
					# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Sold Final Debit'")
					# sys.exit()
			except:
				pass
		
		if isChecked1=="":
			mapp=request.files["Mapping"]
			output=openpyxl.load_workbook(mapp,data_only=True)
		else:
			if isChecked2=="":
				output=openpyxl.load_workbook("home/auditappnexia/output/Other files/tb/Mapping/Mapping Forma Lunga.xlsx",data_only=True)
			else:
				output=openpyxl.load_workbook("home/auditappnexia/output/Other files/tb/Mapping/Mapping Forma Scurta.xlsx",data_only=True)

		PBC_CY =output.create_sheet("PBC_CY")

		PBC_CY.cell(row=1, column=1).value="Class"
		PBC_CY.cell(row=1, column=2).value="Synt3"
		PBC_CY.cell(row=1, column=3).value="Synt4"
		PBC_CY.cell(row=1, column=4).value="Account"
		PBC_CY.cell(row=1, column=5).value="Description"
		PBC_CY.cell(row=1, column=6).value="SID"
		PBC_CY.cell(row=1, column=7).value="SIC"
		PBC_CY.cell(row=1, column=8).value="RCD"
		PBC_CY.cell(row=1, column=9).value="RCC"
		PBC_CY.cell(row=1, column=10).value="SFD"
		PBC_CY.cell(row=1, column=11).value="SFC"


		for i in range (1,10):
			PBC_CY.cell (row=1, column=i).border=doubleborder
			PBC_CY.cell (row=1, column=i).font=font2


		for i in range(1, len(Account)+1):
			PBC_CY.cell(row=i+1, column=4).value=Account[i-1]

		for i in range (1, len(Description)+1):
			PBC_CY.cell(row=i+1, column=5).value= Description[i-1]

		for i in range (1, len(SID)+1):
			PBC_CY.cell(row=i+1, column=6).value=SID[i-1]

		for i in range (1, len(SIC)+1):
			PBC_CY.cell (row=i+1, column =7).value=SIC[i-1]

		for i in range (1,len(RCD)+1):
			PBC_CY.cell (row=i+1, column=8).value=RCD[i-1]

		for i in range (1,len(RCC)+1):
			PBC_CY.cell (row=i+1, column=9).value=RCC[i-1]

		for i in range (1,len(SFD)+1):
			PBC_CY.cell (row=i+1, column=10).value=SFD[i-1]

		for i in range(1,len(SFC)+1):
			PBC_CY.cell (row=i+1, column=11).value=SFC[i-1]

		for i in range (1,12):
			PBC_CY.cell(row=1, column=i).font=font2
			PBC_CY.cell(row=1, column=i).border=doubleborder
			PBC_CY.cell(row=1, column=i).fill=blueFill

		for i in range (1, len(SFD)+1):
			for j in range (6, 12):
				PBC_CY.cell(row=i+1, column=j).number_format='#,##0_);(#,##0)'


		for i in range(1, len(Account)+1):
			PBC_CY.cell(row=i+1,column=1).value='=Left(D{0},1)'.format(i+1)
		for i in range(1,len(Account)+1):
			PBC_CY.cell(row=i+1,column=2).value='=Left(D{0},3)'.format(i+1)
		for i in range(1,len(Account)+1):
			PBC_CY.cell(row=i+1,column=3).value='=Left(D{0},4)'.format(i+1)


		PBC_PY =output.create_sheet("PBC_PY")

		if isChecked4=="":
			try:
				PBC_PY.cell(row=1, column=1).value="Class"
				PBC_PY.cell(row=1, column=2).value="Synt3"
				PBC_PY.cell(row=1, column=3).value="Synt4"
				PBC_PY.cell(row=1, column=4).value="Account"
				PBC_PY.cell(row=1, column=5).value="Description"
				PBC_PY.cell(row=1, column=6).value="CB"


				for i in range (1,8):
					PBC_PY.cell (row=1, column=i).border=doubleborder
					PBC_PY.cell (row=1, column=i).font=font2


				for i in range(1, len(Accountp)+1):
						PBC_PY.cell(row=i+1, column=4).value=Accountp[i-1]

				for i in range (1, len(Descriptionp)+1):
					PBC_PY.cell(row=i+1, column=5).value= Descriptionp[i-1]


				for i in range (1,len(CB)+1):
					PBC_PY.cell (row=i+1, column=6).value=CB[i-1]

				for i in range (1,8):
					PBC_PY.cell(row=1, column=i).font=font2
					PBC_PY.cell(row=1, column=i).border=doubleborder
					PBC_PY.cell(row=1, column=i).fill=blueFill

				for i in range (1, len(CB)+1):
					for j in range (6, 8):
						PBC_PY.cell(row=i+1, column=j).number_format='#,##0_);(#,##0)'


				for i in range(1, len(Accountp)+1):
					PBC_PY.cell(row=i+1,column=1).value='=Left(D{0},1)'.format(i+1)
				for i in range(1,len(Accountp)+1):
					PBC_PY.cell(row=i+1,column=2).value='=Left(D{0},3)'.format(i+1)
				for i in range(1,len(Accountp)+1):
					PBC_PY.cell(row=i+1,column=3).value='=Left(D{0},4)'.format(i+1)
			except:
				pass

		"Define F10 Worksheet"

		F10TB=output.create_sheet("F_10_Trial_Balance")
		F10TB.sheet_view.showGridLines = False
		F10TB.cell(row=1, column=1).value="Client:"
		F10TB.cell(row=1, column=1).font=font
		F10TB.cell(row=1, column=2).value=namec
		F10TB.cell(row=1, column=2).font=font


		F10TB.cell(row=2, column=1).value="Period end:"
		F10TB.cell(row=2, column=1).font=font
		F10TB.cell(row=2, column=2).value=ant
		F10TB.cell(row=2, column=2).font=font
		F10TB.cell(row=2, column=2).number_format="mm/dd/yyyy"


		F10TB.cell(row=1, column=11).value="Prepared by:"
		F10TB.cell(row=1, column=11).font=font
		F10TB.cell(row=1, column=12).value=preparedBy1
		F10TB.cell(row=1, column=12).font=font

		F10TB.cell(row=2, column=11).value="Date:"
		F10TB.cell(row=2, column=11).font=font
		F10TB.cell(row=2, column=12).value=datetime.datetime.now()
		F10TB.cell(row=2, column=12).number_format="mm/dd/yyyy"
		F10TB.cell(row=2, column=12).alignment=Alignment(horizontal='left')

		F10TB.cell(row=3, column=11).value="Ref:"
		F10TB.cell(row=3, column=11).font=font
		F10TB.cell(row=3, column=12).value="F10"
		F10TB.cell(row=3, column=12).font=fontRed

		for i in range(1,4):
			F10TB.cell(row=i, column=11).alignment=Alignment(horizontal='right')

		F10TB.cell(row=4, column=2).value="Trial Balance"
		F10TB.cell(row=4, column=2).font=font

		F10TB.cell(row=6, column=1).value="Work done:"
		F10TB.cell(row=6, column=1).font=font


		F10TB.cell(row=8, column=1).value="(to be adjusted for P&L variation; e.g. if YE is different of 31.12)"
		F10TB.cell(row=8, column=1).font=Font(name='Tahoma', size=8, italic=True)




		F10TB.cell(row=14, column=1).value="Class"
		F10TB.cell(row=14, column=2).value="Synt 1"
		F10TB.cell(row=14, column=3).value="Synt 3"
		F10TB.cell(row=14, column=4).value="Synt 4"
		F10TB.cell(row=14, column=5).value="Account"
		F10TB.cell(row=14, column=6).value="Description"
		F10TB.cell(row=14, column=7).value="OB"
		F10TB.cell(row=14, column=8).value="DM"
		F10TB.cell(row=14, column=9).value="CM"
		F10TB.cell(row=14, column=10).value="CB"
		F10TB.cell(row=14, column=11).value="Check"
		F10TB.cell(row=14, column=13).value="Abs CB-OB"
		F10TB.cell(row=14, column=14).value="VAR %"

		F10TB.cell(row=14, column=16).value="OMF Row"
		F10TB.cell(row=14, column=17).value="OMF Description"
		F10TB.cell(row=14, column=18).value="LS"

		F10TB.cell(row=14, column=20).value="Check OB"



		for i in range(1,12):
			F10TB.cell(row=14, column=i).font=font2
			F10TB.cell(row=14, column=i).fill=blueFill
			F10TB.cell(row=14, column=i).border=doubleborder
			F10TB.cell(row=14, column=i).alignment=Alignment(horizontal='left')

		for i in range(13,15):
			F10TB.cell(row=14, column=i).font=font2
			F10TB.cell(row=14, column=i).fill=blueFill
			F10TB.cell(row=14, column=i).border=doubleborder
			F10TB.cell(row=14, column=i).alignment=Alignment(horizontal='left')

		F10TB.cell(row=14, column=16).font=font2
		F10TB.cell(row=14, column=16).fill=blueFill
		F10TB.cell(row=14, column=16).border=doubleborder
		F10TB.cell(row=14, column=16).alignment=Alignment(horizontal='left')

		F10TB.cell(row=14, column=17).font=font2
		F10TB.cell(row=14, column=17).fill=blueFill
		F10TB.cell(row=14, column=17).border=doubleborder
		F10TB.cell(row=14, column=17).alignment=Alignment(horizontal='left')

		F10TB.cell(row=14, column=18).font=font2
		F10TB.cell(row=14, column=18).fill=blueFill
		F10TB.cell(row=14, column=18).border=doubleborder
		F10TB.cell(row=14, column=18).alignment=Alignment(horizontal='left')

		F10TB.cell(row=8, column=6).value="Check BS"
		F10TB.cell(row=9, column=6).value="Revenues"
		F10TB.cell(row=10, column=6).value="Expenses"
		F10TB.cell(row=11, column=6).value="Result"
		F10TB.cell(row=11, column=6).border=doubleborder
		F10TB.cell(row=12, column=6).value="Check"

		for i in range (8,13):
			F10TB.cell(row=i, column=6).font=font
			F10TB.cell(row=i, column=6).alignment=Alignment(horizontal='right')

		F10TB.cell(row=8, column=7).value = '=SUMIF(A:A,"BS",G:G)'
		F10TB.cell(row=11, column=7).border=doubleborder

		for i in range (8,14):
			F10TB.cell(row=i, column=7).number_format='#,##0_);(#,##0)'


		F10TB.cell(row=8, column=10).value = '=SUMIF(A:A,"BS",J:J)'
		F10TB.cell(row=9, column=10).value='=SUMIF(B:B,"7",J:J)'
		F10TB.cell(row=10, column=10).value='=SUMIF(B:B,"6",J:J)'
		F10TB.cell(row=11, column=10).value='=SUMIF(C:C,"121",J:J)'
		F10TB.cell(row=11, column=10).border=doubleborder
		F10TB.cell(row=12, column=10).value="=SUM(J9:J10)-J11"
		F10TB.cell(row=12, column=10).font=fontRed

		F10TB.cell(row=8, column=7).value = '=SUMIF(A:A,"BS",G:G)'
		F10TB.cell(row=9, column=7).value='=SUMIF(B:B,"7",G:G)'
		F10TB.cell(row=10, column=7).value='=SUMIF(B:B,"6",G:G)'
		F10TB.cell(row=11, column=7).value='=SUMIF(C:C,"121",G:G)'
		F10TB.cell(row=11, column=7).border=doubleborder
		F10TB.cell(row=12, column=7).value="=SUM(G9:G10)-G11"
		F10TB.cell(row=12, column=7).font=fontRed



		F10TB.cell(row=13, column=12).value="=SUM(K:K)"
		F10TB.cell(row=13, column=12).font=fontRed

		F10TB.cell(row=13,column=11).value="Total diff:"
		F10TB.cell(row=13,column=11).alignment=Alignment(horizontal='right')

		for i in range (8,13):
			F10TB.cell(row=i, column=10).number_format='#,##0_);(#,##0)'



		for i in range (1,10):
			F10TB.cell(row=i, column=15).number_format='#,##0_);(#,##0)'

		for i in range(14,16):
			F10TB.cell(row=1, column=i).font=font2
			F10TB.cell(row=1, column=i).fill=blueFill
			F10TB.cell(row=1, column=i).border=doubleborder
			F10TB.cell(row=1, column=i).alignment=Alignment(horizontal='left')

		"Importing Data"

		if isChecked4=="":
			try:

				acc=Account+Accountp

				mylist2 = list(dict.fromkeys(acc))
				mylist=[]
				for xxx in range(0,len(mylist2)):
					mylist.append(str(mylist2[xxx]))
				mylist.sort()

				print(mylist)
			except:
				# acc=Account
				mylist2 = list(dict.fromkeys(Account))
				mylist=[]
				for xxx in range(0,len(mylist2)):
					mylist.append(str(mylist2[xxx]))
				mylist.sort()

				print(mylist)
		else:
			mylist2 = list(dict.fromkeys(Account))
			mylist=[]
			for xxx in range(0,len(mylist2)):
				mylist.append(str(mylist2[xxx]))
			mylist.sort()

			print(mylist)


		for i in range(1, len(mylist)+1):
			F10TB.cell(row=i+14, column=5).value=mylist[i-1]

		for i in range  (  1, len(mylist)+1):
			if(mylist[i-1] in Account):
				F10TB.cell(row=i+14, column=6).value='=VLOOKUP(E{0},PBC_CY!D:E,2,0)'.format(i+14)
			else:
				F10TB.cell(row=i+14, column=6).value='=VLOOKUP(E{0},PBC_PY!D:E,2,0)'.format(i+14)

		for i in range (1,len(mylist)+1):
			x=str(mylist[i-1])
			y=str(x[:4])
			F10TB.cell(row=i+14, column=4).value=str(y)

		for i in range (1,len(mylist)+1):
			x=str(mylist[i-1])
			y=x[:3]
			F10TB.cell(row=i+14, column=3).value=str(y)

		for i in range (1,len(mylist)+1):
			F10TB.cell(row=i+14, column=2).value='=Left(E{0},1)'.format(i+14)


		for i in range(1, len(mylist)+1):
				F10TB.cell(row=i+14, column=1).value='=IF(B{0}<"6","BS",IF(AND(B{0}>"5",B{0}<"8"),"PL","Other Account-Off TB"))'.format(i+14)
		 

		"Calculation"

		for i in range(1, len(mylist)+1):
			if(mylist[i-1] in Account):

				if(int(str(mylist[i-1])[:1])<6):
					F10TB.cell(row=i+14,column=7).value='=SUMIF(PBC_CY!D:D,E{0},PBC_CY!F:F)-SUMIF(PBC_CY!D:D,E{0},PBC_cY!G:G)'.format(i+14)
					F10TB.cell(row=i+14,column=7).number_format='#,##0_);(#,##0)'
			else:
				F10TB.cell(row=i+14,column=7).value='=SUMIF(PBC_PY!D:D,E{0},PBC_PY!F:F)'.format(i+14)
				F10TB.cell(row=i+14,column=7).number_format='#,##0_);(#,##0)'
			if(int(str(mylist[i-1])[:1])==6):
				F10TB.cell(row=i+14,column=7).value='=SUMIF(PBC_PY!D:D,E{0},PBC_PY!F:F)'.format(i+14)
				F10TB.cell(row=i+14,column=7).number_format='#,##0_);(#,##0)'
			if(int(str(mylist[i-1])[:1])==7):
				F10TB.cell(row=i+14,column=7).value='=SUMIF(PBC_PY!D:D,E{0},PBC_PY!F:F)'.format(i+14)
				F10TB.cell(row=i+14,column=7).number_format='#,##0_);(#,##0)'

		for i in range(1, len(mylist)+1):
			F10TB.cell(row=i+14, column=8).value='=SUMIF(PBC_CY!D:D,E{0},PBC_CY!H:H)'.format(i+14)
			F10TB.cell(row=i+14,column=8).number_format='#,##0_);(#,##0)'

		for i in range(1, len(mylist)+1):
			F10TB.cell(row=i+14, column=9).value='=SUMIF(PBC_CY!D:D,E{0},PBC_CY!I:I)'.format(i+14)
			F10TB.cell(row=i+14,column=9).number_format='#,##0_);(#,##0)'

		for i in range(1, len(mylist)+1):
			F10TB.cell(row=i+14, column=10).value='=IF(B{0}<"6",SUMIF(PBC_CY!D:D,E{0},PBC_CY!J:J)-SUMIF(PBC_CY!D:D,E{0},PBC_CY!K:K),IF(AND(B{0}="6",C{0}="609",H{0}>0),-H{0},IF(AND(B{0}="6",H{0}<>I{0}),H{0}-I{0},IF(B{0}="6",H{0},IF(AND(B{0}="7",C{0}="709",I{0}<0),I{0},IF(AND(B{0}="7",C{0}="711"),-$U$6,IF(AND(B{0}="7",C{0}="712"),-$U$8,IF(AND(B{0}="7",H{0}<>I{0}),H{0}-I{0},IF(AND(B{0}="7",I{0}>0),-I{0},IF(AND(B{0}="7",I{0}<0),I{0},0))))))))))'.format(i+14)
			F10TB.cell(row=i+14,column=10).number_format='#,##0_);(#,##0)'


		for i in range(1, len(mylist)+1):
			F10TB.cell(row=i+14, column =11).value='=IF(A{0}="BS",G{0}+H{0}-I{0}-J{0},IF(AND(A{0}="PL",H{0}<>I{0}),H{0}-I{0}-J{0},H{0}-I{0}))'.format(i+14)
			F10TB.cell(row=i+14,column=11).number_format='#,##0_);(#,##0)'
			F10TB.cell(row=i+14, column=11).font=fontRed

		for i in range(1, len(mylist)+1):
			F10TB.cell(row=i+14, column=13).value='=IF(A{0}="BS",J{0}-G{0},"")'.format(i+14)
			F10TB.cell(row=i+14,column=13).number_format='#,##0_);(#,##0)'

		F10TB.cell(row=13, column=12).number_format='#,##0_);(#,##0)'

		for i in range(1, len(mylist)+1):
			F10TB.cell(row=i+14, column=14).value='=IF(A{0}="BS",IF(AND(G{0}=0,J{0}=0),0,IF(AND(G{0}=0,J{0}>0),1,IF(AND(G{0}=0,J{0}<0),-1,IF(AND(J{0}=0,G{0}>0),-1,IF(AND(J{0}=0,G{0}<0),1,J{0}/G{0}-1))))),"")'.format(i+14)
			F10TB.cell(row=i+14,column=14).number_format="0.0%"

		for i in range(1, len(mylist)+1):
			x=str(mylist[i-1])
			F10TB.cell(row=i+14,column=16).value='=if('+x[:1]+'<6'+",iferror(vlookup(D{0},'BS Mapping'!A:C,3,0),vlookup(C{0},'BS Mapping'!A:C,3,0)),iferror(vlookup(D{0},'PL Mapping'!A:C,3,0),vlookup(C{0},'PL Mapping'!A:C,3,0)))".format(i+14)

		for i in range(1, len(mylist)+1):
			x=str(mylist[i-1])
			F10TB.cell(row=i+14,column=17).value='=if('+x[:1]+'<6'+",iferror(vlookup(D{0},'BS Mapping'!A:D,4,0),vlookup(C{0},'BS Mapping'!A:D,4,0)),iferror(vlookup(D{0},'PL Mapping'!A:D,4,0),vlookup(C{0},'PL Mapping'!A:D,4,0)))".format(i+14)
		for i in range(1, len(mylist)+1):
			x=str(mylist[i-1])
			F10TB.cell(row=i+14,column=18).value='=if('+x[:1]+'<6'+",iferror(vlookup(D{0},'BS Mapping'!A:E,5,0),vlookup(C{0},'BS Mapping'!A:E,5,0)),iferror(vlookup(D{0},'PL Mapping'!A:E,5,0),vlookup(C{0},'PL Mapping'!A:E,5,0)))".format(i+14)
		for i in range(1, len(mylist)+1):
			F10TB.cell(row=i+14,column=20).value="=G{0}-SUMIF(PBC_PY!D:D,E{0},PBC_PY!F:F)".format(i+14)
			F10TB.cell(row=i+14,column=20).number_format='#,##0_);(#,##0)'
			F10TB.cell(row=i+14,column=20).font=fontRed
			# F10TB.cell(row=i+14,column=16).value='=if('+x[:1]+'<6,iferror(vlookup('+str(x[0:4])+",'BS Mapping std'!A:E,5,0),vlookup("+str(x[0:3])+",'BS Mapping std'!A:E,5,0)),iferror(vlookup("+str(x[0:4])+",'PL mapping Std'!A:E,5,0),vlookup("+str(x[0:3])+",'PL mapping Std'!A:E,5,0))"
		"Closing 711"

		F10TB.cell(row=1, column=14).value="Closing 711"
		F10TB.cell(row=1, column=14).font=font2
		F10TB.cell(row=1, column=14).alignment=Alignment(horizontal='right')

		F10TB.cell(row=1, column=14).value="Acc."
		F10TB.cell(row=1, column=15).value="OB"
		F10TB.cell(row=1, column=16).value="CB"
		F10TB.cell(row=1, column=17).value="VAR"

		F10TB.cell(row=2, column=14).value="331"
		F10TB.cell(row=3, column=14).value="341"
		F10TB.cell(row=4, column=14).value="345"
		F10TB.cell(row=5, column=14).value="348"
		F10TB.cell(row=6, column=14).value="Total:"
		F10TB.cell(row=6, column=14).font=font
		F10TB.cell(row=6, column=14).alignment=Alignment(horizontal='right' )

		F10TB.cell(row=8, column=14).value="332" 
		F10TB.cell(row=8, column=14).font=font

		for i in range(14,18):
			F10TB.cell(row=1, column=i).font=font2
			F10TB.cell(row=1, column=i).fill=blueFill
			F10TB.cell(row=1, column=i).border=doubleborder
			F10TB.cell(row=1, column=i).alignment=Alignment(horizontal='left')

		for i in range (2,9):
			F10TB.cell(row=i, column=14).font=font
			F10TB.cell(row=i, column=14).alignment=Alignment(horizontal='right')

		for i in range (14,18):
			F10TB.cell(row=5, column=i).border=doubleborder

		for i in range (2, 6):
			F10TB.cell(row=i, column=15).value='=SUMIF(C:C,N{0},G:G)'.format(i)
			F10TB.cell(row=i, column=16).value='=SUMIF(C:C,N{0},J:J)'.format(i)
			F10TB.cell(row=i, column=17).value='=P{0}-O{0}'.format(i)

		F10TB.cell(row=6, column=15).value='=SUM(O2:O5)'
		F10TB.cell(row=6, column=16).value='=SUM(P2:P5)'
		F10TB.cell(row=6, column=17).value='=P6-O6'
		F10TB.cell(row=6, column=17).font=font2
		F10TB.cell(row=6, column=17).fill=blueFill

		for i in range (15,18):
			F10TB.cell(row=6, column=i).font=font2
		F10TB.cell(row=14,column=20).fill=blueFill
		F10TB.cell(row=14,column=20).font=font2

		for i in range(2,14):
			for j in range(15,18):  
				F10TB.cell(row=i, column=j).number_format='#,##0_);(#,##0)'

		F10TB.cell(row=7, column=14).value="Closing 712"
		F10TB.cell(row=7, column=14).font=font2
		F10TB.cell(row=7, column=14).alignment=Alignment(horizontal='right')

		F10TB.cell(row=8, column=15).value='=SUMIF(C:C,N8,G:G)'
		F10TB.cell(row=8, column=16).value='=SUMIF(C:C,N8,J:J)'
		F10TB.cell(row=8, column=17).value='=P8-O8'
		F10TB.cell(row=8, column=17).fill=blueFill
		F10TB.cell(row=8, column=17).font=font2

		for i in range (6,10):
			F10TB.cell(row=11, column=i).border=doubleborder

		F10TB.auto_filter.ref = 'A14:R14'

		c = F10TB['B15']
		F10TB.freeze_panes = c


		x
		"Adjust Column Width"

		for col in F10TB.columns:
			max_length = 0
			for cell in col:
				if cell.coordinate in F10TB.merged_cells:
					continue
				try:
					if len(str(cell.value)) > max_length:
						max_length = len(cell.value)
				except:
					pass
			adjusted_width=(max_length-5)


		listanoua=['F','G','H','I','J','K','M','N','O','L']
		for column in ascii_uppercase:
			for i in listanoua:
				if (column==i):
					F10TB.column_dimensions[column].width =15

		listanoua2=['A']
		for column in ascii_uppercase:
			for i in listanoua2:
				if (column==i):
					F10TB.column_dimensions[column].width = 10


		
		file_path=os.path.join(folderpath, "F100 Trial Balance.xlsx")
		myorder=[4,0,1,2,3]
		output._sheets =[output._sheets[i] for i in myorder]
		output.save(folderpath+"\Trial Balance.xlsx")
		return send_from_directory(folderpath,"Trial Balance.xlsx",as_attachment=True)

		# print(text)
	   

@app.route('/LSFS/Instructions', methods=['GET'])
def downloadLSFS():
		filepath = "/home/auditappnexia/output/LSFS"
 
		return send_from_directory(filepath,"Instructions - Leadsheets.docx", as_attachment=True)

@app.route('/LSFS/GTdjA2adxpt6S8HHJtDhJ2dvPWS')
def LSFS():
	return render_template('TBLS.html')
@app.route('/LSFS/GTdjA2adxpt6S8HHJtDhJ2dvPWS', methods=['POST', 'GET'])
def LSFS_process():

	namec = request.form['client']
	ant= datetime.datetime.strptime(
					 request.form['yearEnd'],
					 '%Y-%m-%d')
	threshol = request.form['threshold']
	preparedBy1 = request.form['preparedBy']
	datePrepared1 = datetime.datetime.strptime(
		request.form['preparedDate'],
		'%Y-%m-%d')
	# isChecked2=request.form.get("gls")
	# isChecked3=request.form.get("gfs")
	# isChecked4=request.form.get("forml")
	# isChecked5=request.form.get("forms")
	isChecked6=request.form.get("mctb")
	
	if isChecked6=="": #daca e bifat
		isChecked6=1
		print("e bine")

	else:
	    isChecked6=0

	
	if(isChecked6==1):

		folderpath="home/auditappnexia/output/LSFS" #aici era inainte fara LSFS
		os.mkdir(folderpath+"LSFS "+namec)
		

		def make_archive(source, destination):
			base = os.path.basename(destination)
			name = base.split('.')[0]
			format = base.split('.')[1]
			archive_from = os.path.dirname(source)
			archive_to = os.path.basename(source.strip(os.sep))
			shutil.make_archive(name, format, archive_from, archive_to)
			shutil.move('%s.%s'%(name,format), destination)
		# yearEnd = str(request.form['yearEnd'])
		# processed_text = client.upper()
		# fisier=request.files.get('monthlyTB')
		if request.method == 'POST':
			workingsblue2= Font(bold=True, italic=True, name='Tahoma', size=8,color='FFFFFF')
			lbluefill = PatternFill(start_color='7030A0',
								end_color='7030A0',
								fill_type='solid')
			grifill=PatternFill(start_color='c4d79b',end_color='c4d79b',fill_type='solid')
			yellow=PatternFill(start_color='ffff00',end_color='ffff00',fill_type='solid')
			blueFill = PatternFill(start_color='00AEAC',
								end_color='00AEAC',
								fill_type='solid')
			doubleborder = Border(bottom=Side(style='double'))
			solidborder = Border(bottom=Side(style='thick'))
			solidborderstanga = Border(left=Side(style='thin'))
			rightborder = Border(right=Side(style='thin'))
			rightdouble = Border (right=Side(style='thin'), bottom=Side(style='double'))
			rightmedium = Border (right=Side(style='thin'), bottom=Side(style='medium'))
			solidborderdreapta = Border(right=Side(style='thin'))
			solidbordersus = Border(top=Side(style='thin'))
			fontitalic = Font(name='Tahoma', size=8, bold=True, italic=True)
			font = Font(name='Tahoma', size=8, bold=True)
			font1 = Font(name='Tahoma', size=8)
			font2 = Font(name='Tahoma', size=10, bold=True)
			fontRed = Font(name='Tahoma', size=10, bold=True, color= 'FF0000')
			fontRedDiff=Font(name="Tahoma", color='FF0000', size=11, )
			fontGT = Font (name='GT Logo', size=8)
			workingsblue = Font(color='2F75B5', bold=True, name='Tahoma', size=8 )
			headers= Font(bold=True, italic=True, name='Tahoma', size=8,color='FFFFFF') 
			headersblue = PatternFill(start_color='7030A0',
							end_color='7030A0',
							fill_type='solid')
			headerspurple= PatternFill(start_color='65CDCC',
								end_color='65CDCC',
								fill_type='solid')
			total=PatternFill(start_color='DDD9C4',
							end_color='DDD9C4',
							fill_type='solid')
			greenbolditalic= Font(bold=True, italic=True,  color='C0504D', name='Tahoma', size=8)
			greenbolditalic= Font(bold=True, italic=True,  color='00af50')
			fontalb = Font(italic=True, color="bfbfbf", size=8, name='Tahoma')
			
			# if isChecked2=="":

				
			blueFill = PatternFill(start_color='4F2D7F',
								end_color='00AEAC',
								fill_type='solid')

			greenfill=PatternFill(start_color='92D050',end_color='92D050',fill_type='solid')
			gtfill=PatternFill(start_color='00AEAC',end_color='00AEAC',fill_type='solid')
			doubleborder = Border(bottom=Side(style='double'))
			font = Font(name='Tahoma', size=8, bold=True)
			normal = Font(name='Tahoma', size=8)
			font2 = Font(name='Tahoma', size=8,bold=True,color='ffffff')
			fontRed = Font(name='Tahoma', size=8, bold=True, color= 'FF0000')
			doubleborder = Border(bottom=Side(style='double'))

			triald=request.files["leadsheets"]
			
			TBls = openpyxl.load_workbook(triald,data_only=True)
			TBls1 = TBls.active
			TBls1.title="Trial Balance" 

			Sumlead=TBls.create_sheet("LS Summary")

			Sumlead.cell(row=1,column=1).value="Client:"
			Sumlead.cell(row=1,column=1).font=font
			Sumlead.cell(row=1,column=2).value=namec
			Sumlead.cell(row=1,column=2).font=font
			Sumlead.cell(row=2,column=1).value="Period end:"
			Sumlead.cell(row=2,column=1).font=font
			Sumlead.cell(row=2,column=2).value=ant
			Sumlead.cell(row=2,column=2).number_format='mm/dd/yyyy'
			Sumlead.cell(row=2,column=2).font=font
			Sumlead.cell(row=1,column=7).value="Prepared by:"
			Sumlead.cell(row=1,column=7).font=font
			Sumlead.cell(row=2,column=7).value="Date:"
			Sumlead.cell(row=2,column=7).font=font
			Sumlead.cell(row=4,column=1).value="LS Summary"
			Sumlead.cell(row=4,column=1).font=font
			Sumlead.cell(row=4,column=4).value="OB"
			Sumlead.cell(row=4,column=4).border=doubleborder
			Sumlead.cell(row=4,column=4).fill=gtfill
			Sumlead.cell(row=4,column=4).font=font2
			Sumlead.cell(row=4,column=5).value="CB"
			Sumlead.cell(row=4,column=5).border=doubleborder
			Sumlead.cell(row=4,column=5).fill=gtfill
			Sumlead.cell(row=4,column=5).font=font2
			Sumlead.cell(row=5,column=3).value="Check BS"
			Sumlead.cell(row=5,column=3).font=font
			Sumlead.cell(row=6,column=3).value="Result"
			Sumlead.cell(row=6,column=3).font=font
			Sumlead.cell(row=7,column=3).value="Acc 121"
			Sumlead.cell(row=7,column=4).border=doubleborder
			Sumlead.cell(row=7,column=5).border=doubleborder
			Sumlead.cell(row=7,column=3).font=font
			Sumlead.cell(row=8,column=3).value="Check Result"
			Sumlead.cell(row=8,column=3).font=font

			Sumlead.cell(row=13,column=3).value="Note that every description from below is an hyperlink to their respective sheet"
			Sumlead.cell(row=13,column=3).font=fontRed
			Sumlead.cell(row=10,column=4).value="Threshold value:"
			Sumlead.cell(row=10,column=4).font=font
			Sumlead.cell(row=10,column=5).value=threshol
			Sumlead.cell(row=10,column=5).font=fontRed
			Sumlead.cell(row=10,column=5).number_format='#,##0_);(#,##0)'
			Sumlead.cell(row=14,column=2).value="Type"
			Sumlead.cell(row=14,column=3).value="LS Description"
			Sumlead.cell(row=14,column=4).value="Total as per LS PY"
			Sumlead.cell(row=14,column=5).value="Total as per LS CY"
			Sumlead.cell(row=14,column=6).value="Var."
			Sumlead.cell(row=14,column=6).fill=greenfill
			Sumlead.cell(row=14,column=6).font=fontRed
			Sumlead.cell(row=14,column=7).value="Var. %"
			Sumlead.cell(row=14,column=7).fill=greenfill
			Sumlead.cell(row=14,column=7).font=fontRed
			Sumlead.cell(row=14,column=8).value="Threshold flag"
			Sumlead.cell(row=14,column=8).fill=greenfill
			Sumlead.cell(row=14,column=8).font=fontRed
			Sumlead.cell(row=14,column=9).value="Assigned to:"
			Sumlead.cell(row=14,column=9).fill=gtfill
			Sumlead.cell(row=14,column=9).font=font2


			for o in range(2,9):
				Sumlead.cell(row=14,column=o).border=doubleborder
			for o in range(2,6):
				Sumlead.cell(row=14,column=o).fill=gtfill
				Sumlead.cell(row=14,column=o).font=font2



			"Iterate from imported PBC's:"


			'Iterate from CY TB'

			for row in TBls1.iter_rows():
					for cell in row:
						if cell.value=="Account":
							Acc=cell.column
							acr=cell.row

			for row in TBls1.iter_rows():
				for cell in row:
					if cell.value=="Description":
						desc=cell.column

			for row in TBls1.iter_rows():
				for cell in row:
					if cell.value=="OB":
						aob=cell.column

			for row in TBls1.iter_rows():

				for cell in row:
					if cell.value=="DM":
						adm=cell.column
					
			for row in TBls1.iter_rows():
				for cell in row:
					if cell.value=="CM":
						acm=cell.column

			for row in TBls1.iter_rows():
				for cell in row:
					if cell.value=="CB":
						acb=cell.column

			for row in TBls1.iter_rows():
				for cell in row:
					if cell.value=="OMF Row":
						omfr=cell.column


			for row in TBls1.iter_rows():
				for cell in row:
					if cell.value=="OMF Description":
						omfd=cell.column

			for row in TBls1.iter_rows():
				for cell in row:
					if cell.value=="LS":
						omfls=cell.column

			for row in TBls1.iter_rows():
				for cell in row:
					if cell.value=="Type":
						omftype=cell.column
			
			luntb=len(TBls1[Acc])

			
			Account=[b.value for b in TBls1[Acc][acr:luntb+1]]
			Description=[b.value for b in TBls1[desc][acr:luntb+1]]
			OB=[b.value for b in TBls1[aob][acr:luntb+1]]
			DM=[b.value for b in TBls1[adm][acr:luntb+1]]
			CM=[b.value for b in TBls1[acm][acr:luntb+1]]
			CB=[b.value for b in TBls1[acb][acr:luntb+1]]
			
			if(isChecked6==1):				
				print("is checked")
				# OMFrow=[b.value for b in TBls1[omfr][acr:luntb+1]]
				# OMFdescrip=[b.value for b in TBls1[omfd][acr:luntb+1]]
				# OMFType=[b.value for b in TBls1[omftype][acr:luntb+1]]
			else:
				OMFrow=[b.value for b in TBls1[omfr][acr:luntb+1]]
				OMFdescrip=[b.value for b in TBls1[omfd][acr:luntb+1]]
				OMFType=[b.value for b in TBls1[omftype][acr:luntb+1]]
			Lead=[b.value for b in TBls1[omfls][acr:luntb+1]]
			# Type=[b.value for b in TBls1[omftype][acr:luntb+1]]
			Sumlead.sheet_view.showGridLines = False
			maplead=list(set(Lead))
			# maptype=list(set(Type))

			for i in range(0,len(maplead)):
				try:
					work=TBls.create_sheet(str(maplead[i][:31]))
				except:
					work=TBls.create_sheet("NA")
				work.sheet_view.showGridLines = False
				work.cell(row=1,column=1).value="Client:"
				work.cell(row=1,column=1).font=font
				work.cell(row=1,column=2).value=namec
				work.cell(row=1,column=2).font=font
				work.cell(row=2,column=1).value="Period end:"
				work.cell(row=2,column=1).font=font
				work.cell(row=2,column=2).value=ant
				work.cell(row=2,column=2).number_format='mm/dd/yyyy'
				work.cell(row=2,column=2).font=font
				work.cell(row=1,column=11).value="Prepared by:"
				work.cell(row = 1, column = 12).value = preparedBy1
				work.cell(row = 2, column = 12).value = datePrepared1
				work.cell(row = 2, column = 12).number_format = 'mm/dd/yyyy'
				work.cell(row = 1, column = 12).font = font
				work.cell(row = 2, column = 12).font = font
				work.cell(row=1,column=11).font=font
				work.cell(row=2,column=11).value="Date:"
				work.cell(row=2,column=11).font=font
				work.cell(row=4,column=1).value=str(maplead[i])
				work.cell(row=4,column=1).font=font
				work.cell(row=8,column=2).value="Account"
				work.cell(row=8,column=3).value="Description"
				work.cell(row=8,column=4).value="OB"
				work.cell(row=8,column=5).value="DM"
				work.cell(row=8,column=6).value="CM"
				work.cell(row=8,column=7).value="CB"
				work.cell(row=8,column=8).value="OMF row"
				work.cell(row=8,column=9).value="Var."
				work.cell(row=8,column=10).value="Var. %"
				work.cell(row=8,column=11).value="Ref Wp"
				for m in range(2,9):
					work.cell(row=8,column=m).fill=gtfill
					work.cell(row=8,column=m).font=font2
				for m in range(9,12):
					work.cell(row=8,column=m).fill=greenfill
					work.cell(row=8,column=m).font=fontRed
				count=0
				for j in range(0,len(Lead)):
					if(Lead[j]==maplead[i]):
						count=count+1
						work.cell(row=8+count,column=2).value=Account[j]
						work.cell(row=8+count,column=2).font=normal
						work.cell(row=8+count,column=3).value=Description[j]
						work.cell(row=8+count,column=3).font=normal
						work.cell(row=8+count,column=4).value=OB[j]
						work.cell(row=8+count,column=4).number_format='#,##0_);(#,##0)'

						work.cell(row=8+count,column=4).font=normal
						work.cell(row=8+count,column=5).value=DM[j]
						work.cell(row=8+count,column=5).number_format='#,##0_);(#,##0)'
						work.cell(row=8+count,column=5).font=normal
						work.cell(row=8+count,column=6).value=CM[j]
						work.cell(row=8+count,column=6).number_format='#,##0_);(#,##0)'
						work.cell(row=8+count,column=6).font=normal
						work.cell(row=8+count,column=7).value=CB[j]
						work.cell(row=8+count,column=7).number_format='#,##0_);(#,##0)'
						work.cell(row=8+count,column=7).font=normal
						if(isChecked6==1):
							print("nada")
							# work.cell(row=8+count,column=8).value=OMFrow[j]
							# work.cell(row=8+count,column=8).font=normal
						else:
							work.cell(row=8+count,column=8).value=OMFrow[j]
							work.cell(row=8+count,column=8).font=normal
						work.cell(row=8+count,column=9).value="=G{0}-D{0}".format(count+8)
						work.cell(row=8+count,column=9).number_format='#,##0_);(#,##0)'
						work.cell(row=8+count,column=10).value="=IFERROR(I{0}/D{0},0)".format(count+8)
						work.cell(row=8+count,column=10).number_format='0.00%'

				work.cell(row=8+count+1,column=3).value="Total"
				work.cell(row=8+count+1,column=3).font=font
				work.cell(row=8+count+1,column=4).value="=sum(D9:D"+str(8+count)+")"
				work.cell(row=8+count+1,column=4).number_format='#,##0_);(#,##0)'
				work.cell(row=8+count+1,column=4).font=font
				work.cell(row=8+count+1,column=5).value="=sum(E9:E"+str(8+count)+")"
				work.cell(row=8+count+1,column=5).font=font
				work.cell(row=8+count+1,column=5).number_format='#,##0_);(#,##0)'
				work.cell(row=8+count+1,column=6).value="=sum(F9:F"+str(8+count)+")"
				work.cell(row=8+count+1,column=6).font=font
				work.cell(row=8+count+1,column=6).number_format='#,##0_);(#,##0)'
				work.cell(row=8+count+1,column=7).value="=sum(G9:G"+str(8+count)+")"
				work.cell(row=8+count+1,column=7).font=font
				work.cell(row=8+count+1,column=7).number_format='#,##0_);(#,##0)'
				
				work.column_dimensions['A'].width=10
				work.column_dimensions['B'].width=20
				work.column_dimensions['C'].width=15
				work.column_dimensions['D'].width=15
				work.column_dimensions['E'].width=15
				work.column_dimensions['F'].width=15
				work.column_dimensions['G'].width=15
				work.column_dimensions['H'].width=15
				work.column_dimensions['I'].width=15
				work.column_dimensions['J'].width=15 


				for p in range(2,12):
					work.cell(row=8,column=p).border=doubleborder
					work.cell(row=8+count,column=p).border=doubleborder

				for k in range(9,9+count):
					work.cell(row=k,column=9).font=fontRed
					work.cell(row=k,column=10).font=fontRed
				link="#'"+str(str(maplead[i])[:31])+"'!A4"
				try:
					Sumlead.cell(row=15+i,column=3).value=str(maplead[i][:31])
					Sumlead.cell(row=15+i,column=3).font=normal
					Sumlead.cell(row=15+i,column=3).hyperlink=link
					Sumlead.cell(row=15+i,column=2).value="=VLOOKUP(C{0},'Trial Balance'!G:H,2,0)".format(i+15)
					Sumlead.cell(row=15+i,column=4).value="='"+str(maplead[i][:31])+"'!D"+str(count+9)
					Sumlead.cell(row=15+i,column=4).font=normal
					Sumlead.cell(row=15+i,column=4).number_format='#,##0_);(#,##0)'
					Sumlead.cell(row=15+i,column=5).value="='"+str(maplead[i][:31])+"'!G"+str(count+9)
					Sumlead.cell(row=15+i,column=5).font=normal
					Sumlead.cell(row=15+i,column=5).number_format='#,##0_);(#,##0)'
					Sumlead.cell(row=15+i,column=6).value="=E{0}-D{0}".format(i+15)
					Sumlead.cell(row=15+i,column=6).font=fontRed
					Sumlead.cell(row=15+i,column=6).number_format='#,##0_);(#,##0)'
					Sumlead.cell(row=15+i,column=7).value="=IFErROR(F{0}/D{0},0)".format(i+15)
					Sumlead.cell(row=15+i,column=7).font=fontRed
					Sumlead.cell(row=15+i,column=7).number_format='0.00%'
					Sumlead.cell(row=15+i,column=8).value='=IF(abs($E$10)<E{0},"Above threshold","Below Threshold")'.format(i+15)
					Sumlead.cell(row=15+i,column=8).font=normal
				except:
					Sumlead.cell(row=15+i,column=3).value="NA"  
					Sumlead.cell(row=15+i,column=4).value='=NA!D'+str(count+9)
					Sumlead.cell(row=15+i,column=5).value='=NA!G'+str(count+9)
				Sumlead.cell(row=5,column=4).value='=SUMIF(B:B,"BS",D:D)'
				Sumlead.cell(row=5,column=4).font=normal
				Sumlead.cell(row=5,column=4).number_format='#,##0_);(#,##0)'
				Sumlead.cell(row=6,column=4).value='=SUMIF(B:B,"PL",D:D)'
				Sumlead.cell(row=6,column=4).font=normal
				Sumlead.cell(row=6,column=4).number_format='#,##0_);(#,##0)'
				Sumlead.cell(row=7,column=4).value="=VLOOKUP(121,'Trial Balance'!a:f,3,0)".format(i+15)
				Sumlead.cell(row=7,column=4).font=normal
				Sumlead.cell(row=7,column=4).number_format='#,##0_);(#,##0)'
				Sumlead.cell(row=8,column=4).value='=D7-D6'
				Sumlead.cell(row=8,column=4).font=fontRed
				Sumlead.cell(row=8,column=4).number_format='#,##0_);(#,##0)'
				Sumlead.cell(row=5,column=5).value='=SUMIF(B:B,"BS",E:E)'
				Sumlead.cell(row=5,column=5).font=normal
				Sumlead.cell(row=5,column=5).number_format='#,##0_);(#,##0)'
				Sumlead.cell(row=6,column=5).value='=SUMIF(B:B,"PL",E:E)'
				Sumlead.cell(row=6,column=5).font=normal
				Sumlead.cell(row=6,column=5).number_format='#,##0_);(#,##0)'
				Sumlead.cell(row=7,column=5).value="=VLOOKUP(121,'Trial Balance'!a:f,6,0)".format(i+15)
				Sumlead.cell(row=7,column=5).font=normal
				Sumlead.cell(row=7,column=5).number_format='#,##0_);(#,##0)'
				Sumlead.cell(row=8,column=5).value='=E7-E6'
				Sumlead.cell(row=8,column=5).font=fontRed
				Sumlead.cell(row=8,column=5).number_format='#,##0_);(#,##0)'

			# for j in range(0,len(maptype)):
			# 	if(maptype[j]=="Banca"):
			# 		Sumlead.cell(row=15+j,column=2).value=str(maptype[j][:31])
			# 	else:
			# 		Sumlead.cell(row=15+j,column=2).value="Na"
			
			Sumlead.column_dimensions['A'].width=10
			Sumlead.column_dimensions['B'].width=30
			Sumlead.column_dimensions['C'].width=15
			Sumlead.column_dimensions['D'].width=15
			Sumlead.column_dimensions['E'].width=15
			Sumlead.column_dimensions['F'].width=15
			Sumlead.column_dimensions['G'].width=15
			Sumlead.column_dimensions['H'].width=15
			Sumlead.column_dimensions['I'].width=15

			if(isChecked6==1):				
				print("is checked")
				# cy=TBls["PBC_CY"]
				# py=TBls["PBC_PY"]
				# bs=TBls['BS Mapping']
				# pl=TBls['PL Mapping']
			else:
				cy=TBls["PBC_CY"]
				py=TBls["PBC_PY"]
				bs=TBls['BS Mapping']
				pl=TBls['PL Mapping']
			
			if(isChecked6==1):				
				print("is checked")
				# cy.sheet_state = 'hidden'
				# py.sheet_state = 'hidden'
				# bs.sheet_state = 'hidden'
				# pl.sheet_state = 'hidden'
			else:
				cy.sheet_state = 'hidden'
				py.sheet_state = 'hidden'
				bs.sheet_state = 'hidden'
				pl.sheet_state = 'hidden'
			Sumlead.auto_filter.ref = 'B14:H14'

			c = Sumlead['B15']
			Sumlead.freeze_panes = c

			

			file_pathLS=os.path.join(folderpath+"LSFS "+namec, "Leadsheets.xlsx")

			TBls.save(file_pathLS)
			
	else:
		
		folderpath="home/auditappnexia/output/LSFS" #aici era inainte fara LSFS
		os.mkdir(folderpath+"LSFS "+namec)
		

		def make_archive(source, destination):
			base = os.path.basename(destination)
			name = base.split('.')[0]
			format = base.split('.')[1]
			archive_from = os.path.dirname(source)
			archive_to = os.path.basename(source.strip(os.sep))
			shutil.make_archive(name, format, archive_from, archive_to)
			shutil.move('%s.%s'%(name,format), destination)
		# yearEnd = str(request.form['yearEnd'])
		# processed_text = client.upper()
		# fisier=request.files.get('monthlyTB')
		if request.method == 'POST':
			workingsblue2= Font(bold=True, italic=True, name='Tahoma', size=8,color='FFFFFF')
			lbluefill = PatternFill(start_color='7030A0',
								end_color='7030A0',
								fill_type='solid')
			grifill=PatternFill(start_color='c4d79b',end_color='c4d79b',fill_type='solid')
			yellow=PatternFill(start_color='ffff00',end_color='ffff00',fill_type='solid')
			blueFill = PatternFill(start_color='00AEAC',
								end_color='00AEAC',
								fill_type='solid')
			doubleborder = Border(bottom=Side(style='double'))
			solidborder = Border(bottom=Side(style='thick'))
			solidborderstanga = Border(left=Side(style='thin'))
			rightborder = Border(right=Side(style='thin'))
			rightdouble = Border (right=Side(style='thin'), bottom=Side(style='double'))
			rightmedium = Border (right=Side(style='thin'), bottom=Side(style='medium'))
			solidborderdreapta = Border(right=Side(style='thin'))
			solidbordersus = Border(top=Side(style='thin'))
			fontitalic = Font(name='Tahoma', size=8, bold=True, italic=True)
			font = Font(name='Tahoma', size=8, bold=True)
			font1 = Font(name='Tahoma', size=8)
			font2 = Font(name='Tahoma', size=10, bold=True)
			fontRed = Font(name='Tahoma', size=10, bold=True, color= 'FF0000')
			fontRedDiff=Font(name="Tahoma", color='FF0000', size=11, )
			fontGT = Font (name='GT Logo', size=8)
			workingsblue = Font(color='2F75B5', bold=True, name='Tahoma', size=8 )
			headers= Font(bold=True, italic=True, name='Tahoma', size=8,color='FFFFFF') 
			headersblue = PatternFill(start_color='7030A0',
							end_color='7030A0',
							fill_type='solid')
			headerspurple= PatternFill(start_color='65CDCC',
								end_color='65CDCC',
								fill_type='solid')
			total=PatternFill(start_color='DDD9C4',
							end_color='DDD9C4',
							fill_type='solid')
			greenbolditalic= Font(bold=True, italic=True,  color='C0504D', name='Tahoma', size=8)
			greenbolditalic= Font(bold=True, italic=True,  color='00af50')
			fontalb = Font(italic=True, color="bfbfbf", size=8, name='Tahoma')
			
			# if isChecked2=="":

				
			blueFill = PatternFill(start_color='4F2D7F',
								end_color='00AEAC',
								fill_type='solid')

			greenfill=PatternFill(start_color='92D050',end_color='92D050',fill_type='solid')
			gtfill=PatternFill(start_color='00AEAC',end_color='00AEAC',fill_type='solid')
			doubleborder = Border(bottom=Side(style='double'))
			font = Font(name='Tahoma', size=8, bold=True)
			normal = Font(name='Tahoma', size=8)
			font2 = Font(name='Tahoma', size=8,bold=True,color='ffffff')
			fontRed = Font(name='Tahoma', size=8, bold=True, color= 'FF0000')
			doubleborder = Border(bottom=Side(style='double'))

			triald=request.files["leadsheets"]
			
			TBls = openpyxl.load_workbook(triald,data_only=True)
			TBls1 = TBls.active

			Sumlead=TBls.create_sheet("LS Summary")

			Sumlead.cell(row=1,column=1).value="Client:"
			Sumlead.cell(row=1,column=1).font=font
			Sumlead.cell(row=1,column=2).value=namec
			Sumlead.cell(row=1,column=2).font=font
			Sumlead.cell(row=2,column=1).value="Period end:"
			Sumlead.cell(row=2,column=1).font=font
			Sumlead.cell(row=2,column=2).value=ant
			Sumlead.cell(row=2,column=2).number_format='mm/dd/yyyy'
			Sumlead.cell(row=2,column=2).font=font
			Sumlead.cell(row=1,column=7).value="Prepared by:"
			Sumlead.cell(row = 1, column = 8).value = preparedBy1
			Sumlead.cell(row = 2, column = 8).value = datePrepared1
			Sumlead.cell(row = 2, column = 8).number_format = 'mm/dd/yyyy'
			Sumlead.cell(row = 1, column = 8).font = font
			Sumlead.cell(row = 2, column = 8).font = font
			Sumlead.cell(row=1,column=7).font=font
			Sumlead.cell(row=2,column=7).value="Date:"
			Sumlead.cell(row=2,column=7).font=font
			Sumlead.cell(row=4,column=1).value="LS Summary"
			Sumlead.cell(row=4,column=1).font=font
			Sumlead.cell(row=4,column=4).value="OB"
			Sumlead.cell(row=4,column=4).border=doubleborder
			Sumlead.cell(row=4,column=4).fill=gtfill
			Sumlead.cell(row=4,column=4).font=font2
			Sumlead.cell(row=4,column=5).value="CB"
			Sumlead.cell(row=4,column=5).border=doubleborder
			Sumlead.cell(row=4,column=5).fill=gtfill
			Sumlead.cell(row=4,column=5).font=font2
			Sumlead.cell(row=5,column=3).value="Check BS"
			Sumlead.cell(row=5,column=3).font=font
			Sumlead.cell(row=6,column=3).value="Result"
			Sumlead.cell(row=6,column=3).font=font
			Sumlead.cell(row=7,column=3).value="Acc 121"
			Sumlead.cell(row=7,column=4).border=doubleborder
			Sumlead.cell(row=7,column=5).border=doubleborder
			Sumlead.cell(row=7,column=3).font=font
			Sumlead.cell(row=8,column=3).value="Check Result"
			Sumlead.cell(row=8,column=3).font=font

			Sumlead.cell(row=13,column=3).value="Note that every description from below is an hyperlink to their respective sheet"
			Sumlead.cell(row=13,column=3).font=fontRed
			Sumlead.cell(row=10,column=4).value="Threshold value:"
			Sumlead.cell(row=10,column=4).font=font
			Sumlead.cell(row=10,column=5).value=threshol
			Sumlead.cell(row=10,column=5).font=fontRed
			Sumlead.cell(row=10,column=5).number_format='#,##0_);(#,##0)'
			Sumlead.cell(row=14,column=2).value="Type"
			Sumlead.cell(row=14,column=3).value="LS Description"
			Sumlead.cell(row=14,column=4).value="Total as per LS PY"
			Sumlead.cell(row=14,column=5).value="Total as per LS CY"
			Sumlead.cell(row=14,column=6).value="Var."
			Sumlead.cell(row=14,column=6).fill=greenfill
			Sumlead.cell(row=14,column=6).font=fontRed
			Sumlead.cell(row=14,column=7).value="Var. %"
			Sumlead.cell(row=14,column=7).fill=greenfill
			Sumlead.cell(row=14,column=7).font=fontRed
			Sumlead.cell(row=14,column=8).value="Threshold flag"
			Sumlead.cell(row=14,column=8).fill=greenfill
			Sumlead.cell(row=14,column=8).font=fontRed
			Sumlead.cell(row=14,column=9).value="Assigned to:"
			Sumlead.cell(row=14,column=9).fill=gtfill
			Sumlead.cell(row=14,column=9).font=font2


			for o in range(2,9):
				Sumlead.cell(row=14,column=o).border=doubleborder
			for o in range(2,6):
				Sumlead.cell(row=14,column=o).fill=gtfill
				Sumlead.cell(row=14,column=o).font=font2



			"Iterate from imported PBC's:"


			'Iterate from CY TB'

			for row in TBls1.iter_rows():
					for cell in row:
						if cell.value=="Account":
							Acc=cell.column
							acr=cell.row

			for row in TBls1.iter_rows():
				for cell in row:
					if cell.value=="Description":
						desc=cell.column

			for row in TBls1.iter_rows():
				for cell in row:
					if cell.value=="OB":
						aob=cell.column

			for row in TBls1.iter_rows():

				for cell in row:
					if cell.value=="DM":
						adm=cell.column
					
			for row in TBls1.iter_rows():
				for cell in row:
					if cell.value=="CM":
						acm=cell.column

			for row in TBls1.iter_rows():
				for cell in row:
					if cell.value=="CB":
						acb=cell.column

			for row in TBls1.iter_rows():
				for cell in row:
					if cell.value=="OMF Row":
						omfr=cell.column


			for row in TBls1.iter_rows():
				for cell in row:
					if cell.value=="OMF Description":
						omfd=cell.column

			for row in TBls1.iter_rows():
				for cell in row:
					if cell.value=="LS":
						omfls=cell.column

			for row in TBls1.iter_rows():
				for cell in row:
					if cell.value=="Type":
						omftype=cell.column
			
			luntb=len(TBls1[Acc])

			
			Account=[b.value for b in TBls1[Acc][acr:luntb+1]]
			Description=[b.value for b in TBls1[desc][acr:luntb+1]]
			OB=[b.value for b in TBls1[aob][acr:luntb+1]]
			DM=[b.value for b in TBls1[adm][acr:luntb+1]]
			CM=[b.value for b in TBls1[acm][acr:luntb+1]]
			CB=[b.value for b in TBls1[acb][acr:luntb+1]]
			
			if(isChecked6==0):				
				print("is checked")
				# OMFrow=[b.value for b in TBls1[omfr][acr:luntb+1]]
				# OMFdescrip=[b.value for b in TBls1[omfd][acr:luntb+1]]
				# OMFType=[b.value for b in TBls1[omftype][acr:luntb+1]]
			else:
				OMFrow=[b.value for b in TBls1[omfr][acr:luntb+1]]
				OMFdescrip=[b.value for b in TBls1[omfd][acr:luntb+1]]
				OMFType=[b.value for b in TBls1[omftype][acr:luntb+1]]
			Lead=[b.value for b in TBls1[omfls][acr:luntb+1]]
			# Type=[b.value for b in TBls1[omftype][acr:luntb+1]]
			Sumlead.sheet_view.showGridLines = False
			maplead=list(set(Lead))
			# maptype=list(set(Type))

			for i in range(0,len(maplead)):
				try:
					work=TBls.create_sheet(str(maplead[i][:31]))
				except:
					work=TBls.create_sheet("NA")
				work.sheet_view.showGridLines = False
				work.cell(row=1,column=1).value="Client:"
				work.cell(row=1,column=1).font=font
				work.cell(row=1,column=2).value=namec
				work.cell(row=1,column=2).font=font
				work.cell(row=2,column=1).value="Period end:"
				work.cell(row=2,column=1).font=font
				work.cell(row=2,column=2).value=ant
				work.cell(row=2,column=2).number_format='mm/dd/yyyy'
				work.cell(row=2,column=2).font=font
				work.cell(row=1,column=11).value="Prepared by:"
				work.cell(row = 1, column = 12).value = preparedBy1
				work.cell(row = 2, column = 12).value = datePrepared1
				work.cell(row = 2, column = 12).number_format = 'mm/dd/yyyy'
				work.cell(row = 1, column = 12).font = font
				work.cell(row = 2, column = 12).font = font
				work.cell(row=1,column=11).font=font
				work.cell(row=2,column=11).value="Date:"
				work.cell(row=2,column=11).font=font
				work.cell(row=4,column=1).value=str(maplead[i])
				work.cell(row=4,column=1).font=font
				work.cell(row=8,column=2).value="Account"
				work.cell(row=8,column=3).value="Description"
				work.cell(row=8,column=4).value="OB"
				work.cell(row=8,column=5).value="DM"
				work.cell(row=8,column=6).value="CM"
				work.cell(row=8,column=7).value="CB"
				work.cell(row=8,column=8).value="OMF row"
				work.cell(row=8,column=9).value="Var."
				work.cell(row=8,column=10).value="Var. %"
				work.cell(row=8,column=11).value="Ref Wp"
				for m in range(2,9):
					work.cell(row=8,column=m).fill=gtfill
					work.cell(row=8,column=m).font=font2
				for m in range(9,12):
					work.cell(row=8,column=m).fill=greenfill
					work.cell(row=8,column=m).font=fontRed
				count=0
				for j in range(0,len(Lead)):
					if(Lead[j]==maplead[i]):
						count=count+1
						work.cell(row=8+count,column=2).value=Account[j]
						work.cell(row=8+count,column=2).font=normal
						work.cell(row=8+count,column=3).value=Description[j]
						work.cell(row=8+count,column=3).font=normal
						work.cell(row=8+count,column=4).value=OB[j]
						work.cell(row=8+count,column=4).number_format='#,##0_);(#,##0)'

						work.cell(row=8+count,column=4).font=normal
						work.cell(row=8+count,column=5).value=DM[j]
						work.cell(row=8+count,column=5).number_format='#,##0_);(#,##0)'
						work.cell(row=8+count,column=5).font=normal
						work.cell(row=8+count,column=6).value=CM[j]
						work.cell(row=8+count,column=6).number_format='#,##0_);(#,##0)'
						work.cell(row=8+count,column=6).font=normal
						work.cell(row=8+count,column=7).value=CB[j]
						work.cell(row=8+count,column=7).number_format='#,##0_);(#,##0)'
						work.cell(row=8+count,column=7).font=normal
						if(isChecked6==0):
							print("nada")
							# work.cell(row=8+count,column=8).value=OMFrow[j]
							# work.cell(row=8+count,column=8).font=normal
						else:
							work.cell(row=8+count,column=8).value=OMFrow[j]
							work.cell(row=8+count,column=8).font=normal
						work.cell(row=8+count,column=9).value="=G{0}-D{0}".format(count+8)
						work.cell(row=8+count,column=9).number_format='#,##0_);(#,##0)'
						work.cell(row=8+count,column=10).value="=IFERROR(I{0}/D{0},0)".format(count+8)
						work.cell(row=8+count,column=10).number_format='0.00%'

				work.cell(row=8+count+1,column=3).value="Total"
				work.cell(row=8+count+1,column=3).font=font
				work.cell(row=8+count+1,column=4).value="=sum(D9:D"+str(8+count)+")"
				work.cell(row=8+count+1,column=4).number_format='#,##0_);(#,##0)'
				work.cell(row=8+count+1,column=4).font=font
				work.cell(row=8+count+1,column=5).value="=sum(E9:E"+str(8+count)+")"
				work.cell(row=8+count+1,column=5).font=font
				work.cell(row=8+count+1,column=5).number_format='#,##0_);(#,##0)'
				work.cell(row=8+count+1,column=6).value="=sum(F9:F"+str(8+count)+")"
				work.cell(row=8+count+1,column=6).font=font
				work.cell(row=8+count+1,column=6).number_format='#,##0_);(#,##0)'
				work.cell(row=8+count+1,column=7).value="=sum(G9:G"+str(8+count)+")"
				work.cell(row=8+count+1,column=7).font=font
				work.cell(row=8+count+1,column=7).number_format='#,##0_);(#,##0)'
				
				work.column_dimensions['A'].width=10
				work.column_dimensions['B'].width=20
				work.column_dimensions['C'].width=15
				work.column_dimensions['D'].width=15
				work.column_dimensions['E'].width=15
				work.column_dimensions['F'].width=15
				work.column_dimensions['G'].width=15
				work.column_dimensions['H'].width=15
				work.column_dimensions['I'].width=15
				work.column_dimensions['J'].width=15 


				for p in range(2,12):
					work.cell(row=8,column=p).border=doubleborder
					work.cell(row=8+count,column=p).border=doubleborder

				for k in range(9,9+count):
					work.cell(row=k,column=9).font=fontRed
					work.cell(row=k,column=10).font=fontRed
				link="#'"+str(str(maplead[i])[:31])+"'!A4"
				try:
					Sumlead.cell(row=15+i,column=3).value=str(maplead[i][:31])
					Sumlead.cell(row=15+i,column=3).font=normal
					Sumlead.cell(row=15+i,column=3).hyperlink=link
					Sumlead.cell(row=15+i,column=2).value="=left(C{0},2)".format(15+i)
					Sumlead.cell(row=15+i,column=4).value="='"+str(maplead[i][:31])+"'!D"+str(count+9)
					Sumlead.cell(row=15+i,column=4).font=normal
					Sumlead.cell(row=15+i,column=4).number_format='#,##0_);(#,##0)'
					Sumlead.cell(row=15+i,column=5).value="='"+str(maplead[i][:31])+"'!G"+str(count+9)
					Sumlead.cell(row=15+i,column=5).font=normal
					Sumlead.cell(row=15+i,column=5).number_format='#,##0_);(#,##0)'
					Sumlead.cell(row=15+i,column=6).value="=E{0}-D{0}".format(i+15)
					Sumlead.cell(row=15+i,column=6).font=fontRed
					Sumlead.cell(row=15+i,column=6).number_format='#,##0_);(#,##0)'
					Sumlead.cell(row=15+i,column=7).value="=IFErROR(F{0}/D{0},0)".format(i+15)
					Sumlead.cell(row=15+i,column=7).font=fontRed
					Sumlead.cell(row=15+i,column=7).number_format='0.00%'
					Sumlead.cell(row=15+i,column=8).value='=IF(abs($E$10)<E{0},"Above threshold","Below Threshold")'.format(i+15)
					Sumlead.cell(row=15+i,column=8).font=normal
				except:
					Sumlead.cell(row=15+i,column=3).value="NA"  
					Sumlead.cell(row=15+i,column=4).value='=NA!D'+str(count+9)
					Sumlead.cell(row=15+i,column=5).value='=NA!G'+str(count+9)
				Sumlead.cell(row=5,column=4).value='=SUMIF(B:B,"BS",D:D)'
				Sumlead.cell(row=5,column=4).font=normal
				Sumlead.cell(row=5,column=4).number_format='#,##0_);(#,##0)'
				Sumlead.cell(row=6,column=4).value='=SUMIF(B:B,"PL",D:D)'
				Sumlead.cell(row=6,column=4).font=normal
				Sumlead.cell(row=6,column=4).number_format='#,##0_);(#,##0)'
				Sumlead.cell(row=7,column=4).value='=F_10_Trial_Balance!G11'
				Sumlead.cell(row=7,column=4).font=normal
				Sumlead.cell(row=7,column=4).number_format='#,##0_);(#,##0)'
				Sumlead.cell(row=8,column=4).value='=D7-D6'
				Sumlead.cell(row=8,column=4).font=fontRed
				Sumlead.cell(row=8,column=4).number_format='#,##0_);(#,##0)'
				Sumlead.cell(row=5,column=5).value='=SUMIF(B:B,"BS",E:E)'
				Sumlead.cell(row=5,column=5).font=normal
				Sumlead.cell(row=5,column=5).number_format='#,##0_);(#,##0)'
				Sumlead.cell(row=6,column=5).value='=SUMIF(B:B,"PL",E:E)'
				Sumlead.cell(row=6,column=5).font=normal
				Sumlead.cell(row=6,column=5).number_format='#,##0_);(#,##0)'
				Sumlead.cell(row=7,column=5).value='=F_10_Trial_Balance!J11'
				Sumlead.cell(row=7,column=5).font=normal
				Sumlead.cell(row=7,column=5).number_format='#,##0_);(#,##0)'
				Sumlead.cell(row=8,column=5).value='=E7-E6'
				Sumlead.cell(row=8,column=5).font=fontRed
				Sumlead.cell(row=8,column=5).number_format='#,##0_);(#,##0)'

			# for j in range(0,len(maptype)):
			# 	if(maptype[j]=="Banca"):
			# 		Sumlead.cell(row=15+j,column=2).value=str(maptype[j][:31])
			# 	else:
			# 		Sumlead.cell(row=15+j,column=2).value="Na"
			
			Sumlead.column_dimensions['A'].width=10
			Sumlead.column_dimensions['B'].width=30
			Sumlead.column_dimensions['C'].width=15
			Sumlead.column_dimensions['D'].width=15
			Sumlead.column_dimensions['E'].width=15
			Sumlead.column_dimensions['F'].width=15
			Sumlead.column_dimensions['G'].width=15
			Sumlead.column_dimensions['H'].width=15
			Sumlead.column_dimensions['I'].width=15

			if(isChecked6==0):				
				print("is checked")
				# cy=TBls["PBC_CY"]
				# py=TBls["PBC_PY"]
				# bs=TBls['BS Mapping']
				# pl=TBls['PL Mapping']
			else:
				cy=TBls["PBC_CY"]
				py=TBls["PBC_PY"]
				bs=TBls['BS Mapping']
				pl=TBls['PL Mapping']
			
			if(isChecked6==0):				
				print("is checked")
				# cy.sheet_state = 'hidden'
				# py.sheet_state = 'hidden'
				# bs.sheet_state = 'hidden'
				# pl.sheet_state = 'hidden'
			else:
				cy.sheet_state = 'hidden'
				py.sheet_state = 'hidden'
				bs.sheet_state = 'hidden'
				pl.sheet_state = 'hidden'
			Sumlead.auto_filter.ref = 'B14:H14'

			c = Sumlead['B15']
			Sumlead.freeze_panes = c

			

			file_pathLS=os.path.join(folderpath+"LSFS "+namec, "Leadsheets.xlsx")

			TBls.save(file_pathLS)
	
	
	make_archive(folderpath+"LSFS "+namec,folderpath+"LSFS "+namec+"/LS+ "+namec+".zip")
	# make_archive(folderpath+"/LSFS",folderpath+"/LSFS/LS+FS"+namec+".zip") ASA ERA INAINTE
					
	return send_from_directory(folderpath+"LSFS "+namec,"LS+ "+ namec+".zip",as_attachment=True)
	return render_template('TBLS.html')

@app.route('/FAR/Instructions', methods=['GET'])
def downloadFAR():
		filepath = "/home/auditappnexia/output/far"
 
		return send_from_directory(filepath,"Instructions - FAR-Reconciliation.docx", as_attachment=True)  
@app.route('/FAR/GTlM4R8dQS5LxdPkpfkPphvCgjq')
def FAR():
	return render_template('FAR.html')

@app.route('/FAR/GTlM4R8dQS5LxdPkpfkPphvCgjq', methods=['POST', 'GET'])
def FAR_process():
	namec=request.form['client']
	ant=datetime.datetime.strptime(
		request.form['yearEnd'],
		'%Y-%m-%d')
	postdate = datetime.datetime.strptime(
		request.form['endDep'],
		'%Y-%m-%d')
	yearentry = datetime.datetime.strptime(
		request.form['startDep'],
		'%Y-%m-%d')
	def make_archive(source, destination):
		base = os.path.basename(destination)
		name = base.split('.')[0]
		format = base.split('.')[1]
		archive_from = os.path.dirname(source)
		archive_to = os.path.basename(source.strip(os.sep))
		shutil.make_archive(name, format, archive_from, archive_to)
		shutil.move('%s.%s'%(name,format), destination)
	if request.method == 'POST':
		print("pas 1")
		file_location_FAR=request.files['far']
		file_location_TB=request.files['TB']
		file_location_ad=request.files['additionsFile']
		file_location_dis=request.files['disposalsFile']
		# file_location_CIP=request.files['cipDetails']
		val1 = request.form.get('Reco')
		val2 = request.form.get('DepReco')
		val3 = request.form.get('tom')
		if val1=="":
			print("Da")  # daca e bifat
			val1 = 1
		else:
			print(val1)            
			val1 = 0
			print("Nu")

		if val2=="":  # daca e bifat
			val2 = 1
		else:
			print(val2)            
			val2 = 0

		if val3=="":  # daca e bifat
			val3 = 1
		else:
			print(val3)            
			val3 = 0


		# appfar.mainloop()
		workingsblue2= Font(bold=True, italic=True, name='Tahoma', size=8,color='FFFFFF')
		lbluefill = PatternFill(start_color='00AEAC',
							   end_color='00AEAC',
							   fill_type='solid')
		grifill=PatternFill(start_color='c4d79b',end_color='c4d79b',fill_type='solid')
		yellow=PatternFill(start_color='ffff00',end_color='ffff00',fill_type='solid')
		blueFill = PatternFill(start_color='00AEAC',
							   end_color='00AEAC',
							   fill_type='solid')
		doubleborder = Border(bottom=Side(style='double'))
		solidborder = Border(bottom=Side(style='thick'))
		solidborderstanga = Border(left=Side(style='thin'))
		rightborder = Border(right=Side(style='thin'))
		rightdouble = Border (right=Side(style='thin'), bottom=Side(style='double'))
		rightmedium = Border (right=Side(style='thin'), bottom=Side(style='medium'))
		solidborderdreapta = Border(right=Side(style='thin'))
		solidbordersus = Border(top=Side(style='thin'))
		fontitalic = Font(name='Tahoma', size=8, bold=True, italic=True)
		font1 = Font(name='Tahoma', size=8)
		font2 = Font(name='Tahoma', size=8, bold=True)
		fontRed = Font(name='Tahoma', size=8, bold=True, color= 'FF0000')
		fontRedDiff=Font(name="Tahoma", color='FF0000', size=8, )
		fontGT = Font (name='GT Logo', size=8)
		workingsblue = Font(color='2F75B5', bold=True, name='Tahoma', size=8 )
		headers= Font(bold=True, italic=True, name='Tahoma', size=8,color='FFFFFF')
		headersblue = PatternFill(start_color='00AEAC',
						   end_color='00AEAC',
						   fill_type='solid')
		headerspurple= PatternFill(start_color='CCC0DA',
							 end_color='CCC0DA',
							 fill_type='solid')
		total=PatternFill(start_color='DDD9C4',
						   end_color='DDD9C4',
						   fill_type='solid')
		greenbolditalic= Font(bold=True, italic=True,  color='C0504D', name='Tahoma', size=8)
		greenbolditalic= Font(bold=True, italic=True,  color='00af50')
		fontalb = Font(italic=True, color="bfbfbf", size=8, name='Tahoma')

		FAR = openpyxl.load_workbook(file_location_FAR,data_only=True)
		FAR1 = FAR.active
		FAR1.title='FAR'
		TB = openpyxl.load_workbook(file_location_TB,data_only=True)
		TB1 = TB.active


		try:
			Add = openpyxl.load_workbook(file_location_ad,data_only=True)
			Add1 = Add.active
		except:
			pass
			# messagebox.showinfo("Information", "File: Addition detail. Please note that the Addition detail cannot be found or opened, the application will continue work without this detail.If a Addition detail is not available, press ok and ignore this message. If a Addition detail is available, please make sure that the file follows the instructions and re-run the application.")
			a=1
		aaaa=0
		try:
			Dis = openpyxl.load_workbook(file_location_dis,data_only=True)
			Disp1 = Dis.active
		except:
			pass
			# messagebox.showinfo("Information", "File: Disposal detail. Please note that the Disposal detail cannot be found or opened, the application will continue work without this detail.If a Disposal detail is not available, press ok and ignore this message. If a Disposal detail is available, please make sure that the file follows the instructions and re-run the application.")
			aaaa=1

		'Iterate from FAR ----------------------------------------------------'




		'Iterate from TB ----------------------------------------------------'


		FAR1.cell(row=1, column=1).fill=blueFill
		FAR1.cell(row=1, column=2).fill=blueFill
		FAR1.cell(row=1, column=3).fill=blueFill
		FAR1.cell(row=1, column=4).fill=blueFill
		FAR1.cell(row=1, column=5).fill=blueFill		
		FAR1.cell(row=1, column=6).fill=blueFill
		FAR1.cell(row=1, column=7).fill=blueFill
		FAR1.cell(row=1, column=8).fill=blueFill
		FAR1.cell(row=1, column=9).fill=blueFill
		FAR1.cell(row=1, column=10).fill=blueFill		
		FAR1.cell(row=1, column=11).fill=blueFill
		FAR1.cell(row=1, column=12).fill=blueFill
		FAR1.cell(row=1, column=13).fill=blueFill
		FAR1.cell(row=1, column=14).fill=blueFill
		FAR1.cell(row=1, column=15).fill=blueFill
		FAR1.cell(row=1, column=16).fill=blueFill			

		
		for row in TB1.iter_rows():
			for cell in row :
				if cell.value=="Account":
					tbacc=cell.column
					tbrow=cell.row

		for row in TB1.iter_rows():
			for cell in row :
				if cell.value=="Description" :

					tbdesc=cell.column

		for row in TB1.iter_rows():
			 for cell in row :
				 if cell.value=="DM" :

					 tbdrm=cell.column

		for row in TB1.iter_rows():
			for cell in row :
				if cell.value=="CM" :

					tbcrm=cell.column


		for row in TB1.iter_rows():
			 for cell in row :
				 if cell.value=="CB" :

					 tbcb=cell.column

		for row in TB1.iter_rows():
			 for cell in row:
				 if cell.value=="OB" :

					 tbob=cell.column
		try:
			for row in Add1.iter_rows():
						for cell in row:
							if cell.value=="Item":
								AddItem=cell.column
								Addrow=cell.row

			for row in Add1.iter_rows():
					for cell in row:
						if cell.value=="Account Add":
							AddAcount=cell.column
							Addrow=cell.row

			for row in Add1.iter_rows():
				for cell in row:
					if cell.value=="GBV":
						AddGBV=cell.column

			for row in Add1.iter_rows():
				for cell in row:
					if cell.value=="Depreciation Charge":
						AddCharge=cell.column

			for row in Add1.iter_rows():
				for cell in row:
					if cell.value=="Accumulated Depreciation":
						AddAccDepr=cell.column

			for row in Add1.iter_rows():
				for cell in row:
					if cell.value=="PIF Date":
						AddPif=cell.column

			for row in Add1.iter_rows():
				for cell in row:
					if cell.value=="Type":
						AddType=cell.column
		except:
			q=1



		'Iterate from Disposals detail ----------------------------------------------------'
		if(aaaa==0):
			for row in Disp1.iter_rows():
			  for cell in row:
				  if cell.value=="Account Disposals":
				  	DispAcount=cell.column
				  	Disprow=cell.row

			for row in Disp1.iter_rows():
			  for cell in row:
				  if cell.value=="Item":
					  DispItem=cell.column

			for row in Disp1.iter_rows():
			  for cell in row:
				  if cell.value=="Description":
					  DispDesc=cell.column

			for row in Disp1.iter_rows():
			  for cell in row:
				  if cell.value=="UL":
					  DispUL=cell.column


			for row in Disp1.iter_rows():
			  for cell in row:
				  if cell.value=="Disposal Date":
				  	DispDate=cell.column

			for row in Disp1.iter_rows():
			  for cell in row:
				  if cell.value=="GBV":
					  DispGBV=cell.column

			for row in Disp1.iter_rows():
			  for cell in row:
				  if cell.value=="Accumulated Depreciation":
					  DispAccDepr=cell.column

			for row in Disp1.iter_rows():
			  for cell in row:
				  if cell.value=="Disposals Charge":
					  DispDeprCh=cell.column

			for row in Disp1.iter_rows():
			 for cell in row:
			  if cell.value=="PIF Date":
				  DispPIF=cell.column
			for row in Disp1.iter_rows():
			 for cell in row:
			  if cell.value=="Type":
				  DispType=cell.column

		'Iterate from Transfer detail ----------------------------------------------------'
		# try:
		#   for row in CIP1.iter_rows():
		#     for cell in row:
		#       if cell.value=="Item":
		#         CIPItem=cell.column
		#         CIProw=cell.row
		#
		#   for row in CIP1.iter_rows():
		#     for cell in row:
		#       if cell.value=="Account":
		#         CIPAccount=cell.column
		#
		#   for row in CIP1.iter_rows():
		#     for cell in row:
		#       if cell.value=="CIP Addition":
		#         CIPAddition=cell.column
		#
		#   for row in CIP1.iter_rows():
		#     for cell in row:
		#       if cell.value=="CIP Disposal":
		#         CIPDisposal=cell.column
		#
		#   for row in CIP1.iter_rows():
		#     for cell in row:
		#       if cell.value=="CIP Type":
		#         CIPType=cell.column
		# except:
		#   a=1

		for row in FAR1.iter_rows():
			for cell in row:
				if cell.value=="Item":
						FARItem=cell.column
						FARrow=cell.row






		for row in FAR1.iter_rows():
				for cell in row:
					if cell.value=="Account FAR":
						FARAcount=cell.column
						FARrow=cell.row

		for row in FAR1.iter_rows():
			for cell in row:
				if cell.value=="GBV":
					FARDescription=cell.column

		for row in FAR1.iter_rows():
			for cell in row:
				if cell.value=="Accumulated Depreciation":
					FARAccDepr=cell.column


		for row in FAR1.iter_rows():
			for cell in row:
				if cell.value=="Depreciation Charge":
					FARCharge=cell.column

		for row in FAR1.iter_rows():
			  for cell in row:
				  if cell.value=="PIF Date":
					  FARPif=cell.column

		for row in FAR1.iter_rows():
		  for cell in row:
			  if cell.value=="UL":
				  FARUL=cell.column
		for row in FAR1.iter_rows():
				  for cell in row:
					  if cell.value=="Description":
						  FARDesc=cell.column

		try:
			lunfar=len(FAR1[FARAcount])
		except:
			flash("Please insert the correct header for Account FAR in FAR file")
			return render_template("index.html")
		try:
		  ItemFAR=[b.value for b in FAR1[FARItem][FARrow:lunfar+1]]
		except:
			flash("Please insert the correct header for Item in FAR file")
			return render_template("index.html")
		try:
		  AccountFAR=[b.value for b in FAR1[FARAcount][FARrow:lunfar+1]]
		except:
			flash("Please insert the correct header for Account FAR in FAR file")
			return render_template("index.html")
		try:
		  GBVFAR=[b.value for b in FAR1[FARDescription][FARrow:lunfar+1]]
		except:
			flash("Please insert the correct header for GBV in FAR file")
			return render_template("index.html")
		try:
		  ChargeFAR=[b.value for b in FAR1[FARCharge][FARrow:lunfar+1]]
		except:
			flash("Please insert the correct header for Depreciation Charge in FAR file")
			return render_template("index.html")
		try:
		  AccDeprFAR=[b.value for b in FAR1[FARAccDepr][FARrow:lunfar+1]]
		except:
			flash("Please insert the correct header for Accumulated Depreciation in FAR file")
			return render_template("index.html")

		
		try:
			DescriptionFAR=[b.value for b in FAR1[FARDesc][FARrow:lunfar+1]]
		except:
			flash("Please insert the correct header for Description in FAR file")
			return render_template("index.html")
		try:
			PIFDate=[b.value for b in FAR1[FARPif][FARrow:lunfar+1]]
		except:
			flash("Please insert the correct header for PIF Date in FAR file")
			return render_template("index.html")
		try:
			UL=[b.value for b in FAR1[FARUL][FARrow:lunfar+1]]
		except:
			flash("Please insert the correct header for UL in FAR file")
			return render_template("index.html")

		try:
			lunAdd=len(Add1[AddAcount])
			ItemAdd=[b.value for b in Add1[AddItem][Addrow:lunAdd+1]]
			AccountAdd=[b.value for b in Add1[AddAcount][Addrow:lunAdd+1]]
			GBVAdd=[b.value for b in Add1[AddGBV][Addrow:lunAdd+1]]
			AccumulatedDepreciation=[b.value for b in Add1[AddAccDepr][Addrow:lunAdd+1]]
			PIFDateadd=[b.value for b in Add1[AddPif][Addrow:lunAdd+1]]
			Charge=[b.value for b in Add1[AddCharge][Addrow:lunAdd+1]]
			TypeAdd=[b.value for b in Add1[AddType][Addrow:lunAdd+1]]
		except:
			a=1

		pur=0
		trans=0
		try:
			for i in range(0,len(TypeAdd)):
			  if(TypeAdd[i]=='Pure'):
				  pur=pur+1
			  elif(TypeAdd[i]=='Transfer'):
				  trans=trans+1
			  # else:
				# messagebox.showerror("Error", "File: Additions detail. Please verify the column Type and make sure that all the rows have one of values: Pure or Transfer!")
				#sys.exit()
		except:
			p=1
		if(aaaa==0):
			try:
			  lundisp=len(Disp1[DispAcount])
			except:
				flash("Please insert the correct header for Account in Disposal file")
				return render_template("index.html")
			  #sys.exit()

			try:
			  ItemDisp=[b.value for b in Disp1[DispItem][Disprow:lundisp+1]]
			except:
				flash("Please insert the correct header for Item in Disposal file")
				return render_template("index.html")
			#   messagebox.showerror("Error", "File: Disposals detail. Please insert the correct header for 'Item'")
			  #sys.exit()
			try:
			  TypeDisp=[b.value for b in Disp1[DispType][Disprow:lundisp+1]]
			except:
				flash("Please insert the correct header for Disposal Type in Disposal file")
				return render_template("index.html")
			#   messagebox.showerror("Error", "File: Disposals detail. Please insert the correct header for 'Item'")
			  #sys.exit()
			try:
			  AccountDisp=[b.value for b in Disp1[DispAcount][Disprow:lundisp+1]]
			except:
				flash("Please insert the correct header for Account Disposals in Disposal file")
				return render_template("index.html")
			#   messagebox.showerror("Error", "File: Disposals detail. Please insert the correct header for 'Account Disposals'")
			  #sys.exit()
			try:
			  DescriptionDisp=[b.value for b in Disp1[DispDesc][Disprow:lundisp+1]]
			except:
				flash("Please insert the correct header for Description in Disposal file")
				return render_template("index.html")
			#   messagebox.showerror("Error", "File: Disposals detail. Please insert the correct header for 'Description'")
			  #sys.exit()
			try:
			  DispDate=[b.value for b in Disp1[DispDate][Disprow:lundisp+1]]
			except:
				flash("Please insert the correct header for Disposal Date in Disposal file")
				return render_template("index.html")
			#   messagebox.showerror("Error", "File: Disposals detail. Please insert the correct header for 'Disposal Date'")
			  #sys.exit()
			try:
			  ULDisp=[b.value for b in Disp1[DispUL][Disprow:lundisp+1]]
			except:
				flash("Please insert the correct header for UL in Disposal file")
				return render_template("index.html")
			#   messagebox.showerror("Error", "File: Disposals detail. Please insert the correct header for 'UL'")
			  #sys.exit()
			try:
			  DeprChDisp=[b.value for b in Disp1[DispDeprCh][Disprow:lundisp+1]]
			except:
				flash("Please insert the correct header for Depreciation Charge in Disposal file")
				return render_template("index.html")
			#   messagebox.showerror("Error", "File: Disposals detail. Please insert the correct header for 'Disposals Charge'")
			  #sys.exit()
			try:
			  PIFDisp=[b.value for b in Disp1[DispPIF][Disprow:lundisp+1]]
			except:
				flash("Please insert the correct header for PIF in Disposal file")
				return render_template("index.html")
			#   messagebox.showerror("Error", "File: Disposals detail. Please insert the correct header for 'PIF'")
			  #sys.exit()
			try:
			  GBVDisp=[b.value for b in Disp1[DispGBV][Disprow:lundisp+1]]
			except:
				flash("Please insert the correct header for GBV in Disposal file")
				return render_template("index.html")
			#   messagebox.showerror("Error", "File: Disposals detail. Please insert the correct header for 'GBV'")
			  #sys.exit()
			try:
			  AccumulatedDepreciationDisp=[b.value for b in Disp1[DispAccDepr][Disprow:lundisp+1]]
			except:
				flash("Please insert the correct header for Accumulated Depreciation in Disposal file")
				return render_template("index.html")
			#   messagebox.showerror("Error", "File: Disposals detail. Please insert the correct header for 'Accumulated Depreciation'")
			  #sys.exit()

		# sale=0
		# scrap=0
		# for i in range(0,len(TypeDis)):
		#   if(TypeDis[i]=='Sale'):
		#     sale=sale+1
		#   elif(TypeDis[i]=='Scrap'):
		#     scrap=scrap+1
		  # else:
		  #   messagebox.showerror("Error", "File: Disposals detail. Please verify the column Type and make sure that all the rows have one of values: Sale or Scrap!")
		  #   #sys.exit()

		# try:
		#   lunCIP=len(CIP1[CIPAccount])
		#
		#   ItemCIP=[b.value for b in CIP1[CIPItem][CIProw:lunCIP+1]]
		#   AccountCIP=[b.value for b in CIP1[CIPAccount][CIProw:lunCIP+1]]
		#   AdditionCIP=[b.value for b in CIP1[CIPAddition][CIProw:lunCIP+1]]
		#   DisposalCIP=[b.value for b in CIP1[CIPDisposal][CIProw:lunCIP+1]]
		#   TypeCIP=[b.value for b in CIP1[CIPType][CIProw:lunCIP+1]]
		#
		#   purCIP=0
		#   transCIP=0
		#   saleCIP=0
		#   scrapCIP=0
		#   for i in range(0,len(TypeCIP)):
		#     if(TypeCIP[i]=='Pure'):
		#       purCIP=purCIP+1
		#     elif(TypeCIP[i]=='Transfer'):
		#       transCIP=transCIP+1
		#     elif(TypeCIP[i]=='Sales'):
		#       saleCIP=saleCIP+1
		#     elif(TypeCIP[i]=='Scrapped'):
		#       scrapCIP=scrapCIP+1
		#     else:
		#       messagebox.showerror("Error", "File: CIP detail. Please verify the column Type and make sure that all the rows have one of values: Pure for Additions or Transfer/Sales/Scrapped for Disposals!")
		#       #sys.exit()
		# except:
		#   a=1
		#   messagebox.showinfo("Information","File: CIP detail. Please note that CIP detail cannot be found or opened, the application will continue work without this detail.If a CIP detail is not available, press ok and ignore this message. If a CIP detail is available, please make sure that the file follows the instructions and re-run the application. ")
		luntb=len(TB1[tbacc])
		try:
		  AccountTB=[b.value for b in TB1[tbacc][tbrow:luntb+1]]
		  DescriptionTB=[b.value for b in TB1[tbdesc][tbrow:luntb+1]]
		  OBTB=[b.value for b in TB1[tbob][tbrow:luntb+1]]
		  DMTB=[b.value for b in TB1[tbdrm][tbrow:luntb+1]]
		  CMTB=[b.value for b in TB1[tbcrm][tbrow:luntb+1]]
		  CBTB=[b.value for b in TB1[tbcb][tbrow:luntb+1]]
		except:
			pass
		#   messagebox.showerror("Error", "File: TB. Please rename the headers following instructions. One or more headers could not be found. Be careful at spaces or key-sensitivity!")

		PBC_FAR =FAR.create_sheet("PBC_FAR")

		F10_TB =FAR.create_sheet("F10_TB")
		PBC_FAR.cell(row=1, column=1).value="Item"
		PBC_FAR.cell(row=1, column=2).value="Account"
		PBC_FAR.cell(row=1, column=3).value="GBV"
		PBC_FAR.cell(row=1, column=4).value="Accumulated Depreciation"
		PBC_FAR.cell(row=1, column=5).value="Accumulated Impairment"
		PBC_FAR.cell(row=1, column=6).value="Depr Account"
		PBC_FAR.cell(row=1, column=7).value="Synt Fixed Assets Accounts"
		PBC_FAR.cell(row=1, column=8).value="Impairment Account"
		PBC_FAR.cell(row=1, column=9).value="Depreciation Charge"
		PBC_FAR.cell(row=1, column=10).value="Impairment Charge"

		for i in range (1,11):
			PBC_FAR.cell(row=1, column= i).border=doubleborder
			PBC_FAR.cell(row=1, column= i).font=font2
			PBC_FAR.cell(row=1, column= i).fill=headersblue

		for i in range (12, 14):
			PBC_FAR.cell(row=1, column= i).border=doubleborder
			PBC_FAR.cell(row=1, column= i).font=font2
			PBC_FAR.cell(row=1, column= i).fill=headersblue



		for i in range(2, 18):
			PBC_FAR.cell(row=i, column=12).alignment = Alignment (horizontal='right')

		for i in range(2, 18):
			PBC_FAR.cell(row=i, column=13).number_format='#,##0_);(#,##0)'

		for i in range (1, len(ItemFAR)+1):
			PBC_FAR.cell(row=i+1, column=1).value=ItemFAR[i-1]


		for i in range (1, len(AccountFAR)+1):
			PBC_FAR.cell(row=i+1, column=2).value=AccountFAR[i-1]

		for i in range (1, len(GBVFAR)+1):
			PBC_FAR.cell(row=i+1, column=3).value=GBVFAR[i-1]

		for i in range (1, len(AccDeprFAR)+1):
			PBC_FAR.cell(row=i+1, column=4).value=AccDeprFAR[i-1]


		for i in range (1, len(AccountFAR)+1):
			PBC_FAR.cell(row=i+1, column=6).value='=IF(G{0}="201",2801,IF(G{0}="203",2803,IF(G{0}="205",2805,IF(G{0}="208",2808,IF(G{0}="206",2806,IF(G{0}="207",2807,IF(G{0}="211",2811,IF(G{0}="212",2812,IF(G{0}="213",2813,IF(G{0}="223",2813,IF(G{0}="214",2814,IF(G{0}="224",2814,IF(G{0}="215",2815,IF(G{0}="216",2816,IF(G{0}="217",2817,0)))))))))))))))'.format(i+1)

		for i in range (1,len(AccountFAR)+1):
			PBC_FAR.cell(row=i+1, column=7).value='=Left(B{0},3)'.format(i+1)

		for i in range (1,len(AccountFAR)+1):
			PBC_FAR.cell(row=i+1, column=8).value='=IF(G{0}="203",2903,IF(G{0}="205",2905,IF(G{0}="206",2906,IF(G{0}="208",2908,IF(G{0}="211",2911,IF(G{0}="212",2912,IF(G{0}="213",2913,IF(G{0}="223",2913,IF(G{0}="214",2914,IF(G{0}="224",2914,IF(G{0}="215",2915,IF(G{0}="216",2916,IF(G{0}="217",2917,IF(G{0}="227",2917,0))))))))))))))'.format(i+1)

		for i in range (1, len(AccountFAR)+1):
			for j in range (3, 6):
				PBC_FAR.cell(row=i+1, column=j).number_format='#,##0_);(#,##0)'

		for i in range (1, len(AccountFAR)+1):
			for j in range (9, 11):
				PBC_FAR.cell(row=i+1, column=j).number_format='#,##0_);(#,##0)'

		for i in range(1, len(ChargeFAR)+1):
			PBC_FAR.cell(row=i+1, column=9).value=ChargeFAR[i-1]



		PBC_FAR.cell(row=14, column=12).font=font2
		PBC_FAR.cell(row=14, column=13).font=fontRed

		PBC_FAR.cell(row=5, column=12).font=font2
		PBC_FAR.cell(row=5, column=13).font=fontRed

		for i in range(2, 5):
			PBC_FAR.cell(row=i, column=13).font=fontRedDiff

		for i in range(7, 14):
			PBC_FAR.cell(row=i, column=13).font=fontRedDiff

		PBC_FAR.cell(row=13, column=12).border=doubleborder
		PBC_FAR.cell(row=13, column=13).border=doubleborder

		PBC_FAR.cell(row=4, column=12).border=doubleborder
		PBC_FAR.cell(row=4, column=13).border=doubleborder



		for i in range(12, 14):
			PBC_FAR.cell(row=16, column=i).fill=total
		PBC_FAR.cell(row=16, column=13).font=fontRed
		PBC_FAR.cell(row=16, column=12).font=font2

		"Adjust Column Width---------------------------------------------------------------------------------------------------------------------------------."

		PBC_FAR.column_dimensions['A'].width=25
		PBC_FAR.column_dimensions['B'].width=33
		PBC_FAR.column_dimensions['C'].width=25
		PBC_FAR.column_dimensions['D'].width=35
		PBC_FAR.column_dimensions['E'].width=25
		PBC_FAR.column_dimensions['F'].width=25
		PBC_FAR.column_dimensions['G'].width=25
		PBC_FAR.column_dimensions['H'].width=25
		PBC_FAR.column_dimensions['I'].width=25
		PBC_FAR.column_dimensions['J'].width=25

		PBC_FAR.column_dimensions['L'].width=35
		PBC_FAR.column_dimensions['M'].width=25


		"Create PBC TB ---------------------------------------------------------------------"

		F10_TB.cell(row=1, column=1).value="Class"
		F10_TB.cell(row=1, column=2).value="Synt 3"
		F10_TB.cell(row=1, column=3).value="Account"
		F10_TB.cell(row=1, column=4).value="Description"
		F10_TB.cell(row=1, column=5).value="OB"
		F10_TB.cell(row=1, column=6).value="DM"
		F10_TB.cell(row=1, column=7).value="CM"
		F10_TB.cell(row=1, column=8).value="CB"
		F10_TB.cell(row=1, column=9).value="Check"
		F10_TB.cell(row=1, column=10).value="Synt 4"

		for i in range (1,11):
			F10_TB.cell(row=1, column= i).border=doubleborder
			F10_TB.cell(row=1, column= i).font=font2
			F10_TB.cell(row=1, column= i).fill=headersblue

		for i in range (1,len(AccountTB)):
			F10_TB.cell(row=i+1, column=3).value=AccountTB[i-1]

		for i in range (1,len(DescriptionTB)):
			F10_TB.cell(row=i+1, column=4).value=DescriptionTB[i-1]

		for i in range (1, len(OBTB)):
			F10_TB.cell(row=i+1, column=5).value=OBTB[i-1]

		for i in range (1, len(DMTB)):
			F10_TB.cell(row=i+1, column=6).value=DMTB[i-1]

		for i in range (1, len(CMTB)):
			F10_TB.cell(row=i+1, column=7).value=CMTB[i-1]

		for i in range (1, len(CBTB)):
			F10_TB.cell(row=i+1, column=8).value=CBTB[i-1]

		for i in range (1, len(AccountTB)):
			F10_TB.cell(row=i+1, column= 9).value='=IF(A{0}<"6",E{0}+F{0}-G{0}-H{0},F{0}-G{0})'.format(i+1)

		for i in range (1, len(AccountTB)+1):
			for j in range (5, 10):
				F10_TB.cell(row=i+1, column=j).number_format='#,##0_);(#,##0)'

		for i in range (1,len(AccountTB)):
			F10_TB.cell(row=i+1, column=2).value='=Left(C{0},3)'.format(i+1)

		for i in range (1,len(AccountTB)):
			F10_TB.cell(row=i+1, column=1).value='=Left(C{0},1)'.format(i+1)

		for i in range(1, len(AccountTB)):
			F10_TB.cell(row=i+1, column=10).value="=LEFT(C{0},4)".format(i+1)

		"Adjust Column Width---------------------------------------------------------------------------------------------------------------------------------."

		F10_TB.column_dimensions['A'].width=25
		F10_TB.column_dimensions['B'].width=25
		F10_TB.column_dimensions['C'].width=25
		F10_TB.column_dimensions['D'].width=35
		F10_TB.column_dimensions['E'].width=25
		F10_TB.column_dimensions['F'].width=25
		F10_TB.column_dimensions['G'].width=25
		F10_TB.column_dimensions['H'].width=25
		F10_TB.column_dimensions['I'].width=25
		F10_TB.column_dimensions['J'].width=25



		PBC_Disposals=FAR.create_sheet("PBC_Disposals")
		PBC_Disposals.cell(row=1, column=1).value="Synt 3"
		PBC_Disposals.cell(row=1, column=2).value="Account"
		PBC_Disposals.cell(row=1, column=3).value="Item"
		PBC_Disposals.cell(row=1, column=4).value="Description"
		PBC_Disposals.cell(row=1, column=5).value="PIF Date"
		PBC_Disposals.cell(row=1, column=6).value="Disposals Date"
		PBC_Disposals.cell(row=1, column=7).value="UL"
		PBC_Disposals.cell(row=1, column=8).value="GBV"
		PBC_Disposals.cell(row=1, column=9).value="Depreciation Charge"
		PBC_Disposals.cell(row=1, column=10).value="Accumulated Depreciation"
		PBC_Disposals.cell(row=1, column=11).value="Type"
		PBC_Disposals.cell(row=1, column=12).value="Depreciation Account"


		for i in range (1,12):
		  PBC_Disposals.cell(row=1, column= i).border=doubleborder
		  PBC_Disposals.cell(row=1, column= i).font=font2
		  PBC_Disposals.cell(row=1, column= i).fill=blueFill
		if(aaaa==0):
		  for i in range(1, len(AccountDisp)+1):
			  PBC_Disposals.cell(row=i+1, column=1).value="=LEFT(TRIM(B{0}),3)".format(i+1)

		  for i in range (1, len(AccountDisp)+1):

			  PBC_Disposals.cell(row=i+1, column=2).value=AccountDisp[i-1]


		  for i in range (1, len(ItemDisp)+1):
			  PBC_Disposals.cell(row=i+1, column=3).value=ItemDisp[i-1]

		  for i in range (1, len(DescriptionDisp)+1):
			  PBC_Disposals.cell(row=i+1, column=4).value=DescriptionDisp[i-1]


		  
		#   if(aa==False):
		#         messagebox.showerror("Error", "File: Disposal detail. Please make sure that the format for PIF Disposals is DATE!")
		#         #sys.exit()

		  for i in range (1, len(PIFDisp)+1):
			  PBC_Disposals.cell(row=i+1, column=5).value=PIFDisp[i-1]
			  PBC_Disposals.cell(row=i+1, column=5).number_format='mm/dd/yyyy'


		#   if(bb==False):
			#     messagebox.showerror("Error", "File: Disposal detail. Please make sure that the format for Disposal Date is DATE!")
				#sys.exit()

		  for i in range (1, len(DispDate)+1):
			  PBC_Disposals.cell(row=i+1, column=6).value=DispDate[i-1]
			  PBC_Disposals.cell(row=i+1, column=6).number_format='mm/dd/yyyy'

		  for i in range (1, len(ULDisp)+1):
			  PBC_Disposals.cell(row=i+1, column=7).value=ULDisp[i-1]

		  for i in range (1, len(GBVDisp)+1):
			  PBC_Disposals.cell(row=i+1, column=8).value=GBVDisp[i-1]
			  PBC_Disposals.cell(row=i+1, column=8).number_format='#,##0_);(#,##0)'
			  PBC_Disposals.cell(row=i+1,column=11).value=TypeDisp[i-1]

		  for i in range (1, len(DeprChDisp)+1):
			  PBC_Disposals.cell(row=i+1, column=9).value=DeprChDisp[i-1]
			  PBC_Disposals.cell(row=i+1,column=9).number_format='#,##0_);(#,##0)'

		  for i in range (1, len(AccumulatedDepreciationDisp)+1):
			  PBC_Disposals.cell(row=i+1, column=10).value=AccumulatedDepreciationDisp[i-1]
			  PBC_Disposals.cell(row=i+1,column=10).number_format='#,##0_);(#,##0)'
		  for i in range (1,len(AccountDisp)+1):
			  PBC_Disposals.cell(row=i+1, column=12).value='=IF(A{0}="201",2801,IF(A{0}="203",2803,IF(A{0}="205",2805,IF(A{0}="208",2808,IF(A{0}="206",2806,IF(A{0}="2071",2807,IF(A{0}="211",2811,IF(A{0}="212",2812,IF(A{0}="213",2813,IF(A{0}="223",2813,IF(A{0}="214",2814,IF(A{0}="224",2814,IF(A{0}="215",2815,IF(A{0}="216",2816,IF(A{0}="217",2817,0)))))))))))))))'.format(i+1)

		PBC_Disposals.column_dimensions['A'].width=25
		PBC_Disposals.column_dimensions['B'].width=25
		PBC_Disposals.column_dimensions['C'].width=25
		PBC_Disposals.column_dimensions['D'].width=35
		PBC_Disposals.column_dimensions['E'].width=25
		PBC_Disposals.column_dimensions['F'].width=25
		PBC_Disposals.column_dimensions['G'].width=25
		PBC_Disposals.column_dimensions['H'].width=25
		PBC_Disposals.column_dimensions['I'].width=25
		PBC_Disposals.column_dimensions['J'].width=25

		for i in range(1,11):
		  for j in range(2,PBC_Disposals.max_row+1):
			  PBC_Disposals.cell(row=j,column=i).font=font1

		if(val1==1):
			G100FAR=FAR.create_sheet("G10 FAR Recon")
			

			G100FAR.title="G10 FAR Recon"

			for i in range(1, 114):
				for j in range(1, 10):
					G100FAR.cell(row=i, column=j).font=font1


			G100FAR.cell(row=1, column=1).value="Client:"
			G100FAR.cell(row=1, column=1).font=font2

			G100FAR.cell(row=1, column=2).value=namec
			G100FAR.cell(row=1, column=2).font=font2

			G100FAR.cell(row=2, column=1).value="Period end:"
			G100FAR.cell(row=2, column=1).font=font2

			G100FAR.cell(row=2, column=2).value=ant
			G100FAR.cell(row=2, column=2).number_format='mm/dd/yyyy'
			G100FAR.cell(row=2, column=2).font=font2

			# G100FAR.cell(row=1, column=6).value="Prepared by:"
			# G100FAR.cell(row=1, column=6).font=font2

			G100FAR.cell(row=2, column=6).value="Date:"
			G100FAR.cell(row=2, column=6).font=font2
			G100FAR.cell(row=2, column=7).value=datetime.datetime.now().date()
			G100FAR.cell(row=2, column=7).number_format="mm/dd/yyyy"
			G100FAR.cell(row=2, column=7).alignment = Alignment (horizontal='left')

			for i in range(1, 4):
				G100FAR.cell(row=i, column=6).alignment=Alignment(horizontal='right')

			G100FAR.cell(row=3, column=6).value="Ref:"
			G100FAR.cell(row=3, column=6).font=font2
			G100FAR.cell(row=3, column=7).value="G10"
			G100FAR.cell(row=3, column=7).font=fontRed

			for i in range(1,4):
				G100FAR.cell(row=i, column=8).alignment = Alignment(horizontal='right')

			G100FAR.cell(row=4, column=3).value="Fixed Assets Register Reconciliation"
			G100FAR.cell(row=4, column=3).font=font2

			G100FAR.cell(row=7, column=1).value="Procedure:"
			G100FAR.cell(row=7, column=1).alignment=Alignment(horizontal='right')
			G100FAR.cell(row=7, column=2).value="Obtain the FAR as at period end and perform the following: "
			G100FAR.cell(row=8, column=2).value="- perform reconciliation for FA GBV, accumulated depreciation, accumulated impairement, depreciation and impairement charge with the TB: "
			G100FAR.cell(row=9, column=2).value="- recompute the NBV  as per FAR and reconcile it with TB"
			G100FAR.cell(row=10, column=1).value="Work done:"
			G100FAR.cell(row=10, column=1).alignment=Alignment(horizontal='right')
			G100FAR.cell(row=10, column=2).value="Please see the work below:"
			G100FAR.cell(row=7, column=1).font=font2
			G100FAR.cell(row=10, column=1).font=font2

			G100FAR.cell(row=12, column=1).value="Working 1: GBV"
			G100FAR.cell(row=12, column=1).font=workingsblue

			G100FAR.cell(row=14,column=1).value="Synt Account"
			G100FAR.cell(row=14,column=2).value="Description"
			G100FAR.cell(row=14,column=3).value="As per FAR(PBC_FAR)"
			G100FAR.cell(row=14,column=4).value="As per TB(F10_TB)"
			G100FAR.cell(row=14,column=5).value="Check"

			for i in range (1,6):
				G100FAR.cell(row=14, column=i).font=headers
				G100FAR.cell(row=14, column=i).border=solidborder
				G100FAR.cell(row=14, column=i).fill=headersblue


			for i in range (1,6):
				G100FAR.cell(row=23, column=i).font=headers
				G100FAR.cell(row=23, column=i).border=solidborder
				G100FAR.cell(row=23, column=i).fill=headersblue


			for i in range (1,6):
				G100FAR.cell(row=47, column=i).font=headers
				G100FAR.cell(row=47, column=i).border=solidborder
				G100FAR.cell(row=47, column=i).fill=headersblue

			"Adjust Column Width---------------------------------------------------------------"

			G100FAR.column_dimensions['A'].width=25
			G100FAR.column_dimensions['B'].width=33
			G100FAR.column_dimensions['C'].width=25
			G100FAR.column_dimensions['D'].width=35
			G100FAR.column_dimensions['E'].width=25
			G100FAR.column_dimensions['F'].width=25
			G100FAR.column_dimensions['G'].width=25
			G100FAR.column_dimensions['H'].width=25
			G100FAR.column_dimensions['I'].width=25



			G100FAR.cell(row=13, column=1).value="'-Intangible Assets-"
			G100FAR.cell(row=13, column=1).font=greenbolditalic

			G100FAR.cell(row=15, column=1).value="201"
			G100FAR.cell(row=16, column=1).value="203"
			G100FAR.cell(row=17, column=1).value="205 & 208"
			G100FAR.cell(row=18, column=1).value="206"
			G100FAR.cell(row=19, column=1).value="2071"
			G100FAR.cell(row=20, column=1).value="Total Intangibles Assets"
			for i in range(15, 20):
				G100FAR.cell(row=i, column=1).alignment=Alignment(horizontal='right')

			G100FAR.cell(row=15, column=2).value="Set-up costs"
			G100FAR.cell(row=16, column=2).value="Development costs"
			G100FAR.cell(row=17, column=2).value="Licenses, Trademarks"
			G100FAR.cell(row=18, column=2).value="Intangible assets for mineral resources"
			G100FAR.cell(row=19, column=2).value="Goodwill"


			G100FAR.cell(row=15, column=3).value='=SUMIF(PBC_FAR!G:G,"201",PBC_FAR!C:C)'
			G100FAR.cell(row=16, column=3).value='=SUMIF(PBC_FAR!G:G,"203",PBC_FAR!C:C)'
			G100FAR.cell(row=17, column=3).value='=SUMIF(PBC_FAR!G:G,"205",PBC_FAR!C:C)+SUMIF(PBC_FAR!G:G,"208",PBC_FAR!C:C)'
			G100FAR.cell(row=18, column=3).value='=SUMIF(PBC_FAR!G:G,"206",PBC_FAR!C:C)'
			G100FAR.cell(row=19, column=3).value='=SUMIF(PBC_FAR!G:G,"207",PBC_FAR!C:C)'

			G100FAR.cell(row=15, column=4).value='=SUMIF(F10_TB!B:B,"201",F10_TB!H:H)'
			G100FAR.cell(row=16, column=4).value='=SUMIF(F10_TB!B:B,"203",F10_TB!H:H)'
			G100FAR.cell(row=17, column=4).value='=SUMIF(F10_TB!B:B,"205",F10_TB!H:H)+SUMIF(F10_TB!B:B,"208",F10_TB!H:H)'
			G100FAR.cell(row=18, column=4).value='=SUMIF(F10_TB!B:B,"206",F10_TB!H:H)'
			G100FAR.cell(row=19, column=4).value='=SUMIF(F10_TB!B:B,"207",F10_TB!H:H)'

			for i in range(14, 20):
				G100FAR.cell(row=i+1, column=5).value="=C{0}-D{0}".format(i+1)

			for i in range(15,20):
				G100FAR.cell(row=i,column=5).font=fontRedDiff

			G100FAR.cell(row=20, column=5).font=fontRed

			for i in range (1,6):
				G100FAR.cell(row=19, column=i).border=doubleborder

			G100FAR.cell(row=20, column=3).value='=SUM(C15:C19)'
			G100FAR.cell(row=20, column=4).value='=SUM(D15:D19)'
			G100FAR.cell(row=20, column=5).value='=SUM(E15:E19)'

			for i in range (1,5):
				G100FAR.cell(row=20, column=i).font=font2



			for i in range(15,99):
				for j in range (3,7):
					G100FAR.cell(row=i, column=j).number_format='#,##0_);(#,##0)'

			G100FAR.cell (row=22, column=1).value="'-Tangible Assets-"
			G100FAR.cell(row=22, column=1).font=greenbolditalic

			G100FAR.cell(row=23,column=1).value="Synt Account"
			G100FAR.cell(row=23,column=2).value="Description"
			G100FAR.cell(row=23,column=3).value="As per FAR(PBC_FAR)"
			G100FAR.cell(row=23,column=4).value="As per TB(F10_TB)"
			G100FAR.cell(row=23,column=5).value="Check"


			G100FAR.cell(row=24, column=1).value="211"
			G100FAR.cell(row=25, column=1).value="212"
			G100FAR.cell(row=26, column=1).value="213 & 223"
			G100FAR.cell(row=27, column=1).value="214 & 224"
			G100FAR.cell(row=28, column=1).value="215"
			G100FAR.cell(row=29, column=1).value="216"
			G100FAR.cell(row=29, column=1).value="217 & 227"
			G100FAR.cell(row=30, column=1).value='Total Tangibles Assets'

			for i in range(24, 31):
				G100FAR.cell(row=i, column=1).alignment=Alignment(horizontal='right')

			G100FAR.cell(row=24, column=2).value="Land"
			G100FAR.cell(row=25, column=2).value="Buildings"
			G100FAR.cell(row=26, column=2).value="Equipments and Machines"
			G100FAR.cell(row=27, column=2).value="Furniture, office equipments"
			G100FAR.cell(row=28, column=2).value="Investment properties"
			G100FAR.cell(row=29, column=2).value="Tangible assets for mineral resources"
			G100FAR.cell(row=30, column=2).value="Biological assets"

			G100FAR.cell(row=24, column=3).value='=SUMIF(PBC_FAR!G:G,"211",PBC_FAR!C:C)'
			G100FAR.cell(row=25, column=3).value='=SUMIF(PBC_FAR!G:G,"212",PBC_FAR!C:C)'
			G100FAR.cell(row=26, column=3).value='=SUMIF(PBC_FAR!G:G,"213",PBC_FAR!C:C)+SUMIF(PBC_FAR!G:G,"223",PBC_FAR!C:C)'
			G100FAR.cell(row=27, column=3).value='=SUMIF(PBC_FAR!G:G,"214",PBC_FAR!C:C)+SUMIF(PBC_FAR!G:G,"224",PBC_FAR!C:C)'
			G100FAR.cell(row=28, column=3).value='=SUMIF(PBC_FAR!G:G,"215",PBC_FAR!C:C)'
			G100FAR.cell(row=29, column=3).value='=SUMIF(PBC_FAR!G:G,"216",PBC_FAR!C:C)'
			G100FAR.cell(row=30, column=3).value='=SUMIF(PBC_FAR!G:G,"217",PBC_FAR!C:C)+SUMIF(PBC_FAR!G:G,"227",PBC_FAR!C:C)'

			G100FAR.cell(row=24, column=4).value='=SUMIF(F10_TB!B:B,"211",F10_TB!H:H)'
			G100FAR.cell(row=25, column=4).value='=SUMIF(F10_TB!B:B,"212",F10_TB!H:H)'
			G100FAR.cell(row=26, column=4).value='=SUMIF(F10_TB!B:B,"213",F10_TB!H:H)+SUMIF(F10_TB!B:B,"223",F10_TB!H:H)'
			G100FAR.cell(row=27, column=4).value='=SUMIF(F10_TB!B:B,"214",F10_TB!H:H)+SUMIF(F10_TB!B:B,"224",F10_TB!H:H)'
			G100FAR.cell(row=28, column=4).value='=SUMIF(F10_TB!B:B,"215",F10_TB!H:H)'
			G100FAR.cell(row=29, column=4).value='=SUMIF(F10_TB!B:B,"216",F10_TB!H:H)'
			G100FAR.cell(row=30, column=4).value='=SUMIF(F10_TB!B:B,"217",F10_TB!H:H)'

			for i in range(23, 30):
				G100FAR.cell(row=i+1, column=5).value="=C{0}-D{0}".format(i+1)

			G100FAR.cell(row=31,column=3).value='=SUM(C24:C30)'
			G100FAR.cell(row=31,column=4).value='=SUM(D24:D30)'
			G100FAR.cell(row=31,column=5).value='=SUM(E24:E30)'

			for i in range (1,5):
				G100FAR.cell(row=31,column=i).font=font2

			G100FAR.cell(row=31, column=5).font=fontRed

			for i in range (1,6):
				G100FAR.cell(row=29, column=i).border=doubleborder

			for i in range (24,31):
				G100FAR.cell(row=i, column=5).font=fontRedDiff

			G100FAR.cell(row=31, column=1).value="Total Tangible Assets"

			G100FAR.cell(row=33, column=1).value='Total FA GBV'
			G100FAR.cell(row=33, column=3).value='=C31+C20'
			G100FAR.cell(row=33, column=4).value='=D31+D20'
			G100FAR.cell(row=33, column=5).value='=E31+E20'

			for i in range(1, 6):
				G100FAR.cell(row=33, column=i).fill=total

			for i in range(1,5):
				G100FAR.cell(row=33, column=i).font=font2

			G100FAR.cell(row=33, column=5).font=fontRed

			G100FAR.cell(row=36, column=1).value='Working 2: Acculumated Depreciation'
			G100FAR.cell(row=36, column=1).font=workingsblue


			G100FAR.cell(row=38,column=1).value="Synt Account"
			G100FAR.cell(row=38,column=2).value="Description"
			G100FAR.cell(row=38,column=3).value="As per FAR(PBC_FAR)"
			G100FAR.cell(row=38,column=4).value="As per TB(F10_TB)"
			G100FAR.cell(row=38,column=5).value="Check"

			for i in range (1,6):
				G100FAR.cell(row=38, column=i).font=headers
				G100FAR.cell(row=38, column=i).border=solidborder
				G100FAR.cell(row=38, column=i).fill=headersblue

			G100FAR.cell(row=37, column=1).value="'-Depreciation of Intangible Assets-"
			G100FAR.cell(row=37, column=1).font=greenbolditalic

			G100FAR.cell(row=39, column=1).value="2801"
			G100FAR.cell(row=40, column=1).value="2803"
			G100FAR.cell(row=41, column=1).value="2805 & 2808"
			G100FAR.cell(row=42, column=1).value="2806"
			G100FAR.cell(row=43, column=1).value="2807"

			for i in range(38, 44):
				G100FAR.cell(row=i, column=1).alignment=Alignment(horizontal='right')

			G100FAR.cell(row=39, column=2).value="Depreciation of Set-up costs"
			G100FAR.cell(row=40, column=2).value="Depreciation of development costs"
			G100FAR.cell(row=41, column=2).value="Depreciation of licenses and trademarks"
			G100FAR.cell(row=42, column=2).value="Depreciation of intangible assets for mineral resources"
			G100FAR.cell(row=43, column=2).value="Depreciation of goodwill"

			G100FAR.cell(row=39, column=3).value='=SUMIF(PBC_FAR!F:F,"2801",PBC_FAR!D:D)'
			G100FAR.cell(row=40, column=3).value='=SUMIF(PBC_FAR!F:F,"2803",PBC_FAR!D:D)'
			G100FAR.cell(row=41, column=3).value='=SUMIF(PBC_FAR!F:F,"2805",PBC_FAR!D:D)+SUMIF(PBC_FAR!F:F,"2808",PBC_FAR!D:D)'
			G100FAR.cell(row=42, column=3).value='=SUMIF(PBC_FAR!F:F,"2806",PBC_FAR!D:D)'
			G100FAR.cell(row=43, column=3).value='=SUMIF(PBC_FAR!F:F,"2807",PBC_FAR!D:D)'

			G100FAR.cell(row=39, column=4).value='=SUMIF(F10_TB!J:J,"2801",F10_TB!H:H)'
			G100FAR.cell(row=40, column=4).value='=SUMIF(F10_TB!J:J,"2803",F10_TB!H:H)'
			G100FAR.cell(row=41, column=4).value='=SUMIF(F10_TB!J:J,"2805",F10_TB!H:H)+SUMIF(F10_TB!J:J,"2808",F10_TB!H:H)'
			G100FAR.cell(row=42, column=4).value='=SUMIF(F10_TB!J:J,"2806",F10_TB!H:H)'
			G100FAR.cell(row=43, column=4).value='=SUMIF(F10_TB!J:J,"2807",F10_TB!H:H)'

			for i in range(38, 43):
				G100FAR.cell(row=i+1, column=5).value='=IF(C{0}<0,C{0}-D{0},C{0}+D{0})'.format(i+1)

			G100FAR.cell(row=44, column=1).value="Total Depreciation of Intangible Assets"
			G100FAR.cell(row=44, column=3).value="=SUM(C39:C43)"
			G100FAR.cell(row=44, column=4).value="=SUM(D39:D43)"
			G100FAR.cell(row=44, column=5).value="=SUM(E39:E43)"

			for i in range (1,5):
				G100FAR.cell(row=44,column=i).font=font2

			G100FAR.cell(row=44, column=5).font=fontRed

			for i in range (1,6):
				G100FAR.cell(row=44, column=i).border=doubleborder

			for i in range (39,45):
				G100FAR.cell(row=i, column=5).font=fontRedDiff


			G100FAR.cell(row=46, column=1).value="'-Depreciation of Tangible Assets-"
			G100FAR.cell(row=46, column=1).font=greenbolditalic


			G100FAR.cell(row=47,column=1).value="Synt Account"
			G100FAR.cell(row=47,column=2).value="Description"
			G100FAR.cell(row=47,column=3).value="As per FAR(PBC_FAR)"
			G100FAR.cell(row=47,column=4).value="As per TB(F10_TB)"
			G100FAR.cell(row=47,column=5).value="Check"


			G100FAR.cell(row=48, column=1).value="2812"
			G100FAR.cell(row=49, column=1).value="2813"
			G100FAR.cell(row=50, column=1).value="2814"
			G100FAR.cell(row=51, column=1).value="2815"
			G100FAR.cell(row=52, column=1).value="2816"
			G100FAR.cell(row=53, column=1).value="2817"

			for i in range(48, 54):
				G100FAR.cell(row=i, column=1).alignment=Alignment(horizontal='right')

			G100FAR.cell(row=48, column=2).value="Depreciation of buildings"
			G100FAR.cell(row=49, column=2).value="Depreciation of equipments and machines"
			G100FAR.cell(row=50, column=2).value="Depreciation of furniture and office equipments"
			G100FAR.cell(row=51, column=2).value="Depreciation of investment properties"
			G100FAR.cell(row=52, column=2).value="Depreciation of tangible assets for mineral resources"
			G100FAR.cell(row=53, column=2).value="Depreciation for biological assets"

			G100FAR.cell(row=48, column=3).value='=SUMIF(PBC_FAR!F:F,"2812",PBC_FAR!D:D)'
			G100FAR.cell(row=49, column=3).value='=SUMIF(PBC_FAR!F:F,"2813",PBC_FAR!D:D)'
			G100FAR.cell(row=50, column=3).value='=SUMIF(PBC_FAR!F:F,"2814",PBC_FAR!D:D)'
			G100FAR.cell(row=51, column=3).value='=SUMIF(PBC_FAR!F:F,"2815",PBC_FAR!D:D)'
			G100FAR.cell(row=52, column=3).value='=SUMIF(PBC_FAR!F:F,"2816",PBC_FAR!D:D)'
			G100FAR.cell(row=53, column=3).value='=SUMIF(PBC_FAR!F:F,"2817",PBC_FAR!D:D)'

			G100FAR.cell(row=48, column=4).value='=SUMIF(F10_TB!J:J,"2812",F10_TB!H:H)'
			G100FAR.cell(row=49, column=4).value='=SUMIF(F10_TB!J:J,"2813",F10_TB!H:H)'
			G100FAR.cell(row=50, column=4).value='=SUMIF(F10_TB!J:J,"2814",F10_TB!H:H)'
			G100FAR.cell(row=51, column=4).value='=SUMIF(F10_TB!J:J,"2815",F10_TB!H:H)'
			G100FAR.cell(row=52, column=4).value='=SUMIF(F10_TB!J:J,"2816",F10_TB!H:H)'
			G100FAR.cell(row=53, column=4).value='=SUMIF(F10_TB!J:J,"2817",F10_TB!H:H)'

			for i in range(47, 53):
				G100FAR.cell(row=i+1, column=5).value="=IF(C{0}<0,C{0}-D{0},C{0}+D{0})".format(i+1)

			G100FAR.cell(row=54, column=1).value="Total Depreciation of Tangible Assets"
			G100FAR.cell(row=54, column=3).value="=SUM(C48:C53)"
			G100FAR.cell(row=54, column=4).value="=SUM(D48:D53)"
			G100FAR.cell(row=54, column=5).value="=SUM(E48:E53)"


			G100FAR.cell(row=56, column=1).value="Total FA Depreciation "
			G100FAR.cell(row=56, column=3).value="=C54+C44"
			G100FAR.cell(row=56, column=4).value="=D54+D44"
			G100FAR.cell(row=56, column=5).value="=E54+E44"

			for i in range (1,5):
				G100FAR.cell(row=54,column=i).font=font2
				G100FAR.cell(row=56,column=i).font=font2


			G100FAR.cell(row=54, column=5).font=fontRed
			G100FAR.cell(row=56, column=5).font=fontRed

			for i in range (1,6):
				G100FAR.cell(row=54, column=i).border=doubleborder

			for i in range (48,54):
				G100FAR.cell(row=i, column=5).font=fontRedDiff

			for i in range(1, 6):
				G100FAR.cell(row=56, column=i).fill=total


			G100FAR.cell(row=60,column=1).value="Working 3: NBV"
			G100FAR.cell(row=60,column=1).font=workingsblue

			G100FAR.cell(row=62,column=1).value="Account"
			G100FAR.cell(row=62,column=2).value="Description"
			G100FAR.cell(row=62,column=3).value="GBV"

			G100FAR.cell(row=62,column=4).value="Accumulated Depreciation"
			G100FAR.cell(row=62,column=5).value="NBV"


			for i in range (1,6):
				G100FAR.cell(row=62, column=i).font=headers
				G100FAR.cell(row=62, column=i).border=solidborder
				G100FAR.cell(row=62, column=i).fill=headersblue

			for i in range (1,6):
				G100FAR.cell(row=71, column=i).font=headers
				G100FAR.cell(row=71, column=i).border=solidborder
				G100FAR.cell(row=71, column=i).fill=headersblue


			G100FAR.cell(row=61, column=1).value="'-Intangible Assets-"
			G100FAR.cell(row=61, column=1).font=greenbolditalic

			G100FAR.cell(row=63, column=1).value="201"
			G100FAR.cell(row=64, column=1).value="203"
			G100FAR.cell(row=65, column=1).value="205 & 208"
			G100FAR.cell(row=66, column=1).value="206"
			G100FAR.cell(row=67, column=1).value="2071"

			for i in range(63, 68):
				G100FAR.cell(row=i, column=1).alignment=Alignment(horizontal='right')

			G100FAR.cell(row=63, column=2).value="Set-up costs"
			G100FAR.cell(row=64, column=2).value="Development costs"
			G100FAR.cell(row=65, column=2).value="Licenses, Trademarks"
			G100FAR.cell(row=66, column=2).value="Intangible assets for mineral resources"
			G100FAR.cell(row=67, column=2).value="Goodwill"

			G100FAR.cell(row=63, column=3).value="=C15"
			G100FAR.cell(row=64, column=3).value="=C16"
			G100FAR.cell(row=65, column=3).value="=C17"
			G100FAR.cell(row=66, column=3).value="=C18"
			G100FAR.cell(row=67, column=3).value="=C19"

			G100FAR.cell(row=63, column=4).value="=IF(C39<0,C39,-C39)"
			G100FAR.cell(row=64, column=4).value="=IF(C40<0,C40,-C40)"
			G100FAR.cell(row=65, column=4).value="=IF(C41<0,C41,-C41)"
			G100FAR.cell(row=66, column=4).value="=IF(C42<0,C42,-C42)"
			G100FAR.cell(row=67, column=4).value="=IF(C43<0,C43,-C43)"

			for i in range(62, 67):
				G100FAR.cell(row=i+1, column=5).value="=C{0}+D{0}".format(i+1)


			G100FAR.cell(row=68, column=1).value="Total Intangible Assets"
			G100FAR.cell(row=68, column=3).value="=SUM(C63:C67)"
			G100FAR.cell(row=68, column=4).value="=SUM(D63:D67)"
			G100FAR.cell(row=68, column=5).value="=SUM(E63:E67)"


			for i in range (1,6):
				G100FAR.cell(row=68,column=i).font=font2

			for i in range (1,6):
				G100FAR.cell(row=67, column=i).border=doubleborder


			G100FAR.cell(row=70, column=1).value="'-Tangible Assets-"
			G100FAR.cell(row=70, column=1).font=greenbolditalic

			G100FAR.cell(row=71,column=1).value="Account"
			G100FAR.cell(row=71,column=2).value="Description"
			G100FAR.cell(row=71,column=3).value="GBV"

			G100FAR.cell(row=71,column=4).value="Accumulated Depreciation"
			G100FAR.cell(row=71,column=5).value="NBV"


			G100FAR.cell(row=72, column=1).value="211"
			G100FAR.cell(row=73, column=1).value="212"
			G100FAR.cell(row=74, column=1).value="213 & 223"
			G100FAR.cell(row=75, column=1).value="214 & 224"
			G100FAR.cell(row=76, column=1).value="215"
			G100FAR.cell(row=77, column=1).value="216"
			G100FAR.cell(row=78, column=1).value="217 & 227"

			for i in range(72, 79):
				G100FAR.cell(row=i, column=1).alignment=Alignment(horizontal='right')

			G100FAR.cell(row=72, column=2).value="Land"
			G100FAR.cell(row=73, column=2).value="Buildings"
			G100FAR.cell(row=74, column=2).value="Equipments and Machines"
			G100FAR.cell(row=75, column=2).value="Furniture, office equipments"
			G100FAR.cell(row=76, column=2).value="Investment properties"
			G100FAR.cell(row=77, column=2).value="Tangible assets for mineral resources"
			G100FAR.cell(row=78, column=2).value="Biological assets"

			G100FAR.cell(row=72, column=3).value="=C24"
			G100FAR.cell(row=73, column=3).value="=C25"
			G100FAR.cell(row=74, column=3).value="=C26"
			G100FAR.cell(row=75, column=3).value="=C27"
			G100FAR.cell(row=76, column=3).value="=C28"
			G100FAR.cell(row=77, column=3).value="=C29"
			G100FAR.cell(row=78, column=3).value="=C30"

			G100FAR.cell(row=72, column=4).value=0
			G100FAR.cell(row=73, column=4).value="=IF(C48<0,C48,-C48)"
			G100FAR.cell(row=74, column=4).value="=IF(C49<0,C49,-C49)"
			G100FAR.cell(row=75, column=4).value="=IF(C50<0,C50,-C50)"
			G100FAR.cell(row=76, column=4).value="=IF(C51<0,C51,-C51)"
			G100FAR.cell(row=77, column=4).value="=IF(C52<0,C52,-C52)"
			G100FAR.cell(row=78, column=4).value="=IF(C53<0,C53,-C53)"


			for i in range(72, 78):
				G100FAR.cell(row=i+1, column=5).value="=C{0}+D{0}".format(i+1)

			G100FAR.cell(row=79, column=1).value="Total Tangible Assets"
			G100FAR.cell(row=79, column=3).value="=SUM(C72:C78)"
			G100FAR.cell(row=79, column=4).value="=SUM(D72:D78)"
			G100FAR.cell(row=79, column=5).value="=SUM(E72:E78)"

			G100FAR.cell(row=81, column=1).value="Total FA NBV"
			G100FAR.cell(row=81, column=3).value="=SUM(C68,C79)"
			G100FAR.cell(row=81, column=4).value="=SUM(D68,D79)"
			G100FAR.cell(row=81, column=5).value="=SUM(E68,E79)"

			for i in range (1,6):
				G100FAR.cell(row=79,column=i).font=font2
				G100FAR.cell(row=81,column=i).font=font2


			for i in range (1,6):
				G100FAR.cell(row=78, column=i).border=doubleborder


			for i in range(1, 6):
				G100FAR.cell(row=81, column=i).fill=total

			if (len(ChargeFAR)==0):
				print("list is empty")
			else:

				G100FAR.cell(row=83, column=1).value="Working 4: Depreciation charge reconciliation"
				G100FAR.cell(row=83, column=1).font=workingsblue

				G100FAR.cell(row=84, column=1).value="As per FAR(PBC_FAR)"
				G100FAR.cell(row=84, column=2).value="As per TB(PBC_TB)"
				G100FAR.cell(row=84, column=3).value="Check"

				for i in range(1, 4):
					G100FAR.cell(row=84, column=i).font=headers
					G100FAR.cell(row=84, column=i).fill=headersblue
					G100FAR.cell(row=84, column=i).border=solidborder

				G100FAR.cell(row=85, column=1).value="=SUM(PBC_FAR!I:I)"
				G100FAR.cell(row=85, column=2).value='=SUMIF(F10_TB!J:J,"6811",F10_TB!H:H)'
				G100FAR.cell(row=85, column=3).value='=IF(A85<0,B85+A85,A85-B85)'

			G100FAR.sheet_view.showGridLines = False

		if(val2==1):
			  G20=FAR.create_sheet("G20 Depr Recomp")


			  G20.column_dimensions['A'].width=15
			  G20.column_dimensions['B'].width=15
			  G20.column_dimensions['C'].width=30
			  G20.column_dimensions['D'].width=15
			  G20.column_dimensions['E'].width=15
			  G20.column_dimensions['F'].width=25
			  G20.column_dimensions['G'].width=25
			  G20.column_dimensions['H'].width=15
			  G20.column_dimensions['I'].width=20
			  G20.column_dimensions['J'].width=15
			  G20.column_dimensions['K'].width=15
			  G20.column_dimensions['L'].width=15
			  G20.column_dimensions['M'].width=15
			  G20.column_dimensions['N'].width=25
			  G20.column_dimensions['O'].width=15
			  G20.column_dimensions['P'].width=35
			  G20.column_dimensions['Q'].width=25
			  G20.column_dimensions['R'].width=15
			  G20.column_dimensions['S'].width=25
			  G20.column_dimensions['T'].width=15
			  G20.column_dimensions['U'].width=15
			  G20.column_dimensions['V'].width=25
			  G20.column_dimensions['W'].width=15
			  G20.column_dimensions['X'].width=25


			  G20.cell(row=1, column=1).value="Client:"
			  G20.cell(row=1, column=1).font=font2
			  G20.cell(row=1, column=2).value=namec
			  G20.cell(row=1, column=2).font=font2
			  G20.cell(row=2, column=1).value="Period end:"
			  G20.cell(row=2, column=1).font=font2
			  G20.cell(row=2, column=2).value=ant
			  G20.cell(row=2, column=2).number_format='mm/dd/yyyy'
			  G20.cell(row=2, column=2).font=font2
			  postdate = str(postdate)
			  anp = postdate[0]+postdate[1]+postdate[2]+postdate[3]
			  lunap = postdate[5]+postdate[6]
			  ziip = postdate[8]+postdate[9]

			  yearentry = str(yearentry)
			  print(yearentry)
			  an = yearentry[0]+yearentry[1]+yearentry[2]+yearentry[3]
			  luna = yearentry[5]+yearentry[6]
			  zi = yearentry[8]+yearentry[9]
			  an = int(an)
			  zi = int(zi)
			  luna=int(luna)

			  #yearentry=datetime.date(int(an),int(luna),int(zi))
			  #print(type(yearentry))
			  #yearentry=str(yearentry)
			  ##yearentry = yearentry.strftime('%m/%d/%Y')
			  ##yearentry = datetime.datetime.strptime(yearentry,'%m/%d/%Y')
			  #yearentry = yearentry.date('%m/%d/%Y')
			  #print(type(yearentry))
			  #print("qwjdnqwibdqw", yearentry)

			  G20.cell(row=1, column=8).value="Prepared by:"
			  G20.cell(row=1, column=8).font=font2

			  G20.cell(row=2, column=8).value="Date:"
			  G20.cell(row=2, column=8).font=font2
			  G20.cell(row=2, column=9).value=datetime.datetime.now().date()
			  G20.cell(row=2, column=9).number_format="mm/dd/yyyy"
			  G20.cell(row=2, column=9).alignment = Alignment (horizontal='left')
			  G20.cell(row=2, column=9).font=font1


			  G20.cell(row=3, column=8).value="Ref:"
			  G20.cell(row=3, column=8).font=font2
			  G20.cell(row=3, column=9).value="G20"
			  G20.cell(row=3, column=9).font=fontRed

			  for i in range(1,4):
				  G20.cell(row=i, column=8).alignment = Alignment(horizontal='right')

			  G20.cell(row=4, column=4).value="Tangible and Intangible Assets Depreciation Recomputation"
			  G20.cell(row=4, column=4).font=font2

			  G20.cell(row=7, column=1).value="Procedure:"
			  G20.cell(row=7, column=1).alignment=Alignment(horizontal='right')
			  G20.cell(row=7, column=2).value="Obtain the Fixed asset Register and list of tangible and intangible assets and detail of disposals for the period and recompute the depreciation charge in the period and accumulated depreciation at period end."
			  G20.cell(row=9, column=1).value="Work done:"
			  G20.cell(row=9, column=1).alignment=Alignment(horizontal='right')
			  G20.cell(row=8, column=2).value="Agree the recomputed depreciation charge and accumulated depreciation with the FAR and Disposal list and with TB."

			  G20.cell(row=7, column=2).font=font1
			  G20.cell(row=9, column=1).font=font1
			  G20.cell(row=8, column=2).font=font1

			  G20.cell(row=9, column=2).value="Please see the work below:"
			  G20.cell(row=7, column=1).font=font2
			  G20.cell(row=9, column=1).font=font2

			  G20.cell(row=9, column=2).font=font1


			  G20.cell(row=40,column=1).value="Synt"
			  G20.cell(row=40,column=2).value="Account"
			  G20.cell(row=40,column=3).value="Item"
			  G20.cell(row=40, column=4).value="Description"
			  G20.cell(row=40,column=5).value="PIF Year"
			  G20.cell(row=40,column=6).value="PIF Date"
			  G20.cell(row=40,column=7).value="Disposal Date"
			  G20.cell(row=40,column=8).value="UL in months"
			  G20.cell(row=40,column=9).value="GBV"
			  G20.cell(row=40,column=10).value="Depr Charge"
			  G20.cell(row=40,column=11).value="Acc Depreciation"


			  for i in range (1,12):
				  G20.cell(row=40, column=i).font=headers
				  G20.cell(row=40, column=i).border=solidborder
				  G20.cell(row=40, column=i).fill=headersblue


			  "Adjust Column Width---------------------------------------------------------------"

			  G20.cell(row=11, column=4).value="As per Client(FAR and Disposal list)"
			  G20.cell(row=11, column=4).font=font2

			  G20.cell(row=11, column=2).value="Summary:"
			  G20.cell(row=11, column=2).font=fontitalic

			  G20.cell(row=12, column=2).value="Account"
			  G20.cell(row=12, column=3).value="Description"
			  G20.cell(row=12, column=4).value="GBV (FAR)"
			  G20.cell(row=12, column=5).value="Depreciation charge (FAR and Disposal list)"
			  G20.cell(row=12, column=6).value="Accumulated depreciation (FAR)"
			  G20.cell(row=12, column=7).value="Accumulated depreciation (Disposals)"



			  for i in range(2, 8):
				  G20.cell(row=12, column=i).font=headers
				  G20.cell(row=12, column=i).fill=headersblue
				  G20.cell(row=12, column=i).border=solidborder
			  G20.cell(row=13, column=7).border=solidbordersus

			  G20.cell(row=13, column=2).value=201
			  G20.cell(row=14, column=2).value=203
			  G20.cell(row=15, column=2).value=205
			  G20.cell(row=16, column=2).value=208
			  G20.cell(row=17, column=2).value=206
			  G20.cell(row=18, column=2).value=2071


			  G20.cell(row=13, column=3).value="Set-up costs"
			  G20.cell(row=14, column=3).value="Development costs"
			  G20.cell(row=15, column=3).value="Licenses, Trademarks"
			  G20.cell(row=16, column=3).value="Other Intangibles"
			  G20.cell(row=17, column=3).value="Intangible assets for mineral resources"
			  G20.cell(row=18, column=3).value="Goodwill"

			  G20.cell(row=13, column=4).value='=SUMIF(A:A,"201",I:I)'
			  G20.cell(row=14, column=4).value='=SUMIF(A:A,"203",I:I)'
			  G20.cell(row=15, column=4).value='=SUMIF(A:A,"205",I:I)'
			  G20.cell(row=16, column=4).value='=SUMIF(A:A,"208",I:I)'
			  G20.cell(row=17, column=4).value='=SUMIF(A:A,"206",I:I)'
			  G20.cell(row=18, column=4).value='=SUMIF(A:A,"207",I:I)'

			  G20.cell(row=13, column=5).value='=SUMIF(A:A,"201",J:J)+SUMIF(A:A,"201D",J:J)'
			  G20.cell(row=14, column=5).value='=SUMIF(A:A,"203",J:J)+SUMIF(A:A,"203D",J:J)'
			  G20.cell(row=15, column=5).value='=SUMIF(A:A,"205",J:J)+SUMIF(A:A,"205D",J:J)'
			  G20.cell(row=16, column=5).value='=SUMIF(A:A,"208",J:J)+SUMIF(A:A,"208D",J:J)'
			  G20.cell(row=17, column=5).value='=SUMIF(A:A,"206",J:J)+SUMIF(A:A,"206D",J:J)'
			  G20.cell(row=18, column=5).value='=SUMIF(A:A,"207",J:J)+SUMIF(A:A,"207D",J:J)'

			  G20.cell(row=13, column=6).value='=SUMIF(A:A,"201",K:K)'
			  G20.cell(row=14, column=6).value='=SUMIF(A:A,"203",K:K)'
			  G20.cell(row=15, column=6).value='=SUMIF(A:A,"205",K:K)'
			  G20.cell(row=16, column=6).value='=SUMIF(A:A,"208",K:K)'
			  G20.cell(row=17, column=6).value='=SUMIF(A:A,"206",K:K)'
			  G20.cell(row=18, column=6).value='=SUMIF(A:A,"207",K:K)'

			  G20.cell(row=13, column=7).value='=SUMIF(A:A,"201D",K:K)'
			  G20.cell(row=14, column=7).value='=SUMIF(A:A,"203D",K:K)'
			  G20.cell(row=15, column=7).value='=SUMIF(A:A,"205D",K:K)'
			  G20.cell(row=16, column=7).value='=SUMIF(A:A,"208D",K:K)'
			  G20.cell(row=17, column=7).value='=SUMIF(A:A,"206D",K:K)'
			  G20.cell(row=18, column=7).value='=SUMIF(A:A,"207D",K:K)'

			  G20.cell(row=19, column=2).value="Total Intangibles Assets"
			  G20.cell(row=19, column=4).value="=SUM(D13:D18)"
			  G20.cell(row=19, column=5).value="=SUM(E13:E18)"
			  G20.cell(row=19, column=6).value="=SUM(F13:F18)"
			  G20.cell(row=19, column=7).value="=SUM(G13:G18)"


			  for i in range(2,8):
				  G20.cell(row=19, column=i).font=font2

			  #G20.cell(row=20, column=2).value='=IF(SUMIF(PBC_FAR!G:G,"211",PBC_FAR!E:E)+SUMIF(PBC_FAR!G:G,"211",PBC_FAR!J:J)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!I:I)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!J:J)=0,"","211")'
			  G20.cell(row=21, column=2).value=212
			  G20.cell(row=22, column=2).value=213
			  G20.cell(row=23, column=2).value=214
			  G20.cell(row=24, column=2).value=215
			  G20.cell(row=25, column=2).value=216
			  G20.cell(row=26, column=2).value=217

			  #G20.cell(row=20, column=3).value='=IF(SUMIF(PBC_FAR!G:G,"211",PBC_FAR!E:E)+SUMIF(PBC_FAR!G:G,"211",PBC_FAR!J:J)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!I:I)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!J:J)=0,"","Land")'
			  G20.cell(row=21, column=3).value="Buildings"
			  G20.cell(row=22, column=3).value="Equipments and Machines"
			  G20.cell(row=23, column=3).value="Furniture, office equipments"
			  G20.cell(row=24, column=3).value="Investment properties"
			  G20.cell(row=25, column=3).value="Tangible assets for mineral resources"
			  G20.cell(row=26, column=3).value="Biological assets"


			  #G20.cell(row=20, column=4).value='=IF(SUMIF(PBC_FAR!G:G,"211",PBC_FAR!E:E)+SUMIF(PBC_FAR!G:G,"211",PBC_FAR!J:J)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!I:I)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!J:J)=0,"",SUMIF(A:A,"211",I:I))'
			  G20.cell(row=21, column=4).value='=SUMIF(A:A,"212",I:I)'
			  G20.cell(row=22, column=4).value='=SUMIF(A:A,"213",I:I)'
			  G20.cell(row=23, column=4).value='=SUMIF(A:A,"214",I:I)'
			  G20.cell(row=24, column=4).value='=SUMIF(A:A,"215",I:I)'
			  G20.cell(row=25, column=4).value='=SUMIF(A:A,"216",I:I)'
			  G20.cell(row=26, column=4).value='=SUMIF(A:A,"217",I:I)'

			  #G20.cell(row=20, column=5).value='=IF(SUMIF(PBC_FAR!G:G,"211",PBC_FAR!E:E)+SUMIF(PBC_FAR!G:G,"211",PBC_FAR!J:J)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!I:I)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!J:J)=0,"",SUMIF(A:A,"211",J:J)+SUMIF(A:A,"211D",J:J))'
			  G20.cell(row=21, column=5).value='=SUMIF(A:A,"212",J:J)+SUMIF(A:A,"212D",J:J)'
			  G20.cell(row=22, column=5).value='=SUMIF(A:A,"213",J:J)+SUMIF(A:A,"213D",J:J)'
			  G20.cell(row=23, column=5).value='=SUMIF(A:A,"214",J:J)+SUMIF(A:A,"214D",J:J)'
			  G20.cell(row=24, column=5).value='=SUMIF(A:A,"215",J:J)+SUMIF(A:A,"215D",J:J)'
			  G20.cell(row=25, column=5).value='=SUMIF(A:A,"216",J:J)+SUMIF(A:A,"216D",J:J)'
			  G20.cell(row=26, column=5).value='=SUMIF(A:A,"217",J:J)+SUMIF(A:A,"217D",J:J)'

			  #G20.cell(row=20, column=6).value='=IF(SUMIF(PBC_FAR!G:G,"211",PBC_FAR!E:E)+SUMIF(PBC_FAR!G:G,"211",PBC_FAR!J:J)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!I:I)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!J:J)=0,"",SUMIF(A:A,"211",K:K))'
			  G20.cell(row=21, column=6).value='=SUMIF(A:A,"212",K:K)'
			  G20.cell(row=22, column=6).value='=SUMIF(A:A,"213",K:K)'
			  G20.cell(row=23, column=6).value='=SUMIF(A:A,"214",K:K)'
			  G20.cell(row=24, column=6).value='=SUMIF(A:A,"215",K:K)'
			  G20.cell(row=25, column=6).value='=SUMIF(A:A,"216",K:K)'
			  G20.cell(row=26, column=6).value='=SUMIF(A:A,"217",K:K)'

			  #G20.cell(row=20, column=7).value='=IF(SUMIF(PBC_FAR!G:G,"211",PBC_FAR!E:E)+SUMIF(PBC_FAR!G:G,"211",PBC_FAR!J:J)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!I:I)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!J:J)=0,"",SUMIF(A:A,"211D",K:K))'
			  G20.cell(row=21, column=7).value='=SUMIF(A:A,"212D",K:K)'
			  G20.cell(row=22, column=7).value='=SUMIF(A:A,"213D",K:K)'
			  G20.cell(row=23, column=7).value='=SUMIF(A:A,"214D",K:K)'
			  G20.cell(row=24, column=7).value='=SUMIF(A:A,"215D",K:K)'
			  G20.cell(row=25, column=7).value='=SUMIF(A:A,"216D",K:K)'
			  G20.cell(row=26, column=7).value='=SUMIF(A:A,"217D",K:K)'

			  G20.cell(row=27, column=2).value="Total Tangible Assets"

			  G20.cell(row=27, column=4).value="=SUM(D20:D26)"
			  G20.cell(row=27, column=5).value="=SUM(E20:E26)"
			  G20.cell(row=27, column=6).value="=SUM(F20:F26)"
			  G20.cell(row=27, column=7).value="=SUM(G20:G26)"

			  G20.cell(row=29, column=2).value="Total Tangible and Intangible Assets"
			  G20.cell(row=29, column=4).value="=D19+D27"
			  G20.cell(row=29, column=5).value='=E19+E27'
			  G20.cell(row=29, column=6).value='=F19+F27'

			  G20.cell(row=29, column=7).value="=G19+G27"

			  G20.cell(row=30,column=2).value="As per TB"
			  G20.cell(row=30,column=4).value='=SUMIF(F10_TB!B:B,"201",F10_TB!H:H)+SUMIF(F10_TB!B:B,"203",F10_TB!H:H)+SUMIF(F10_TB!B:B,"205",F10_TB!H:H)+SUMIF(F10_TB!B:B,"208",F10_TB!H:H)+SUMIF(F10_TB!B:B,"206",F10_TB!H:H)+SUMIF(F10_TB!B:B,"207",F10_TB!H:H)+SUMIF(F10_TB!B:B,"212",F10_TB!H:H)+SUMIF(F10_TB!B:B,"213",F10_TB!H:H)+SUMIF(F10_TB!B:B,"214",F10_TB!H:H)+SUMIF(F10_TB!B:B,"215",F10_TB!H:H)+SUMIF(F10_TB!B:B,"216",F10_TB!H:H)+SUMIF(F10_TB!B:B,"217",F10_TB!H:H)'
			  G20.cell(row=30,column=5).value="=Q14"
			  G20.cell(row=30,column=6).value='=SUMIF(F10_TB!B:B,"280",F10_TB!H:H)+SUMIF(F10_TB!B:B,"281",F10_TB!H:H)'

			  G20.cell(row=31,column=2).value="Check"
			  G20.cell(row=31,column=4).value="=D29-D30"
			  G20.cell(row=31,column=5).value="=IF(E29<0,E29+E30,E29-E30)"
			  G20.cell(row=31,column=6).value="=IF(F29<0,F29-F30,F29+F30)"



			  G20.cell(row=21,column=13).font=font2

			  for i in range(2, 8):
				  G20.cell(row=27, column=i).font=font2
				  G20.cell(row=29, column=i).font=font2
				  G20.cell(row=30, column=i).font=font2
				  G20.cell(row=31, column=i).font=fontRed
				  G20.cell(row=29,column=i).number_format='#,##0_);(#,##0)'
				  G20.cell(row=30,column=i).number_format='#,##0_);(#,##0)'
				  G20.cell(row=31,column=i).number_format='#,##0_);(#,##0)'

			  for i in range(2, 8):
				  G20.cell(row=18, column=i).border=doubleborder
				  G20.cell(row=26, column=i).border=doubleborder


			  G20.cell(row=33, column=2).value="As per Disposal list"
			  G20.cell(row=33, column=2).font=font2
			  G20.cell(row=34, column=2).value="Check"
			  G20.cell(row=34, column=2).font=fontRed
			  G20.cell(row=36, column=1).value="Recomputation:"
			  G20.cell(row=36, column=1).font=fontitalic


			  G20.cell(row=33, column=7).value="=SUM(PBC_Disposals!J:J)"
			  G20.cell(row=33, column=7).font=font2
			  G20.cell(row=34, column=7).value="=G29-G33"
			  G20.cell(row=34, column=7).font=fontRed


			  G20.cell(row=40,column=13).value="Cost"
			  G20.cell(row=40,column=14).value="UL"

			  G20.cell(row=40,column=15).value="Start Depr"
			  G20.cell(row=40,column=16).value="End Depr"
			  G20.cell(row=40,column=17).value="Month in year"
			  G20.cell(row=40,column=18).value="Monthly depr"
			  G20.cell(row=40,column=19).value="Depreciation charge"
			  G20.cell(row=40,column=20).value="Diff"

			  G20.cell(row=40,column=22).value="Accumulated depr"
			  G20.cell(row=40,column=23).value="Diff"
			  G20.cell(row=40,column=24).value="Check if accumulated depr < GBV"

			  for i in range (13,21):
				  G20.cell(row=40, column=i).font=headers
				  G20.cell(row=40, column=i).border=solidborder
				  G20.cell(row=40, column=i).fill=headerspurple

			  for i in range (22,25):
				  G20.cell(row=40, column=i).font=headers
				  G20.cell(row=40, column=i).border=solidborder
				  G20.cell(row=40, column=i).fill=headerspurple

			  "noul tabel-------------------------------------------------------------------------------------------------"

			  G20.cell(row=11, column=9).value="As per Nexia"
			  G20.cell(row=11, column=9).font=font2
			  G20.cell(row=12, column=9).value="Depreciation charge"
			  G20.cell(row=12, column=10).value="Diff"
			  G20.cell(row=12, column=10).font=fontRed
			  G20.cell(row=12, column=11).value="Accumulated depreciation"
			  G20.cell(row=12, column=12).value="Diff"
			  G20.cell(row=12, column=12).font=fontRed
			  G20.cell(row=12, column=13).value="Accumulated Depreciation (Disposals)"
			  G20.cell(row=12, column=14).value="Diff"

			  for i in range (9,15):
				  G20.cell(row=12, column=i).font=headers
				  G20.cell(row=12, column=i).border=solidborder
				  G20.cell(row=12, column=i).fill=headersblue

			  for i in range(9, 15):
				  G20.cell(row=18, column=i).border=doubleborder

			  for i in range(9, 15):
				  G20.cell(row=26, column=i).border=doubleborder



			  G20.cell(row=13, column=9).value='=SUMIF(A:A,"201",S:S)+SUMIF(A:A,"201D",S:S)'
			  G20.cell(row=14, column=9).value='=SUMIF(A:A,"203",S:S)+SUMIF(A:A,"203D",S:S)'
			  G20.cell(row=15, column=9).value='=SUMIF(A:A,"205",S:S)+SUMIF(A:A,"205D",S:S)'
			  G20.cell(row=16, column=9).value='=SUMIF(A:A,"208",S:S)+SUMIF(A:A,"208D",S:S)'
			  G20.cell(row=17, column=9).value='=SUMIF(A:A,"206",S:S)+SUMIF(A:A,"206D",S:S)'
			  G20.cell(row=18, column=9).value='=SUMIF(A:A,"207",S:S)+SUMIF(A:A,"207D",S:S)'

			  G20.cell(row=13, column=11).value='=SUMIF(A:A,"201",V:V)'
			  G20.cell(row=14, column=11).value='=SUMIF(A:A,"203",V:V)'
			  G20.cell(row=15, column=11).value='=SUMIF(A:A,"205",V:V)'
			  G20.cell(row=16, column=11).value='=SUMIF(A:A,"208",V:V)'
			  G20.cell(row=17, column=11).value='=SUMIF(A:A,"206",V:V)'
			  G20.cell(row=18, column=11).value='=SUMIF(A:A,"207",V:V)'

			  G20.cell(row=13, column=13).value='=SUMIF(A:A,"201D",V:V)'
			  G20.cell(row=14, column=13).value='=SUMIF(A:A,"203D",V:V)'
			  G20.cell(row=15, column=13).value='=SUMIF(A:A,"205D",V:V)'
			  G20.cell(row=16, column=13).value='=SUMIF(A:A,"208D",V:V)'
			  G20.cell(row=17, column=13).value='=SUMIF(A:A,"206D",V:V)'
			  G20.cell(row=18, column=13).value='=SUMIF(A:A,"207D",V:V)'

			  #G20.cell(row=20, column=9).value='=IF(SUMIF(PBC_FAR!G:G,"211",PBC_FAR!E:E)+SUMIF(PBC_FAR!G:G,"211",PBC_FAR!J:J)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!I:I)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!J:J)=0,"",SUMIF(A:A,"211",S:S)+SUMIF(A:A,"211D",S:S))'
			  G20.cell(row=21, column=9).value='=SUMIF(A:A,"212",S:S)+SUMIF(A:A,"212D",S:S)'
			  G20.cell(row=22, column=9).value='=SUMIF(A:A,"213",S:S)+SUMIF(A:A,"213D",S:S)'
			  G20.cell(row=23, column=9).value='=SUMIF(A:A,"214",S:S)+SUMIF(A:A,"214D",S:S)'
			  G20.cell(row=24, column=9).value='=SUMIF(A:A,"215",S:S)+SUMIF(A:A,"215D",S:S)'
			  G20.cell(row=25, column=9).value='=SUMIF(A:A,"216",S:S)+SUMIF(A:A,"216D",S:S)'
			  G20.cell(row=26, column=9).value='=SUMIF(A:A,"217",S:S)+SUMIF(A:A,"217D",S:S)'

			  #G20.cell(row=20, column=13).value='=IF(SUMIF(PBC_FAR!G:G,"211",PBC_FAR!E:E)+SUMIF(PBC_FAR!G:G,"211",PBC_FAR!J:J)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!I:I)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!J:J)=0,"",SUMIF(A:A,"211D",V:V))'
			  G20.cell(row=21, column=13).value='=SUMIF(A:A,"212D",V:V)'
			  G20.cell(row=22, column=13).value='=SUMIF(A:A,"213D",V:V)'
			  G20.cell(row=23, column=13).value='=SUMIF(A:A,"214D",V:V)'
			  G20.cell(row=24, column=13).value='=SUMIF(A:A,"215D",V:V)'
			  G20.cell(row=25, column=13).value='=SUMIF(A:A,"216D",V:V)'
			  G20.cell(row=26, column=13).value='=SUMIF(A:A,"217D ",V:V)'

			  #G20.cell(row=20, column=11).value='=IF(SUMIF(PBC_FAR!G:G,"211",PBC_FAR!E:E)+SUMIF(PBC_FAR!G:G,"211",PBC_FAR!J:J)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!I:I)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!J:J)=0,"",SUMIF(A:A,"211",V:V))'
			  G20.cell(row=21, column=11).value='=SUMIF(A:A,"212",V:V)'
			  G20.cell(row=22, column=11).value='=SUMIF(A:A,"213",V:V)'
			  G20.cell(row=23, column=11).value='=SUMIF(A:A,"214",V:V)'
			  G20.cell(row=24, column=11).value='=SUMIF(A:A,"215",V:V)'
			  G20.cell(row=25, column=11).value='=SUMIF(A:A,"216",V:V)'
			  G20.cell(row=26, column=11).value='=SUMIF(A:A,"217",V:V)'

			  #G20.cell(row=20, column=10).value='=IF(SUMIF(PBC_FAR!G:G,"211",PBC_FAR!E:E)+SUMIF(PBC_FAR!G:G,"211",PBC_FAR!J:J)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!I:I)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!J:J)=0,"",I20-E20)'
			  #G20.cell(row=20, column=12).value='=IF(SUMIF(PBC_FAR!G:G,"211",PBC_FAR!E:E)+SUMIF(PBC_FAR!G:G,"211",PBC_FAR!J:J)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!I:I)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!J:J)=0,"",K20-F20)'
			  #G20.cell(row=20, column=14).value='=IF(SUMIF(PBC_FAR!G:G,"211",PBC_FAR!E:E)+SUMIF(PBC_FAR!G:G,"211",PBC_FAR!J:J)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!I:I)+SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!J:J)=0,"",M20-G20)'

			  "PARTEA NUUA"
			  G20.cell(row=19, column=13).value="=SUM(M13:M18)"
			  G20.cell(row=27, column=13).value="=SUM(M20:M26)"
			  G20.cell(row=29, column=13).value="=M19+M27"
			  G20.cell(row=33, column=13).value="=SUM(PBC_Disposals!J:J)"
			  G20.cell(row=34, column=13).value="=M29-M33"

			  for i in range(12, 18):
				  G20.cell(row=i+1, column=14).value='=M{0}-G{0}'.format(i+1)

			  for i in range(20, 26):
				  G20.cell(row=i+1, column=14).value='=M{0}-G{0}'.format(i+1)

			  G20.cell(row=19, column=14).value="=SUM(N13:N18)"
			  G20.cell(row=27, column=14).value="=SUM(N20:N26)"
			  G20.cell(row=29, column=14).value="=M29-G29"

			  for i in range(13, 19):
				  for j in range(13, 14):
					  G20.cell(row=i, column=j).font=font1

			  for i in range(21, 27):
				  for j in range(13, 14):
					  G20.cell(row=i, column=j).font=font1

			  for i in range(13, 35):
				  for j in range(13, 15):
					  G20.cell(row=i, column=j).number_format='#,##0_);(#,##0)'

			  G20.cell(row=19, column=13).font=font2
			  G20.cell(row=27, column=13).font=font2
			  G20.cell(row=29, column=13).font=font2
			  G20.cell(row=33, column=13).font=font2
			  G20.cell(row=34, column=13).font=fontRed

			  for i in range(13, 19):
				  G20.cell(row=i, column=14).font=fontRedDiff

			  G20.cell(row=19, column=19).font=fontRed


			  for i in range(21, 27):
				  G20.cell(row=i, column=14).font=fontRedDiff

			  G20.cell(row=27, column=14).font=fontRed
			  G20.cell(row=29, column=14).font=fontRed
			  "PARTEA NOUA"

			  G20.cell(row=19, column=9).value='=SUM(I13:I18)'
			  G20.cell(row=27, column=9).value='=SUM(I20:I26)'

			  G20.cell(row=19, column=10).value='=SUM(J13:J18)'
			  G20.cell(row=27, column=10).value='=SUM(J20:J26)'

			  G20.cell(row=19, column=11).value='=SUM(K13:K18)'
			  G20.cell(row=27, column=11).value='=SUM(K20:K26)'

			  G20.cell(row=19, column=12).value='=SUM(L13:L18)'
			  G20.cell(row=27, column=12).value='=SUM(L20:L26)'

			  for i in range(12, 18):
				  G20.cell(row=i+1, column=10).value='=if($E$29=0,"N/A – no depr expense in FAR PBC",I{0}-E{0})'.format(i+1)
				  G20.cell(row=i+1, column=10).font=fontRedDiff

			  for i in range(20, 26):
				  G20.cell(row=i+1, column=10).value='=if($E$29=0,"N/A – no depr expense in FAR PBC",I{0}-E{0})'.format(i+1)
				  G20.cell(row=i+1, column=10).font=fontRedDiff

			  for i in range(12, 18):
				  G20.cell(row=i+1, column=12).value='=if(F{0}>=0,K{0}-F{0},K{0}+F{0})'.format(i+1)
				  G20.cell(row=i+1, column=12).font=fontRedDiff

			  for i in range(20, 26):
				  G20.cell(row=i+1, column=12).value='=if(F{0}>=0,K{0}-F{0},K{0}+F{0})'.format(i+1)
				  G20.cell(row=i+1, column=12).font=fontRedDiff

			  G20.cell(row=19, column=10).font=fontRed
			  G20.cell(row=19, column=12).font=fontRed

			  G20.cell(row=27, column=10).font=fontRed
			  G20.cell(row=27, column=12).font=fontRed

			  for i in range(27, 30):
				  G20.cell(row=i, column=9).font=font2

			  for i in range(27, 31):
				  G20.cell(row=i, column=11).font=font2

			  for i in range(13, 20):
				  for j in range(9, 13):
					  G20.cell(row=i, column=j).number_format='#,##0_);(#,##0)'

			  for i in range(21, 32):
				  for j in range(9, 13):
					  G20.cell(row=i, column=j).number_format='#,##0_);(#,##0)'

			  for i in range(13, 20):
				  G20.cell(row=i, column=9).font=font1

			  for i in range(13, 20):
				  G20.cell(row=i, column=11).font=font1

			  for i in range(21, 27):
				  G20.cell(row=i, column=9).font=font1

			  for i in range(21, 27):
				  G20.cell(row=i, column=11).font=font1

			  G20.cell(row=29, column=9).value='=I19+I27'
			  G20.cell(row=29, column=11).value='=K19+K27'

			  G20.cell(row=29, column=10).value='=E29-I29'
			  G20.cell(row=29, column=12).value='=F29-K29'

			  G20.cell(row=30, column=9).value='=Q14'
			  G20.cell(row=30, column=11).value='=SUMIF(F10_TB!B:B,"280",F10_TB!H:H)+SUMIF(F10_TB!B:B,"281",F10_TB!H:H)'

			  G20.cell(row=31, column=9).value='=IF(I29<0,I29+I30,I29-I30)'
			  G20.cell(row=31, column=9).font=fontRed
			  G20.cell(row=31, column=11).value='=IF(K29<0,K29-K30,K29+K30)'
			  G20.cell(row=31, column=11).font=fontRed

			  G20.cell(row=29, column=10).font=fontRed
			  G20.cell(row=29, column=12).font=fontRed

			  G20.cell(row=30, column=9).border=solidborder
			  G20.cell(row=30, column=11).border=solidborder

			  for i in range(4, 8):
				  G20.cell(row=30, column=i).border=solidborder
				  G20.cell(row=33, column=i).border=solidborder
			  for i in range(9, 14):
				  G20 .cell(row=33, column=i).border=solidborder
			  G20.cell(row=29 ,column=15).value='=IF(ISERROR(L29),"One of the FAR PIF Dates exceed the period end date, therefore the fixed assets should not be presented in FAR","")'

			  G20.cell(row=12, column=16).value="Reconciliation #28x vs #68x as per TB"
			  G20.cell(row=13, column=16).value="Depreciation CM #28x"
			  G20.cell(row=14, column=16).value="Depreciation charge 681"
			  G20.cell(row=15, column=16).value="Diff"

			  G20.cell(row=12, column=16).font=font2

			  for i in range (12,16):
				  G20.cell(row=i, column=16).font=font2

			  for i in range(12,16):
				  G20.cell(row=i, column=16).alignment=Alignment(horizontal='center', vertical='center')

			  for i in range(18,21):
				  G20.cell(row=i, column=16).alignment=Alignment(horizontal='center', vertical='center')

			  for i in range(16,18):
				  G20.cell(row=12, column=i).font=headers
				  G20.cell(row=13, column=i).border=solidbordersus
				  G20.cell(row=12, column=i).fill=headersblue

			  G20.cell(row=13, column=17).value='=SUMIF(F10_TB!B:B,"280",F10_TB!G:G)+SUMIF(F10_TB!B:B,"281",F10_TB!G:G)'
			  G20.cell(row=14, column=17).value='=SUMIF(F10_TB!J:J,"6811",F10_TB!H:H)+SUMIF(F10_TB!J:J,"6817",F10_TB!H:H)'
			  G20.cell(row=15, column=17).value='=IF(AND(Q13<0,Q14<0),Q13-Q14,IF(OR(Q13<0,Q14<0),Q13+Q14,Q13-Q14))'
			  G20.cell(row=15, column=16).font=fontRed
			  G20.cell(row=15, column=17).font=fontRed

			  G20.cell(row=13, column=17).font=font1
			  G20.cell(row=14, column=17).font=font1

			  for i in range(16, 18):
				  G20.cell(row=14, column=i).border=doubleborder


			  for i in range (13,17):
				  G20.cell(row=i, column=17).number_format='#,##0_);(#,##0)'


			  G20.cell(row=36, column=13).value=datetime.date(int(anp), int(lunap), int(ziip))
			  G20.cell(row=37, column=13).value=datetime.date(an, luna, zi)
			  G20.cell(row=36, column=13).number_format="mm/dd/yyyy"
			  G20.cell(row=37, column=13).number_format="mm/dd/yyyy"
			  G20.cell(row=36, column=13).font=fontalb
			  G20.cell(row=37, column=13).font=fontalb

			  G20.cell(row=38, column=1).value="As per Client (FAR and Disposal list)"
			  G20.cell(row=38, column=1).font=font2

			  G20.cell(row=38, column=13).value="As per Nexia Recomputation"
			  G20.cell(row=38, column=13).font=font2
			  print(len(AccountDisp))			  

			  lenTotal=len(AccountFAR)+len(AccountDisp)
			  for i in range (1,lenTotal+1):
				  G20.cell(row=i+40, column=1).value='=Left(TRIM(B{0}),3)'.format(i+40)

			  print(lenTotal)
			  print(len(AccountDisp))
			  # print(len(AccountFar))
			  h = lenTotal - len(AccountDisp)
			  print(h)

			  for i in range(1, len(AccountDisp)+1):
				  G20.cell(row=i+h+40, column=1).value='=Left(TRIM(B{0}),3)&"D"'.format(i+h+40)

			  for i in range (1, len(AccountFAR)+1):
				  G20.cell(row=i+40, column=2).value=AccountFAR[i-1]
			  ###
			  for i in range(1,len(AccountDisp)+1):
				  G20.cell(row=i+40+len(AccountFAR),column=2).value=AccountDisp[i-1]
			  ##
			  for i in range (1, len(ItemFAR)+1):
				  G20.cell(row=i+40, column=3).value=ItemFAR[i-1]

			  for i in range(1, len(DescriptionFAR)+1):
				  G20.cell(row=i+40, column=4).value=DescriptionFAR[i-1]

			  for i in range(1,len(DescriptionDisp)+1):
				  G20.cell(row=i+40+len(AccountFAR),column=4).value=DescriptionDisp[i-1]

			  for i in range (1,lenTotal+1):
				  G20.cell(row=i+40, column=5).value='=Year(F{0})'.format(i+40)

			  for i in range(1,len(PIFDate)+1):
				  G20.cell(row=i+40, column=6).value=PIFDate[i-1]

			  for i in range(1,len(PIFDisp)+1):
				  G20.cell(row=i+40+len(AccountFAR),column=6).value=PIFDisp[i-1]

			  for i in range(1,lenTotal+1):
				  G20.cell(row=i+40, column=6).number_format='mm/dd/yyyy'

			  for i in range(1,len(PIFDate)+len(DispDate)+2):
				  G20.cell(row=i+40, column=7).number_format='mm/dd/yyyy'

			  for i in range (1,len (UL)+1):
				  G20.cell(row=i+40, column=8).value=UL[i-1]

			  for i in range(1,len(ULDisp)+1):
				  G20.cell(row=i+40+len(AccountFAR),column=8).value=ULDisp[i-1]

			  for i in range (1, len(GBVFAR)+1):
				  G20.cell(row=i+40, column=9).value=GBVFAR[i-1]

			  for i in range(1,len(GBVDisp)+1):
				  G20.cell(row=i+40+len(AccountFAR),column=9).value=GBVDisp[i-1]

			  for i in range (1,len(ChargeFAR)+1):
				  G20.cell(row=i+40, column=10).value=ChargeFAR[i-1]

			  for i in range(1,len(DeprChDisp)+1):
				  G20.cell(row=i+40+len(AccountFAR),column=10).value=DeprChDisp[i-1]

			  for i in range (1, len(AccDeprFAR)+1):
				  G20.cell(row=i+40, column=11).value=AccDeprFAR[i-1]

			  for i in range(1,len(AccumulatedDepreciationDisp)+1):
				  G20.cell(row=i+40+len(AccountFAR),column=11).value=AccumulatedDepreciationDisp[i-1]

			  for i in range(1, len(AccDeprFAR)+len(AccumulatedDepreciationDisp)+2):
				  for j in range(8, 12):
					  G20.cell(row=i+40, column=j).number_format='#,##0_);(#,##0)'


			  for i in range (1,lenTotal+1):
				  G20.cell(row=i+40, column=13).value='=I{0}'.format(i+40)

			  for i in range (1,lenTotal+1):
				  G20.cell(row=i+40, column=14).value='=H{0}'.format(i+40)


			  # for i in range (1,len(PIFDate)+1):
			  #   G20.cell(row=i+33, column=6).number_format='mm/dd/yyyy'

			  # for i in range (1,len(DispDate)+1):
			  #   G20.cell(row=i+33, column=7).number_format='mm/dd/yyyy'

			  for i in range (1,lenTotal+1):
				  G20.cell(row=i+40, column=15).value='=DATE(YEAR(F{0}),MONTH(F{0})+1,1)'.format(i+40)

			  for i in range (1,lenTotal+1):
				  G20.cell(row=i+40, column=15).number_format='mm/dd/yyyy'

			  #=IF(DAY(E{0})=1,E{0},DATE(YEAR(N{0}),MONTH(N{0})+1,1))

			  for i in range (1,lenTotal+1):
				  G20.cell(row=i+40, column=16).value='=IF(MID(A{0},4,1)="D",MIN(G{0},DATE(YEAR(O{0}),MONTH(O{0})+N{0},DAY(O{0}))),(DATE(YEAR(O{0}),MONTH(O{0})+N{0},1)))'.format(i+40)

			  for i in range (1,lenTotal+1):
				  G20.cell(row=i+40, column=16).number_format='mm/dd/yyyy'


			  # for i in range (1,len(PIFDate)+1):
			  #   G20.cell(row=i+33, column=17).number_format='mm/dd/yyyy'

			  # for i in range (1,len(PIFDate)+1):
			  #   G20.cell(row=i+33, column=16).number_format='mm/dd/yyyy'

			  for i in range (1,lenTotal+1):
				  G20.cell(row=i+40, column=17).value='=DATEDIF(IF(MAX(O{0},$M$37)>$M$36+1,$M$36+1,MAX(O{0},$M$37)),MAX($M$37,MIN($M$36+1,P{0})),"m")'.format(i+40)

			  for i in range (1,lenTotal+1):
				  G20.cell(row=i+40, column=18).value='=IF(N{0}>0,M{0}/N{0},0)'.format(i+40)

			  # for i in range (1,len(PIFDate)+1):
			  #   G20.cell(row=i+33, column=18).number_format='#,##0_);(#,##0)'

			  for i in range (1,lenTotal+1):
				  G20.cell(row=i+40, column=19).value='=Q{0}*R{0}'.format(i+40)



			  for i in range (1,lenTotal+1):
				  G20.cell(row=i+40, column=20).value='=if($e$29=0,"N/A – no depr expense in FAR PBC",IF(J{0}<0,J{0}+S{0},J{0}-S{0}))'.format(i+40)



			  for i in range (1,lenTotal+1):
				  G20.cell(row=i+40, column=20).font=fontRedDiff

			  for i in range (1,lenTotal+1):
				  G20.cell(row=i+40, column=22).value='=iferror(DATEDIF(O{0},MIN(P{0},MAX($M$37,MIN($M$36+1,P{0}))),"m")*R{0},0)'.format(i+40)

			  for i in range (1,lenTotal+1):
				  G20.cell(row=i+40, column=23).value='=IF(K{0}<0,K{0}+V{0},K{0}-V{0})'.format(i+40)
			  for i in range(41,lenTotal+42):
				  G20.cell(row=i,column=13).number_format='#,##0_);(#,##0)'

			  for i in range (1,lenTotal+1):
				  G20.cell(row=i+40, column=23).font=fontRedDiff

			  for i in range(1,len(PIFDate)+len(AccumulatedDepreciationDisp)+2):
				  for j in range(17, 24):
				    G20.cell(row=i+40, column=j).number_format='#,##0_);(#,##0)'

			  for i in range(1, len(ItemDisp)+1):
				  G20.cell(row=i+len(ItemFAR)+40, column=3).value=ItemDisp[i-1]

			  for i in range(1, len(DispDate)+1):
				  G20.cell(row=i+len(ItemFAR)+40, column=7).value=DispDate[i-1]

			  # for i in range(1, len(AccumulatedDepreciationDisp)+1):
			  #   G20.cell(row=i+len(ItemFAR)+32, column=19).value=AccumulatedDepreciationDisp[i-1]

			  for i in range(4,7):
				  for j in range(15, 30):
				    G20.cell(row=j, column=i).number_format='#,##0_);(#,##0)'

			  # c=G20['A33']
			  # G20.freeze_panes = c

			  for i in range(2,7):
				  for j in range(13,19):
				    G20.cell(row=j,column=i).font=font1

			  for i in range(2,7):
				  for j in range(21,27):
				    G20.cell(row=j,column=i).font=font1


			  for i in range(1,20):
				  for j in range(41,G20.max_row):
				    G20.cell(row=j,column=i).font=font1

			  for j in range(41,G20.max_row):
				  G20.cell(row=j,column=22).font=font1

			  for i in range(41,G20.max_row):
				  G20.cell(row=i,column=24).value='=IF(V{0}<=I{0},"OK","Not OK")'.format(i)
				  G20.cell(row=i,column=24).font=font1



			  G20.cell(row=17, column=16).value="Reconciliation of Disposals with FA CM"
			  G20.cell(row=18, column=16).value="Disposals as per detail"
			  G20.cell(row=19, column=16).value="As per TB"
			  for i in range(18, 20):
				  G20.cell(row=i, column=16).font=font2
			  G20.cell(row=20, column=16).value="Diff"
			  G20.cell(row=20, column=16).alignment=Alignment(horizontal='center')
			  for i in range(16, 18):
				  G20.cell(row=20, column=i).font=fontRed

			  for i in range(16, 18):
				  G20.cell(row=17, column=i).font=headers
				  G20.cell(row=18, column=i).border=solidbordersus
				  G20.cell(row=17, column=i).fill=headersblue
				  G20.cell(row=19, column=i).border=doubleborder

			  G20.cell(row=18, column=17).value='=SUM(PBC_Disposals!H:H)'
			  G20.cell(row=19, column=17).value='=SUMIF(F10_TB!B:B,"201",F10_TB!G:G)+SUMIF(F10_TB!B:B,"203",F10_TB!G:G)+SUMIF(F10_TB!B:B,"205",F10_TB!G:G)+SUMIF(F10_TB!B:B,"208",F10_TB!G:G)+SUMIF(F10_TB!B:B,"206",F10_TB!G:G)+SUMIF(F10_TB!B:B,"207",F10_TB!G:G)+SUMIF(F10_TB!B:B,"212",F10_TB!G:G)+SUMIF(F10_TB!B:B,"213",F10_TB!G:G)+SUMIF(F10_TB!B:B,"214",F10_TB!G:G)+SUMIF(F10_TB!B:B,"215",F10_TB!G:G)+SUMIF(F10_TB!B:B,"216",F10_TB!G:G)+SUMIF(F10_TB!B:B,"217",F10_TB!G:G)'
			  G20.cell(row=20, column=17).value='=IF(AND(Q18<0,Q19<0),Q18-Q19,IF(OR(Q18<0,Q19<0),Q18+Q19,Q18-Q19))'

			  for i in range(18, 21):
				  G20.cell(row=i, column=17).number_format='#,##0_);(#,##0)'
				  G20.cell(row=i, column=17).font=font1

			  # for i in range(4, 7):
			  #   for j in range(13, 16):
			  #     G20.cell(row=j, column=i).number_format='#,##0_);(#,##0)'

			  G20.cell(row=20, column=17).font=fontRed

			  for i in range(13, 15):
				  G20.cell(row=i, column=14).font=font2

			  for i in range(13, 35):
				  G20.cell(row=i, column=7).number_format='#,##0_);(#,##0)'

			  G20.cell(row=30, column=9).font=font2
			  G20.cell(row=19, column=9).font=font2
			  G20.cell(row=19, column=11).font=font2

			  for i in range(2,8):
				  G20.cell(row=12, column=i).alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)

			  for i in range(9,15):
				  G20.cell(row=12, column=i).alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)

			  for i in range(1, 25):
				  G20.cell(row=40, column=i).alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)


			  G20.row_dimensions[39].height = 4
			  G20.row_dimensions[28].height = 4

			  G20.merge_cells('P12:Q12')
			  G20.merge_cells('P17:Q17')
			  G20.merge_cells('D11:G11')
			  G20.merge_cells('I11:N11')

			  G20.cell(row=11, column=4).alignment=Alignment(horizontal='center', vertical='center')
			  G20.cell(row=11, column=9).alignment=Alignment(horizontal='center', vertical='center')
			  G20.cell(row=12, column=16).alignment=Alignment(horizontal='center', vertical='center')
			  G20.cell(row=17, column=16).alignment=Alignment(horizontal='center', vertical='center')

			  for i in range(4, 8):
				  G20.cell(row=11, column=i).border=solidborder

			  for i in range(4, 8):
				  G20.cell(row=10, column=i).border=solidborder

			  for i in range(9, 15):
				  G20.cell(row=11, column=i).border=solidborder

			  for i in range(11, 15):
				  G20.cell(row=10, column=i).border=solidborder

			  G20.cell(row=11, column=3).border=solidborderdreapta
			  G20.cell(row=11, column=7).border=solidborderdreapta
			  G20.cell(row=10, column=9).border=solidborder
			  G20.cell(row=10, column=10).border=solidborder
			  G20.cell(row=11, column=7).border=solidborder
			  G20.cell(row=11, column=7).border=solidborderdreapta

			  G20.cell(row=11, column=8).border=solidborderdreapta
			  G20.cell(row=12, column=7).border=solidbordersus
			  G20.cell(row=11, column=15).border=solidborderstanga
			  G20.cell(row=13, column=14).font=fontRedDiff
			  G20.cell(row=14, column=14).font=fontRedDiff
			  G20.cell(row=19, column=14).font=fontRed

			  for i in range(13, 19):
				  G20.cell(row=i ,column=7).font=font1

			  for i in range(21, 27):
				  G20.cell(row=i ,column=7).font=font1

			  for i in range(9, 14):
				  G20.cell(row=30, column=i).border=solidborder

			  for i in range(2, 14):
				  G20.cell(row=20, column=i).font=font1
			  for i in range(4, 14):
				  G20.cell(row=20, column=i).number_format='#,##0_);(#,##0)'

			  G20.cell(row=20, column=10).font=fontRedDiff
			  G20.cell(row=20, column=12).font=fontRedDiff
			  G20.cell(row=20, column=14).font=fontRedDiff

			  for i in range(4, 7):
				  for j in range(13, 15):
				    G20.cell(row=j, column=i).number_format='#,##0_);(#,##0)'

			  G20.cell(row=20, column=2).alignment = Alignment (horizontal='right')
			  G20.sheet_view.showGridLines = False

		if(val3==1):
			G30=FAR.create_sheet("G30 FA TOM")

			# G30.title="G30 FA TOM"


			PBC_Addition =FAR.create_sheet("PBC_Add")

			# PBC_CIP =FAR.create_sheet("PBC_CIP")



			"Create PBC FAR ---------------------------------------------------------------------"




			"Create Add Detail ---------------------------------------------------------------------"


			PBC_Addition.cell(row=1, column=1).value="Item"
			PBC_Addition.cell(row=1, column=2).value="Account"
			PBC_Addition.cell(row=1, column=3).value="GBV"
			PBC_Addition.cell(row=1, column=4).value="Depreciation Charge"
			PBC_Addition.cell(row=1, column=5).value="Accumulated Depreciation"
			PBC_Addition.cell(row=1, column=6).value="PIF Date"
			PBC_Addition.cell(row=1, column=7).value="Synt Account"
			PBC_Addition.cell(row=1, column=8).value="Type"
			PBC_Addition.cell(row=1, column=9).value="Depr Account"
			PBC_Addition.cell(row=1,column=10).value="Synt 4"
			try:
				for i in range (1,11):
				  PBC_Addition.cell(row=1, column= i).border=doubleborder
				  PBC_Addition.cell(row=1, column= i).font=font2
				  PBC_Addition.cell(row=1, column= i).fill=blueFill

				for i in range (1,len(ItemAdd)+1):
				  PBC_Addition.cell(row=i+1, column=1).value=ItemAdd[i-1]

				for i in range (1,len(AccountAdd)+1):
				  PBC_Addition.cell(row=i+1, column=2).value=AccountAdd[i-1]

				for i in range (1,len(GBVAdd)+1):
				  PBC_Addition.cell(row=i+1, column=3).value=GBVAdd[i-1]

				for i in range (1,len(AccumulatedDepreciation)+1):
				  PBC_Addition.cell(row=i+1, column=5).value=AccumulatedDepreciation[i-1]


				for i in range (1,len(Charge)+1):
				  PBC_Addition.cell(row=i+1, column=4).value=Charge[i-1]

				for i in range (1,len(PIFDateadd)+1):
				  PBC_Addition.cell(row=i+1, column=6).value=PIFDateadd[i-1]

				for i in range (1,len(PIFDateadd)+1):
				  for j in range (3,6):
					  PBC_Addition.cell(row=i+1, column=j).number_format='#,##0_);(#,##0)'

				for i in range (1,len(PIFDateadd)+1):
				  PBC_Addition.cell(row=i+1, column=6).number_format='mm/dd/yyyy'

				for i in range (1,len(PIFDateadd)+1):
				  PBC_Addition.cell(row=i+1, column=7).value='=left(B{0},3)'.format(i+1)

				for i in range (1,len(AccountAdd)+1):
				  PBC_Addition.cell(row=i+1, column=8).value=TypeAdd[i-1]

				for i in range (1,len(AccountAdd)+1):
				  PBC_Addition.cell(row=i+1, column=9).value='=IF(G{0}<"210","280","281")'.format(i+1)

				for i in range (1,len(AccountAdd)+1):
				  PBC_Addition.cell(row=i+1, column=10).value='=left(B{0},4)'.format(i+1)
			except:
				l=1

			"Adjust Column Width---------------------------------------------------------------------------------------------------------------------------------."

			for col in PBC_Addition.columns:
				max_length = 10
				for cell in col:
					if cell.coordinate in PBC_Addition.merged_cells:
						continue
					try:
						if len(str(cell.value)) > max_length:
							max_length = len(cell.value)
					except:
						pass

				adjusted_width = (max_length + 5)


			listanoua=['A','B','C','D','E','F','G']

			for column in listanoua:
			  for i in listanoua:
				  if (column==i):
				    PBC_Addition.column_dimensions[column].width = adjusted_width







			"Create CIP Detail ---------------------------------------------------------------------"

			# PBC_CIP.cell(row=1, column=1).value="Item"
			# PBC_CIP.cell(row=1, column=2).value="Account"
			# PBC_CIP.cell(row=1, column=3).value="Addition"
			# PBC_CIP.cell(row=1, column=4).value="Disposals"
			# PBC_CIP.cell(row=1, column=5).value="Type"
			# PBC_CIP.cell(row=1, column=6).value="Synt"


			# for i in range (1,7):
			#   PBC_CIP.cell(row=1, column= i).border=doubleborder
			#   PBC_CIP.cell(row=1, column= i).font=font2
			#   PBC_CIP.cell(row=1, column= i).fill=blueFill

			# try:
			#   for i in range (1,len(ItemCIP)+1):
			#     PBC_CIP.cell(row=i+1, column=1).value=ItemCIP[i-1]
			#
			#   for i in range (1,len(AccountCIP)+1):
			#     PBC_CIP.cell(row=i+1, column=2).value=AccountCIP[i-1]
			#
			#   for i in range (1,len(AdditionCIP)+1):
			#     PBC_CIP.cell(row=i+1, column=3).value=AdditionCIP[i-1]
			#
			#   for i in range (1,len(DisposalCIP)+1):
			#     PBC_CIP.cell(row=i+1, column=4).value=DisposalCIP[i-1]
			#
			#   for i in range (1,len(TypeCIP)+1):
			#     PBC_CIP.cell(row=i+1, column=5).value=TypeCIP[i-1]
			#
			#   for i in range (1,len(AccountCIP)+1):
			#     PBC_CIP.cell(row=i+1, column=6).value='=left(B{0},3)'.format(i+1);
			#
			#   for i in range (1,len(AccountCIP)+1):
			#     for j in range (3,5):
			#       PBC_CIP.cell(row=i+1,column=j).number_format='#,##0_);(#,##0)'
			#
			# except:
			#   a=1
			"Adjust Column Width---------------------------------------------------------------------------------------------------------------------------------."

			# for col in PBC_CIP.columns:
			# 	max_length = 10
			# 	for cell in col:
			# 		if cell.coordinate in PBC_CIP.merged_cells:
			# 			continue
			# 		try:
			# 			if len(str(cell.value)) > max_length:
			# 				max_length = len(cell.value)
			# 		except:
			# 			pass

			# 	adjusted_width = (max_length + 5)


			# listanoua=['A','B','C','D','E','F','G']

			# for column in listanoua:
			#   for i in listanoua:
			# 	  if (column==i):
			# 	    PBC_CIP.column_dimensions[column].width = adjusted_width

			"Create Advances Detail ---------------------------------------------------------------------"

			"Adjust Column Width---------------------------------------------------------------------------------------------------------------------------------."


			"Adjust Column Width---------------------------------------------------------------------------------------------------------------------------------."

			# for col in F10_TB.columns:
			#     max_length = 0
			#     for cell in col:
			#         if cell.coordinate in F10_TB.merged_cells:
			#             continue
			#         try:
			#             if len(str(cell.value)) > max_length:
			#                 max_length = len(cell.value)
			#         except:
			#             pass
			#     adjusted_width = (max_length - 20)




			G30.cell(row=1, column=1).value="Client:"
			G30.cell(row=1, column=1).font=font2
			G30.cell(row=1, column=2).value=namec
			G30.cell(row=1, column=2).font=font2
			G30.cell(row=2, column=1).value="Period end:"
			G30.cell(row=2, column=1).font=font2
			G30.cell(row=2, column=2).value=ant
			G30.cell(row=2, column=2).number_format='mm/dd/yyyy'
			G30.cell(row=2, column=2).font=font2
			G30.cell(row=1, column=11).value="Prepared by:"
			G30.cell(row=1, column=11).font=font2

			G30.cell(row=2, column=11).value="Date:"
			G30.cell(row=2, column=11).font=font2
			G30.cell(row=2, column=12).value=datetime.datetime.now().date()
			G30.cell(row=2, column=12).number_format="mm/dd/yyyy"
			G30.cell(row=2, column=12).font=font1
			G30.cell(row=2, column=12).alignment = Alignment (horizontal='left')

			G30.cell(row=3, column=11).value="Ref:"
			G30.cell(row=3, column=11).font=font2
			G30.cell(row=3, column=12).value="G30"
			G30.cell(row=3, column=12).font=fontRed

			for i in range(1,4):
				G30.cell(row=i, column=11).alignment = Alignment(horizontal='right')

			G30.cell(row=4, column=3).value="Tangible and Intangible Assets Table of Movement"
			G30.cell(row=4, column=3).font=font2

			G30.cell(row=6, column=1).value="Procedure:"
			G30.cell(row=6, column=1).alignment=Alignment(horizontal='right')
			G30.cell(row=6, column=2).value="Based on the Tangible and Intangible Movement Details from the period received from client (Additions, Disposals, Assets in progress, including identification of Transfers and Advances) and the Fixed Asset Register (FAR), perform the following:"
			G30.cell(row=7, column=2).value="- prepare table of movement of tangible and intangible assets for GBV, Accumulated depreciation and Impairment(based on TB) and reconcile the closing amount with the TB"
			G30.cell(row=8, column=2).value="- compute the closing NBV of tangible and intangible assets and reconcile it with TB"
			G30.cell(row=9, column=2).value="- prepare a summary of opening and closing NBV"
			G30.cell(row=6, column=1).font=font2

			G30.cell(row=6, column=2).font=font1
			G30.cell(row=7, column=2).font=font1
			G30.cell(row=8, column=2).font=font1
			G30.cell(row=9, column=2).font=font1

			G30.cell(row=11, column=1).value="Work done"
			G30.cell(row=11, column=1).alignment=Alignment(horizontal='right')
			G30.cell(row=11, column=1).font=font2
			G30.cell(row=11, column=2).value="Please see work below:"
			G30.cell(row=11, column=2).font=font1

			for i in range(14,21):
				G30.cell(row=i,column=3).border=rightborder
				G30.cell(row=i,column=6).border=rightborder
				G30.cell(row=i,column=9).border=rightborder
			for i in range(1,7):
			  G30.cell(row=79,column=i).border=doubleborder
			  G30.cell(row=91,column=i).border=doubleborder

			for i in range(25,32):
			  G30.cell(row=i,column=3).border=rightborder
			  G30.cell(row=i,column=6).border=rightborder
			  G30.cell(row=i,column=9).border=rightborder
			for i in range(37,41):
			  G30.cell(row=i,column=3).border=rightborder
			  G30.cell(row=i,column=6).border=rightborder
			  G30.cell(row=i,column=9).border=rightborder
			for i in range(49,55):
			  G30.cell(row=i,column=3).border=rightborder
			  G30.cell(row=i,column=6).border=rightborder
			  G30.cell(row=i,column=9).border=rightborder
			for i in range(58,66):
			  G30.cell(row=i,column=3).border=rightborder
			  G30.cell(row=i,column=6).border=rightborder
			  G30.cell(row=i,column=9).border=rightborder


			for i in range(1,13):
			  G30.cell(row=20,column=i).border=doubleborder
			  G30.cell(row=31,column=i).border=doubleborder
			  G30.cell(row=40,column=i).border=doubleborder
			  G30.cell(row=54,column=i).border=doubleborder
			  G30.cell(row=65,column=i).border=doubleborder


			G30.cell(row=13, column=1).value="Working 1: GBV"
			G30.cell(row=13, column=1).font=workingsblue

			G30.cell(row=14,column=1).value="'-Intangible assets-"
			G30.cell(row=14,column=1).font=greenbolditalic
			G30.cell(row=14, column=6).value="Additions(PBC_Add)"
			G30.cell(row=14, column=9).value="Disposals(PBC_Disposals)"

			# for i in range (4,7):
			#     G30.cell(row=14,column=i).alignment=Alignment (horizontal='centerContinuous')


			# for i in range (7,10):
			#     G30.cell(row=14,column=i).alignment=Alignment (horizontal='centerContinuous')

			"Adjust Column Width---------------------------------------------------------------"

			for col in G30.columns:
				max_length = 20
				for cell in col:
					if cell.coordinate in G30.merged_cells:
						continue
					try:
						if len(str(cell.value)) > max_length:
							max_length = len(cell.value)
					except:
						pass
				adjusted_width = (max_length - 2)

			listanoua=['A','B','C','D','E','F','G','H','J','I','K']
			for column in listanoua:
				for i in listanoua:
					if (column==i):
						G30.column_dimensions[column].width = adjusted_width


			'---------'


			G30.cell(row=15, column=1).value="Account"
			G30.cell(row=15, column=2).value="Description"
			G30.cell(row=15, column=3).value="OB 01.01."+str(yearentry)
			G30.cell(row=15, column=4).value="Pure"
			G30.cell(row=15, column=5).value="Transfer"
			G30.cell(row=15, column=6).value="Total Additions"
			# G30.cell(row=15, column=7).value="Sales"
			# G30.cell(row=15, column=8).value="Scrapped"
			G30.cell(row=15, column=9).value="Total Disposals"
			G30.cell(row=15, column=10).value="Recomputed CB as per Nexia"
			G30.cell(row=15, column=11).value="CB 31.12."+str(yearentry)+" as per TB"
			G30.cell(row=15, column=12).value="Diff"
			for i in range (1,13):
			  G30.cell(row=15,column=i).font=workingsblue2
			  G30.cell(row=15,column=i).fill=lbluefill
			  G30.cell(row=15,column=i).border=solidborder

			for i in range (4,10):
			  G30.cell(row=14, column=i).font=workingsblue2
			  G30.cell(row=14, column=i).border=solidborder
			  G30.cell(row=14, column=i).fill=lbluefill


			G30.cell(row=16, column=1).value="201"
			G30.cell(row=17, column=1).value="203"
			G30.cell(row=18, column=1).value="205 & 208"
			G30.cell(row=19, column=1).value="2071"
			G30.cell(row=20, column=1).value="206"

			G30.cell(row=21,column=1).value="Total GBV intangible assets in function"
			G30.cell(row=21,column=1).font=font2

			G30.cell(row=23,column=1).value='-Tangible assets-'
			G30.cell(row=23,column=1).font=greenbolditalic

			G30.cell(row=24, column=1).value="Account"
			G30.cell(row=24, column=2).value="Description"
			G30.cell(row=24, column=3).value="OB 01.01."+str(yearentry)
			G30.cell(row=24, column=4).value="Pure"
			G30.cell(row=24, column=5).value="Transfer"
			G30.cell(row=24, column=6).value="Total Additions"
			# G30.cell(row=15, column=7).value="Sales"
			# G30.cell(row=15, column=8).value="Scrapped"
			G30.cell(row=24, column=9).value="Total Disposals"
			G30.cell(row=24, column=10).value="Recomputed CB as per Nexia"
			G30.cell(row=24, column=11).value="CB 31.12."+str(yearentry)+" as per TB"
			G30.cell(row=24, column=12).value="Diff"

			for i in range (1,13):
			  G30.cell(row=24,column=i).font=workingsblue2
			  G30.cell(row=24,column=i).fill=lbluefill
			  G30.cell(row=24,column=i).border=solidborder


			G30.cell(row=25, column=1).value="211"
			G30.cell(row=26, column=1).value="212"
			G30.cell(row=1+26, column=1).value="213 & 223"
			G30.cell(row=1+27, column=1).value="214 & 224"
			G30.cell(row=1+28, column=1).value="215"
			G30.cell(row=1+29, column=1).value="216"
			G30.cell(row=1+30, column=1).value="217 & 227"
			G30.cell(row=1+31, column=1).value="Total GBV tangible assets in function"
			G30.cell(row=1+31, column=1).font=font2
			# G30.cell(row=1+33, column=1).value="Total Fixed Assets"
			# G30.cell(row=1+33, column=1).font=font2
			# for i in range(1,13):
			#   G30.cell(row=1+33,column=i).fill=grifill


			G30.cell(row=1+36, column=1).value="-CIP-"
			G30.cell(row=1+36, column=1).font=greenbolditalic

			# G30.cell(row=1+36, column=6).value="Additions(PBC_CIP)"
			# G30.cell(row=1+36, column=9).value="Disposals(PBC_CIP)"

			# for i in range (4,7):
			#     G30.cell(row=1+36,column=i).alignment=Alignment (horizontal='centerContinuous')


			# for i in range (7,10):
			#     G30.cell(row=1+36,column=i).alignment=Alignment (horizontal='centerContinuous')


			G30.cell(row=1+37, column=1).value="Account"
			G30.cell(row=1+37, column=2).value="Description"
			G30.cell(row=1+37, column=3).value="OB 01.01."+str(yearentry)
			G30.cell(row=1+37, column=4).value="Pure"
			G30.cell(row=1+37, column=6).value="Total Additions"
			G30.cell(row=1+37, column=7).value="Sales/Scrapped"
			G30.cell(row=1+37, column=8).value="Transfer"
			G30.cell(row=1+37, column=9).value="Total Disposals"
			G30.cell(row=1+37, column=10).value="Recomputed CB as per Nexia"
			G30.cell(row=1+37, column=11).value="CB 31.12."+str(yearentry)+" as per TB"
			G30.cell(row=1+37, column=12).value="Diff"
			for i in range (1,13):
			  G30.cell(row=1+37,column=i).font=workingsblue2
			  G30.cell(row=1+37,column=i).fill=lbluefill
			  G30.cell(row=1+37,column=i).border=solidborder

			for i in range (4,10):
			  G30.cell(row=1+36, column=i).font=workingsblue2
			  G30.cell(row=1+36, column=i).border=solidborder
			  G30.cell(row=1+36, column=i).fill=lbluefill


			G30.cell(row=1+38, column=1).value="231"
			G30.cell(row=1+39, column=1).value="235"
			G30.cell(row=1+40, column=1).value="Total GBV CIP"
			G30.cell(row=1+40, column=1).font=font2
			# for i in range(1,13):
			#   G30.cell(row=1+40, column=i).fill=grifill
			###

			# for i in range(1,13):
			#   G30.cell(row=47, column=i).fill=grifill

			G30.cell(row=44, column=1).value="TOTAL FA GBV"
			G30.cell(row=44, column=1).font=font2

			G30.cell(row=47, column=1).value="Working 2: Accumulated Depreciation and Impairment"
			G30.cell(row=47, column=1).font=workingsblue

			G30.cell(row=48,column=1).value="'-Depreciation of intangible assets-"
			G30.cell(row=48,column=1).font=greenbolditalic
			G30.cell(row=49, column=1).value="Account"
			G30.cell(row=49, column=2).value="Description"
			G30.cell(row=49, column=3).value="OB 01.01."+str(yearentry)
			G30.cell(row=49, column=6).value="Depreciation Charge"
			G30.cell(row=49, column=9).value="Accumulated depreciation of disposals"
			G30.cell(row=49, column=10).value="Recomputed CB as per Nexia"
			G30.cell(row=49, column=11).value="CB 31.12."+str(yearentry)+" as per TB"
			G30.cell(row=49, column=12).value="Diff"

			for i in range (1,13):
			  G30.cell(row=49,column=i).font=workingsblue2
			  G30.cell(row=49,column=i).fill=lbluefill
			  G30.cell(row=49,column=i).border= solidborder

			# G30.cell(row=52, column=4).value="Additions(FAR)"
			# G30.cell(row=52, column=7).value="Disposals(PBC_Disposals)"

			# for i in range (4,7):
			#     G30.cell(row=52,column=i).alignment=Alignment (horizontal='centerContinuous')

			# for i in range (7,10):
			#     G30.cell(row=52,column=i).alignment=Alignment (horizontal='centerContinuous')

			# for i in range (4,10):
			#   G30.cell(row=52, column=i).font=font2
			#   G30.cell(row=52, column=i).border=solidborder
			#   G30.cell(row=52, column=i).fill=lbluefill


			G30.cell(row=50, column=1).value="2801"
			G30.cell(row=51, column=1).value="2803"
			G30.cell(row=52, column=1).value="2805 & 2808"
			G30.cell(row=53, column=1).value="2807"
			G30.cell(row=54, column=1).value="2806"

			G30.cell(row=55, column=1).value="Total depreciation of intangible assets"
			G30.cell(row=55, column=1).font=font2

			G30.cell(row=57,column=1).value='-Depreciation of tangible assets'
			G30.cell(row=57,column=1).font=greenbolditalic

			G30.cell(row=58, column=1).value="Account"
			G30.cell(row=58, column=2).value="Description"
			G30.cell(row=58, column=3).value="OB 01.01."+str(yearentry)
			G30.cell(row=58, column=6).value="Depreciation Charge"
			G30.cell(row=58, column=9).value="Accumulated depreciation of disposals"
			G30.cell(row=58, column=10).value="Recomputed CB as per Nexia"
			G30.cell(row=58, column=11).value="CB 31.12."+str(yearentry)+" as per TB"
			G30.cell(row=58, column=12).value="Diff"

			for i in range (1,13):
			  G30.cell(row=58,column=i).font=workingsblue2
			  G30.cell(row=58,column=i).fill=lbluefill
			  G30.cell(row=58,column=i).border= solidborder


			G30.cell(row=59, column=1).value="2811"
			G30.cell(row=60, column=1).value="2812"
			G30.cell(row=61, column=1).value="2813"
			G30.cell(row=62, column=1).value="2814"
			G30.cell(row=63, column=1).value="2815"
			G30.cell(row=64, column=1).value="2816"
			G30.cell(row=65, column=1).value="2817"
			G30.cell(row=66, column=1).value="Total Depreciation of tangible assets"
			G30.cell(row=66, column=1).font=font2

			G30.cell(row=67, column=1).value="Total depreciation"
			G30.cell(row=67, column=1).font=font2

			# for i in range(1,13):
			#   G30.cell(row=95, column=i).fill=grifill

			# G30.cell(row=98, column=1).value="Working 5: Depreciation, impairment and goodwill check"
			# G30.cell(row=98, column=1).font=workingsblue

			# G30.cell(row=99, column=1).value="Check"
			# G30.cell(row=99, column=2).value="As per FAR"
			# G30.cell(row=99, column=3).value="As per TB"
			# G30.cell(row=99, column=6).value="Diff"

			# for i in range (1,7):
			#   G30.cell(row=99,column=i).font=font2
			#   G30.cell(row=99,column=i).fill=lbluefill
			#   G30.cell(row=99,column=i).border= solidborder
			#   # G30.cell(row=105,column=i).border= solidborder


			# G30.cell(row=100, column=1).value="Depreciation charge"
			# G30.cell(row=101, column=1).value="Impairment charge"
			# G30.cell(row=102, column=1).value="Goodwill charge"


			G30.cell(row=72, column=1).value="Working 3: NBV"
			G30.cell(row=72, column=1).font=workingsblue

			G30.cell(row=73, column=1).value="Account"
			G30.cell(row=73, column=2).value="Description"
			G30.cell(row=73, column=3).value="Recomputed NBV 01.01."+str(yearentry)+" as per Nexia"
			# G30.cell(row=105, column=4).value="Accumulated depreciation"
			# G30.cell(row=105, column=5).value="Accumulated impairment"
			G30.cell(row=73, column=6).value="Recomputed NBV 31.12."+str(yearentry)+" as per Nexia"

			for i in range (1,7):
			  G30.cell(row=73,column=i).font=workingsblue2
			  G30.cell(row=73,column=i).fill=lbluefill
			  G30.cell(row=73,column=i).border= solidborder

			G30.cell(row=74, column=1).value="201,2801"
			G30.cell(row=75, column=1).value="203,2803"
			G30.cell(row=76, column=1).value="205 & 208, 2805 & 2808"
			G30.cell(row=77, column=1).value="2071,2807"
			G30.cell(row=78, column=1).value="206, 2806"
			G30.cell(row=79, column=1).value="4094"
			G30.cell(row=80, column=1).value="Total NBV intangibles"

			G30.cell(row=83, column=1).value="211, 212, 2811, 2812"
			G30.cell(row=84, column=1).value="213 & 223, 2813"
			G30.cell(row=85, column=1).value="214 & 224, 2814"
			G30.cell(row=86, column=1).value="215, 2815"
			G30.cell(row=87, column=1).value="216, 2816"
			G30.cell(row=88, column=1).value="217, 2817"
			G30.cell(row=89, column=1).value="4093"
			G30.cell(row=90, column=1).value="231"
			G30.cell(row=91, column=1).value="235"
			G30.cell(row=92, column=1).value="Total NBV tangibles"

			G30.cell(row=94, column=1).value="Total Intangible and Tangible Assets NBV"
			G30.cell(row=94, column=1).font=font2

			G30.cell(row=96,column=1).value="Findings:"
			G30.cell(row=96,column=1).font=font2
			#mai punem niste borduri

			# G30.cell(row=15,column=3).border=rightmedium
			# G30.cell(row=15,column=6).border=rightmedium
			# G30.cell(row=15,column=9).border=rightmedium

			# G30.cell(row=14,column=6).border=rightmedium
			# G30.cell(row=14,column=9).border=rightmedium

			# G30.cell(row=20,column=3).border=rightdouble
			# G30.cell(row=20,column=6).border=rightdouble
			# G30.cell(row=20,column=9).border=rightdouble

			# G30.cell(row=31,column=3).border=rightdouble
			# G30.cell(row=31,column=6).border=rightdouble
			# G30.cell(row=31,column=9).border=rightdouble


			# G30.cell(row=24,column=3).border=rightdouble
			# G30.cell(row=24,column=6).border=rightdouble
			# G30.cell(row=24,column=9).border=rightdouble


			# G30.cell(row=38,column=3).border=rightmedium
			# G30.cell(row=38,column=6).border=rightmedium
			# G30.cell(row=38,column=9).border=rightmedium

			# G30.cell(row=40,column=3).border=rightdouble
			# G30.cell(row=40,column=6).border=rightdouble
			# G30.cell(row=40,column=9).border=rightdouble


			# # G30.cell(row=52,column=3).border=rightborder
			# # G30.cell(row=52,column=6).border=rightmedium
			# # G30.cell(row=52,column=9).border=rightmedium
			# G30.cell(row=49,column=3).border=rightmedium
			# G30.cell(row=49,column=6).border=rightmedium
			# G30.cell(row=49,column=9).border=rightmedium

			# G30.cell(row=58,column=3).border=rightdouble
			# G30.cell(row=58,column=6).border=rightdouble
			# G30.cell(row=58,column=9).border=rightdouble




			#descrieri
			G30.cell(row=16, column=2).value="Set-up costs"
			G30.cell(row=17, column=2).value="Development costs"
			G30.cell(row=18, column=2).value="Licenses, Trademarks and Other intangibles"
			G30.cell(row=19, column=2).value="Goodwill"
			G30.cell(row=20, column=2).value="Intangible assets for mineral resources"

			G30.cell(row=25, column=2).value="Land"
			G30.cell(row=26, column=2).value="Buildings"
			G30.cell(row=1+26, column=2).value="Equipments and Machines"
			G30.cell(row=1+27, column=2).value="Furniture, office equipments"
			G30.cell(row=1+28, column=2).value="Investment properties"
			G30.cell(row=1+29, column=2).value="Tangible assets for mineral resources"
			G30.cell(row=1+30, column=2).value="Biological assets"

			G30.cell(row=1+38, column=2).value="WIP Tangible Assets"
			G30.cell(row=1+39, column=2).value="WIP Investment properties"

			G30.cell(row=50, column=2).value="Depreciation for set-up costs"
			G30.cell(row=51, column=2).value="Depreciation for development costs"
			G30.cell(row=52, column=2).value="Depreciation for Licenses, Trademarks and Other intangibles "
			G30.cell(row=53, column=2).value="Depreciation for Goodwill"
			G30.cell(row=54, column=2).value="Depreciation for intangible assets for mineral resources "

			G30.cell(row=59, column=2).value="Depreciation for Land"
			G30.cell(row=60, column=2).value="Depreciation for Buildings"
			G30.cell(row=61, column=2).value="Depreciation for Equipments and Machines"
			G30.cell(row=62, column=2).value="Depreciation for Furniture, office equipments"
			G30.cell(row=63, column=2).value="Depreciation for investment properties"
			G30.cell(row=64, column=2).value="Depreciation for tangible assets for mineral resources"
			G30.cell(row=65, column=2).value="Depreciation for biological assets"



			G30.cell(row=74, column=2).value="Set-up costs"
			G30.cell(row=75, column=2).value="Development costs"
			G30.cell(row=76, column=2).value="Licenses, Trademarks and Other intangibles"
			G30.cell(row=77, column=2).value="Goodwill"
			G30.cell(row=78, column=2).value="Intangible assets for mineral resources"
			G30.cell(row=79, column=2).value="Advances for intangibles"

			G30.cell(row=83, column=2).value="Land and Buildings"
			G30.cell(row=84, column=2).value="Equipments and Machines"
			G30.cell(row=85, column=2).value="Furniture, office equipments"
			G30.cell(row=86, column=2).value="Investment properties"
			G30.cell(row=87, column=2).value="Tangible assets for mineral resources"
			G30.cell(row=88, column=2).value="Biological assets"
			G30.cell(row=89, column=2).value="Advances for tangibles"

			G30.cell(row=90, column=2).value="WIP Tangible Assets"
			G30.cell(row=91, column=2).value="WIP investment properties"

			#formule

			G30.cell(row=16, column=3).value='=SUMIF(F10_TB!B:B,"201",F10_TB!E:E)'
			G30.cell(row=17, column=3).value='=SUMIF(F10_TB!B:B,"203",F10_TB!E:E)'
			G30.cell(row=18, column=3).value='=SUMIF(F10_TB!B:B,"205",F10_TB!E:E)+SUMIF(F10_TB!B:B,"208",F10_TB!E:E)'
			G30.cell(row=19, column=3).value='=SUMIF(F10_TB!B:B,"207",F10_TB!E:E)'
			G30.cell(row=20, column=3).value='=SUMIF(F10_TB!B:B,"206",F10_TB!E:E)'
			G30.cell(row=21, column=3).value='=SUM(C16:C20)'

			G30.cell(row=16, column=4).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"201",PBC_Add!H:H,"Pure")'
			G30.cell(row=17, column=4).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"203",PBC_Add!H:H,"Pure")'
			G30.cell(row=18, column=4).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"205",PBC_Add!H:H,"Pure")+SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"208",PBC_Add!H:H,"Pure")'
			G30.cell(row=19, column=4).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"207",PBC_Add!H:H,"Pure")'
			G30.cell(row=20, column=4).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"206",PBC_Add!H:H,"Pure")'
			G30.cell(row=21, column=4).value='=SUM(D16:D20)'

			G30.cell(row=16, column=5).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"201",PBC_Add!H:H,"Transfer")'
			G30.cell(row=17, column=5).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"203",PBC_Add!H:H,"Transfer")'
			G30.cell(row=18, column=5).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"205",PBC_Add!H:H,"Transfer")+SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"208",PBC_Add!H:H,"Transfer")'
			G30.cell(row=19, column=5).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"207",PBC_Add!H:H,"Transfer")'
			G30.cell(row=20, column=5).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"206",PBC_Add!H:H,"Transfer")'
			G30.cell(row=21, column=5).value='=SUM(E16:E20)'

			G30.cell(row=16, column=6).value="=SUM(D16:E16)"
			G30.cell(row=17, column=6).value="=SUM(D17:E17)"
			G30.cell(row=18, column=6).value="=SUM(D18:E18)"
			G30.cell(row=19, column=6).value="=SUM(D19:E19)"
			G30.cell(row=20, column=6).value="=SUM(D20:E20)"
			G30.cell(row=21, column=6).value="=SUM(F16:F20)"

			# G30.cell(row=17, column=7).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"201",PBC_Disposals!K:K,"Sale")'
			# G30.cell(row=18, column=7).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"203",PBC_Disposals!K:K,"Sale")'
			# G30.cell(row=19, column=7).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"205",PBC_Disposals!K:K,"Sale")+SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"208",PBC_Disposals!K:K,"Sale")'
			# G30.cell(row=20, column=7).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"207",PBC_Disposals!K:K,"Sale")'
			# G30.cell(row=21, column=7).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"206",PBC_Disposals!K:K,"Sale")'
			# G30.cell(row=22, column=7).value="=SUM(G17:G21)"


			# # G30.cell(row=17, column=8).value=0
			# # G30.cell(row=18, column=8).value=0
			# # G30.cell(row=19, column=8).value=0
			# # G30.cell(row=20, column=8).value=0
			# # G30.cell(row=21, column=8).value=0
			# # G30.cell(row=22, column=8).value="=SUM(H17:H21)"
			# G30.cell(row=17, column=8).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"201",PBC_Disposals!K:K,"Scrap")'
			# G30.cell(row=18, column=8).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"203",PBC_Disposals!K:K,"Scrap")'
			# G30.cell(row=19, column=8).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"205",PBC_Disposals!K:K,"Scrap")+SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"208",PBC_Disposals!K:K,"Scrap")'
			# G30.cell(row=20, column=8).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"207",PBC_Disposals!K:K,"Scrap")'
			# G30.cell(row=21, column=8).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"206",PBC_Disposals!K:K,"Scrap")'
			# G30.cell(row=22, column=8).value="=SUM(H17:H21)"

			G30.cell(row=16, column=9).value='=SUMIF(PBC_Disposals!A:A,"201",PBC_Disposals!H:H)'
			G30.cell(row=17, column=9).value='=SUMIF(PBC_Disposals!A:A,"203",PBC_Disposals!H:H)'
			G30.cell(row=18, column=9).value='=SUMIF(PBC_Disposals!A:A,"205",PBC_Disposals!H:H)+SUMIF(PBC_Disposals!A:A,"208",PBC_Disposals!H:H)'
			G30.cell(row=19, column=9).value='=SUMIF(PBC_Disposals!A:A,"207",PBC_Disposals!H:H)'
			G30.cell(row=20, column=9).value='=SUMIF(PBC_Disposals!A:A,"206",PBC_Disposals!H:H)'
			G30.cell(row=21, column=9).value="=SUM(I16:I20)"

			G30.cell(row=16, column=10).value='=IF(I16<0,C16+F16+I16,C16+F16-I16)'
			G30.cell(row=17, column=10).value='=IF(I17<0,C17+F17+I17,C17+F17-I17)'
			G30.cell(row=18, column=10).value='=IF(I18<0,C18+F18+I18,C18+F18-I18)'
			G30.cell(row=19, column=10).value='=IF(I19<0,C19+F19+I19,C19+F19-I19)'
			G30.cell(row=20, column=10).value='=IF(I20<0,C20+F20+I20,C20+F20-I20)'
			G30.cell(row=21, column=10).value="=SUM(J16:J20)"


			G30.cell(row=16, column=11).value='=SUMIF(F10_TB!B:B,"201",F10_TB!H:H)'
			G30.cell(row=17, column=11).value='=SUMIF(F10_TB!B:B,"203",F10_TB!H:H)'
			G30.cell(row=18, column=11).value='=SUMIF(F10_TB!B:B,"205",F10_TB!H:H)+SUMIF(F10_TB!B:B,"208",F10_TB!H:H)'
			G30.cell(row=19, column=11).value='=SUMIF(F10_TB!B:B,"207",F10_TB!H:H)'
			G30.cell(row=20, column=11).value='=SUMIF(F10_TB!B:B,"206",F10_TB!H:H)'
			G30.cell(row=21, column=11).value="=SUM(K16:K20)"

			G30.cell(row=16, column=12).value='=J16-K16'
			G30.cell(row=17, column=12).value='=J17-K17'
			G30.cell(row=18, column=12).value='=J18-K18'
			G30.cell(row=19, column=12).value='=J19-K19'
			G30.cell(row=20, column=12).value='=J20-K20'
			G30.cell(row=21, column=12).value="=SUM(L16:L20)"

			G30.cell(row=25, column=3).value='=SUMIF(F10_TB!B:B,"211",F10_TB!E:E)'
			G30.cell(row=26, column=3).value='=SUMIF(F10_TB!B:B,"212",F10_TB!E:E)'
			G30.cell(row=27, column=3).value='=SUMIF(F10_TB!B:B,"213",F10_TB!E:E)+SUMIF(F10_TB!B:B,"223",F10_TB!E:E)'
			G30.cell(row=28, column=3).value='=SUMIF(F10_TB!B:B,"214",F10_TB!E:E)+SUMIF(F10_TB!B:B,"224",F10_TB!E:E)'
			G30.cell(row=29, column=3).value='=SUMIF(F10_TB!B:B,"215",F10_TB!E:E)'
			G30.cell(row=30, column=3).value='=SUMIF(F10_TB!B:B,"216",F10_TB!E:E)'
			G30.cell(row=31, column=3).value='=SUMIF(F10_TB!B:B,"217",F10_TB!E:E)+SUMIF(F10_TB!B:B,"227",F10_TB!E:E)'
			G30.cell(row=32, column=3).value='=SUM(C25:C31)'

			G30.cell(row=25, column=4).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"211",PBC_Add!H:H,"Pure")'
			G30.cell(row=26, column=4).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"212",PBC_Add!H:H,"Pure")'
			G30.cell(row=27, column=4).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"213",PBC_Add!H:H,"Pure")+SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"223",PBC_Add!H:H,"Pure")'
			G30.cell(row=28, column=4).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"214",PBC_Add!H:H,"Pure")+SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"224",PBC_Add!H:H,"Pure")'
			G30.cell(row=29, column=4).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"215",PBC_Add!H:H,"Pure")'
			G30.cell(row=30, column=4).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"216",PBC_Add!H:H,"Pure")'
			G30.cell(row=31, column=4).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"217",PBC_Add!H:H,"Pure")+SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"227",PBC_Add!H:H,"Pure")'
			G30.cell(row=32, column=4).value='=SUM(D25:D31)'

			G30.cell(row=25, column=5).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"211",PBC_Add!H:H,"Transfer")'
			G30.cell(row=26, column=5).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"212",PBC_Add!H:H,"Transfer")'
			G30.cell(row=27, column=5).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"213",PBC_Add!H:H,"Transfer")+SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"223",PBC_Add!H:H,"Transfer")'
			G30.cell(row=28, column=5).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"214",PBC_Add!H:H,"Transfer")+SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"224",PBC_Add!H:H,"Transfer")'
			G30.cell(row=29, column=5).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"215",PBC_Add!H:H,"Transfer")'
			G30.cell(row=30, column=5).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"216",PBC_Add!H:H,"Transfer")'
			G30.cell(row=31, column=5).value='=SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"217",PBC_Add!H:H,"Transfer")+SUMIFS(PBC_Add!C:C,PBC_Add!G:G,"227",PBC_Add!H:H,"Transfer")'
			G30.cell(row=32, column=5).value='=SUM(E25:E31)'

			G30.cell(row=25, column=6).value="=SUM(D25:E25)"
			G30.cell(row=26, column=6).value="=SUM(D26:E26)"
			G30.cell(row=27, column=6).value="=SUM(D27:E27)"
			G30.cell(row=28, column=6).value="=SUM(D28:E28)"
			G30.cell(row=29, column=6).value="=SUM(D29:E29)"
			G30.cell(row=30, column=6).value="=SUM(D30:E30)"
			G30.cell(row=31, column=6).value="=SUM(D31:E31)"
			G30.cell(row=32, column=6).value="=SUM(F25:F31)"

			G30.cell(row=25, column=7).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"211",PBC_Disposals!K:K,"Sale")'
			G30.cell(row=26, column=7).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"212",PBC_Disposals!K:K,"Sale")'
			G30.cell(row=27, column=7).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"213",PBC_Disposals!K:K,"Sale")+SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"223",PBC_Disposals!K:K,"Sale")'
			G30.cell(row=28, column=7).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"214",PBC_Disposals!K:K,"Sale")+SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"224",PBC_Disposals!K:K,"Sale")'
			G30.cell(row=29, column=7).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"215",PBC_Disposals!K:K,"Sale")'
			G30.cell(row=30, column=7).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"216",PBC_Disposals!K:K,"Sale")'
			G30.cell(row=31, column=7).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"217",PBC_Disposals!K:K,"Sale")+SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"227",PBC_Disposals!K:K,"Sale")'
			G30.cell(row=32, column=7).value="=SUM(G25:G31)"


			# G30.cell(row=25, column=8).value=0
			# G30.cell(row=26, column=8).value=0
			# G30.cell(row=27, column=8).value=0
			# G30.cell(row=28, column=8).value=0
			# G30.cell(row=29, column=8).value=0
			# G30.cell(row=30, column=8).value=0
			# G30.cell(row=31, column=8).value="=SUM(H25:H30)"

			G30.cell(row=25, column=8).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"211",PBC_Disposals!K:K,"Scrap")'
			G30.cell(row=26, column=8).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"212",PBC_Disposals!K:K,"Scrap")'
			G30.cell(row=27, column=8).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"213",PBC_Disposals!K:K,"Scrap")+SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"223",PBC_Disposals!K:K,"Scrap")'
			G30.cell(row=28, column=8).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"214",PBC_Disposals!K:K,"Scrap")+SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"224",PBC_Disposals!K:K,"Scrap")'
			G30.cell(row=29, column=8).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"215",PBC_Disposals!K:K,"Scrap")'
			G30.cell(row=30, column=8).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"216",PBC_Disposals!K:K,"Scrap")'
			G30.cell(row=31, column=8).value='=SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"217",PBC_Disposals!K:K,"Scrap")+SUMIFS(PBC_Disposals!H:H,PBC_Disposals!A:A,"227",PBC_Disposals!K:K,"Scrap")'
			G30.cell(row=32, column=8).value="=SUM(H25:H31)"

			G30.cell(row=25, column=9).value='=SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!H:H)'
			G30.cell(row=26, column=9).value='=SUMIF(PBC_Disposals!A:A,"212",PBC_Disposals!H:H)'
			G30.cell(row=27, column=9).value='=SUMIF(PBC_Disposals!A:A,"213",PBC_Disposals!H:H)+SUMIF(PBC_Disposals!A:A,"223",PBC_Disposals!H:H)'
			G30.cell(row=28, column=9).value='=SUMIF(PBC_Disposals!A:A,"214",PBC_Disposals!H:H)+SUMIF(PBC_Disposals!A:A,"224",PBC_Disposals!H:H)'
			G30.cell(row=29, column=9).value='=SUMIF(PBC_Disposals!A:A,"215",PBC_Disposals!H:H)'
			G30.cell(row=30, column=9).value='=SUMIF(PBC_Disposals!A:A,"216",PBC_Disposals!H:H)'
			G30.cell(row=31, column=9).value='=SUMIF(PBC_Disposals!A:A,"217",PBC_Disposals!H:H)+SUMIF(PBC_Disposals!A:A,"227",PBC_Disposals!H:H)'
			G30.cell(row=32, column=9).value="=SUM(I25:I31)"

			G30.cell(row=25, column=10).value='=IF(I25<0,C25+F25+I25,C25+F25-I25)'
			G30.cell(row=26, column=10).value='=IF(I26<0,C26+F26+I26,C26+F26-I26)'
			G30.cell(row=27, column=10).value='=IF(I27<0,C27+F27+I27,C27+F27-I27)'
			G30.cell(row=28, column=10).value='=IF(I28<0,C28+F28+I28,C28+F28-I28)'
			G30.cell(row=29, column=10).value='=IF(I29<0,C29+F29+I29,C29+F29-I29)'
			G30.cell(row=30, column=10).value='=IF(I30<0,C30+F30+I30,C30+F30-I30)'
			G30.cell(row=31, column=10).value='=IF(I31<0,C31+F31+I31,C31+F31-I31)'
			G30.cell(row=32, column=10).value="=SUM(J25:J31)"


			G30.cell(row=25, column=11).value='=SUMIF(F10_TB!B:B,"211",F10_TB!H:H)'
			G30.cell(row=26, column=11).value='=SUMIF(F10_TB!B:B,"212",F10_TB!H:H)'
			G30.cell(row=27, column=11).value='=SUMIF(F10_TB!B:B,"213",F10_TB!H:H)+SUMIF(F10_TB!B:B,"223",F10_TB!H:H)'
			G30.cell(row=28, column=11).value='=SUMIF(F10_TB!B:B,"214",F10_TB!H:H)+SUMIF(F10_TB!B:B,"224",F10_TB!H:H)'
			G30.cell(row=29, column=11).value='=SUMIF(F10_TB!B:B,"215",F10_TB!H:H)'
			G30.cell(row=30, column=11).value='=SUMIF(F10_TB!B:B,"216",F10_TB!H:H)'
			G30.cell(row=31, column=11).value='=SUMIF(F10_TB!B:B,"217",F10_TB!H:H)+SUMIF(F10_TB!B:B,"227",F10_TB!H:H)'
			G30.cell(row=32, column=11).value="=SUM(K25:K31)"

			G30.cell(row=25, column=12).value='=J25-K25'
			G30.cell(row=26, column=12).value='=J26-K26'
			G30.cell(row=27, column=12).value='=J27-K27'
			G30.cell(row=28, column=12).value='=J28-K28'
			G30.cell(row=29, column=12).value='=J29-K29'
			G30.cell(row=30, column=12).value='=J30-K30'
			G30.cell(row=31, column=12).value='=J31-K31'
			G30.cell(row=32, column=12).value="=SUM(L25:L31)"

			# G30.cell(row=34,column=3).value="=C22+C32"
			# G30.cell(row=34,column=4).value="=D22+D32"
			# G30.cell(row=34,column=5).value="=E22+E32"
			# G30.cell(row=34,column=6).value="=F22+F32"
			# # G30.cell(row=34,column=7).value="=G22+G32"
			# # G30.cell(row=34,column=8).value="=H22+H32"
			# G30.cell(row=34,column=9).value="=I22+I32"
			# G30.cell(row=34,column=10).value="=J22+J32"
			# G30.cell(row=34,column=11).value="=K22+K32"
			# G30.cell(row=34,column=12).value="=L22+L32"


			G30.cell(row=39, column=3).value='=SUMIF(F10_TB!B:B,"231",F10_TB!E:E)'
			G30.cell(row=40, column=3).value='=SUMIF(F10_TB!B:B,"235",F10_TB!E:E)'
			G30.cell(row=41,column=3).value='=SUM(C39:C40)'

			# G30.cell(row=39,column=4).value='=SUMIF(PBC_CIP!F:F,"231",PBC_CIP!C:C)'
			# G30.cell(row=40,column=4).value='=SUMIF(PBC_CIP!F:F,"235",PBC_CIP!C:C)'
			G30.cell(row=41,column=4).value='=SUM(D39:D40)'

			G30.cell(row=39,column=6).value='=SUM(D39:E39)'
			G30.cell(row=40,column=6).value='=SUM(D40:E40)'
			G30.cell(row=41,column=6).value='=SUM(F39:F40)'

			# G30.cell(row=39,column=7).value='=SUMIFS(PBC_CIP!D:D,PBC_CIP!F:F,"231",PBC_CIP!E:E,"Sales")+SUMIFS(PBC_CIP!D:D,PBC_CIP!F:F,"231",PBC_CIP!E:E,"Scrapped")'
			# G30.cell(row=40,column=7).value='=SUMIFS(PBC_CIP!D:D,PBC_CIP!F:F,"235",PBC_CIP!E:E,"Sales")+SUMIFS(PBC_CIP!D:D,PBC_CIP!F:F,"235",PBC_CIP!E:E,"Scrapped")'
			# G30.cell(row=41,column=7).value='=SUM(G39:G40)'

			# G30.cell(row=39,column=8).value='=SUMIFS(PBC_CIP!D:D,PBC_CIP!F:F,"231",PBC_CIP!E:E,"Transfer")'
			# G30.cell(row=40,column=8).value='=SUMIFS(PBC_CIP!D:D,PBC_CIP!F:F,"235",PBC_CIP!E:E,"Transfer")'
			G30.cell(row=41,column=8).value='=SUM(H39:H40)'

			G30.cell(row=39,column=9).value='=SUM(G39:H39)'
			G30.cell(row=40,column=9).value='=SUM(G40:H40)'
			G30.cell(row=41,column=9).value='=SUM(I39:I40)'

			G30.cell(row=39,column=10).value='=IF(I39<0,C39+F39+I39,C39+F39-I39)'
			G30.cell(row=40,column=10).value='=IF(I40<0,C40+F40+I40,C40+F40-I40)'
			G30.cell(row=41,column=10).value='=SUM(J39:J40)'

			G30.cell(row=39,column=11).value='=SUMIF(F10_TB!B:B,"231",F10_TB!H:H)'
			G30.cell(row=40,column=11).value='=SUMIF(F10_TB!B:B,"235",F10_TB!H:H)'
			G30.cell(row=41,column=11).value='=SUM(K39:K40)'

			G30.cell(row=39,column=12).value='=J39-K39'
			G30.cell(row=40,column=12).value='=J40-K40'
			G30.cell(row=41,column=12).value='=SUM(L39:L40)'
			#################################################




			G30.cell(row=44, column=3).value='=C21+C32+C41'
			G30.cell(row=44, column=6).value='=F21+F32+F41'
			G30.cell(row=44, column=9).value='=I21+I32+I41'
			G30.cell(row=44, column=10).value='=J21+J32+J41'
			G30.cell(row=44, column=11).value='=K21+K32+K41'
			G30.cell(row=44, column=12).value='=L21+L32+L41'
			for i in range(1,13):
				G30.cell(row=44,column=i).fill=total
			G30.cell(row=50,column=3).value='=SUMIF(F10_TB!J:J,"2801",F10_TB!E:E)'
			G30.cell(row=51,column=3).value='=SUMIF(F10_TB!J:J,"2803",F10_TB!E:E)'
			G30.cell(row=52,column=3).value='=SUMIF(F10_TB!J:J,"2805",F10_TB!E:E)+SUMIF(F10_TB!J:J,"2808",F10_TB!E:E)'
			G30.cell(row=53,column=3).value='=SUMIF(F10_TB!J:J,"2807",F10_TB!E:E)'
			G30.cell(row=54,column=3).value='=SUMIF(F10_TB!J:J,"2806",F10_TB!E:E)'
			G30.cell(row=55,column=3).value='=SUM(C50:C54)'

			G30.cell(row=50,column=6).value='=IF(SUMIF(PBC_FAR!F:F,"2801",PBC_FAR!I:I)>0,-SUMIF(PBC_FAR!F:F,"2801",PBC_FAR!I:I),SUMIF(PBC_FAR!F:F,"2801",PBC_FAR!I:I))+IF(SUMIF(PBC_Disposals!A:A,"201",PBC_Disposals!I:I)>0,-SUMIF(PBC_Disposals!A:A,"201",PBC_Disposals!I:I),SUMIF(PBC_Disposals!A:A,"201",PBC_Disposals!I:I))'
			G30.cell(row=51,column=6).value='=IF(SUMIF(PBC_FAR!F:F,"2803",PBC_FAR!I:I)>0,-SUMIF(PBC_FAR!F:F,"2803",PBC_FAR!I:I),SUMIF(PBC_FAR!F:F,"2803",PBC_FAR!I:I))+IF(SUMIF(PBC_Disposals!A:A,"203",PBC_Disposals!I:I)>0,-SUMIF(PBC_Disposals!A:A,"201",PBC_Disposals!I:I),SUMIF(PBC_Disposals!A:A,"203",PBC_Disposals!I:I))'
			G30.cell(row=52,column=6).value='=IF(SUMIF(PBC_FAR!F:F,"2805",PBC_FAR!I:I)+SUMIF(PBC_FAR!F:F,"2808",PBC_FAR!I:I)>0,-SUMIF(PBC_FAR!F:F,"2805",PBC_FAR!I:I)-SUMIF(PBC_FAR!F:F,"2808",PBC_FAR!I:I),SUMIF(PBC_FAR!F:F,"2805",PBC_FAR!I:I)+SUMIF(PBC_FAR!F:F,"2808",PBC_FAR!I:I))+IF(SUMIF(PBC_Disposals!A:A,"205",PBC_Disposals!I:I)>0,-SUMIF(PBC_Disposals!A:A,"205",PBC_Disposals!I:I),SUMIF(PBC_Disposals!A:A,"205",PBC_Disposals!I:I))+IF(SUMIF(PBC_Disposals!A:A,"208",PBC_Disposals!I:I)>0,-SUMIF(PBC_Disposals!A:A,"208",PBC_Disposals!I:I),SUMIF(PBC_Disposals!A:A,"208",PBC_Disposals!I:I))'
			G30.cell(row=53,column=6).value='=IF(SUMIF(PBC_FAR!F:F,"2807",PBC_FAR!I:I)>0,-SUMIF(PBC_FAR!F:F,"2807",PBC_FAR!I:I),SUMIF(PBC_FAR!F:F,"2807",PBC_FAR!I:I))+IF(SUMIF(PBC_Disposals!A:A,"207",PBC_Disposals!I:I)>0,-SUMIF(PBC_Disposals!A:A,"207",PBC_Disposals!I:I),SUMIF(PBC_Disposals!A:A,"207",PBC_Disposals!I:I))'
			G30.cell(row=54,column=6).value='=IF(SUMIF(PBC_FAR!F:F,"2806",PBC_FAR!I:I)>0,-SUMIF(PBC_FAR!F:F,"2806",PBC_FAR!I:I),SUMIF(PBC_FAR!F:F,"2806",PBC_FAR!I:I))+IF(SUMIF(PBC_Disposals!A:A,"206",PBC_Disposals!I:I)>0,-SUMIF(PBC_Disposals!A:A,"206",PBC_Disposals!I:I),SUMIF(PBC_Disposals!A:A,"206",PBC_Disposals!I:I))'
			G30.cell(row=55,column=6).value='=SUM(F50:F54)'

			G30.cell(row=50,column=9).value='=SUMIF(PBC_Disposals!L:L,"2801",PBC_Disposals!J:J)'
			G30.cell(row=51,column=9).value='=SUMIF(PBC_Disposals!L:L,"2803",PBC_Disposals!J:J)'
			G30.cell(row=52,column=9).value='=SUMIF(PBC_Disposals!L:L,"2805",PBC_Disposals!J:J)+SUMIF(PBC_Disposals!L:L,"2808",PBC_Disposals!J:J)'
			G30.cell(row=53,column=9).value='=SUMIF(PBC_Disposals!L:L,"2807",PBC_Disposals!J:J)'
			G30.cell(row=54,column=9).value='=SUMIF(PBC_Disposals!L:L,"2806",PBC_Disposals!J:J)'
			G30.cell(row=55,column=9).value='=SUM(I50:I54)'

			G30.cell(row=50,column=10).value='=C50+F50+I50'
			G30.cell(row=51,column=10).value='=C51+F51+I51'
			G30.cell(row=52,column=10).value='=C52+F52+I52'
			G30.cell(row=53,column=10).value='=C53+F53+I53'
			G30.cell(row=54,column=10).value='=C54+F54+I54'
			G30.cell(row=55,column=10).value='=SUM(J50:J54)'


			G30.cell(row=50,column=11).value='=SUMIF(F10_TB!J:J,"2801",F10_TB!H:H)'
			G30.cell(row=51,column=11).value='=SUMIF(F10_TB!J:J,"2803",F10_TB!H:H)'
			G30.cell(row=52,column=11).value='=SUMIF(F10_TB!J:J,"2805",F10_TB!H:H)+SUMIF(F10_TB!J:J,"2808",F10_TB!H:H)'
			G30.cell(row=53,column=11).value='=SUMIF(F10_TB!J:J,"2807",F10_TB!H:H)'
			G30.cell(row=54,column=11).value='=SUMIF(F10_TB!J:J,"2806",F10_TB!H:H)'
			G30.cell(row=55,column=11).value='=SUM(K50:K54)'


			G30.cell(row=50,column=12).value='=J50-K50'
			G30.cell(row=51,column=12).value='=J51-K51'
			G30.cell(row=52,column=12).value='=J52-K52'
			G30.cell(row=53,column=12).value='=J53-K53'
			G30.cell(row=54,column=12).value='=J54-K54'
			G30.cell(row=55,column=12).value='=SUM(L50:L54)'

			G30.cell(row=59,column=3).value='=SUMIF(F10_TB!J:J,"2811",F10_TB!E:E)'
			G30.cell(row=60,column=3).value='=SUMIF(F10_TB!J:J,"2812",F10_TB!E:E)'
			G30.cell(row=61,column=3).value='=SUMIF(F10_TB!J:J,"2813",F10_TB!E:E)'
			G30.cell(row=62,column=3).value='=SUMIF(F10_TB!J:J,"2814",F10_TB!E:E)'
			G30.cell(row=63,column=3).value='=SUMIF(F10_TB!J:J,"2815",F10_TB!E:E)'
			G30.cell(row=64,column=3).value='=SUMIF(F10_TB!J:J,"2816",F10_TB!E:E)'
			G30.cell(row=65,column=3).value='=SUMIF(F10_TB!J:J,"2817",F10_TB!E:E)'
			G30.cell(row=66,column=3).value='=SUM(C59:C65)'

			G30.cell(row=59,column=6).value='=IF(SUMIF(PBC_FAR!F:F,"2811",PBC_FAR!I:I)>0,-SUMIF(PBC_FAR!F:F,"2811",PBC_FAR!I:I),SUMIF(PBC_FAR!F:F,"2811",PBC_FAR!I:I))+IF(SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!I:I)>0,-SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!I:I),SUMIF(PBC_Disposals!A:A,"211",PBC_Disposals!I:I))'
			G30.cell(row=60,column=6).value='=IF(SUMIF(PBC_FAR!F:F,"2812",PBC_FAR!I:I)>0,-SUMIF(PBC_FAR!F:F,"2812",PBC_FAR!I:I),SUMIF(PBC_FAR!F:F,"2812",PBC_FAR!I:I))+IF(SUMIF(PBC_Disposals!A:A,"212",PBC_Disposals!I:I)>0,-SUMIF(PBC_Disposals!A:A,"212",PBC_Disposals!I:I),SUMIF(PBC_Disposals!A:A,"212",PBC_Disposals!I:I))'
			G30.cell(row=61,column=6).value='=IF(SUMIF(PBC_FAR!F:F,"2813",PBC_FAR!I:I)>0,-SUMIF(PBC_FAR!F:F,"2813",PBC_FAR!I:I),SUMIF(PBC_FAR!F:F,"2813",PBC_FAR!I:I))+IF(SUMIF(PBC_Disposals!A:A,"213",PBC_Disposals!I:I)>0,-SUMIF(PBC_Disposals!A:A,"213",PBC_Disposals!I:I),SUMIF(PBC_Disposals!A:A,"213",PBC_Disposals!I:I))'
			G30.cell(row=62,column=6).value='=IF(SUMIF(PBC_FAR!F:F,"2814",PBC_FAR!I:I)>0,-SUMIF(PBC_FAR!F:F,"2814",PBC_FAR!I:I),SUMIF(PBC_FAR!F:F,"2814",PBC_FAR!I:I))+IF(SUMIF(PBC_Disposals!A:A,"214",PBC_Disposals!I:I)>0,-SUMIF(PBC_Disposals!A:A,"214",PBC_Disposals!I:I),SUMIF(PBC_Disposals!A:A,"214",PBC_Disposals!I:I))'
			G30.cell(row=63,column=6).value='=IF(SUMIF(PBC_FAR!F:F,"2815",PBC_FAR!I:I)>0,-SUMIF(PBC_FAR!F:F,"2815",PBC_FAR!I:I),SUMIF(PBC_FAR!F:F,"2815",PBC_FAR!I:I))+IF(SUMIF(PBC_Disposals!A:A,"215",PBC_Disposals!I:I)>0,-SUMIF(PBC_Disposals!A:A,"215",PBC_Disposals!I:I),SUMIF(PBC_Disposals!A:A,"215",PBC_Disposals!I:I))'
			G30.cell(row=64,column=6).value='=IF(SUMIF(PBC_FAR!F:F,"2816",PBC_FAR!I:I)>0,-SUMIF(PBC_FAR!F:F,"2816",PBC_FAR!I:I),SUMIF(PBC_FAR!F:F,"2816",PBC_FAR!I:I))+IF(SUMIF(PBC_Disposals!A:A,"216",PBC_Disposals!I:I)>0,-SUMIF(PBC_Disposals!A:A,"216",PBC_Disposals!I:I),SUMIF(PBC_Disposals!A:A,"216",PBC_Disposals!I:I))'
			G30.cell(row=65,column=6).value='=IF(SUMIF(PBC_FAR!F:F,"2817",PBC_FAR!I:I)>0,-SUMIF(PBC_FAR!F:F,"2817",PBC_FAR!I:I),SUMIF(PBC_FAR!F:F,"2817",PBC_FAR!I:I))+IF(SUMIF(PBC_Disposals!A:A,"217",PBC_Disposals!I:I)>0,-SUMIF(PBC_Disposals!A:A,"217",PBC_Disposals!I:I),SUMIF(PBC_Disposals!A:A,"217",PBC_Disposals!I:I))'
			G30.cell(row=66,column=6).value='=SUM(F59:F65)'#AICI STOP

			G30.cell(row=59,column=9).value='=SUMIF(PBC_Disposals!L:L,"2811",PBC_Disposals!J:J)'
			G30.cell(row=60,column=9).value='=SUMIF(PBC_Disposals!L:L,"2812",PBC_Disposals!J:J)'
			G30.cell(row=61,column=9).value='=SUMIF(PBC_Disposals!L:L,"2813",PBC_Disposals!J:J)'
			G30.cell(row=62,column=9).value='=SUMIF(PBC_Disposals!L:L,"2814",PBC_Disposals!J:J)'
			G30.cell(row=63,column=9).value='=SUMIF(PBC_Disposals!L:L,"2815",PBC_Disposals!J:J)'
			G30.cell(row=64,column=9).value='=SUMIF(PBC_Disposals!L:L,"2816",PBC_Disposals!J:J)'
			G30.cell(row=65,column=9).value='=SUMIF(PBC_Disposals!L:L,"2817",PBC_Disposals!J:J)'
			G30.cell(row=66,column=9).value='=SUM(I59:I65)'

			G30.cell(row=59,column=10).value='=C59+F59+I59'
			G30.cell(row=60,column=10).value='=C60+F60+I60'
			G30.cell(row=61,column=10).value='=C61+F61+I61'
			G30.cell(row=62,column=10).value='=C62+F62+I62'
			G30.cell(row=63,column=10).value='=C63+F63+I63'
			G30.cell(row=64,column=10).value='=C64+F64+I64'
			G30.cell(row=65,column=10).value='=C65+F65+I65'
			G30.cell(row=66,column=10).value='=SUM(J59:J65)'

			G30.cell(row=59,column=11).value='=SUMIF(F10_TB!J:J,"2811",F10_TB!H:H)'
			G30.cell(row=60,column=11).value='=SUMIF(F10_TB!J:J,"2812",F10_TB!H:H)'
			G30.cell(row=61,column=11).value='=SUMIF(F10_TB!J:J,"2813",F10_TB!H:H)'
			G30.cell(row=62,column=11).value='=SUMIF(F10_TB!J:J,"2814",F10_TB!H:H)'
			G30.cell(row=63,column=11).value='=SUMIF(F10_TB!J:J,"2815",F10_TB!H:H)'
			G30.cell(row=64,column=11).value='=SUMIF(F10_TB!J:J,"2816",F10_TB!H:H)'
			G30.cell(row=65,column=11).value='=SUMIF(F10_TB!J:J,"2817",F10_TB!H:H)'
			G30.cell(row=66,column=11).value='=SUM(K59:K65)'

			G30.cell(row=59,column=12).value='=J59-K59'
			G30.cell(row=60,column=12).value='=J60-K60'
			G30.cell(row=61,column=12).value='=J61-K61'
			G30.cell(row=62,column=12).value='=J62-K62'
			G30.cell(row=63,column=12).value='=J63-K63'
			G30.cell(row=64,column=12).value='=J64-K64'
			G30.cell(row=65,column=12).value='=J65-K65'
			G30.cell(row=66,column=12).value='=SUM(L59:L65)'

			G30.cell(row=67,column=3).value="=C55+C66"
			G30.cell(row=67,column=6).value="=F55+F66"
			G30.cell(row=67,column=9).value="=I55+I66"
			G30.cell(row=67,column=10).value="=J55+J66"
			G30.cell(row=67,column=11).value="=K55+K66"
			G30.cell(row=67,column=12).value="=L55+L66"







			G30.cell(row=74,column=3).value="=C16+C50"
			G30.cell(row=75,column=3).value="=C17+C51"
			G30.cell(row=76,column=3).value="=C18+C52"
			G30.cell(row=77,column=3).value="=C19+C53"
			G30.cell(row=78,column=3).value="=C20+C54"


			G30.cell(row=80,column=3).value="=SUM(C74:C79)"
			G30.cell(row=80,column=6).value="=SUM(F74:F79)"

			G30.cell(row=83,column=3).value="=C25+C59+C26+C60"
			G30.cell(row=84,column=3).value="=C27+C61"
			G30.cell(row=85,column=3).value="=C28+C62"
			G30.cell(row=86,column=3).value="=C29+C63"
			G30.cell(row=87,column=3).value="=C30+C64"
			G30.cell(row=88,column=3).value="=C31+C65"

			G30.cell(row=90,column=3).value="=C39"
			G30.cell(row=91,column=3).value="=C40"

			G30.cell(row=92,column=3).value="=SUM(C83:C91)"
			G30.cell(row=92,column=6).value="=SUM(F83:F91)"

			G30.cell(row=74,column=6).value="=J16+J50"
			G30.cell(row=75,column=6).value="=J17+J51"
			G30.cell(row=76,column=6).value="=J18+J52"
			G30.cell(row=77,column=6).value="=J19+J53"
			G30.cell(row=78,column=6).value="=J20+J54"

			# G30.cell(row=111,column=6).value="=J45"

			G30.cell(row=83,column=6).value="=J25+J59+J26+J60"
			G30.cell(row=84,column=6).value="=J27+J61"
			G30.cell(row=85,column=6).value="=J28+J62"
			G30.cell(row=86,column=6).value="=J29+J63"
			G30.cell(row=87,column=6).value="=J30+J64"
			G30.cell(row=88,column=6).value="=J31+J65"

			# G30.cell(row=120,column=6).value="=J46"
			G30.cell(row=90,column=6).value="=J39"
			G30.cell(row=91,column=6).value="=J40"

			G30.cell(row=94,column=3).value="=C80+C92"
			G30.cell(row=94,column=6).value="=F80+F92"

			for i in range(17,126):
			  G30.cell(row=i,column=3).number_format="#,##0_);(#,##0)"
			  G30.cell(row=i,column=4).number_format="#,##0_);(#,##0)"
			  G30.cell(row=i,column=5).number_format="#,##0_);(#,##0)"
			  G30.cell(row=i,column=6).number_format="#,##0_);(#,##0)"
			  G30.cell(row=i,column=7).number_format="#,##0_);(#,##0)"
			  G30.cell(row=i,column=8).number_format="#,##0_);(#,##0)"
			  G30.cell(row=i,column=9).number_format="#,##0_);(#,##0)"
			  G30.cell(row=i,column=10).number_format="#,##0_);(#,##0)"
			  G30.cell(row=i,column=11).number_format="#,##0_);(#,##0)"
			  G30.cell(row=i,column=12).number_format="#,##0_);(#,##0)"

			for i in range(1,13):
			  G30.cell(row=21,column=i).font=font2
			  G30.cell(row=32,column=i).font=font2
			  G30.cell(row=41,column=i).font=font2
			  G30.cell(row=47,column=i).font=font2
			  G30.cell(row=55,column=i).font=font2
			  G30.cell(row=66,column=i).font=font2
			  G30.cell(row=67,column=i).font=font2
			  G30.cell(row=80,column=i).font=font2
			  G30.cell(row=92,column=i).font=font2
			  G30.cell(row=94,column=i).font=font2
			  G30.cell(row=96,column=i).font=font2

			for i in range(15,103):
			  G30.cell(row=i,column=12).font=fontRed



			G30.cell(row=37,column=6).border=rightmedium
			G30.cell(row=37,column=9).border=rightmedium

			for i in range(17,22):
			  G30.cell(row=i,column=12).font=fontRedDiff
			for i in range(25,32):
			  G30.cell(row=i,column=12).font=fontRedDiff
			for i in range(39,41):
			  G30.cell(row=i,column=12).font=fontRedDiff
			for i in range(45,47):
			  G30.cell(row=i,column=12).font=fontRedDiff
			for i in range(55,60):
			  G30.cell(row=i,column=12).font=fontRedDiff
			for i in range(63,70):
			  G30.cell(row=i,column=12).font=fontRedDiff
			for i in range(79,82):
			  G30.cell(row=i,column=12).font=fontRedDiff
			for i in range(85,94):
			  G30.cell(row=i,column=12).font=fontRedDiff


			for i in range(16,21):
			  for j in range(1,12):
				  G30.cell(row=i,column=j).font=font1
			for i in range(25,32):
			  for j in range(1,12):
				  G30.cell(row=i,column=j).font=font1
			for i in range(39,41):
			  for j in range(1,12):
				  G30.cell(row=i,column=j).font=font1
			for i in range(50,55):
			  for j in range(1,12):
				  G30.cell(row=i,column=j).font=font1
			for i in range(59,66):
			  for j in range(1,12):
				  G30.cell(row=i,column=j).font=font1
			for i in range(74,80):
			  for j in range(1,12):
				  G30.cell(row=i,column=j).font=font1
			for i in range(83,92):
			  for j in range(1,12):
				  G30.cell(row=i,column=j).font=font1


			G30.column_dimensions.group('D','E',hidden=True)
			G30.column_dimensions.group('G','H',hidden=True)

			G30.cell(row=42,column=7).value="Check"
			G30.cell(row=42,column=8).value="=IF(OR(H41<0),E32+E22+H41,E32+E22-H41)"

			G30.cell(row=42,column=7).font=fontRed
			G30.cell(row=42,column=8).font=fontRed



			for i in range(1,13):
			  G30.cell(row=15,column=i).alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
			  G30.cell(row=24,column=i).alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
			  G30.cell(row=38,column=i).alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
			  G30.cell(row=44,column=i).alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
			  G30.cell(row=49,column=i).alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
			  G30.cell(row=58,column=i).alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
			  G30.cell(row=73,column=i).alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)


			G30.column_dimensions['A'].width=27
			G30.column_dimensions['B'].width=32

			G30.sheet_view.showGridLines = False


		balanta=FAR.create_sheet('Trial Balance')
		mr = TB1.max_row
		mc = TB1.max_column
		# copying the cell values from source
		# excel file to destination excel file
		for i in range (1, mr + 1):
			for j in range (1, mc + 1):
			# reading cell value from source excel file
				c = TB1.cell(row = i, column = j)

			# writing the read value to destination excel file
				balanta.cell(row = i, column = j).value = c.value

		folderpath = "home/auditappnexia/output/far/"
		folder_path=""
		os.mkdir(folderpath+namec)
		print(val1,val2,val3)
		if(val1==1 and val2==1 and val3==1):
			print("Yes")
			folder_path=os.path.join(folderpath+namec, 'G10, 20, 30 FA Reconciliation, TOM, Depreciation Expenses.xlsx')
			myorder=[4,5,6,0,9,1,2,3,7,8]
		if(val1==1 and val2==0 and val3==0):
			print("Yes")            
			myorder=[4,0,1,2,3]
			folder_path=os.path.join(folderpath+namec, 'G10 FA Reconciliation.xlsx')
		if(val1==0 and val2==1 and val3==0):
			print("Yes")            
			myorder=[4,0,1,2,3]
			folder_path=os.path.join(folderpath+namec, 'G20 FA Depreciation Expenses.xlsx')
		if(val1==0 and val2==0 and val3==1):
			print("Yes")            
			myorder=[4,0,1,2,3,5,6,7]
			folder_path=os.path.join(folderpath+namec, 'G30 TOM.xlsx')
		if(val1==1 and val2==1 and val3==0):
			print("Yes")            
			folder_path=os.path.join(folderpath+namec, 'G10,20 FA Reconciliation,Depreciation Expenses.xlsx')
			myorder=[4,5,0,1,2,3,6]
		if(val1==1 and val2==0 and val3==1):
			print("Yes")            
			myorder=[4,5,0,1,2,3,6,7,8]
			folder_path=os.path.join(folderpath+namec, 'G10,30 FA Reconciliation,TOM.xlsx')
		if(val1==0 and val2==1 and val3==1):
			print("Yes")            
			myorder=[4,5,0,1,2,3,6,7,8]
			folder_path=os.path.join(folderpath+namec, 'G20,30 Depreciation Expenses,TOM.xlsx')
		# FAR._sheets =[FAR._sheets[i] for i in myorder]



		FAR.save(folder_path)
		make_archive(folder_path,"home/auditappnexia/output/far/FAR "+namec+".zip")
		return send_from_directory(folderpath,"FAR "+namec+".zip", as_attachment=True)
	return "0"      


@app.route('/JournalEntries/Instructions', methods=['GET'])
def downloadJE():
		filepath = "/home/auditappnexia/output/je"
 
		return send_from_directory(filepath,"Instructions - Journal Entries.docx", as_attachment=True)  
@app.route('/JournalEntries/GTv5Wjk1PboB5kRZ8elnBUECePO')
def JE():
	return render_template('JE.html')
@app.route('/JournalEntries/GTv5Wjk1PboB5kRZ8elnBUECePO', methods=['POST', 'GET'])
def JE_process():

	namec = request.form['client']
	ant= datetime.datetime.strptime(
					 request.form['yearEnd'],
					 '%Y-%m-%d')
	preparedBy1=request.form['preparedBy']
	datePrepared1= datetime.datetime.strptime(
					 request.form['preparedDate'],
					 '%Y-%m-%d')
	refference1 = request.form['reff']
	threshol = request.form['thers']
	isChecked1=request.form.get("JEReco")
	isChecked2=request.form.get("MPL")
	isChecked3=request.form.get("AccountDetail")
	print(isChecked3,isChecked2,isChecked1,namec,ant,threshol)
	# if isChecked1=="": #daca e bifat
	#     isChecked=1
	# else:
	#     isChecked=0
	
	folderpath="home/auditappnexia/output/je"
	os.mkdir(folderpath+"\\"+namec)
	def make_archive(source, destination):
		base = os.path.basename(destination)
		name = base.split('.')[0]
		format = base.split('.')[1]
		archive_from = os.path.dirname(source)
		archive_to = os.path.basename(source.strip(os.sep))
		shutil.make_archive(name, format, archive_from, archive_to)
		shutil.move('%s.%s'%(name,format), destination)
	# yearEnd = str(request.form['yearEnd'])
	# processed_text = client.upper()
	# fisier=request.files.get('monthlyTB')
	if request.method == 'POST':
		workingsblue2= Font(bold=True, italic=True, name='Tahoma', size=8,color='FFFFFF')
		lbluefill = PatternFill(start_color='7030A0',
							end_color='7030A0',
							fill_type='solid')
		grifill=PatternFill(start_color='c4d79b',end_color='c4d79b',fill_type='solid')
		yellow=PatternFill(start_color='ffff00',end_color='ffff00',fill_type='solid')
		blueFill = PatternFill(start_color='00AEAC',
							end_color='00AEAC',
							fill_type='solid')
		doubleborder = Border(bottom=Side(style='double'))
		solidborder = Border(bottom=Side(style='thick'))
		solidborderstanga = Border(left=Side(style='thin'))
		rightborder = Border(right=Side(style='thin'))
		rightdouble = Border (right=Side(style='thin'), bottom=Side(style='double'))
		rightmedium = Border (right=Side(style='thin'), bottom=Side(style='medium'))
		solidborderdreapta = Border(right=Side(style='thin'))
		solidbordersus = Border(top=Side(style='thin'))
		fontitalic = Font(name='Tahoma', size=8, bold=True, italic=True)
		font1 = Font(name='Tahoma', size=8)
		font2 = Font(name='Tahoma', size=8, bold=True)
		fontRed = Font(name='Tahoma', size=8, bold=True, color= 'FF0000')
		fontRedDiff=Font(name="Tahoma", color='FF0000', size=11, )
		fontGT = Font (name='GT Logo', size=8)
		workingsblue = Font(color='2F75B5', bold=True, name='Tahoma', size=8 )
		headers= Font(bold=True, italic=True, name='Tahoma', size=8,color='FFFFFF') 
		headersblue = PatternFill(start_color='7030A0',
						end_color='7030A0',
						fill_type='solid')
		headerspurple= PatternFill(start_color='65CDCC',
							end_color='65CDCC',
							fill_type='solid')
		total=PatternFill(start_color='DDD9C4',
						end_color='DDD9C4',
						fill_type='solid')
		greenbolditalic= Font(bold=True, italic=True,  color='C0504D', name='Tahoma', size=8)
		greenbolditalic= Font(bold=True, italic=True,  color='00af50')
		fontalb = Font(italic=True, color="bfbfbf", size=8, name='Tahoma')
		trialb=request.files["TB"]
		je=request.files["JE"]

		tb=openpyxl.load_workbook(je,data_only='True')
		tb1=tb.active

		TB=openpyxl.load_workbook(trialb,data_only='True')
		TBls1=TB.active
		for row in TBls1.iter_rows():
					for cell in row:
						if cell.value=="Account":
							Acc=cell.column
							acr=cell.row

		for row in TBls1.iter_rows():
			for cell in row:
				if cell.value=="Description":
					desc=cell.column

		for row in TBls1.iter_rows():
			for cell in row:
				if cell.value=="OB":
					aob=cell.column

		for row in TBls1.iter_rows():

			for cell in row:
				if cell.value=="DM":
					adm=cell.column
				
		for row in TBls1.iter_rows():
			for cell in row:
				if cell.value=="CM":
					acm=cell.column

		for row in TBls1.iter_rows():
			for cell in row:
				if cell.value=="CB":
					acb=cell.column

		try:
			lungl=len(TBls1[Acc])
		except:
			flash("Please insert the correct header for Account in Trial Balance file")
			return render_template("index.html")



		try:
			lungl=len(TBls1[desc])
		except:
			flash("Please insert the correct header for Description in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: TB. Please insert the correct header for 'Description'")
			# sys.exit()

		try:
			lungl=len(TBls1[aob])
		except:
			flash("Please insert the correct header for OB in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: TB. Please insert the correct header for 'Opening Balance'")
			# sys.exit()

		try:
			lungl=len(TBls1[adm])
		except:
			flash("Please insert the correct header for DM in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: TB. Please insert the correct header for 'Debit Movement'")
			# sys.exit()

		try:
			lungl=len(TBls1[acm])
		except:
			flash("Please insert the correct header for CM in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: TB. Please insert the correct header for 'Credit Movement'")
			# sys.exit()

		try:
			lungl=len(TBls1[acb])
		except:
			flash("Please insert the correct header for CB in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: TB. Please insert the correct header for 'Closing Balance'")
			# sys.exit()
		try:
			luntb=len(TBls1[Acc])
		except:
			flash("Please insert the correct header for Account in Trial Balance file")
			return render_template("index.html")

		Account=[b.value for b in TBls1[Acc][acr:luntb+1]]
		Description=[b.value for b in TBls1[desc][acr:luntb+1]]
		OB=[b.value for b in TBls1[aob][acr:luntb+1]]
		DM=[b.value for b in TBls1[adm][acr:luntb+1]]
		CM=[b.value for b in TBls1[acm][acr:luntb+1]]
		CB=[b.value for b in TBls1[acb][acr:luntb+1]]


		for row in tb1.iter_rows():
			for cell in row :

				if cell.value=="Account Debit" :
					
					glac=cell.column
					glrow=cell.row
		for row in tb1.iter_rows():
			for cell in row :

				if cell.value=="Account Credit" :
					
					glacc=cell.column
					glrow=cell.row


		for row in tb1.iter_rows():
			for cell in row :

				if cell.value=="Description" :
					
					gld=cell.column
					gldrow=cell.row
		for row in tb1.iter_rows():
			for cell in row :

				if cell.value=="JE Number" :
					
					glje=cell.column
					gldrow=cell.row

		for row in tb1.iter_rows():
			for cell in row :

				if cell.value=="Date" :
					
					glde=cell.column
					gldrow=cell.row

		for row in tb1.iter_rows():
			for cell in row :

				if cell.value=="Amount Debit" :
					
					glamd=cell.column
					glarow=cell.row

		for row in tb1.iter_rows():
			for cell in row :

				if cell.value=="Amount Credit" :
					
					glamc=cell.column
					glarow=cell.row


		try:
			lungl=len(tb1[glac])
		except:
			flash("Please insert the correct header for Account GL Debit in Journal Entries file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: G/L. Please insert the correct header for 'Account'")
			# sys.exit()



		try:
			accountgldebit=[b.value for b in tb1[glac][glrow:lungl]]
		except:
			flash("Please insert the correct header for Account GL Debit in Journal Entries file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: G/L. Please insert the correct header for 'Account Debit'")
			# sys.exit()

		try:
			accountglcredit=[b.value for b in tb1[glacc][glrow:lungl]]
		except:
			flash("Please insert the correct header for Account GL Debit in Journal Entries file")
			return render_template("index.html")


		try:
			gldate=[b.value for b in tb1[glde][glrow:lungl]]
		except:
			flash("Please insert the correct header for Date in Journal Entries file")
			return render_template("index.html")

		try:
			gljenr=[b.value for b in tb1[glje][glrow:lungl]]
		except:
			flash("Please insert the correct header for JE Number in Journal Entries file")
			return render_template("index.html")

		try:
			glamountdebit=[b.value for b in tb1[glamd][glrow:lungl]]
			glamountcredit=[b.value for b in tb1[glamc][glrow:lungl]]
		except:
			for row in tb1.iter_rows():
				for cell in row :

					if cell.value=="Amount" :
						
						glamc=cell.column
						glarow=cell.row
			glamountdebit=[b.value for b in tb1[glamc][glrow:lungl]]
			glamountcredit=[b.value for b in tb1[glamc][glrow:lungl]]



		syntaccgldebit1=[]
		syntaccglcredit1=[]


		for i in range(0,len(accountgldebit)):
				syntaccgldebit1.append(str(accountgldebit[i])[0:3])
		for i in range(0,len(accountglcredit)):
				syntaccglcredit1.append(str(accountglcredit[i])[0:3])

		syntaccgldebit=[]
		syntaccglcredit=[]

		for i in range(0,len(syntaccgldebit1)):
			try:
				syntaccgldebit1[i]=int(syntaccgldebit1[i])
				syntaccgldebit.append(syntaccgldebit1[i])
			except:
				pass
		for i in range(0,len(syntaccglcredit1)):
			try:
				syntaccglcredit1[i]=int(syntaccglcredit1[i])
				syntaccglcredit.append(syntaccglcredit1[i])
			except:
				pass
		lisunicaacd=list(set(syntaccgldebit))
		lisunicaacc=list(set(syntaccglcredit))
		listaunica=list(set(lisunicaacc+lisunicaacd))
		listasumec=[]
		listasumed=[]
		for i in range(0,len(listaunica)):
			sumad=0
			sumac=0
			for j in range(0,len(accountglcredit)):
				if(str(listaunica[i])==str(accountglcredit[j])[0:3]):
					try:
						sumac=sumac+glamountcredit[j]
					except:
						pass
				if(str(listaunica[i])==str(accountgldebit[j])[0:3]):
					try:
						sumad=sumad+glamountdebit[j]
					except:
						pass
			listasumed.append(sumad)
			listasumec.append(sumac)
		if isChecked1=="":
		 
			output=Workbook()
			out=output.active
			out.sheet_view.showGridLines = False

			out.title="Journal Entries Reconciliation"

			out.cell(row=1,column=1).value="Client:"
			out.cell(row=2,column=1).value="Period End:"

			out.cell(row=1,column=2).value=namec
			out.cell(row=2,column=2).value=ant
			out.cell(row=2,column=2).number_format='mm/dd/yyyy'

			out.cell(row=1,column=1).font=font2
			out.cell(row=2,column=1).font=font2
			out.cell(row=4,column=4).value="Reconciliation Trial Balance vs Journal Entries"
			out.cell(row=4,column=4).font=font2

			out.cell(row=7,column=3).value="As per JE"
			out.cell(row=7,column=3).font=font2
			out.cell(row=7,column=6).value="As per TB"
			out.cell(row=7,column=6).font=font2

			out.cell(row=7,column=9).value="Difference"
			out.cell(row=7,column=9).font=font2



			out.cell(row=10,column=2).value="Account"
			out.cell(row=10,column=3).value="Amount Debit"
			out.cell(row=10,column=4).value="Amount Credit"

			out.cell(row=10,column=6).value="Amount Debit"
			out.cell(row=10,column=7).value="Amount Credit"

			out.cell(row=10,column=9).value="Amount Debit"
			out.cell(row=10,column=10).value="Amount Credit"

			for o in range(2,5):
				out.cell(row=10,column=o).border=doubleborder
				out.cell(row=10,column=o).font=font2
				out.cell(row=10,column=o).fill=headerspurple

			for o in range(6,8):
				out.cell(row=10,column=o).border=doubleborder
				out.cell(row=10,column=o).font=font2
				out.cell(row=10,column=o).fill=headerspurple

			for o in range(9,11):
				out.cell(row=10,column=o).border=doubleborder
				out.cell(row=10,column=o).font=font2
				out.cell(row=10,column=o).fill=headerspurple


			out2=output.create_sheet("Trial Balance")
			out2.sheet_view.showGridLines = False

			out2.cell(row=1,column=1).value="Synt(3)"
			out2.cell(row=1,column=2).value="Account"
			out2.cell(row=1,column=3).value="Description"
			out2.cell(row=1,column=4).value="OB"
			out2.cell(row=1,column=5).value="DM"
			out2.cell(row=1,column=6).value="CM"
			out2.cell(row=1,column=7).value="CB"

			for o in range(1,8):
				out2.cell(row=1,column=o).border=doubleborder
				out2.cell(row=1,column=o).font=font2
				out2.cell(row=1,column=o).fill=headerspurple

			for k in range(0,len(Account)):
				out2.cell(row=2+k,column=2).value=Account[k]
				out2.cell(row=2+k,column=1).value=str(Account[k])[:3]
				out2.cell(row=2+k,column=3).value=Description[k]
				out2.cell(row=2+k,column=4).value=OB[k]
				out2.cell(row=2+k,column=5).value=DM[k]
				out2.cell(row=2+k,column=6).value=CM[k]
				out2.cell(row=2+k,column=7).value=CB[k]

			listaunica, listasumed, listasumec = (list(t) for t in zip(*sorted(zip(listaunica, listasumed,listasumec))))

			for l in range(0,len(listaunica)):
				out.cell(row=11+l,column=2).value=listaunica[l]
				out.cell(row=11+l,column=3).value=listasumed[l]
				out.cell(row=11+l,column=3).number_format='#,##0_);(#,##0)'
				out.cell(row=11+l,column=4).value=listasumec[l]
				out.cell(row=11+l,column=4).number_format='#,##0_);(#,##0)'
				out.cell(row=11+l,column=6).value="=SUMIF('Trial Balance'!A:A,B"+str(11+l)+",'Trial Balance'!E:E)"
				out.cell(row=11+l,column=6).number_format='#,##0_);(#,##0)'
				out.cell(row=11+l,column=7).value="=SUMIF('Trial Balance'!A:A,B"+str(11+l)+",'Trial Balance'!F:F)"
				out.cell(row=11+l,column=7).number_format='#,##0_);(#,##0)'
				out.cell(row=11+l,column=9).value="=F"+str(11+l)+"-C"+str(11+l)
				out.cell(row=11+l,column=9).font=fontRedDiff
				out.cell(row=11+l,column=9).number_format='#,##0_);(#,##0)'
				out.cell(row=11+l,column=10).value="=G"+str(11+l)+"-D"+str(11+l)
				out.cell(row=11+l,column=10).number_format='#,##0_);(#,##0)'
				out.cell(row=11+l,column=10).font=fontRedDiff

			out.cell(row=6,column=2).value="Total"
			out.cell(row=6,column=2).font=font2

			out.cell(row=7,column=3).value="Check"
			out.cell(row=7,column=3).font=fontRedDiff

			out.cell(row=7,column=6).value="Check"
			out.cell(row=7,column=6).font=fontRedDiff

			out.cell(row=7,column=9).value="Check"
			out.cell(row=7,column=9).font=fontRedDiff


			out.cell(row=6,column=3).value="=sum(C11:C"+str(len(listaunica)+11)+")"
			out.cell(row=6,column=3).font=font2
			out.cell(row=6,column=3).number_format='#,##0_);(#,##0)'

			out.cell(row=6,column=4).value="=sum(D11:D"+str(len(listaunica)+11)+")"
			out.cell(row=6,column=4).font=font2
			out.cell(row=6,column=4).number_format='#,##0_);(#,##0)'

			out.cell(row=6,column=6).value="=sum(F11:F"+str(len(listaunica)+11)+")"
			out.cell(row=6,column=6).font=font2
			out.cell(row=6,column=6).number_format='#,##0_);(#,##0)'

			out.cell(row=6,column=7).value="=sum(G11:G"+str(len(listaunica)+11)+")"
			out.cell(row=6,column=7).font=font2
			out.cell(row=6,column=7).number_format='#,##0_);(#,##0)'

			out.cell(row=6,column=9).value="=sum(I11:I"+str(len(listaunica)+11)+")"
			out.cell(row=6,column=9).font=fontRedDiff
			out.cell(row=6,column=9).number_format='#,##0_);(#,##0)'

			out.cell(row=6,column=10).value="=sum(J11:J"+str(len(listaunica)+11)+")"
			out.cell(row=6,column=10).font=fontRedDiff
			out.cell(row=6,column=10).number_format='#,##0_);(#,##0)'


			out.cell(row=7,column=4).value="=C6-D6"
			out.cell(row=7,column=4).font=fontRedDiff
			out.cell(row=7,column=4).number_format='#,##0_);(#,##0)'
			out.cell(row=7,column=7).value="=F6-G6"
			out.cell(row=7,column=7).font=fontRedDiff
			out.cell(row=7,column=7).number_format='#,##0_);(#,##0)'

			out.cell(row=7,column=10).value="=I6-J6"
			out.cell(row=7,column=10).font=fontRedDiff
			out.cell(row=7,column=10).number_format='#,##0_);(#,##0)'

			c=out['B11']
			out.freeze_panes = c
			out.column_dimensions['B'].width=15
			out.column_dimensions['C'].width=15
			out.column_dimensions['D'].width=15
			out.column_dimensions['E'].width=15
			out.column_dimensions['F'].width=15
			out.column_dimensions['G'].width=15
			out.column_dimensions['H'].width=15
			out.column_dimensions['I'].width=15
			out.column_dimensions['J'].width=15

			out.auto_filter.ref = "B10:J"+str(out.max_row)
			
			output.save(folderpath+"\\"+str(namec)+"\\Reco JE "+namec+".xlsx")
			
		if isChecked2=="":
			
			outputmpl=Workbook()
			listaunicaanalitic=list(set(accountgldebit+accountglcredit))
			listaunicaanaliticint=[]
			for i in range(0,len(listaunicaanalitic)):
				try:
					listaunicaanalitic[i]=int(listaunicaanalitic[i])
					listaunicaanaliticint.append(listaunicaanalitic[i])
				except:
					pass

			listaunicaanaliticint.sort()

			mpl=outputmpl.active
			mpl.title="Monthly P&L"
			mpl.cell(row=1,column=1).value="Client"
			mpl.cell(row=1,column=1).font=font2
			mpl.cell(row=2,column=1).value="Period end:"
			mpl.cell(row=2,column=1).font=font2

			c=mpl['A14']
			mpl.freeze_panes = c
			mpl.sheet_view.showGridLines = False

			mpl.cell(row=4,column=1).value="Monthly P&L"
			mpl.cell(row=4,column=1).font=font2

			mpl.cell(row=1,column=2).value=namec
			mpl.cell(row=2,column=2).value=ant
			mpl.cell(row=2,column=2).number_format='mm/dd/yyyy'

			mpl.cell(row=6,column=4).value="Expenses"
			mpl.cell(row=7,column=4).value="Revenues"
			mpl.cell(row=8,column=4).value="Result"
			mpl.cell(row=9,column=4).value="Acc 121"
			mpl.cell(row=10,column=4).value="Check"
			mpl.cell(row=10,column=4).font=fontRedDiff

			mpl.cell(row=5,column=5).value="January"
			mpl.cell(row=5,column=6).value="February"
			mpl.cell(row=5,column=7).value="March"
			mpl.cell(row=5,column=8).value="April"
			mpl.cell(row=5,column=9).value="May"
			mpl.cell(row=5,column=10).value="June"
			mpl.cell(row=5,column=11).value="July"
			mpl.cell(row=5,column=12).value="August"
			mpl.cell(row=5,column=13).value="September"
			mpl.cell(row=5,column=14).value="October"
			mpl.cell(row=5,column=15).value="November"
			mpl.cell(row=5,column=16).value="December"
			mpl.cell(row=5,column=17).value="Total"
			mpl.cell(row=5,column=18).value="As per TB"
			mpl.cell(row=5,column=19).value="Difference"
			for p in range(5,20):
				mpl.cell(row=5,column=p).border=doubleborder
				mpl.cell(row=5,column=p).font=font2
				mpl.cell(row=5,column=p).fill=headerspurple

			for p in range(1,20):
				mpl.cell(row=13,column=p).border=doubleborder
				mpl.cell(row=13,column=p).font=font2
				mpl.cell(row=13,column=p).fill=headerspurple
			mpl.cell(row=13,column=1).value="Class"
			mpl.cell(row=13,column=2).value="Synt(3)"
			mpl.cell(row=13,column=3).value="Account"
			mpl.cell(row=13,column=4).value="Description"
			mpl.cell(row=13,column=5).value="January"
			mpl.cell(row=13,column=6).value="February"
			mpl.cell(row=13,column=7).value="March"
			mpl.cell(row=13,column=8).value="April"
			mpl.cell(row=13,column=9).value="May"
			mpl.cell(row=13,column=10).value="June"
			mpl.cell(row=13,column=11).value="July"
			mpl.cell(row=13,column=12).value="August"
			mpl.cell(row=13,column=13).value="September"
			mpl.cell(row=13,column=14).value="October"
			mpl.cell(row=13,column=15).value="November"
			mpl.cell(row=13,column=16).value="December"
			mpl.cell(row=13,column=17).value="Total"
			mpl.cell(row=13,column=18).value="As per TB"
			mpl.cell(row=13,column=19).value="Difference"
			rand=0

			mpltb=outputmpl.create_sheet("Trial Balance")
			mpltb.cell(row=1,column=1).value="Synt(3)"
			mpltb.cell(row=1,column=2).value="Account"
			mpltb.cell(row=1,column=3).value="Description"
			mpltb.cell(row=1,column=4).value="OB"
			mpltb.cell(row=1,column=5).value="DM"
			mpltb.cell(row=1,column=6).value="CM"
			mpltb.cell(row=1,column=7).value="CB"
			mpltb.sheet_view.showGridLines = False

			for k in range(0,len(Account)):
						mpltb.cell(row=2+k,column=2).value=Account[k]
						mpltb.cell(row=2+k,column=1).value=int(str(Account[k])[:3])
						mpltb.cell(row=2+k,column=3).value=Description[k]
						mpltb.cell(row=2+k,column=4).value=OB[k]
						mpltb.cell(row=2+k,column=5).value=DM[k]
						mpltb.cell(row=2+k,column=6).value=CM[k]
						mpltb.cell(row=2+k,column=7).value=CB[k]
						mpltb.cell(row=2+k,column=8).value=str(Account[k])[:1]

			mplje=outputmpl.create_sheet("JE")
			mplje.sheet_view.showGridLines = False

			mplje.cell(row=1,column=1).value="JE number"
			mplje.cell(row=1,column=2).value="Date"
			mplje.cell(row=1,column=3).value="Account Debit"
			mplje.cell(row=1,column=4).value="Account Credit"
			mplje.cell(row=1,column=5).value="Amount Debit"
			mplje.cell(row=1,column=6).value="Amount Credit"
			mplje.cell(row=1,column=7).value="Month"

			for k in range(0,len(gljenr)):
				mplje.cell(row=2+k,column=1).value=gljenr[k]
				mplje.cell(row=2+k,column=2).value=gldate[k]
				mplje.cell(row=2+k,column=3).value=str(accountgldebit[k])[:3]
				mplje.cell(row=2+k,column=4).value=str(accountglcredit[k])[:3]

				mplje.cell(row=2+k,column=5).value=glamountdebit[k]
				mplje.cell(row=2+k,column=6).value=glamountcredit[k]
				mplje.cell(row=2+k,column=7).value="=month(B"+str(2+k)+")"



			rand=0
			for j in range(0,len(listaunica)):
				if(int(str(listaunica[j])[:1])>5):
					rand=rand+1
					mpl.cell(row=13+rand,column=5).number_format='#,##0_);(#,##0)'
					mpl.cell(row=13+rand,column=6).number_format='#,##0_);(#,##0)'
					mpl.cell(row=13+rand,column=7).number_format='#,##0_);(#,##0)'
					mpl.cell(row=13+rand,column=8).number_format='#,##0_);(#,##0)'
					mpl.cell(row=13+rand,column=9).number_format='#,##0_);(#,##0)'
					mpl.cell(row=13+rand,column=10).number_format='#,##0_);(#,##0)'
					mpl.cell(row=13+rand,column=11).number_format='#,##0_);(#,##0)'
					mpl.cell(row=13+rand,column=12).number_format='#,##0_);(#,##0)'
					mpl.cell(row=13+rand,column=13).number_format='#,##0_);(#,##0)'
					mpl.cell(row=13+rand,column=14).number_format='#,##0_);(#,##0)'
					mpl.cell(row=13+rand,column=15).number_format='#,##0_);(#,##0)'
					mpl.cell(row=13+rand,column=16).number_format='#,##0_);(#,##0)'
					mpl.cell(row=13+rand,column=17).number_format='#,##0_);(#,##0)'
					mpl.cell(row=13+rand,column=18).number_format='#,##0_);(#,##0)'
					mpl.cell(row=13+rand,column=19).number_format='#,##0_);(#,##0)'
					

					mpl.cell(row=13+rand,column=3).value=listaunica[j]
					mpl.cell(row=13+rand,column=2).value=str(listaunica[j])[:3]
					mpl.cell(row=13+rand,column=1).value=str(listaunica[j])[:1]
					mpl.cell(row=13+rand,column=4).value="=VLOOKUP(C"+str(13+rand)+",'Trial Balance'!A:C,3,0)"
					if(str(listaunica[j])[:1]=="6"):
						mpl.cell(row=13+rand,column=5).value="=SUMIFS(JE!E:E,JE!G:G,1,JE!C:C,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=6).value="=SUMIFS(JE!E:E,JE!G:G,2,JE!C:C,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=7).value="=SUMIFS(JE!E:E,JE!G:G,3,JE!C:C,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=8).value="=SUMIFS(JE!E:E,JE!G:G,4,JE!C:C,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=9).value="=SUMIFS(JE!E:E,JE!G:G,5,JE!C:C,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=10).value="=SUMIFS(JE!E:E,JE!G:G,6,JE!C:C,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=11).value="=SUMIFS(JE!E:E,JE!G:G,7,JE!C:C,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=12).value="=SUMIFS(JE!E:E,JE!G:G,8,JE!C:C,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=13).value="=SUMIFS(JE!E:E,JE!G:G,9,JE!C:C,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=14).value="=SUMIFS(JE!E:E,JE!G:G,10,JE!C:C,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=15).value="=SUMIFS(JE!E:E,JE!G:G,11,JE!C:C,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=16).value="=SUMIFS(JE!E:E,JE!G:G,12,JE!C:C,C"+str(rand+13)+")"
					else:
						mpl.cell(row=13+rand,column=5).value="=-SUMIFS(JE!F:F,JE!G:G,1,JE!D:D,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=6).value="=-SUMIFS(JE!F:F,JE!G:G,2,JE!D:D,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=7).value="=-SUMIFS(JE!F:F,JE!G:G,3,JE!D:D,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=8).value="=-SUMIFS(JE!F:F,JE!G:G,4,JE!D:D,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=9).value="=-SUMIFS(JE!F:F,JE!G:G,5,JE!D:D,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=10).value="=-SUMIFS(JE!F:F,JE!G:G,6,JE!D:D,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=11).value="=-SUMIFS(JE!F:F,JE!G:G,7,JE!D:D,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=12).value="=-SUMIFS(JE!F:F,JE!G:G,8,JE!D:D,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=13).value="=-SUMIFS(JE!F:F,JE!G:G,9,JE!D:D,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=14).value="=-SUMIFS(JE!F:F,JE!G:G,10,JE!D:D,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=15).value="=-SUMIFS(JE!F:F,JE!G:G,11,JE!D:D,C"+str(rand+13)+")"
						mpl.cell(row=13+rand,column=16).value="=-SUMIFS(JE!F:F,JE!G:G,12,JE!D:D,C"+str(rand+13)+")"
					mpl.cell(row=13+rand,column=17).value="=sum(E"+str(rand+13)+":P"+str(rand+13)+")"
					mpl.cell(row=13+rand,column=18).value="=SUMIF('Trial Balance'!A:A,+B"+str(rand+13)+",'Trial Balance'!G:G)"
					mpl.cell(row=13+rand,column=19).value="=Q"+str(rand+13)+"-R"+str(rand+13)
					mpl.cell(row=13+rand,column=19).font=fontRedDiff
			mpl.cell(row=6,column=5).value="=SUMIF($A:$A,6,E:E)"
			mpl.cell(row=6,column=6).value="=SUMIF($A:$A,6,F:F)"
			mpl.cell(row=6,column=7).value="=SUMIF($A:$A,6,G:G)"
			mpl.cell(row=6,column=8).value="=SUMIF($A:$A,6,H:H)"
			mpl.cell(row=6,column=9).value="=SUMIF($A:$A,6,I:I)"
			mpl.cell(row=6,column=10).value="=SUMIF($A:$A,6,J:J)"
			mpl.cell(row=6,column=11).value="=SUMIF($A:$A,6,K:K)"
			mpl.cell(row=6,column=12).value="=SUMIF($A:$A,6,L:L)"
			mpl.cell(row=6,column=13).value="=SUMIF($A:$A,6,M:M)"
			mpl.cell(row=6,column=14).value="=SUMIF($A:$A,6,N:N)"
			mpl.cell(row=6,column=15).value="=SUMIF($A:$A,6,O:O)"
			mpl.cell(row=6,column=16).value="=SUMIF($A:$A,6,P:P)"
			mpl.cell(row=7,column=5).value="=SUMIF($A:$A,7,E:E)"
			mpl.cell(row=7,column=6).value="=SUMIF($A:$A,7,F:F)"
			mpl.cell(row=7,column=7).value="=SUMIF($A:$A,7,G:G)"
			mpl.cell(row=7,column=8).value="=SUMIF($A:$A,7,H:H)"
			mpl.cell(row=7,column=9).value="=SUMIF($A:$A,7,I:I)"
			mpl.cell(row=7,column=10).value="=SUMIF($A:$A,7,J:J)"
			mpl.cell(row=7,column=11).value="=SUMIF($A:$A,7,K:K)"
			mpl.cell(row=7,column=12).value="=SUMIF($A:$A,7,L:L)"
			mpl.cell(row=7,column=13).value="=SUMIF($A:$A,7,M:M)"
			mpl.cell(row=7,column=14).value="=SUMIF($A:$A,7,N:N)"
			mpl.cell(row=7,column=15).value="=SUMIF($A:$A,7,O:O)"
			mpl.cell(row=7,column=16).value="=SUMIF($A:$A,7,P:P)"

			mpl.cell(row=8,column=5).value="=SUM(E6:E7)"
			mpl.cell(row=8,column=6).value="=SUM(F6:F7)"
			mpl.cell(row=8,column=7).value="=SUM(G6:G7)"
			mpl.cell(row=8,column=8).value="=SUM(H6:H7)"
			mpl.cell(row=8,column=9).value="=SUM(I6:I7)"
			mpl.cell(row=8,column=10).value="=SUM(J6:J7)"
			mpl.cell(row=8,column=11).value="=SUM(K6:K7)"
			mpl.cell(row=8,column=12).value="=SUM(L6:L7)"
			mpl.cell(row=8,column=13).value="=SUM(M6:M7)"
			mpl.cell(row=8,column=14).value="=SUM(N6:N7)"
			mpl.cell(row=8,column=15).value="=SUM(O6:O7)"
			mpl.cell(row=8,column=16).value="=SUM(P6:P7)"

			mpl.cell(row=9,column=5).value='=SUMIFS(JE!E:E,JE!G:G,1,JE!C:C,"121")-SUMIFS(JE!F:F,JE!G:G,1,JE!D:D,"121")'
			mpl.cell(row=9,column=6).value='=SUMIFS(JE!E:E,JE!G:G,2,JE!C:C,"121")-SUMIFS(JE!F:F,JE!G:G,2,JE!D:D,"121")'
			mpl.cell(row=9,column=7).value='=SUMIFS(JE!E:E,JE!G:G,3,JE!C:C,"121")-SUMIFS(JE!F:F,JE!G:G,3,JE!D:D,"121")'
			mpl.cell(row=9,column=8).value='=SUMIFS(JE!E:E,JE!G:G,4,JE!C:C,"121")-SUMIFS(JE!F:F,JE!G:G,4,JE!D:D,"121")'
			mpl.cell(row=9,column=9).value='=SUMIFS(JE!E:E,JE!G:G,5,JE!C:C,"121")-SUMIFS(JE!F:F,JE!G:G,5,JE!D:D,"121")'
			mpl.cell(row=9,column=10).value='=SUMIFS(JE!E:E,JE!G:G,6,JE!C:C,"121")-SUMIFS(JE!F:F,JE!G:G,6,JE!D:D,"121")'
			mpl.cell(row=9,column=11).value='=SUMIFS(JE!E:E,JE!G:G,7,JE!C:C,"121")-SUMIFS(JE!F:F,JE!G:G,7,JE!D:D,"121")'
			mpl.cell(row=9,column=12).value='=SUMIFS(JE!E:E,JE!G:G,8,JE!C:C,"121")-SUMIFS(JE!F:F,JE!G:G,8,JE!D:D,"121")'
			mpl.cell(row=9,column=13).value='=SUMIFS(JE!E:E,JE!G:G,9,JE!C:C,"121")-SUMIFS(JE!F:F,JE!G:G,9,JE!D:D,"121")'
			mpl.cell(row=9,column=14).value='=SUMIFS(JE!E:E,JE!G:G,10,JE!C:C,"121")-SUMIFS(JE!F:F,JE!G:G,10,JE!D:D,"121")'
			mpl.cell(row=9,column=15).value='=SUMIFS(JE!E:E,JE!G:G,11,JE!C:C,"121")-SUMIFS(JE!F:F,JE!G:G,11,JE!D:D,"121")'
			mpl.cell(row=9,column=16).value='=SUMIFS(JE!E:E,JE!G:G,12,JE!C:C,"121")-SUMIFS(JE!F:F,JE!G:G,12,JE!D:D,"121")'



			mpl.cell(row=10,column=5).value="=E8-E9"
			mpl.cell(row=10,column=6).value="=F8-F9"
			mpl.cell(row=10,column=7).value="=G8-G9"
			mpl.cell(row=10,column=8).value="=H8-H9"
			mpl.cell(row=10,column=9).value="=I8-I9"
			mpl.cell(row=10,column=10).value="=J8-J9"
			mpl.cell(row=10,column=11).value="=K8-K9"
			mpl.cell(row=10,column=12).value="=L8-L9"
			mpl.cell(row=10,column=13).value="=M8-M9"
			mpl.cell(row=10,column=14).value="=N8-N9"
			mpl.cell(row=10,column=15).value="=O8-O9"
			mpl.cell(row=10,column=16).value="=P8-P9"
			mpl.cell(row=10,column=17).value="=Q8-Q9"
			mpl.cell(row=10,column=18).value="=R8-R9"
			mpl.cell(row=6,column=17).value="=sum(E6:P6)"
			mpl.cell(row=7,column=17).value="=sum(E7:P7)"
			mpl.cell(row=8,column=17).value="=sum(E8:P8)"
			mpl.cell(row=9,column=17).value="=sum(E9:P9)"
			mpl.cell(row=6,column=18).value="=SUMIF('Trial Balance'!H:H,6,'Trial Balance'!G:G)"
			mpl.cell(row=7,column=18).value="=SUMIF('Trial Balance'!H:H,7,'Trial Balance'!G:G)"
			mpl.cell(row=8,column=18).value="=SUM(R6:R7)"
			mpl.cell(row=9,column=18).value="=SUMIF('Trial Balance'!A:A,121,'Trial Balance'!G:G)"
			mpl.cell(row=6,column=19).value="=Q6-R6"
			mpl.cell(row=6,column=19).font=fontRedDiff
			mpl.cell(row=7,column=19).value="=Q7-R7"
			mpl.cell(row=7,column=19).font=fontRedDiff
			mpl.cell(row=8,column=19).font=fontRedDiff
			mpl.cell(row=9,column=19).font=fontRedDiff

			mpl.cell(row=8,column=19).value="=Q8-R8"
			mpl.cell(row=9,column=19).value="=Q9-R9"

			mpl.cell(row=1,column=1).value="Client:"
			mpl.cell(row=2,column=1).value="Period End:"

			mpl.cell(row=1,column=2).value=namec
			mpl.cell(row=2,column=2).value=ant
			mpl.cell(row=2,column=2).number_format='mm/dd/yyyy'

			mpl.cell(row=1,column=1).font=font2
			mpl.cell(row=2,column=1).font=font2
			for ko in range(5,20):
				mpl.cell(row=10,column=ko).font=fontRedDiff
				mpl.cell(row=9,column=ko).border=doubleborder
			for po in range(6,11):
				for pop in range(5,20):
					mpl.cell(row=po,column=pop).number_format='#,##0_);(#,##0)'
			mpl.auto_filter.ref = "A13:S"+str(mpl.max_row)
			
			
			outputmpl.save(folderpath+"\\"+str(namec)+"\\MonthlyPL JE" + " " + namec + ".xlsx")
			

		if isChecked3=="":
			
			for i in range(0,len(listaunica)):
				if(int(listasumed[i])>int(threshol) and int(str(listaunica[i])[:1])==6):
					excel=Workbook()
					ws=excel.active
					ws.title="Database"
					ws1=excel.create_sheet("Trial Balance")
					ws3=excel.create_sheet("Overview")
					ws.sheet_view.showGridLines = False
					ws1.sheet_view.showGridLines = False
					ws3.sheet_view.showGridLines = False

					for o in range(1,9):
						ws.cell(row=1,column=o).border=doubleborder
						ws.cell(row=1,column=o).font=font2
						ws.cell(row=1,column=o).fill=headerspurple
					ws.cell(row=1,column=1).value="JE Number"
					ws.cell(row=1,column=2).value="Date"
					ws.cell(row=1,column=3).value="Description"
					ws.cell(row=1,column=4).value="Account Debit"
					ws.cell(row=1,column=5).value="Account Credit"
					ws.cell(row=1,column=6).value="Amount Debit"
					ws.cell(row=1,column=7).value="Amount Credit"
					ws.cell(row=1,column=8).value="Month"
					row=1

					ws1.cell(row=1,column=1).value="Synt(3)"
					ws1.cell(row=1,column=2).value="Account"
					ws1.cell(row=1,column=3).value="Description"
					ws1.cell(row=1,column=4).value="OB"
					ws1.cell(row=1,column=5).value="DM"
					ws1.cell(row=1,column=6).value="CM"
					ws1.cell(row=1,column=7).value="CB"


					for k in range(0,len(Account)):
						ws1.cell(row=2+k,column=2).value=Account[k]
						ws1.cell(row=2+k,column=1).value=str(Account[k])[:3]
						ws1.cell(row=2+k,column=3).value=Description[k]
						ws1.cell(row=2+k,column=4).value=OB[k]
						ws1.cell(row=2+k,column=5).value=DM[k]
						ws1.cell(row=2+k,column=6).value=CM[k]
						ws1.cell(row=2+k,column=7).value=CB[k]
						ws1.cell(row=2+k,column=8).value=str(Account[k])[:1]

					listagldebit=[]
					for j in range(0,len(accountgldebit)):
						if(str(listaunica[i])==str(accountgldebit[j])[0:3]):
							ws.cell(row=row+1,column=1).value=gljenr[j]
							ws.cell(row=row+1,column=2).value=gldate[j]
							# ws.cell(row=row+1,column=3).value=gldate[j]
							ws.cell(row=row+1,column=4).value=accountgldebit[j]
							listagldebit.append(accountgldebit[j])
							ws.cell(row=row+1,column=5).value=accountglcredit[j]
							ws.cell(row=row+1,column=6).value=glamountdebit[j]
							ws.cell(row=row+1,column=7).value=glamountcredit[j]
							ws.cell(row=row+1,column=8).value="=month("+"B"+str(row+1)+")"
							row=row+1

					ws3.cell(row=1,column=1).value="Client"
					ws3.cell(row=1,column=1).font=font2
					ws3.cell(row=2,column=1).font=font2
					ws3.cell(row=5,column=2).font=font2
					ws3.cell(row=2,column=1).value="Period End"

					ws3.cell(row=1,column=2).value=namec
					ws3.cell(row=2,column=2).value=ant
					ws3.cell(row=2,column=2).number_format='mm/dd/yyyy'
					ws3.cell(row=5,column=2).value="Detail of account  " + str(listaunica[i])
					ws3.cell(row=8,column=1).value="Account"
					ws3.cell(row=8,column=2).value="Amount as per Detail"
					ws3.cell(row=8,column=3).value="Amount as per TB"
					ws3.cell(row=8,column=4).value="Difference"
					uniclistgldebit=list(set(listagldebit))
					for kk in range(0,len(uniclistgldebit)):
						ws3.cell(row=9+kk,column=1).value=uniclistgldebit[kk]
						ws3.cell(row=9+kk,column=2).value="=SUMIF(Database!D:D,"+"A"+str(kk+9)+",Database!F:F)"
						ws3.cell(row=9+kk,column=3).value="=SUMIF('Trial Balance'!B:B,"+"A"+str(kk+9)+",'Trial Balance'!G:G)"
						ws3.cell(row=9+kk,column=4).value="=C"+str(kk+9)+"-"+"B"+str(kk+9)
						ws3.cell(row=9+kk,column=4).font=fontRedDiff
						ws3.cell(row=9+kk,column=2).number_format='#,##0_);(#,##0)'
						ws3.cell(row=9+kk,column=3).number_format='#,##0_);(#,##0)'
						ws3.cell(row=9+kk,column=4).number_format='#,##0_);(#,##0)'
					ws3.cell(row=9+len(uniclistgldebit),column=1).value="Total"
					ws3.cell(row=9+len(uniclistgldebit),column=1).font=font2
					# print(uniclistgldebit)
					ws3.cell(row=9+len(uniclistgldebit),column=2).value="=sum(B9:B"+str(len(uniclistgldebit)+8)+")"
					ws3.cell(row=9+len(uniclistgldebit),column=2).font=font2
					ws3.cell(row=9+len(uniclistgldebit),column=2).number_format='#,##0_);(#,##0)'

					ws3.cell(row=9+len(uniclistgldebit),column=3).value="=sum(C9:C"+str(len(uniclistgldebit)+8)+")"
					ws3.cell(row=9+len(uniclistgldebit),column=3).font=font2
					ws3.cell(row=9+len(uniclistgldebit),column=3).number_format='#,##0_);(#,##0)'

					ws3.cell(row=9+len(uniclistgldebit),column=4).value="=sum(D9:D"+str(len(uniclistgldebit)+8)+")"
					ws3.cell(row=9+len(uniclistgldebit),column=4).font=fontRedDiff
					ws3.cell(row=9+len(uniclistgldebit),column=4).number_format='#,##0_);(#,##0)'

					for ii in range(1,5):
						ws3.cell(row=8+len(uniclistgldebit),column=ii).border=doubleborder
						

					randuri=len(uniclistgldebit)
					for ii in range(1,5):
						ws3.cell(row=8,column=ii).border=doubleborder
						ws3.cell(row=8,column=ii).font=font2
						ws3.cell(row=8,column=ii).fill=headerspurple

					for ii in range(1,15):
						ws3.cell(row=11+randuri,column=ii).border=doubleborder
						ws3.cell(row=11+randuri,column=ii).font=font2
						ws3.cell(row=11+randuri,column=ii).fill=headerspurple
						
					ws3.cell(row=11+randuri,column=1).value="Account"
					ws3.cell(row=11+randuri,column=2).value="January"
					ws3.cell(row=11+randuri,column=3).value="February"
					ws3.cell(row=11+randuri,column=4).value="March"
					ws3.cell(row=11+randuri,column=5).value="April"
					ws3.cell(row=11+randuri,column=6).value="May"
					ws3.cell(row=11+randuri,column=7).value="June"
					ws3.cell(row=11+randuri,column=8).value="July"
					ws3.cell(row=11+randuri,column=9).value="August"
					ws3.cell(row=11+randuri,column=10).value="September"
					ws3.cell(row=11+randuri,column=11).value="October"
					ws3.cell(row=11+randuri,column=12).value="November"
					ws3.cell(row=11+randuri,column=13).value="December"
					ws3.cell(row=11+randuri,column=14).value="Total"


					for z in range(0,len(uniclistgldebit)):
						ws3.cell(row=12+randuri+z,column=1).value=uniclistgldebit[z]
						ws3.cell(row=12+randuri+z,column=2).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=3).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=4).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=5).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=6).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=8).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=9).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=10).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=11).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=12).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=13).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=7).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=14).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=2).value="=SUMIFs(Database!F:F,Database!H:H,1,Database!D:D,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=3).value="=SUMIFs(Database!F:F,Database!H:H,2,Database!D:D,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=4).value="=SUMIFs(Database!F:F,Database!H:H,3,Database!D:D,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=5).value="=SUMIFs(Database!F:F,Database!H:H,4,Database!D:D,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=6).value="=SUMIFs(Database!F:F,Database!H:H,5,Database!D:D,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=7).value="=SUMIFs(Database!F:F,Database!H:H,6,Database!D:D,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=8).value="=SUMIFs(Database!F:F,Database!H:H,7,Database!D:D,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=9).value="=SUMIFs(Database!F:F,Database!H:H,8,Database!D:D,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=10).value="=SUMIFs(Database!F:F,Database!H:H,9,Database!D:D,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=11).value="=SUMIFs(Database!F:F,Database!H:H,10,Database!D:D,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=12).value="=SUMIFs(Database!F:F,Database!H:H,11,Database!D:D,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=13).value="=SUMIFs(Database!F:F,Database!H:H,12,Database!D:D,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=14).value="=sum(B"+str(12+randuri+z)+":M"+str(12+randuri+z)+")"
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=1).value="Total"
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=1).font=font2

					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=2).value="=sum(B"+str(12+randuri)+":B"+str(11+len(uniclistgldebit)+randuri)+")"
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=2).font=font2
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=2).number_format='#,##0_);(#,##0)'
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=3).value="=sum(C"+str(12+randuri)+":C"+str(11+len(uniclistgldebit)+randuri)+")"
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=3).font=font2
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=3).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=4).value="=sum(D"+str(12+randuri)+":D"+str(11+len(uniclistgldebit)+randuri)+")"
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=4).font=font2
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=4).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=5).value="=sum(E"+str(12+randuri)+":E"+str(11+len(uniclistgldebit)+randuri)+")"
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=5).font=font2
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=5).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=6).value="=sum(F"+str(12+randuri)+":F"+str(11+len(uniclistgldebit)+randuri)+")"
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=6).font=font2
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=6).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=7).value="=sum(G"+str(12+randuri)+":G"+str(11+len(uniclistgldebit)+randuri)+")"
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=7).font=font2
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=7).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=8).value="=sum(H"+str(12+randuri)+":H"+str(11+len(uniclistgldebit)+randuri)+")"
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=8).font=font2
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=8).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=9).value="=sum(I"+str(12+randuri)+":I"+str(11+len(uniclistgldebit)+randuri)+")"
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=9).font=font2
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=9).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=10).value="=sum(J"+str(12+randuri)+":J"+str(11+len(uniclistgldebit)+randuri)+")"
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=10).font=font2
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=10).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=11).value="=sum(K"+str(12+randuri)+":K"+str(11+len(uniclistgldebit)+randuri)+")"
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=11).font=font2
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=11).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=12).value="=sum(L"+str(12+randuri)+":L"+str(11+len(uniclistgldebit)+randuri)+")"
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=12).font=font2
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=12).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=13).value="=sum(M"+str(12+randuri)+":M"+str(11+len(uniclistgldebit)+randuri)+")"
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=13).font=font2
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=13).number_format='#,##0_);(#,##0)'
					
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=14).value="=sum(N"+str(12+randuri)+":N"+str(11+len(uniclistgldebit)+randuri)+")"
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=14).font=font2
					ws3.cell(row=12+len(uniclistgldebit)+randuri,column=14).number_format='#,##0_);(#,##0)'
					
					for ii in range(1,15):
						ws3.cell(row=11+len(uniclistgldebit)+randuri,column=ii).border=doubleborder

					ws3.column_dimensions['A'].width=15
					ws3.column_dimensions['B'].width=15
					ws3.column_dimensions['C'].width=15
					ws3.column_dimensions['D'].width=15
					ws3.column_dimensions['E'].width=15
					ws3.column_dimensions['F'].width=15
					ws3.column_dimensions['G'].width=15
					ws3.column_dimensions['H'].width=15
					ws3.column_dimensions['I'].width=15
					ws3.column_dimensions['J'].width=15
					ws3.column_dimensions['K'].width=15
					ws3.column_dimensions['L'].width=15
					ws3.column_dimensions['M'].width=15



					ws4=excel.create_sheet("Sheet")
					ws4.sheet_view.showGridLines = False
					ws4.cell(row=1,column=1).value="Month"
					ws4.cell(row=2,column=1).value="January"
					ws4.cell(row=3,column=1).value="February"
					ws4.cell(row=4,column=1).value="March"
					ws4.cell(row=5,column=1).value="April"
					ws4.cell(row=6,column=1).value="May"
					ws4.cell(row=7,column=1).value="June"
					ws4.cell(row=8,column=1).value="July"
					ws4.cell(row=9,column=1).value="August"
					ws4.cell(row=10,column=1).value="September"
					ws4.cell(row=11,column=1).value="October"
					ws4.cell(row=12,column=1).value="November"
					ws4.cell(row=13,column=1).value="December"
					ws4.cell(row=1,column=2).value="Monthly Trend"
					ws4.cell(row=2,column=2).value="=Overview!B"+str(12+len(uniclistgldebit)+randuri)
					ws4.cell(row=3,column=2).value="=Overview!C"+str(12+len(uniclistgldebit)+randuri)
					ws4.cell(row=4,column=2).value="=Overview!D"+str(12+len(uniclistgldebit)+randuri)
					ws4.cell(row=5,column=2).value="=Overview!E"+str(12+len(uniclistgldebit)+randuri)
					ws4.cell(row=6,column=2).value="=Overview!F"+str(12+len(uniclistgldebit)+randuri)
					ws4.cell(row=7,column=2).value="=Overview!G"+str(12+len(uniclistgldebit)+randuri)
					ws4.cell(row=8,column=2).value="=Overview!H"+str(12+len(uniclistgldebit)+randuri)
					ws4.cell(row=9,column=2).value="=Overview!I"+str(12+len(uniclistgldebit)+randuri)
					ws4.cell(row=10,column=2).value="=Overview!J"+str(12+len(uniclistgldebit)+randuri)
					ws4.cell(row=11,column=2).value="=Overview!K"+str(12+len(uniclistgldebit)+randuri)
					ws4.cell(row=12,column=2).value="=Overview!L"+str(12+len(uniclistgldebit)+randuri)
					ws4.cell(row=13,column=2).value="=Overview!M"+str(12+len(uniclistgldebit)+randuri)


					values = Reference(ws4,
							min_col=2,  # I
							max_col=2,  # T
							min_row=1,
							max_row=13)
					labels=Reference(ws4,
							min_col=1,  # I
							max_col=1,  # T
							min_row=2,
							max_row=13)

					chart = LineChart()
					chart.add_data(values, titles_from_data=True)
					chart.set_categories(labels)

					chart.title = "Monthly Amount for account " + str(listaunica[i])
					chart.x_axis.title = ""
					chart.y_axis.title = "" 
					ws3.add_chart(chart, "A"+str(14+len(uniclistgldebit)+randuri))
					ws4.sheet_state='hidden'
					myorder=[2,0,1,3]
					
					excel._sheets =[excel._sheets[i] for i in myorder]
					excel.save(folderpath+"/"+str(namec)+"/"+str(listaunica[i])+"db.xlsx")
					
				if(int(listasumec[i])>int(threshol) and int(str(listaunica[i])[:1])==7):
					excel=Workbook()
					ws=excel.active
					ws.title="Database"
					ws1=excel.create_sheet("Trial Balance")
					ws3=excel.create_sheet("Overview")

					for o in range(1,9):
						ws.cell(row=1,column=o).border=doubleborder
						ws.cell(row=1,column=o).font=font2
						ws.cell(row=1,column=o).fill=headerspurple
					ws.cell(row=1,column=1).value="JE Number"
					ws.cell(row=1,column=2).value="Date"
					ws.cell(row=1,column=3).value="Description"
					ws.cell(row=1,column=4).value="Account Debit"
					ws.cell(row=1,column=5).value="Account Credit"
					ws.cell(row=1,column=6).value="Amount Debit"
					ws.cell(row=1,column=7).value="Amount Credit"
					ws.cell(row=1,column=8).value="Month"
					row=1

					ws1.cell(row=1,column=1).value="Synt(3)"
					ws1.cell(row=1,column=2).value="Account"
					ws1.cell(row=1,column=3).value="Description"
					ws1.cell(row=1,column=4).value="OB"
					ws1.cell(row=1,column=5).value="DM"
					ws1.cell(row=1,column=6).value="CM"
					ws1.cell(row=1,column=7).value="CB"

					for k in range(0,len(Account)):
						ws1.cell(row=2+k,column=2).value=Account[k]
						ws1.cell(row=2+k,column=1).value=str(Account[k])[:3]
						ws1.cell(row=2+k,column=3).value=Description[k]
						ws1.cell(row=2+k,column=4).value=OB[k]
						ws1.cell(row=2+k,column=5).value=DM[k]
						ws1.cell(row=2+k,column=6).value=CM[k]
						ws1.cell(row=2+k,column=7).value=CB[k]
					listaglcredit=[]
					for j in range(0,len(accountglcredit)):
						if(str(listaunica[i])==str(accountglcredit[j])[0:3]):
							ws.cell(row=row+1,column=1).value=gljenr[j]
							ws.cell(row=row+1,column=2).value=gldate[j]
							ws.cell(row=row+1,column=4).value=accountgldebit[j]
							ws.cell(row=row+1,column=5).value=accountglcredit[j]
							listaglcredit.append(accountglcredit[j])
							ws.cell(row=row+1,column=6).value=glamountdebit[j]
							ws.cell(row=row+1,column=7).value=glamountcredit[j]
							ws.cell(row=row+1,column=8).value="=month("+"B"+str(row+1)+")"
							row=row+1
					ws3.cell(row=1,column=1).value="Client"	
					ws3.cell(row=1,column=1).font=font2
					ws3.cell(row=2,column=1).font=font2
					ws3.cell(row=5,column=2).font=font2	
					ws3.cell(row=2,column=1).value="Period End"
					ws3.cell(row=1,column=2).value=namec
					ws3.cell(row=2,column=2).value=ant
					ws3.cell(row=2,column=2).number_format='mm/dd/yyyy'
					ws3.cell(row=5,column=2).value="Detail of account " + str(listaunica[i])
					ws3.cell(row=8,column=1).value="Account"
					ws3.cell(row=8,column=2).value="Amount as per Detail"
					ws3.cell(row=8,column=3).value="Amount as per TB"
					ws3.cell(row=8,column=4).value="Difference"
					uniclistglcredit=list(set(listaglcredit))
					for z in range(0,len(uniclistglcredit)):
						ws3.cell(row=9+z,column=1).value=uniclistglcredit[z]
						ws3.cell(row=9+z,column=2).value="=-SUMIF(Database!E:E,"+"A"+str(z+9)+",Database!G:G)"
						ws3.cell(row=9+z,column=2).number_format='#,##0_);(#,##0)'
						ws3.cell(row=9+z,column=3).value="=SUMIF('Trial Balance'!B:B,"+"A"+str(z+9)+",'Trial Balance'!G:G)"
						ws3.cell(row=9+z,column=3).number_format='#,##0_);(#,##0)'
						ws3.cell(row=9+z,column=4).value="=C"+str(z+9)+"-"+"B"+str(z+9)
						ws3.cell(row=9+z,column=4).font=fontRedDiff
						ws3.cell(row=9+z,column=4).number_format='#,##0_);(#,##0)'
					for ii in range(1,5):
						ws3.cell(row=8,column=ii).border=doubleborder
						ws3.cell(row=8,column=ii).font=font2
						ws3.cell(row=8,column=ii).fill=headerspurple
					ws3.cell(row=9+len(uniclistglcredit),column=1).value="Total"
					ws3.cell(row=9+len(uniclistglcredit),column=1).font=font2

					ws3.cell(row=9+len(uniclistglcredit),column=2).value="=sum(B9:B"+str(len(uniclistglcredit)+8)+")"
					ws3.cell(row=9+len(uniclistglcredit),column=2).font=font2
					ws3.cell(row=9+len(uniclistglcredit),column=2).number_format='#,##0_);(#,##0)'

					ws3.cell(row=9+len(uniclistglcredit),column=3).value="=sum(C9:C"+str(len(uniclistglcredit)+8)+")"
					ws3.cell(row=9+len(uniclistglcredit),column=3).font=font2
					ws3.cell(row=9+len(uniclistglcredit),column=3).number_format='#,##0_);(#,##0)'

					ws3.cell(row=9+len(uniclistglcredit),column=4).value="=sum(D9:D"+str(len(uniclistglcredit)+8)+")"
					ws3.cell(row=9+len(uniclistglcredit),column=4).font=fontRedDiff
					ws3.cell(row=9+len(uniclistglcredit),column=4).number_format='#,##0_);(#,##0)'

					for ii in range(1,5):
						ws3.cell(row=8+len(uniclistglcredit),column=ii).border=doubleborder
					randuri=len(uniclistglcredit)

					for ii in range(1,15):
						ws3.cell(row=11+randuri,column=ii).border=doubleborder
						ws3.cell(row=11+randuri,column=ii).font=font2
						ws3.cell(row=11+randuri,column=ii).fill=headerspurple

					ws3.cell(row=11+randuri,column=1).value="Account"
					ws3.cell(row=11+randuri,column=2).value="January"
					ws3.cell(row=11+randuri,column=3).value="February"
					ws3.cell(row=11+randuri,column=4).value="March"
					ws3.cell(row=11+randuri,column=5).value="April"
					ws3.cell(row=11+randuri,column=6).value="May"
					ws3.cell(row=11+randuri,column=7).value="June"
					ws3.cell(row=11+randuri,column=8).value="July"
					ws3.cell(row=11+randuri,column=9).value="August"
					ws3.cell(row=11+randuri,column=10).value="September"
					ws3.cell(row=11+randuri,column=11).value="October"
					ws3.cell(row=11+randuri,column=12).value="November"
					ws3.cell(row=11+randuri,column=13).value="December"
					ws3.cell(row=11+randuri,column=14).value="Total"

					for z in range(0,len(uniclistglcredit)):
						ws3.cell(row=12+randuri+z,column=2).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=3).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=4).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=5).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=6).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=8).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=9).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=10).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=11).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=12).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=13).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=14).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=7).number_format='#,##0_);(#,##0)'
						ws3.cell(row=12+randuri+z,column=1).value=uniclistglcredit[z]
						ws3.cell(row=12+randuri+z,column=2).value="=SUMIFs(Database!G:G,Database!H:H,1,Database!E:E,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=3).value="=SUMIFs(Database!G:G,Database!H:H,2,Database!E:E,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=4).value="=SUMIFs(Database!G:G,Database!H:H,3,Database!E:E,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=5).value="=SUMIFs(Database!G:G,Database!H:H,4,Database!E:E,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=6).value="=SUMIFs(Database!G:G,Database!H:H,5,Database!E:E,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=7).value="=SUMIFs(Database!G:G,Database!H:H,6,Database!E:E,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=8).value="=SUMIFs(Database!G:G,Database!H:H,7,Database!E:E,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=9).value="=SUMIFs(Database!G:G,Database!H:H,8,Database!E:E,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=10).value="=SUMIFs(Database!G:G,Database!H:H,9,Database!E:E,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=11).value="=SUMIFs(Database!G:G,Database!H:H,10,Database!E:E,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=12).value="=SUMIFs(Database!G:G,Database!H:H,11,Database!E:E,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=13).value="=SUMIFs(Database!G:G,Database!H:H,12,Database!E:E,A"+str(12+randuri+z)+")"
						ws3.cell(row=12+randuri+z,column=14).value="=sum(B"+str(12+randuri+z)+":M"+str(12+randuri+z)+")"
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=1).value="Total"
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=1).font=font2

					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=2).value="=sum(B"+str(12+randuri)+":B"+str(11+len(uniclistglcredit)+randuri)+")"
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=2).font=font2
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=2).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=3).value="=sum(C"+str(12+randuri)+":C"+str(11+len(uniclistglcredit)+randuri)+")"
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=3).font=font2
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=3).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=4).value="=sum(D"+str(12+randuri)+":D"+str(11+len(uniclistglcredit)+randuri)+")"
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=4).font=font2
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=4).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=5).value="=sum(E"+str(12+randuri)+":E"+str(11+len(uniclistglcredit)+randuri)+")"
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=5).font=font2
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=5).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=6).value="=sum(F"+str(12+randuri)+":F"+str(11+len(uniclistglcredit)+randuri)+")"
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=6).font=font2
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=6).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=7).value="=sum(G"+str(12+randuri)+":G"+str(11+len(uniclistglcredit)+randuri)+")"
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=7).font=font2
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=7).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=8).value="=sum(H"+str(12+randuri)+":H"+str(11+len(uniclistglcredit)+randuri)+")"
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=8).font=font2
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=8).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=9).value="=sum(I"+str(12+randuri)+":I"+str(11+len(uniclistglcredit)+randuri)+")"
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=9).font=font2
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=9).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=10).value="=sum(J"+str(12+randuri)+":J"+str(11+len(uniclistglcredit)+randuri)+")"
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=10).font=font2
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=10).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=11).value="=sum(K"+str(12+randuri)+":K"+str(11+len(uniclistglcredit)+randuri)+")"
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=11).font=font2
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=11).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=12).value="=sum(L"+str(12+randuri)+":L"+str(11+len(uniclistglcredit)+randuri)+")"
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=12).font=font2
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=12).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=13).value="=sum(M"+str(12+randuri)+":M"+str(11+len(uniclistglcredit)+randuri)+")"
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=13).font=font2
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=13).number_format='#,##0_);(#,##0)'

					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=14).value="=sum(N"+str(12+randuri)+":N"+str(11+len(uniclistglcredit)+randuri)+")"
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=14).font=font2
					ws3.cell(row=12+len(uniclistglcredit)+randuri,column=14).number_format='#,##0_);(#,##0)'
					

					ws4=excel.create_sheet("Sheet")
					ws4.cell(row=1,column=1).value="Month"
					ws4.cell(row=2,column=1).value="January"
					ws4.cell(row=3,column=1).value="February"
					ws4.cell(row=4,column=1).value="March"
					ws4.cell(row=5,column=1).value="April"
					ws4.cell(row=6,column=1).value="May"
					ws4.cell(row=7,column=1).value="June"
					ws4.cell(row=8,column=1).value="July"
					ws4.cell(row=9,column=1).value="August"
					ws4.cell(row=10,column=1).value="September"
					ws4.cell(row=11,column=1).value="October"
					ws4.cell(row=12,column=1).value="November"
					ws4.cell(row=13,column=1).value="December"
					ws4.cell(row=1,column=2).value="Monthly Trend"
					ws4.cell(row=2,column=2).value="=Overview!B"+str(12+len(uniclistglcredit)+randuri)
					ws4.cell(row=3,column=2).value="=Overview!C"+str(12+len(uniclistglcredit)+randuri)
					ws4.cell(row=4,column=2).value="=Overview!D"+str(12+len(uniclistglcredit)+randuri)
					ws4.cell(row=5,column=2).value="=Overview!E"+str(12+len(uniclistglcredit)+randuri)
					ws4.cell(row=6,column=2).value="=Overview!F"+str(12+len(uniclistglcredit)+randuri)
					ws4.cell(row=7,column=2).value="=Overview!G"+str(12+len(uniclistglcredit)+randuri)
					ws4.cell(row=8,column=2).value="=Overview!H"+str(12+len(uniclistglcredit)+randuri)
					ws4.cell(row=9,column=2).value="=Overview!I"+str(12+len(uniclistglcredit)+randuri)
					ws4.cell(row=10,column=2).value="=Overview!J"+str(12+len(uniclistglcredit)+randuri)
					ws4.cell(row=11,column=2).value="=Overview!K"+str(12+len(uniclistglcredit)+randuri)
					ws4.cell(row=12,column=2).value="=Overview!L"+str(12+len(uniclistglcredit)+randuri)
					ws4.cell(row=13,column=2).value="=Overview!M"+str(12+len(uniclistglcredit)+randuri)




					for ii in range(1,15):
						ws3.cell(row=11+len(uniclistglcredit)+randuri,column=ii).border=doubleborder
						
					ws3.column_dimensions['A'].width=15
					ws3.column_dimensions['B'].width=15
					ws3.column_dimensions['C'].width=15
					ws3.column_dimensions['D'].width=15
					ws3.column_dimensions['E'].width=15
					ws3.column_dimensions['F'].width=15
					ws3.column_dimensions['G'].width=15
					ws3.column_dimensions['H'].width=15
					ws3.column_dimensions['I'].width=15
					ws3.column_dimensions['J'].width=15
					ws3.column_dimensions['K'].width=15
					ws3.column_dimensions['L'].width=15
					ws3.column_dimensions['M'].width=15



					values = Reference(ws4,
							min_col=2,  # I
							max_col=2,  # T
							min_row=1,
							max_row=13)
					labels=Reference(ws4,
							min_col=1,  # I
							max_col=1,  # T
							min_row=2,
							max_row=13)

					chart = LineChart()
					chart.add_data(values, titles_from_data=True)
					chart.set_categories(labels)

					chart.title = "Monthly Amount for account " + str(listaunica[i])
					chart.x_axis.title = ""
					chart.y_axis.title = "" 
					ws3.add_chart(chart, "A"+str(14+len(uniclistglcredit)+randuri))
					ws4.sheet_state='hidden'
					myorder=[2,0,1,3]
					folderpath="home/auditappnexia/output/je"
					excel._sheets =[excel._sheets[i] for i in myorder]
					excel.save(folderpath+"\\"+str(namec)+"\\"+str(listaunica[i])+"db.xlsx")
		make_archive("home/auditappnexia/output/je/"+str(namec),"home/auditappnexia/output/je/"+str(namec)+"/JE "+namec+".zip")                
				# shutil.make_archive(name, format, archive_from, archive_to)
		file_pathFS = os.path.join(folderpath, namec)    
		return send_from_directory(file_pathFS,"JE "+ namec+".zip",as_attachment=True)    
		
			
	return render_template('JE.html')
@app.route('/Payroll/Instructions', methods=['GET'])
def downloadPayroll():
		filepath = "/home/auditappnexia/output/payroll"
 
		return send_from_directory(filepath,"Instructions - Payroll.docx", as_attachment=True)


@app.route('/Payroll/GTaNuOQQZKTcQDsHcaAcqBLKmv3')
def payroll():
	return render_template('Payroll.html')

@app.route('/Payroll/GTaNuOQQZKTcQDsHcaAcqBLKmv3', methods=['POST', 'GET'])
def payroll_process():
	clientname1 = request.form['client']
	periodEnd1= datetime.datetime.strptime(
					 request.form['yearEnd'],
					 '%Y-%m-%d')
	preparedBy1=request.form['preparedBy']
	datePrepared1= datetime.datetime.strptime(
					 request.form['preparedDate'],
					 '%Y-%m-%d')
	refference1 = request.form['reff']
	currYear1 = request.form['CY']
	prevYear1 = request.form['PY']

	if request.method == 'POST':
		# file_Details = request.file('monthlyTB')
		openTB = request.files["TB"]
		openPBC = request.files['PBC']
		openMPL = request.files['MPL']
		D112XML = request.files['D112XML']

		# for i in file_Details:
		#     i.save(secure_filename(i.filename))

		# butoane

		# B1 = Button(app, text='About app', fg='white', height=5, width=20, bg="#8A2BE2", command=clickabout)
		# B1.grid(row=1, column=1, pady=2)
		#
		# B2 = Button(app, text='Import TB', fg='white', height=5, width=20, bg="#8A2BE2", command=openTB)
		# B2.grid(row=2, column=1, pady=2)
		#
		# B3 = Button(app, text='Import D112', fg='white', height=5, width=20, bg="#8A2BE2", command=openXML)
		# B3.grid(row=3, column=1, pady=2)
		#
		# B4 = Button(app, text='Import PBC', fg='white', height=5, width=20, bg="#8A2BE2", command=openPBC)
		# B4.grid(row=4, column=1, pady=2)
		#
		# B5 = Button(app, text='Import Monthly P&L', fg='white', height=5, width=20, bg="#8A2BE2", command=openMonPL)
		# B5.grid(row=5, column=1, pady=2)
		#
		# B6 = Button(app, text='Run and Save', fg='white', height=5, width=20, bg="#228B22", command=runAndSave)
		# B6.grid(row=6, column=1, pady=2)
		#
		# clname = Label(app, text="Enter the client name").grid(row=1, column=2, pady=2, sticky=W)
		# cl_name = Entry(app, textvariable=clientname, width=20).grid(row=1, column=3, pady=2, sticky=W)
		#
		# yrEnd = Label(app, text="Year End (mm/dd/yyyy)").grid(row=2, column=2, pady=2, sticky=W)
		# yr_End = Entry(app, textvariable=periodEnd, width=20).grid(row=2, column=3, pady=2, sticky=W)
		#
		# currYr = Label(app, text="Current Year (YYYY)").grid(row=3, column=2, pady=2, sticky=W)
		# curr_Yr = Entry(app, textvariable=currYear, width=20).grid(row=3, column=3, pady=2, sticky=W)
		#
		# prevYr = Label(app, text="Previous Year (YYYY)").grid(row=4, column=2, pady=2, sticky=W)
		# prev_Yr = Entry(app, textvariable=prevYear, width=20).grid(row=4, column=3, pady=2, sticky=W)
		#
		# prepBy = Label(app, text="Preparer name").grid(row=5, column=2, pady=2, sticky=W)
		# prep_By = Entry(app, textvariable=preparedBy, width=20).grid(row=5, column=3, pady=2, sticky=W)
		#
		# datePrep = Label(app, text="Prepared date").grid(row=6, column=2, pady=2, sticky=W)
		# date_Prep = Entry(app, textvariable=datePrepared, width=20).grid(row=6, column=3, pady=2, sticky=W)

		# reff = Label(app, text="Refference").grid(row=6, column=2, pady=2, sticky=W)
		# reff_1 = Entry(app, textvariable=refference, width=20).grid(row=6, column=3, pady=2, sticky=W)

		# fonts and colors
		ft1 = Font(name='Arial', size=10, bold=True)
		f_testname = Font(name='Arial', size=15, color='614C77', bold=True)
		f_info = Font(name='Arial', size=10, color='614C77', bold=True)
		cap_tabel = Font(name='Arial', size=10, color="FFFFFF", bold=True)
		cap_tabel_color_PBC = PatternFill(start_color='808080', end_color='808080', fill_type='solid')  # grey
		cap_tabel_color_GT = PatternFill(start_color='00AEAC', end_color='00AEAC', fill_type='solid')  # indigo #B1A0C7
		fprocentaj = Font(name='Arial', size=10, color="FF0000", bold=True)
		font_worksheet = Font(name='Arial', size=10)
		check_font = Font(name='Arial', size=10, color="FF0000", bold=True)
		check_font_1 = Font(name='Arial', size=10, color="FF0000", bold=False)
		cap_tabel_color_GT_movdeschis = PatternFill(start_color='00AEAC', end_color='00AEAC', fill_type='solid')
		cap_tabel_color_GT_movinchis = PatternFill(start_color='00AEAC', end_color='00AEAC', fill_type='solid')
		TB_font = Font(name='Arial', size=10, color="0070C0", bold=True)

		thin = Side(border_style='thin', color='000000')
		border = Border(left=thin, right=thin, top=thin, bottom=thin)

		thin = Side(border_style='thin', color='000000')
		border_left = Border(left=thin, right=None, top=thin, bottom=thin)

		thin = Side(border_style='thin', color='000000')
		border_right = Border(left=None, right=thin, top=thin, bottom=thin)

		thin = Side(border_style='thin', color='000000')
		border_centered = Border(left=None, right=None, top=thin, bottom=thin)

		thin = Side(border_style='thin', color='000000')
		border_upperleft = Border(left=thin, top=thin)

		thin = Side(border_style='thin', color='000000')
		border_lowerleft = Border(left=thin, right=None, top=None, bottom=thin)

		thin = Side(border_style='thin', color='000000')
		border_upperright = Border(right=thin, top=thin)

		thin = Side(border_style='thin', color='000000')
		border_lowerright = Border(right=thin, bottom=thin)

		thin = Side(border_style='thin', color='000000')
		border_left1 = Border(left=thin)

		thin = Side(border_style='thin', color='000000')
		border_right1 = Border(right=thin)

		thin = Side(border_style='thin', color='000000')
		border_top = Border(top=thin)

		thin = Side(border_style='thin', color='000000')
		border_bottom = Border(bottom=thin)

		# im = tk.PhotoImage(data='''iVBORw0KGgoAAAANSUhEUgAAAS0AAACnCAMAAABzYfrWAAAAzFBMVEX///8AAAAjICFSKYSVg7abhLiLc62FbKseGxyTfLSCaKePe7IgHR50WJ6cm5t8YqVrSpaSkZFkQ5MQCAs5NjdGQ0QXExTAv78IAAD4+PhNS0u6uboMAAVWMInMzMxvUptcOY3l5eWWlZYoJSZpZ2eGhYV0c3TY19hnRpRNIYGqqanw8PBiPZBaNIvj4+NAAHrt6vI0MTJcWlvUzOCkk7/d1+bl4exLHIB0cnJFEX2lpKSysbFYVlZjYWK0p8nAttGhj726rs6vn8bKwddFjzy8AAAK/klEQVR4nO2bCXuiOheACSCLoqhUBGvdlyra6V6r3fv//9N3sgARtO2d6b32e+a8zzNTt5DwkpycRFQUBEEQBEEQBEEQBEEQBEEQBEEQBEEQBEEQBEEQBEEQBEEQBEEQBEEQBEEQBEEQBEEQ5P+C8Pbq+e365eXt9f3m/NCN+dlE79eGaRhGiQJ/9aebQzfpx3L7ZprUlC1h6M/Yw3Zwo7tAudztuoYuUbJfw0O37acRvrnFYpHaKnerp6duoVDQCwLbfj90834WNxCuTLPI+1a1errouQUJ+xq7V8orDe2yrXpvubAlXbp+e+g2/hjejFJqi+o6rS96S624Fb6uDt3KHwLIytiqUlsnM1fWZaAuyhOkDMJWkUd5Zqt3ot25upxMYO6lKM8lfZ8tbWZKtkoGZl63JT1jq5za0rSShKEfurEHh853YKuU61tLZuuUWmJhjfp8PXRrD8yrvc9W74TqmhUNieJHecTZcWU8rs2n35SadZo5Osr0GP47GOd2pm+5yZzIbWmL1JVpmnvHYjS4IMQDAuKPvuWEHkiOB2VMSOM7Dv57vBW2bRVFCpHamtEkP8HdMy9WAs9RLd/zfEt1fNI6+/OmXbZ9hqOqKn/UvlRqntX680P/JueGZMswy6eQk2rLJTW1FLa0riTLLL7sOkx4SRzVIw/jWm3SasNjn2z+uG211YRx4ahD/mg1OKytJz2xVXJ7sxnTo93dv1y9F++ErR7tcgnursjV8lQrGET8STQftlWVHP9ei8JomnmlYVlHyZPvshX2+/+8kMFtgS5zKVRp2i+eV11V7/jz2BNnx7S4ClT/Qq59TLyH3zuLxyG5zLz0L9gKWyoZ/ONSN7EtvZq4mt0/x2+//GIvlF2ZUu4oTaI6arT90lHuU1/jyPJHmZf+DVuW5VX+calXW9jqpbI0aag9/WK2tjByQ/HIUsl3Tev/kS3V+Q1btrB1oiWyTrZWN1TXti3XyO4MdojqP/5J22V+sK3QZLYKy0SWdp/pOS/32qzLPXUZrv2WOcrK39u1QkBRpvN5kz2NOvNBZd6MB20U0UdhZz6f8/JhFB05/iP8kRPcXbamm3l6HHasJhx6k+QtScWdHdVEUR9s1dJqouZ6UDlOC/M3snXcclt17STmPrel3J1p1W5K1SxlE1THcazdspRj4o2UCQlIQNu0gr80ew3G/N0jQqZKzSdBEJAhPZE5gaRNtTyPzD+0tRkSKESCpHucjQhpQzHS4JdFqZD2WHmEiiGVbWSrWbFqIDfkF3n6mCtMKsrmIlOHcmVQWy6VtVzSf5qRO+HbXzNmSXBaso3tlU0fBmJ2GkttWRc1mAMsFT7ne5CGAY5KRkKDt34kNO0ERXBGYMuCTFS1rA9tPQwIT1odsuYvzonv0EP7cBh+JSpt/3HVhjwZpuaW1Z4/yNVMqCxolMVsraXCNV448EeVbB2K8g62dBuyUUhHGXc7tmSu76oS3YJtbE9/ELa82l5b0GCPDFVPUS5gwI6n/f7xRTwpNCzI/iGlHY9gAUCNb9QjOI3h0YUq57YZW9ArSGM8nqg+nAp7bQO1kMlxpzmhV4K1peLRhUUwpMO2FVfzCJZohB0PIeF1oBqHXSFIFSfNzvGKFh5w1TBQWR1DEEzizvFMbVW5KFjp9Jb57ADSfQ16lKC+KBZsc1spGGmvd5QT74GiKAwjeOiLSxfBtZwIW47KctH+EFrKLsLnUR6OyFyGDyJeRrTz8pS2E8BxzrgtOHVWMbXlOOz9M6jPY0EsifJnIGvII1aTNvaM21JF76Z1iPGpvBq2bSwXnHq9vkwyLRmXvsVZLAq6ndkTPA5UaZkTDiqCQchs+SJIrVerkbhKE996ELbipszbKpl+0Vbck6fQqSvscOksA93MX3FbyaTXShOctagmtUWnqHjxAN2MXUaw5YlWn5H0OGCr0K2nLF52yAqFTdb5XP0TWxHbh6AQbovk19cVz2oIDfF+Apw5P6Mv2CJxJBAhYCuBoZ00ZLaSEdRKq+mIahJbYVuV6lP5fAW2SLwyCdIwAyPRrp+mQen0fsfW1POsl0LXlJm4tT0SI761IsY7tZU5JMzsqS0+IhU2VXzVljNMzdAzoaLTNemY95WK9LFWvprEFtXXTArTXnpGbTlqKjCxdWWUTDmEL+53bF8tlr2lQDOYre3zp1F+nMo4Ztt2Iz+xlYrcrFpHF0NAjW0lLYmCL9tKOoqwRQdx2n/nATt9uCBJmVZaTc7WOki7ER+otHA7LXyUFr419HJVyqYW2q/c9tX1XZKMaV22X5ExCi2wckvocd5WJYBJmY1R689sJbk8tzWQxiYL1e15zla8gs7ZGkNWkxamzd3stXVulqTEs1tewioxM3Cu7qkmvudc57s715mzGTqOlx3BeVtj4vhtmMaBB+s7bdW+z1bzQ1uKAbakNSA4uStvtfTmLl1BLnSd2cpOnHRaye795WxBdLFaol1p3PoOW2tPHokbHsW+aquypXpDPhqJypuxtWCm68X74pasmSyLLcFz38FC4LKyGzQ5W3ARkwAx+FZbza1ATcfl9Ou2jom8xq15IsrvtnVlCk+cOltXl5MM4fk+ldUtCFul3LwJeVOQyeZzti79dJ659L/TViSypPjDTlv5uq3+1hQFhR1lv62wuLXRdyp2uPjS+tz4lbhaFm1b2MpuQbDOpW6t7ECIlbEFSbEfmyHfGuVZkhuPpk6cYH7RFhXkS4WZun22lOuyK+25l3lXOrkvX4Xnb3ciZs1my7JNb69ktko7vvQZU10TKVrCmtCHpaFkS0q4Qdx+W6AmO8F+ZqtJIwHv8BFNTvtfsOU4Yo1BVxsNXrivxoX32bopp67MYlHYOlnOetqMiYJnVZfdtiv6ViEvC1yALo+MNlNgMx4SGJpsnz61RXejrU0URs1G8JGtB0v1jsNIXgB8ZksZQULuVKBqCIgq3zP4vG85aodV8wiFh2soXKOF2QjZa0sxivL3Xz1hq7dYnNJEjEY00zQlW3vuqVzTLSNf5PGQUZDkysUZRMtTHUICeL9R+yBu1QLopoTIswYMlo9tKZeEVkmgDZb4cuITWxA6oTUBoUvKh6SwT/jo3G/ryk1UgRN3Jmz16okt6HTcFh+Je76274894vkW4HtEHYvpbwOtEB8IR4R+m0rIKuwQwpYlF4TEEbYPH2SnEV3QjaW2/H00fCyZIOh31cmZBEn5tQqV+3RDT4z2gaiC2SYkPuGzuJppm1ZD2OywtkTh1jRfeJi2kaLLtzkYJzttmYkte/8db2Gz9thqNFqjSicxGkEPTz4wXU9G4zl0/lC8Cn/ipIK+xEuFm8nl41jO36byQfrSE6m80qG3YKQ7zVsf63Ti19NqovUKqhEb3B8VluoAzl3ZFutc9Hvqen2nrWwe/9fxbPJ7jhh6XfStxU5bJbzd7Tq5m822S7Zkq0pXRbKtXdnD30aob/1SZaadLJc7bZV2bq3+bYS2+GUPv3WZ6kpGYmyL3h74dOiG/gwifetOb02jtuJJMbaFsmLCl1Iiq6DbixPJlsttmTgMU57kn6kU3GVv25aJPy3Y4qYg+7LLcgpRdF+iz4/wVxE+2za/l4v/rKfYZZMi3cnBjrWD89dSHL7oDFmwzaJbdF0dXe0mvLq2IaLb7D5U9stq/RV/afcB4c3724vOVkHXT1e40vkK7GYxBEEQBEEQBEEQBEEQBEEQBEEQBEEQBEEQBEEQBEEQBEEQBEEQBEEQBEEQBEEQBEEQ5D/kf5n9VgB9zbfsAAAAAElFTkSuQmCC''')
		# tk.Label(app, image=im).grid(row = 3, column = 4,columnspan = 2, rowspan = 2, padx = 2, pady = 2)
		#
		# app.mainloop()

		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "Summary"

		Sheet1 = wb['Summary']
		# Sheet1 = wb.create_sheet("Summary")

		ws.sheet_view.showGridLines = False
		Sheet1.cell(row=1, column=1).value = "Client Name:"
		Sheet1.cell(row=2, column=1).value = "Period ended:"
		Sheet1.cell(row=3, column=1).value = "Audit area:"
		Sheet1.cell(row=3, column=2).value = "Payroll"
		Sheet1.cell(row=5, column=1).value = "Payroll and related taxes"

		Sheet1.cell(row=1, column=13).value = "Prepared by:"
		Sheet1.cell(row=2, column=13).value = "Date:"
		Sheet1.cell(row=3, column=13).value = "Reference"
		Sheet1.cell(row=3, column=14).value = refference1
		Sheet1.cell(row=4, column=13).value = "Date:"
		Sheet1.cell(row=5, column=13).value = "Reviewed by:"

		Sheet1.cell(row=1, column=2).value = clientname1
		Sheet1.cell(row=2, column=2).value = periodEnd1
		Sheet1.cell(row=2, column=2).number_format='mm/dd/yyyy'
		Sheet1.cell(row=1, column=14).value = preparedBy1
		Sheet1.cell(row=2, column=14).value = datePrepared1
		Sheet1.cell(row=2, column=14).number_format = 'mm/dd/yyyy'

		Sheet1.cell(row=7, column=1).value = "Source"
		Sheet1.cell(row=8, column=1).value ="Trial Balance @ 31.12." + currYear1 +";"
		Sheet1.cell(row=9, column=1).value ="Trial Balance @ 31.12."+prevYear1+";"
		Sheet1.cell(row=10, column=1).value ="D112 Statement @ December, " + currYear1 +";"
		Sheet1.cell(row=11, column=1).value ="Payroll sheets @ FY "+currYear1+";"
		Sheet1.cell(row=12, column=1).value ="Untaken holiday database;"
		Sheet1.cell(row=13, column=1).value ="Other info provided by the client; "

		Sheet1.cell(row=15, column=1).value = "Purpose"
		Sheet1.cell(row=17, column=1).value ="Perform analytical procedures on an account basis for payroll and related taxes (expenses and liabilities)."

		Sheet1.cell(row=20, column=1).value = "Relevant assertion"
		Sheet1.cell(row=22, column=1).value = "Completeness & Accuracy;"

		Sheet1.cell(row=25, column=1).value = "Risk"
		Sheet1.cell(row=27, column=1).value ="Payroll and related taxes understated or not recorded in correct period (e.g., accruals, prepaids, etc.)."

		Sheet1.cell(row=30, column=1).value = "Work done (description of procedures performed)"
		Sheet1.cell(row=32, column=1).value = "V10.1 Payroll Analysis"
		Sheet1.cell(row=33, column=1).value ="We made variations for this account by comparing current year with the prior year;"
		Sheet1.cell(row=34, column=1).value ="We obtained the monthly number of the employees for "+prevYear1 +" and "+currYear1 +"; "
		Sheet1.cell(row=35, column=1).value ="We divided the value of payroll costs to the no. of employees to obtain the monthly averge salary for "+prevYear1 + " and " +currYear1+";"

		Sheet1.cell(row=37, column=1).value = "V10.2 Constribution Reasonableness"
		Sheet1.cell(row=38, column=1).value ="We divided the contribution costs paid to the payroll costs in order to see if these are in accordance with the law percent;"
		Sheet1.cell(row=39, column=1).value ="We checked and analyzed the percentage obtained with the law percent;"
		Sheet1.cell(row=40, column=1).value ="The Labor Insurement Contribution (CAM) of 2.25% is due by employers and  the basis of calculation is the gross salary fund;"

		Sheet1.cell(row=42, column=1).value = "V10.3 D11 Reconciliation"
		Sheet1.cell(row=43, column=1).value ="We received the monthly trial balances from client;"
		Sheet1.cell(row=44, column=1).value ="We extracted from trial balances the social contributions;"
		Sheet1.cell(row=45, column=1).value ="We compared the amounts from trial balance with the amounts from D112 ""Social contributions statement"" for 31.12."+currYear1+";"
		Sheet1.cell(row=46, column=1).value ="We reconcilied the amount from D112 ""Social contribution statement"" for 31.12."+currYear1+" with the payment and the payment centralizer;"

		Sheet1.cell(row=48, column=1).value = "V10.4 Payroll Sheets Reconciliation"
		Sheet1.cell(row=49, column=1).value ="We obtained the montly payroll expenses for current year;"
		Sheet1.cell(row=50, column=1).value ="We checked if the monthly payroll expense recorded by the Company is equal to the salary from the monthly payroll sheet;"

		Sheet1.cell(row=52, column=1).value = "V10.5 Meal Tickets Reasonableness"
		Sheet1.cell(row=53, column=1).value ="We obtained the number of monthy meal tickets provided to Company employees;"
		Sheet1.cell(row=54, column=1).value ="According to Romanian Legislation the Companies can spend between 9.57 and 20 RON/ ticket;"
		Sheet1.cell(row=55, column=1).value ="The value of the tickets is established, internally at RON XX/ticket;"
		Sheet1.cell(row=56, column=1).value ="We have recomputed the monthly M.T expense based on the info received;"

		Sheet1.cell(row=58, column=1).value = "V10.6 Untaken Holiday "
		Sheet1.cell(row=59, column=1).value ="We obtained the detail of untaken holidays from the client;"
		Sheet1.cell(row=60, column=1).value ="We recomputed the provision amount, based on the information received;"
		Sheet1.cell(row=61, column=1).value ="We reconciled the recalculated amounts with the ones in the Trial Balance. "

		Sheet1.cell(row=63, column=1).value = "Findings & Conclusion"
		Sheet1.cell(row=65, column=1).value ="Based on our tests, no material differences were identified."
		Sheet1.cell(row=66, column=1).value ="We obtained sufficient and appropriate audit evidence that the completeness and accuracy of the ""Payroll and related taxes"" records are not significantly misstated."

		Sheet1.cell(row=69, column=1).value = "Materiality"
		Sheet1.cell(row=70, column=1).value = "Per. Materiality"
		Sheet1.cell(row=71, column=1).value = "Trivial Amount"

		#FORMAT + culori

		Sheet1.cell(row=1, column=1).font = f_info
		Sheet1.cell(row=2, column=1).font = f_info
		Sheet1.cell(row=3, column=1).font = f_info
		Sheet1.cell(row=3, column=2).font = ft1
		Sheet1.cell(row=5, column=1).font = f_testname

		Sheet1.cell(row=1, column=13).font = ft1
		Sheet1.cell(row=2, column=13).font = ft1
		Sheet1.cell(row=3, column=13).font = ft1
		Sheet1.cell(row=4, column=13).font = ft1
		Sheet1.cell(row=5, column=13).font = ft1

		Sheet1.cell(row=1, column=2).font = font_worksheet
		Sheet1.cell(row=2, column=2).font = font_worksheet

		Sheet1.cell(row=1, column=14).font = font_worksheet
		Sheet1.cell(row=2, column=14).font = font_worksheet

		#primul tabel
		Sheet1.cell(row=7, column=1).font = ft1

		for row in Sheet1['A8:N13']:
			for cell in row:
				cell.font = font_worksheet

		# al doilea tabel
		Sheet1.cell(row=15, column=1).font = ft1

		for row in Sheet1['A16:N18']:
			for cell in row:
				cell.font = font_worksheet

		# al treilea tabel
		Sheet1.cell(row=20, column=1).font = ft1

		for row in Sheet1['A21:N23']:
			for cell in row:
				cell.font = font_worksheet

		# al patrulea tabel
		Sheet1.cell(row=25, column=1).font = ft1

		for row in Sheet1['A26:N28']:
			for cell in row:
				cell.font = font_worksheet

		# al cincelea tabel
		Sheet1.cell(row=30, column=1).font = ft1

		for row in Sheet1['A31:N61']:
			for cell in row:
				cell.font = font_worksheet
		Sheet1.cell(row=32, column=1).font = ft1
		Sheet1.cell(row=37, column=1).font = ft1
		Sheet1.cell(row=42, column=1).font = ft1
		Sheet1.cell(row=48, column=1).font = ft1
		Sheet1.cell(row=52, column=1).font = ft1
		Sheet1.cell(row=58, column=1).font = ft1

		# al saselea tabel
		Sheet1.cell(row=63, column=1).font = ft1

		for row in Sheet1['A64:N67']:
			for cell in row:
				cell.font = font_worksheet

		#al saptelea tabel
		for row in Sheet1['A69:A71']:
			for cell in row:
				cell.font = ft1
		for row in Sheet1['B69:B71']:
			for cell in row:
				cell.font = font_worksheet


		#COLORS
		for row in Sheet1['A7:N7']:
			for cell in row:
				cell.fill = cap_tabel_color_GT_movinchis

		for row in Sheet1['A8:N13']:
			for cell in row:
				cell.fill = cap_tabel_color_GT_movdeschis

		#al doilea tabel
		for row in Sheet1['A15:N15']:
			for cell in row:
				cell.fill = cap_tabel_color_GT_movinchis

		for row in Sheet1['A16:N18']:
			for cell in row:
				cell.fill = cap_tabel_color_GT_movdeschis

		# al treilea tabel
		for row in Sheet1['A20:N20']:
			for cell in row:
				cell.fill = cap_tabel_color_GT_movinchis

		for row in Sheet1['A21:N23']:
			for cell in row:
				cell.fill = cap_tabel_color_GT_movdeschis

		# al patrulea tabel
		for row in Sheet1['A25:N25']:
			for cell in row:
				cell.fill = cap_tabel_color_GT_movinchis

		for row in Sheet1['A26:N28']:
			for cell in row:
				cell.fill = cap_tabel_color_GT_movdeschis

		# al cincelea tabel
		for row in Sheet1['A30:N30']:
			for cell in row:
				cell.fill = cap_tabel_color_GT_movinchis

		for row in Sheet1['A31:N61']:
			for cell in row:
				cell.fill = cap_tabel_color_GT_movdeschis

		# al saselea tabel
		for row in Sheet1['A63:N63']:
			for cell in row:
				cell.fill = cap_tabel_color_GT_movinchis

		for row in Sheet1['A64:N67']:
			for cell in row:
				cell.fill = cap_tabel_color_GT_movdeschis

		# al saptelea tabel
		for row in Sheet1['A69:A71']:
			for cell in row:
				cell.fill = cap_tabel_color_GT_movinchis
		for row in Sheet1['B69:B71']:
			for cell in row:
				cell.fill = cap_tabel_color_GT_movdeschis

		#borders
		#primul tabel
		Sheet1['A7'].border = border_left
		for row in Sheet1['B7:M7']:
			for cell in row:
				cell.border = border_centered
		Sheet1["N7"].border = border_right

		for row in Sheet1['A8:A12']:
			for cell in row:
				cell.border = border_left1

		Sheet1['A13'].border = border_lowerleft

		for row in Sheet1['B13:M13']:
			for cell in row:
				cell.border = border_bottom

		Sheet1['N13'].border = border_lowerright

		for row in Sheet1['N8:N12']:
			for cell in row:
				cell.border = border_right1

		#al doilea tabel
		Sheet1['A15'].border = border_left
		for row in Sheet1['B15:M15']:
			for cell in row:
				cell.border = border_centered
		Sheet1["N15"].border = border_right

		for row in Sheet1['A16:A17']:
			for cell in row:
				cell.border = border_left1

		Sheet1['A18'].border = border_lowerleft

		for row in Sheet1['B18:M18']:
			for cell in row:
				cell.border = border_bottom

		Sheet1['N18'].border = border_lowerright

		for row in Sheet1['N16:N17']:
			for cell in row:
				cell.border = border_right1

			# al treilea tabel
		Sheet1['A20'].border = border_left
		for row in Sheet1['B20:M20']:
			for cell in row:
				cell.border = border_centered
		Sheet1["N20"].border = border_right

		for row in Sheet1['A21:A22']:
			for cell in row:
				cell.border = border_left1

		Sheet1['A23'].border = border_lowerleft

		for row in Sheet1['B23:M23']:
			for cell in row:
				cell.border = border_bottom

		Sheet1['N23'].border = border_lowerright

		for row in Sheet1['N21:N22']:
			for cell in row:
				cell.border = border_right1

		# al patrulea tabel
		Sheet1['A25'].border = border_left
		for row in Sheet1['B25:M25']:
			for cell in row:
				cell.border = border_centered
		Sheet1["N25"].border = border_right

		for row in Sheet1['A26:A27']:
			for cell in row:
				cell.border = border_left1

		Sheet1['A28'].border = border_lowerleft

		for row in Sheet1['B28:M28']:
			for cell in row:
				cell.border = border_bottom

		Sheet1['N28'].border = border_lowerright

		for row in Sheet1['N26:N27']:
			for cell in row:
				cell.border = border_right1

			# al cincelea tabel
		Sheet1['A30'].border = border_left
		for row in Sheet1['B30:M30']:
			for cell in row:
				cell.border = border_centered
		Sheet1["N30"].border = border_right

		for row in Sheet1['A31:A60']:
			for cell in row:
				cell.border = border_left1

		Sheet1['A28'].border = border_lowerleft

		for row in Sheet1['B61:M61']:
			for cell in row:
				cell.border = border_bottom

		Sheet1['A61'].border = border_lowerleft
		Sheet1['N61'].border = border_lowerright

		for row in Sheet1['N31:N60']:
			for cell in row:
				cell.border = border_right1

		# al saselea tabel
		Sheet1['A63'].border = border_left
		for row in Sheet1['B63:M63']:
			for cell in row:
				cell.border = border_centered
		Sheet1["N63"].border = border_right

		for row in Sheet1['A64:A66']:
			for cell in row:
				cell.border = border_left1

		Sheet1['A67'].border = border_lowerleft

		for row in Sheet1['B67:M67']:
			for cell in row:
				cell.border = border_bottom

		Sheet1['N67'].border = border_lowerright

		for row in Sheet1['N64:N66']:
			for cell in row:
				cell.border = border_right1

		# al saptelea
		for row in Sheet1['A69:B71']:
			for cell in row:
				cell.border = border

		Sheet1.column_dimensions['A'].width = 16
		Sheet1.column_dimensions['B'].width = 10

		Sheet2 = wb.create_sheet("Lead")

		Sheet2.sheet_view.showGridLines = False
		Sheet2.cell(row=1, column=1).value = "Client Name:"
		Sheet2.cell(row=2, column=1).value = "Period ended:"
		Sheet2.cell(row=3, column=1).value = "Audit area:"
		Sheet2.cell(row=3, column=2).value = "Payroll"
		Sheet2.cell(row=5, column=1).value = "Accounts Reconciliation & SA"

		try:
			Sheet2.cell(row=1, column=2).value = clientname1
			Sheet2.cell(row=2, column=2).value = periodEnd1
		except:
			None

		# FORMATARE
		Sheet2.cell(row=1, column=1).font = f_info
		Sheet2.cell(row=2, column=1).font = f_info
		Sheet2.cell(row=3, column=1).font = f_info
		Sheet2.cell(row=3, column=2).font = font_worksheet
		Sheet2.cell(row=5, column=1).font = f_testname

		Sheet2.cell(row=1, column=2).font = font_worksheet
		Sheet2.cell(row=2, column=2).font = font_worksheet

		Sheet3 = wb.create_sheet("V10.1 Payroll Analysis")

		Sheet3.sheet_view.showGridLines = False

		Sheet3.cell(row=1, column=1).value = "Client Name:"
		Sheet3.cell(row=2, column=1).value = "Period ended:"
		Sheet3.cell(row=3, column=1).value = "Audit area:"
		Sheet3.cell(row=3, column=2).value = "Payroll"
		Sheet3.cell(row=5, column=1).value = "Payroll Analysis"

		try:
			Sheet3.cell(row=1, column=2).value = clientname1
			Sheet3.cell(row=2, column=2).value = periodEnd1
		except:
			None
		#
		# Sheet3.cell(row=7, column=2).value = "Work done"
		# Sheet3.cell(row=12, column=2).value = "Conclusion"

		Sheet3.cell(row=22, column=2).value = "Description"
		Sheet3.cell(row=22, column=3).value = "CY"
		Sheet3.cell(row=22, column=4).value = "PY"
		Sheet3.cell(row=22, column=5).value = "Variation"
		Sheet3.cell(row=22, column=6).value = "Variation %"

		Sheet3.cell(row=23, column=2).value = "No of employee"
		Sheet3.cell(row=24, column=2).value = "Gross salary"
		Sheet3.cell(row=25, column=2).value = "Employees Advantages"
		Sheet3.cell(row=26, column=2).value = "Total Salary"
		Sheet3.cell(row=27, column=2).value = "Average salary"

		Sheet3.cell(row=22, column=8).value = "Average salary"
		Sheet3.cell(row=22, column=9).value = "January"
		Sheet3.cell(row=22, column=10).value = "February"
		Sheet3.cell(row=22, column=11).value = "March"
		Sheet3.cell(row=22, column=12).value = "April"
		Sheet3.cell(row=22, column=13).value = "May"
		Sheet3.cell(row=22, column=14).value = "June"
		Sheet3.cell(row=22, column=15).value = "July"
		Sheet3.cell(row=22, column=16).value = "August"
		Sheet3.cell(row=22, column=17).value = "September"
		Sheet3.cell(row=22, column=18).value = "Octomber"
		Sheet3.cell(row=22, column=19).value = "November"
		Sheet3.cell(row=22, column=20).value = "December"
		Sheet3.cell(row=22, column=21).value = "Average"
		Sheet3.cell(row=22, column=22).value = "Variation"
		Sheet3.cell(row=22, column=23).value = "Variation %"

		Sheet3.cell(row=23, column=8).value = "No of employees " + currYear1
		Sheet3.cell(row=24, column=8).value = "No of employees " + prevYear1
		Sheet3.cell(row=26, column=8).value = "Average salary of " + currYear1
		Sheet3.cell(row=27, column=8).value = "Average salary of " + prevYear1
		Sheet3.cell(row=28, column=8).value = "Variation"
		Sheet3.cell(row=29, column=8).value = "Variation %"

		Sheet3.cell(row=31, column=2).value = "As per Monthly PL"
		Sheet3.cell(row=32, column=2).value = "Period"
		Sheet3.cell(row=32, column=3).value = "Account"
		Sheet3.cell(row=32, column=4).value = "Descriere"
		Sheet3.cell(row=38, column=2).value = currYear1
		Sheet3.cell(row=33, column=2).value = currYear1
		Sheet3.cell(row=34, column=2).value = currYear1
		Sheet3.cell(row=35, column=2).value = prevYear1
		Sheet3.cell(row=36, column=2).value = currYear1
		Sheet3.cell(row=37, column=2).value = prevYear1
		Sheet3.cell(row=38, column=2).value = currYear1
		Sheet3.cell(row=39, column=2).value = prevYear1

		Sheet3.cell(row=33, column=3).value = "641"
		Sheet3.cell(row=34, column=3).value = "641"
		Sheet3.cell(row=35, column=3).value = "641"
		Sheet3.cell(row=36, column=3).value = "642"
		Sheet3.cell(row=37, column=3).value = "642"
		Sheet3.cell(row=38, column=3).value = "645"
		Sheet3.cell(row=39, column=3).value = "645"

		Sheet3.cell(row=33, column=4).value = "Cheltuieli cu salariile personalului"
		Sheet3.cell(row=34, column=4).value = "Bonus"
		Sheet3.cell(row=35, column=4).value = "Cheltuieli cu salariile personalului"
		Sheet3.cell(row=36, column=4).value = "Cheltuieli cu avantajele în natura acordate salariatilor"
		Sheet3.cell(row=37, column=4).value = "Cheltuieli cu avantajele în natura acordate salariatilor"
		Sheet3.cell(row=38, column=4).value = "Alte cheltuieli privind asigurarile si protectia sociala"
		Sheet3.cell(row=39, column=4).value = "Alte cheltuieli privind asigurarile si protectia sociala"

		Sheet3.cell(row=32, column=5).value = "January"
		Sheet3.cell(row=32, column=6).value = "February"
		Sheet3.cell(row=32, column=7).value = "March"
		Sheet3.cell(row=32, column=8).value = "April"
		Sheet3.cell(row=32, column=9).value = "May"
		Sheet3.cell(row=32, column=10).value = "June"
		Sheet3.cell(row=32, column=11).value = "July"
		Sheet3.cell(row=32, column=12).value = "August"
		Sheet3.cell(row=32, column=13).value = "September"
		Sheet3.cell(row=32, column=14).value = "Octomber"
		Sheet3.cell(row=32, column=15).value = "November"
		Sheet3.cell(row=32, column=16).value = "December"
		Sheet3.cell(row=32, column=17).value = "TOTAL"
		Sheet3.cell(row=32, column=18).value = "As per TB"
		Sheet3.cell(row=32, column=19).value = "Check"
		Sheet3.cell(row=32, column=20).value = "Variation"
		Sheet3.cell(row=32, column=21).value = "Variation %"

		# FORMULAS
		# numar salariati
		#CY
		Sheet3.cell(row=23, column=9).value = "=PBC!B4"
		Sheet3.cell(row=23, column=10).value = "=PBC!C4"
		Sheet3.cell(row=23, column=11).value = "=PBC!D4"
		Sheet3.cell(row=23, column=12).value = "=PBC!E4"
		Sheet3.cell(row=23, column=13).value = "=PBC!F4"
		Sheet3.cell(row=23, column=14).value = "=PBC!G4"
		Sheet3.cell(row=23, column=15).value = "=PBC!H4"
		Sheet3.cell(row=23, column=16).value = "=PBC!I4"
		Sheet3.cell(row=23, column=17).value = "=PBC!J4"
		Sheet3.cell(row=23, column=18).value = "=PBC!K4"
		Sheet3.cell(row=23, column=19).value = "=PBC!L4"
		Sheet3.cell(row=23, column=20).value = "=PBC!M4"

		Sheet3.cell(row=24, column=9).value = "=PBC!B3"
		Sheet3.cell(row=24, column=10).value = "=PBC!C3"
		Sheet3.cell(row=24, column=11).value = "=PBC!D3"
		Sheet3.cell(row=24, column=12).value = "=PBC!E3"
		Sheet3.cell(row=24, column=13).value = "=PBC!F3"
		Sheet3.cell(row=24, column=14).value = "=PBC!G3"
		Sheet3.cell(row=24, column=15).value = "=PBC!H3"
		Sheet3.cell(row=24, column=16).value = "=PBC!I3"
		Sheet3.cell(row=24, column=17).value = "=PBC!J3"
		Sheet3.cell(row=24, column=18).value = "=PBC!K3"
		Sheet3.cell(row=24, column=19).value = "=PBC!L3"
		Sheet3.cell(row=24, column=20).value = "=PBC!M3"

		Sheet3.cell(row=23, column=3).value = "=U23"
		Sheet3.cell(row=24, column=3).value = "=Q33"
		Sheet3.cell(row=25, column=3).value = "=Q36"
		Sheet3.cell(row=26, column=3).value = "=SUM(C24:C25)"
		Sheet3.cell(row=27, column=3).value = "=C26/C23/12"

		Sheet3.cell(row=23, column=4).value = "=U24"
		Sheet3.cell(row=24, column=4).value = "=Q35"
		Sheet3.cell(row=25, column=4).value = "=Q37"
		Sheet3.cell(row=26, column=4).value = "=SUM(D24:D25)"
		Sheet3.cell(row=27, column=4).value = "=D26/D23/12"

		Sheet3.cell(row=23, column=5).value = "=C23-D23"
		Sheet3.cell(row=24, column=5).value = "=C24-D24"
		Sheet3.cell(row=25, column=5).value = "=C25-D25"
		Sheet3.cell(row=26, column=5).value = "=C26-D26"
		Sheet3.cell(row=27, column=5).value = "=C27-D27"

		Sheet3.cell(row=23, column=6).value = "=E23/D23"
		Sheet3.cell(row=24, column=6).value = "=E24/D24"
		Sheet3.cell(row=25, column=6).value = "=E25/D25"
		Sheet3.cell(row=26, column=6).value = "=E26/D26"
		Sheet3.cell(row=27, column=6).value = "=E27/D27"

		Sheet3.cell(row=23, column=21).value = "=AVERAGE(I23:T23)"
		Sheet3.cell(row=23, column=22).value = "=U23-U24"
		Sheet3.cell(row=23, column=23).value = "=V23/U24"
		Sheet3.cell(row=24, column=21).value = "=AVERAGE(I24:T24)"

		Sheet3.cell(row=26, column=9).value = "=(E33-E34)/I23"
		Sheet3.cell(row=26, column=10).value = "=(F33-F34)/J23"
		Sheet3.cell(row=26, column=11).value = "=(G33-G34)/K23"
		Sheet3.cell(row=26, column=12).value = "==(H33-H34)/L23"
		Sheet3.cell(row=26, column=13).value = "=(I33-I34)/M23"
		Sheet3.cell(row=26, column=14).value = "=(J33-J34)/N23"
		Sheet3.cell(row=26, column=15).value = "=(K33-K34)/O23"
		Sheet3.cell(row=26, column=16).value = "=(L33-L34)/P23"
		Sheet3.cell(row=26, column=17).value = "=(M33-M34)/Q23"
		Sheet3.cell(row=26, column=18).value = "=(N33-N34)/R23"
		Sheet3.cell(row=26, column=19).value = "=(O33-O34)/S23"
		Sheet3.cell(row=26, column=20).value = "=(P33-P34)/T23"
		Sheet3.cell(row=26, column=21).value = "=AVERAGE(I26:T26)"

		Sheet3.cell(row=27, column=9).value = "=E35/I24"
		Sheet3.cell(row=27, column=10).value = "=F35/J24"
		Sheet3.cell(row=27, column=11).value = "=G35/K24"
		Sheet3.cell(row=27, column=12).value = "=H35/L24"
		Sheet3.cell(row=27, column=13).value = "=I35/M24"
		Sheet3.cell(row=27, column=14).value = "=J35/N24"
		Sheet3.cell(row=27, column=15).value = "=K35/O24"
		Sheet3.cell(row=27, column=16).value = "=L35/P24"
		Sheet3.cell(row=27, column=17).value = "=M35/Q24"
		Sheet3.cell(row=27, column=18).value = "=N35/R24"
		Sheet3.cell(row=27, column=19).value = "=O35/S24"
		Sheet3.cell(row=27, column=20).value = "=P35/T24"
		Sheet3.cell(row=27, column=21).value = "=AVERAGE(I27:T27)"

		Sheet3.cell(row=28, column=9).value = "=I26-I27"
		Sheet3.cell(row=28, column=10).value = "=J26-J27"
		Sheet3.cell(row=28, column=11).value = "=K26-K27"
		Sheet3.cell(row=28, column=12).value = "=L26-L27"
		Sheet3.cell(row=28, column=13).value = "=M26-M27"
		Sheet3.cell(row=28, column=14).value = "=N26-N27"
		Sheet3.cell(row=28, column=15).value = "=O26-O27"
		Sheet3.cell(row=28, column=16).value = "=P26-P27"
		Sheet3.cell(row=28, column=17).value = "=Q26-Q27"
		Sheet3.cell(row=28, column=18).value = "=R26-R27"
		Sheet3.cell(row=28, column=19).value = "=S26-S27"
		Sheet3.cell(row=28, column=20).value = "=T26-T27"
		Sheet3.cell(row=28, column=21).value = "=U26-U27"

		# procentaj
		Sheet3.cell(row=29, column=9).value = "=I28/I27"
		Sheet3.cell(row=29, column=10).value = "=J28/J27"
		Sheet3.cell(row=29, column=11).value = "=K28/K27"
		Sheet3.cell(row=29, column=12).value = "=L28/L27"
		Sheet3.cell(row=29, column=13).value = "=M28/M27"
		Sheet3.cell(row=29, column=14).value = "=N28/N27"
		Sheet3.cell(row=29, column=15).value = "=O28/O27"
		Sheet3.cell(row=29, column=16).value = "=P28/P27"
		Sheet3.cell(row=29, column=17).value = "=Q28/Q27"
		Sheet3.cell(row=29, column=18).value = "=R28/R27"
		Sheet3.cell(row=29, column=19).value = "=S28/S27"
		Sheet3.cell(row=29, column=20).value = "=T28/T27"
		Sheet3.cell(row=29, column=21).value = "=U28/U27"

		# 2020
		Sheet3.cell(row=33, column=5).value = "=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!D:D)"
		Sheet3.cell(row=33, column=6).value = "=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!E:E)"
		Sheet3.cell(row=33, column=7).value = "=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!F:F)"
		Sheet3.cell(row=33, column=8).value = "=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!G:G)"
		Sheet3.cell(row=33, column=9).value = "=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!H:H)"
		Sheet3.cell(row=33, column=10).value = "=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!I:I)"
		Sheet3.cell(row=33, column=11).value = "=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!J:J)"
		Sheet3.cell(row=33, column=12).value = "=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!K:K)"
		Sheet3.cell(row=33, column=13).value = "=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!L:L)"
		Sheet3.cell(row=33, column=14).value = "=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!M:M)"
		Sheet3.cell(row=33, column=15).value = "=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!N:N)"
		Sheet3.cell(row=33, column=16).value = "=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!O:O)"

		#bonuses
		Sheet3.cell(row=34, column=5).value ="=PBC!B7"
		Sheet3.cell(row=34, column=6).value ="=PBC!C7"
		Sheet3.cell(row=34, column=7).value ="=PBC!D7"
		Sheet3.cell(row=34, column=8).value ="=PBC!E7"
		Sheet3.cell(row=34, column=9).value ="=PBC!F7"
		Sheet3.cell(row=34, column=10).value ="=PBC!G7"
		Sheet3.cell(row=34, column=11).value ="=PBC!H7"
		Sheet3.cell(row=34, column=12).value ="=PBC!I7"
		Sheet3.cell(row=34, column=13).value ="=PBC!J7"
		Sheet3.cell(row=34, column=14).value ="=PBC!K7"
		Sheet3.cell(row=34, column=15).value ="=PBC!L7"
		Sheet3.cell(row=34, column=16).value ="=PBC!M7"

		Sheet3.cell(row=36, column=5).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!D:D)"
		Sheet3.cell(row=36, column=6).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!E:E)"
		Sheet3.cell(row=36, column=7).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!F:F)"
		Sheet3.cell(row=36, column=8).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!G:G)"
		Sheet3.cell(row=36, column=9).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!H:H)"
		Sheet3.cell(row=36, column=10).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!I:I)"
		Sheet3.cell(row=36, column=11).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!J:J)"
		Sheet3.cell(row=36, column=12).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!K:K)"
		Sheet3.cell(row=36, column=13).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!L:L)"
		Sheet3.cell(row=36, column=14).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!M:M)"
		Sheet3.cell(row=36, column=15).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!N:N)"
		Sheet3.cell(row=36, column=16).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!O:O)"

		Sheet3.cell(row=38, column=5).value = "=SUMIF('Monthly P&L CY'!B:B,645,'Monthly P&L CY'!D:D)"
		Sheet3.cell(row=38, column=6).value = "=SUMIF('Monthly P&L CY'!B:B,645,'Monthly P&L CY'!E:E)"
		Sheet3.cell(row=38, column=7).value = "=SUMIF('Monthly P&L CY'!B:B,645,'Monthly P&L CY'!F:F)"
		Sheet3.cell(row=38, column=8).value = "=SUMIF('Monthly P&L CY'!B:B,645,'Monthly P&L CY'!G:G)"
		Sheet3.cell(row=38, column=9).value = "=SUMIF('Monthly P&L CY'!B:B,645,'Monthly P&L CY'!H:H)"
		Sheet3.cell(row=38, column=10).value = "=SUMIF('Monthly P&L CY'!B:B,645,'Monthly P&L CY'!I:I)"
		Sheet3.cell(row=38, column=11).value = "=SUMIF('Monthly P&L CY'!B:B,645,'Monthly P&L CY'!J:J)"
		Sheet3.cell(row=38, column=12).value = "=SUMIF('Monthly P&L CY'!B:B,645,'Monthly P&L CY'!K:K)"
		Sheet3.cell(row=38, column=13).value = "=SUMIF('Monthly P&L CY'!B:B,645,'Monthly P&L CY'!L:L)"
		Sheet3.cell(row=38, column=14).value = "=SUMIF('Monthly P&L CY'!B:B,645,'Monthly P&L CY'!M:M)"
		Sheet3.cell(row=38, column=15).value = "=SUMIF('Monthly P&L CY'!B:B,645,'Monthly P&L CY'!N:N)"
		Sheet3.cell(row=38, column=16).value = "=SUMIF('Monthly P&L CY'!B:B,645,'Monthly P&L CY'!O:O)"

		# 2019
		Sheet3.cell(row=35, column=5).value = "=SUMIF('Monthly P&L PY'!B:B,641,'Monthly P&L PY'!D:D)"
		Sheet3.cell(row=35, column=6).value = "=SUMIF('Monthly P&L PY'!B:B,641,'Monthly P&L PY'!E:E)"
		Sheet3.cell(row=35, column=7).value = "=SUMIF('Monthly P&L PY'!B:B,641,'Monthly P&L PY'!F:F)"
		Sheet3.cell(row=35, column=8).value = "=SUMIF('Monthly P&L PY'!B:B,641,'Monthly P&L PY'!G:G)"
		Sheet3.cell(row=35, column=9).value = "=SUMIF('Monthly P&L PY'!B:B,641,'Monthly P&L PY'!H:H)"
		Sheet3.cell(row=35, column=10).value = "=SUMIF('Monthly P&L PY'!B:B,641,'Monthly P&L PY'!I:I)"
		Sheet3.cell(row=35, column=11).value = "=SUMIF('Monthly P&L PY'!B:B,641,'Monthly P&L PY'!J:J)"
		Sheet3.cell(row=35, column=12).value = "=SUMIF('Monthly P&L PY'!B:B,641,'Monthly P&L PY'!K:K)"
		Sheet3.cell(row=35, column=13).value = "=SUMIF('Monthly P&L PY'!B:B,641,'Monthly P&L PY'!L:L)"
		Sheet3.cell(row=35, column=14).value = "=SUMIF('Monthly P&L PY'!B:B,641,'Monthly P&L PY'!M:M)"
		Sheet3.cell(row=35, column=15).value = "=SUMIF('Monthly P&L PY'!B:B,641,'Monthly P&L PY'!N:N)"
		Sheet3.cell(row=35, column=16).value = "=SUMIF('Monthly P&L PY'!B:B,641,'Monthly P&L PY'!O:O)"

		Sheet3.cell(row=37, column=5).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!D:D)"
		Sheet3.cell(row=37, column=6).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!E:E)"
		Sheet3.cell(row=37, column=7).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!F:F)"
		Sheet3.cell(row=37, column=8).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!G:G)"
		Sheet3.cell(row=37, column=9).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!H:H)"
		Sheet3.cell(row=37, column=10).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!I:I)"
		Sheet3.cell(row=37, column=11).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!J:J)"
		Sheet3.cell(row=37, column=12).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!K:K)"
		Sheet3.cell(row=37, column=13).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!L:L)"
		Sheet3.cell(row=37, column=14).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!M:M)"
		Sheet3.cell(row=37, column=15).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!N:N)"
		Sheet3.cell(row=37, column=16).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!O:O)"

		Sheet3.cell(row=39, column=5).value = "=SUMIF('Monthly P&L PY'!B:B,645,'Monthly P&L PY'!D:D)"
		Sheet3.cell(row=39, column=6).value = "=SUMIF('Monthly P&L PY'!B:B,645,'Monthly P&L PY'!E:E)"
		Sheet3.cell(row=39, column=7).value = "=SUMIF('Monthly P&L PY'!B:B,645,'Monthly P&L PY'!F:F)"
		Sheet3.cell(row=39, column=8).value = "=SUMIF('Monthly P&L PY'!B:B,645,'Monthly P&L PY'!G:G)"
		Sheet3.cell(row=39, column=9).value = "=SUMIF('Monthly P&L PY'!B:B,645,'Monthly P&L PY'!H:H)"
		Sheet3.cell(row=39, column=10).value = "=SUMIF('Monthly P&L PY'!B:B,645,'Monthly P&L PY'!I:I)"
		Sheet3.cell(row=39, column=11).value = "=SUMIF('Monthly P&L PY'!B:B,645,'Monthly P&L PY'!J:J)"
		Sheet3.cell(row=39, column=12).value = "=SUMIF('Monthly P&L PY'!B:B,645,'Monthly P&L PY'!K:K)"
		Sheet3.cell(row=39, column=13).value = "=SUMIF('Monthly P&L PY'!B:B,645,'Monthly P&L PY'!L:L)"
		Sheet3.cell(row=39, column=14).value = "=SUMIF('Monthly P&L PY'!B:B,645,'Monthly P&L PY'!M:M)"
		Sheet3.cell(row=39, column=15).value = "=SUMIF('Monthly P&L PY'!B:B,645,'Monthly P&L PY'!N:N)"
		Sheet3.cell(row=39, column=16).value = "=SUMIF('Monthly P&L PY'!B:B,645,'Monthly P&L PY'!O:O)"

		Sheet3.cell(row=33, column=17).value = "=SUM(E33:P33)"
		Sheet3.cell(row=34, column=17).value = "=SUM(E34:P34)"
		Sheet3.cell(row=35, column=17).value = "=SUM(E35:P35)"
		Sheet3.cell(row=36, column=17).value = "=SUM(E36:P36)"
		Sheet3.cell(row=37, column=17).value = "=SUM(E37:P37)"
		Sheet3.cell(row=38, column=17).value = "=SUM(E38:P38)"
		Sheet3.cell(row=39, column=17).value = "=SUM(E39:P39)"

		Sheet3.cell(row=33, column=18).value = "=SUMIF('TB 12'!A:A,641,'TB 12'!H:H)"
		# Sheet3.cell(row=34, column=18).value =
		# Sheet3.cell(row=35, column=18).value =
		Sheet3.cell(row=36, column=18).value = "=SUMIF('TB 12'!A:A,642,'TB 12'!H:H)"
		# Sheet3.cell(row=37, column=18).value =
		Sheet3.cell(row=38, column=18).value = "=SUMIF('TB 12'!A:A,645,'TB 12'!H:H)"
		# Sheet3.cell(row=39, column=18).value =

		Sheet3.cell(row=33, column=19).value = "=R33-Q33"
		# Sheet3.cell(row=34, column=19).value =
		# Sheet3.cell(row=35, column=19).value =
		Sheet3.cell(row=36, column=19).value = "=R36-Q36"
		# Sheet3.cell(row=37, column=19).value =
		Sheet3.cell(row=38, column=19).value = "=R38-Q38"
		# Sheet3.cell(row=39, column=19).value =

		Sheet3.cell(row=33, column=20).value = "=Q33-Q35"
		Sheet3.cell(row=33, column=21).value = "=IFERROR(T33/Q35,0)"
		Sheet3.cell(row=36, column=20).value = "=Q36-Q37"
		Sheet3.cell(row=36, column=21).value = "=IFERROR(T36/Q37,0)"
		Sheet3.cell(row=38, column=20).value = "=Q38-Q39"
		Sheet3.cell(row=38, column=21).value = "=IFERROR(T38/Q39,0)"

		# FORMAT
		Sheet3.cell(row=1, column=1).font = f_info
		Sheet3.cell(row=2, column=1).font = f_info
		Sheet3.cell(row=3, column=1).font = f_info
		Sheet3.cell(row=3, column=2).font = font_worksheet
		Sheet3.cell(row=5, column=1).font = f_testname

		Sheet3.cell(row=1, column=2).font = font_worksheet
		Sheet3.cell(row=2, column=2).font = font_worksheet

		Sheet3.cell(row=7, column=2).font = font_worksheet
		Sheet3.cell(row=12, column=2).font = font_worksheet

		Sheet3.cell(row=22, column=2).font = ft1
		Sheet3.cell(row=22, column=3).font = ft1
		Sheet3.cell(row=22, column=4).font = ft1
		Sheet3.cell(row=22, column=5).font = ft1
		Sheet3.cell(row=22, column=6).font = ft1

		Sheet3.cell(row=23, column=2).font = ft1
		Sheet3.cell(row=24, column=2).font = ft1
		Sheet3.cell(row=25, column=2).font = ft1
		Sheet3.cell(row=26, column=2).font = ft1
		Sheet3.cell(row=27, column=2).font = ft1

		Sheet3.cell(row=22, column=8).font = ft1
		Sheet3.cell(row=22, column=9).font = ft1
		Sheet3.cell(row=22, column=10).font = ft1
		Sheet3.cell(row=22, column=11).font = ft1
		Sheet3.cell(row=22, column=12).font = ft1
		Sheet3.cell(row=22, column=13).font = ft1
		Sheet3.cell(row=22, column=14).font = ft1
		Sheet3.cell(row=22, column=15).font = ft1
		Sheet3.cell(row=22, column=16).font = ft1
		Sheet3.cell(row=22, column=17).font = ft1
		Sheet3.cell(row=22, column=18).font = ft1
		Sheet3.cell(row=22, column=19).font = ft1
		Sheet3.cell(row=22, column=20).font = ft1
		Sheet3.cell(row=22, column=21).font = ft1
		Sheet3.cell(row=22, column=22).font = ft1
		Sheet3.cell(row=22, column=23).font = ft1

		Sheet3.cell(row=23, column=8).font = ft1
		Sheet3.cell(row=24, column=8).font = ft1
		Sheet3.cell(row=26, column=8).font = ft1
		Sheet3.cell(row=27, column=8).font = ft1
		Sheet3.cell(row=28, column=8).font = ft1
		Sheet3.cell(row=29, column=8).font = ft1

		Sheet3.cell(row=31, column=2).font = ft1
		Sheet3.cell(row=32, column=2).font = ft1
		Sheet3.cell(row=32, column=3).font = ft1
		Sheet3.cell(row=32, column=4).font = ft1
		Sheet3.cell(row=38, column=2).font = ft1

		Sheet3.cell(row=33, column=2).font = font_worksheet
		Sheet3.cell(row=34, column=2).font = font_worksheet
		Sheet3.cell(row=35, column=2).font = font_worksheet
		Sheet3.cell(row=36, column=2).font = font_worksheet
		Sheet3.cell(row=37, column=2).font = font_worksheet
		Sheet3.cell(row=38, column=2).font = font_worksheet
		Sheet3.cell(row=39, column=2).font = font_worksheet

		Sheet3.cell(row=33, column=3).font = ft1
		Sheet3.cell(row=34, column=3).font = ft1
		Sheet3.cell(row=35, column=3).font = ft1
		Sheet3.cell(row=36, column=3).font = ft1
		Sheet3.cell(row=37, column=3).font = ft1
		Sheet3.cell(row=38, column=3).font = ft1
		Sheet3.cell(row=39, column=3).font = ft1

		Sheet3.cell(row=33, column=4).font = font_worksheet
		Sheet3.cell(row=34, column=4).font = font_worksheet
		Sheet3.cell(row=35, column=4).font = font_worksheet
		Sheet3.cell(row=36, column=4).font = font_worksheet
		Sheet3.cell(row=37, column=4).font = font_worksheet
		Sheet3.cell(row=38, column=4).font = font_worksheet
		Sheet3.cell(row=39, column=4).font = font_worksheet

		Sheet3.cell(row=32, column=5).font = ft1
		Sheet3.cell(row=32, column=6).font = ft1
		Sheet3.cell(row=32, column=7).font = ft1
		Sheet3.cell(row=32, column=8).font = ft1
		Sheet3.cell(row=32, column=9).font = ft1
		Sheet3.cell(row=32, column=10).font = ft1
		Sheet3.cell(row=32, column=11).font = ft1
		Sheet3.cell(row=32, column=12).font = ft1
		Sheet3.cell(row=32, column=13).font = ft1
		Sheet3.cell(row=32, column=14).font = ft1
		Sheet3.cell(row=32, column=15).font = ft1
		Sheet3.cell(row=32, column=16).font = ft1
		Sheet3.cell(row=32, column=17).font = ft1
		Sheet3.cell(row=32, column=18).font = TB_font
		Sheet3.cell(row=32, column=19).font = check_font
		Sheet3.cell(row=32, column=20).font = ft1
		Sheet3.cell(row=32, column=21).font = ft1

		# formule
		Sheet3.cell(row=23, column=9).font = font_worksheet
		Sheet3.cell(row=23, column=10).font = font_worksheet
		Sheet3.cell(row=23, column=11).font = font_worksheet
		Sheet3.cell(row=23, column=12).font = font_worksheet
		Sheet3.cell(row=23, column=13).font = font_worksheet
		Sheet3.cell(row=23, column=14).font = font_worksheet
		Sheet3.cell(row=23, column=15).font = font_worksheet
		Sheet3.cell(row=23, column=16).font = font_worksheet
		Sheet3.cell(row=23, column=17).font = font_worksheet
		Sheet3.cell(row=23, column=18).font = font_worksheet
		Sheet3.cell(row=23, column=19).font = font_worksheet
		Sheet3.cell(row=23, column=20).font = font_worksheet

		Sheet3.cell(row=24, column=9).font = font_worksheet
		Sheet3.cell(row=24, column=10).font = font_worksheet
		Sheet3.cell(row=24, column=11).font = font_worksheet
		Sheet3.cell(row=24, column=12).font = font_worksheet
		Sheet3.cell(row=24, column=13).font = font_worksheet
		Sheet3.cell(row=24, column=14).font = font_worksheet
		Sheet3.cell(row=24, column=15).font = font_worksheet
		Sheet3.cell(row=24, column=16).font = font_worksheet
		Sheet3.cell(row=24, column=17).font = font_worksheet
		Sheet3.cell(row=24, column=18).font = font_worksheet
		Sheet3.cell(row=24, column=19).font = font_worksheet
		Sheet3.cell(row=24, column=20).font = font_worksheet

		Sheet3.cell(row=23, column=3).font = font_worksheet
		Sheet3.cell(row=24, column=3).font = font_worksheet
		Sheet3.cell(row=25, column=3).font = font_worksheet
		Sheet3.cell(row=26, column=3).font = font_worksheet
		Sheet3.cell(row=27, column=3).font = font_worksheet

		Sheet3.cell(row=23, column=4).font = font_worksheet
		Sheet3.cell(row=24, column=4).font = font_worksheet
		Sheet3.cell(row=25, column=4).font = font_worksheet
		Sheet3.cell(row=26, column=4).font = font_worksheet
		Sheet3.cell(row=27, column=4).font = font_worksheet

		Sheet3.cell(row=23, column=5).font = font_worksheet
		Sheet3.cell(row=24, column=5).font = font_worksheet
		Sheet3.cell(row=25, column=5).font = font_worksheet
		Sheet3.cell(row=26, column=5).font = font_worksheet
		Sheet3.cell(row=27, column=5).font = font_worksheet

		Sheet3.cell(row=23, column=6).font = font_worksheet
		Sheet3.cell(row=24, column=6).font = font_worksheet
		Sheet3.cell(row=25, column=6).font = font_worksheet
		Sheet3.cell(row=26, column=6).font = font_worksheet
		Sheet3.cell(row=27, column=6).font = font_worksheet

		Sheet3.cell(row=23, column=21).font = font_worksheet
		Sheet3.cell(row=23, column=22).font = font_worksheet
		Sheet3.cell(row=23, column=23).font = font_worksheet
		Sheet3.cell(row=24, column=21).font = font_worksheet

		Sheet3.cell(row=26, column=9).font = font_worksheet
		Sheet3.cell(row=26, column=10).font = font_worksheet
		Sheet3.cell(row=26, column=11).font = font_worksheet
		Sheet3.cell(row=26, column=12).font = font_worksheet
		Sheet3.cell(row=26, column=13).font = font_worksheet
		Sheet3.cell(row=26, column=14).font = font_worksheet
		Sheet3.cell(row=26, column=15).font = font_worksheet
		Sheet3.cell(row=26, column=16).font = font_worksheet
		Sheet3.cell(row=26, column=17).font = font_worksheet
		Sheet3.cell(row=26, column=18).font = font_worksheet
		Sheet3.cell(row=26, column=19).font = font_worksheet
		Sheet3.cell(row=26, column=20).font = font_worksheet
		Sheet3.cell(row=26, column=21).font = font_worksheet

		Sheet3.cell(row=27, column=9).font = font_worksheet
		Sheet3.cell(row=27, column=10).font = font_worksheet
		Sheet3.cell(row=27, column=11).font = font_worksheet
		Sheet3.cell(row=27, column=12).font = font_worksheet
		Sheet3.cell(row=27, column=13).font = font_worksheet
		Sheet3.cell(row=27, column=14).font = font_worksheet
		Sheet3.cell(row=27, column=15).font = font_worksheet
		Sheet3.cell(row=27, column=16).font = font_worksheet
		Sheet3.cell(row=27, column=17).font = font_worksheet
		Sheet3.cell(row=27, column=18).font = font_worksheet
		Sheet3.cell(row=27, column=19).font = font_worksheet
		Sheet3.cell(row=27, column=20).font = font_worksheet
		Sheet3.cell(row=27, column=21).font = font_worksheet

		Sheet3.cell(row=28, column=9).font = ft1
		Sheet3.cell(row=28, column=10).font = ft1
		Sheet3.cell(row=28, column=11).font = ft1
		Sheet3.cell(row=28, column=12).font = ft1
		Sheet3.cell(row=28, column=13).font = ft1
		Sheet3.cell(row=28, column=14).font = ft1
		Sheet3.cell(row=28, column=15).font = ft1
		Sheet3.cell(row=28, column=16).font = ft1
		Sheet3.cell(row=28, column=17).font = ft1
		Sheet3.cell(row=28, column=18).font = ft1
		Sheet3.cell(row=28, column=19).font = ft1
		Sheet3.cell(row=28, column=20).font = ft1
		Sheet3.cell(row=28, column=21).font = ft1

		# procentaj
		Sheet3.cell(row=29, column=9).font = ft1
		Sheet3.cell(row=29, column=10).font = ft1
		Sheet3.cell(row=29, column=11).font = ft1
		Sheet3.cell(row=29, column=12).font = ft1
		Sheet3.cell(row=29, column=13).font = ft1
		Sheet3.cell(row=29, column=14).font = ft1
		Sheet3.cell(row=29, column=15).font = ft1
		Sheet3.cell(row=29, column=16).font = ft1
		Sheet3.cell(row=29, column=17).font = ft1
		Sheet3.cell(row=29, column=18).font = ft1
		Sheet3.cell(row=29, column=19).font = ft1
		Sheet3.cell(row=29, column=20).font = ft1
		Sheet3.cell(row=29, column=21).font = ft1

		# 2020
		Sheet3.cell(row=33, column=5).font = font_worksheet
		Sheet3.cell(row=33, column=6).font = font_worksheet
		Sheet3.cell(row=33, column=7).font = font_worksheet
		Sheet3.cell(row=33, column=8).font = font_worksheet
		Sheet3.cell(row=33, column=9).font = font_worksheet
		Sheet3.cell(row=33, column=10).font = font_worksheet
		Sheet3.cell(row=33, column=11).font = font_worksheet
		Sheet3.cell(row=33, column=12).font = font_worksheet
		Sheet3.cell(row=33, column=13).font = font_worksheet
		Sheet3.cell(row=33, column=14).font = font_worksheet
		Sheet3.cell(row=33, column=15).font = font_worksheet
		Sheet3.cell(row=33, column=16).font = font_worksheet

		Sheet3.cell(row=36, column=5).font = font_worksheet
		Sheet3.cell(row=36, column=6).font = font_worksheet
		Sheet3.cell(row=36, column=7).font = font_worksheet
		Sheet3.cell(row=36, column=8).font = font_worksheet
		Sheet3.cell(row=36, column=9).font = font_worksheet
		Sheet3.cell(row=36, column=10).font = font_worksheet
		Sheet3.cell(row=36, column=11).font = font_worksheet
		Sheet3.cell(row=36, column=12).font = font_worksheet
		Sheet3.cell(row=36, column=13).font = font_worksheet
		Sheet3.cell(row=36, column=14).font = font_worksheet
		Sheet3.cell(row=36, column=15).font = font_worksheet
		Sheet3.cell(row=36, column=16).font = font_worksheet

		Sheet3.cell(row=38, column=5).font = font_worksheet
		Sheet3.cell(row=38, column=6).font = font_worksheet
		Sheet3.cell(row=38, column=7).font = font_worksheet
		Sheet3.cell(row=38, column=8).font = font_worksheet
		Sheet3.cell(row=38, column=9).font = font_worksheet
		Sheet3.cell(row=38, column=10).font = font_worksheet
		Sheet3.cell(row=38, column=11).font = font_worksheet
		Sheet3.cell(row=38, column=12).font = font_worksheet
		Sheet3.cell(row=38, column=13).font = font_worksheet
		Sheet3.cell(row=38, column=14).font = font_worksheet
		Sheet3.cell(row=38, column=15).font = font_worksheet
		Sheet3.cell(row=38, column=16).font = font_worksheet

		# 2019
		Sheet3.cell(row=35, column=5).font = font_worksheet
		Sheet3.cell(row=35, column=6).font = font_worksheet
		Sheet3.cell(row=35, column=7).font = font_worksheet
		Sheet3.cell(row=35, column=8).font = font_worksheet
		Sheet3.cell(row=35, column=9).font = font_worksheet
		Sheet3.cell(row=35, column=10).font = font_worksheet
		Sheet3.cell(row=35, column=11).font = font_worksheet
		Sheet3.cell(row=35, column=12).font = font_worksheet
		Sheet3.cell(row=35, column=13).font = font_worksheet
		Sheet3.cell(row=35, column=14).font = font_worksheet
		Sheet3.cell(row=35, column=15).font = font_worksheet
		Sheet3.cell(row=35, column=16).font = font_worksheet

		Sheet3.cell(row=37, column=5).font = font_worksheet
		Sheet3.cell(row=37, column=6).font = font_worksheet
		Sheet3.cell(row=37, column=7).font = font_worksheet
		Sheet3.cell(row=37, column=8).font = font_worksheet
		Sheet3.cell(row=37, column=9).font = font_worksheet
		Sheet3.cell(row=37, column=10).font = font_worksheet
		Sheet3.cell(row=37, column=11).font = font_worksheet
		Sheet3.cell(row=37, column=12).font = font_worksheet
		Sheet3.cell(row=37, column=13).font = font_worksheet
		Sheet3.cell(row=37, column=14).font = font_worksheet
		Sheet3.cell(row=37, column=15).font = font_worksheet
		Sheet3.cell(row=37, column=16).font = font_worksheet

		Sheet3.cell(row=35, column=5).font = font_worksheet
		Sheet3.cell(row=35, column=6).font = font_worksheet
		Sheet3.cell(row=35, column=7).font = font_worksheet
		Sheet3.cell(row=35, column=8).font = font_worksheet
		Sheet3.cell(row=35, column=9).font = font_worksheet
		Sheet3.cell(row=35, column=10).font = font_worksheet
		Sheet3.cell(row=35, column=11).font = font_worksheet
		Sheet3.cell(row=35, column=12).font = font_worksheet
		Sheet3.cell(row=35, column=13).font = font_worksheet
		Sheet3.cell(row=35, column=14).font = font_worksheet
		Sheet3.cell(row=35, column=15).font = font_worksheet
		Sheet3.cell(row=35, column=16).font = font_worksheet

		Sheet3.cell(row=33, column=17).font = font_worksheet
		Sheet3.cell(row=34, column=17).font = font_worksheet
		Sheet3.cell(row=35, column=17).font = font_worksheet
		Sheet3.cell(row=36, column=17).font = font_worksheet
		Sheet3.cell(row=37, column=17).font = font_worksheet
		Sheet3.cell(row=38, column=17).font = font_worksheet
		Sheet3.cell(row=39, column=17).font = font_worksheet

		Sheet3.cell(row=33, column=18).font = TB_font
		# Sheet3.cell(row=34, column=18).font =
		# Sheet3.cell(row=35, column=18).font =
		Sheet3.cell(row=36, column=18).font = TB_font
		# Sheet3.cell(row=37, column=18).font =
		Sheet3.cell(row=38, column=18).font = TB_font
		# Sheet3.cell(row=39, column=18).font =

		Sheet3.cell(row=33, column=19).font = check_font_1
		# Sheet3.cell(row=34, column=19).font =
		# Sheet3.cell(row=35, column=19).font =
		Sheet3.cell(row=36, column=19).font = check_font_1
		# Sheet3.cell(row=37, column=19).font =
		Sheet3.cell(row=38, column=19).font = check_font_1
		# Sheet3.cell(row=39, column=19).font =

		Sheet3.cell(row=33, column=20).font = font_worksheet
		Sheet3.cell(row=33, column=21).font = font_worksheet
		Sheet3.cell(row=36, column=20).font = font_worksheet
		Sheet3.cell(row=36, column=21).font = font_worksheet
		Sheet3.cell(row=38, column=20).font = font_worksheet
		Sheet3.cell(row=38, column=21).font = font_worksheet

		# number formats
		Sheet3.cell(row=23, column=9).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=23, column=10).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=23, column=11).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=23, column=12).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=23, column=13).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=23, column=14).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=23, column=15).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=23, column=16).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=23, column=17).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=23, column=18).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=23, column=19).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=23, column=20).number_format = '#,##0_);(#,##0)'

		Sheet3.cell(row=24, column=9).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=24, column=10).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=24, column=11).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=24, column=12).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=24, column=13).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=24, column=14).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=24, column=15).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=24, column=16).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=24, column=17).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=24, column=18).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=24, column=19).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=24, column=20).number_format = '#,##0_);(#,##0)'

		Sheet3.cell(row=23, column=3).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=24, column=3).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=25, column=3).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=26, column=3).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=27, column=3).number_format = '#,##0_);(#,##0)'

		Sheet3.cell(row=23, column=4).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=24, column=4).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=25, column=4).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=26, column=4).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=27, column=4).number_format = '#,##0_);(#,##0)'

		Sheet3.cell(row=23, column=5).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=24, column=5).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=25, column=5).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=26, column=5).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=27, column=5).number_format = '#,##0_);(#,##0)'

		Sheet3.cell(row=23, column=6).style = 'Percent'
		Sheet3.cell(row=24, column=6).style = 'Percent'
		Sheet3.cell(row=25, column=6).style = 'Percent'
		Sheet3.cell(row=26, column=6).style = 'Percent'
		Sheet3.cell(row=27, column=6).style = 'Percent'

		Sheet3.cell(row=23, column=21).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=23, column=22).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=23, column=23).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=24, column=21).number_format = '#,##0_);(#,##0)'

		Sheet3.cell(row=26, column=9).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=26, column=10).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=26, column=11).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=26, column=12).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=26, column=13).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=26, column=14).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=26, column=15).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=26, column=16).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=26, column=17).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=26, column=18).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=26, column=19).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=26, column=20).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=26, column=21).number_format = '#,##0_);(#,##0)'

		Sheet3.cell(row=27, column=9).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=27, column=10).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=27, column=11).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=27, column=12).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=27, column=13).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=27, column=14).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=27, column=15).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=27, column=16).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=27, column=17).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=27, column=18).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=27, column=19).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=27, column=20).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=27, column=21).number_format = '#,##0_);(#,##0)'

		Sheet3.cell(row=28, column=9).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=28, column=10).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=28, column=11).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=28, column=12).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=28, column=13).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=28, column=14).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=28, column=15).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=28, column=16).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=28, column=17).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=28, column=18).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=28, column=19).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=28, column=20).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=28, column=21).number_format = '#,##0_);(#,##0)'

		# procentaj
		Sheet3.cell(row=29, column=9).style = 'Percent'
		Sheet3.cell(row=29, column=10).style = 'Percent'
		Sheet3.cell(row=29, column=11).style = 'Percent'
		Sheet3.cell(row=29, column=12).style = 'Percent'
		Sheet3.cell(row=29, column=13).style = 'Percent'
		Sheet3.cell(row=29, column=14).style = 'Percent'
		Sheet3.cell(row=29, column=15).style = 'Percent'
		Sheet3.cell(row=29, column=16).style = 'Percent'
		Sheet3.cell(row=29, column=17).style = 'Percent'
		Sheet3.cell(row=29, column=18).style = 'Percent'
		Sheet3.cell(row=29, column=19).style = 'Percent'
		Sheet3.cell(row=29, column=20).style = 'Percent'
		Sheet3.cell(row=29, column=21).style = 'Percent'

		# 2020
		Sheet3.cell(row=33, column=5).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=33, column=6).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=33, column=7).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=33, column=8).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=33, column=9).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=33, column=10).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=33, column=11).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=33, column=12).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=33, column=13).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=33, column=14).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=33, column=15).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=33, column=16).number_format = '#,##0_);(#,##0)'

		Sheet3.cell(row=36, column=5).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=36, column=6).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=36, column=7).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=36, column=8).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=36, column=9).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=36, column=10).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=36, column=11).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=36, column=12).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=36, column=13).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=36, column=14).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=36, column=15).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=36, column=16).number_format = '#,##0_);(#,##0)'

		Sheet3.cell(row=38, column=5).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=38, column=6).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=38, column=7).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=38, column=8).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=38, column=9).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=38, column=10).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=38, column=11).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=38, column=12).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=38, column=13).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=38, column=14).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=38, column=15).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=38, column=16).number_format = '#,##0_);(#,##0)'

		# 2019
		Sheet3.cell(row=35, column=5).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=6).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=7).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=8).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=9).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=10).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=11).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=12).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=13).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=14).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=15).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=16).number_format = '#,##0_);(#,##0)'

		Sheet3.cell(row=37, column=5).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=37, column=6).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=37, column=7).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=37, column=8).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=37, column=9).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=37, column=10).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=37, column=11).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=37, column=12).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=37, column=13).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=37, column=14).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=37, column=15).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=37, column=16).number_format = '#,##0_);(#,##0)'

		Sheet3.cell(row=35, column=5).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=6).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=7).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=8).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=9).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=10).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=11).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=12).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=13).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=14).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=15).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=16).number_format = '#,##0_);(#,##0)'

		Sheet3.cell(row=33, column=17).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=34, column=17).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=35, column=17).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=36, column=17).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=37, column=17).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=38, column=17).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=39, column=17).number_format = '#,##0_);(#,##0)'

		Sheet3.cell(row=33, column=18).number_format = '#,##0_);(#,##0)'
		# Sheet3.cell(row=34, column=18).font =
		# Sheet3.cell(row=35, column=18).font =
		Sheet3.cell(row=36, column=18).number_format = '#,##0_);(#,##0)'
		# Sheet3.cell(row=37, column=18).font =
		Sheet3.cell(row=38, column=18).number_format = '#,##0_);(#,##0)'
		# Sheet3.cell(row=39, column=18).font =

		Sheet3.cell(row=33, column=19).number_format = '#,##0_);(#,##0)'
		# Sheet3.cell(row=34, column=19).font =
		# Sheet3.cell(row=35, column=19).font =
		Sheet3.cell(row=36, column=19).number_format = '#,##0_);(#,##0)'
		# Sheet3.cell(row=37, column=19).font =
		Sheet3.cell(row=38, column=19).number_format = '#,##0_);(#,##0)'
		# Sheet3.cell(row=39, column=19).font =

		Sheet3.cell(row=33, column=20).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=33, column=21).style = 'Percent'
		Sheet3.cell(row=36, column=20).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=36, column=21).style = 'Percent'
		Sheet3.cell(row=38, column=20).number_format = '#,##0_);(#,##0)'
		Sheet3.cell(row=38, column=21).style = 'Percent'

		Sheet3['B22'].border = border_left
		Sheet3['C22'].border = border_centered
		Sheet3['D22'].border = border_right

		Sheet3['E22'].border = border_left
		Sheet3['F22'].border = border_right

		Sheet3['B23'].border = border_upperleft

		for row in Sheet3['B24:B26']:
			for cell in row:
				cell.border = border_left1

		Sheet3['B27'].border = border_lowerleft
		Sheet3['C27'].border = border_bottom
		Sheet3['D27'].border = border_lowerright

		for row in Sheet3['D24:D26']:
			for cell in row:
				cell.border = border_right1

		Sheet3['D23'].border = border_upperright

		Sheet3['E27'].border = border_bottom
		Sheet3['F27'].border = border_lowerright

		for row in Sheet3['F23:F26']:
			for cell in row:
				cell.border = border_right1

		Sheet3['H22'].border = border_left

		for row in Sheet3['I22:T22']:
			for cell in row:
				cell.border = border_centered

		Sheet3['U22'].border = border_right

		Sheet3['V22'].border = border_left
		Sheet3['W22'].border = border_right

		Sheet3['H23'].border = border_upperleft

		for row in Sheet3['H24:H28']:
			for cell in row:
				cell.border = border_left1

		Sheet3['H29'].border = border_lowerleft

		for row in Sheet3['I29:T29']:
			for cell in row:
				cell.border = border_bottom

		Sheet3['U29'].border = border_lowerright

		for row in Sheet3['U23:U28']:
			for cell in row:
				cell.border = border_right1

		Sheet3['V29'].border = border_bottom
		Sheet3['W29'].border = border_lowerright

		for row in Sheet3['W23:W28']:
			for cell in row:
				cell.border = border_right1

		for row in Sheet3['B32:U32']:
			for cell in row:
				cell.border = border_centered

		for row in Sheet3['B33:U33']:
			for cell in row:
				cell.border = border_bottom

		for row in Sheet3['B35:U35']:
			for cell in row:
				cell.border = border_centered

		for row in Sheet3['B37:U37']:
			for cell in row:
				cell.border = border_bottom

		for row in Sheet3['B39:U39']:
			for cell in row:
				cell.border = border_centered

		# for row in Sheet3['R32:R39']:
		#     for cell in row:
		#         cell.fill = TB_font

		Sheet3.column_dimensions['A'].width = 10
		Sheet3.column_dimensions['B'].width = 20
		Sheet3.column_dimensions['C'].width = 11
		Sheet3.column_dimensions['D'].width = 12
		Sheet3.column_dimensions['E'].width = 14
		Sheet3.column_dimensions['F'].width = 12
		Sheet3.column_dimensions['G'].width = 10
		Sheet3.column_dimensions['H'].width = 23
		Sheet3.column_dimensions['I'].width = 9
		Sheet3.column_dimensions['J'].width = 9
		Sheet3.column_dimensions['K'].width = 9
		Sheet3.column_dimensions['L'].width = 9
		Sheet3.column_dimensions['M'].width = 9
		Sheet3.column_dimensions['N'].width = 9
		Sheet3.column_dimensions['O'].width = 12
		Sheet3.column_dimensions['P'].width = 15
		Sheet3.column_dimensions['Q'].width = 15
		Sheet3.column_dimensions['R'].width = 15
		Sheet3.column_dimensions['S'].width = 15
		Sheet3.column_dimensions['T'].width = 15
		Sheet3.column_dimensions['U'].width = 15
		Sheet3.column_dimensions['V'].width = 15
		Sheet3.column_dimensions['W'].width = 15

		# Sheet3 = wb.active

		pbc = openpyxl.load_workbook(openPBC, data_only=True)
		pbc_1 = pbc.active

		Sheet4 = wb.create_sheet("V10.2 Contribution Reaso")

		Sheet4.sheet_view.showGridLines = False

		Sheet4.cell(row=1, column=1).value = "Client Name:"
		Sheet4.cell(row=2, column=1).value = "Period ended:"
		Sheet4.cell(row=3, column=1).value = "Audit area:"
		Sheet4.cell(row=3, column=2).value = "Payroll"
		Sheet4.cell(row=5, column=1).value = "Contribution Reasonableness"

		try:
			Sheet4.cell(row=1, column=2).value = clientname1
			Sheet4.cell(row=2, column=2).value = periodEnd1
		except:
			None

		Sheet4.cell(row=7, column=2).value = "Work done"
		Sheet4.cell(row=8, column=2).value = "We have divided the contribution costs paid to the payroll costs in order to see if these are in accordance with the law percent."
		Sheet4.cell(row=9, column=2).value = "We have checked and analyzed the percentage obtained with the law percent."

		Sheet4.cell(row=11, column=2).value = "Conclusion"

		Sheet4.cell(row=16, column=2).value = "Synt"
		Sheet4.cell(row=16, column=3).value = "Account"
		Sheet4.cell(row=16, column=4).value = "Denumirea contului"
		Sheet4.cell(row=16, column=5).value = "January"
		Sheet4.cell(row=16, column=6).value = "February"
		Sheet4.cell(row=16, column=7).value = "March"
		Sheet4.cell(row=16, column=8).value = "April"
		Sheet4.cell(row=16, column=9).value = "May"
		Sheet4.cell(row=16, column=10).value = "June"
		Sheet4.cell(row=16, column=11).value = "July"
		Sheet4.cell(row=16, column=12).value = "August"
		Sheet4.cell(row=16, column=13).value = "September"
		Sheet4.cell(row=16, column=14).value = "October"
		Sheet4.cell(row=16, column=15).value = "November"
		Sheet4.cell(row=16, column=16).value = "December"
		Sheet4.cell(row=16, column=17).value = "Total"
		Sheet4.cell(row=16, column=18).value = "As per TB"
		Sheet4.cell(row=16, column=19).value = "Check"

		Sheet4.cell(row=17, column=2).value = "646"
		Sheet4.cell(row=17, column=3).value = "646000"
		Sheet4.cell(row=17, column=4).value = "Cheltuieli privind contributia asiguratorie pentru munca"

		Sheet4.cell(row=20, column=2).value = "Synt"
		Sheet4.cell(row=20, column=3).value = "Account"
		Sheet4.cell(row=20, column=4).value = "Denumirea contului"
		Sheet4.cell(row=20, column=5).value = "January"
		Sheet4.cell(row=20, column=6).value = "February"
		Sheet4.cell(row=20, column=7).value = "March"
		Sheet4.cell(row=20, column=8).value = "April"
		Sheet4.cell(row=20, column=9).value = "May"
		Sheet4.cell(row=20, column=10).value = "June"
		Sheet4.cell(row=20, column=11).value = "July"
		Sheet4.cell(row=20, column=12).value = "August"
		Sheet4.cell(row=20, column=13).value = "September"
		Sheet4.cell(row=20, column=14).value = "October"
		Sheet4.cell(row=20, column=15).value = "November"
		Sheet4.cell(row=20, column=16).value = "December"
		Sheet4.cell(row=20, column=17).value = "Total"
		Sheet4.cell(row=20, column=18).value = "As per TB"
		Sheet4.cell(row=20, column=19).value = "Check"

		Sheet4.cell(row=21, column=2).value = "641"
		Sheet4.cell(row=21, column=3).value = "641"
		Sheet4.cell(row=21, column=4).value = "Cheltuieli cu salariile personalului "

		Sheet4.cell(row=24, column=1).value = "Check for reasonableness"
		Sheet4.cell(row=24, column=3).value = "Please check if the percentage is the same as provided by the legislation on the Labor Law in force."

		Sheet4.cell(row=26, column=4).value = "Contrib Angajator"
		Sheet4.cell(row=27, column=4).value = "2.25%"
		Sheet4.cell(row=26, column=5).value = "January"
		Sheet4.cell(row=26, column=6).value = "February"
		Sheet4.cell(row=26, column=7).value = "March"
		Sheet4.cell(row=26, column=8).value = "April"
		Sheet4.cell(row=26, column=9).value = "May"
		Sheet4.cell(row=26, column=10).value = "June"
		Sheet4.cell(row=26, column=11).value = "July"
		Sheet4.cell(row=26, column=12).value = "August"
		Sheet4.cell(row=26, column=13).value = "September"
		Sheet4.cell(row=26, column=14).value = "October"
		Sheet4.cell(row=26, column=15).value = "November"
		Sheet4.cell(row=26, column=16).value = "December"
		Sheet4.cell(row=26, column=17).value = "Average"

		Sheet4.cell(row=30, column=5).value = "January"
		Sheet4.cell(row=30, column=6).value = "February"
		Sheet4.cell(row=30, column=7).value = "March"
		Sheet4.cell(row=30, column=8).value = "April"
		Sheet4.cell(row=30, column=9).value = "May"
		Sheet4.cell(row=30, column=10).value = "June"
		Sheet4.cell(row=30, column=11).value = "July"
		Sheet4.cell(row=30, column=12).value = "August"
		Sheet4.cell(row=30, column=13).value = "September"
		Sheet4.cell(row=30, column=14).value = "October"
		Sheet4.cell(row=30, column=15).value = "November"
		Sheet4.cell(row=30, column=16).value = "December"
		Sheet4.cell(row=30, column=17).value = "Total"

		Sheet4.cell(row=31, column=4).value = "CAM"
		Sheet4.cell(row=32, column=4).value = "2.25%"
		Sheet4.cell(row=33, column=4).value = "Check"

		Sheet4.cell(row=24, column=3).value = "Please check if the percentage is the same as provided by the legislation on the Labor Law in force."

		# FORMULE
		Sheet4.cell(row=17, column=5).value = "=SUMIF('Monthly P&L CY'!B:B,646,'Monthly P&L CY'!D:D)"
		Sheet4.cell(row=17, column=6).value = "=SUMIF('Monthly P&L CY'!B:B,646,'Monthly P&L CY'!E:E)"
		Sheet4.cell(row=17, column=7).value = "=SUMIF('Monthly P&L CY'!B:B,646,'Monthly P&L CY'!F:F)"
		Sheet4.cell(row=17, column=8).value = "=SUMIF('Monthly P&L CY'!B:B,646,'Monthly P&L CY'!G:G)"
		Sheet4.cell(row=17, column=9).value = "=SUMIF('Monthly P&L CY'!B:B,646,'Monthly P&L CY'!H:H)"
		Sheet4.cell(row=17, column=10).value = "=SUMIF('Monthly P&L CY'!B:B,646,'Monthly P&L CY'!I:I)"
		Sheet4.cell(row=17, column=11).value = "=SUMIF('Monthly P&L CY'!B:B,646,'Monthly P&L CY'!J:J)"
		Sheet4.cell(row=17, column=12).value = "=SUMIF('Monthly P&L CY'!B:B,646,'Monthly P&L CY'!K:K)"
		Sheet4.cell(row=17, column=13).value = "=SUMIF('Monthly P&L CY'!B:B,646,'Monthly P&L CY'!L:L)"
		Sheet4.cell(row=17, column=14).value = "=SUMIF('Monthly P&L CY'!B:B,646,'Monthly P&L CY'!M:M)"
		Sheet4.cell(row=17, column=15).value = "=SUMIF('Monthly P&L CY'!B:B,646,'Monthly P&L CY'!N:N)"
		Sheet4.cell(row=17, column=16).value = "=SUMIF('Monthly P&L CY'!B:B,646,'Monthly P&L CY'!O:O)"

		Sheet4.cell(row=17, column=17).value = "=SUM(E17:P17)"
		Sheet4.cell(row=17, column=18).value = "=SUMIF('TB 12'!A:A,646,'TB 12'!H:H)"
		Sheet4.cell(row=17, column=19).value = "=R17-Q17"

		Sheet4.cell(row=21, column=5).value = "='V10.1 Payroll Analysis'!E33"
		Sheet4.cell(row=21, column=6).value = "='V10.1 Payroll Analysis'!F33"
		Sheet4.cell(row=21, column=7).value = "='V10.1 Payroll Analysis'!G33"
		Sheet4.cell(row=21, column=8).value = "='V10.1 Payroll Analysis'!H33"
		Sheet4.cell(row=21, column=9).value = "='V10.1 Payroll Analysis'!I33"
		Sheet4.cell(row=21, column=10).value = "='V10.1 Payroll Analysis'!J33"
		Sheet4.cell(row=21, column=11).value = "='V10.1 Payroll Analysis'!K33"
		Sheet4.cell(row=21, column=12).value = "='V10.1 Payroll Analysis'!L33"
		Sheet4.cell(row=21, column=13).value = "='V10.1 Payroll Analysis'!M33"
		Sheet4.cell(row=21, column=14).value = "='V10.1 Payroll Analysis'!N33"
		Sheet4.cell(row=21, column=15).value = "='V10.1 Payroll Analysis'!O33"
		Sheet4.cell(row=21, column=16).value = "='V10.1 Payroll Analysis'!P33"

		Sheet4.cell(row=21, column=17).value = "=SUM(E21:P21)"
		Sheet4.cell(row=21, column=18).value = "=SUMIF('TB 12'!A:A,641,'TB 12'!H:H)"
		Sheet4.cell(row=21, column=19).value = "=R21-Q21"

		Sheet4.cell(row=27, column=5).value = "=E17/E21"
		Sheet4.cell(row=27, column=6).value = "=F17/F21"
		Sheet4.cell(row=27, column=7).value = "=G17/G21"
		Sheet4.cell(row=27, column=8).value = "=H17/H21"
		Sheet4.cell(row=27, column=9).value = "=I17/I21"
		Sheet4.cell(row=27, column=10).value = "=J17/J21"
		Sheet4.cell(row=27, column=11).value = "=K17/K21"
		Sheet4.cell(row=27, column=12).value = "=L17/L21"
		Sheet4.cell(row=27, column=13).value = "=M17/M21"
		Sheet4.cell(row=27, column=14).value = "=N17/N21"
		Sheet4.cell(row=27, column=15).value = "=O17/O21"
		Sheet4.cell(row=27, column=16).value = "=P17/P21"
		Sheet4.cell(row=27, column=17).value = "=AVERAGE(E27:P27)"

		Sheet4.cell(row=31, column=5).value = "=D32*E21"
		Sheet4.cell(row=31, column=6).value = "=D32*F21"
		Sheet4.cell(row=31, column=7).value = "=D32*G21"
		Sheet4.cell(row=31, column=8).value = "=D32*H21"
		Sheet4.cell(row=31, column=9).value = "=D32*I21"
		Sheet4.cell(row=31, column=10).value = "=D32*J21"
		Sheet4.cell(row=31, column=11).value = "=D32*K21"
		Sheet4.cell(row=31, column=12).value = "=D32*L21"
		Sheet4.cell(row=31, column=13).value = "=D32*M21"
		Sheet4.cell(row=31, column=14).value = "=D32*N21"
		Sheet4.cell(row=31, column=15).value = "=D32*O21"
		Sheet4.cell(row=31, column=16).value = "=D32*P21"
		Sheet4.cell(row=31, column=17).value = "=SUM(E31:P31)"

		Sheet4.cell(row=33, column=5).value = "=E31-E17"
		Sheet4.cell(row=33, column=6).value = "=F31-F17"
		Sheet4.cell(row=33, column=7).value = "=G31-G17"
		Sheet4.cell(row=33, column=8).value = "=H31-H17"
		Sheet4.cell(row=33, column=9).value = "=I31-I17"
		Sheet4.cell(row=33, column=10).value = "=J31-J17"
		Sheet4.cell(row=33, column=11).value = "=K31-K17"
		Sheet4.cell(row=33, column=12).value = "=L31-L17"
		Sheet4.cell(row=33, column=13).value = "=M31-M17"
		Sheet4.cell(row=33, column=14).value = "=N31-N17"
		Sheet4.cell(row=33, column=15).value = "=O31-O17"
		Sheet4.cell(row=33, column=16).value = "=P31-P17"
		Sheet4.cell(row=33, column=17).value = "=SUM(E33:P33)"

		# FORMAT

		Sheet4.cell(row=1, column=1).font = f_info
		Sheet4.cell(row=2, column=1).font = f_info
		Sheet4.cell(row=3, column=1).font = f_info
		Sheet4.cell(row=3, column=2).font = font_worksheet
		Sheet4.cell(row=5, column=1).font = f_testname

		Sheet4.cell(row=1, column=2).font = font_worksheet
		Sheet4.cell(row=2, column=2).font = font_worksheet

		Sheet4.cell(row=7, column=2).font = ft1
		Sheet4.cell(row=8, column=2).font = font_worksheet
		Sheet4.cell(row=9, column=2).font = font_worksheet
		Sheet4.cell(row=11, column=2).font = ft1

		Sheet4.cell(row=16, column=2).font = ft1
		Sheet4.cell(row=16, column=3).font = ft1
		Sheet4.cell(row=16, column=4).font = ft1
		Sheet4.cell(row=16, column=5).font = ft1
		Sheet4.cell(row=16, column=6).font = ft1
		Sheet4.cell(row=16, column=7).font = ft1
		Sheet4.cell(row=16, column=8).font = ft1
		Sheet4.cell(row=16, column=9).font = ft1
		Sheet4.cell(row=16, column=10).font = ft1
		Sheet4.cell(row=16, column=11).font = ft1
		Sheet4.cell(row=16, column=12).font = ft1
		Sheet4.cell(row=16, column=13).font = ft1
		Sheet4.cell(row=16, column=14).font = ft1
		Sheet4.cell(row=16, column=15).font = ft1
		Sheet4.cell(row=16, column=16).font = ft1
		Sheet4.cell(row=16, column=17).font = ft1
		Sheet4.cell(row=16, column=18).font = TB_font
		Sheet4.cell(row=16, column=19).font = check_font

		Sheet4.cell(row=17, column=2).font = ft1
		Sheet4.cell(row=17, column=3).font = ft1
		Sheet4.cell(row=17, column=4).font = ft1

		Sheet4.cell(row=20, column=2).font = ft1
		Sheet4.cell(row=20, column=3).font = ft1
		Sheet4.cell(row=20, column=4).font = ft1
		Sheet4.cell(row=20, column=5).font = ft1
		Sheet4.cell(row=20, column=6).font = ft1
		Sheet4.cell(row=20, column=7).font = ft1
		Sheet4.cell(row=20, column=8).font = ft1
		Sheet4.cell(row=20, column=9).font = ft1
		Sheet4.cell(row=20, column=10).font = ft1
		Sheet4.cell(row=20, column=11).font = ft1
		Sheet4.cell(row=20, column=12).font = ft1
		Sheet4.cell(row=20, column=13).font = ft1
		Sheet4.cell(row=20, column=14).font = ft1
		Sheet4.cell(row=20, column=15).font = ft1
		Sheet4.cell(row=20, column=16).font = ft1
		Sheet4.cell(row=20, column=17).font = ft1
		Sheet4.cell(row=20, column=18).font = TB_font
		Sheet4.cell(row=20, column=19).font = check_font

		Sheet4.cell(row=21, column=2).font = ft1
		Sheet4.cell(row=21, column=3).font = ft1
		Sheet4.cell(row=21, column=4).font = ft1

		Sheet4.cell(row=24, column=1).font = ft1

		Sheet4.cell(row=26, column=4).font = ft1
		Sheet4.cell(row=27, column=4).font = ft1
		Sheet4.cell(row=26, column=5).font = ft1
		Sheet4.cell(row=26, column=6).font = ft1
		Sheet4.cell(row=26, column=7).font = ft1
		Sheet4.cell(row=26, column=8).font = ft1
		Sheet4.cell(row=26, column=9).font = ft1
		Sheet4.cell(row=26, column=10).font = ft1
		Sheet4.cell(row=26, column=11).font = ft1
		Sheet4.cell(row=26, column=12).font = ft1
		Sheet4.cell(row=26, column=13).font = ft1
		Sheet4.cell(row=26, column=14).font = ft1
		Sheet4.cell(row=26, column=15).font = ft1
		Sheet4.cell(row=26, column=16).font = ft1
		Sheet4.cell(row=26, column=17).font = ft1

		Sheet4.cell(row=30, column=5).font = ft1
		Sheet4.cell(row=30, column=6).font = ft1
		Sheet4.cell(row=30, column=7).font = ft1
		Sheet4.cell(row=30, column=8).font = ft1
		Sheet4.cell(row=30, column=9).font = ft1
		Sheet4.cell(row=30, column=10).font = ft1
		Sheet4.cell(row=30, column=11).font = ft1
		Sheet4.cell(row=30, column=12).font = ft1
		Sheet4.cell(row=30, column=13).font = ft1
		Sheet4.cell(row=30, column=14).font = ft1
		Sheet4.cell(row=30, column=15).font = ft1
		Sheet4.cell(row=30, column=16).font = ft1
		Sheet4.cell(row=30, column=17).font = ft1

		Sheet4.cell(row=31, column=4).font = ft1
		Sheet4.cell(row=32, column=4).font = ft1
		Sheet4.cell(row=33, column=4).font = check_font

		Sheet4.cell(row=24, column=3).font = check_font_1

		# FORMULE
		Sheet4.cell(row=17, column=5).font = font_worksheet
		Sheet4.cell(row=17, column=6).font = font_worksheet
		Sheet4.cell(row=17, column=7).font = font_worksheet
		Sheet4.cell(row=17, column=8).font = font_worksheet
		Sheet4.cell(row=17, column=9).font = font_worksheet
		Sheet4.cell(row=17, column=10).font = font_worksheet
		Sheet4.cell(row=17, column=11).font = font_worksheet
		Sheet4.cell(row=17, column=12).font = font_worksheet
		Sheet4.cell(row=17, column=13).font = font_worksheet
		Sheet4.cell(row=17, column=14).font = font_worksheet
		Sheet4.cell(row=17, column=15).font = font_worksheet
		Sheet4.cell(row=17, column=16).font = font_worksheet

		Sheet4.cell(row=17, column=17).font = font_worksheet
		Sheet4.cell(row=17, column=18).font = TB_font
		Sheet4.cell(row=17, column=19).font = check_font_1

		Sheet4.cell(row=21, column=5).font = font_worksheet
		Sheet4.cell(row=21, column=6).font = font_worksheet
		Sheet4.cell(row=21, column=7).font = font_worksheet
		Sheet4.cell(row=21, column=8).font = font_worksheet
		Sheet4.cell(row=21, column=9).font = font_worksheet
		Sheet4.cell(row=21, column=10).font = font_worksheet
		Sheet4.cell(row=21, column=11).font = font_worksheet
		Sheet4.cell(row=21, column=12).font = font_worksheet
		Sheet4.cell(row=21, column=13).font = font_worksheet
		Sheet4.cell(row=21, column=14).font = font_worksheet
		Sheet4.cell(row=21, column=15).font = font_worksheet
		Sheet4.cell(row=21, column=16).font = font_worksheet

		Sheet4.cell(row=21, column=17).font = font_worksheet
		Sheet4.cell(row=21, column=18).font = TB_font
		Sheet4.cell(row=21, column=19).font = check_font_1

		Sheet4.cell(row=27, column=5).font = fprocentaj
		Sheet4.cell(row=27, column=6).font = fprocentaj
		Sheet4.cell(row=27, column=7).font = fprocentaj
		Sheet4.cell(row=27, column=8).font = fprocentaj
		Sheet4.cell(row=27, column=9).font = fprocentaj
		Sheet4.cell(row=27, column=10).font = fprocentaj
		Sheet4.cell(row=27, column=11).font = fprocentaj
		Sheet4.cell(row=27, column=12).font = fprocentaj
		Sheet4.cell(row=27, column=13).font = fprocentaj
		Sheet4.cell(row=27, column=14).font = fprocentaj
		Sheet4.cell(row=27, column=15).font = fprocentaj
		Sheet4.cell(row=27, column=16).font = fprocentaj
		Sheet4.cell(row=27, column=17).font = fprocentaj

		Sheet4.cell(row=31, column=5).font = font_worksheet
		Sheet4.cell(row=31, column=6).font = font_worksheet
		Sheet4.cell(row=31, column=7).font = font_worksheet
		Sheet4.cell(row=31, column=8).font = font_worksheet
		Sheet4.cell(row=31, column=9).font = font_worksheet
		Sheet4.cell(row=31, column=10).font = font_worksheet
		Sheet4.cell(row=31, column=11).font = font_worksheet
		Sheet4.cell(row=31, column=12).font = font_worksheet
		Sheet4.cell(row=31, column=13).font = font_worksheet
		Sheet4.cell(row=31, column=14).font = font_worksheet
		Sheet4.cell(row=31, column=15).font = font_worksheet
		Sheet4.cell(row=31, column=16).font = font_worksheet
		Sheet4.cell(row=31, column=17).font = font_worksheet

		Sheet4.cell(row=33, column=5).font = check_font_1
		Sheet4.cell(row=33, column=6).font = check_font_1
		Sheet4.cell(row=33, column=7).font = check_font_1
		Sheet4.cell(row=33, column=8).font = check_font_1
		Sheet4.cell(row=33, column=9).font = check_font_1
		Sheet4.cell(row=33, column=10).font = check_font_1
		Sheet4.cell(row=33, column=11).font = check_font_1
		Sheet4.cell(row=33, column=12).font = check_font_1
		Sheet4.cell(row=33, column=13).font = check_font_1
		Sheet4.cell(row=33, column=14).font = check_font_1
		Sheet4.cell(row=33, column=15).font = check_font_1
		Sheet4.cell(row=33, column=16).font = check_font_1
		Sheet4.cell(row=33, column=17).font = check_font_1

		# for row in Sheet4['R16:R21']:
		#     for cell in row:
		#         cell.fill = TB_font


		# number fornats

		Sheet4.cell(row=17, column=5).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=17, column=6).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=17, column=7).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=17, column=8).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=17, column=9).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=17, column=10).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=17, column=11).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=17, column=12).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=17, column=13).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=17, column=14).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=17, column=15).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=17, column=16).number_format = '#,##0_);(#,##0)'

		Sheet4.cell(row=17, column=17).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=17, column=18).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=17, column=19).number_format = '#,##0_);(#,##0)'

		Sheet4.cell(row=21, column=5).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=21, column=6).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=21, column=7).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=21, column=8).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=21, column=9).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=21, column=10).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=21, column=11).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=21, column=12).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=21, column=13).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=21, column=14).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=21, column=15).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=21, column=16).number_format = '#,##0_);(#,##0)'

		Sheet4.cell(row=21, column=17).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=21, column=18).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=21, column=19).number_format = '#,##0_);(#,##0)'

		Sheet4.cell(row=27, column=5).style = 'Percent'
		Sheet4.cell(row=27, column=6).style = 'Percent'
		Sheet4.cell(row=27, column=7).style = 'Percent'
		Sheet4.cell(row=27, column=8).style = 'Percent'
		Sheet4.cell(row=27, column=9).style = 'Percent'
		Sheet4.cell(row=27, column=10).style = 'Percent'
		Sheet4.cell(row=27, column=11).style = 'Percent'
		Sheet4.cell(row=27, column=12).style = 'Percent'
		Sheet4.cell(row=27, column=13).style = 'Percent'
		Sheet4.cell(row=27, column=14).style = 'Percent'
		Sheet4.cell(row=27, column=15).style = 'Percent'
		Sheet4.cell(row=27, column=16).style = 'Percent'
		Sheet4.cell(row=27, column=17).style = 'Percent'

		Sheet4.cell(row=27, column=5).number_format='0.00%'
		Sheet4.cell(row=27, column=6).number_format='0.00%'
		Sheet4.cell(row=27, column=7).number_format='0.00%'
		Sheet4.cell(row=27, column=8).number_format='0.00%'
		Sheet4.cell(row=27, column=9).number_format='0.00%'
		Sheet4.cell(row=27, column=10).number_format='0.00%'
		Sheet4.cell(row=27, column=11).number_format='0.00%'
		Sheet4.cell(row=27, column=12).number_format='0.00%'
		Sheet4.cell(row=27, column=13).number_format='0.00%'
		Sheet4.cell(row=27, column=14).number_format='0.00%'
		Sheet4.cell(row=27, column=15).number_format='0.00%'
		Sheet4.cell(row=27, column=16).number_format='0.00%'
		Sheet4.cell(row=27, column=17).number_format='0.00%'


		Sheet4.cell(row=31, column=5).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=31, column=6).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=31, column=7).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=31, column=8).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=31, column=9).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=31, column=10).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=31, column=11).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=31, column=12).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=31, column=13).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=31, column=14).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=31, column=15).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=31, column=16).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=31, column=17).number_format = '#,##0_);(#,##0)'

		Sheet4.cell(row=33, column=5).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=33, column=6).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=33, column=7).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=33, column=8).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=33, column=9).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=33, column=10).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=33, column=11).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=33, column=12).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=33, column=13).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=33, column=14).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=33, column=15).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=33, column=16).number_format = '#,##0_);(#,##0)'
		Sheet4.cell(row=33, column=17).number_format = '#,##0_);(#,##0)'

		for row in Sheet4['B16:S16']:
			for cell in row:
				cell.border = border_centered

		for row in Sheet4['B20:S20']:
			for cell in row:
				cell.border = border_centered

		for row in Sheet4['A24:S24']:
			for cell in row:
				cell.border = border_bottom

		for row in Sheet4['S25:S34']:
			for cell in row:
				cell.border = border_right1

		for row in Sheet4['A34:S34']:
			for cell in row:
				cell.border = border_bottom

		Sheet4['S34'].border = border_lowerright

		for row in Sheet4['E26:Q26']:
			for cell in row:
				cell.border = border_centered

		for row in Sheet4['E30:Q30']:
			for cell in row:
				cell.border = border_centered

		Sheet4.column_dimensions['A'].width = 10
		Sheet4.column_dimensions['B'].width = 20
		Sheet4.column_dimensions['C'].width = 11
		Sheet4.column_dimensions['D'].width = 12
		Sheet4.column_dimensions['E'].width = 14
		Sheet4.column_dimensions['F'].width = 12
		Sheet4.column_dimensions['G'].width = 10
		Sheet4.column_dimensions['H'].width = 9
		Sheet4.column_dimensions['I'].width = 9
		Sheet4.column_dimensions['J'].width = 9
		Sheet4.column_dimensions['K'].width = 9
		Sheet4.column_dimensions['L'].width = 9
		Sheet4.column_dimensions['M'].width = 9
		Sheet4.column_dimensions['N'].width = 9
		Sheet4.column_dimensions['O'].width = 12
		Sheet4.column_dimensions['P'].width = 15
		Sheet4.column_dimensions['Q'].width = 15
		Sheet4.column_dimensions['R'].width = 15
		Sheet4.column_dimensions['S'].width = 15
		Sheet4.column_dimensions['T'].width = 15
		Sheet4.column_dimensions['U'].width = 15
		Sheet4.column_dimensions['V'].width = 15
		Sheet4.column_dimensions['W'].width = 15

		Sheet5 = wb.create_sheet("V10.3 D112")

		Sheet5.sheet_view.showGridLines = False

		Sheet5.cell(row=1, column=1).value = "Client Name:"
		Sheet5.cell(row=2, column=1).value = "Period ended:"
		Sheet5.cell(row=3, column=1).value = "Audit area:"
		Sheet5.cell(row=3, column=2).value = "Payroll"
		Sheet5.cell(row=5, column=1).value = "D112 Reconciliation"

		try:
			Sheet5.cell(row=1, column=2).value = clientname1
			Sheet5.cell(row=2, column=2).value = periodEnd1
		except:
			None

		Sheet5.cell(row=7, column=2).value = "Work done"
		Sheet5.cell(row=8, column=2).value ="We have received the monthly trial balances from client."
		Sheet5.cell(row=9, column=2).value ="We have extracted from trial balances the social contributions."
		Sheet5.cell(row=10, column=2).value ="We have compared the amounts from trial balance with the amounts from 112 ""Social contributions statement"" for 31.12." +currYear1 +"."
		Sheet5.cell(row=11, column=2).value ="We have reconcilied the amount from 112 ""Social contribution statement"" for 31.12." + currYear1 +" with the payment and the payment centralizer."
		Sheet5.cell(row=13, column=2).value = "Conclusion"

		# 1st bucket
		Sheet5.cell(row=16, column=2).value = "December " + currYear1

		Sheet5.cell(row=17, column=2).value = "Description"
		Sheet5.cell(row=18, column=2).value = "# 431 - Social Security"
		Sheet5.cell(row=18, column=3).value = "As per Statement(112)"
		Sheet5.cell(row=19, column=2).value = "Employee contribution to social security"
		Sheet5.cell(row=19, column=3).value = "2.412"
		Sheet5.cell(row=20, column=2).value = "Employee contribution to social security (# 43150)"
		Sheet5.cell(row=20, column=3).value = "As per TB"
		Sheet5.cell(row=22, column=3).value = "Difference"

		# 2nd
		Sheet5.cell(row=23, column=2).value = "Description"
		Sheet5.cell(row=24, column=2).value = "# 431 - Social health contribution"
		Sheet5.cell(row=25, column=2).value = "Employee contribution to health insurance"
		Sheet5.cell(row=26, column=2).value = "Employee contribution to health insurance (# 43160)"
		Sheet5.cell(row=24, column=3).value = "As per Statement(112)"
		Sheet5.cell(row=25, column=3).value = "7.432"
		Sheet5.cell(row=26, column=3).value = "As per TB"
		Sheet5.cell(row=28, column=3).value = "Difference"

		# 3rd bucket
		Sheet5.cell(row=29, column=2).value = "Description"
		Sheet5.cell(row=30, column=2).value = "# 436 - Work insurance contribution "
		Sheet5.cell(row=31, column=2).value = "Employee contribution to health insurance"
		Sheet5.cell(row=32, column=2).value = "Company contribution to work insurance (# 43600)"
		Sheet5.cell(row=30, column=3).value = "As per Statement(112)"
		Sheet5.cell(row=31, column=3).value = "46.480"
		Sheet5.cell(row=32, column=3).value = "As per TB"
		Sheet5.cell(row=34, column=3).value = "Difference"

		# 4th bucket
		Sheet5.cell(row=35, column=2).value = "Description"
		Sheet5.cell(row=36, column=2).value = "# 444 Wage tax"
		Sheet5.cell(row=38, column=2).value = "# 444 Wage tax"
		Sheet5.cell(row=36, column=3).value = "As per Statement(112)"
		Sheet5.cell(row=37, column=3).value = "1.602"
		Sheet5.cell(row=38, column=3).value = "As per TB"
		Sheet5.cell(row=40, column=3).value = "Difference"
		Sheet5.cell(row=41, column=3).value = "Total diff"

		Sheet5.cell(row=43, column=2).value = "Reconciliation of taxes payment"
		Sheet5.cell(row=45, column=2).value = "As per D112 Statement"
		Sheet5.cell(row=46, column=2).value = "As per TB"
		Sheet5.cell(row=47, column=2).value = "Check"

		# XML
		try:
			tree = ET.parse(D112XML)
			root = tree.getroot()

			codoblig = []
			adatorat = []

			# for child in root:
			#     print(child.tag)
			for elem in root.findall('{mfp:anaf:dgti:declaratie_unica:declaratie:v6}angajator'):
				for item in elem.findall('{mfp:anaf:dgti:declaratie_unica:declaratie:v6}angajatorA'):
					tagcod = item.get('A_codOblig')
					codoblig.append(tagcod)
					tagsuma = item.get('A_datorat')
					adatorat.append(tagsuma)

			for i in range(0, len(codoblig)):
				if (codoblig[i] == '412'):
					Sheet5.cell(row=19, column=4).value = int(adatorat[i].format(19))
					Sheet5.cell(row=19, column=4).font = font_worksheet
					Sheet5.cell(row=19, column=4).number_format = '#,##0_);(#,##0)'

				if (codoblig[i] == '432'):
					Sheet5.cell(row=25, column=4).value = int(adatorat[i].format(19))
					Sheet5.cell(row=25, column=4).font = font_worksheet
					Sheet5.cell(row=25, column=4).number_format = '#,##0_);(#,##0)'

				if (codoblig[i] == '480'):
					Sheet5.cell(row=31, column=4).value = int(adatorat[i])
					Sheet5.cell(row=31, column=4).font = font_worksheet
					Sheet5.cell(row=31, column=4).number_format = '#,##0_);(#,##0)'

				if (codoblig[i] == '602'):
					Sheet5.cell(row=37, column=4).value = int(adatorat[i])
					Sheet5.cell(row=37, column=4).font = font_worksheet
					Sheet5.cell(row=37, column=4).number_format = '#,##0_);(#,##0)'
		except:
			pass

		Sheet5.cell(row=20, column=4).value = "=SUMIF('TB 12'!B:B,4315,'TB 12'!K:K)"
		Sheet5.cell(row=22, column=4).value = "=D19-D20"
		Sheet5.cell(row=26, column=4).value = "=SUMIF('TB 12'!B:B,4316,'TB 12'!K:K)"
		Sheet5.cell(row=28, column=4).value = "=D25-D26"
		Sheet5.cell(row=32, column=4).value = "=SUMIF('TB 12'!B:B,436,'TB 12'!K:K)"
		Sheet5.cell(row=34, column=4).value = "=D31-D32"
		Sheet5.cell(row=38, column=4).value = "=SUMIF('TB 12'!B:B,444,'TB 12'!K:K)"
		Sheet5.cell(row=40, column=4).value = "=D37-D38"
		Sheet5.cell(row=41, column=4).value = "=D40+D34+D28+D22"
		Sheet5.cell(row=46, column=4).value ="=D38+D32+D26+D20"
		Sheet5.cell(row=47, column=4).value = "=D45-D46"

		# FORMAT
		Sheet5.cell(row=1, column=1).font = f_info
		Sheet5.cell(row=2, column=1).font = f_info
		Sheet5.cell(row=3, column=1).font = f_info
		Sheet5.cell(row=3, column=2).font = font_worksheet
		Sheet5.cell(row=5, column=1).font = f_testname

		Sheet5.cell(row=1, column=2).font = font_worksheet
		Sheet5.cell(row=2, column=2).font = font_worksheet

		Sheet5.cell(row=7, column=2).font = ft1
		Sheet5.cell(row=8, column=2).font = font_worksheet
		Sheet5.cell(row=9, column=2).font = font_worksheet
		Sheet5.cell(row=10, column=2).font = font_worksheet
		Sheet5.cell(row=11, column=2).font = font_worksheet
		Sheet5.cell(row=13, column=2).font = ft1

		# 1st bucket
		Sheet5.cell(row=16, column=2).font = ft1

		Sheet5.cell(row=17, column=2).font = ft1
		Sheet5.cell(row=18, column=2).font = ft1
		Sheet5.cell(row=18, column=3).font = ft1
		Sheet5.cell(row=19, column=2).font = ft1
		Sheet5.cell(row=19, column=3).font = ft1
		Sheet5.cell(row=20, column=2).font = TB_font
		Sheet5.cell(row=20, column=3).font = TB_font
		Sheet5.cell(row=20, column=4).font = TB_font
		Sheet5.cell(row=22, column=3).font = check_font_1

		# 2nd
		Sheet5.cell(row=23, column=2).font = ft1
		Sheet5.cell(row=24, column=2).font = ft1
		Sheet5.cell(row=25, column=2).font = ft1
		Sheet5.cell(row=26, column=2).font = TB_font
		Sheet5.cell(row=24, column=3).font = ft1
		Sheet5.cell(row=25, column=3).font = ft1
		Sheet5.cell(row=26, column=3).font = TB_font
		Sheet5.cell(row=26, column=4).font = TB_font
		Sheet5.cell(row=28, column=3).font = check_font_1

		# 3rd bucket
		Sheet5.cell(row=29, column=2).font = ft1
		Sheet5.cell(row=30, column=2).font = ft1
		Sheet5.cell(row=31, column=2).font = ft1
		Sheet5.cell(row=32, column=2).font = TB_font
		Sheet5.cell(row=30, column=3).font = ft1
		Sheet5.cell(row=31, column=3).font = ft1
		Sheet5.cell(row=32, column=3).font = TB_font
		Sheet5.cell(row=32, column=4).font = TB_font
		Sheet5.cell(row=34, column=3).font = check_font_1

		# 4th bucket
		Sheet5.cell(row=35, column=2).font = ft1
		Sheet5.cell(row=36, column=2).font = ft1
		Sheet5.cell(row=38, column=2).font = TB_font
		Sheet5.cell(row=36, column=3).font = ft1
		Sheet5.cell(row=37, column=3).font = ft1
		Sheet5.cell(row=38, column=3).font = TB_font
		Sheet5.cell(row=38, column=4).font = TB_font
		Sheet5.cell(row=40, column=3).font = check_font_1
		Sheet5.cell(row=41, column=3).font = check_font

		Sheet5.cell(row=43, column=2).font = ft1
		Sheet5.cell(row=45, column=2).font = ft1
		Sheet5.cell(row=46, column=2).font = TB_font
		Sheet5.cell(row=46, column=4).font = TB_font
		Sheet5.cell(row=47, column=2).font = ft1
		Sheet5.cell(row=47, column=3).font = check_font_1

		Sheet5.cell(row=20, column=4).font = TB_font
		Sheet5.cell(row=22, column=4).font = check_font_1
		Sheet5.cell(row=26, column=4).font = TB_font
		Sheet5.cell(row=28, column=4).font = check_font_1
		Sheet5.cell(row=32, column=4).font = TB_font
		Sheet5.cell(row=34, column=4).font = check_font_1
		Sheet5.cell(row=38, column=4).font = TB_font
		Sheet5.cell(row=40, column=4).font = check_font_1
		Sheet5.cell(row=41, column=4).font = font_worksheet
		Sheet5.cell(row=47, column=2).font = check_font
		Sheet5.cell(row=47, column=4).font = check_font_1

		# for row in Sheet5['B20:D20']:
		#     for cell in row:
		#         cell.fill = TB_font
		#
		# for row in Sheet5['B26:D26']:
		#     for cell in row:
		#         cell.fill = TB_font
		#
		# for row in Sheet5['B32:D32']:
		#     for cell in row:
		#         cell.fill = TB_font
		#
		# for row in Sheet5['B38:D38']:
		#     for cell in row:
		#         cell.fill = TB_font

		# NUMBER FORMAT
		Sheet5.cell(row=19, column=4).number_format = '#,##0_);(#,##0)'
		Sheet5.cell(row=20, column=4).number_format = '#,##0_);(#,##0)'
		Sheet5.cell(row=22, column=4).number_format = '#,##0_);(#,##0)'
		Sheet5.cell(row=25, column=4).number_format = '#,##0_);(#,##0)'
		Sheet5.cell(row=26, column=4).number_format = '#,##0_);(#,##0)'
		Sheet5.cell(row=28, column=4).number_format = '#,##0_);(#,##0)'
		Sheet5.cell(row=31, column=4).number_format = '#,##0_);(#,##0)'
		Sheet5.cell(row=32, column=4).number_format = '#,##0_);(#,##0)'
		Sheet5.cell(row=34, column=4).number_format = '#,##0_);(#,##0)'
		Sheet5.cell(row=37, column=4).number_format = '#,##0_);(#,##0)'
		Sheet5.cell(row=38, column=4).number_format = '#,##0_);(#,##0)'
		Sheet5.cell(row=40, column=4).number_format = '#,##0_);(#,##0)'
		Sheet5.cell(row=41, column=4).number_format = '#,##0_);(#,##0)'
		Sheet5.cell(row=45, column=4).number_format = '#,##0_);(#,##0)'
		Sheet5.cell(row=46, column=4).number_format = '#,##0_);(#,##0)'
		Sheet5.cell(row=47, column=4).number_format = '#,##0_);(#,##0)'

		Sheet5.column_dimensions['A'].width = 10
		Sheet5.column_dimensions['B'].width = 39
		Sheet5.column_dimensions['C'].width = 30
		Sheet5.column_dimensions['D'].width = 20
		Sheet5.column_dimensions['E'].width = 14

		for row in Sheet5['B17:C17']:
			for cell in row:
				cell.border = border_centered

		for row in Sheet5['B23:C23']:
			for cell in row:
				cell.border = border_centered

		for row in Sheet5['B29:C29']:
			for cell in row:
				cell.border = border_centered

		for row in Sheet5['B44:C44']:
			for cell in row:
				cell.border = border_centered

		Sheet6 = wb.create_sheet("V10.4 Payroll Reconciliation")

		Sheet6.sheet_view.showGridLines = False

		Sheet6.cell(row=1, column=1).value = "Client Name:"
		Sheet6.cell(row=2, column=1).value = "Period ended:"
		Sheet6.cell(row=3, column=1).value = "Audit area:"
		Sheet6.cell(row=3, column=2).value = "Payroll"
		Sheet6.cell(row=5, column=1).value = "Payroll Sheets Reconciliation"

		try:
			Sheet6.cell(row=1, column=2).value = clientname1
			Sheet6.cell(row=2, column=2).value = periodEnd1
		except:
			None

		Sheet6.cell(row=7, column=2).value = "Work done"
		Sheet6.cell(row=8, column=2).value ="We have obtained the montly payroll expenses for current year."
		Sheet6.cell(row=9, column=2).value ="We have checked if the monthly payroll expense recorded by the Company is equal to the salary from the monthly payroll sheet."
		Sheet6.cell(row=11, column=2).value = "Conclusion"

		Sheet6.cell(row=14, column=3).value = "Month"

		Sheet6.cell(row=15, column=3).value = "January"
		Sheet6.cell(row=16, column=3).value = "February"
		Sheet6.cell(row=17, column=3).value = "March"
		Sheet6.cell(row=18, column=3).value = "April"
		Sheet6.cell(row=19, column=3).value = "May"
		Sheet6.cell(row=20, column=3).value = "June"
		Sheet6.cell(row=21, column=3).value = "July"
		Sheet6.cell(row=22, column=3).value = "August"
		Sheet6.cell(row=23, column=3).value = "September"
		Sheet6.cell(row=24, column=3).value = "October"
		Sheet6.cell(row=25, column=3).value = "November"
		Sheet6.cell(row=26, column=3).value = "December"
		Sheet6.cell(row=27, column=3).value = "Total"

		Sheet6.cell(row=14, column=4).value = "Monthly payroll expenses acc.#641"
		Sheet6.cell(row=14, column=5).value = "Monthly payroll from payroll statement"
		Sheet6.cell(row=14, column=6).value = "Check"

		Sheet6.cell(row=30, column=3).value = "Reconciliation of salary payment"
		Sheet6.cell(row=32, column=3).value = "Dec 20"
		Sheet6.cell(row=33, column=3).value = "As per bank statement"
		Sheet6.cell(row=34, column=3).value = "As per TB"
		Sheet6.cell(row=35, column=3).value = "Check"
		Sheet6.cell(row=34, column=5).value = "doar 421 sold"

		# FORMULE
		Sheet6.cell(row=15, column=4).value ="=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!D:D)"
		Sheet6.cell(row=16, column=4).value ="=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!E:E)"
		Sheet6.cell(row=17, column=4).value ="=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!F:F)"
		Sheet6.cell(row=18, column=4).value ="=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!G:G)"
		Sheet6.cell(row=19, column=4).value ="=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!H:H)"
		Sheet6.cell(row=20, column=4).value ="=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!I:I)"
		Sheet6.cell(row=21, column=4).value ="=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!J:J)"
		Sheet6.cell(row=22, column=4).value ="=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!K:K)"
		Sheet6.cell(row=23, column=4).value ="=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!L:L)"
		Sheet6.cell(row=24, column=4).value ="=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!M:M)"
		Sheet6.cell(row=25, column=4).value ="=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!N:N)"
		Sheet6.cell(row=26, column=4).value ="=SUMIF('Monthly P&L CY'!B:B,641,'Monthly P&L CY'!O:O)"
		Sheet6.cell(row=27, column=4).value = "=SUM(D15:D26)"

		Sheet6.cell(row=15, column=5).value ="=PBC!B9"
		Sheet6.cell(row=16, column=5).value ="=PBC!C9"
		Sheet6.cell(row=17, column=5).value ="=PBC!D9"
		Sheet6.cell(row=18, column=5).value ="=PBC!E9"
		Sheet6.cell(row=19, column=5).value ="=PBC!F9"
		Sheet6.cell(row=20, column=5).value ="=PBC!G9"
		Sheet6.cell(row=21, column=5).value ="=PBC!H9"
		Sheet6.cell(row=22, column=5).value ="=PBC!I9"
		Sheet6.cell(row=23, column=5).value ="=PBC!J9"
		Sheet6.cell(row=24, column=5).value ="=PBC!K9"
		Sheet6.cell(row=25, column=5).value ="=PBC!L9"
		Sheet6.cell(row=26, column=5).value ="=PBC!M9"
		Sheet6.cell(row=27, column=5).value = "=SUM(E15:E26)"


		Sheet6.cell(row=15, column=6).value = "=D15-SUM(E15:E15)"
		Sheet6.cell(row=16, column=6).value = "=D16-SUM(E16:E16)"
		Sheet6.cell(row=17, column=6).value = "=D17-SUM(E17:E17)"
		Sheet6.cell(row=18, column=6).value = "=D18-SUM(E18:E18)"
		Sheet6.cell(row=19, column=6).value = "=D19-SUM(E19:E19)"
		Sheet6.cell(row=20, column=6).value = "=D20-SUM(E20:E20)"
		Sheet6.cell(row=21, column=6).value = "=D21-SUM(E21:E21)"
		Sheet6.cell(row=22, column=6).value = "=D22-SUM(E22:E22)"
		Sheet6.cell(row=23, column=6).value = "=D23-SUM(E23:E23)"
		Sheet6.cell(row=24, column=6).value = "=D24-SUM(E24:E24)"
		Sheet6.cell(row=25, column=6).value = "=D25-SUM(E25:E25)"
		Sheet6.cell(row=26, column=6).value = "=D26-SUM(E26:E26)"
		Sheet6.cell(row=27, column=6).value = "=SUM(F15:F26)"

		Sheet6.cell(row=34, column=5).value ="=-SUMIF('TB 12'!A:A,421,'TB 12'!H:H)"
		Sheet6.cell(row=35, column=5).value ="=E33-E34"

		# FORMAT
		Sheet6.cell(row=1, column=1).font = f_info
		Sheet6.cell(row=2, column=1).font = f_info
		Sheet6.cell(row=3, column=1).font = f_info
		Sheet6.cell(row=3, column=2).font = font_worksheet
		Sheet6.cell(row=5, column=1).font = f_testname

		Sheet6.cell(row=1, column=2).font = font_worksheet
		Sheet6.cell(row=2, column=2).font = font_worksheet

		Sheet6.cell(row=7, column=2).font = ft1
		Sheet6.cell(row=8, column=2).font = font_worksheet
		Sheet6.cell(row=9, column=2).font = font_worksheet
		Sheet6.cell(row=11, column=2).font = ft1

		Sheet6.cell(row=14, column=3).font = ft1

		Sheet6.cell(row=15, column=3).font = ft1
		Sheet6.cell(row=16, column=3).font = ft1
		Sheet6.cell(row=17, column=3).font = ft1
		Sheet6.cell(row=18, column=3).font = ft1
		Sheet6.cell(row=19, column=3).font = ft1
		Sheet6.cell(row=20, column=3).font = ft1
		Sheet6.cell(row=21, column=3).font = ft1
		Sheet6.cell(row=22, column=3).font = ft1
		Sheet6.cell(row=23, column=3).font = ft1
		Sheet6.cell(row=24, column=3).font = ft1
		Sheet6.cell(row=25, column=3).font = ft1
		Sheet6.cell(row=26, column=3).font = ft1
		Sheet6.cell(row=27, column=3).font = ft1

		Sheet6.cell(row=14, column=4).font = ft1
		Sheet6.cell(row=14, column=5).font = ft1
		Sheet6.cell(row=14, column=6).font = check_font

		Sheet6.cell(row=30, column=3).font = ft1
		Sheet6.cell(row=32, column=3).font = ft1
		Sheet6.cell(row=33, column=3).font = ft1
		Sheet6.cell(row=34, column=3).font = TB_font
		Sheet6.cell(row=35, column=3).font = check_font
		Sheet6.cell(row=33, column=5).font = ft1
		Sheet6.cell(row=34, column=5).font = TB_font
		Sheet6.cell(row=35, column=5).font =check_font_1

		# FORMULE
		Sheet6.cell(row=15, column=4).font = font_worksheet
		Sheet6.cell(row=16, column=4).font = font_worksheet
		Sheet6.cell(row=17, column=4).font = font_worksheet
		Sheet6.cell(row=18, column=4).font = font_worksheet
		Sheet6.cell(row=19, column=4).font = font_worksheet
		Sheet6.cell(row=20, column=4).font = font_worksheet
		Sheet6.cell(row=21, column=4).font = font_worksheet
		Sheet6.cell(row=22, column=4).font = font_worksheet
		Sheet6.cell(row=23, column=4).font = font_worksheet
		Sheet6.cell(row=24, column=4).font = font_worksheet
		Sheet6.cell(row=25, column=4).font = font_worksheet
		Sheet6.cell(row=26, column=4).font = font_worksheet
		Sheet6.cell(row=27, column=4).font = font_worksheet

		Sheet6.cell(row=15, column=6).font = check_font
		Sheet6.cell(row=16, column=6).font = check_font
		Sheet6.cell(row=17, column=6).font = check_font
		Sheet6.cell(row=18, column=6).font = check_font
		Sheet6.cell(row=19, column=6).font = check_font
		Sheet6.cell(row=20, column=6).font = check_font
		Sheet6.cell(row=21, column=6).font = check_font
		Sheet6.cell(row=22, column=6).font = check_font
		Sheet6.cell(row=23, column=6).font = check_font
		Sheet6.cell(row=24, column=6).font = check_font
		Sheet6.cell(row=25, column=6).font = check_font
		Sheet6.cell(row=26, column=6).font = check_font
		Sheet6.cell(row=27, column=6).font = check_font

		# for row in Sheet6['C34:D34']:
		#     for cell in row:
		#         cell.fill = TB_font

		# number format
		Sheet6.cell(row=15, column=4).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=16, column=4).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=17, column=4).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=18, column=4).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=19, column=4).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=20, column=4).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=21, column=4).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=22, column=4).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=23, column=4).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=24, column=4).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=25, column=4).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=26, column=4).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=27, column=4).number_format = '#,##0_);(#,##0)'

		Sheet6.cell(row=15, column=5).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=16, column=5).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=17, column=5).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=18, column=5).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=19, column=5).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=20, column=5).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=21, column=5).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=22, column=5).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=23, column=5).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=24, column=5).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=25, column=5).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=26, column=5).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=27, column=5).number_format = '#,##0_);(#,##0)'

		Sheet6.cell(row=15, column=6).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=16, column=6).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=17, column=6).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=18, column=6).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=19, column=6).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=20, column=6).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=21, column=6).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=22, column=6).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=23, column=6).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=24, column=6).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=25, column=6).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=26, column=6).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=27, column=6).number_format = '#,##0_);(#,##0)'

		Sheet6.cell(row=34, column=5).number_format = '#,##0_);(#,##0)'
		Sheet6.cell(row=35, column=5).number_format = '#,##0_);(#,##0)'

		Sheet6['C14'].border = border_left

		for row in Sheet6['D14:E14']:
			for cell in row:
				cell.border = border_centered

		Sheet6['F14'].border = border_right

		for row in Sheet6['C15:C26']:
			for cell in row:
				cell.border = border_left1

		for row in Sheet6['F15:F26']:
			for cell in row:
				cell.border = border_right1

		Sheet6['C27'].border = border_left

		for row in Sheet6['D27:E27']:
			for cell in row:
				cell.border = border_centered

		Sheet6['F27'].border = border_right

		for row in Sheet6['C30:F30']:
			for cell in row:
				cell.border = border_centered

		Sheet6.column_dimensions['A'].width = 10
		Sheet6.column_dimensions['B'].width = 20
		Sheet6.column_dimensions['C'].width = 20
		Sheet6.column_dimensions['D'].width = 20
		Sheet6.column_dimensions['E'].width = 20
		Sheet6.column_dimensions['F'].width = 20
		Sheet6.column_dimensions['G'].width = 10

		Sheet7 = wb.create_sheet("V10.5 Meal tickets ")

		Sheet7.sheet_view.showGridLines = False

		Sheet7.cell(row=1, column=1).value = "Client Name:"
		Sheet7.cell(row=2, column=1).value = "Period ended:"
		Sheet7.cell(row=3, column=1).value = "Audit area:"
		Sheet7.cell(row=3, column=2).value = "Payroll"
		Sheet7.cell(row=5, column=1).value = "Meal Tickets Reasonableness"

		try:
			Sheet7.cell(row=1, column=2).value = clientname1
			Sheet7.cell(row=2, column=2).value = periodEnd1
		except:
			None

		Sheet7.cell(row=7, column=2).value = "Work done"
		Sheet7.cell(row=8, column=2).value ="We have obtained the number of monthy meal tickes provided to Company employees;"
		Sheet7.cell(row=9, column=2).value ="According to Romanian Legislation the Companies can spend between 9.57 and 20 RON/ ticket"
		Sheet7.cell(row=10, column=2).value ="The value of the tickets is established, internally at RON 15/ticket."
		Sheet7.cell(row=11, column=2).value ="We have recomputed the monthly M.T expense based on the info received;"
		Sheet7.cell(row=13, column=2).value = "Conclusion"

		Sheet7.cell(row=18, column=2).value = "Month"
		Sheet7.cell(row=19, column=2).value = "Avg no of employees"
		Sheet7.cell(row=20, column=2).value = "Value of meal tickets "
		Sheet7.cell(row=21, column=2).value = "Working days"
		Sheet7.cell(row=22, column=2).value = "Less 20 Avg CO days"
		Sheet7.cell(row=23, column=2).value = "Less avg legal holidays"
		Sheet7.cell(row=24, column=2).value = "Working days (after legal holidays and CO)"
		Sheet7.cell(row=25, column=2).value = "Recomputed"
		Sheet7.cell(row=26, column=2).value = "Check"

		Sheet7.cell(row=20, column=3).value ="=PBC!N5"
		Sheet7.cell(row=21, column=3).value =int(261)
		Sheet7.cell(row=22, column=3).value =int(20)
		Sheet7.cell(row=23, column=3).value =int(7)

		Sheet7.cell(row=19, column=4).value ="Please assess with client if all employees or only a part are entitled for Meal Tickets"
		Sheet7.cell(row=21, column=4).value ="If the audited period is not 2021 – please modify accordingly"
		Sheet7.cell(row=23, column=4).value ="If the audited period is not 2021 – please modify accordingly"

		Sheet7.cell(row=28, column=1).value = "As per Monthly PL"
		Sheet7.cell(row=29, column=1).value = "Period"
		Sheet7.cell(row=30, column=1).value = currYear1
		Sheet7.cell(row=31, column=1).value = prevYear1
		Sheet7.cell(row=29, column=2).value = "Account"
		Sheet7.cell(row=30, column=2).value = "642"
		Sheet7.cell(row=31, column=2).value = "642"
		Sheet7.cell(row=29, column=3).value = "Descriere"
		Sheet7.cell(row=30, column=3).value = "Cheltuieli cu avantajele în natura acordate salariatilor"
		Sheet7.cell(row=31, column=3).value = "Cheltuieli cu avantajele în natura acordate salariatilor"


		Sheet7.cell(row=29, column=4).value = "January"
		Sheet7.cell(row=29, column=5).value = "February"
		Sheet7.cell(row=29, column=6).value = "March"
		Sheet7.cell(row=29, column=7).value = "April"
		Sheet7.cell(row=29, column=8).value = "May"
		Sheet7.cell(row=29, column=9).value = "June"
		Sheet7.cell(row=29, column=10).value = "July"
		Sheet7.cell(row=29, column=11).value = "August"
		Sheet7.cell(row=29, column=12).value = "September"
		Sheet7.cell(row=29, column=13).value = "October"
		Sheet7.cell(row=29, column=14).value = "November"
		Sheet7.cell(row=29, column=15).value = "December"
		Sheet7.cell(row=29, column=16).value = "Total"
		Sheet7.cell(row=29, column=17).value = "As per TB"
		Sheet7.cell(row=29, column=18).value = "Check"
		Sheet7.cell(row=29, column=19).value = "Variation"
		Sheet7.cell(row=29, column=20).value = "Variation %"

		# formule
		Sheet7.cell(row=19, column=3).value = "='V10.1 Payroll Analysis'!U23"
		Sheet7.cell(row=24, column=3).value = "=C21-C22"
		Sheet7.cell(row=25, column=3).value = "=C19*C20*C24"
		Sheet7.cell(row=26, column=3).value = "=C25-P30"

		Sheet7.cell(row=30, column=4).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!D:D)"
		Sheet7.cell(row=30, column=5).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!E:E)"
		Sheet7.cell(row=30, column=6).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!F:F)"
		Sheet7.cell(row=30, column=7).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!G:G)"
		Sheet7.cell(row=30, column=8).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!H:H)"
		Sheet7.cell(row=30, column=9).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!I:I)"
		Sheet7.cell(row=30, column=10).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!J:J)"
		Sheet7.cell(row=30, column=11).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!K:K)"
		Sheet7.cell(row=30, column=12).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!L:L)"
		Sheet7.cell(row=30, column=13).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!M:M)"
		Sheet7.cell(row=30, column=14).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!N:N)"
		Sheet7.cell(row=30, column=15).value = "=SUMIF('Monthly P&L CY'!B:B,642,'Monthly P&L CY'!O:O)"
		Sheet7.cell(row=30, column=16).value = "=SUM(D30:O30)"
		Sheet7.cell(row=30, column=17).value = "=SUMIF('TB 12'!A:A,642,'TB 12'!H:H)"
		Sheet7.cell(row=30, column=18).value = "=P30-Q30"
		Sheet7.cell(row=30, column=19).value = "=P30-P31"
		Sheet7.cell(row=30, column=20).value = "=S30/P31"

		Sheet7.cell(row=31, column=4).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!D:D)"
		Sheet7.cell(row=31, column=5).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!E:E)"
		Sheet7.cell(row=31, column=6).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!F:F)"
		Sheet7.cell(row=31, column=7).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!G:G)"
		Sheet7.cell(row=31, column=8).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!H:H)"
		Sheet7.cell(row=31, column=9).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!I:I)"
		Sheet7.cell(row=31, column=10).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!J:J)"
		Sheet7.cell(row=31, column=11).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!K:K)"
		Sheet7.cell(row=31, column=12).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!L:L)"
		Sheet7.cell(row=31, column=13).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!M:M)"
		Sheet7.cell(row=31, column=14).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!N:N)"
		Sheet7.cell(row=31, column=15).value = "=SUMIF('Monthly P&L PY'!B:B,642,'Monthly P&L PY'!O:O)"
		Sheet7.cell(row=31, column=16).value = "=SUM(D31:O31)"

		# FORMAT
		Sheet7.cell(row=1, column=1).font = f_info
		Sheet7.cell(row=2, column=1).font = f_info
		Sheet7.cell(row=3, column=1).font = f_info
		Sheet7.cell(row=3, column=2).font = font_worksheet
		Sheet7.cell(row=5, column=1).font = f_testname

		Sheet7.cell(row=1, column=2).font = font_worksheet
		Sheet7.cell(row=2, column=2).font = font_worksheet

		Sheet7.cell(row=7, column=2).font = ft1
		Sheet7.cell(row=8, column=2).font =font_worksheet
		Sheet7.cell(row=9, column=2).font =font_worksheet
		Sheet7.cell(row=10, column=2).font =font_worksheet
		Sheet7.cell(row=11, column=2).font =font_worksheet
		Sheet7.cell(row=13, column=2).font = ft1

		Sheet7.cell(row=18, column=3).font = ft1
		Sheet7.cell(row=18, column=4).font = ft1
		Sheet7.cell(row=18, column=5).font = ft1
		Sheet7.cell(row=18, column=6).font = ft1
		Sheet7.cell(row=18, column=7).font = ft1
		Sheet7.cell(row=18, column=8).font = ft1
		Sheet7.cell(row=18, column=9).font = ft1
		Sheet7.cell(row=18, column=10).font = ft1
		Sheet7.cell(row=18, column=11).font = ft1
		Sheet7.cell(row=18, column=12).font = ft1
		Sheet7.cell(row=18, column=13).font = ft1
		Sheet7.cell(row=18, column=14).font = ft1
		Sheet7.cell(row=18, column=15).font = ft1

		Sheet7.cell(row=19, column=4).font = check_font_1
		Sheet7.cell(row=21, column=4).font = check_font_1
		Sheet7.cell(row=23, column=4).font = check_font_1

		Sheet7.cell(row=19, column=3).font = font_worksheet
		Sheet7.cell(row=20, column=3).font = font_worksheet
		Sheet7.cell(row=21, column=3).font = font_worksheet
		Sheet7.cell(row=22, column=3).font = font_worksheet
		Sheet7.cell(row=23, column=3).font = font_worksheet
		Sheet7.cell(row=24, column=3).font = font_worksheet
		Sheet7.cell(row=25, column=3).font = font_worksheet
		Sheet7.cell(row=26, column=3).font = check_font

		Sheet7.cell(row=18, column=2).font = ft1
		Sheet7.cell(row=19, column=2).font = ft1
		Sheet7.cell(row=20, column=2).font = ft1
		Sheet7.cell(row=21, column=2).font = ft1
		Sheet7.cell(row=22, column=2).font = ft1
		Sheet7.cell(row=23, column=2).font = ft1
		Sheet7.cell(row=24, column=2).font = ft1
		Sheet7.cell(row=25, column=2).font = ft1
		Sheet7.cell(row=26, column=2).font = check_font

		Sheet7.cell(row=28, column=1).font = ft1
		Sheet7.cell(row=29, column=1).font = ft1
		Sheet7.cell(row=30, column=1).font = font_worksheet
		Sheet7.cell(row=31, column=1).font = font_worksheet
		Sheet7.cell(row=29, column=2).font = ft1
		Sheet7.cell(row=30, column=2).font = font_worksheet
		Sheet7.cell(row=31, column=2).font = font_worksheet
		Sheet7.cell(row=29, column=3).font = ft1
		Sheet7.cell(row=30, column=3).font = ft1
		Sheet7.cell(row=31, column=3).font = ft1

		Sheet7.cell(row=29, column=4).font = ft1
		Sheet7.cell(row=29, column=5).font = ft1
		Sheet7.cell(row=29, column=6).font = ft1
		Sheet7.cell(row=29, column=7).font = ft1
		Sheet7.cell(row=29, column=8).font = ft1
		Sheet7.cell(row=29, column=9).font = ft1
		Sheet7.cell(row=29, column=10).font = ft1
		Sheet7.cell(row=29, column=11).font = ft1
		Sheet7.cell(row=29, column=12).font = ft1
		Sheet7.cell(row=29, column=13).font = ft1
		Sheet7.cell(row=29, column=14).font = ft1
		Sheet7.cell(row=29, column=15).font = ft1
		Sheet7.cell(row=29, column=16).font = ft1
		Sheet7.cell(row=29, column=17).font = TB_font
		Sheet7.cell(row=29, column=18).font = check_font
		Sheet7.cell(row=29, column=19).font = ft1
		Sheet7.cell(row=29, column=20).font = ft1

		Sheet7.cell(row=30, column=4).font = font_worksheet
		Sheet7.cell(row=30, column=5).font = font_worksheet
		Sheet7.cell(row=30, column=6).font = font_worksheet
		Sheet7.cell(row=30, column=7).font = font_worksheet
		Sheet7.cell(row=30, column=8).font = font_worksheet
		Sheet7.cell(row=30, column=9).font = font_worksheet
		Sheet7.cell(row=30, column=10).font = font_worksheet
		Sheet7.cell(row=30, column=11).font = font_worksheet
		Sheet7.cell(row=30, column=12).font = font_worksheet
		Sheet7.cell(row=30, column=13).font = font_worksheet
		Sheet7.cell(row=30, column=14).font = font_worksheet
		Sheet7.cell(row=30, column=15).font = font_worksheet
		Sheet7.cell(row=30, column=16).font = font_worksheet
		Sheet7.cell(row=30, column=17).font = TB_font
		Sheet7.cell(row=30, column=18).font = check_font_1
		Sheet7.cell(row=30, column=19).font = font_worksheet

		Sheet7.cell(row=31, column=4).font = font_worksheet
		Sheet7.cell(row=31, column=5).font = font_worksheet
		Sheet7.cell(row=31, column=6).font = font_worksheet
		Sheet7.cell(row=31, column=7).font = font_worksheet
		Sheet7.cell(row=31, column=8).font = font_worksheet
		Sheet7.cell(row=31, column=9).font = font_worksheet
		Sheet7.cell(row=31, column=10).font = font_worksheet
		Sheet7.cell(row=31, column=11).font = font_worksheet
		Sheet7.cell(row=31, column=12).font = font_worksheet
		Sheet7.cell(row=31, column=13).font = font_worksheet
		Sheet7.cell(row=31, column=14).font = font_worksheet
		Sheet7.cell(row=31, column=15).font = font_worksheet
		Sheet7.cell(row=31, column=16).font = font_worksheet
		Sheet7.cell(row=31, column=17).font = font_worksheet
		Sheet7.cell(row=31, column=18).font = check_font_1
		Sheet7.cell(row=31, column=19).font = font_worksheet

		# for row in Sheet7['Q29:Q32']:
		#     for cell in row:
		#         cell.fill = TB_font

		# NUMBER FORMATS
		Sheet7.cell(row=20, column=3).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=21, column=3).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=22, column=3).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=23, column=3).number_format = '#,##0_);(#,##0)'

		Sheet7.cell(row=19, column=3).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=20, column=3).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=21, column=3).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=22, column=3).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=23, column=3).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=24, column=3).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=25, column=3).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=26, column=3).number_format = '#,##0_);(#,##0)'

		Sheet7.cell(row=30, column=4).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=30, column=5).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=30, column=6).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=30, column=7).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=30, column=8).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=30, column=9).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=30, column=10).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=30, column=11).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=30, column=12).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=30, column=13).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=30, column=14).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=30, column=15).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=30, column=16).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=30, column=17).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=30, column=18).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=30, column=19).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=30, column=20).style = 'Percent'

		Sheet7.cell(row=31, column=4).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=31, column=5).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=31, column=6).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=31, column=7).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=31, column=8).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=31, column=9).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=31, column=10).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=31, column=11).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=31, column=12).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=31, column=13).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=31, column=14).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=31, column=15).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=31, column=16).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=31, column=17).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=31, column=18).number_format = '#,##0_);(#,##0)'
		Sheet7.cell(row=31, column=19).number_format = '#,##0_);(#,##0)'

		#TABLE
		Sheet7['B18'].border = border_centered
		Sheet7['C18'].border = border_centered

		for row in Sheet7['A29:T29']:
			for cell in row:
				cell.border = border_centered

		for row in Sheet7['A32:T32']:
			for cell in row:
				cell.border = border_bottom

		Sheet7.column_dimensions['A'].width = 20
		Sheet7.column_dimensions['B'].width = 20
		Sheet7.column_dimensions['C'].width = 20
		Sheet7.column_dimensions['D'].width = 10
		Sheet7.column_dimensions['E'].width = 10
		Sheet7.column_dimensions['F'].width = 10
		Sheet7.column_dimensions['G'].width = 15
		Sheet7.column_dimensions['H'].width = 15
		Sheet7.column_dimensions['I'].width = 15
		Sheet7.column_dimensions['J'].width = 15
		Sheet7.column_dimensions['G'].width = 15
		Sheet7.column_dimensions['H'].width = 15
		Sheet7.column_dimensions['I'].width = 15
		Sheet7.column_dimensions['J'].width = 15
		Sheet7.column_dimensions['K'].width = 15
		Sheet7.column_dimensions['L'].width = 15
		Sheet7.column_dimensions['M'].width = 15
		Sheet7.column_dimensions['N'].width = 15
		Sheet7.column_dimensions['O'].width = 15
		Sheet7.column_dimensions['P'].width = 15
		Sheet7.column_dimensions['R'].width = 15

		Sheet8 = wb.create_sheet("V10.6 Untaken Holiday")

		Sheet8.sheet_view.showGridLines = False

		Sheet8.cell(row=1, column=1).value = "Client Name:"
		Sheet8.cell(row=2, column=1).value = "Period ended:"
		Sheet8.cell(row=3, column=1).value = "Audit area:"
		Sheet8.cell(row=3, column=2).value = "Payroll"
		Sheet8.cell(row=5, column=1).value = "Untaken Holiday"

		try:
			Sheet8.cell(row=1, column=2).value = clientname1
			Sheet8.cell(row=2, column=2).value = periodEnd1
		except:
			None

		Sheet8.cell(row=7, column=1).value ="Objective"
		Sheet8.cell(row=7, column=2).value ="To ensure that holidays not taken in "+currYear1+" had been correctly included in the payroll accrual, "
		Sheet8.cell(row=8, column=2).value ="or that transactions were consistently complete, valid and compliant with applicable policies, procedures and regulations;"

		Sheet8.cell(row=10, column=1).value ="Work done"
		Sheet8.cell(row=10, column=2).value ="We obtained the detail of untaken holidays from the client;"
		Sheet8.cell(row=11, column=2).value ="We recomputed the provision amount, based on the information received;"
		Sheet8.cell(row=12, column=2).value ="We reconciled the recalculated amounts with the ones in the Trial Balance. "
		Sheet8.cell(row=13, column=2).value ="Please see below:"

		Sheet8.cell(row=15, column=2).value ="a. Provision recomputation for untaken holidays:"
		Sheet8.cell(row=17, column=2).value ="Gross salary/period "
		Sheet8.cell(row=18, column=2).value ="Average no. of employees"
		Sheet8.cell(row=19, column=2).value ="Working days/period"
		Sheet8.cell(row=19, column=5).value = "If the audited period is not 2021 – please modify accordingly"
		Sheet8.cell(row=20, column=2).value ="Average salary per employee/day"
		Sheet8.cell(row=21, column=2).value ="Total untaken holidays"
		Sheet8.cell(row=22, column=2).value ="Untaken holiday provision"
		Sheet8.cell(row=23, column=2).value ="Contribution"
		Sheet8.cell(row=24, column=2).value ="Total provision"

		Sheet8.cell(row=27, column=2).value ="b. Agree totals to Trial Balance"
		Sheet8.cell(row=29, column=2).value ="Nexia recomputation"
		Sheet8.cell(row=30, column=2).value ="Trial Balance:"
		Sheet8.cell(row=31, column=5).value = "Please asses if the Company booked an untaken provision, otherwise, please asses if any ADJE is required"
		Sheet8.cell(row=31, column=2).value ="Difference:"

		Sheet8.cell(row=33, column=1).value ="Conclusion:"

		#FORMULE
		Sheet8.cell(row=17, column=3).value = "=SUMIF('TB 12'!A:A,641,'TB 12'!H:H)"
		Sheet8.cell(row=18, column=3).value = "='V10.1 Payroll Analysis'!U23"
		Sheet8.cell(row=19, column=3).value = int(261)
		Sheet8.cell(row=20, column=3).value = "=C17/C18/C19"
		Sheet8.cell(row=21, column=3).value = "=PBC!B8"
		Sheet8.cell(row=22, column=3).value = "=C20*C21"
		Sheet8.cell(row=23, column=3).value = "=C22*2.25%"
		Sheet8.cell(row=24, column=3).value = "=SUM(C22:C23)"

		Sheet8.cell(row=29, column=3).value = "=C24"
		Sheet8.cell(row=30, column=3).value = "=SUMIF('TB 12'!C:C,1518,'TB 12'!H:H)"
		Sheet8.cell(row=31, column=3).value = "=C30-C29"

		# FORMAT
		Sheet8.cell(row=1, column=1).font = f_info
		Sheet8.cell(row=2, column=1).font = f_info
		Sheet8.cell(row=3, column=1).font = f_info
		Sheet8.cell(row=3, column=2).font = font_worksheet
		Sheet8.cell(row=5, column=1).font = f_testname

		Sheet8.cell(row=1, column=2).font = font_worksheet
		Sheet8.cell(row=2, column=2).font = font_worksheet

		Sheet8.cell(row=7, column=1).font = ft1
		Sheet8.cell(row=7, column=2).font = font_worksheet
		Sheet8.cell(row=8, column=2).font = font_worksheet

		Sheet8.cell(row=10, column=1).font = ft1
		Sheet8.cell(row=10, column=2).font = font_worksheet
		Sheet8.cell(row=11, column=2).font = font_worksheet
		Sheet8.cell(row=12, column=2).font = font_worksheet
		Sheet8.cell(row=13, column=2).font = font_worksheet

		Sheet8.cell(row=15, column=2).font= f_info
		Sheet8.cell(row=17, column=2).font= font_worksheet
		Sheet8.cell(row=17, column=5).font= font_worksheet
		Sheet8.cell(row=18, column=2).font= font_worksheet
		Sheet8.cell(row=19, column=2).font= font_worksheet
		Sheet8.cell(row=20, column=2).font= font_worksheet
		Sheet8.cell(row=21, column=2).font= font_worksheet
		Sheet8.cell(row=22, column=2).font= font_worksheet
		Sheet8.cell(row=23, column=2).font= font_worksheet
		Sheet8.cell(row=24, column=2).font= ft1

		Sheet8.cell(row=27, column=2).font= f_info
		Sheet8.cell(row=29, column=2).font= font_worksheet
		Sheet8.cell(row=30, column=2).font= TB_font
		Sheet8.cell(row=31, column=2).font= check_font

		Sheet8.cell(row=17, column=3).font= font_worksheet
		Sheet8.cell(row=18, column=3).font= font_worksheet
		Sheet8.cell(row=19, column=3).font= font_worksheet
		Sheet8.cell(row=19, column=5).font = check_font_1
		Sheet8.cell(row=20, column=3).font= font_worksheet
		Sheet8.cell(row=21, column=3).font= font_worksheet
		Sheet8.cell(row=22, column=3).font= font_worksheet
		Sheet8.cell(row=23, column=3).font= font_worksheet
		Sheet8.cell(row=24, column=3).font= ft1

		Sheet8.cell(row=29, column=3).font= font_worksheet
		Sheet8.cell(row=30, column=3).font= TB_font
		Sheet8.cell(row=31, column=3).font= check_font_1
		Sheet8.cell(row=31, column=5).font= check_font_1

		Sheet8.cell(row=33, column=1).font= ft1

		# FORMAT
		Sheet8.cell(row=17, column=3).number_format = '#,##0_);(#,##0)'
		Sheet8.cell(row=18, column=3).number_format = '#,##0_);(#,##0)'
		Sheet8.cell(row=19, column=3).number_format = '#,##0_);(#,##0)'
		Sheet8.cell(row=20, column=3).number_format = '#,##0_);(#,##0)'
		Sheet8.cell(row=21, column=3).number_format = '#,##0_);(#,##0)'
		Sheet8.cell(row=22, column=3).number_format = '#,##0_);(#,##0)'
		Sheet8.cell(row=23, column=3).number_format = '#,##0_);(#,##0)'
		Sheet8.cell(row=24, column=3).number_format = '#,##0_);(#,##0)'

		Sheet8.cell(row=29, column=3).number_format = '#,##0_);(#,##0)'
		Sheet8.cell(row=30, column=3).number_format = '#,##0_);(#,##0)'
		Sheet8.cell(row=31, column=3).number_format = '#,##0_);(#,##0)'

		#table
		#TABLE1
		Sheet8['B15'].border = border_left
		Sheet8['C15'].border = border_centered
		Sheet8['D15'].border = border_right

		Sheet8['B16'].border = border_upperleft
		Sheet8['D16'].border = border_upperright

		for row in Sheet8['B16:B23']:
			for cell in row:
				cell.border = border_left1

		Sheet8['C24'].border = border_bottom
		Sheet8['B24'].border = border_lowerleft
		Sheet8['D24'].border = border_lowerright

		for row in Sheet8['D16:D23']:
			for cell in row:
				cell.border = border_right1

		#TEBLE2
		Sheet8['B27'].border = border_left
		Sheet8['C27'].border = border_centered
		Sheet8['D27'].border = border_right

		Sheet8['B28'].border = border_upperleft
		Sheet8['D28'].border = border_upperright

		for row in Sheet8['B29:B30']:
			for cell in row:
				cell.border = border_left1

		Sheet8['C31'].border = border_bottom

		Sheet8['B31'].border = border_lowerleft
		Sheet8['D31'].border = border_lowerright

		for row in Sheet8['D29:D30']:
			for cell in row:
				cell.border = border_right1

		Sheet8.column_dimensions['A'].width = 10
		Sheet8.column_dimensions['B'].width = 55
		Sheet8.column_dimensions['C'].width = 10
		Sheet8.column_dimensions['E'].width = 30

		Sheet9 = wb.create_sheet("TB 12")

		Sheet9.cell(row=1, column=1).value = "Synt(3)"
		Sheet9.cell(row=1, column=1).font = cap_tabel
		Sheet9.cell(row=1, column=1).fill = cap_tabel_color_GT
		Sheet9.cell(row=1, column=2).value = "Synt(4)"
		Sheet9.cell(row=1, column=2).font = cap_tabel
		Sheet9.cell(row=1, column=2).fill = cap_tabel_color_GT
		Sheet9.cell(row=1, column=3).value = "Account"
		Sheet9.cell(row=1, column=3).font = cap_tabel
		Sheet9.cell(row=1, column=3).fill = cap_tabel_color_GT
		Sheet9.cell(row=1, column=4).value = "Description"
		Sheet9.cell(row=1, column=4).font = cap_tabel
		Sheet9.cell(row=1, column=4).fill = cap_tabel_color_GT
		Sheet9.cell(row=1, column=5).value = "Opening Balance"
		Sheet9.cell(row=1, column=5).font = cap_tabel
		Sheet9.cell(row=1, column=5).fill = cap_tabel_color_GT
		Sheet9.cell(row=1, column=6).value = "Debit Movement"
		Sheet9.cell(row=1, column=6).font = cap_tabel
		Sheet9.cell(row=1, column=6).fill = cap_tabel_color_GT
		Sheet9.cell(row=1, column=7).value = "Credit Movement"
		Sheet9.cell(row=1, column=7).font = cap_tabel
		Sheet9.cell(row=1, column=7).fill = cap_tabel_color_GT
		Sheet9.cell(row=1, column=8).value = "Closing Balance"
		Sheet9.cell(row=1, column=8).font = cap_tabel
		Sheet9.cell(row=1, column=8).fill = cap_tabel_color_GT
		Sheet9.cell(row=1, column=9).value = "Synth(2)"
		Sheet9.cell(row=1, column=9).font = cap_tabel
		Sheet9.cell(row=1, column=9).fill = cap_tabel_color_GT
		Sheet9.cell(row=1, column=10).value = "RDC"
		Sheet9.cell(row=1, column=10).font = cap_tabel
		Sheet9.cell(row=1, column=10).fill = cap_tabel_color_GT
		Sheet9.cell(row=1, column=11).value = "RCC"
		Sheet9.cell(row=1, column=11).font = cap_tabel
		Sheet9.cell(row=1, column=11).fill = cap_tabel_color_GT

		tb = openpyxl.load_workbook(openTB, data_only=True)  # deschidem TB-ul
		tb1 = tb.active

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "Account":
					row_tb = cell.row
					column_acc = cell.column
					lun = len(tb1[cell.column])
		try:
			account = [b.value for b in tb1[column_acc][row_tb:lun]]
		except:
			flash("Please insert the correct header for Account in Trial Balance file")
			return render_template("index.html")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "Description":
					row_tb = cell.row
					column_descr = cell.column
					lun = len(tb1[cell.column])
		try:
			descr = [b.value for b in tb1[column_descr][row_tb:lun]]
		except:
			flash("Please insert the correct header for Description in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "The Description table header was not correctly created.")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "OB":
					row_tb = cell.row
					coloana_opTB_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			opTB = [b.value for b in tb1[coloana_opTB_tb][row_tb:lun]]
		except:
			flash("Please insert the correct header for OB in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "The OB table header was not correctly created.")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "RDC":
					row_tb = cell.row
					coloana_RDC_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			RDC = [b.value for b in tb1[coloana_RDC_tb][row_tb:lun]]
		except:
			flash("Please insert the correct header for RDC in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "The RDC table header was not correctly created.")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "RCC":
					row_tb = cell.row
					coloana_RCC_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			RCC = [b.value for b in tb1[coloana_RCC_tb][row_tb:lun]]
		except:
			flash("Please insert the correct header for RCC in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "The RCC table header was not correctly created.")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "CM":
					row_tb = cell.row
					coloana_CM_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			cr_mv = [b.value for b in tb1[coloana_CM_tb][row_tb:lun]]
		except:
			flash("Please insert the correct header for CM in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "The RCC table header was not correctly created.")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "DM":
					row_tb = cell.row
					coloana_db_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			dr_mv = [b.value for b in tb1[coloana_db_tb][row_tb:lun]]
		except:
			flash("Please insert the correct header for DM in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "The DM table header was not correctly created.")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "CB":
					row_tb = cell.row
					coloana_clTB_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			clTB = [b.value for b in tb1[coloana_clTB_tb][row_tb:lun]]
		except:
			flash("Please insert the correct header for CB in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "The CB table header was not correctly created.")

		for i in range(0, len(account)):
			Sheet9.cell(row=2 + i, column=3).value = account[i]
			Sheet9.cell(row=2 + i, column=2).value = str(account[i])[:4]  # in Excel =left("celula", 4)

		for i in range(0, len(account)):
			Sheet9.cell(row=2 + i, column=1).value = str(account[i])[:3]  # in Excel =left("celula", 3)

		for i in range(0, len(descr)):
			Sheet9.cell(row=2 + i, column=4).value = descr[i]

		for i in range(0, len(opTB)):
			Sheet9.cell(row=2 + i, column=5).value = opTB[i]
			Sheet9.cell(row=2 + i, column=5).font = font_worksheet
			Sheet9.cell(row=2 + i, column=5).number_format = '#,##0_);(#,##0)'

		for i in range(0, len(cr_mv)):
			Sheet9.cell(row=2 + i, column=6).value = cr_mv[i]
			Sheet9.cell(row=2 + i, column=6).font = font_worksheet
			Sheet9.cell(row=2 + i, column=6).number_format = '#,##0_);(#,##0)'

		for i in range(0, len(dr_mv)):
			Sheet9.cell(row=2 + i, column=7).value = dr_mv[i]
			Sheet9.cell(row=2 + i, column=7).font = font_worksheet
			Sheet9.cell(row=2 + i, column=7).number_format = '#,##0_);(#,##0)'

		for i in range(0, len(clTB)):
			Sheet9.cell(row=2 + i, column=8).value = clTB[i]
			Sheet9.cell(row=2 + i, column=8).font = font_worksheet
			Sheet9.cell(row=2 + i, column=8).number_format = '#,##0_);(#,##0)'

		for i in range(0, len(account)):
			Sheet9.cell(row=2 + i, column=9).value = str(account[i])[:2]

		for i in range(0, len(RDC)):
			Sheet9.cell(row=2 + i, column=10).value = RDC[i]
			Sheet9.cell(row=2 + i, column=10).font = font_worksheet
			Sheet9.cell(row=2 + i, column=10).number_format = '#,##0_);(#,##0)'

		for i in range(0, len(RCC)):
			Sheet9.cell(row=2 + i, column=11).value = RCC[i]
			Sheet9.cell(row=2 + i, column=11).font = font_worksheet
			Sheet9.cell(row=2 + i, column=11).number_format = '#,##0_);(#,##0)'

		monPlPBC = openpyxl.load_workbook(openMPL, data_only=True)

		Sheet9.column_dimensions['A'].width = 15
		Sheet9.column_dimensions['B'].width = 15
		Sheet9.column_dimensions['C'].width = 15
		Sheet9.column_dimensions['D'].width = 15
		Sheet9.column_dimensions['E'].width = 15
		Sheet9.column_dimensions['F'].width = 15
		Sheet9.column_dimensions['G'].width = 15
		Sheet9.column_dimensions['H'].width = 15
		Sheet9.column_dimensions['I'].width = 15
		Sheet9.column_dimensions['J'].width = 15
		Sheet9.column_dimensions['K'].width = 15

		Sheet10 = wb.create_sheet('Monthly P&L CY')

		Sheet10.cell(row=1, column=2).value = "Synth"
		Sheet10.cell(row=1, column=3).value = "Account"
		Sheet10.cell(row=1, column=4).value = "January"
		Sheet10.cell(row=1, column=5).value = "February"
		Sheet10.cell(row=1, column=6).value = "March"
		Sheet10.cell(row=1, column=7).value = "April"
		Sheet10.cell(row=1, column=8).value = "May"
		Sheet10.cell(row=1, column=9).value = "June"
		Sheet10.cell(row=1, column=10).value = "July"
		Sheet10.cell(row=1, column=11).value = "August"
		Sheet10.cell(row=1, column=12).value = "September"
		Sheet10.cell(row=1, column=13).value = "October"
		Sheet10.cell(row=1, column=14).value = "November"
		Sheet10.cell(row=1, column=15).value = "December"

		monPlPBC_1 = monPlPBC["MP CY"]
		# lun=monPlPBC_1.max_row
		for row in monPlPBC_1.iter_rows():
			for cell in row:
				if cell.value == "Synth":
					row_monplCY = cell.row
					column_synth = cell.column
					lun = len(monPlPBC_1[cell.column])
		try:
			synth_CY = [b.value for b in monPlPBC_1[column_synth][row_monplCY:lun]]
		except:
			flash("Please insert the correct header for Synth in Monthly PL Current Year")
			return render_template("index.html")

		for row in monPlPBC_1.iter_rows():
			for cell in row:
				if cell.value == "Account":
					row_monplCY = cell.row
					column_acc_CY = cell.column
					lun = len(monPlPBC_1[cell.column])
		try:
			acc_CY = [b.value for b in monPlPBC_1[column_acc_CY][row_monplCY:lun]]
		except:
			flash("Please insert the correct header for Account in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "Account field from header for current year was not created correctly.")

		for row in monPlPBC_1.iter_rows():
			for cell in row:
				if cell.value == "January":
					row_monplCY = cell.row
					column_jan = cell.column
					lun = len(monPlPBC_1[cell.column])
		try:
			January_CY = [b.value for b in monPlPBC_1[column_jan][row_monplCY:lun]]
		except:
			flash("Please insert the correct header for January in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "January field from header for current year was not created correctly.")

		for row in monPlPBC_1.iter_rows():
			for cell in row:
				if cell.value == "February":
					row_monplCY = cell.row
					column_feb = cell.column
					lun = len(monPlPBC_1[cell.column])
		try:
			February_CY = [b.value for b in monPlPBC_1[column_feb][row_monplCY:lun]]
		except:
			flash("Please insert the correct header for February in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "February field from header for current year was not created correctly.")

		for row in monPlPBC_1.iter_rows():
			for cell in row:
				if cell.value == "March":
					row_monplCY = cell.row
					column_mar = cell.column
					lun = len(monPlPBC_1[cell.column])
		try:
			March_CY = [b.value for b in monPlPBC_1[column_mar][row_monplCY:lun]]
		except:
			flash("Please insert the correct header for March in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "March field from header for current year was not created correctly.")

		for row in monPlPBC_1.iter_rows():
			for cell in row:
				if cell.value == "April":
					row_monplCY = cell.row
					column_apr = cell.column
					lun = len(monPlPBC_1[cell.column])
		try:
			April_CY = [b.value for b in monPlPBC_1[column_apr][row_monplCY:lun]]
		except:
			flash("Please insert the correct header for April in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "April field from header for current year was not created correctly.")

		for row in monPlPBC_1.iter_rows():
			for cell in row:
				if cell.value == "May":
					row_monplCY = cell.row
					column_may = cell.column
					lun = len(monPlPBC_1[cell.column])
		try:
			May_CY = [b.value for b in monPlPBC_1[column_may][row_monplCY:lun]]
		except:
			flash("Please insert the correct header for May in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "May field from header for current year was not created correctly.")

		for row in monPlPBC_1.iter_rows():
			for cell in row:
				if cell.value == "June":
					row_monplCY = cell.row
					column_jun = cell.column
					lun = len(monPlPBC_1[cell.column])
		try:
			June_CY = [b.value for b in monPlPBC_1[column_jun][row_monplCY:lun]]
		except:
			flash("Please insert the correct header for June in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "June field from header for current year was not created correctly.")

		for row in monPlPBC_1.iter_rows():
			for cell in row:
				if cell.value == "July":
					row_monplCY = cell.row
					column_jul = cell.column
					lun = len(monPlPBC_1[cell.column])
		try:
			July_CY = [b.value for b in monPlPBC_1[column_jul][row_monplCY:lun]]
		except:
			flash("Please insert the correct header for July in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "July field from header for current year was not created correctly.")

		for row in monPlPBC_1.iter_rows():
			for cell in row:
				if cell.value == "August":
					row_monplCY = cell.row
					column_aug = cell.column
					lun = len(monPlPBC_1[cell.column])
		try:
			August_CY = [b.value for b in monPlPBC_1[column_aug][row_monplCY:lun]]
		except:
			flash("Please insert the correct header for August in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "August field from header for current year was not created correctly.")

		for row in monPlPBC_1.iter_rows():
			for cell in row:
				if cell.value == "September":
					row_monplCY = cell.row
					column_sep = cell.column
					lun = len(monPlPBC_1[cell.column])
		try:
			September_CY = [b.value for b in monPlPBC_1[column_sep][row_monplCY:lun]]
		except:
			flash("Please insert the correct header for September in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "September field from header for current year was not created correctly.")

		for row in monPlPBC_1.iter_rows():
			for cell in row:
				if cell.value == "October":
					row_monplCY = cell.row
					column_oct = cell.column
					lun = len(monPlPBC_1[cell.column])
		try:
			October_CY = [b.value for b in monPlPBC_1[column_oct][row_monplCY:lun]]
		except:
			flash("Please insert the correct header for October in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "October field from header for current year was not created correctly.")

		for row in monPlPBC_1.iter_rows():
			for cell in row:
				if cell.value == "November":
					row_monplCY = cell.row
					column_nov = cell.column
					lun = len(monPlPBC_1[cell.column])
		try:
			November_CY = [b.value for b in monPlPBC_1[column_nov][row_monplCY:lun]]
		except:
			flash("Please insert the correct header for November in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "November field from header for current year was not created correctly.")

		for row in monPlPBC_1.iter_rows():
			for cell in row:
				if cell.value == "December":
					row_monplCY = cell.row
					column_dec = cell.column
					lun = len(monPlPBC_1[cell.column])
		try:
			December_CY = [b.value for b in monPlPBC_1[column_dec][row_monplCY:lun]]
		except:
			flash("Please insert the correct header for December in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "December field from header for current year was not created correctly.")

		for i in range(0, len(synth_CY)):
			Sheet10.cell(row=2+i, column=2).value = synth_CY[i]

		for i in range(0, len(acc_CY)):
			Sheet10.cell(row=2+i, column=3).value = acc_CY[i]

		for i in range(0, len(January_CY)):
			Sheet10.cell(row=2+i, column=4).value = January_CY[i]

		for i in range(0, len(February_CY)):
			Sheet10.cell(row=2+i, column=5).value = February_CY[i]

		for i in range(0, len(March_CY)):
			Sheet10.cell(row=2+i, column=6).value = March_CY[i]

		for i in range(0, len(April_CY)):
			Sheet10.cell(row=2+i, column=7).value = April_CY[i]

		for i in range(0, len(May_CY)):
			Sheet10.cell(row=2+i, column=8).value = May_CY[i]

		for i in range(0, len(June_CY)):
			Sheet10.cell(row=2+i, column=9).value = June_CY[i]

		for i in range(0, len(July_CY)):
			Sheet10.cell(row=2+i, column=10).value = July_CY[i]

		for i in range(0, len(August_CY)):
			Sheet10.cell(row=2+i, column=11).value = August_CY[i]

		for i in range(0, len(September_CY)):
			Sheet10.cell(row=2+i, column=12).value = September_CY[i]

		for i in range(0, len(October_CY)):
			Sheet10.cell(row=2+i, column=13).value = October_CY[i]

		for i in range(0, len(November_CY)):
			Sheet10.cell(row=2+i, column=14).value = November_CY[i]

		for i in range(0, len(December_CY)):
			Sheet10.cell(row=2+i, column=15).value = December_CY[i]

		Sheet10.column_dimensions['A'].width = 15
		Sheet10.column_dimensions['B'].width = 15
		Sheet10.column_dimensions['C'].width = 15
		Sheet10.column_dimensions['D'].width = 15
		Sheet10.column_dimensions['E'].width = 15
		Sheet10.column_dimensions['F'].width = 15
		Sheet10.column_dimensions['G'].width = 15
		Sheet10.column_dimensions['H'].width = 15
		Sheet10.column_dimensions['I'].width = 15
		Sheet10.column_dimensions['J'].width = 10
		Sheet10.column_dimensions['K'].width = 15
		Sheet10.column_dimensions['L'].width = 15
		Sheet10.column_dimensions['M'].width = 15
		Sheet10.column_dimensions['N'].width = 15
		Sheet10.column_dimensions['O'].width = 15

		for row in Sheet10['B1:O1']:
			for cell in row:
				cell.font = cap_tabel

		for row in Sheet10['B1:O1']:
			for cell in row:
				cell.fill = cap_tabel_color_GT

		Sheet11 = wb.create_sheet('Monthly P&L PY')
		monPlPBC_2 = monPlPBC["MP PY"]

		Sheet11.cell(row=1, column=2).value = "Synth"
		Sheet11.cell(row=1, column=3).value = "Account"
		Sheet11.cell(row=1, column=4).value = "January"
		Sheet11.cell(row=1, column=5).value = "February"
		Sheet11.cell(row=1, column=6).value = "March"
		Sheet11.cell(row=1, column=7).value = "April"
		Sheet11.cell(row=1, column=8).value = "May"
		Sheet11.cell(row=1, column=9).value = "June"
		Sheet11.cell(row=1, column=10).value = "July"
		Sheet11.cell(row=1, column=11).value = "August"
		Sheet11.cell(row=1, column=12).value = "September"
		Sheet11.cell(row=1, column=13).value = "October"
		Sheet11.cell(row=1, column=14).value = "November"
		Sheet11.cell(row=1, column=15).value = "December"

		for row in monPlPBC_2.iter_rows():
			for cell in row:
				if cell.value == "Synth":
					row_monplPY = cell.row
					column_synth = cell.column
					lun = len(monPlPBC_2[cell.column])
		try:
			synth_PY = [b.value for b in monPlPBC_2[column_synth][row_monplPY:lun]]
		except:
			flash("Please insert the correct header for Synt in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "Synthetic field from header for previous year was not created correctly.")

		for row in monPlPBC_2.iter_rows():
			for cell in row:
				if cell.value == "Account":
					row_monplPY = cell.row
					column_acc_PY = cell.column
					lun = len(monPlPBC_2[cell.column])
		try:
			acc_PY = [b.value for b in monPlPBC_2[column_acc_PY][row_monplPY:lun]]
		except:
			flash("Please insert the correct header for Account in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "Account field from header for previous year was not created correctly.")

		for row in monPlPBC_2.iter_rows():
			for cell in row:
				if cell.value == "January":
					row_monplPY = cell.row
					column_jan = cell.column
					lun = len(monPlPBC_2[cell.column])
		try:
			January_PY = [b.value for b in monPlPBC_2[column_jan][row_monplPY:lun]]
		except:
			flash("Please insert the correct header for January in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "January field from header for previous year was not created correctly.")

		for row in monPlPBC_2.iter_rows():
			for cell in row:
				if cell.value == "February":
					row_monplPY = cell.row
					column_feb = cell.column
					lun = len(monPlPBC_2[cell.column])
		try:
			February_PY = [b.value for b in monPlPBC_2[column_feb][row_monplPY:lun]]
		except:
			flash("Please insert the correct header for February in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "February field from header for previous year was not created correctly.")

		for row in monPlPBC_2.iter_rows():
			for cell in row:
				if cell.value == "March":
					row_monplPY = cell.row
					column_mar = cell.column
					lun = len(monPlPBC_2[cell.column])
		try:
			March_PY = [b.value for b in monPlPBC_2[column_mar][row_monplPY:lun]]
		except:
			flash("Please insert the correct header for March in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "March field from header for previous year was not created correctly.")

		for row in monPlPBC_2.iter_rows():
			for cell in row:
				if cell.value == "April":
					row_monplPY = cell.row
					column_apr = cell.column
					lun = len(monPlPBC_2[cell.column])
		try:
			April_PY = [b.value for b in monPlPBC_2[column_apr][row_monplPY:lun]]
		except:
			flash("Please insert the correct header for April in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "April field from header for previous year was not created correctly.")

		for row in monPlPBC_2.iter_rows():
			for cell in row:
				if cell.value == "May":
					row_monplPY = cell.row
					column_may = cell.column
					lun = len(monPlPBC_2[cell.column])
		try:
			May_PY = [b.value for b in monPlPBC_2[column_may][row_monplPY:lun]]
		except:
			flash("Please insert the correct header for May in Monthly PL Current Year")
			return render_template("index.html")

		for row in monPlPBC_2.iter_rows():
			for cell in row:
				if cell.value == "June":
					row_monplPY = cell.row
					column_jun = cell.column
					lun = len(monPlPBC_2[cell.column])
		try:
			June_PY = [b.value for b in monPlPBC_2[column_jun][row_monplPY:lun]]
		except:
			flash("Please insert the correct header for June in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "June field from header for previous year was not created correctly.")

		for row in monPlPBC_2.iter_rows():
			for cell in row:
				if cell.value == "July":
					row_monplPY = cell.row
					column_jul = cell.column
					lun = len(monPlPBC_2[cell.column])
		try:
			July_PY = [b.value for b in monPlPBC_2[column_jul][row_monplPY:lun]]
		except:
			flash("Please insert the correct header for July in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "July field from header for previous year was not created correctly.")

		for row in monPlPBC_2.iter_rows():
			for cell in row:
				if cell.value == "August":
					row_monplPY = cell.row
					column_aug = cell.column
					lun = len(monPlPBC_2[cell.column])
		try:
			August_PY = [b.value for b in monPlPBC_2[column_aug][row_monplPY:lun]]
		except:
			flash("Please insert the correct header for August in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "August field from header for previous year was not created correctly.")

		for row in monPlPBC_2.iter_rows():
			for cell in row:
				if cell.value == "September":
					row_monplPY = cell.row
					column_sep = cell.column
					lun = len(monPlPBC_2[cell.column])
		try:
			September_PY = [b.value for b in monPlPBC_2[column_sep][row_monplPY:lun]]
		except:
			flash("Please insert the correct header for September in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "September field from header for previous year was not created correctly.")

		for row in monPlPBC_2.iter_rows():
			for cell in row:
				if cell.value == "October":
					row_monplPY = cell.row
					column_oct = cell.column
					lun = len(monPlPBC_2[cell.column])
		try:
			October_PY = [b.value for b in monPlPBC_2[column_oct][row_monplPY:lun]]
		except:
			flash("Please insert the correct header for October in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "October field from header for previous year was not created correctly.")

		for row in monPlPBC_2.iter_rows():
			for cell in row:
				if cell.value == "November":
					row_monplPY = cell.row
					column_nov = cell.column
					lun = len(monPlPBC_2[cell.column])
		try:
			November_PY = [b.value for b in monPlPBC_2[column_nov][row_monplPY:lun]]
		except:
			flash("Please insert the correct header for November in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "November field from header for previous year was not created correctly.")

		for row in monPlPBC_2.iter_rows():
			for cell in row:
				if cell.value == "December":
					row_monplPY = cell.row
					column_dec = cell.column
					lun = len(monPlPBC_2[cell.column])
		try:
			December_PY = [b.value for b in monPlPBC_2[column_dec][row_monplPY:lun]]
		except:
			flash("Please insert the correct header for December in Monthly PL Current Year")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "December field from header for previous year was not created correctly.")

		for i in range(0, len(synth_PY)):
			Sheet11.cell(row=2+i, column=2).value = synth_PY[i]

		for i in range(0, len(acc_PY)):
			Sheet11.cell(row=2+i, column=3).value = acc_PY[i]

		for i in range(0, len(January_PY)):
			Sheet11.cell(row=2+i, column=4).value = January_PY[i]

		for i in range(0, len(February_PY)):
			Sheet11.cell(row=2+i, column=5).value = February_PY[i]

		for i in range(0, len(March_PY)):
			Sheet11.cell(row=2+i, column=6).value = March_PY[i]

		for i in range(0, len(April_PY)):
			Sheet11.cell(row=2+i, column=7).value = April_PY[i]

		for i in range(0, len(May_PY)):
			Sheet11.cell(row=2+i, column=8).value = May_PY[i]

		for i in range(0, len(June_PY)):
			Sheet11.cell(row=2+i, column=9).value = June_PY[i]

		for i in range(0, len(July_PY)):
			Sheet11.cell(row=2+i, column=10).value = July_PY[i]

		for i in range(0, len(August_PY)):
			Sheet11.cell(row=2+i, column=11).value = August_PY[i]

		for i in range(0, len(September_PY)):
			Sheet11.cell(row=2+i, column=12).value = September_PY[i]

		for i in range(0, len(October_PY)):
			Sheet11.cell(row=2+i, column=13).value = October_PY[i]

		for i in range(0, len(November_PY)):
			Sheet11.cell(row=2+i, column=14).value = November_PY[i]

		for i in range(0, len(December_PY)):
			Sheet11.cell(row=2+i, column=15).value = December_PY[i]

		Sheet11.column_dimensions['A'].width = 15
		Sheet11.column_dimensions['B'].width = 15
		Sheet11.column_dimensions['C'].width = 15
		Sheet11.column_dimensions['D'].width = 15
		Sheet11.column_dimensions['E'].width = 15
		Sheet11.column_dimensions['F'].width = 15
		Sheet11.column_dimensions['G'].width = 15
		Sheet11.column_dimensions['H'].width = 15
		Sheet11.column_dimensions['I'].width = 15
		Sheet11.column_dimensions['J'].width = 10
		Sheet11.column_dimensions['K'].width = 15
		Sheet11.column_dimensions['L'].width = 15
		Sheet11.column_dimensions['M'].width = 15
		Sheet11.column_dimensions['N'].width = 15
		Sheet11.column_dimensions['O'].width = 15

		for row in Sheet11['B1:O1']:
			for cell in row:
				cell.font = cap_tabel

		for row in Sheet11['B1:O1']:
			for cell in row:
				cell.fill = cap_tabel_color_GT

		Sheet12 = wb.create_sheet("PBC")

		Sheet12.cell(row=2, column=1).value = "Month"
		Sheet12.cell(row=2, column=2).value = "January"
		Sheet12.cell(row=2, column=3).value = "February"
		Sheet12.cell(row=2, column=4).value = "March"
		Sheet12.cell(row=2, column=5).value = "April"
		Sheet12.cell(row=2, column=6).value = "May"
		Sheet12.cell(row=2, column=7).value = "June"
		Sheet12.cell(row=2, column=8).value = "July"
		Sheet12.cell(row=2, column=9).value = "August"
		Sheet12.cell(row=2, column=10).value = "September"
		Sheet12.cell(row=2, column=11).value = "October"
		Sheet12.cell(row=2, column=12).value = "November"
		Sheet12.cell(row=2, column=13).value = "December"
		Sheet12.cell(row=2, column=14).value = "Average"

		Sheet12.cell(row=3, column=1).value = "No of employees PY"
		Sheet12.cell(row=4, column=1).value = "No of employees CY"
		Sheet12.cell(row=5, column=1).value = "Value of meal tickets"
		Sheet12.cell(row=6, column=1).value = "Number of meal tickets"
		Sheet12.cell(row=7, column=1).value = "Bonus"
		Sheet12.cell(row=8, column=1).value = "Total no of untaken holidays"
		Sheet12.cell(row=9, column=1).value = "Monthly Payroll Statement"

		for row in pbc_1.iter_rows():
			for cell in row:
				if cell.value == "January":
					row_pbc = cell.row
					column_janpbc = cell.column
					lun = len(pbc_1[cell.column])
		try:
			January_PBC = [b.value for b in pbc_1[column_janpbc][row_pbc:lun]]
		except:
			flash("Please insert the correct header for January in PBC file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "January field from PBC template was not created.")

		for row in pbc_1.iter_rows():
			for cell in row:
				if cell.value == "February":
					row_pbc = cell.row
					column_febpbc = cell.column
					lun = len(pbc_1[cell.column])
		try:
			February_PBC = [b.value for b in pbc_1[column_febpbc][row_pbc:lun]]
		except:
			flash("Please insert the correct header for February in PBC file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "February field from PBC template was not created.")

		for row in pbc_1.iter_rows():
			for cell in row:
				if cell.value == "March":
					row_pbc = cell.row
					column_marpbc = cell.column
					lun = len(pbc_1[cell.column])
		try:
			March_PBC = [b.value for b in pbc_1[column_marpbc][row_pbc:lun]]
		except:
			flash("Please insert the correct header for March in PBC file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "March field from PBC template was not created.")

		for row in pbc_1.iter_rows():
			for cell in row:
				if cell.value == "April":
					row_pbc = cell.row
					column_aprpbc = cell.column
					lun = len(pbc_1[cell.column])
		try:
			April_PBC = [b.value for b in pbc_1[column_aprpbc][row_pbc:lun]]
		except:
			flash("Please insert the correct header for April in PBC file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "April field from PBC template was not created.")

		for row in pbc_1.iter_rows():
			for cell in row:
				if cell.value == "May":
					row_pbc = cell.row
					column_maypbc = cell.column
					lun = len(monPlPBC_2[cell.column])
		try:
			May_PBC = [b.value for b in pbc_1[column_maypbc][row_pbc:lun]]
		except:
			flash("Please insert the correct header for May in PBC file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "May field from PBC template was not created.")

		for row in pbc_1.iter_rows():
			for cell in row:
				if cell.value == "June":
					row_pbc = cell.row
					column_junpbc = cell.column
					lun = len(pbc_1[cell.column])
		try:
			June_PBC = [b.value for b in pbc_1[column_junpbc][row_pbc:lun]]
		except:
			flash("Please insert the correct header for June in PBC file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "June field from PBC template was not created.")

		for row in pbc_1.iter_rows():
			for cell in row:
				if cell.value == "July":
					row_pbc = cell.row
					column_julpbc = cell.column
					lun = len(pbc_1[cell.column])
		try:
			July_PBC = [b.value for b in pbc_1[column_julpbc][row_pbc:lun]]
		except:
			flash("Please insert the correct header for July in PBC file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "July field from PBC template was not created.")

		for row in pbc_1.iter_rows():
			for cell in row:
				if cell.value == "August":
					row_pbc = cell.row
					column_augpbc = cell.column
					lun = len(pbc_1[cell.column])
		try:
			August_PBC = [b.value for b in pbc_1[column_augpbc][row_pbc:lun]]
		except:
			flash("Please insert the correct header for August in PBC file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "August field from PBC template was not created.")

		for row in pbc_1.iter_rows():
			for cell in row:
				if cell.value == "September":
					row_pbc = cell.row
					column_seppbc = cell.column
					lun = len(pbc_1[cell.column])
		try:
			September_PBC = [b.value for b in pbc_1[column_seppbc][row_pbc:lun]]
		except:
			flash("Please insert the correct header for September in PBC file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "September field from PBC template was not created.")

		for row in pbc_1.iter_rows():
			for cell in row:
				if cell.value == "October":
					row_pbc = cell.row
					column_octpbc = cell.column
					lun = len(pbc_1[cell.column])
		try:
			October_PBC = [b.value for b in pbc_1[column_octpbc][row_pbc:lun]]
		except:
			flash("Please insert the correct header for October in PBC file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "October field from PBC template was not created.")

		for row in pbc_1.iter_rows():
			for cell in row:
				if cell.value == "November":
					row_pbc = cell.row
					column_novpbc = cell.column
					lun = len(pbc_1[cell.column])
		try:
			November_PBC = [b.value for b in pbc_1[column_novpbc][row_pbc:lun]]
		except:
			flash("Please insert the correct header for November in PBC file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "November field from PBC template was not created.")

		for row in pbc_1.iter_rows():
			for cell in row:
				if cell.value == "December":
					row_pbc = cell.row
					column_decpbc = cell.column
					lun = len(pbc_1[cell.column])
		try:
			December_PBC = [b.value for b in pbc_1[column_decpbc][row_pbc:lun]]
		except:
			flash("Please insert the correct header for December in PBC file")
			return render_template("index.html")
			# messagebox.showerror("ERROR!", "December field from PBC template was not created.")

		for i in range(0, len(January_PBC)):
			Sheet12.cell(row=3+i, column=2).value = January_PBC[i]
			Sheet12.cell(row=3 + i, column=3).value = February_PBC[i]
			Sheet12.cell(row=3 + i, column=4).value = March_PBC[i]
			Sheet12.cell(row=3 + i, column=5).value = April_PBC[i]
			Sheet12.cell(row=3 + i, column=6).value = May_PBC[i]
			Sheet12.cell(row=3 + i, column=7).value = June_PBC[i]
			Sheet12.cell(row=3 + i, column=8).value = July_PBC[i]
			Sheet12.cell(row=3 + i, column=9).value = August_PBC[i]
			Sheet12.cell(row=3 + i, column=10).value = September_PBC[i]
			Sheet12.cell(row=3 + i, column=11).value = October_PBC[i]
			Sheet12.cell(row=3 + i, column=12).value = November_PBC[i]
			Sheet12.cell(row=3 + i, column=13).value = December_PBC[i]

		Sheet12.cell(row=3, column=14).value ="=AVERAGE(B3:M3)"
		Sheet12.cell(row=4, column=14).value ="=AVERAGE(B4:M4)"
		Sheet12.cell(row=5, column=14).value ="=AVERAGE(B5:M5)"
		Sheet12.cell(row=6, column=14).value ="=AVERAGE(B6:M6)"

		for row in Sheet12['A2:N2']:
			for cell in row:
				cell.font = cap_tabel

		for row in Sheet12['A2:N2']:
			for cell in row:
				cell.fill = cap_tabel_color_GT

		Sheet13 = wb.create_sheet('V10.2 Payroll Analysis Graphic')

		Sheet13.sheet_state = 'hidden'
		Sheet13.protection.sheet = True

		Sheet13.cell(row=3, column=2).value ="='V10.1 Payroll Analysis'!H22"
		Sheet13.cell(row=4, column=2).value ="='V10.1 Payroll Analysis'!I22"
		Sheet13.cell(row=5, column=2).value ="='V10.1 Payroll Analysis'!J22"
		Sheet13.cell(row=6, column=2).value ="='V10.1 Payroll Analysis'!K22"
		Sheet13.cell(row=7, column=2).value ="='V10.1 Payroll Analysis'!L22"
		Sheet13.cell(row=8, column=2).value ="='V10.1 Payroll Analysis'!M22"
		Sheet13.cell(row=9, column=2).value ="='V10.1 Payroll Analysis'!N22"
		Sheet13.cell(row=10, column=2).value ="='V10.1 Payroll Analysis'!O22"
		Sheet13.cell(row=11, column=2).value ="='V10.1 Payroll Analysis'!P22"
		Sheet13.cell(row=12, column=2).value ="='V10.1 Payroll Analysis'!Q22"
		Sheet13.cell(row=13, column=2).value ="='V10.1 Payroll Analysis'!R22"
		Sheet13.cell(row=14, column=2).value ="='V10.1 Payroll Analysis'!S22"
		Sheet13.cell(row=15, column=2).value ="='V10.1 Payroll Analysis'!T22"

		Sheet13.cell(row=3, column=3).value = "='V10.1 Payroll Analysis'!H26"
		Sheet13.cell(row=4, column=3).value = "='V10.1 Payroll Analysis'!I26"
		Sheet13.cell(row=5, column=3).value = "='V10.1 Payroll Analysis'!J26"
		Sheet13.cell(row=6, column=3).value = "='V10.1 Payroll Analysis'!K26"
		Sheet13.cell(row=7, column=3).value = "='V10.1 Payroll Analysis'!L26"
		Sheet13.cell(row=8, column=3).value = "='V10.1 Payroll Analysis'!M26"
		Sheet13.cell(row=9, column=3).value = "='V10.1 Payroll Analysis'!N26"
		Sheet13.cell(row=10, column=3).value = "='V10.1 Payroll Analysis'!O26"
		Sheet13.cell(row=11, column=3).value = "='V10.1 Payroll Analysis'!P26"
		Sheet13.cell(row=12, column=3).value = "='V10.1 Payroll Analysis'!Q26"
		Sheet13.cell(row=13, column=3).value = "='V10.1 Payroll Analysis'!R26"
		Sheet13.cell(row=14, column=3).value = "='V10.1 Payroll Analysis'!S26"
		Sheet13.cell(row=15, column=3).value = "='V10.1 Payroll Analysis'!T26"

		Sheet13.cell(row=3, column=4).value = "='V10.1 Payroll Analysis'!H27"
		Sheet13.cell(row=4, column=4).value = "='V10.1 Payroll Analysis'!I27"
		Sheet13.cell(row=5, column=4).value = "='V10.1 Payroll Analysis'!J27"
		Sheet13.cell(row=6, column=4).value = "='V10.1 Payroll Analysis'!K27"
		Sheet13.cell(row=7, column=4).value = "='V10.1 Payroll Analysis'!L27"
		Sheet13.cell(row=8, column=4).value = "='V10.1 Payroll Analysis'!M27"
		Sheet13.cell(row=9, column=4).value = "='V10.1 Payroll Analysis'!N27"
		Sheet13.cell(row=10, column=4).value = "='V10.1 Payroll Analysis'!O27"
		Sheet13.cell(row=11, column=4).value = "='V10.1 Payroll Analysis'!P27"
		Sheet13.cell(row=12, column=4).value = "='V10.1 Payroll Analysis'!Q27"
		Sheet13.cell(row=13, column=4).value = "='V10.1 Payroll Analysis'!R27"
		Sheet13.cell(row=14, column=4).value = "='V10.1 Payroll Analysis'!S27"
		Sheet13.cell(row=15, column=4).value = "='V10.1 Payroll Analysis'!T27"

		# Data for plotting
		values = Reference(Sheet13,
						   min_col=3,  # I
						   max_col=4,  # T
						   min_row=3,
						   max_row=15)

		chart = LineChart()
		chart.add_data(values, titles_from_data=True)
		dates = Reference(Sheet13, min_col=2, max_col=2, min_row=4, max_row=15)
		chart.set_categories(dates)

		chart.title = "AVERAGE GROSS SALARY " + currYear1 + " vs  " + prevYear1
		chart.x_axis.title = ""
		chart.y_axis.title = ""
		Sheet3.add_chart(chart, "K3")


		folderpath = "home/auditappnexia/output/Payroll"
		file_pathFS = os.path.join(folderpath, "V10 Payroll Analysis " + clientname1 + ".xlsx")
		# messagebox.showinfo("Success!", "Succesful task.")


		wb.save(file_pathFS)

	# return clientname1
		return send_from_directory(folderpath, "V10 Payroll Analysis " + clientname1 + ".xlsx")
	return render_template("Payroll.html")

@app.route('/Equity/GTRWhSrzitah41yuSieMuQ6KVpn/Instructions', methods=['GET'])
def downloadEquity():
		filepath = "/home/auditappnexia/output/Equity"
 
		return send_from_directory(filepath,"Instructions - Equity.docx", as_attachment=True)


@app.route('/Equity/GTRWhSrzitah41yuSieMuQ6KVpn')
def equity():
	return render_template('Equity.html')

@app.route('/Equity/GTRWhSrzitah41yuSieMuQ6KVpn', methods=['POST', 'GET'])
def equity_process():
	clientname1 = request.form['client']
	yearEnd = datetime.datetime.strptime(request.form['yearEnd'],'%Y-%m-%d')
	yearEnd1 = datetime.datetime.strptime(
		request.form['yearEnd'],
		'%Y-%m-%d')
	preparedBy1 = request.form['preparedBy']
	# datePrepared1 = datetime.datetime.strptime(
	# 	request.form['preparedDate'],
	# 	'%Y-%m-%d')
	refference1 = request.form['reff']

	if request.method == 'POST':
		file_location_TB = request.files["TB"]




		"Formating Codes--------------------------------------------------------------------"
		italic=Font(name='Tahoma',size=8,italic=True)
		doubleborder = Border(bottom=Side(style='double'))
		solidborder = Border(bottom=Side(style='thin'))
		rightborder = Border(right=Side(style='thin'))
		leftborder = Border(left=Side(style='thin'))
		rightdouble = Border (right=Side(style='thin'), bottom=Side(style='double'))
		rightmedium = Border (right=Side(style='thin'), bottom=Side(style='medium'))
		font1 = Font(name='Tahoma', size=8)
		ft2 = Font(name='Tahoma', size=8, bold=True)
		fontRedBold = Font(name='Tahoma', size=8, bold=True, color= 'FF0000')
		fontRedNormal=Font(name='Tahoma',size=8,color='FF0000')
		fontKPMG = Font (name='KPMG Logo', size=8)
		workingsblue = Font(color='2F75B5', bold=True)
		headers= Font(bold=True, italic=True) 
		# headersblue= PatternFill(start_color='44546A',
		#                   end_color='44546A',
		#                   fill_type='solid')
		headersblue= PatternFill(start_color='4B0082',
						end_color='4B0082',
						fill_type='solid')

		headersorange= PatternFill(start_color='EE2A1C',
						end_color='EE2A1C',
						fill_type='solid')

		headersblack = PatternFill(start_color='00AEAC',
						end_color='00AEAC',
						fill_type='solid')
		grifill=PatternFill(start_color='c4d79b',end_color='c4d79b',fill_type='solid')

		equity = Font(name='Tahoma', size=8, bold=True, italic=True)

		tah=Font(name='Tahoma',size=8)
		white = Font(name="Tahoma", size=8, bold=True,color="FFFFFF")

		"Import files--------------------------------------------------------------------------"



		try:
			TB = openpyxl.load_workbook(file_location_TB,data_only=True)
			TB1 = TB.active
		except:
			messagebox.showerror("Error", "File: Trial Balance. Can't find or open the TB file!")
		#   sys.exit()


		'Iterate from FAR ----------------------------------------------------'

		'Iterate from TB ----------------------------------------------------'


		for row in TB1.iter_rows():
			for cell in row :
				if cell.value=="Account" :
					tbacc=cell.column
					tbrow=cell.row

		for row in TB1.iter_rows():
			for cell in row :
				if cell.value=="Description" :
					
					tbdesc=cell.column
						
		for row in TB1.iter_rows():
			for cell in row :

				if cell.value=="DM" :
					
					tbdrm=cell.column

		for row in TB1.iter_rows():
			for cell in row :
				if cell.value=="CM" :
					tbcrm=cell.column


		for row in TB1.iter_rows():
			for cell in row :
				if cell.value=="CB" :
					tbcb=cell.column

		for row in TB1.iter_rows():
			for cell in row :

				if cell.value=="OB" :
					
					tbob=cell.column


		'Save Input DATA -----------------------------------------------------------------------------------------------'



		luntb=len(TB1[tbacc])
		try:
			AccountTB=[b.value for b in TB1[tbacc][tbrow:luntb+1]]
		except:
			messagebox.showerror("Error", "File: TB. Please insert the correct header for 'Account'")
		try:
			DescriptionTB=[b.value for b in TB1[tbdesc][tbrow:luntb+1]]
		except:
			messagebox.showerror("Error", "File: TB. Please insert the correct header for 'Description'")
		try:
			OBTB=[b.value for b in TB1[tbob][tbrow:luntb+1]]
		except:
			messagebox.showerror("Error", "File: TB. Please insert the correct header for 'OB'")
		try:
			DMTB=[b.value for b in TB1[tbdrm][tbrow:luntb+1]]
		except:
			messagebox.showerror("Error", "File: TB. Please insert the correct header for 'DM'")
		try:
			CMTB=[b.value for b in TB1[tbcrm][tbrow:luntb+1]]
		except:
			messagebox.showerror("Error", "File: TB. Please insert the correct header for 'CM'")
		try:
			CBTB=[b.value for b in TB1[tbcb][tbrow:luntb+1]]
		except:
			messagebox.showerror("Error", "File: TB. Please insert the correct header for 'CB'")


		"Create Excel Files---------------------------------------------------------------------"

		output=openpyxl.Workbook()

		tom=output.active

		tom.font=font1
		tom.title=refference1+" TOM Equity"

		F10_TB =output.create_sheet("F10_TB")


		"Create PBC FAR ---------------------------------------------------------------------"




		"Create PBC TB ---------------------------------------------------------------------"

		F10_TB.cell(row=1, column=1).value="Synt3"
		F10_TB.cell(row=1, column=2).value="Account"
		F10_TB.cell(row=1, column=3).value="Description"
		F10_TB.cell(row=1, column=4).value="OB"
		F10_TB.cell(row=1, column=5).value="DM"
		F10_TB.cell(row=1, column=6).value="CM"
		F10_TB.cell(row=1, column=7).value="CB"
		F10_TB.cell(row=1, column=8).value="Class"
		F10_TB.cell(row=1, column=9).value="Synt4"


		for i in range (1,10):
			F10_TB.cell(row=1, column= i).border=doubleborder
			F10_TB.cell(row=1, column= i).fill=headersblack
			F10_TB.cell(row=1, column= i).font=white

		for i in range (1,len(AccountTB)+1):
			F10_TB.cell(row=i+1, column=2).value=AccountTB[i-1]

		for i in range (1,len(DescriptionTB)+1):
			F10_TB.cell(row=i+1, column=3).value=DescriptionTB[i-1]

		for i in range (1, len(OBTB)+1):
			F10_TB.cell(row=i+1, column=4).value=OBTB[i-1]

		for i in range (1, len(DMTB)+1):
			F10_TB.cell(row=i+1, column=5).value=DMTB[i-1]

		for i in range (1, len(CMTB)+1):
			F10_TB.cell(row=i+1, column=6).value=CMTB[i-1]

		for i in range (1, len(CBTB)+1):
			F10_TB.cell(row=i+1, column=7).value=CBTB[i-1]

		for i in range (1, len(AccountTB)+1):
			F10_TB.cell(row=i+1, column=1).value="=LEFT(TRIM(B{0}),3)".format(i+1)

		for i in range (1, len(AccountTB)+1):
			F10_TB.cell(row=i+1, column=8).value="=LEFT(TRIM(B{0}),1)".format(i+1)

		for i in range (1, len(AccountTB)+1):
			F10_TB.cell(row=i+1, column=9).value="=LEFT(TRIM(B{0}),4)".format(i+1)

		for i in range (1, len(AccountTB)+1):
			for j in range (3, 7):
				F10_TB.cell(row=i+1, column=j).number_format='#,##0_);(#,##0)'


		"Adjust Column Width---------------------------------------------------------------------------------------------------------------------------------."

		# for col in F10_TB.columns:
		#     max_length = 0
		#     for cell in col:
		#         if cell.coordinate in F10_TB.merged_cells:
		#             continue
		#         try:
		#             if len(str(cell.value)) > max_length:
		#                 max_length = len(cell.value)
		#         except:
		#             pass
		#     adjusted_width = (max_length - 20)


		listanoua=['A','B','C','D','E','F','G','H','I','J']
		for column in listanoua:
			for i in listanoua:
				if (column==i):
					F10_TB.column_dimensions[column].width = 15

		"Create TOM Equity ---------------------------------------------------------------------"
		tom.column_dimensions['A'].width=1
		tom.column_dimensions['B'].width=40
		tom.column_dimensions['C'].width=45
		tom.column_dimensions['D'].width=15
		tom.column_dimensions['E'].width=15
		tom.column_dimensions['F'].width=15
		tom.column_dimensions['G'].width=15
		tom.column_dimensions['H'].width=12
		tom.column_dimensions['I'].width=12
		tom.column_dimensions['J'].width=30
		tom.sheet_view.showGridLines = False



		for i in range(1, 87):
			for j in range(1, 13):
				tom.cell(row=i, column=j).font=tah

		tom.cell(row=1, column=2).value="Client:"
		tom.cell(row=1, column=2).font=ft2
		tom.cell(row=1, column=3).value=clientname1
		tom.cell(row=1, column=3).font=font1

		tom.cell(row=2, column=2).value="Period end:"
		tom.cell(row=2, column=2).font=ft2
		tom.cell(row=2, column=3).value=yearEnd1
		tom.cell(row=2, column=3).number_format='mm/dd/yyyy'
		tom.cell(row=3, column=2).value="All amounts are in RON if not otherwise stated"
		tom.cell(row=3, column=2).font=italic
		tom.cell(row=5, column=2).value="Equity Table of Movement"
		tom.cell(row=5, column=2).font=equity
		tom.cell(row=1, column=8).value="Ref:"
		tom.cell(row=1, column=8).font=ft2
		tom.cell(row=1, column=9).value=refference1
		tom.cell(row=1, column=9).font=fontRedBold
		for i in range(1, 4):
			tom.cell(row=i, column=8).alignment=Alignment(horizontal='right', vertical='center', wrap_text=True)

		tom.cell(row=2, column=8).value="Prepared by:"
		tom.cell(row=2, column=8).font=ft2
		tom.cell(row=2, column=9).value=preparedBy1
		tom.cell(row=3, column=8).value="Date:"
		tom.cell(row=3, column=8).font=ft2
		tom.cell(row=3, column=9).value=datetime.date.today()
		tom.cell(row=3, column=9).number_format="dd/mm/yyyy"

		tom.cell(row=7, column=2).value="Procedure:"
		tom.cell(row=7, column=2).font=ft2
		tom.cell(row=7, column=3).value="Based on the TB, prepare a TOM for equity accounts."
		tom.cell(row=8, column=3).value="Explain the equity movements and trace them to the AGA minutes and other supporting documents, if applicable."
		tom.cell(row=9, column=3).value="Perform checks on equity movements - see below workings."

		tom.cell(row=11, column=2).value="Work done:"
		tom.cell(row=11, column=2).font=ft2
		tom.cell(row=11, column=3).value="Please see below:"

		tom.cell(row=13, column=2).value="Account Value"
		tom.cell(row=13, column=3).value="Account Description"
		tom.cell(row=13, column=4).value="OB as at 01.01. " + str(yearEnd1.year)
		tom.cell(row=13, column=5).value="Dr Mvm"
		tom.cell(row=13, column=6).value="Cr Mvm"
		tom.cell(row=13, column=7).value="CB as at 31.12." + str(yearEnd1.year) 
		tom.cell(row=13, column=8).value="Variance"
		tom.cell(row=13, column=9).value="%"
		tom.cell(row=13, column=10).value="Comments and Ref to supporting documents"

		tom.cell(row=14, column=2).value=1011
		tom.cell(row=15, column=2).value=1012
		tom.cell(row=16, column=2).value=1015
		tom.cell(row=17, column=2).value=1018
		tom.cell(row=18, column=2).value=1031

		tom.cell(row=19, column=2).value="Subtotal capital"

		tom.cell(row=20, column=2).value=104

		tom.cell(row=21, column=2).value="Subtotal share premium"

		tom.cell(row=22, column=2).value=105

		tom.cell(row=23, column=2).value="Subtotal revaluation reserve"

		tom.cell(row=24, column=2).value=1061
		tom.cell(row=25, column=2).value=1063
		tom.cell(row=26, column=2).value=1068

		tom.cell(row=27, column=2).value="Subtotal reserves"

		tom.cell(row=28, column=2).value=109
		tom.cell(row=29, column=2).value=141
		tom.cell(row=30, column=2).value=149
		tom.cell(row=31, column=2).value=117
		tom.cell(row=32, column=2).value=121
		tom.cell(row=33, column=2).value=129

		tom.cell(row=34, column=2).value="Subtotal Own Equity"

		tom.cell(row=35, column=2).value=1016
		tom.cell(row=36, column=2).value=1017

		tom.cell(row=37, column=2).value="Total EQUITY"

		tom.cell(row=14, column=3).value="Subscribed unpaid capial"
		tom.cell(row=15, column=3).value="Subscribed paid in capital"
		tom.cell(row=16, column=3).value="“Regii” patrimony"
		tom.cell(row=17, column=3).value="Patrimony of national research and development institutes"
		tom.cell(row=18, column=3).value="Other equity items"

		tom.cell(row=20, column=3).value="Share premium"

		tom.cell(row=22, column=3).value=" Revaluation reserve "

		tom.cell(row=24, column=3).value="Legal reserves"
		tom.cell(row=25, column=3).value="Statutory or contractual reserves"
		tom.cell(row=26, column=3).value="Other reserves"

		tom.cell(row=28, column=3).value="Own shares"
		tom.cell(row=29, column=3).value="Gains related to equity items"
		tom.cell(row=30, column=3).value="Losses related to equity items"
		tom.cell(row=31, column=3).value="Retained earnings"
		tom.cell(row=32, column=3).value="Profit or loss for the year"
		tom.cell(row=33, column=3).value="Profit appropriation"

		tom.cell(row=35, column=3).value="Public patrimony"
		tom.cell(row=36, column=3).value="Private patrimony"

		tom.cell(row=39, column=2).value="Working 1) Reconciliation Transfer of OB Profit & Loss to Retained Earnings"
		tom.cell(row=39, column=2).font=ft2

		tom.cell(row=41, column=2).value="OB 121 - profit from PY"
		tom.cell(row=42, column=2).value="OB 129 - PY profit"
		tom.cell(row=43, column=2).value="CM 117 - PY profit "
		tom.cell(row=44, column=2).value="Check OB P&L with CMvm account 117"
		tom.cell(row=44, column=2).font=fontRedNormal
		tom.cell(row=44, column=3).font=fontRedNormal
		tom.cell(row=45, column=2).value="Other movements in CMvm 117 if case, please detail"
		tom.cell(row=46, column=2).value="<p.y. accounting losses covered>"
		tom.cell(row=47, column=2).value="<correction of errors>"
		tom.cell(row=48, column=2).value="<transfer from 105>"
		# tom.cell(row=49, column=2).value="Check CMvm account 117"
		# tom.cell(row=49, column=2).font=fontRedNormal
		tom.cell(row=49, column=3).font=fontRedNormal

		tom.cell(row=51, column=2).value="Working 2) Reconciliation of Profit and Loss"
		tom.cell(row=51, column=2).font=ft2

		tom.cell(row=53, column=2).value="Total revenues as per TB"
		tom.cell(row=54, column=2).value="Total expenses as per TB"
		tom.cell(row=55, column=2).value="Profit/Loss recomputed"
		tom.cell(row=56, column=2).value="Check CB Profit and Loss"
		tom.cell(row=56, column=2).font=fontRedNormal
		tom.cell(row=56, column=3).font=fontRedNormal

		tom.cell(row=58, column=2).value="Working 3) Reconciliation of Dividends Distributed"
		tom.cell(row=58, column=2).font=ft2

		tom.cell(row=60, column=2).value="CMvm Account 457"
		tom.cell(row=61, column=2).value="DMvm Account 117"
		tom.cell(row=62, column=2).value="Check Dividends distributed"
		tom.cell(row=62, column=2).font=fontRedNormal
		tom.cell(row=62, column=3).font=fontRedNormal
		tom.cell(row=63, column=2).value="Other movements in DMvm 117 if case, please detail"
		tom.cell(row=64, column=2).value="<p.y. accounting lossses>"
		tom.cell(row=65, column=2).value="<other distributions from p.y. profit>"
		tom.cell(row=66, column=2).value="Check DMvm account 117"
		tom.cell(row=66, column=2).font=fontRedNormal
		tom.cell(row=66, column=3).font=fontRedNormal

		tom.cell(row=68, column=2).value="Working 4) Legal reserves reconciliation"
		tom.cell(row=68, column=2).font=ft2

		tom.cell(row=70, column=2).value="Share capital registered and cashed"
		tom.cell(row=71, column=2).value="20%  share capital"
		tom.cell(row=72, column=2).value="Legal reserve in balance"
		tom.cell(row=73, column=2).value="20%  of share capital reached?"
		tom.cell(row=74, column=2).value="Accounting profit"
		tom.cell(row=75, column=2).value="5%  of accounting profit if case (20%  of share capital not reached)"
		tom.cell(row=76, column=2).value="Check legal reserve in balance"
		tom.cell(row=76, column=2).font=fontRedNormal
		tom.cell(row=76, column=3).font=fontRedNormal

		tom.cell(row=78, column=2).value="Working 5) Minimum capital requirements check"
		tom.cell(row=78, column=2).font=ft2

		tom.cell(row=80, column=2).value="Subscribed share capital"
		tom.cell(row=81, column=2).value="Net assets"
		tom.cell(row=82, column=2).value="Check net assets > 0.5 Subscribed share capital"
		tom.cell(row=83, column=2).value="Check"
		tom.cell(row=83, column=2).font=fontRedNormal
		tom.cell(row=83, column=3).font=fontRedNormal

		tom.cell(row=86, column=2).value="Findings:"
		tom.cell(row=86, column=2).font=ft2

		tom.cell(row=19, column=4).value="=SUM(D14:D18)"
		tom.cell(row=19, column=5).value="=SUM(E14:E18)"
		tom.cell(row=19, column=6).value="=SUM(F14:F18)"
		tom.cell(row=19, column=7).value="=SUM(G14:G18)"

		tom.cell(row=21, column=4).value="=D20"
		tom.cell(row=21, column=5).value="=E20"
		tom.cell(row=21, column=6).value="=F20"
		tom.cell(row=21, column=7).value="=G20"

		tom.cell(row=23, column=4).value="=D22"
		tom.cell(row=23, column=5).value="=E22"
		tom.cell(row=23, column=6).value="=F22"
		tom.cell(row=23, column=7).value="=G22"

		tom.cell(row=27, column=4).value="=SUM(D24:D26)"
		tom.cell(row=27, column=5).value="=SUM(E24:E26)"
		tom.cell(row=27, column=6).value="=SUM(F24:F26)"
		tom.cell(row=27, column=7).value="=SUM(G24:G26)"

		tom.cell(row=34, column=4).value="=D19+D21+D23+D27+SUM(D28:D33)"
		tom.cell(row=34, column=5).value="=E19+E21+E23+E27+SUM(E28:E33)"
		tom.cell(row=34, column=6).value="=F19+F21+F23+F27+SUM(F28:F33)"
		tom.cell(row=34, column=7).value="=G19+G21+G23+G27+SUM(G28:G33)"

		tom.cell(row=37, column=4).value="=D34+SUM(D35:D36)"
		tom.cell(row=37, column=5).value="=E34+SUM(E35:E36)"
		tom.cell(row=37, column=6).value="=F34+SUM(F35:F36)"
		tom.cell(row=37, column=7).value="=G34+SUM(G35:G36)"

		tom.cell(row=14, column=4).value='=SUMIF(F10_TB!I:I,"1011",F10_TB!D:D)'
		tom.cell(row=15, column=4).value='=SUMIF(F10_TB!I:I,"1012",F10_TB!D:D)'
		tom.cell(row=16, column=4).value='=SUMIF(F10_TB!I:I,"1015",F10_TB!D:D)'
		tom.cell(row=17, column=4).value='=SUMIF(F10_TB!I:I,"1018",F10_TB!D:D)'
		tom.cell(row=18, column=4).value='=SUMIF(F10_TB!I:I,"1031",F10_TB!D:D)'

		tom.cell(row=14, column=5).value='=SUMIF(F10_TB!I:I,"1011",F10_TB!E:E)'
		tom.cell(row=15, column=5).value='=SUMIF(F10_TB!I:I,"1012",F10_TB!E:E)'
		tom.cell(row=16, column=5).value='=SUMIF(F10_TB!I:I,"1015",F10_TB!E:E)'
		tom.cell(row=17, column=5).value='=SUMIF(F10_TB!I:I,"1018",F10_TB!E:E)'
		tom.cell(row=18, column=5).value='=SUMIF(F10_TB!I:I,"1031",F10_TB!E:E)'

		tom.cell(row=14, column=6).value='=SUMIF(F10_TB!I:I,"1011",F10_TB!F:F)'
		tom.cell(row=15, column=6).value='=SUMIF(F10_TB!I:I,"1012",F10_TB!F:F)'
		tom.cell(row=16, column=6).value='=SUMIF(F10_TB!I:I,"1015",F10_TB!F:F)'
		tom.cell(row=17, column=6).value='=SUMIF(F10_TB!I:I,"1018",F10_TB!F:F)'
		tom.cell(row=18, column=6).value='=SUMIF(F10_TB!I:I,"1031",F10_TB!F:F)'

		tom.cell(row=14, column=7).value='=SUMIF(F10_TB!I:I,"1011",F10_TB!G:G)'
		tom.cell(row=15, column=7).value='=SUMIF(F10_TB!I:I,"1012",F10_TB!G:G)'
		tom.cell(row=16, column=7).value='=SUMIF(F10_TB!I:I,"1015",F10_TB!G:G)'
		tom.cell(row=17, column=7).value='=SUMIF(F10_TB!I:I,"1018",F10_TB!G:G)'
		tom.cell(row=18, column=7).value='=SUMIF(F10_TB!I:I,"1031",F10_TB!G:G)'

		tom.cell(row=20, column=4).value='=SUMIF(F10_TB!A:A,"104",F10_TB!D:D)'
		tom.cell(row=20, column=5).value='=SUMIF(F10_TB!A:A,"104",F10_TB!E:E)'
		tom.cell(row=20, column=6).value='=SUMIF(F10_TB!A:A,"104",F10_TB!F:F)'
		tom.cell(row=20, column=7).value='=SUMIF(F10_TB!A:A,"104",F10_TB!G:G)'

		tom.cell(row=22, column=4).value='=SUMIF(F10_TB!A:A,"105",F10_TB!D:D)'
		tom.cell(row=22, column=5).value='=SUMIF(F10_TB!A:A,"105",F10_TB!E:E)'
		tom.cell(row=22, column=6).value='=SUMIF(F10_TB!A:A,"105",F10_TB!F:F)'
		tom.cell(row=22, column=7).value='=SUMIF(F10_TB!A:A,"105",F10_TB!G:G)'

		tom.cell(row=24, column=4).value='=SUMIF(F10_TB!I:I,"1061",F10_TB!D:D)'
		tom.cell(row=25, column=4).value='=SUMIF(F10_TB!I:I,"1063",F10_TB!D:D)'
		tom.cell(row=26, column=4).value='=SUMIF(F10_TB!I:I,"1068",F10_TB!D:D)'

		tom.cell(row=24, column=5).value='=SUMIF(F10_TB!I:I,"1061",F10_TB!E:E)'
		tom.cell(row=25, column=5).value='=SUMIF(F10_TB!I:I,"1063",F10_TB!E:E)'
		tom.cell(row=26, column=5).value='=SUMIF(F10_TB!I:I,"1068",F10_TB!E:E)'

		tom.cell(row=24, column=6).value='=SUMIF(F10_TB!I:I,"1061",F10_TB!F:F)'
		tom.cell(row=25, column=6).value='=SUMIF(F10_TB!I:I,"1063",F10_TB!F:F)'
		tom.cell(row=26, column=6).value='=SUMIF(F10_TB!I:I,"1068",F10_TB!F:F)'

		tom.cell(row=24, column=7).value='=SUMIF(F10_TB!I:I,"1061",F10_TB!G:G)'
		tom.cell(row=25, column=7).value='=SUMIF(F10_TB!I:I,"1063",F10_TB!G:G)'
		tom.cell(row=26, column=7).value='=SUMIF(F10_TB!I:I,"1068",F10_TB!G:G)'

		tom.cell(row=28, column=4).value='=SUMIF(F10_TB!A:A,"109",F10_TB!D:D)'
		tom.cell(row=29, column=4).value='=SUMIF(F10_TB!A:A,"141",F10_TB!D:D)'
		tom.cell(row=30, column=4).value='=SUMIF(F10_TB!A:A,"149",F10_TB!D:D)'
		tom.cell(row=31, column=4).value='=SUMIF(F10_TB!A:A,"117",F10_TB!D:D)'
		tom.cell(row=32, column=4).value='=SUMIF(F10_TB!A:A,"121",F10_TB!D:D)'
		tom.cell(row=33, column=4).value='=SUMIF(F10_TB!A:A,"129",F10_TB!D:D)'

		tom.cell(row=28, column=5).value='=SUMIF(F10_TB!A:A,"109",F10_TB!E:E)'
		tom.cell(row=29, column=5).value='=SUMIF(F10_TB!A:A,"141",F10_TB!E:E)'
		tom.cell(row=30, column=5).value='=SUMIF(F10_TB!A:A,"149",F10_TB!E:E)'
		tom.cell(row=31, column=5).value='=SUMIF(F10_TB!A:A,"117",F10_TB!E:E)'
		tom.cell(row=32, column=5).value='=SUMIF(F10_TB!A:A,"121",F10_TB!E:E)'
		tom.cell(row=33, column=5).value='=SUMIF(F10_TB!A:A,"129",F10_TB!E:E)'

		tom.cell(row=28, column=6).value='=SUMIF(F10_TB!A:A,"109",F10_TB!F:F)'
		tom.cell(row=29, column=6).value='=SUMIF(F10_TB!A:A,"141",F10_TB!F:F)'
		tom.cell(row=30, column=6).value='=SUMIF(F10_TB!A:A,"149",F10_TB!F:F)'
		tom.cell(row=31, column=6).value='=SUMIF(F10_TB!A:A,"117",F10_TB!F:F)'
		tom.cell(row=32, column=6).value='=SUMIF(F10_TB!A:A,"121",F10_TB!F:F)'
		tom.cell(row=33, column=6).value='=SUMIF(F10_TB!A:A,"129",F10_TB!F:F)'

		tom.cell(row=28, column=7).value='=SUMIF(F10_TB!A:A,"109",F10_TB!G:G)'
		tom.cell(row=29, column=7).value='=SUMIF(F10_TB!A:A,"141",F10_TB!G:G)'
		tom.cell(row=30, column=7).value='=SUMIF(F10_TB!A:A,"149",F10_TB!G:G)'
		tom.cell(row=31, column=7).value='=SUMIF(F10_TB!A:A,"117",F10_TB!G:G)'
		tom.cell(row=32, column=7).value='=SUMIF(F10_TB!A:A,"121",F10_TB!G:G)'
		tom.cell(row=33, column=7).value='=SUMIF(F10_TB!A:A,"129",F10_TB!G:G)'

		tom.cell(row=35, column=4).value='=SUMIF(F10_TB!I:I,"1016",F10_TB!D:D)'
		tom.cell(row=36, column=4).value='=SUMIF(F10_TB!I:I,"1017",F10_TB!D:D)'

		tom.cell(row=35, column=5).value='=SUMIF(F10_TB!I:I,"1016",F10_TB!E:E)'
		tom.cell(row=36, column=5).value='=SUMIF(F10_TB!I:I,"1017",F10_TB!E:E)'

		tom.cell(row=35, column=6).value='=SUMIF(F10_TB!I:I,"1016",F10_TB!F:F)'
		tom.cell(row=36, column=6).value='=SUMIF(F10_TB!I:I,"1017",F10_TB!F:F)'

		tom.cell(row=35, column=7).value='=SUMIF(F10_TB!I:I,"1016",F10_TB!G:G)'
		tom.cell(row=36, column=7).value='=SUMIF(F10_TB!I:I,"1017",F10_TB!G:G)'

		for i in range(13, 37):
			tom.cell(row=i+1, column=8).value="=G{0}-D{0}".format(i+1)

		for i in range(13, 37):
			tom.cell(row=i+1, column=9).value="=IFERROR(H{0}/D{0},0)".format(i+1)
			tom.cell(row=i+1, column=9).number_format='0.00%'

		tom.cell(row=41, column=3).value="=D32"
		tom.cell(row=42, column=3).value="=D33"
		tom.cell(row=43, column=3).value="=F31"
		tom.cell(row=44, column=3).value="=C41+C42+C43"
		# tom.cell(row=49, column=3).value="=C43+SUM(C44:C47)"

		tom.cell(row=53, column=3).value='=SUMIF(F10_TB!H:H,"7",F10_TB!G:G)'
		tom.cell(row=54, column=3).value='=SUMIF(F10_TB!H:H,"6",F10_TB!G:G)'
		tom.cell(row=55, column=3).value='=C53+C54'
		tom.cell(row=56, column=3).value='=G32-C55'

		tom.cell(row=60, column=3).value='=SUMIF(F10_TB!A:A,"457",F10_TB!F:F)'
		tom.cell(row=61, column=3).value="=E31"
		tom.cell(row=62, column=3).value="=C60-C61"
		tom.cell(row=66, column=3).value="=C62+SUM(C63:C65)"

		tom.cell(row=70, column=3).value="=G15"
		tom.cell(row=71, column=3).value="=C70*0.2  "
		tom.cell(row=72, column=3).value="=G24"
		tom.cell(row=73, column=3).value='=IF(ABS(C72)>=ABS(C71),"yes","no")'
		tom.cell(row=74, column=3).value="=SUMIF(F10_TB!A:A,121,F10_TB!G:G)"

		tom.cell(row=75, column=3).value="=C74*0.05"
		tom.cell(row=76, column=3).value='=IF(C73="yes","ok",IF(C75>C71+C72,"ok","not ok"))'

		tom.cell(row=80, column=3).value="=-G15"
		tom.cell(row=81, column=3).value="=-G37"
		tom.cell(row=82, column=3).value="=IFERROR(C80/C81,0)"
		tom.cell(row=83, column=3).value='=IF(C81>0.5*C80,"ok","not ok")'

		#sa fac synt de 3 pt conturile cu 3 cifre! - pt maine 

		for i in range(41, 84):
			tom.cell(row=i, column=3).number_format='#,##0_);(#,##0)'

		for i in range(4, 9):
			for j in range(14, 38):
				tom.cell(row=j, column=i).number_format='#,##0_);(#,##0)'

		for i in range(80, 84):
			tom.cell(row=i, column=1).border=rightborder
			tom.cell(row=i, column=3).border=rightborder

		for j in range(2, 4):
			tom.cell(row=79, column=j).border=solidborder
			tom.cell(row=83, column=j).border=solidborder


		for i in range(70, 77):
			tom.cell(row=i, column=1).border=rightborder
			tom.cell(row=i, column=3).border=rightborder

		for j in range(2, 4):
			tom.cell(row=69, column=j).border=solidborder
			tom.cell(row=76, column=j).border=solidborder

		for i in range(60, 67):
			tom.cell(row=i, column=1).border=rightborder
			tom.cell(row=i, column=3).border=rightborder

		for j in range(2, 4):
			tom.cell(row=59, column=j).border=solidborder
			tom.cell(row=66, column=j).border=solidborder


		for i in range(53, 57):
			tom.cell(row=i, column=1).border=rightborder
			tom.cell(row=i, column=3).border=rightborder

		for j in range(2, 4):
			tom.cell(row=52, column=j).border=solidborder
			tom.cell(row=56, column=j).border=solidborder


		for i in range(41, 50):
			tom.cell(row=i, column=1).border=rightborder
			tom.cell(row=i, column=3).border=rightborder

		for j in range(2, 4):
			tom.cell(row=40, column=j).border=solidborder
			tom.cell(row=49, column=j).border=solidborder

		for i in range(2, 8):
			tom.cell(row=13, column=i).fill=headersblack
			tom.cell(row=19, column=i).fill=headersblack
			tom.cell(row=21, column=i).fill=headersblack
			tom.cell(row=23, column=i).fill=headersblack
			tom.cell(row=27, column=i).fill=headersblack
			tom.cell(row=34, column=i).fill=headersblack
			tom.cell(row=37, column=i).fill=headersblack
			tom.cell(row=13, column=i).font=white
			tom.cell(row=19, column=i).font=white
			tom.cell(row=21, column=i).font=white
			tom.cell(row=23, column=i).font=white
			tom.cell(row=27, column=i).font=white
			tom.cell(row=34, column=i).font=white
			tom.cell(row=37, column=i).font=white

		for i in range(8, 10):
			tom.cell(row=13, column=i).fill=headersorange
			tom.cell(row=19, column=i).fill=headersorange
			tom.cell(row=21, column=i).fill=headersorange
			tom.cell(row=23, column=i).fill=headersorange
			tom.cell(row=27, column=i).fill=headersorange
			tom.cell(row=34, column=i).fill=headersorange
			tom.cell(row=37, column=i).fill=headersorange
			tom.cell(row=13, column=i).font=white
			tom.cell(row=19, column=i).font=white
			tom.cell(row=21, column=i).font=white
			tom.cell(row=23, column=i).font=white
			tom.cell(row=27, column=i).font=white
			tom.cell(row=34, column=i).font=white
			tom.cell(row=37, column=i).font=white

		tom.cell(row=13, column=10).fill=headersblack
		tom.cell(row=19, column=10).fill=headersblack
		tom.cell(row=21, column=10).fill=headersblack
		tom.cell(row=23, column=10).fill=headersblack
		tom.cell(row=27, column=10).fill=headersblack
		tom.cell(row=34, column=10).fill=headersblack
		tom.cell(row=37, column=10).fill=headersblack
		tom.cell(row=13, column=10).font=white
		tom.cell(row=19, column=10).font=white
		tom.cell(row=21, column=10).font=white
		tom.cell(row=23, column=10).font=white
		tom.cell(row=27, column=10).font=white
		tom.cell(row=34, column=10).font=white
		tom.cell(row=37, column=10).font=white

		for i in range(2, 11):
			tom.cell(row=13, column=i).alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)

		tom.cell(row=49, column=4).border=leftborder
		tom.cell(row=56, column=4).border=leftborder
		tom.cell(row=66, column=4).border=leftborder
		tom.cell(row=76, column=4).border=leftborder
		tom.cell(row=83, column=4).border=leftborder




		def Mbox (title, text, style):
			return ctypes.windll.users32.MessageBoxW (0, text, title, style)

		# save1=tk.Tk()
		# save1.withdraw()
		folderpath="home/auditappnexia/output/Equity"
		file_pathFS = os.path.join(folderpath, 'TOM Equity.xlsx')
		output.save(file_pathFS)


	return send_from_directory(folderpath, "TOM Equity.xlsx", as_attachment=True)
@app.route('/MergeFiles/Instructions', methods=['GET'])
def downloadMerge():
		filepath = "/home/auditappnexia/output/Merge"
 
		return send_from_directory(filepath,"Instructions - Merge Files.docx", as_attachment=True)
@app.route('/MergeFiles/GT2hAORxnRTx8HrE2mSyYRiqgOp')
def Merge():
	return render_template('Merge.html')
@app.route('/MergeFiles/GT2hAORxnRTx8HrE2mSyYRiqgOp', methods=['POST', 'GET'])
def Merge_process():
	# global column_description, coloana_opTB_tb, coloana_cr_tb, coloana_db_tb, coloana_clTB_tb, column_tb, row_tb, clientName
	clientname1 = request.form['client']
	isChecked1=request.form.get("multipleFiles")
	# print(isChecked1, 'ce mortii tai')
	
	if request.method == 'POST':
		file_location_db = request.files.getlist("database")
		isChecked=request.form.get("multipleFiles")

		if str(isChecked1)=="": #daca e bifat
			isChecked=0
			
		else:
			isChecked=1
		# print(isChecked, "fmm")


		wb = openpyxl.Workbook()
		ws = wb.active

		Sheet1 = wb.create_sheet("Database")

		Sheet1.cell(row=1, column=2).value = "File name"
		Sheet1.cell(row=1, column=3).value = "Sheet name"
		Sheet1.column_dimensions['B'].width = 35
		Sheet1.column_dimensions['C'].width = 35


		if isChecked==0:
			print("Is checked")

			db = openpyxl.load_workbook(file_location_db[0], data_only='True')
			db1 = db.active
			# head, tail = os.path.split(file_location_db[0])
			# print(tail)

			for a in db.sheetnames:
				# print(a)

				mr = db[str(a)].max_row
				mc = db[str(a)].max_column
				for i in range(1, mr + 1):
					for j in range(1, mc + 1):
						c = db[str(a)].cell(row=i, column=j)
						Sheet1.cell(row=i + 1, column=4 + j).value = c.value
						Sheet1.cell(row=i + 1, column=2).value = str(file_location_db[0].filename)
						Sheet1.cell(row=i + 1, column=3).value = a

		else:
			print("is NOT checked")
			startRand=0
			startCol=0
			for a in range(0, len(file_location_db)):
				# print(a.filename, 'NUME FISIERE')
				# head, tail = os.path.split(a)
				# print(tail)
				
				db = openpyxl.load_workbook(file_location_db[a], data_only='True')
				db1 = db.active
				sheetnames = str(db.sheetnames)
				# print(sheetnames)


				mr = db1.max_row
				mc = db1.max_column
				for i in range(1, mr + 1):
					startRand=startRand+1
					for j in range(1, mc + 1):

						c = db1.cell(row=i, column=j)
						Sheet1.cell(row=startRand + 1, column=4 + j).value = c.value
						Sheet1.cell(row=startRand + 1, column=2).value = str(file_location_db[a].filename)
						Sheet1.cell(row=startRand + 1, column=3).value = sheetnames

		std = wb["Sheet"]
		wb.remove(std)
		# folderpath = filedialog.askdirectory()
		# file_pathFS = os.path.join(folderpath, "Dabatase "+" "+clientname1+".xlsx")
		# tk.messagebox.showinfo('Notice', 'Merging data is complete.')
		folderpath="home/auditappnexia/output/Merge"
		file_pathFS = os.path.join(folderpath, 'Database '+ str(clientname1) +'.xlsx')
		wb.save(file_pathFS)

	return send_from_directory(folderpath, "Database " + str(clientname1) + ".xlsx", as_attachment=True)

@app.route('/JournalEntriesRoutines/GTRk1PboB5knZ8elePOv5WjBUEC/Instructions', methods=['GET'])
def downloadJET():
		filepath = "/home/auditappnexia/output/jet"
 
		return send_from_directory(filepath,"Instructions - Journal Entries Routines.docx", as_attachment=True)

@app.route('/SplitFiles/Instructions', methods=['GET'])
def downloadSplit():
		filepath = "/home/auditappnexia/output/Split"
 
		return send_from_directory(filepath,"Instructions - Split Files.docx", as_attachment=True)
@app.route('/SplitFiles/GTuKs87BZEXuVAXP57qjmBZkwj2')
def Split():
	return render_template('split.html')
@app.route('/SplitFiles/GTuKs87BZEXuVAXP57qjmBZkwj2', methods=['POST', 'GET'])
def Split_process():
	namec = request.form['client']
	folderpath="home/auditappnexia/output/Split"
	os.mkdir(folderpath+"\\"+namec)
	def make_archive(source, destination):
		base = os.path.basename(destination)
		name = base.split('.')[0]
		format = base.split('.')[1]
		archive_from = os.path.dirname(source)
		archive_to = os.path.basename(source.strip(os.sep))
		shutil.make_archive(name, format, archive_from, archive_to)
		shutil.move('%s.%s'%(name,format), destination)

	if request.method == 'POST':

		workingsblue2= Font(bold=True, italic=True, name='Tahoma', size=8,color='FFFFFF')
		lbluefill = PatternFill(start_color='7030A0',
							end_color='7030A0',
							fill_type='solid')
		grifill=PatternFill(start_color='c4d79b',end_color='c4d79b',fill_type='solid')
		yellow=PatternFill(start_color='ffff00',end_color='ffff00',fill_type='solid')
		blueFill = PatternFill(start_color='00AEAC',
							end_color='00AEAC',
							fill_type='solid')
		doubleborder = Border(bottom=Side(style='double'))
		solidborder = Border(bottom=Side(style='thick'))
		solidborderstanga = Border(left=Side(style='thin'))
		rightborder = Border(right=Side(style='thin'))
		rightdouble = Border (right=Side(style='thin'), bottom=Side(style='double'))
		rightmedium = Border (right=Side(style='thin'), bottom=Side(style='medium'))
		solidborderdreapta = Border(right=Side(style='thin'))
		solidbordersus = Border(top=Side(style='thin'))
		fontitalic = Font(name='Tahoma', size=8, bold=True, italic=True)
		font1 = Font(name='Tahoma', size=8)
		font2 = Font(name='Tahoma', size=8, bold=True)
		fontRed = Font(name='Tahoma', size=8, bold=True, color= 'FF0000')
		fontRedDiff=Font(name="Tahoma", color='FF0000', size=11, )
		fontGT = Font (name='GT Logo', size=8)
		workingsblue = Font(color='2F75B5', bold=True, name='Tahoma', size=8 )
		headers= Font(bold=True, italic=True, name='Tahoma', size=8,color='FFFFFF') 
		headersblue = PatternFill(start_color='7030A0',
						end_color='7030A0',
						fill_type='solid')
		headerspurple= PatternFill(start_color='65CDCC',
							end_color='65CDCC',
							fill_type='solid')
		total=PatternFill(start_color='DDD9C4',
						end_color='DDD9C4',
						fill_type='solid')
		greenbolditalic= Font(bold=True, italic=True,  color='C0504D', name='Tahoma', size=8)
		greenbolditalic= Font(bold=True, italic=True,  color='00af50')
		fontalb = Font(italic=True, color="bfbfbf", size=8, name='Tahoma')
		# trialb=request.files["TB"]
		je=request.files["JE"]

		tb=openpyxl.load_workbook(je,data_only='True')
		tb1=tb.active

		for row in tb1.iter_rows():
			for cell in row :

				if cell.value=="Coloana Split" :
					
					glac=cell.column
					glrow=cell.row


		try:
			lungl=len(tb1[glac])
		except:
			flash("Please insert the correct header for Account GL Debit in Journal Entries file")
			return render_template("index.html")

		try:
			accountgldebit=[b.value for b in tb1[glac][glrow:lungl]]
		except:
			flash("Please insert the correct header for Account GL Debit in Journal Entries file")
			return render_template("index.html")



		syntaccgldebit1=[]
		syntaccglcredit1=[]
		listadenis=[]
		
		row_denis=tb1['1']
		column_denis=tb1.max_column


		for i in range(0,len(accountgldebit)):
				syntaccgldebit1.append(str(accountgldebit[i]))
		# for i in range(0,len(accountglcredit)):
		# 		syntaccglcredit1.append(str(accountglcredit[i]))

		syntaccgldebit=[]
		syntaccglcredit=[]

		for i in range(0,len(syntaccgldebit1)):
			try:
				syntaccgldebit1[i]=str(syntaccgldebit1[i])
				syntaccgldebit.append(syntaccgldebit1[i])
			except:
				pass
		lisunicaacd=list(set(syntaccgldebit))
		listaunica=list(set(lisunicaacd))
		print()
		for i in range(0,len(listaunica)):

			for i in range(0,len(listaunica)):
				if(str(listaunica[i])!="6"):
					excel=Workbook()
					ws=excel.active
					ws.title="Database"
					ws.sheet_view.showGridLines = False
					ws.cell(row=1,column=1).value="Coloana Split"
					row=1

					listagldebit=[]
					for j in range(0,len(accountgldebit)):
						if(str(listaunica[i])==str(accountgldebit[j])):
							ws.cell(row=row+1,column=1).value=accountgldebit[j]
							listagldebit.append(accountgldebit[j])		
							row=row+1
							row_value=[]
							column_value=[]
							for cell in row_denis:
								row_value.append(str(cell.value))
							for z in range(0,len(row_value)):
								ws.cell(row=1,column=z+1).value=row_value[z]
							for o in range(1,len(row_value)+1):
								ws.cell(row=1,column=o).border=doubleborder
								ws.cell(row=1,column=o).font=font2
								ws.cell(row=1,column=o).fill=headerspurple
								for g in range(1,column_denis):
									ws.cell(row=row	,column=1+g).value=tb1.cell(row=j+glrow+1,column=g+1).value
									# ws.save()
							# listanoua=['A',column_denis]
							# for column in listanoua:
							# 	column_value.append(row.value)
							# 	for g in listanoua:
							# 		ws.cell(row=2+g,column=g).value=column_value[g]
					# 	print(row_value)
					folderpath="home/auditappnexia/output/Split"
					# excel._sheets =[excel._sheets[i] for i in myorder]
					excel.save(folderpath+"\\"+str(namec)+"\\"+str(listaunica[i])+".xlsx")
		make_archive("home/auditappnexia/output/Split/"+str(namec),"home/auditappnexia/output/Split/"+str(namec)+"/Split "+namec+".zip")                
				# shutil.make_archive(name, format, archive_from, archive_to)
		file_pathFS = os.path.join(folderpath, namec)    
		return send_from_directory(file_pathFS,"Split "+ namec+".zip",as_attachment=True)    



@app.route('/JournalEntriesRoutines/GTRk1PboB5knZ8elePOv5WjBUEC/')
def JET():
	return render_template('JET.html')

@app.route('/JournalEntriesRoutines/GTRk1PboB5knZ8elePOv5WjBUEC/', methods=['POST', 'GET'])
def JET_process():
    clientname1 = request.form['client']
    # try:
    yearEnd1 = datetime.datetime.strptime(
        request.form['yearEnd'],
        '%Y-%m-%d')
    # except:
    #     pass
    # print(yearEnd1)
    # preparedBy1 = request.form['preparedBy']
    # datePrepared1 = datetime.datetime.now().date()
    # refference1 = request.form['reff']

    if request.method == 'POST':
        file_Details = request.files["uploadJET"]
        # manag = request.form["manag"].split()
        corFrd = request.form["xyz"].split("; ")
        largeAm = request.form["largeAm"]
        try:
            spDays = request.form['specD'].split(",")
            listaCostica = [datetime.datetime.strptime(x,'%m/%d/%Y') for x in spDays]
        except:
            pass
        
        misJE=request.form.get("misJE")
        largeAM = request.form.get("largeAM")
        speCom = request.form.get("speCom")
        poWE = request.form.get("poWE")
        poSD = request.form.get("poSD")
        poUSER = request.form.get("poUSER")
        endZero = request.form.get("endZero")
        endNine = request.form.get("endNine")
        missExplanation = request.form.get("missExp")
        oobal = request.form.get("oobal")

        if misJE=="":
            print("Da misJE")  # daca e bifat
            misJE = 1
        else:
            print(misJE)            
            misJE = 0
            print("Nu misJE")

        if largeAM=="":
            print("Da largeAM")  # daca e bifat
            largeAM = 1
        else:
            print(largeAM)            
            largeAM = 0
            print("Nu largeAM")

        if speCom=="":
            print("Da speCom")  # daca e bifat
            speCom = 1
        else:
            print(speCom)            
            speCom = 0
            print("Nu speCom")

        if poWE=="":
            print("Da poWE")  # daca e bifat
            poWE = 1
        else:
            print(poWE)            
            poWE = 0
            print("Nu poWE")

        if poSD=="":
            print("Da poSD")  # daca e bifat
            poSD = 1
        else:
            print(poSD)            
            poSD = 0
            print("Nu poSD")

        if poUSER=="":
            print("Da poUSER")  # daca e bifat
            poUSER = 1
        else:
            print(poUSER)            
            poUSER = 0
            print("Nu poUSER")

        if endZero=="":
            print("Da endZero")  # daca e bifat
            endZero = 1
        else:
            print(endZero)            
            endZero = 0
            print("Nu endZero")

        if endNine=="":
            print("Da endNine")  # daca e bifat
            endNine = 1
        else:
            print(endNine)            
            endNine = 0
            print("Nu endNine")

        if missExplanation=="":
            print("Da missExplanation")  # daca e bifat
            missExplanation = 1
        else:
            print(missExplanation)            
            missExplanation = 0
            print("Nu missExplanation")

        if oobal=="":
            print("Da oobal")  # daca e bifat
            oobal = 1
        else:
            print(oobal)            
            oobal = 0
            print("Nu oobal")

        ft1 = Font(name='Arial', size=10, bold=True)
        f_testname = Font(name='Arial', size=15, color='614C77', bold=True)
        f_info = Font(name='Arial', size=10, color='614C77', bold=True)
        cap_tabel = Font(name='Arial', size=10, color="FFFFFF", bold=True)
        cap_tabel_color_PBC = PatternFill(start_color='808080', end_color='808080', fill_type='solid')  # grey
        cap_tabel_color_GT = PatternFill(start_color='00AEAC', end_color='00AEAC', fill_type='solid')  # indigo #B1A0C7
        fprocentaj = Font(name='Arial', size=10, color="FF0000", bold=True)
        font_worksheet = Font(name='Arial', size=10)
        check_font = Font(name='Arial', size=10, color="FF0000", bold=True)
        check_font_1 = Font(name='Arial', size=10, color="FF0000", bold=False)
        cap_tabel_color_GT_movdeschis = PatternFill(start_color='00AEAC', end_color='00AEAC', fill_type='solid')
        cap_tabel_color_GT_movinchis = PatternFill(start_color='00AEAC', end_color='00AEAC', fill_type='solid')
        blue_bold_font = Font(name='Arial', size=10, color="0070C0", bold=True)
        blue_thin_font = Font(name='Arial', size=10, color="0070C0", bold=False)
        # hyperlink_color=PatternFill(start_color='0563C1', end_color='0563C1', fill_type='solid')
        hyperlink_font=Font(name='Arial', size=10, color="0563C1", bold=False)


        wb = openpyxl.Workbook()
        ws = wb.active
        summary = wb.create_sheet("Summary")
        summary.sheet_view.showGridLines = False

        summary.cell(row=1, column=1).value="Client"
        summary.cell(row=2, column=1).value="Period End"
        summary.cell(row=4, column=1).value="Work Done"
        summary.cell(row=7, column=1).value="Purpose"
        summary.cell(row=7, column=2).value="To obtain sufficient and appropriate audit evidence that the completeness, accuracy and valuation regarding journal entries are not significantly misstated."
        summary.cell(row=11, column=1).value="Work Done"
        summary.cell(row=11, column=2).value="We obtained from the client the general ledger at 31.12.2021"
        summary.cell(row=12, column=2).value="We reconciled the  DR and CR mvm from general ledger with the DR and CR mvm from TB in order to ensure completeness."

        summary.cell(row=1, column=2).value=clientname1
        summary.cell(row=2, column=2).value=yearEnd1

        #-----------------------------------FORMATS AND COLORS--------------------------------------------

        summary.cell(row=2, column=2).number_format="mm/dd/yyyy"

        summary.cell(row=1, column=1).font=ft1
        summary.cell(row=2, column=1).font=ft1
        summary.cell(row=4, column=1).font=font_worksheet
        summary.cell(row=7, column=1).font=font_worksheet
        summary.cell(row=7, column=2).font=font_worksheet
        summary.cell(row=11, column=1).font=font_worksheet
        summary.cell(row=11, column=2).font=font_worksheet
        summary.cell(row=12, column=2).font=font_worksheet

        summary.cell(row=14, column=3).font=font_worksheet
        
        #---------------------------IMPORTING JOURNAL ENTRY----------------------------------------------------

        # details=openpyxl.load_workbook('D:\\WEBDEV\\Platforma apps web\\Audit\\JET\\testing\\JE test.xlsx', data_only = 'True')
        details=openpyxl.load_workbook(file_Details, data_only='True')
        details1 = details.active

        je = wb.create_sheet("JE")
        je.sheet_view.showGridLines = False

        summary.cell(row=14, column=3).value="JE"
        summary.cell(row=14, column=3).hyperlink='#JE!A1'
        summary.cell(row=14, column=3).font=hyperlink_font
        
        je.cell(row=1, column=1).value="Client"
        je.cell(row=2, column=1).value="Period End"

        je.cell(row=1, column=2).value=clientname1
        je.cell(row=2, column=2).value=yearEnd1

        je.cell(row=2, column=2).number_format="mm/dd/yyyy"

        je.cell(row=3, column=9).value="M"
        je.cell(row=4, column=9).value="PM"
        je.cell(row=5, column=9).value="TA"
        

        #de facut campuri in intefata pentru introducerea materialitatilor
        # M=409000
        try:
            PM=int(largeAm)
        except:
            PM=0
        # TA=20450

        # je.cell(row=3, column=10).value=M
        je.cell(row=4, column=10).value=PM
        je.cell(row=4, column=10).number_format='#,##0_);(#,##0)'
        je.column_dimensions['J'].width = 12
        # je.cell(row=5, column=10).value=TA    

        je.cell(row=8, column=1).value="JE Number"
        je.cell(row=8, column=2).value="Posting Date"
        je.cell(row=8, column=3).value="Description"
        je.cell(row=8, column=4).value="Amount"
        je.cell(row=8, column=5).value="Account"
        je.cell(row=8, column=6).value="User"
        # je.cell(row=8, column=7).value="Debit Amount (Python)"
        # je.cell(row=8, column=8).value="Credit Amount (Python)"
        je.cell(row=8, column=7).value="Debit/Credit"

        try:
            for row in details1.iter_rows():
                for cell in row:
                    if cell.value == "Account Debit":
                        row_tb = cell.row
                        column_acDB = cell.column
                        lun = len(details1[cell.column])
            acDeb = [b.value for b in details1[column_acDB][row_tb:lun]]

            for row in details1.iter_rows():
                for cell in row:
                    if cell.value == "Account Credit":
                        row_tb = cell.row
                        column_acCR = cell.column
                        lun = len(details1[cell.column])
            acCre = [b.value for b in details1[column_acCR][row_tb:lun]]

            for row in details1.iter_rows():
                for cell in row:
                    if cell.value == "Amount Debit":
                        row_tb = cell.row
                        column_amDB = cell.column
                        lun = len(details1[cell.column])
            amDeb = [b.value for b in details1[column_amDB][row_tb:lun]]
            # amDeb=int(amDeb)
            amDeb1=[]
            for i in range(0, len(amDeb)):
                if amDeb[i] == None:
                    amDeb1.append(0)
                else:
                    amDeb1.append(amDeb[i])
            # print(amDeb1)

            for row in details1.iter_rows():
                for cell in row:
                    if cell.value == "Amount Credit":
                        row_tb = cell.row
                        column_amCR = cell.column
                        lun = len(details1[cell.column])
            amCre = [b.value for b in details1[column_amCR][row_tb:lun]]
            # amCre=int(amCre)
            amCre1=[]
            for i in range(0, len(amCre)):
                if amCre[i] == None:
                    amCre1.append(0)
                else:
                    amCre1.append(amCre[i])
            # print(amCre1)

            for row in details1.iter_rows():
                for cell in row:
                    if cell.value == "JE Number":
                        row_tb = cell.row
                        column_je = cell.column
                        lun = len(details1[cell.column])
            jeNo = [b.value for b in details1[column_je][row_tb:lun]]

            jeNoo=[]
            for i in range(0, len(jeNo)):
                if jeNo[i] == None:
                    jeNoo.append(0)
                else:
                    jeNoo.append(jeNo[i])


            for row in details1.iter_rows():
                for cell in row:
                    if cell.value == "Posting Date":
                        row_tb = cell.row
                        column_date = cell.column
                        lun = len(details1[cell.column])
            postDate = [b.value for b in details1[column_date][row_tb:lun]]

            for row in details1.iter_rows():
                for cell in row:
                    if cell.value == "Description":
                        row_tb = cell.row
                        column_desc = cell.column
                        lun = len(details1[cell.column])
            description = [b.value for b in details1[column_desc][row_tb:lun]]

            for row in details1.iter_rows():
                for cell in row:
                    if cell.value == "User":
                        row_tb = cell.row
                        column_usr = cell.column
                        lun = len(details1[cell.column])
            users = [b.value for b in details1[column_usr][row_tb:lun]]
            for row in details1.iter_rows():
                for cell in row:
                    if cell.value == "Debit/Credit":
                        row_tb = cell.row
                        column_usr = cell.column
                        lun = len(details1[cell.column])
            dbcd = [b.value for b in details1[column_usr][row_tb:lun]]
            descriptionFull=[]
            for c in range(0, len(description)):
                if description[c] == None:
                    descriptionFull.append("Nu exista descriere")
                else:
                    descriptionFull.append(description[c])
            
            account=[]
            amount=[]
            flag=[]
            for i in range(0, len(acDeb)):
                if acDeb[i] == None:
                    # print(acCre[i])
                    account.append(acCre[i])
                    flag.append("Cr")
                else:
                    account.append(acDeb[i])
                    flag.append("Db")
            # print(account)
            

            for i in range(0, len(amDeb1)):    
                if amDeb1[i] == None or amDeb1[i] == 0:
                    amount.append(-amCre1[i])
                else:
                    amount.append(amDeb1[i])
            # print(amount)
            amountDb=amDeb1
            amountCr=amCre1
            # amountss=amount
            # for i in range(0, len(acDeb)):
            #     if acDeb[i] == None:
            #         flag.append("Credit")
            #     else:
            #         flag.append("Debit")
            # print(flag)

            # for i in range(0,len(flag)):
            #     je.cell(row=9+i, column=7).value=flag[i]

            typeOfTr=[]
            for c in range(0, len(amount)):
                if amount[c] <= 0:
                    typeOfTr.append("Credit")
                else:
                    typeOfTr.append("Debit")

            
        except:
            try:
                for row in details1.iter_rows():
                    for cell in row:
                        if cell.value == "JE Number":
                            row_tb = cell.row
                            column_je = cell.column
                            lun = len(details1[cell.column])
                jeNo = [b.value for b in details1[column_je][row_tb:lun]]

                jeNoo=[]
                for i in range(0, len(jeNo)):
                    if jeNo[i] == None:
                        jeNoo.append(0)
                    else:
                        jeNoo.append(jeNo[i])

                for row in details1.iter_rows():
                    for cell in row:
                        if cell.value == "Posting Date":
                            row_tb = cell.row
                            column_date = cell.column
                            lun = len(details1[cell.column])
                postDate = [b.value for b in details1[column_date][row_tb:lun]]

                for row in details1.iter_rows():
                    for cell in row:
                        if cell.value == "Description":
                            row_tb = cell.row
                            column_desc = cell.column
                            lun = len(details1[cell.column])
                description = [b.value for b in details1[column_desc][row_tb:lun]]

                descriptionFull=[]
                for c in range(0, len(description)):
                    if description[c] == None:
                        descriptionFull.append("Nu exista descriere")
                    else:
                        descriptionFull.append(description[c])

                for row in details1.iter_rows():
                    for cell in row:
                        if cell.value == "Account":
                            row_tb = cell.row
                            column_AmC = cell.column
                            lun = len(details1[cell.column])
                account = [b.value for b in details1[column_AmC][row_tb:lun]]
                # print(amountCredit)
                # print("RULEAZA CAZUL ASTA")
                for row in details1.iter_rows():
                    for cell in row:
                        if cell.value == "Amount":
                            row_tb = cell.row
                            column_AmD = cell.column
                            lun = len(details1[cell.column])
                amount = [b.value for b in details1[column_AmD][row_tb:lun]]

                for row in details1.iter_rows():
                    for cell in row:
                        if cell.value == "User":
                            row_tb = cell.row
                            column_usr = cell.column
                            lun = len(details1[cell.column])
                users = [b.value for b in details1[column_usr][row_tb:lun]]
                for row in details1.iter_rows():
                    for cell in row:
                        if cell.value == "Debit/Credit":
                            row_tb = cell.row
                            column_usr = cell.column
                            lun = len(details1[cell.column])
                dbcd = [b.value for b in details1[column_usr][row_tb:lun]]
                amountDb=[]
                amountCr=[]

                for c in range(0, len(amount)):
                    if amount[c] <= 0:
                        amountCr.append(amount[c])
                    else:
                        amountCr.append(0)

                for c in range(0, len(amount)):
                    if amount[c] >0:
                        amountDb.append(amount[c])
                    else:
                        amountDb.append(0)

                typeOfTr=[]
                for i in range(0, len(amount)):
                    if amount[i] > 0:
                        typeOfTr.append("Debit")
                    else:
                        typeOfTr.append("Credit")
                # amountss=amount
                typeAccount=typeOfTr
                    
            except:
                try:
                    for row in details1.iter_rows():
                        for cell in row:
                            if cell.value == "Account":
                                row_tb = cell.row
                                column_AmC = cell.column
                                lun = len(details1[cell.column])
                    account = [b.value for b in details1[column_AmC][row_tb:lun]]

                    for row in details1.iter_rows():
                        for cell in row:
                            if cell.value == "JE Number":
                                row_tb = cell.row
                                column_je = cell.column
                                lun = len(details1[cell.column])
                    jeNo = [b.value for b in details1[column_je][row_tb:lun]]

                    jeNoo=[]
                    for i in range(0, len(jeNo)):
                        if jeNo[i] == None:
                            jeNoo.append(0)
                        else:
                            jeNoo.append(jeNo[i])

                    for row in details1.iter_rows():
                        for cell in row:
                            if cell.value == "Posting Date":
                                row_tb = cell.row
                                column_date = cell.column
                                lun = len(details1[cell.column])
                    postDate = [b.value for b in details1[column_date][row_tb:lun]]

                    for row in details1.iter_rows():
                        for cell in row:
                            if cell.value == "Description":
                                row_tb = cell.row
                                column_desc = cell.column
                                lun = len(details1[cell.column])
                    description = [b.value for b in details1[column_desc][row_tb:lun]]

                    descriptionFull=[]
                    for c in range(0, len(description)):
                        if description[c] == None:
                            descriptionFull.append("Nu exista descriere")
                        else:
                            descriptionFull.append(description[c])

                    for row in details1.iter_rows():
                        for cell in row:
                            if cell.value == "Amount Debit":
                                row_tb = cell.row
                                column_amDB = cell.column
                                lun = len(details1[cell.column])
                    amDeb = [b.value for b in details1[column_amDB][row_tb:lun]]
                    

                    for row in details1.iter_rows():
                        for cell in row:
                            if cell.value == "Amount Credit":
                                row_tb = cell.row
                                column_amCR = cell.column
                                lun = len(details1[cell.column])
                    amCre = [b.value for b in details1[column_amCR][row_tb:lun]]

                    amount=[]
                    typeOfTr=[]
                    for i in range(0, len(amDeb)):    
                        if amDeb[i] == None or amDeb[i] == 0:
                            amount.append(amCre[i])
                            typeOfTr.append("Credit")
                        else:
                            amount.append(amDeb[i])
                            typeOfTr.append("Debit")
                    # print(typeOfTr)
                    # amountss=amount
                    typeAccount=typeOfTr

                    for row in details1.iter_rows():
                        for cell in row:
                            if cell.value == "User":
                                row_tb = cell.row
                                column_usr = cell.column
                                lun = len(details1[cell.column])
                    users = [b.value for b in details1[column_usr][row_tb:lun]]
                    for row in details1.iter_rows():
                        for cell in row:
                            if cell.value == "Debit/Credit":
                                row_tb = cell.row
                                column_usr = cell.column
                                lun = len(details1[cell.column])
                    dbcd = [b.value for b in details1[column_usr][row_tb:lun]]
                    amountDb=[]
                    amountCr=[]

                    for c in range(0, len(amount)):
                        if amount[c] <= 0:
                            amountCr.append(amount[c])
                        else:
                            amountCr.append(0)

                    for c in range(0, len(amount)):
                        if amount[c] >0:
                            amountDb.append(amount[c])
                        else:
                            amountDb.append(0)
                    
                except:
                    # pass
                    for row in details1.iter_rows():
                        for cell in row:
                            if cell.value == "Amount":
                                row_tb = cell.row
                                column_AmD = cell.column
                                lun = len(details1[cell.column])
                    amount = [b.value for b in details1[column_AmD][row_tb:lun]]
                    

                    # suma=sum(amount)
                    # print("suma este ",suma)

                    for row in details1.iter_rows():
                        for cell in row:
                            if cell.value == "Account Debit":
                                row_tb = cell.row
                                column_acDB = cell.column
                                lun = len(details1[cell.column])
                    acDeb = [b.value for b in details1[column_acDB][row_tb:lun]]
                    typeOTDB=[]
                    for i in range(0, len(acDeb)):
                        typeOTDB.append("Debit")

                    for row in details1.iter_rows():
                        for cell in row:
                            if cell.value == "Account Credit":
                                row_tb = cell.row
                                column_acCR = cell.column
                                lun = len(details1[cell.column])
                    acCre = [b.value for b in details1[column_acCR][row_tb:lun]]
                    typeOTCr=[]
                    for i in range(0, len(acCre)):
                        typeOTCr.append("Credit")

                    account=acDeb+acCre
                    typeAccount=typeOTDB+typeOTCr

                    for row in details1.iter_rows():
                        for cell in row:
                            if cell.value == "JE Number":
                                row_tb = cell.row
                                column_je = cell.column
                                lun = len(details1[cell.column])
                    jeNo = [b.value for b in details1[column_je][row_tb:lun]]

                    jeNoo=[]
                    for i in range(0, len(jeNo)):
                        if jeNo[i] == None:
                            jeNoo.append(0)
                        else:
                            jeNoo.append(jeNo[i])
                    print(len(jeNo), "jeNo")
                    print(len(jeNoo), "jeNoo")

                    for row in details1.iter_rows():
                        for cell in row:
                            if cell.value == "Posting Date":
                                row_tb = cell.row
                                column_date = cell.column
                                lun = len(details1[cell.column])
                    postDate = [b.value for b in details1[column_date][row_tb:lun]]

                    for row in details1.iter_rows():
                        for cell in row:
                            if cell.value == "Description":
                                row_tb = cell.row
                                column_desc = cell.column
                                lun = len(details1[cell.column])
                    description = [b.value for b in details1[column_desc][row_tb:lun]]

                    descriptionFull=[]
                    for c in range(0, len(description)):
                        if description[c] == None:
                            descriptionFull.append("Nu exista descriere")
                        else:
                            descriptionFull.append(description[c])

                    for row in details1.iter_rows():
                        for cell in row:
                            if cell.value == "User":
                                row_tb = cell.row
                                column_usr = cell.column
                                lun = len(details1[cell.column])
                    users = [b.value for b in details1[column_usr][row_tb:lun]]
                    for row in details1.iter_rows():
                        for cell in row:
                            if cell.value == "Debit/Credit":
                                row_tb = cell.row
                                column_usr = cell.column
                                lun = len(details1[cell.column])
                    dbcd = [b.value for b in details1[column_usr][row_tb:lun]]
                    amountsDeb=[]
                    amountsCre=[]
                    jeNoje=[]
                    descriptionFullje=[]
                    postDateje=[]
                    usersJe=[]
                    typeOfTr=[]
                    typeOfTrDb=[]
                    typeOfTrCr=[]

                    for i in range(0, len(acDeb)):
                        amountsDeb.append(amount[i])
                        # suma3=sum(amountsDeb)                        
                        descriptionFullje.append(descriptionFull[i])
                        # jeNoje.append(jeNoo[i])
                        jeNoje.append(jeNo[i])
                        postDateje.append(postDate[i])
                        usersJe.append(users[i])
                        typeOfTrDb.append("Debit")

                    for x in range(0, len(acCre)):
                        
                        if amount[x] < 0:
                            descriptionFullje.append(descriptionFull[x])
                            amountsCre.append(amount[x])
                            typeOfTrCr.append("Credit")
                            # sum4=sum(amountsCre)
                            # jeNoje.append(jeNoo[x])
                            jeNoje.append(jeNo[x])
                            postDateje.append(postDate[x])
                            usersJe.append(users[x])
                        
                    
                    # typeOfTr = typeOfTrDb + typeOfTrCr
                    
                    # amountss=amountsDeb + amountsCre
                    # amount=amountss
                    
                    jeNo=jeNoje
                    # postDate=postDateje
                    # descriptionFull=descriptionFullje
                    # users=usersJe
                    
                    amountDb=[]
                    amountCr=[]

                    for c in range(0, len(amount)):
                        if amount[c] < 0:
                            amountCr.append(amount[c])
                        else:
                            amountCr.append(0)

                    for c in range(0, len(amount)):
                        if amount[c] >=0:
                            amountDb.append(amount[c])
                        else:
                            amountDb.append(0)
            

        for i in range(0,len(jeNoo)):
            je.cell(row=9+i, column=1).value=jeNoo[i]
        
        # for i in range(0,len(jeNo)):
        #     je.cell(row=9+i, column=1).value=jeNo[i]

        for i in range(0,len(postDate)):
            je.cell(row=9+i, column=2).value=postDate[i]

        for i in range(0,len(descriptionFull)):
            je.cell(row=9+i, column=3).value=descriptionFull[i]

        for i in range(0,len(amount)):
            je.cell(row=9+i, column=4).value=amount[i]
            
        for i in range(0,len(jeNoo)):
            je.cell(row=9+i, column=5).value=account[i]

        for i in range(0,len(users)):
            je.cell(row=9+i, column=6).value=users[i]

        for i in range(0,len(dbcd)):
            je.cell(row=9+i, column=7).value=dbcd[i]

        # try:
        #     for i in range(0,len(users)):
        #         je.cell(row=9+i, column=7).value=amountDb[i]

        #     for i in range(0,len(users)):
        #         je.cell(row=9+i, column=8).value=amountCr[i]
        # except:
        #     for i in range(0,len(users)):
        #         je.cell(row=9+i, column=7).value=flag[i]

        #--------------------------------------------FORMATS AND FONTS--------------------------------------------------------

        for row in je["A8:G8"]:
            for cell in row:
                # cell.number_format = '#,##0_);(#,##0)'
                cell.font = cap_tabel
                cell.fill = cap_tabel_color_GT
        
        # for i in range(0,len(jeNoo)):
        for i in range(0,len(jeNo)):
            je.cell(row=9+i, column=2).number_format='MM/DD/YYYY'
            je.cell(row=9+i, column=4).number_format='#,##0_);(#,##0)'
            je.cell(row=9+i, column=7).number_format='#,##0_);(#,##0)'
            je.cell(row=9+i, column=8).number_format='#,##0_);(#,##0)'

        listaColoane=['A','B','C','D','E','G','H']
        for column in ascii_uppercase:
            for i in listaColoane:
                if (column==i):
                    je.column_dimensions[column].width = 12
        print("a facut JE")

        #---------------------------------------GREATER THAN PM---------------------------------------------------------------
        if(largeAM==1):
            greaterPM = wb.create_sheet("Large_Amounts")
            greaterPM.sheet_view.showGridLines = False

            summary.cell(row=15, column=3).value="Large Amounts"
            summary.cell(row=15, column=3).hyperlink='#Large_Amounts!A1'
            summary.cell(row=15, column=3).font=hyperlink_font

            greaterPM.cell(row=1, column=1).value="Client:"
            greaterPM.cell(row=2, column=1).value="Period end:"

            greaterPM.cell(row=1, column=2).value=clientname1
            try:
                greaterPM.cell(row=2, column=2).value=yearEnd1
            except:
                pass
            greaterPM.cell(row=5, column=1).value="Large Amounts"

            greaterPM.cell(row=7, column=1).value="JE Number"
            greaterPM.cell(row=7, column=2).value="Posting date"
            greaterPM.cell(row=7, column=3).value="Description"
            greaterPM.cell(row=7, column=5).value="Amount"
            greaterPM.cell(row=7, column=4).value="Account"

            # greaterPM.cell(row=2, column=10).value="Total records:"
            # greaterPM.cell(row=3, column=10).value="Total amount:"

            amounts=[]
            jeNo1=[]
            no1=[]
            postDate1=[]
            docDate1=[]
            descript1=[]
            syntD=[]
            acct=[]
            syntC=[]
            acCred=[]

            absAmount=[abs(number) for number in amount]

            for i in range (0, len(amount)):
                if amount[i] > PM or amount[i]< -(PM):
                    amounts.append(amount[i])
                    try:
                        # jeNo1.append(jeNo[i])
                        jeNo1.append(jeNoo[i])
                    except:
                        None
                    postDate1.append(postDate[i])
                    descript1.append(descriptionFull[i])
                    acct.append(account[i])

            for k in range(0, len(amounts)):
                try:
                    greaterPM.cell(row=8+k, column=1).value=jeNo1[k]
                except:
                    None
                greaterPM.cell(row=8+k, column=2).value=postDate1[k]
                greaterPM.cell(row=8+k, column=3).value=descript1[k]
                greaterPM.cell(row=8+k, column=5).value=amounts[k]
                greaterPM.cell(row=8+k, column=4).value=acct[k]
                # greaterPM.cell(row=8+k, column=6).value=typeAccount[k]

            # sumElements=sum(amounts)
            # greaterPM.cell(row=2, column=11).value=str(len(amounts))
            # greaterPM.cell(row=3, column=11).value=sumElements

            #------------------------------------------------------FORMATS AND FONTS-----------------------------------------

            greaterPM.cell(row=1, column=1).font=ft1
            greaterPM.cell(row=2, column=1).font=ft1

            greaterPM.cell(row=1, column=2).font=font_worksheet
            greaterPM.cell(row=2, column=2).font=font_worksheet

            greaterPM.cell(row=5, column=1).font=f_testname

            for row in greaterPM["A7:E7"]:
                for cell in row:
                    cell.font = cap_tabel
                    cell.fill = cap_tabel_color_GT
            
            for k in range(0, len(amounts)):
                        greaterPM.cell(row=8+k, column=2).number_format='MM/DD/YYYY'
                        greaterPM.cell(row=8+k, column=5).number_format='#,##0_);(#,##0)'

            for k in range(0, len(amounts)):
                        greaterPM.cell(row=8+k, column=1).font=font_worksheet
                        greaterPM.cell(row=8+k, column=2).font=font_worksheet
                        greaterPM.cell(row=8+k, column=3).font=font_worksheet
                        greaterPM.cell(row=8+k, column=5).font=font_worksheet
                        greaterPM.cell(row=8+k, column=4).font=font_worksheet


            listaColoane=['A','B','C','D','E']
            for column in ascii_uppercase:
                for i in listaColoane:
                    if (column==i):
                        greaterPM.column_dimensions[column].width = 12
            greaterPM.column_dimensions["J"].width = 12
            greaterPM.column_dimensions["K"].width = 12
            greaterPM.cell(row=3, column=11).number_format='#,##0_);(#,##0)'

            print("greater PM")

            #---------------------------------------GREATER PM PIVOT TABLE-----------------------------------------------

            # greaterPMpivot = wb.create_sheet("1. Pivot greater than PM")

        #---------------------------------------WEEKEND TRANSACTIONS-----------------------------------------------
        if(poWE==1):
            # try:
            weekendPost = wb.create_sheet("Posted_in_weekend")
            weekendPost.sheet_view.showGridLines = False
            summary.cell(row=16, column=3).value="Posted in weekend"
            summary.cell(row=16, column=3).hyperlink='#Posted_in_weekend!A1'
            summary.cell(row=16, column=3).font=hyperlink_font

            weekDays = {0:"Monday",1:"Tuesday",2:"Wednesday",3:"Thursday",4:"Friday",5:"Saturday",6:"Sunday"}

            weekendPost.cell(row=1, column=1).value="Client name:"
            weekendPost.cell(row=2, column=1).value="Period end:"

            weekendPost.cell(row=1, column=2).value=clientname1
            weekendPost.cell(row=2, column=2).value=yearEnd1
            
            weekendPost.cell(row=5, column=1).value="Transactions posted on weekend"

            weekendPost.cell(row=9, column=1).value="JE Number"
            weekendPost.cell(row=9, column=2).value="Posting Date"
            weekendPost.cell(row=9, column=3).value="Description"
            weekendPost.cell(row=9, column=4).value="Account"
            weekendPost.cell(row=9, column=5).value="Amount"

            weekendPost.cell(row=3, column=7).value="Total records"
            weekendPost.cell(row=4, column=7).value="Total amount"

            jeNo2=[]
            postDate2=[]
            descript2=[]
            acct2=[]
            amount2=[]
            # print(postDate)
            for i in range(0, len(jeNoo)):
                date=postDate[i]
                date1=date.weekday()
                if date1 in weekDays:
                    if date1 == 5 or date1 == 6:
                        jeNo2.append(jeNoo[i])
                        postDate2.append(postDate[i])
                        descript2.append(descriptionFull[i])
                        acct2.append(account[i])
                        amount2.append(amount[i])
            print("jeNo2", len(jeNo2))
            for k in range(0, len(postDate2)):
                weekendPost.cell(row=10+k, column=1).value=jeNo2[k]
                weekendPost.cell(row=10+k, column=2).value=postDate2[k]
                weekendPost.cell(row=10+k, column=3).value=descript2[k]
                weekendPost.cell(row=10+k, column=4).value=acct2[k]
                weekendPost.cell(row=10+k, column=5).value=amount2[k]

            sum2=sum(amount2)
            weekendPost.cell(row=3, column=8).value=str(len(postDate2))
            weekendPost.cell(row=4, column=8).value=sum2

            print("posted weekend")

            #---------------------------------------------FORMATS AND COLORS--------------------------

            weekendPost.cell(row=1, column=1).font=ft1
            weekendPost.cell(row=2, column=1).font=ft1

            weekendPost.cell(row=1, column=2).font=font_worksheet
            weekendPost.cell(row=2, column=2).font=font_worksheet

            weekendPost.cell(row=5, column=1).font=f_testname

            for row in weekendPost["A9:E9"]:
                for cell in row:
                    cell.font = cap_tabel
                    cell.fill = cap_tabel_color_GT
            
            for k in range(0, len(amount2)):
                weekendPost.cell(row=10+k, column=2).number_format='MM/DD/YYYY'
                weekendPost.cell(row=10+k, column=5).number_format='#,##0_);(#,##0)'

            for k in range(0, len(amount2)):
                weekendPost.cell(row=10+k, column=1).font=font_worksheet
                weekendPost.cell(row=10+k, column=2).font=font_worksheet
                weekendPost.cell(row=10+k, column=3).font=font_worksheet
                weekendPost.cell(row=10+k, column=5).font=font_worksheet
                weekendPost.cell(row=10+k, column=4).font=font_worksheet


            listaColoane=['A','B','C','D','E']
            for column in ascii_uppercase:
                for i in listaColoane:
                    if (column==i):
                        weekendPost.column_dimensions[column].width = 12

            weekendPost.column_dimensions['G'].width = 12
            weekendPost.column_dimensions['H'].width = 12
            weekendPost.cell(row=4, column=8).number_format='#,##0_);(#,##0)'        
            # except:
            #     pass

            #---------------------------------------PIVOT POSTED IN WEEKEND------------------------------------------------
            # weekendPostPivot = wb.create_sheet("2. Pivot posted in weekend")

        #----------------------------------------------- MISSING EXPLANATION------------------------------------------
        if(missExplanation==1):
            try:
                missExp = wb.create_sheet("Missing_explanation")
                missExp.sheet_view.showGridLines = False
                summary.cell(row=17, column=3).value="Missing explanation"
                summary.cell(row=17, column=3).hyperlink='#Missing_explanation!A1'
                summary.cell(row=17, column=3).font=hyperlink_font

                missExp.cell(row=1, column=1).value="Client name:"
                missExp.cell(row=2, column=1).value="Period end:"

                missExp.cell(row=1, column=2).value=clientname1
                try:
                    missExp.cell(row=2, column=2).value=yearEnd1
                except:
                    pass

                missExp.cell(row=5, column=1).value="Missing explanation"

                missExp.cell(row=3, column=7).value="Total records"
                missExp.cell(row=4, column=7).value="Total amount"

                missExp.cell(row=9, column=1).value="JE Number"
                missExp.cell(row=9, column=2).value="Posting Date"
                missExp.cell(row=9, column=3).value="Description"
                missExp.cell(row=9, column=4).value="Account"
                missExp.cell(row=9, column=5).value="Amount"

                jeNo3=[]
                postDate3=[]
                descript3=[]
                acct3=[]
                amount3=[]

                for i in range(0, len(descriptionFull)):
                    if descriptionFull[i] == "Nu exista descriere":
                        jeNo3.append(jeNo[i])
                        postDate3.append(postDate[i])
                        descript3.append(descriptionFull[i])
                        acct3.append(account[i])
                        amount3.append(amount[i])

                for k in range(0, len(descript3)):
                    missExp.cell(row=10+k, column=1).value=jeNo3[k]
                    missExp.cell(row=10+k, column=2).value=postDate3[k]
                    missExp.cell(row=10+k, column=3).value=descript3[k]
                    missExp.cell(row=10+k, column=4).value=acct3[k]
                    missExp.cell(row=10+k, column=5).value=amount3[k]
                sum3=sum(amount3)
                missExp.cell(row=3, column=8).value=str(len(descript3))
                missExp.cell(row=4, column=8).value=sum3

                print("missing exp")

                #-----------------------------------------FORMATS AND COLORS------------------------------------------------

                missExp.cell(row=1, column=1).font=ft1
                missExp.cell(row=2, column=1).font=ft1

                missExp.cell(row=1, column=2).font=font_worksheet
                missExp.cell(row=2, column=2).font=font_worksheet

                missExp.cell(row=5, column=1).font=f_testname

                for row in missExp["A9:E9"]:
                    for cell in row:
                        cell.font = cap_tabel
                        cell.fill = cap_tabel_color_GT
                
                for k in range(0, len(amounts)):
                    missExp.cell(row=10+k, column=2).number_format='MM/DD/YYYY'
                    missExp.cell(row=10+k, column=5).number_format='#,##0_);(#,##0)'

                for k in range(0, len(amounts)):
                    missExp.cell(row=10+k, column=1).font=font_worksheet
                    missExp.cell(row=10+k, column=2).font=font_worksheet
                    missExp.cell(row=10+k, column=3).font=font_worksheet
                    missExp.cell(row=10+k, column=5).font=font_worksheet
                    missExp.cell(row=10+k, column=4).font=font_worksheet


                listaColoane=['A','B','C','D','E']
                for column in ascii_uppercase:
                    for i in listaColoane:
                        if (column==i):
                            missExp.column_dimensions[column].width = 12

                missExp.column_dimensions['G'].width = 12
                missExp.column_dimensions['H'].width = 12
                missExp.cell(row=4, column=8).number_format='#,##0_);(#,##0)'  
            except:
                pass

        #---------------------------------------CORRECTION FRAUD SHEET------------------------------------------------
        if(speCom==1):
            try:
                corrFraud = wb.create_sheet("Correction_or_fraud")
                corrFraud.sheet_view.showGridLines = False
                summary.cell(row=18, column=3).value='Correction or fraud'
                summary.cell(row=18, column=3).hyperlink='#Correction_or_fraud!A1'
                summary.cell(row=18, column=3).font=hyperlink_font 
            
                corrFraud.cell(row=1, column=1).value="Client name:"
                corrFraud.cell(row=2, column=1).value="Period end:"

                corrFraud.cell(row=1, column=2).value=clientname1
                try:
                    corrFraud.cell(row=2, column=2).value=yearEnd1
                except:
                    pass

                corrFraud.cell(row=3, column=10).value="Total records:"
                corrFraud.cell(row=4, column=10).value="Total amount:"
                corrFraud.cell(row=5, column=1).value="Correction or fraud"

                corrFraud.cell(row=9, column=1).value="JE Number"
                corrFraud.cell(row=9, column=2).value="Posting Date"
                corrFraud.cell(row=9, column=3).value="Description"
                corrFraud.cell(row=9, column=4).value="Account"
                corrFraud.cell(row=9, column=5).value="Amount"

                jeNo4=[]
                postDate4=[]
                descript4=[]
                acct4=[]
                amount4=[]

                
                for j in range(0, len(corFrd)):
                    for i in range(0, len(descriptionFull)):
                        # if "corectie" in descriptionFull[i] or "correction" in descriptionFull[i] or "fraud" in descriptionFull[i] or "frauda" in descriptionFull[i]:
                        #eroare prost 1
                        if corFrd[j] == descriptionFull[i]:
                            # print("avem tranzactii suspecte")
                            jeNo4.append(jeNo[i])
                            postDate4.append(postDate[i])
                            descript4.append(descriptionFull[i])
                            acct4.append(account[i])
                            amount4.append(amount[i])

                for k in range(0, len(descript4)):
                    corrFraud.cell(row=10+k, column=1).value=jeNo4[k]
                    corrFraud.cell(row=10+k, column=2).value=postDate4[k]
                    corrFraud.cell(row=10+k, column=3).value=descript4[k]
                    corrFraud.cell(row=10+k, column=4).value=acct4[k]
                    corrFraud.cell(row=10+k, column=5).value=amount4[k]

                sumCorr=sum(amount4)

                corrFraud.cell(row=3, column=11).value=str(len(descript4))
                corrFraud.cell(row=4, column=11).value=sumCorr

                print("correction or fraud")
            
                #-----------------------------------------FORMATS AND COLORS------------------------------------------------

                corrFraud.cell(row=1, column=1).font=ft1
                corrFraud.cell(row=2, column=1).font=ft1

                corrFraud.cell(row=1, column=2).font=font_worksheet
                corrFraud.cell(row=2, column=2).font=font_worksheet

                corrFraud.cell(row=5, column=1).font=f_testname

                for row in corrFraud["A9:E9"]:
                    for cell in row:
                        cell.font = cap_tabel
                        cell.fill = cap_tabel_color_GT
                
                for k in range(0, len(amounts)):
                    corrFraud.cell(row=10+k, column=2).number_format='MM/DD/YYYY'
                    corrFraud.cell(row=10+k, column=5).number_format='#,##0_);(#,##0)'

                for k in range(0, len(amounts)):
                    corrFraud.cell(row=10+k, column=1).font=font_worksheet
                    corrFraud.cell(row=10+k, column=2).font=font_worksheet
                    corrFraud.cell(row=10+k, column=3).font=font_worksheet
                    corrFraud.cell(row=10+k, column=5).font=font_worksheet
                    corrFraud.cell(row=10+k, column=4).font=font_worksheet


                listaColoane=['A','B','C','D','E']
                for column in ascii_uppercase:
                    for i in listaColoane:
                        if (column==i):
                            corrFraud.column_dimensions[column].width = 12

                corrFraud.column_dimensions['J'].width = 12
                corrFraud.column_dimensions['K'].width = 12
                corrFraud.cell(row=4, column=11).number_format='#,##0_);(#,##0)' 
            except:
                pass

        #--------------------------------------- PIVOT CORRECTION FRAUD SHEET------------------------------------------------
        # corrFraudPivot = wb.create_sheet("4. Pivot correction or fraud")

        #---------------------------------------MANAGEMENT SHEET------------------------------------------
        # management = wb.create_sheet("5. Management")
        # management.sheet_view.showGridLines = False

        # management.cell(row=1, column=1).value="Client name:"
        # management.cell(row=2, column=1).value="Period end:"

        # management.cell(row=1, column=2).value=clientname1
        # management.cell(row=2, column=2).value=yearEnd1

        # management.cell(row=5, column=1).value="5. Management Transactions"

        # management.cell(row=9, column=1).value="JE Number"
        # management.cell(row=9, column=2).value="Posting Date"
        # management.cell(row=9, column=3).value="Description"
        # management.cell(row=9, column=4).value="Account"
        # management.cell(row=9, column=5).value="Amount"

        # managementList=['Soare', 'Alina', 'Cristi', 'Costica']
        # # managementList = manag
        # print(managementList)
        # jeNo5=[]
        # postDate5=[]
        # descript5=[]
        # acct5=[]
        # amount5=[]
        # users5=[]

        # for i in range(0, len(managementList)):
        #     # print(managementList[i])
        #     for k in range(0, len(users)):
        #         if managementList[i] == users[k]:
        #             # print("avem management")
        #             jeNo5.append(jeNo[k])
        #             postDate5.append(postDate[k])
        #             descript5.append(description[k])
        #             acct5.append(account[k])
        #             amount5.append(amount[k])
        #             users5.append(users[k])

        # # print(users5)

        # for k in range(0, len(users5)):
        #     management.cell(row=10+k, column=1).value=jeNo5[k]
        #     management.cell(row=10+k, column=2).value=postDate5[k]
        #     management.cell(row=10+k, column=3).value=descript5[k]
        #     management.cell(row=10+k, column=4).value=acct5[k]
        #     management.cell(row=10+k, column=5).value=amount5[k]
        #     management.cell(row=10+k, column=6).value=users5[k]



        #---------------------------------------ENDING 000 --------------------------------------------
        if(endZero==1):
            try:
                endingZero = wb.create_sheet("Ending_in_000")
                endingZero.sheet_view.showGridLines = False
                summary.cell(row=19, column=3).value='Ending in 000'
                summary.cell(row=19, column=3).hyperlink='#Ending_in_000!A1'
                summary.cell(row=19, column=3).font=hyperlink_font 
            
                endingZero.cell(row=1, column=1).value="Client name:"
                endingZero.cell(row=2, column=1).value="Period end:"

                endingZero.cell(row=1, column=2).value=clientname1
                try:
                    endingZero.cell(row=2, column=2).value=yearEnd1
                except:
                    pass

                endingZero.cell(row=5, column=1).value="Transactions Ending in 000"

                endingZero.cell(row=9, column=1).value="JE Number"
                endingZero.cell(row=9, column=2).value="Posting Date"
                endingZero.cell(row=9, column=3).value="Description"
                endingZero.cell(row=9, column=4).value="Account"
                endingZero.cell(row=9, column=5).value="Amount"
                endingZero.cell(row=9, column=6).value="User"

                # intAmount=[int(number) for number in amount]
                # strAmount=[str(no) for no in intAmount]

                jeNo6=[]
                postDate6=[]
                descript6=[]
                acct6=[]
                amount6=[]
                users6=[]

                # for i in range(0, len(strAmount)):
                #     if strAmount[i][-3:] == "000":
                #         intAmt=[int(float(no1)) for no1 in strAmount]
                #         jeNo6.append(jeNo[i])
                #         postDate6.append(postDate[i])
                #         descript6.append(descriptionFull[i])
                #         acct6.append(account[i])
                #         amount6.append(intAmt[i])

                for i in range(0, len(amount)):
                    if amount[i] % 1000 == 0 and amount[i] != 0:
                        jeNo6.append(jeNo[i])
                        postDate6.append(postDate[i])
                        descript6.append(descriptionFull[i])
                        acct6.append(account[i])
                        amount6.append(amount[i])
                        users6.append(users[i])

                for k in range(0, len(amount6)):
                    endingZero.cell(row=10+k, column=1).value=jeNo6[k]
                    endingZero.cell(row=10+k, column=2).value=postDate6[k]
                    endingZero.cell(row=10+k, column=3).value=descript6[k]
                    endingZero.cell(row=10+k, column=4).value=acct6[k]
                    endingZero.cell(row=10+k, column=5).value=amount6[k]
                    endingZero.cell(row=10+k, column=6).value=users6[k]

                print("ending zero")

                #-----------------------------------------FORMATS AND COLORS------------------------------------------------

                endingZero.cell(row=1, column=1).font=ft1
                endingZero.cell(row=2, column=1).font=ft1

                endingZero.cell(row=1, column=2).font=font_worksheet
                endingZero.cell(row=2, column=2).font=font_worksheet

                endingZero.cell(row=5, column=1).font=f_testname

                for row in endingZero["A9:F9"]:
                    for cell in row:
                        cell.font = cap_tabel
                        cell.fill = cap_tabel_color_GT
                
                for k in range(0, len(amounts)):
                    endingZero.cell(row=10+k, column=2).number_format='MM/DD/YYYY'
                    endingZero.cell(row=10+k, column=5).number_format='#,##0_);(#,##0)'

                for k in range(0, len(amounts)):
                    endingZero.cell(row=10+k, column=1).font=font_worksheet
                    endingZero.cell(row=10+k, column=2).font=font_worksheet
                    endingZero.cell(row=10+k, column=3).font=font_worksheet
                    endingZero.cell(row=10+k, column=5).font=font_worksheet
                    endingZero.cell(row=10+k, column=4).font=font_worksheet
                    endingZero.cell(row=10+k, column=6).font=font_worksheet


                listaColoane=['A','B','C','D','E']
                for column in ascii_uppercase:
                    for i in listaColoane:
                        if (column==i):
                            endingZero.column_dimensions[column].width = 12
            except:
                pass

            # corrFraud.column_dimensions['J'].width = 12
            # corrFraud.column_dimensions['K'].width = 12

            #--------------------------------------- PIVOT ENDING 000 --------------------------------------------
            # endingPivot = wb.create_sheet("6. Pivot ending in 000")

        #-------------------------------------------------------ENDING 999------------------------------------
        if(endNine==1):
            try:
                endingNine = wb.create_sheet("Ending_in_999")
                endingNine.sheet_view.showGridLines = False
                summary.cell(row=20, column=3).value="Ending in 999"
                summary.cell(row=20, column=3).hyperlink='#Ending_in_999!A1'
                summary.cell(row=20, column=3).font=hyperlink_font
            
                endingNine.cell(row=1, column=1).value="Client name:"
                endingNine.cell(row=2, column=1).value="Period end:"

                endingNine.cell(row=1, column=2).value=clientname1
                try:
                    endingNine.cell(row=2, column=2).value=yearEnd1
                except:
                    pass

                endingNine.cell(row=5, column=1).value="Transactions Ending in 999"

                endingNine.cell(row=9, column=1).value="JE Number"
                endingNine.cell(row=9, column=2).value="Posting Date"
                endingNine.cell(row=9, column=3).value="Description"
                endingNine.cell(row=9, column=4).value="Account"
                endingNine.cell(row=9, column=5).value="Amount"
                endingNine.cell(row=9, column=6).value="User"

                intAmount=[int(number) for number in amount]
                strAmount=[str(no) for no in intAmount]

                jeNo7=[]
                postDate7=[]
                descript7=[]
                acct7=[]
                amount7=[]
                users7=[]

                for i in range(0, len(strAmount)):
                    if strAmount[i][-3:] == "999":
                        intAmt=[int(no2) for no2 in strAmount]
                        jeNo7.append(jeNo[i])
                        postDate7.append(postDate[i])
                        descript7.append(descriptionFull[i])
                        acct7.append(account[i])
                        amount7.append(intAmt[i])
                        users7.append(users[i])

                for k in range(0, len(amount7)):
                    endingNine.cell(row=10+k, column=1).value=jeNo7[k]
                    endingNine.cell(row=10+k, column=2).value=postDate7[k]
                    endingNine.cell(row=10+k, column=3).value=descript7[k]
                    endingNine.cell(row=10+k, column=4).value=acct7[k]
                    endingNine.cell(row=10+k, column=5).value=amount7[k]
                    endingNine.cell(row=10+k, column=6).value=users7[k]
                print("ending 9")

                #-----------------------------------------FORMATS AND COLORS------------------------------------------------

                endingNine.cell(row=1, column=1).font=ft1
                endingNine.cell(row=2, column=1).font=ft1

                endingNine.cell(row=1, column=2).font=font_worksheet
                endingNine.cell(row=2, column=2).font=font_worksheet

                endingNine.cell(row=5, column=1).font=f_testname

                for row in endingNine["A9:F9"]:
                    for cell in row:
                        cell.font = cap_tabel
                        cell.fill = cap_tabel_color_GT
                
                for k in range(0, len(amounts)):
                    endingNine.cell(row=10+k, column=2).number_format='MM/DD/YYYY'
                    endingNine.cell(row=10+k, column=5).number_format='#,##0_);(#,##0)'

                for k in range(0, len(amounts)):
                    endingNine.cell(row=10+k, column=1).font=font_worksheet
                    endingNine.cell(row=10+k, column=2).font=font_worksheet
                    endingNine.cell(row=10+k, column=3).font=font_worksheet
                    endingNine.cell(row=10+k, column=5).font=font_worksheet
                    endingNine.cell(row=10+k, column=4).font=font_worksheet
                    endingNine.cell(row=10+k, column=6).font=font_worksheet


                listaColoane=['A','B','C','D','E']
                for column in ascii_uppercase:
                    for i in listaColoane:
                        if (column==i):
                            endingNine.column_dimensions[column].width = 12

                # corrFraud.column_dimensions['J'].width = 12
                # corrFraud.column_dimensions['K'].width = 12
            except:
                pass

        #--------------------------------------- TRANSACTIONS BY USERS --------------------------------------------
        if(poUSER==1):
            try:
                userSheet = wb.create_sheet("Transactions_by_users")
                userSheet.sheet_view.showGridLines = False
                summary.cell(row=21, column=3).value="Transactions by users"
                summary.cell(row=21, column=3).hyperlink='#Transactions_by_users!A1'
                summary.cell(row=21, column=3).font=hyperlink_font

                userSheet.cell(row=1, column=1).value="Client name:"
                userSheet.cell(row=2, column=1).value="Period end:"

                userSheet.cell(row=1, column=2).value=clientname1
                try:
                    userSheet.cell(row=2, column=2).value=yearEnd1
                except:
                    pass

                userSheet.cell(row=3, column=8).value="Total number of users:"
                userSheet.cell(row=4, column=8).value="Total amount Debit:"
                userSheet.cell(row=5, column=8).value="Total amount Credit:"

                userSheet.cell(row=5, column=1).value="Transactions by users"

                userSheet.cell(row=9, column=1).value="User"
                userSheet.cell(row=9, column=2).value="Number of transactions"
                userSheet.cell(row=9, column=3).value="Debit Amount"
                userSheet.cell(row=9, column=4).value="Credit Amount"

                listaUseri=set(users)
                listaUnicaUseri=list(listaUseri)
                # print(listaUnicaUseri)
                listaJE=set(jeNo)
                listaUnicaJE=list(listaJE)

                # print(listaUnicaJE)
                # print(len(listaUnicaJE))

                users8=[]

                for i in range(0, len(listaUnicaUseri)):
                    users8.append(listaUnicaUseri[i])
                    userSheet.cell(row=10+i, column=1).value=users8[i]
                    dataAparitie=users.count(listaUnicaUseri[i])
                    userSheet.cell(row=10+i, column=2).value=dataAparitie

                sumaUseriDb=[]
                for p in range(0,len(users8)):
                    count=0
                    for k in range(0,len(users)):
                        if(users8[p] == users[k]):
                            count=count+amountDb[k]
                    sumaUseriDb.append(count)

                sumaUseriCr=[]
                for p in range(0,len(users8)):
                    count=0
                    for k in range(0,len(users)):
                        if(users8[p] == users[k]):
                            count=count+amountCr[k]
                    sumaUseriCr.append(count)
                # print(users7, sumaUseriDb)
                for i in range(0, len(listaUnicaUseri)):
                    userSheet.cell(row=10+i, column=1).value=users8[i]
                    userSheet.cell(row=10+i, column=3).value=sumaUseriDb[i]
                    userSheet.cell(row=10+i, column=4).value=sumaUseriCr[i]

                sum8Db=sum(sumaUseriDb)
                sum8Cr=sum(sumaUseriCr)
                userSheet.cell(row=3, column=9).value=str(len(users8))
                userSheet.cell(row=4, column=9).value=sum8Db
                userSheet.cell(row=5, column=9).value=sum8Cr

                print("users")

                #-----------------------------------------FORMATS AND COLORS------------------------------------------------

                userSheet.cell(row=1, column=1).font=ft1
                userSheet.cell(row=2, column=1).font=ft1

                userSheet.cell(row=1, column=2).font=font_worksheet
                userSheet.cell(row=2, column=2).font=font_worksheet

                userSheet.cell(row=5, column=1).font=f_testname

                for row in userSheet["A9:D9"]:
                    for cell in row:
                        cell.font = cap_tabel
                        cell.fill = cap_tabel_color_GT
                
                for k in range(0, len(amounts)):
                    userSheet.cell(row=10+k, column=3).number_format='#,##0_);(#,##0)'
                    userSheet.cell(row=10+k, column=4).number_format='#,##0_);(#,##0)'

                for k in range(0, len(amounts)):
                    userSheet.cell(row=10+k, column=1).font=font_worksheet
                    userSheet.cell(row=10+k, column=2).font=font_worksheet
                    userSheet.cell(row=10+k, column=3).font=font_worksheet
                    userSheet.cell(row=10+k, column=5).font=font_worksheet
                    userSheet.cell(row=10+k, column=4).font=font_worksheet

                userSheet.cell(row=3, column=9).number_format='#,##0_);(#,##0)'
                userSheet.cell(row=4, column=9).number_format='#,##0_);(#,##0)'
                userSheet.cell(row=5, column=9).number_format='#,##0_);(#,##0)'

                listaColoane=['A','B','C','D','E']
                for column in ascii_uppercase:
                    for i in listaColoane:
                        if (column==i):
                            userSheet.column_dimensions[column].width = 12

                userSheet.column_dimensions['H'].width = 22
                userSheet.column_dimensions['I'].width = 22
            except:
                pass

        #----------------------------------------------------POSTED ON SPECIFIC DATE------------------------------------

        if(poSD==1):
            try:
                specificDate = wb.create_sheet("Specific_date")
                specificDate.sheet_view.showGridLines = False
                summary.cell(row=22, column=3).value="Specific date"
                summary.cell(row=22, column=3).hyperlink='#Specific_date!A1'
                summary.cell(row=22, column=3).font=hyperlink_font
            
                specificDate.cell(row=1, column=1).value="Client name:"
                specificDate.cell(row=2, column=1).value="Period end:"

                specificDate.cell(row=1, column=2).value=clientname1
                try:
                    specificDate.cell(row=2, column=2).value=yearEnd1
                    specificDate.cell(row=2, column=2).number_format="mm/dd/yyyy"
                except:
                    pass

                specificDate.cell(row=5, column=1).value="Transactions by specific date"

                specificDate.cell(row=9, column=1).value="JE Number"
                specificDate.cell(row=9, column=2).value="Posting Date"
                specificDate.cell(row=9, column=3).value="Description"
                specificDate.cell(row=9, column=4).value="Account"
                specificDate.cell(row=9, column=5).value="Amount"

                # date=datetime.datetime(2021, 1, 1)
                date=[]
                # print(listaCostica)
                # date=spDays
                # print(type(datee))
                # specificDate.cell(row=9, column=6).value=spDays
                jeNo9=[]
                postDate9=[]
                descript9=[]
                acct9=[]
                amount9=[]
                
                for x in range(0, len(listaCostica)):
                    # print(listaCostica[x])
                    for i in range(0, len(postDate)):
                        if listaCostica[x] == postDate[i]:
                            # print("sunt la fel")
                            jeNo9.append(jeNo[i])
                            postDate9.append(postDate[i])
                            descript9.append(descriptionFull[i])
                            acct9.append(account[i])
                            amount9.append(amount[i])
                
                    
                for k in range(0, len(postDate9)):
                    specificDate.cell(row=10+k, column=1).value=jeNo9[k]
                    specificDate.cell(row=10+k, column=2).value=postDate9[k]
                    specificDate.cell(row=10+k, column=3).value=descript9[k]
                    specificDate.cell(row=10+k, column=4).value=acct9[k]
                    specificDate.cell(row=10+k, column=5).value=amount9[k]
            
            
                #-----------------------------------------FORMATS AND COLORS------------------------------------------------

                # specificDate.cell(row=1, column=1).font=ft1
                # specificDate.cell(row=2, column=1).font=ft1

                specificDate.cell(row=1, column=2).font=font_worksheet
                specificDate.cell(row=2, column=2).font=font_worksheet

                specificDate.cell(row=5, column=1).font=f_testname

                for row in specificDate["A9:E9"]:
                    for cell in row:
                        cell.font = cap_tabel
                        cell.fill = cap_tabel_color_GT
                
                for k in range(0, len(amount)):
                    specificDate.cell(row=10+k, column=2).number_format='MM/DD/YYYY'
                    specificDate.cell(row=10+k, column=5).number_format='#,##0_);(#,##0)'

                for k in range(0, len(amount)):
                    specificDate.cell(row=10+k, column=1).font=font_worksheet
                    specificDate.cell(row=10+k, column=2).font=font_worksheet
                    specificDate.cell(row=10+k, column=3).font=font_worksheet
                    specificDate.cell(row=10+k, column=5).font=font_worksheet
                    specificDate.cell(row=10+k, column=4).font=font_worksheet


                listaColoane=['A','B','C','D','E']
                for column in ascii_uppercase:
                    for i in listaColoane:
                        if (column==i):
                            specificDate.column_dimensions[column].width = 12

                specificDate.column_dimensions['H'].width = 12
                specificDate.column_dimensions['I'].width = 12

                print("specific date")
            except:
                specificDate.cell(row=1, column=1).font=ft1
                specificDate.cell(row=2, column=1).font=ft1

                specificDate.cell(row=1, column=2).font=font_worksheet
                specificDate.cell(row=2, column=2).font=font_worksheet

                specificDate.cell(row=5, column=1).font=f_testname

                for row in specificDate["A9:E9"]:
                    for cell in row:
                        cell.font = cap_tabel
                        cell.fill = cap_tabel_color_GT
                
                for k in range(0, len(amount)):
                    specificDate.cell(row=10+k, column=2).number_format='MM/DD/YYYY'
                    specificDate.cell(row=10+k, column=5).number_format='#,##0_);(#,##0)'

                for k in range(0, len(amount)):
                    specificDate.cell(row=10+k, column=1).font=font_worksheet
                    specificDate.cell(row=10+k, column=2).font=font_worksheet
                    specificDate.cell(row=10+k, column=3).font=font_worksheet
                    specificDate.cell(row=10+k, column=5).font=font_worksheet
                    specificDate.cell(row=10+k, column=4).font=font_worksheet


                listaColoane=['A','B','C','D','E']
                for column in ascii_uppercase:
                    for i in listaColoane:
                        if (column==i):
                            specificDate.column_dimensions[column].width = 12

                specificDate.column_dimensions['H'].width = 12
                specificDate.column_dimensions['I'].width = 12

        #----------------------------------------------------------------MISSING JOURNAL ENTRIES------------------------------------
        if(misJE==1):
            try:
                missJE = wb.create_sheet("Missing_Journal_Entries")
                missJE.sheet_view.showGridLines = False
                summary.cell(row=23, column=3).value="Missing Journal Entries"
                summary.cell(row=23, column=3).hyperlink='#Missing_Journal_Entries!A1'
                summary.cell(row=23, column=3).font=hyperlink_font
            
                missJE.cell(row=1, column=1).value="Client name:"
                missJE.cell(row=2, column=1).value="Period end:"

                # missJE.cell(row=1, column=2).value=clientname1
                # try:
                #     missJE.cell(row=2, column=2).value=yearEnd1
                # except:
                #     pass

                # missJE.cell(row=4, column=5).value="Number of records"
                missJE.cell(row=5, column=1).value="Missing Journal Entries"

                missJE.cell(row=10, column=1).value="JE Number From"
                missJE.cell(row=10, column=2).value="JE Number To"
                missJE.cell(row=10, column=3).value="Missing JE"

                listaJE=set(jeNo)
                listaUnicaJE=list(listaJE)
                # print(listaUnicaJE)
                listaUnicaJE.sort()
                listaTranz=[]
                listaTranz.append(1)
                for i in range(1, len(listaUnicaJE)):
                    listaTranz.append(listaUnicaJE[i]-listaUnicaJE[i-1])
                poz=10
                listaJEmissing=[]
                for k in range(0, len(listaTranz)):
                    if listaTranz[k]>1:
                        poz=poz+1
                        missJE.cell(row=poz,column=1).value=listaUnicaJE[k-1]
                        missJE.cell(row=poz,column=2).value=listaUnicaJE[k]
                        missJE.cell(row=poz,column=3).value=listaTranz[k]
                        
                # missJE.cell(row=4, column=6).value=len(listaTranz)

                #-----------------------------------------FORMATS AND COLORS------------------------------------------------

                missJE.cell(row=1, column=1).font=ft1
                missJE.cell(row=2, column=1).font=ft1

                missJE.cell(row=1, column=2).font=font_worksheet
                missJE.cell(row=2, column=2).font=font_worksheet

                missJE.cell(row=5, column=1).font=f_testname

                for row in missJE["A10:C10"]:
                    for cell in row:
                        cell.font = cap_tabel
                        cell.fill = cap_tabel_color_GT
                
                # for k in range(0, len(amounts)):
                #     missJE.cell(row=11+k, column=1).number_format='#,##0_);(#,##0)'
                #     missJE.cell(row=11+k, column=2).number_format='#,##0_);(#,##0)'

                for k in range(0, len(amounts)):
                    missJE.cell(row=11+k, column=1).font=font_worksheet
                    missJE.cell(row=11+k, column=2).font=font_worksheet
                    missJE.cell(row=11+k, column=3).font=font_worksheet
                    missJE.cell(row=11+k, column=5).font=font_worksheet
                    missJE.cell(row=11+k, column=4).font=font_worksheet


                listaColoane=['A','B','C','D','E']
                for column in ascii_uppercase:
                    for i in listaColoane:
                        if (column==i):
                            missJE.column_dimensions[column].width = 12

                missJE.column_dimensions['E'].width = 17
                missJE.column_dimensions['F'].width = 12
                print("missing JE")
            except:
                pass
        #-------------------------------------------------------------OUT OF BALANCE--------------------------
        if(oobal==1):
            try:
                outOfBal = wb.create_sheet("Out_of_Balance")
                outOfBal.sheet_view.showGridLines = False
                summary.cell(row=24, column=3).value="Out of Balance"
                summary.cell(row=24, column=3).hyperlink='#Out_of_Balance!A1'
                summary.cell(row=24, column=3).font=hyperlink_font

                outOfBal.cell(row=1, column=1).value="Client name:"
                outOfBal.cell(row=2, column=1).value="Period end:"

                outOfBal.cell(row=4, column=5).value="Number of records"
                outOfBal.cell(row=5, column=1).value="Out of Balance Journal Entries"

                outOfBal.cell(row=10, column=1).value="JE Number"
                outOfBal.cell(row=10, column=2).value="Amount"

                listaJE=set(jeNo)
                listaUnicaJE=list(listaJE)
                
                outBal=[]
                outAmt=[]
                for i in range(0, len(listaUnicaJE)):
                    s=0
                    for x in range(0, len(jeNo)):
                        
                        if listaUnicaJE[i]== jeNo[x]:
                            s=s+amount[x]
                    if(s>0.1 or s<-0.1):
                        outBal.append(listaUnicaJE[i])
                        outAmt.append(s)
                        # print(outBal, s)
                for k in range(0, len(outBal)):
                    outOfBal.cell(row=11+k, column=1).value=outBal[k]
                    outOfBal.cell(row=11+k, column=2).value=outAmt[k]

                outOfBal.cell(row=4, column=6).value=len(outBal)
                
                #-----------------------------------------FORMATS AND COLORS------------------------------------------------

                outOfBal.cell(row=1, column=1).font=ft1
                outOfBal.cell(row=2, column=1).font=ft1

                outOfBal.cell(row=1, column=2).font=font_worksheet
                outOfBal.cell(row=2, column=2).font=font_worksheet

                outOfBal.cell(row=5, column=1).font=f_testname

                for row in outOfBal["A10:B10"]:
                    for cell in row:
                        cell.font = cap_tabel
                        cell.fill = cap_tabel_color_GT
                
                # for k in range(0, len(amounts)):
                #     outOfBal.cell(row=11+k, column=2).number_format='MM/DD/YYYY'
                #     outOfBal.cell(row=11+k, column=5).number_format='#,##0_);(#,##0)'

                for k in range(0, len(outBal)):
                    outOfBal.cell(row=11+k, column=1).font=font_worksheet
                    outOfBal.cell(row=11+k, column=2).font=font_worksheet
                    # outOfBal.cell(row=11+k, column=3).font=font_worksheet
                    # outOfBal.cell(row=asdasda11+k, column=5).font=font_worksheet
                    # outOfBal.cell(row=11+k, column=4).font=font_worksheet


                listaColoane=['A','B']
                for column in ascii_uppercase:
                    for i in listaColoane:
                        if (column==i):
                            outOfBal.column_dimensions[column].width = 12

                outOfBal.column_dimensions['E'].width = 17
                outOfBal.column_dimensions['F'].width = 12

                print("oobal")
            except:
                pass

                
        

        wr = wb["Sheet"]
        wb.remove(wr)
       
        

        for i in range(1,12):
            summary.cell(row=13+i, column=2).value=i
            summary.cell(row=13+i, column=2).font=hyperlink_font
        # folderpath = ('D:\\WEBDEV\\Platforma apps web\\Audit\\JET\\output')
        # folderpath = ('D:\\Python Projects\\JET\\JET\\output')
        folderpath = ('home/auditappnexia/output/jet/')
        file_pathFS = os.path.join(folderpath, "JET "+clientname1+".xlsx")
        wb.save(file_pathFS)
    return send_from_directory(folderpath, "JET "+clientname1+".xlsx", as_attachment=True)
    # return render_template('JET.html')

# @app.route('/password')
# def passw():
	
	

# 	return render_template('password.html')

# @app.route('/password', methods=['POST', 'GET'])
# def pasw_process():
# 	# d = var_list.pop()
# 	# print(var_list)
# 		# var='%'+d+'%'
# 		# result = Mylist.query.filter(Mylist.type.like(var)
	
# 	if request.method == 'POST':
# 		password =request.form['pass']
# 		# print(d, "parola")
# 		if password in var_list:
# 			return redirect(url_for('home', user=password))
# 		else:

# 			flash("The password you have entered does not match the provided provided one! Please try again!")
# 			return render_template("index.html")
# @app.route('/home')
# def home2():
	


# 	return render_template('dashboard.html')

# @app.route('/', methods=['POST', 'GET'])

# def LogIn_process():
# 	# folderpath="D:/Projects/Python code-modele/PDF/PDF/"
# 	password_char="GT"
# 	email =request.form['email']
# 	# print(email)
# 	if request.method == 'POST':
# 		# flash("Please wait!")
# 	#     # print("abc")
		
# 		listaApproved=["cristian.iordache@ro.gt.com", "bogdan.constantinescu@ro.gt.com", "alina.boarca@ro.gt.com","andrei.soare@ro.gt.com", "stefan.ciochinaru@ro.gt.com", "anamaria.petrican@ro.gt.com", "stefan.vizireanu@ro.gt.com", "denis.david@ro.gt.com", "robert.dorobantu@ro.gt.com", "andrei.paunescu@ro.gt.com", "liliana.veneticu@ro.gt.com", "anca.bolontoc@ro.gt.com", "catalin.mazilu@ro.gt.com" ]
# 		# for i in listaApproved:
# 		Chars = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ123456789!@#$%^&*()"
# 		if str(email) not in listaApproved:
# 			ok=1
			
# 			flash("You do not have access to this section! Please contact the Automation Team for further details!")
# 			return render_template("index.html")
# 		else: 
# 			ok=2
# 		print(ok)
# 		if(ok==2):
# 			nume=email.split(".")[0].capitalize()
# 			for x in range(0, 5):
# 				password_char = password_char+random.choice(Chars)
# 			var_list.append(password_char)
# 			# password_login=password_char
# 			smtp_server = "smtp.office365.com"
# 			port = 587  # For starttls
# 			sender_email = "GTRDigital@ro.gt.com"
# 			password = "g[&vuBR9WQqr=7>D"
			
			

# 			# Create a secure SSL context
# 			context = ssl.create_default_context()
# 			message_text = "Dear "+nume+",\n\nYour temporary password is "+password_char+"\n\nIf you did not initiate this request, please ignore this e-mail.\n\nThank you,\nGTRDigital"
# 			subj = "Generated Password for GTR Digital"
# 			date = datetime.datetime.now().strftime( "%d/%m/%Y %H:%M" )

# 			msg = "From: %s\nTo: %s\nSubject: %s\nDate: %s\n\n%s" % ( "GTRDigital@ro.gt.com", email, subj, date, message_text )

# 			# Try to log in to server and send email
# 			try:
# 				server = smtplib.SMTP(smtp_server,port)
# 				server.ehlo() # Can be omitted
# 				server.starttls(context=context) # Secure the connection
# 				server.ehlo() # Can be omitted
# 				server.login(sender_email, password)
# 				# TODO: Send email here
# 				server.sendmail(
# 				sender_email, 
# 				email, 
# 				msg)
# 			except Exception as e:
# 				# Print any error messages to stdout
# 				print(e)
# 			finally:
# 				server.quit() 
			
# 			# f=open(folderpath+"pass"+str(datetime.now().strftime("%H%M%S"))+".txt","w").write(password_char)
# 			# f.close()
# 		# flash("Success!")
# 		# sleep(3)
		
# 		# return send_from_directory(folderpath,"SAF-T Reporting GT Romania.xml", as_attachment=True)
# 			# session['password_login'] = password_char
		 
	
# 			return redirect("/password")
# 			# return redirect(url_for('pasw_process', password_char=password_char))

# 			# return render_template("password.html", password_login=password_char)
# 	# else:

# 	#     return render_template("password.html")
# 	return redirect("/")
if __name__ == '__main__':
   app.run()
